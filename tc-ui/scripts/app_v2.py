#!/usr/bin/env python3
"""
TC 자동화 v2 — Flask + SSE + Claude AI
PDF/GitHub URL/텍스트 → 실시간 TC 생성 → Excel 출력
실행: python3 scripts/app_v2.py
접속: http://localhost:5001
"""

import os
import sys
import json
import uuid
import queue
import threading
import subprocess
import re
import time
from pathlib import Path
from datetime import datetime

# ── .env 로딩 ──────────────────────────────────────────────────────────────────
# Windows 메모장 저장 시 BOM이 붙을 수 있어 utf-8-sig 사용 (없으면 무해).
env_file = Path(__file__).parent.parent / ".env"
if env_file.exists():
    for line in env_file.read_text(encoding='utf-8-sig').splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        key, val = k.strip(), v.strip()
        # 값이 "..." 또는 '...'로 감싸져 있으면 벗겨내기 (Windows 사용자 흔한 실수 방어)
        if len(val) >= 2 and ((val[0] == val[-1] == '"') or (val[0] == val[-1] == "'")):
            val = val[1:-1]
        if key and not os.environ.get(key):
            os.environ[key] = val

from flask import Flask, request, jsonify, render_template_string, send_file, Response

# ── 경로 설정 ──────────────────────────────────────────────────────────────────
BASE_DIR       = Path(__file__).parent.parent   # /tc-ui/
AGENT_DIR      = BASE_DIR.parent / "tc-agent"
RULES_FILE     = AGENT_DIR / "common" / "tc-rules.md"
BUILD_EXCEL    = AGENT_DIR / "scripts" / "build_excel.py"
WORKSPACE_ROOT   = BASE_DIR / "workspace"
OUTPUTS_DIR      = BASE_DIR / "outputs"
SPECS_DIR        = BASE_DIR / "specs"
TC_FILES_DIR     = BASE_DIR / "tc_files"    # 저장된 TC 마크다운
PROJECTS_FILE    = BASE_DIR / "projects.json"  # 프로젝트 레지스트리
CONFIG_FILE      = BASE_DIR / "config.json"
DRIVE_CREDS_FILE = BASE_DIR / "credentials.json"
DRIVE_TOKEN_FILE = BASE_DIR / ".drive_token.json"
PORT             = int(os.environ.get("PORT", 5001))
MODEL          = "claude-opus-4-5"

# ── 앱 버전 (단일 소스 — 여기 한 곳만 수정하면 UI 배지/배너/모달/JS 상수 모두 자동 반영) ──
APP_VERSION         = "v0.12.1"
APP_VERSION_DATE    = "2026-05-14"
APP_VERSION_TAGLINE = "TC Update 모드 — 기획서 변경 기반 기존 TC 자동 갱신"
# 릴리즈 요약 — UI 배너/모달용 (4~5줄 권장)
APP_VERSION_HIGHLIGHTS = [
    "🆕 TC Update 모드 — 기획서 버전업 시 기존 Google Sheets TC 를 사본에 자동 갱신",
    "🗂️ 후보 리스트 SCR 단위 그룹화 — 기본 접힘 + 헤더에 TC 수/체크 카운트/일괄 적용 버튼",
    "✨ 신규 SCR 은 별도 '신규 TC 생성 모드' 로 분기 — 1단계 분석 결과에서 바로 진입",
    "🛡️ 통합 게이트 네비게이션 박스 + 진행 중 재시작 가드 + SSE 정상 종료",
    "🤖 propose AI 6 원칙 강화 — UI 검증 TC 인식 + 보수적 판단 + no_change 남용 방지",
    "⚡ Sheets API 쿼터 방어 — 헤더/TC 본문 세션 캐시 + 429 명시 안내",
    "📁 TC Update Drive 폴더 사용자 설정 (~/.tc-update-config.json)",
    "📦 SCR 일괄 적용 — 해당 SCR 의 체크된 modify 만 AI 호출 + 사본 셀 갱신",
]

WORKSPACE_ROOT.mkdir(exist_ok=True)
OUTPUTS_DIR.mkdir(exist_ok=True)
SPECS_DIR.mkdir(exist_ok=True)
TC_FILES_DIR.mkdir(exist_ok=True)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB

# ── 세션 저장소 ────────────────────────────────────────────────────────────────
SESSIONS: dict[str, dict] = {}

def new_session() -> dict:
    sid = str(uuid.uuid4())[:8]
    ws  = WORKSPACE_ROOT / sid
    ws.mkdir(exist_ok=True)
    sess = {
        "id":               sid,
        "workspace":        ws,
        "events":           queue.Queue(),
        "gate_event":       threading.Event(),
        "approved":         None,
        "selected_domains": None,  # None=전체, list=선택된 대분류 코드
        "status":           "idle",
        "result":           None,
        "thread":           None,
        "project_name":     "",
        "stop_requested":   False,
    }
    SESSIONS[sid] = sess
    return sess


class PipelineStopError(Exception):
    """사용자가 중단 버튼을 눌렀을 때 발생"""
    pass


def check_stop(sess: dict):
    """중단 요청 시 PipelineStopError 를 발생시킨다."""
    if sess.get("stop_requested"):
        raise PipelineStopError("사용자가 파이프라인을 중단했습니다.")

# ── 프로젝트 레지스트리 ────────────────────────────────────────────────────────
def load_projects() -> list:
    if PROJECTS_FILE.exists():
        try:
            data = json.loads(PROJECTS_FILE.read_text(encoding="utf-8"))
        except Exception:
            return []
        # 이름이 비어있는 유령 레코드 자동 필터 (과거 버그로 남은 데이터 방어)
        return [p for p in data if isinstance(p, dict) and (p.get("name") or "").strip()]
    return []

def save_project(project_name: str, tc_file: str = "", excel_file: str = "", **extra):
    # 빈 이름 저장 방지 — "단발성 작업"이나 프로젝트 미선택 케이스는 레코드를 남기지 않음
    project_name = (project_name or "").strip()
    if not project_name:
        return
    projects = load_projects()
    existing = next((p for p in projects if p["name"] == project_name), None)
    entry = {
        "name":       project_name,
        "tc_file":    tc_file or (existing or {}).get("tc_file", ""),
        "excel_file": excel_file or (existing or {}).get("excel_file", ""),
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    # 기존 필드 보존 (pipeline_state 등)
    if existing:
        for k, v in existing.items():
            if k not in entry:
                entry[k] = v
    # extra 필드 병합
    entry.update(extra)

    # 최근 사용 spec 폴더 이력 누적 (last_sources 에서 spec_folder 타입 추출)
    if "last_sources" in extra:
        recent = list(entry.get("recent_spec_folders") or [])
        for src in extra["last_sources"]:
            if src.get("type") in ("spec_folder", "spec_folder_prev") and src.get("content"):
                path = src["content"]
                if path in recent:
                    recent.remove(path)
                recent.insert(0, path)  # 최신을 맨 앞에
        # 최대 10개
        entry["recent_spec_folders"] = recent[:10]

    if existing:
        projects = [entry if p["name"] == project_name else p for p in projects]
    else:
        projects.insert(0, entry)
    PROJECTS_FILE.write_text(json.dumps(projects, ensure_ascii=False, indent=2), encoding="utf-8")


# ── TC Update 작업 이력 (프로젝트별 최근 N개) ──────────────────────────────
UPDATE_HISTORY_MAX = 5


def _derive_update_label(prev_folder: str, new_folder: str) -> str:
    """폴더명에서 버전 추출해 'v0.52.0 → v0.54.0' 같은 짧은 라벨 생성."""
    import re as _re
    def _ver(p: str) -> str:
        m = _re.search(r"v\d+\.\d+\.\d+", p or "")
        return m.group(0) if m else (Path(p).name[:20] if p else "?")
    return f"{_ver(prev_folder)} → {_ver(new_folder)}"


def save_update_history(project_name: str, prev_folder: str, new_folder: str,
                          existing_tc_url: str) -> None:
    """프로젝트의 update 작업 이력에 한 건 추가. 같은 (prev,new,url) 조합은 최상위로 이동.

    유효성 — 세 필드 모두 있어야 저장. 빈 프로젝트명도 skip.
    """
    project_name = (project_name or "").strip()
    if not project_name:
        return
    if not (prev_folder and new_folder and existing_tc_url):
        return
    projects = load_projects()
    proj = next((p for p in projects if p["name"] == project_name), None)
    if not proj:
        return

    history = list(proj.get("update_history") or [])
    # 같은 조합이 있으면 제거 (맨 위로 올림)
    history = [h for h in history if not (
        h.get("prev_folder") == prev_folder
        and h.get("new_folder") == new_folder
        and h.get("existing_tc_url") == existing_tc_url
    )]
    entry = {
        "prev_folder": prev_folder,
        "new_folder": new_folder,
        "existing_tc_url": existing_tc_url,
        "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "label": _derive_update_label(prev_folder, new_folder),
    }
    history.insert(0, entry)
    proj["update_history"] = history[:UPDATE_HISTORY_MAX]
    PROJECTS_FILE.write_text(
        json.dumps(projects, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def get_update_history(project_name: str) -> list:
    """프로젝트의 update 작업 이력 반환 (최신 순). 없으면 []."""
    project_name = (project_name or "").strip()
    if not project_name:
        return []
    projects = load_projects()
    proj = next((p for p in projects if p["name"] == project_name), None)
    return list((proj or {}).get("update_history") or [])


def delete_update_history(project_name: str, index: int) -> bool:
    """프로젝트의 update_history 에서 index 번째 항목 삭제.
    Returns: True 면 삭제 성공, False 면 실패 (프로젝트 없음 / index 범위 초과).
    """
    project_name = (project_name or "").strip()
    if not project_name:
        return False
    projects = load_projects()
    proj = next((p for p in projects if p["name"] == project_name), None)
    if not proj:
        return False
    history = list(proj.get("update_history") or [])
    if index < 0 or index >= len(history):
        return False
    history.pop(index)
    proj["update_history"] = history
    PROJECTS_FILE.write_text(
        json.dumps(projects, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return True


# ── 프로젝트별 파이프라인 상태 저장 (이어서 작업용) ─────────────────────────────
PIPELINE_STATES_DIR = BASE_DIR / "pipeline_states"
PIPELINE_STATES_DIR.mkdir(exist_ok=True)

def save_pipeline_state(project_name: str, stage: str, data: dict):
    """파이프라인 중간 상태를 프로젝트별 JSON으로 저장"""
    safe = re.sub(r"[^\w\-_]", "_", project_name)[:40]
    state_file = PIPELINE_STATES_DIR / f"{safe}.json"
    state = {
        "project_name": project_name,
        "stage":        stage,
        "saved_at":     datetime.now().strftime("%Y-%m-%d %H:%M"),
        "data":         data,
    }
    state_file.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    # 프로젝트 레지스트리에도 상태 요약 기록
    save_project(project_name, pipeline_stage=stage, pipeline_saved_at=state["saved_at"])

def load_pipeline_state(project_name: str) -> dict | None:
    """저장된 파이프라인 상태 로드"""
    safe = re.sub(r"[^\w\-_]", "_", project_name)[:40]
    state_file = PIPELINE_STATES_DIR / f"{safe}.json"
    if state_file.exists():
        try:
            return json.loads(state_file.read_text(encoding="utf-8"))
        except Exception:
            return None
    return None

def clear_pipeline_state(project_name: str):
    """완료 시 파이프라인 상태 파일 삭제"""
    safe = re.sub(r"[^\w\-_]", "_", project_name)[:40]
    state_file = PIPELINE_STATES_DIR / f"{safe}.json"
    if state_file.exists():
        state_file.unlink()
    # 프로젝트 레지스트리에서 상태 필드 제거
    projects = load_projects()
    for p in projects:
        if p["name"] == project_name:
            p.pop("pipeline_stage", None)
            p.pop("pipeline_saved_at", None)
            break
    PROJECTS_FILE.write_text(json.dumps(projects, ensure_ascii=False, indent=2), encoding="utf-8")


CHECKPOINT_FILE = BASE_DIR / ".bkit" / "state" / "modify_checkpoint.json"
CHECKPOINT_FILE.parent.mkdir(parents=True, exist_ok=True)

def save_checkpoint(project_name: str, change_desc: str, stage: str, partial: dict = None):
    """TC 수정 파이프라인 체크포인트 저장"""
    data = {
        "project_name": project_name,
        "change_desc": change_desc,
        "stage": stage,
        "partial": partial or {},
        "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    CHECKPOINT_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def load_checkpoint() -> dict:
    if CHECKPOINT_FILE.exists():
        try:
            return json.loads(CHECKPOINT_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}

def clear_checkpoint():
    if CHECKPOINT_FILE.exists():
        CHECKPOINT_FILE.unlink()


def push(sess: dict, etype: str, data: dict):
    sess["events"].put({"type": etype, "data": data})

def push_stage(sess, stage: int, label: str, pct: int, eta_sec: int | None = None):
    data = {"stage": stage, "label": label, "pct": pct}
    if eta_sec is not None:
        data["eta_sec"] = eta_sec
    push(sess, "stage", data)

def push_log(sess, msg: str):
    push(sess, "log", {"msg": msg})

def push_error(sess, msg: str):
    sess["status"] = "error"
    push(sess, "error", {"msg": msg})


def count_unique_tc_ids(tc_content: str) -> int:
    """tc_content에서 유니크 TC ID 개수를 반환.
    build_excel의 parse_tc_markdown이 TC ID 기반 dedup을 수행하므로,
    웹에 표시하는 값도 동일한 유니크 기준으로 맞춘다.
    """
    if not tc_content:
        return 0
    ids = re.findall(r"^###\s+\*?\*?([A-Z]{2,}-[A-Z0-9]+-\d+)", tc_content, re.MULTILINE)
    return len(set(ids)) if ids else 0


# ── Claude API 헬퍼 ────────────────────────────────────────────────────────────
def call_claude(system_prompt: str, user_prompt: str, max_tokens: int = 8192) -> str:
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY", ""))
        msg = client.messages.create(
            model=MODEL,
            max_tokens=max_tokens,
            temperature=0,
            system=system_prompt,
            messages=[{"role": "user", "content": user_prompt}],
        )
        return msg.content[0].text
    except Exception as e:
        raise RuntimeError(f"Claude API 오류: {e}")


def call_claude_cached(system_blocks: list, user_prompt: str, max_tokens: int = 8192) -> str:
    """system_blocks: list of {"type":"text","text":..., "cache_control":{"type":"ephemeral"} 선택}.
    첫 호출은 캐시 미스(오버헤드 ~25%), 이후 5분 내 동일 블록 호출은 캐시 히트(약 90% 비용↓).
    화면별 1:1 호출처럼 system 부분이 동일하고 user만 바뀔 때 효과 큼.
    """
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY", ""))
        msg = client.messages.create(
            model=MODEL,
            max_tokens=max_tokens,
            temperature=0,
            system=system_blocks,
            messages=[{"role": "user", "content": user_prompt}],
        )
        return msg.content[0].text
    except Exception as e:
        raise RuntimeError(f"Claude API 오류 (cached): {e}")

# ── TC 규칙 로딩 ───────────────────────────────────────────────────────────────
FEWSHOT_FILE = AGENT_DIR / "common" / "tc-sample-fewshot.md"

def load_tc_rules() -> str:
    if RULES_FILE.exists():
        return RULES_FILE.read_text(encoding="utf-8")
    return ""

def load_fewshot_examples() -> str:
    if FEWSHOT_FILE.exists():
        return FEWSHOT_FILE.read_text(encoding="utf-8")
    return ""

# ── 프로젝트별 정책 파일 로딩 ──────────────────────────────────────────────────
PROJECTS_RULES_DIR = AGENT_DIR / "projects"

def load_project_policies(project_name: str) -> str:
    """프로젝트명으로 매칭되는 정책 파일들을 로드.
    projects/ 하위 폴더명과 프로젝트명을 대소문자 무시로 매칭."""
    if not PROJECTS_RULES_DIR.exists():
        return ""
    # 프로젝트명 → 폴더명 매칭 (supercycl, tc-manager 등)
    pname_lower = project_name.lower().replace(" ", "").replace("-", "").replace("_", "")
    for folder in PROJECTS_RULES_DIR.iterdir():
        if not folder.is_dir():
            continue
        fname_lower = folder.name.lower().replace(" ", "").replace("-", "").replace("_", "")
        if fname_lower in pname_lower or pname_lower in fname_lower:
            # 해당 폴더의 .md 파일 모두 읽기
            texts = []
            for md_file in sorted(folder.glob("*.md")):
                texts.append(f"### 📋 {md_file.stem}\n\n{md_file.read_text(encoding='utf-8')[:5000]}")
            if texts:
                return "\n\n---\n\n".join(texts)
    return ""

# ── 1단계: 문서 파싱 ──────────────────────────────────────────────────────────
def step_parse(sess: dict, input_type: str, content: str) -> str:
    """단일 소스 파싱 (하위 호환용)"""
    return step_parse_sources(sess, [{"type": input_type, "content": content}])


def extract_scr_from_source(filename: str, content: str) -> str:
    """입력 소스에서 화면 식별자(SCR-NNN[A]) 추출 — Phase 1 단순 패턴.

    우선순위:
      1) 파일명 매칭: SCR-003.md → 'SCR-003'
      2) 본문 H1 첫 줄: '# SCR-003: ...' → 'SCR-003'
      3) 본문 첫 번째 패턴 등장
      4) 위 모두 실패 → '' (빈 문자열, AI 환각/추측 금지)
    """
    pattern = re.compile(r'SCR-\d+[A-Z]?')

    # 1) 파일명
    if filename:
        m = pattern.search(filename)
        if m:
            return m.group(0)

    if not content:
        return ''

    # 2) H1 첫 줄
    first_line = content.lstrip().split('\n', 1)[0] if content.lstrip() else ''
    if first_line.startswith('#'):
        m = pattern.search(first_line)
        if m:
            return m.group(0)

    # 3) 본문 첫 번째 등장
    m = pattern.search(content)
    if m:
        return m.group(0)

    # 4) 폴백
    return ''


def strip_appendix_tables(filename: str, content: str, primary_scr: str) -> tuple[str, list]:
    """입력 소스 본문에서 다른 SCR 화면의 부록 표를 자동 감지하여 제거 — v0.9.26.

    동료 보고 케이스: SCR-810.md 본문 끝에 SCR-403/SCR-221/SCR-410 의 에러 케이스
    부록 표가 붙어있어 AI 가 무관한 분류를 만드는 문제.

    감지 패턴:
      1. 부록 헤더: '**에러 케이스 (...)**:' 또는 '## 부록' 또는 '---' 다음에
         '| 화면 |' 컬럼이 있는 표
      2. 표 안의 '화면' 컬럼이 입력 파일의 primary_scr 와 다른 SCR 코드
      3. '---' 구분자 + 별도 SCR 코드들이 연이어 등장

    안전장치 (제거 안 함):
      - primary_scr 가 비어있으면 (식별자 없는 임의 마크다운) 부록 감지 시도 안 함
      - 표 안에 primary_scr 가 다수 등장하면 본문 핵심으로 판단해 보존

    Returns:
        (정제된 content, [{"reason": str, "scr_codes": [...]}, ...] 제거 정보)
    """
    if not content or not primary_scr:
        return content, []

    pattern_scr = re.compile(r'SCR-\d+[A-Z]?')

    # 줄 단위로 쪼개서 부록 후보 찾기
    lines = content.split('\n')
    removed_segments = []  # 제거된 정보 (사용자 안내용)
    keep_lines = []
    i = 0
    n = len(lines)

    while i < n:
        line = lines[i]

        # 부록 헤더 패턴 — 본문 끝에서 자주 발견되는 형식
        is_appendix_header = (
            re.match(r'^\*\*에러\s*케이스\s*\(.+?\)\*\*\s*:?\s*$', line.strip()) or
            re.match(r'^\*\*부록.+?\*\*\s*:?\s*$', line.strip()) or
            re.match(r'^##\s+부록', line.strip()) or
            re.match(r'^##\s+다른\s*화면.+', line.strip())
        )

        # 부록 헤더 발견 → 다음 표의 SCR 들이 primary_scr 와 다른지 검사
        if is_appendix_header:
            # 헤더부터 표 끝까지 (또는 다음 ---) 의 영역 미리보기
            preview_end = min(n, i + 50)
            preview_block = '\n'.join(lines[i:preview_end])
            scrs_in_block = set(pattern_scr.findall(preview_block))
            other_scrs = [s for s in scrs_in_block if s != primary_scr]
            if other_scrs and primary_scr not in scrs_in_block:
                # 부록 영역이 다른 화면만 다룸 → 제거
                removed_segments.append({
                    "reason": "부록 표 — 다른 화면의 정보",
                    "scr_codes": sorted(other_scrs),
                    "header": line.strip(),
                })
                # 다음 빈 줄 + '---' 까지 또는 다음 H1/H2 까지 skip
                j = i + 1
                while j < n:
                    nxt = lines[j].strip()
                    # 다음 ---  / 다음 H1 / 다음 H2 / 다음 부록 헤더 만나면 종료 (그 줄은 keep 로 진입)
                    if nxt == '---':
                        # --- 자체는 보존하지 말고 skip (부록 마무리 구분선)
                        j += 1
                        break
                    if re.match(r'^#{1,2}\s+', nxt) and not re.match(r'^##\s+부록', nxt):
                        break  # 다음 일반 섹션 — keep 시작
                    j += 1
                i = j
                continue

        # 일반 라인 — 보존
        keep_lines.append(line)
        i += 1

    cleaned = '\n'.join(keep_lines)
    # 연속 빈 줄 정리 (3개 이상 → 2개)
    cleaned = re.sub(r'\n{3,}', '\n\n', cleaned)
    return cleaned, removed_segments


def step_parse_sources(sess: dict, sources: list) -> str:
    """복수 소스를 각각 파싱한 뒤 하나의 텍스트로 합친다."""
    parts = []
    total = len(sources)
    # 입력 소스의 SCR 식별자 매핑 — TC 작성 단계에서 AI 에게 전달됨
    sess["_source_scr_map"] = {}  # {filename: scr_id}
    for i, src in enumerate(sources, 1):
        src_type = src.get("type", "text")
        content  = src.get("content", "").strip()
        push_log(sess, f"[파싱] 소스 {i}/{total} — {src_type}")

        if src_type == "pdf":
            pdf_path = SPECS_DIR / content
            if not pdf_path.exists():
                raise FileNotFoundError(f"PDF 파일 없음: {pdf_path}")
            text = extract_pdf_text(pdf_path)
            label = f"📄 PDF: {content}"
            push_log(sess, f"[파싱] PDF 추출 완료 ({len(text):,}자)")
        elif src_type == "url":
            selected_files = src.get("selected_files") or None
            text  = fetch_url_content(sess, content, selected_files=selected_files)
            label = f"🔗 URL: {content}"
            push_log(sess, f"[파싱] URL 추출 완료 ({len(text):,}자)")
        elif src_type == "web":
            text  = fetch_web_page(sess, content)
            label = f"🌐 웹: {content}"
            push_log(sess, f"[파싱] 웹 크롤링 완료 ({len(text):,}자)")
        elif src_type == "md":
            md_path = SPECS_DIR / content
            if not md_path.exists():
                raise FileNotFoundError(f"마크다운 파일 없음: {md_path}")
            text = md_path.read_text(encoding="utf-8")
            label = f"📝 마크다운: {content}"
            push_log(sess, f"[파싱] 마크다운 읽기 완료 ({len(text):,}자)")
            # SCR 식별자 추출 (마크다운 입력만 우선 적용)
            scr_id = extract_scr_from_source(content, text)
            if scr_id:
                sess["_source_scr_map"][content] = scr_id
                push_log(sess, f"[파싱] 화면 식별자 발견: {content} → {scr_id}")
                # v0.9.26: 부록 표 자동 감지 + 제거 — 다른 화면의 무관한 정보 차단
                cleaned_text, removed = strip_appendix_tables(content, text, scr_id)
                if removed:
                    text = cleaned_text
                    sess.setdefault("_appendix_removed", []).append({
                        "filename": content,
                        "primary_scr": scr_id,
                        "removed": removed,
                    })
                    for r in removed:
                        push_log(sess, f"[파싱] 부록 자동 제거 — {content} 의 '{r['header']}' (다른 화면: {', '.join(r['scr_codes'])})")
        else:
            text  = content
            label = "✏️ 텍스트 입력"
            push_log(sess, f"[파싱] 텍스트 입력 ({len(text):,}자)")

        parts.append(f"===== 소스 {i}/{total} — {label} =====\n\n{text}")

    raw_text = "\n\n".join(parts)
    parsed_path = sess["workspace"] / "01_parsed.md"
    parsed_path.write_text(f"# 파싱 결과\n\n{raw_text}", encoding="utf-8")
    # SCR 매핑 요약 로그
    if sess["_source_scr_map"]:
        push_log(sess, f"[파싱] 입력 소스 화면 식별자 매핑: {len(sess['_source_scr_map'])}개")
    else:
        push_log(sess, "[파싱] 입력 소스에서 화면 식별자(SCR-NNN) 미발견 — 화면 코드 컬럼은 빈 칸으로 출력됩니다")
    return raw_text


# ── 구조화 spec 폴더 처리 (v0.10.0+) ─────────────────────────────────────────
# 새 기획서 형식: 폴더 1개에 overview/policy/design/scr/*.md 가 역할별로 분리됨.
# 기존 "raw_text concat" 모드와 병행. 사용자가 폴더 경로로 입력하면 이 흐름을 탄다.

def classify_spec_files(folder: Path) -> dict:
    """spec 폴더 안의 md 파일들을 역할별로 자동 분류.
    반환: {"overview": Path|None, "policy": [Path...], "design": [Path...], "screens": [Path...]}
    분류 규칙(파일명/경로 우선순위):
      - scr/ 또는 SCR-* 패턴 → screens
      - *error*, *policy*, *biz*, *rule* → policy
      - *design*, *ux*, *theme*, *token* → design
      - 01_*, *spec*, *overview* → overview (첫 매칭 1개)
      - 그 외 .md → overview 폴백 (있으면 policy 으로)
    """
    out = {"overview": None, "policy": [], "design": [], "screens": []}
    if not folder.exists() or not folder.is_dir():
        return out

    # 1) screens — scr/ 하위 또는 파일명이 SCR-* 패턴
    scr_dir = folder / "scr"
    if scr_dir.exists() and scr_dir.is_dir():
        out["screens"] = sorted([p for p in scr_dir.glob("*.md") if p.is_file()])
    # 루트에 SCR-XXX.md 가 직접 있는 경우도 수용
    for p in sorted(folder.glob("SCR-*.md")):
        if p.is_file() and p not in out["screens"]:
            out["screens"].append(p)

    # 2) 루트 md 분류 (scr/ 안의 파일은 위에서 처리됨)
    for p in sorted(folder.glob("*.md")):
        if not p.is_file():
            continue
        name = p.name.lower()
        # 생성물(generated) 은 참고용으로만 — design 으로 분류
        if "error" in name or "policy" in name or "biz-rule" in name or "business-rule" in name:
            out["policy"].append(p)
        elif "design" in name or name.startswith("02_ux") or "theme" in name or "token" in name or "ux_design" in name:
            out["design"].append(p)
        elif name.startswith("01_") or "spec" in name or "overview" in name:
            if out["overview"] is None:
                out["overview"] = p
            else:
                out["policy"].append(p)
        else:
            # 미상 → policy 폴백 (overview 가 없으면 overview 로)
            if out["overview"] is None:
                out["overview"] = p
            else:
                out["policy"].append(p)
    return out


def parse_screen_list_table(overview_md_text: str) -> list[dict]:
    """01_spec.md 의 '화면 목록' 표를 파싱해 분류표 시드를 만든다.
    헤더 형식: | ID | 화면명 | 대분류 | 중분류 | 상태 | 설명 | 진입 경로 |
    상태(status) 컬럼이 'deprecated'/'미사용' 같으면 제외.
    반환: [{"id":"SCR-001","name":"Login","major":"Onboarding","middle":"Login","status":"active","desc":"...","entry":"..."}]
    """
    rows = []
    in_table = False
    headers: list[str] = []
    for line in overview_md_text.splitlines():
        line = line.rstrip()
        # 표 시작 감지: ID/화면명/대분류 같은 헤더가 있어야 함
        if line.startswith("|") and "ID" in line and ("대분류" in line or "Major" in line):
            cells = [c.strip() for c in line.strip("|").split("|")]
            headers = [c.lower() for c in cells]
            in_table = True
            continue
        if in_table and line.startswith("|---"):
            continue
        if in_table:
            if not line.startswith("|"):
                in_table = False
                headers = []
                continue
            cells = [c.strip() for c in line.strip("|").split("|")]
            if len(cells) < 4:
                continue
            row = {}
            for i, h in enumerate(headers):
                if i >= len(cells):
                    break
                row[h] = cells[i]
            # 키 정규화
            sid = row.get("id") or row.get("화면id") or ""
            name = row.get("화면명") or row.get("name") or ""
            major = row.get("대분류") or row.get("major") or ""
            middle = row.get("중분류") or row.get("middle") or ""
            status = row.get("상태") or row.get("status") or "active"
            desc = row.get("설명") or row.get("description") or ""
            entry = row.get("진입 경로") or row.get("진입경로") or row.get("entry") or ""
            if not sid or not re.match(r"^SCR[-_]?\w+", sid, re.IGNORECASE):
                continue
            if status.lower() in ("deprecated", "미사용", "obsolete", "삭제"):
                continue
            rows.append({
                "id": sid.upper().replace("_", "-"),
                "name": name,
                "major": major,
                "middle": middle,
                "status": status,
                "desc": desc,
                "entry": entry,
            })
    return rows


def extract_minors_from_screen_md(md_text: str, max_minors: int = 20) -> list[str]:
    """SCR-XXX.md 본문에서 소분류(세부 시나리오) 시드를 규칙 기반으로 추출.
    LLM 호출 없음. 우선순위:
      1. 상태(Status) 표 케이스 — 가장 구조화된 시나리오
      2. 에러 케이스 표
      3. [dev] 인터랙션 항목
      4. 비고의 [정책]/[제약]/[접근성]/[세션] 마커
      5. 계산 공식 / 수치 변환 — ROE, Est., 공식, "= X × Y" 패턴 (정확성 검증 누락 방지)
    너무 많아지면 max_minors 까지 자른다.
    각 항목은 짧은 한국어 라벨(50자 이내)로 정규화.
    """
    minors: list[str] = []
    seen_keys: set[str] = set()  # 중복 제거용 (소문자 정규화 비교)

    def add(label: str):
        label = label.strip().rstrip(".·")
        if not label:
            return
        # 정규화 — 시인성 강화 (v0.10.x):
        #   1) Markdown ** 강조 제거 / 백틱 / HTML 태그 제거
        #   2) 메타 마커 제거 ([정책]/[보안]/[제약]/...) — 시드 라벨이 아니라 분류 메타
        #   3) dash(공백 — 공백) 뒤 부연 설명 cutoff
        #   4) 콜론·괄호 뒤 부연 cutoff (trigger label 은 보호)
        label = re.sub(r"\*\*([^*]+)\*\*", r"\1", label)  # **xxx** → xxx
        label = re.sub(r"`[^`]+`", "", label)              # 백틱 코드 제거
        label = re.sub(r"<[^>]+>", "", label)              # HTML 태그 제거
        # 메타 마커 제거 — '[정책] Cancel All 버튼 없음' → 'Cancel All 버튼 없음'
        label = re.sub(
            r"\s*\[(통합|미결|그룹|메모|TODO|보류|보안|정책|제약|접근성|세션|타이밍|dev|"
            r"성능|UX|UI|에러|edge|규약|규칙)[^\]]*\]\s*",
            " ", label, flags=re.IGNORECASE,
        ).strip()
        # dash 뒤 부연 cutoff — 양쪽 공백 포함 dash (— / – / -- / -)
        # 예: 'Cancel All 버튼 없음 — Lite 는 Training Wheels futures 로 초보자...'
        #      → 'Cancel All 버튼 없음'
        m_dash = re.match(r"^(.+?)\s+[—–\-]{1,2}\s+(.+)$", label)
        if m_dash:
            head_dash = m_dash.group(1).strip()
            if len(head_dash) >= 4:  # 너무 짧으면 보존
                label = head_dash
        # 콜론·괄호 뒤 부연설명 cutoff — '키' 부분만 남김 (단, 너무 짧아지면 원본 유지)
        # Trigger label guard: head 가 trigger 라벨('… 시', '… 시점', '… 탭', '… 후',
        # '… 클릭') 로 끝나면 cutoff 시 의미 손실 → 전체 라벨 보존
        # 길이 제한 없음 — 'Reconnect with OKX 버튼 탭 시' 같은 긴 trigger 도 보호.
        for sep in [":", "("]:
            if sep in label:
                head = label.split(sep, 1)[0].strip()
                if len(head) >= 6:
                    head_clean = re.sub(r'["\'`]', "", head).strip()
                    is_trigger = (
                        head_clean.endswith("시") or
                        head_clean.endswith("시점") or
                        head_clean.endswith("탭") or
                        head_clean.endswith("후") or
                        head_clean.endswith("클릭") or
                        head_clean.endswith("진입") or
                        head_clean.endswith("선택")
                    )
                    if not is_trigger:
                        label = head
                        break
        # 짧은 라벨 (32자) 보장 — 길면 '…' 없이 깔끔하게 잘라냄
        label = re.sub(r"\s+", " ", label).strip(" .·-—")
        if len(label) > 32:
            label = label[:32].rstrip()
        if len(label) < 3:
            return
        key = re.sub(r"\s+", "", label.lower())
        if key in seen_keys:
            return
        seen_keys.add(key)
        minors.append(label)

    # 1) 상태 표 — | 케이스 | 조건 | 설명/동작 |
    in_status = False
    for line in md_text.splitlines():
        if not in_status:
            if re.search(r"상태\s*\(\s*Status\s*\)|^##+\s*상태", line):
                in_status = "header_pending"
            continue
        if in_status == "header_pending":
            if line.strip().startswith("|") and "케이스" in line:
                in_status = "rows"
            continue
        if in_status == "rows":
            if line.strip().startswith("|---"):
                continue
            if not line.strip().startswith("|"):
                in_status = False
                continue
            cells = [c.strip() for c in line.strip("|").split("|")]
            if len(cells) >= 1 and cells[0]:
                # 케이스 라벨 (Normal, Error.network) 을 소분류 시드로
                add(f"{cells[0]} 상태 표시 및 동작")

    # 1.5) 계산 정확성 — 상태 직후, 다른 카테고리보다 먼저 (max_minors 캡 도달 시 보장)
    #      도메인 본질이라 우선순위 가장 높음.
    #      정의는 본 함수 하단에 모아둠 → 여기서는 헬퍼 호출.
    _add_calc_accuracy_minors(md_text, add, domain_max=8, general_max=6)

    # 2) 에러 케이스 표 — | 에러 케이스 | 표시 패턴 | 메시지 | 동작 |
    in_err = False
    err_headers: list[str] = []
    for line in md_text.splitlines():
        if not in_err:
            if re.search(r"에러\s*케이스|^##+\s*에러", line):
                in_err = "header_pending"
            continue
        if in_err == "header_pending":
            if line.strip().startswith("|") and ("에러" in line or "케이스" in line):
                err_headers = [c.strip().lower() for c in line.strip("|").split("|")]
                in_err = "rows"
            continue
        if in_err == "rows":
            if line.strip().startswith("|---"):
                continue
            if not line.strip().startswith("|"):
                in_err = False
                continue
            cells = [c.strip() for c in line.strip("|").split("|")]
            if len(cells) >= 1 and cells[0]:
                add(f"에러 처리 — {cells[0]}")

    # 3) [dev] 인터랙션 항목 — bullet list (- xxx → yyy)
    in_inter = False
    for line in md_text.splitlines():
        if not in_inter:
            if re.search(r"\[dev\]\s*인터랙션|^##+\s*인터랙션|\*\*인터랙션\*\*", line, re.IGNORECASE):
                in_inter = True
            continue
        if not line.strip():
            continue
        # bullet 끝났음을 다음 굵은 헤더(**) 또는 다음 ## 으로 판단
        if line.startswith("**") or line.startswith("##") or line.startswith("---"):
            in_inter = False
            continue
        # bullet 추출
        m = re.match(r"^[\-\*]\s+(.+)", line)
        if m:
            body = m.group(1)
            # "<요소> <이벤트> → <결과>" 형식이 흔함. 화살표 앞부분만 사용
            label = body.split("→")[0]
            # 백틱·HTML 코드 제거
            label = re.sub(r"`[^`]+`", "", label)
            label = re.sub(r"<[^>]+>", "", label)
            label = re.sub(r"\s+", " ", label).strip(" .·:")
            if label and len(label) >= 3:
                add(label)

    # 4) 비고의 [정책]/[제약]/[접근성]/[세션]/[타이밍] 마커
    for m in re.finditer(r"\[(정책|제약|접근성|세션|타이밍|보안|성능)\]\s*([^\n]{5,80})", md_text):
        marker, body = m.group(1), m.group(2)
        # 첫 문장만
        body = re.split(r"[.。]", body)[0].strip(" ·-")
        if body:
            add(f"[{marker}] {body}")

    # 한도 잘라내기 — 가장 중요한 것 먼저 (위에서 이미 우선순위 순)
    if len(minors) > max_minors:
        minors = minors[:max_minors]
    return minors


# ── 계산·표시 정확성 패턴 (extract_minors_from_screen_md + checklist 공유) ──
# 두 함수가 같은 패턴 셋을 사용하도록 모듈 상수로 분리.
CALC_DOMAIN_PATTERNS_C = [
    # (감지 정규식, 라벨 — 소분류용 / 체크리스트용 둘 다 호환되는 단순 라벨)
    (r"\bROE\b",                                       "ROE 계산 정확성"),
    (r"\bPnL\b",                                       "PnL 계산 정확성"),
    (r"P&L|P\\&L",                                     "P&L 계산 정확성"),
    (r"Liquidation\s*price",                           "Liquidation price 계산 정확성"),
    (r"(Liquidation|청산)\s*(distance|거리)",          "Liquidation distance 계산 정확성"),
    (r"Position\s*size",                               "Position size 계산 정확성"),
    (r"Notional\s*(value)?",                           "Notional value 계산 정확성"),
    (r"Initial\s*margin|초기\s*증거금",                "Initial margin 계산 정확성"),
    (r"Maintenance\s*margin|유지\s*증거금",            "Maintenance margin 계산 정확성"),
    (r"Margin\s*ratio",                                "Margin ratio 계산 정확성"),
    (r"\b(MMR|IMR)\b",                                 "MMR/IMR 비율 계산 정확성"),
    (r"Effective\s*leverage|실효\s*레버리지",          "Effective leverage 계산 정확성"),
    (r"Est\.\s*receive",                               "Est. receive 계산 정확성 (비율별 재계산)"),
    (r"Est\.\s*fee",                                   "Est. fee 계산 정확성"),
    (r"Maker\s*fee|maker_fee",                         "Maker fee 계산 정확성"),
    (r"Taker\s*fee|taker_fee",                         "Taker fee 계산 정확성"),
    (r"Funding\s*fee",                                 "Funding fee 계산 정확성"),
    (r"Funding\s*rate|펀딩\s*레이트",                  "Funding rate 계산 정확성"),
    (r"Mark\s*price",                                  "Mark price 갱신 정확성"),
    (r"Index\s*price",                                 "Index price 갱신 정확성"),
    (r"Last\s*price",                                  "Last price 갱신 정확성"),
    (r"\bSpread\b",                                    "Spread 계산 정확성"),
    (r"Available\s*balance|가용\s*잔고",               "Available balance 정확성"),
    (r"Total\s*equity|총\s*자산",                      "Total equity 정확성"),
    (r"Locked\s*margin",                               "Locked margin 정확성"),
    (r"Free\s*margin",                                 "Free margin 정확성"),
    (r"Total\s*Unrealized\s*P&?L",                     "Total Unrealized P&L 정확성"),
    (r"Daily\s*(PnL|ROI)|일일\s*(수익|손익)",          "Daily PnL/ROI 정확성"),
    (r"\bSlippage\b",                                  "Slippage 계산 정확성"),
    (r"Filled\s*(amount|qty)",                         "Filled amount 정확성"),
    (r"(Average|Avg)\s*price",                         "Avg price 계산 정확성"),
    (r"Δ\s*[%％]",                                     "Δ% 산출 정확성"),
]

CALC_GENERAL_CATEGORIES_C_PLUS = [
    ("정렬 동작 정확성 (asc/desc 토글, 기본 정렬)", [
        r"asc\s*↔\s*desc|asc\s*/\s*desc|toggle.*Sort\b|정렬\s*토글",
        r"기본\s*정렬|default\s*sort\b",
        r"sort\s*indicator|정렬\s*[표기지표]",
    ]),
    ("표시 형식 변환 정확성 (K/M/B 단위, 통화 포맷)", [
        r"\$[0-9]+\.?[0-9]*\s*[KMB]\b",
        r"_format\w+|format\w*Volume|format\w*Price|format\w*Number",
        r"K\s*/\s*M\s*/\s*B|단위\s*변환",
    ]),
    ("색상 분기 정확성 (양수·음수·danger 임계)", [
        r"양수[^\n]{0,15}(초록|green|accent-(primary|buy))",
        r"음수[^\n]{0,15}(빨강|red|accent-sell|sell)",
        r"위험|danger|경고[^\n]{0,15}(빨강|red|색상)",
        r"≤\s*\d+\s*[%％]|≥\s*\d+\s*[%％]",
    ]),
    ("검색·필터 매칭 정확성", [
        r"filterList|검색.*(매칭|일치|결과)|search.*(match|filter)",
        r"부분\s*일치|prefix\s*match|substring",
    ]),
    ("실시간 갱신 주기 정확성 (tick rate, throttle)", [
        r"초당\s*\d+(\s*[~∼\-]\s*\d+)?\s*회",
        r"\d+\s*ms\s*(throttle|debounce|간격)",
        r"tick\b.*수신|WebSocket.*tick",
    ]),
    ("시간 기반 트리거 정확성 (타임아웃·카운트다운·백오프)", [
        r"\d+\s*초\s*(후|동안|이내|타임아웃|쿨다운|countdown)",
        r"\d+\s*분\s*(후|동안|이내|타임아웃)",
        r"카운트다운|countdown|backoff",
    ]),
]


def _add_calc_accuracy_minors(md_text: str, add_fn, domain_max: int = 8, general_max: int = 6) -> None:
    """계산·표시 정확성 항목을 add_fn 으로 등록.
    extract_minors_from_screen_md 안에서 상태 표 직후 호출하면 우선순위 가장 높음.
    domain_max: 코인 선물 도메인 패턴 최대 갯수 (한 화면당)
    general_max: 일반 카테고리 최대 갯수 (한 화면당)
    """
    # C 도메인 (코인 선물)
    domain_count = 0
    for pat, label in CALC_DOMAIN_PATTERNS_C:
        if re.search(pat, md_text, re.IGNORECASE):
            add_fn(f"계산 정확성 — {label.replace(' 계산 정확성', '').replace(' 정확성', '')}")
            domain_count += 1
            if domain_count >= domain_max:
                break

    # C+ 일반 카테고리
    general_count = 0
    for label, patterns in CALC_GENERAL_CATEGORIES_C_PLUS:
        for pat in patterns:
            if re.search(pat, md_text, re.IGNORECASE):
                add_fn(label)
                general_count += 1
                break  # 카테고리당 1번만 (다른 패턴 매칭은 카테고리 1개로 통합)
        if general_count >= general_max:
            break

    # 자연어 폴백
    if re.search(r"(계산|산출)\s*(정확성|정확도|검증)", md_text):
        add_fn("본문 명시 계산 정확성 검증")
    if re.search(r"실시간\s*(갱신|업데이트|계산|반영)", md_text):
        add_fn("실시간 데이터 갱신 동작 검증")


def build_classification_from_screen_list(screen_rows: list[dict], project_name: str,
                                            screens_meta: list[dict] | None = None) -> str:
    """parse_screen_list_table() 의 결과를 분류표 마크다운(extract_domains() 가 읽는 형식)으로 변환.
    LLM 호출 없이 바로 Human Gate 로 갈 수 있는 분류표를 생성한다.

    구조: 대분류(영역) > 중분류(화면) > 소분류(세부 시나리오)
      - 중분류 = 화면 (SCR-XXX) — 사용자가 자주 단위로 인식하는 레벨
      - 소분류 = SCR md 본문에서 규칙 기반 추출한 시나리오 (extract_minors_from_screen_md)

    Args:
        screen_rows: parse_screen_list_table() 결과 (id/name/major/middle 등)
        screens_meta: parse_screen_md() 결과 list — 있으면 본문에서 소분류 추출.
                      없으면 폴백(소분류 = '기본 동작 검증'만).
    """
    # 화면 ID → meta 매핑
    meta_by_id = {sc["id"]: sc for sc in (screens_meta or [])}

    # 대분류 → [화면 행 list]  (중분류는 화면 자체로 매핑)
    from collections import defaultdict
    tree: dict[str, list[dict]] = defaultdict(list)
    for r in screen_rows:
        major = r["major"] or "공통"
        tree[major].append(r)

    lines = [f"# TC 분류표 — {project_name}", ""]
    for major, screens in tree.items():
        lines.append(f"## 대분류: {major}")
        lines.append("")
        for s in screens:
            # 중분류 = 화면명만 (화면 코드는 별도 '화면 코드' 컬럼에서 보여짐 → 중복 제거).
            # 화면명이 비어있는 폴백 케이스에서만 ID 를 사용.
            screen_label = (s["name"] or s["id"]).strip()
            lines.append(f"### 중분류: {screen_label}")
            # 메타 주석 — build_excel 의 대분류 검증용 SCR 매핑 ground truth.
            # 사용자에게 보이는 분류표 시각에는 영향 없음 (HTML 주석).
            lines.append(f"<!-- SCR: {s['id']} -->")
            lines.append("")
            lines.append("#### 소분류")

            # 소분류 = SCR md 본문에서 추출 (규칙 기반)
            sc_meta = meta_by_id.get(s["id"])
            minors = []
            if sc_meta and sc_meta.get("raw"):
                minors = extract_minors_from_screen_md(sc_meta["raw"])

            if minors:
                for label in minors:
                    lines.append(f"- {label}")
            else:
                # 폴백 — 본문이 없거나 파싱 실패. 최소 1개라도 시드.
                desc_short = (s.get("desc") or "")[:60]
                lines.append(f"- 기본 UI 표시 및 동작 검증{(': ' + desc_short) if desc_short else ''}")

            lines.append("")
    return "\n".join(lines)


def parse_screen_md(md_path: Path) -> dict:
    """SCR-XXX.md 한 파일을 파싱해 메타 정보 추출.
    반환: {
      "id": "SCR-001",
      "title": "Login",
      "group": "로그인_공통",
      "description": "...",
      "states": [{"case":"Normal","cond":"...","desc":"..."}, ...],
      "ref_keywords": ["error.network", ...],   # cross-ref 감지용
      "raw": <원문 전체>
    }
    """
    text = md_path.read_text(encoding="utf-8")
    # ID — 파일명 우선, 본문 H1 보조
    m = re.match(r"^(SCR[-_]?\w+)", md_path.stem, re.IGNORECASE)
    sid = m.group(1).upper().replace("_", "-") if m else md_path.stem.upper()

    # H1 제목
    title = ""
    h1 = re.search(r"^#\s+([^\n]+)", text, re.MULTILINE)
    if h1:
        # "SCR-001: Login" → "Login"
        t = h1.group(1).strip()
        t = re.sub(r"^SCR[-_]?\w+\s*[:\-]\s*", "", t, flags=re.IGNORECASE)
        title = t.strip()

    # 그룹/설명 — **그룹**: xxx, **설명**: xxx 패턴
    group = ""
    g = re.search(r"\*\*그룹\*\*\s*[:：]\s*([^\n]+)", text)
    if g:
        group = g.group(1).strip()
    desc = ""
    d = re.search(r"\*\*설명\*\*\s*[:：]\s*([^\n]+(?:\n(?![*#\-|]).+)*)", text)
    if d:
        desc = d.group(1).strip()[:300]

    # 상태(Status) 표 — | 케이스 | 조건 | 설명 / 동작 |
    states: list[dict] = []
    in_status_tbl = False
    status_headers: list[str] = []
    for line in text.splitlines():
        if not in_status_tbl:
            # "**상태 (Status)**" 또는 "## 상태" 등 키워드 뒤의 첫 표
            if re.search(r"상태\s*\(\s*Status\s*\)|^##+\s*상태", line):
                in_status_tbl = "header_pending"
            continue
        if in_status_tbl == "header_pending":
            if line.strip().startswith("|") and ("케이스" in line or "case" in line.lower()):
                cells = [c.strip().lower() for c in line.strip("|").split("|")]
                status_headers = cells
                in_status_tbl = "rows"
            continue
        if in_status_tbl == "rows":
            if line.strip().startswith("|---"):
                continue
            if not line.strip().startswith("|"):
                in_status_tbl = False
                continue
            cells = [c.strip() for c in line.strip("|").split("|")]
            if len(cells) < 2:
                continue
            row = {}
            for i, h in enumerate(status_headers):
                if i < len(cells):
                    row[h] = cells[i]
            states.append({
                "case": row.get("케이스") or row.get("case") or "",
                "cond": row.get("조건") or row.get("condition") or "",
                "desc": (row.get("설명 / 동작") or row.get("설명") or row.get("동작")
                         or row.get("description") or ""),
            })

    # cross-ref 키워드 감지 (정책/디자인 자동 inject 용)
    ref_kw = []
    if re.search(r"error\.\w+|에러|네트워크|타임아웃|세션", text, re.IGNORECASE):
        ref_kw.append("error")
    if re.search(r"세션|토큰|로그아웃|만료", text):
        ref_kw.append("session")
    if re.search(r"toast|modal|alert|inline error|full-?screen error", text, re.IGNORECASE):
        ref_kw.append("error_ui")
    if re.search(r"empty\s*state|빈\s*상태", text, re.IGNORECASE):
        ref_kw.append("empty_state")
    if re.search(r"loading|로딩|skeleton", text, re.IGNORECASE):
        ref_kw.append("loading")

    return {
        "id": sid,
        "title": title,
        "group": group,
        "description": desc,
        "states": states,
        "ref_keywords": ref_kw,
        "raw": text,
    }


def _split_checklist_report(tc_md: str) -> tuple[str, str]:
    """AI 응답의 말미에서 '### 체크리스트 처리 결과' 섹션을 분리.
    반환: (TC 본문, 체크리스트 보고 텍스트)
    못 찾으면 (원본, "") 반환.
    """
    # ### 체크리스트 처리 결과 헤더 찾기 (대소문자/이모지 변형 허용)
    m = re.search(r"\n#{1,4}\s*(?:체크리스트\s*처리\s*결과|체크리스트\s*검증|Checklist\s*(?:Verification|Result))\s*\n",
                  tc_md, re.IGNORECASE)
    if not m:
        return tc_md, ""
    body = tc_md[:m.start()].rstrip()
    report = tc_md[m.start():].strip()
    return body, report


def _extract_screen_classification_section(classification: str, screen_id: str, screen_name: str) -> str:
    """전체 분류표 마크다운에서 해당 SCR 의 중분류 섹션만 추출 (A-1 분류표 동적 축소).

    build_classification_from_screen_list() 결과는 다음 구조:
      ## 대분류: Trade
      ### 중분류: Order Confirm     ← 화면명 = screen_name
      #### 소분류
      - ...

    찾기 우선순위: screen_name 일치 → screen_id 일치.
    못 찾으면 빈 문자열 반환 (호출부에서 폴백 처리).
    """
    lines = classification.splitlines()
    out: list[str] = []
    in_target = False
    current_major = ""
    for line in lines:
        if line.startswith("## 대분류:"):
            current_major = line
            in_target = False
            continue
        if line.startswith("### 중분류:"):
            label = line[len("### 중분류:"):].strip()
            # 화면명 또는 ID 매칭
            if screen_name and label == screen_name.strip():
                in_target = True
                if current_major:
                    out.append(current_major)
                out.append(line)
                continue
            if screen_id and screen_id in label:
                in_target = True
                if current_major:
                    out.append(current_major)
                out.append(line)
                continue
            in_target = False
            continue
        if line.startswith("## ") or line.startswith("### "):
            in_target = False
            continue
        if in_target:
            out.append(line)
    return "\n".join(out).strip()


def _build_checklist_from_screen_meta(screen_meta: dict) -> str:
    """C-2 체크리스트 강제 — 화면 본문의 모든 상태/인터랙션/비고 마커에 ID 부여.
    AI 가 응답 끝에 각 항목별 [✓ 처리 / ✗ 누락 + 사유] 보고하도록 함.
    """
    items: list[str] = []
    counter_s = 0
    counter_i = 0
    counter_r = 0

    # 상태 케이스 — S1, S2, ...
    for st in screen_meta.get("states", []):
        counter_s += 1
        case = st.get("case", "").strip() or f"상태{counter_s}"
        items.append(f"S{counter_s}. {case} 상태")

    # 인터랙션 — I1, I2, ... — raw 에서 [dev] 인터랙션 bullet 발췌 (extract_minors 와 같은 패턴)
    in_inter = False
    for line in screen_meta.get("raw", "").splitlines():
        if not in_inter:
            if re.search(r"\[dev\]\s*인터랙션|\*\*인터랙션\*\*", line, re.IGNORECASE):
                in_inter = True
            continue
        if line.startswith("**") or line.startswith("##") or line.startswith("---"):
            in_inter = False
            continue
        m = re.match(r"^[\-\*]\s+(.+)", line)
        if m:
            body = m.group(1).split("→")[0]
            body = re.sub(r"`[^`]+`", "", body)
            body = re.sub(r"<[^>]+>", "", body)
            body = re.sub(r"\s+", " ", body).strip(" .·:")
            if body and len(body) >= 3:
                counter_i += 1
                items.append(f"I{counter_i}. {body[:50]}")
                if counter_i >= 10:  # 너무 많으면 자름
                    break

    # 비고 마커 — R1, R2, ...
    for m in re.finditer(r"\[(정책|제약|접근성|세션|타이밍|보안|성능)\]\s*([^\n]{5,80})",
                         screen_meta.get("raw", "")):
        marker, body = m.group(1), m.group(2)
        body = re.split(r"[.。]", body)[0].strip(" ·-")
        if body:
            counter_r += 1
            items.append(f"R{counter_r}. [{marker}] {body[:45]}")
            if counter_r >= 6:
                break

    # 계산·표시 정확성 — C1, C2, ... (extract_minors 와 동일한 모듈 상수 재사용)
    # 응답 말미 자체 검증 강제 → AI 가 "표시된다" 한 줄로 흡수하지 못하게 함.
    counter_c = 0
    raw_text = screen_meta.get("raw", "")

    # C 도메인 (코인 선물) — 한 화면당 최대 12개
    for pat, label in CALC_DOMAIN_PATTERNS_C:
        if counter_c >= 12:
            break
        if re.search(pat, raw_text, re.IGNORECASE):
            counter_c += 1
            items.append(f"C{counter_c}. {label}")

    # C+ 일반 6 카테고리 (각 카테고리당 1번)
    for label, patterns in CALC_GENERAL_CATEGORIES_C_PLUS:
        for pat in patterns:
            if re.search(pat, raw_text, re.IGNORECASE):
                counter_c += 1
                items.append(f"C{counter_c}. {label}")
                break

    # 자연어 폴백
    if re.search(r"(계산|산출)\s*(정확성|정확도|검증)", raw_text):
        counter_c += 1
        items.append(f"C{counter_c}. 본문 명시 계산 항목 정확성 검증")
    if re.search(r"실시간\s*(갱신|업데이트|계산|반영)", raw_text):
        counter_c += 1
        items.append(f"C{counter_c}. 실시간 데이터 갱신 동작 검증")

    if not items:
        return ""
    return "\n".join(items)


def build_screen_user_prompt(screen_meta: dict, project_name: str, project_code: str,
                              suite_code: str, starting_seq: int,
                              policy_excerpts: list[str] = None,
                              design_excerpts: list[str] = None,
                              screen_classification_section: str = "") -> str:
    """화면별 1:1 호출용 user prompt.

    v0.10.x 변경:
    - A-1: 전체 분류표 대신 해당 SCR 의 중분류 섹션만 주입 (다른 화면 정보 격리)
    - C-2: 화면 본문에서 추출한 체크리스트 + 응답 말미 자체 검증 강제
    """
    seq_str = f"{starting_seq:03d}"
    next_str = f"{starting_seq+1:03d}"
    example_id = f"{project_code}-{suite_code}-{seq_str}"

    # A-1: 분류표 — 해당 화면 섹션만
    classification_block = ""
    if screen_classification_section:
        classification_block = (
            f"## 분류표 (이 화면만 — 사용자 승인된 최종 진실 소스)\n"
            f"{screen_classification_section}\n"
        )

    states_block = ""
    if screen_meta["states"]:
        lines = ["## 이 화면의 상태 케이스 (각각 TC 후보)"]
        for st in screen_meta["states"]:
            lines.append(f"- **{st['case']}**: 조건={st['cond']} / 동작={st['desc']}")
        states_block = "\n".join(lines) + "\n"

    refs_block = ""
    if policy_excerpts or design_excerpts:
        parts = ["## 이 화면에 적용되는 정책·디자인 발췌 (자동 cross-ref)"]
        if policy_excerpts:
            parts.append("### 관련 정책")
            for x in policy_excerpts:
                parts.append(x)
        if design_excerpts:
            parts.append("### 관련 디자인")
            for x in design_excerpts:
                parts.append(x)
        refs_block = "\n".join(parts) + "\n"

    # C-2: 체크리스트
    checklist = _build_checklist_from_screen_meta(screen_meta)
    checklist_block = ""
    if checklist:
        checklist_block = (
            "\n## 체크리스트 — 반드시 모두 처리 (누락 시 응답 끝에 사유 명시)\n\n"
            "다음 항목 각각에 대해 최소 1개의 TC 를 작성하세요. 항목은 화면 명세에서 자동 추출됨.\n\n"
            f"{checklist}\n"
        )

    selfcheck = (
        "\n## 응답 말미 자체 검증 (필수)\n\n"
        "TC 작성 완료 후, 응답 마지막에 아래 형식으로 체크리스트 처리 결과를 반드시 명시하세요:\n\n"
        "```\n"
        "### 체크리스트 처리 결과\n"
        "- S1: ✓ TC SC-LIT-001, SC-LIT-005\n"
        "- S2: ✓ TC SC-LIT-002\n"
        "- I1: ✗ 누락 / 사유: 화면 명세에 충분한 정보 없음\n"
        "- ...\n"
        "```\n"
        "이 자체 검증 표는 후처리에서 분리되며 TC 본문에 영향 없습니다. 누락이 발견되면 솔직히 명시하세요.\n"
    )

    return f"""프로젝트: {project_name}
화면: {screen_meta['id']} — {screen_meta['title']}
그룹: {screen_meta['group']}

⚠️ TC ID 형식: `{example_id}`, `{project_code}-{suite_code}-{next_str}`, ... (시작 번호 `{seq_str}`부터 연속 증가)
⚠️ 시스템 프롬프트의 'TC 생성 카테고리 5가지(UI/주요/예외/에러/연결시나리오)'와 비율(18/32/25/18/5~10)을 따르세요.
   카테고리 3(예외)·4(에러)는 반드시 포함합니다 (Positive 만으로 구성 금지).
   카테고리 5(연결 시나리오)는 시작점/종착지 화면에만 1~3개 — 그 외는 skip.
⚠️ **대분류 값은 아래 분류표의 `## 대분류:` 헤더 그대로 사용** — 임의 prefix/접두자/번호 추가 절대 금지
   (예: 분류표가 `Trade` 면 모든 TC 의 대분류 컬럼도 정확히 `Trade`. `02.Trade`, `[Trade]`,
   `Group-Trade` 같이 변형하면 시트가 잘못 분리됨. 원칙 G 통합 TC 도 동일.)

{classification_block}
## 화면 명세 (전문 — 이 문서의 내용에서만 TC 작성)

{screen_meta['raw']}

{states_block}
{refs_block}
{checklist_block}
{selfcheck}

위 화면의 모든 상태 케이스·인터랙션·비고 마커를 빠짐없이 TC 로 변환하세요.
- Normal 상태 → 카테고리 1 (UI/UX) + 카테고리 2 (주요 기능) Positive TC
- Error.* 상태 → 카테고리 3 (예외) Negative + 카테고리 4 (에러) Edge TC
- 비고의 [정책]/[제약]/[접근성]/[세션]/[타이밍]/[dev] 마커는 각각 별도 TC 후보
- 인터랙션의 모든 진입 경로/탭 동작/네비게이션 → 각각 별도 TC
- 화면 명세의 "**진입**" 메타가 "(없음/온보딩/탭바/외부)"이거나 그룹의 첫/마지막 화면이면
  → 카테고리 5 (연결 시나리오) 1~3개 추가. 다음 단계 1~2개를 거치는 E2E happy path + 핵심 분기 또는 round-trip
- 누락 없이 가능한 모든 시나리오를 작성. 응답이 길어지더라도 끝까지 작성하세요.

## ⚠️ 계산 정확성 별도 TC 의무 (자주 누락되는 항목)

화면 명세에 **수치 계산·공식·환산** 이 등장하면 **고유 정확성 TC를 별도로** 작성하세요.
"표시된다" 같은 한 줄 검증으로 흡수하지 말고 독립 TC 로 분리:

대표 패턴:
- `ROE = ...`, `PnL = ...`, `Liquidation price = ...`, `Position size = Margin × Leverage` 같은 공식
- `Est. receive = (포지션 × 비율%) − 수수료` 같은 추정값 재계산
- `실시간 갱신`, `Δ% 산출`, `백분율 변환` 등 동적 계산

각 공식별 TC 형식 예시:
```
### {{ID}} — ROE 계산 정확성 확인

| 분류 | Positive |
| 우선순위 | Medium |

**사전 조건**
1. 포지션이 1개 이상 존재하는 상태
2. 입력 증거금 100 USDT, Leverage 10x 인 상태

**테스트 단계**
포지션의 ROE 표시값과 (현재 PnL / 투입 증거금 × 100) 공식 결과를 비교한다.

**예상 결과**
- 표시된 ROE % 값이 공식 계산 결과와 ±0.01% 오차 이내로 일치한다.
- 가격 변동 시 ROE % 가 실시간으로 재계산된다.
```

이런 정확성 TC 는 우선순위 Medium 이라도 반드시 포함하세요.

## ⚠️ 중복 통합 원칙 (반드시 준수)

**같은 화면 안에서 동일한 검증 결과를 갖는 트리거가 여러 개라면 1개의 TC 로 통합** 하세요.
대표 예시:
- "Cancel 버튼 탭" 과 "배경 오버레이 탭" 둘 다 → "시트 닫힘 + SCR-102 복귀"
  → ❌ 별도 TC 2개로 만들지 말고, ✅ TC 1개로 통합 + 테스트 단계에 두 동작 모두 명시

통합 형식 예시:
```
### {{ID}} — Order Confirm 시트 닫기 (Cancel 버튼 또는 배경 오버레이 탭)

**테스트 단계**
다음 중 하나를 수행한다:
1. Cancel 버튼을 탭한다
2. 시트 외부 배경 오버레이를 탭한다

**예상 결과**
- 시트가 닫히고 SCR-102 Lite Trade 화면으로 복귀한다
- 두 진입 모두 동일한 결과 (AppState 변경 없음)
```

판단 기준:
- 검증 결과(예상 결과)가 100% 동일 → 통합 권장
- 검증 결과가 일부 다르거나 부수 효과(로그·이벤트 등)가 다름 → 별도 TC 유지
- 사전 조건이 다른 경우 → 별도 TC 유지 (사전 조건이 다르면 같은 결과여도 별도 검증 필요)

이 통합 원칙은 양을 줄이는 것이 목적이 아니라 **테스터 가독성과 유지보수성** 을 위한 것입니다.
의미 있는 검증이 사라지면 안 되고, 단지 "동일 검증의 트리거 중복" 만 합치는 것입니다.
"""


def extract_policy_excerpts(policy_full_text: str, ref_keywords: list[str]) -> list[str]:
    """정책 전문에서 화면이 참조할 섹션만 발췌. 키워드별 최대 800자.
    너무 길면 화면별 호출이 무거워지므로 압축해서 user prompt 에 넣는다.
    """
    if not ref_keywords or not policy_full_text:
        return []

    excerpts = []
    sections = re.split(r"\n(?=#{1,3}\s)", policy_full_text)
    seen_titles = set()
    for kw in ref_keywords:
        kw_pat = {
            "error": r"error|에러|패턴",
            "session": r"세션|토큰|만료|로그아웃",
            "error_ui": r"toast|modal|alert|inline|full-?screen",
            "empty_state": r"empty|빈\s*상태",
            "loading": r"loading|로딩|skeleton",
        }.get(kw, kw)
        for sec in sections:
            head = sec.split("\n", 1)[0]
            if head in seen_titles:
                continue
            if re.search(kw_pat, sec, re.IGNORECASE):
                excerpt = sec.strip()[:800]
                excerpts.append(excerpt)
                seen_titles.add(head)
                if len(excerpts) >= 4:  # 너무 많으면 잘라냄
                    return excerpts
    return excerpts


def parse_scr_filter_from_focus(focus_area: str, available_scrs: set[str] | None = None) -> set[str] | None:
    """focus_area 텍스트에서 SCR ID 패턴을 관대하게 추출해 필터로 사용.

    지원 입력 형식:
      - 명시: "SCR-104", "scr-104", "scr_104", "SCR104", "SCR 104"
      - 일괄 (앵커 모드): "SCR-102, 104, 106, 116" → 첫 SCR 발견 후 따라오는 숫자도 SCR ID 로
      - 범위: "SCR-102~116", "SCR-102 to 116", "SCR-102 - 116"
        → available_scrs 가 주어지면 그 범위 내 실제 존재 SCR 만 매칭
      - 영문 접미: "SCR-007A" 그대로 인식
      - 자연어 섞임: "주문 확인 화면 (SCR-104) 만 만들어줘" OK

    Args:
        focus_area:      사용자 입력 텍스트
        available_scrs:  폴더에서 실제 발견된 SCR ID 집합 (예: {"SCR-001","SCR-102",...})
                         주어지면 범위 표기·숫자 흡수의 유효성 검증에 사용.
                         None 이면 명시적으로 보인 ID 만 신뢰.

    Returns:
        - set[str]: 매칭된 SCR ID 들 (예: {"SCR-102","SCR-104"})
        - None:     SCR 패턴이 하나도 없으면 (필터 미적용 = 전체 처리)
    """
    if not focus_area or not focus_area.strip():
        return None

    text = focus_area
    # 1) 명시적 SCR-XXX 매칭 (영문 접미 가능)
    explicit = re.findall(r"SCR[\s\-_]*([0-9]+[A-Z]?)", text, re.IGNORECASE)
    explicit_norm = {f"SCR-{m.upper()}" for m in explicit}

    # 앵커가 하나도 없으면 SCR 모드 아님 → None
    if not explicit_norm:
        return None

    # available_scrs 가 주어지면 명시 ID 도 폴더 검증 (잘못 입력한 ID 흡수 방지)
    if available_scrs:
        result = explicit_norm & available_scrs
    else:
        result = set(explicit_norm)

    # 2) 범위 표기 지원: "SCR-102~116", "102-116", "102 to 116"
    #    숫자~숫자 / 숫자-숫자 / 숫자 to 숫자 패턴
    range_matches = re.findall(
        r"(?:SCR[\s\-_]*)?([0-9]{2,4})\s*(?:~|to|\-\-|—|–)\s*([0-9]{2,4})",
        text, re.IGNORECASE
    )
    if range_matches and available_scrs:
        for start_s, end_s in range_matches:
            start, end = int(start_s), int(end_s)
            if start > end:
                start, end = end, start
            # available_scrs 안에서 숫자 부분이 [start, end] 범위에 들어가는 것 추가
            for sid in available_scrs:
                m = re.match(r"^SCR-([0-9]+)([A-Z]?)$", sid)
                if not m:
                    continue
                num = int(m.group(1))
                if start <= num <= end:
                    result.add(sid)

    # 3) 앵커 모드 — 텍스트에 SCR 가 보였으니 뒤따르는 "그냥 숫자"도 SCR ID 후보로
    #    예: "SCR-102, 104, 106, 116" → 102, 104, 106, 116 모두
    #    단, available_scrs 가 주어진 경우에만 (미지의 숫자 흡수 방지).
    if available_scrs:
        # 텍스트 안의 모든 2~4자리 숫자 토큰 (앞뒤가 단어 경계)
        # 단, 범위 표기에서 이미 처리한 숫자는 흡수해도 무방 (set 이라 중복 제거됨)
        all_nums = re.findall(r"\b([0-9]{2,4})([A-Z]?)\b", text)
        for num_s, suffix in all_nums:
            num = int(num_s)
            # SCR ID 와 매칭되는 형식으로 후보 생성
            # available_scrs 에는 보통 "SCR-001", "SCR-102", "SCR-007A" 형식
            candidate_plain = f"SCR-{num_s}{suffix.upper()}"
            # zero-padding 변형: 102 → 102, 1 → 001 / 12 → 012 도 시도
            candidate_padded = f"SCR-{num_s.zfill(3)}{suffix.upper()}"
            for cand in (candidate_plain, candidate_padded):
                if cand in available_scrs:
                    result.add(cand)
                    break

    return result


def step_parse_structured_spec(sess: dict, folder_path: str) -> dict:
    """구조화 spec 폴더 1개를 받아 overview/policy/design/screens 로 분리하고 메타 추출.
    반환: {
      "overview_text": str,
      "policy_text":   str,    # 모든 정책 md concat
      "design_text":   str,    # 모든 디자인 md concat
      "screens":       [parse_screen_md() 결과 ...],
      "screen_rows":   [parse_screen_list_table() 결과 ...],
    }
    LLM 호출은 0회 — 순수 파싱만.
    """
    folder = Path(folder_path).expanduser()
    if not folder.exists():
        raise FileNotFoundError(f"폴더 없음: {folder}")
    if not folder.is_dir():
        raise NotADirectoryError(f"폴더가 아님: {folder}")

    push_log(sess, f"[구조화 spec] 폴더 스캔 중: {folder}")
    cls = classify_spec_files(folder)

    overview_text = cls["overview"].read_text(encoding="utf-8") if cls["overview"] else ""
    policy_text = "\n\n".join(p.read_text(encoding="utf-8") for p in cls["policy"])
    design_text = "\n\n".join(p.read_text(encoding="utf-8") for p in cls["design"])

    screens = [parse_screen_md(p) for p in cls["screens"]]
    screen_rows = parse_screen_list_table(overview_text) if overview_text else []

    push_log(sess,
        f"[구조화 spec] 분류 완료 — overview={'있음' if cls['overview'] else '없음'} / "
        f"policy={len(cls['policy'])}개 / design={len(cls['design'])}개 / "
        f"화면 md={len(cls['screens'])}개 / 화면 목록 표={len(screen_rows)}행")

    return {
        "overview_text": overview_text,
        "policy_text":   policy_text,
        "design_text":   design_text,
        "screens":       screens,
        "screen_rows":   screen_rows,
        "_files":        cls,  # 디버깅용
    }


def diff_spec_folders(prev_folder: Path, new_folder: Path) -> dict:
    """이전 spec 폴더와 신규 spec 폴더를 비교해 SCR 단위로 변경분 분류.

    반환: {
      "added":     [SCR-ID, ...],   # 신규 추가된 화면
      "modified":  [SCR-ID, ...],   # 본문이 바뀐 화면
      "removed":   [SCR-ID, ...],   # 신규에서 사라진 화면
      "unchanged": [SCR-ID, ...],   # 동일 (재생성 불필요)
      "common_changed": bool,       # 정책/디자인/overview 가 바뀌었는지
    }
    """
    import hashlib

    def _hash_file(p: Path) -> str:
        return hashlib.sha256(p.read_bytes()).hexdigest()

    def _scr_map(folder: Path) -> dict[str, Path]:
        cls = classify_spec_files(folder)
        out: dict[str, Path] = {}
        for p in cls["screens"]:
            m = re.match(r"^(SCR[-_]?\w+)", p.stem, re.IGNORECASE)
            if m:
                sid = m.group(1).upper().replace("_", "-")
                out[sid] = p
        return out

    prev_scr = _scr_map(prev_folder)
    new_scr = _scr_map(new_folder)

    added = sorted(set(new_scr) - set(prev_scr))
    removed = sorted(set(prev_scr) - set(new_scr))
    modified = []
    unchanged = []
    for sid in sorted(set(new_scr) & set(prev_scr)):
        if _hash_file(prev_scr[sid]) != _hash_file(new_scr[sid]):
            modified.append(sid)
        else:
            unchanged.append(sid)

    # 공통 문서(overview/policy/design) 변경 여부
    prev_cls = classify_spec_files(prev_folder)
    new_cls = classify_spec_files(new_folder)
    common_changed = False
    for key in ("overview",):
        a = prev_cls.get(key)
        b = new_cls.get(key)
        if (a is None) != (b is None) or (a and b and _hash_file(a) != _hash_file(b)):
            common_changed = True
            break
    if not common_changed:
        prev_common = sorted([p.name for p in prev_cls["policy"] + prev_cls["design"]])
        new_common = sorted([p.name for p in new_cls["policy"] + new_cls["design"]])
        if prev_common != new_common:
            common_changed = True
        else:
            for plist_a, plist_b in [
                (sorted(prev_cls["policy"], key=lambda p: p.name),
                 sorted(new_cls["policy"], key=lambda p: p.name)),
                (sorted(prev_cls["design"], key=lambda p: p.name),
                 sorted(new_cls["design"], key=lambda p: p.name)),
            ]:
                for a, b in zip(plist_a, plist_b):
                    if _hash_file(a) != _hash_file(b):
                        common_changed = True
                        break
                if common_changed:
                    break

    return {
        "added": added,
        "modified": modified,
        "removed": removed,
        "unchanged": unchanged,
        "common_changed": common_changed,
    }


# ── TC Update 분석 — SCR 그룹화 + 영속 캐시 ─────────────────────────────────

# 그룹 묶음 크기 (5개씩)
DIFF_GROUP_SIZE = 5

# 캐시 디렉토리 (tc-ui 프로젝트 루트 기준)
_DIFF_CACHE_DIR = BASE_DIR / ".update_cache"
_DIFF_CACHE_DIR.mkdir(exist_ok=True)


def _group_scr_changes(diff: dict) -> list:
    """SCR diff 를 종류별 + 5개 단위 그룹으로 묶음.

    Args:
      diff: {added, modified, removed, unchanged, ...}

    Returns: [{type, scrs, index, label}, ...] — index 0..N
      type: 'added' | 'modified' | 'removed'
      scrs: ['SCR-222', ...] (최대 5개)
      label: 사람이 읽을 한 줄 설명
    """
    groups = []
    idx = 0

    def _chunked(items: list, n: int):
        for i in range(0, len(items), n):
            yield items[i:i + n]

    # 신규 먼저
    for chunk in _chunked(list(diff.get("added", []) or []), DIFF_GROUP_SIZE):
        groups.append({
            "index": idx,
            "type": "added",
            "scrs": chunk,
            "label": f"신규 SCR {len(chunk)}개",
        })
        idx += 1

    # 그 다음 수정
    for chunk in _chunked(list(diff.get("modified", []) or []), DIFF_GROUP_SIZE):
        groups.append({
            "index": idx,
            "type": "modified",
            "scrs": chunk,
            "label": f"수정 SCR {len(chunk)}개",
        })
        idx += 1

    # 마지막 삭제
    for chunk in _chunked(list(diff.get("removed", []) or []), DIFF_GROUP_SIZE):
        groups.append({
            "index": idx,
            "type": "removed",
            "scrs": chunk,
            "label": f"삭제 SCR {len(chunk)}개",
        })
        idx += 1

    return groups


def _diff_cache_key(prev_folder: str, new_folder: str, group_index: int) -> str:
    """캐시 키 = SHA256(prev|new|group_index) 의 hex 12자리."""
    import hashlib
    raw = f"{prev_folder}|{new_folder}|{group_index}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]


def _diff_cache_path(prev_folder: str, new_folder: str, group_index: int) -> Path:
    key = _diff_cache_key(prev_folder, new_folder, group_index)
    return _DIFF_CACHE_DIR / f"{key}.json"


def _load_diff_cache(prev_folder: str, new_folder: str, group_index: int) -> dict | None:
    """캐시 hit 면 dict 반환, miss 면 None."""
    p = _diff_cache_path(prev_folder, new_folder, group_index)
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None


def _save_diff_cache(prev_folder: str, new_folder: str, group_index: int,
                       data: dict) -> None:
    """캐시 저장 — 실패해도 조용히 무시."""
    p = _diff_cache_path(prev_folder, new_folder, group_index)
    try:
        from datetime import datetime
        payload = {
            **data,
            "_cached_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "_prev_folder": prev_folder,
            "_new_folder": new_folder,
            "_group_index": group_index,
        }
        p.write_text(json.dumps(payload, ensure_ascii=False, indent=2),
                       encoding="utf-8")
    except Exception:
        pass


def _read_scr_pair(prev_folder: str, new_folder: str, scr_id: str) -> tuple:
    """주어진 SCR 의 (prev_text, new_text) 반환. 없으면 빈 문자열."""
    try:
        prev_cls = classify_spec_files(Path(prev_folder).expanduser())
        new_cls = classify_spec_files(Path(new_folder).expanduser())
    except Exception:
        return "", ""
    prev_map = {p.stem.split(".")[0]: p for p in prev_cls.get("screens", [])}
    new_map = {p.stem.split(".")[0]: p for p in new_cls.get("screens", [])}
    prev_p = prev_map.get(scr_id)
    new_p = new_map.get(scr_id)
    prev_text = prev_p.read_text(encoding="utf-8") if prev_p and prev_p.exists() else ""
    new_text = new_p.read_text(encoding="utf-8") if new_p and new_p.exists() else ""
    return prev_text, new_text


# ── SCR 단위 캐시 (그룹 캐시와 별개) ─────────────────────────────────────

def _scr_cache_path(prev_folder: str, new_folder: str, scr_id: str) -> Path:
    import hashlib
    raw = f"scr|{prev_folder}|{new_folder}|{scr_id}"
    key = hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]
    return _DIFF_CACHE_DIR / f"scr_{key}.json"


def _load_scr_cache(prev_folder: str, new_folder: str, scr_id: str) -> dict | None:
    p = _scr_cache_path(prev_folder, new_folder, scr_id)
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None


def _save_scr_cache(prev_folder: str, new_folder: str, scr_id: str, data: dict) -> None:
    p = _scr_cache_path(prev_folder, new_folder, scr_id)
    try:
        from datetime import datetime
        payload = {
            **data,
            "_cached_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "_scr_id": scr_id,
        }
        p.write_text(json.dumps(payload, ensure_ascii=False, indent=2),
                       encoding="utf-8")
    except Exception:
        pass


def step_analyze_scr_one(prev_folder: str, new_folder: str, scr_id: str,
                            change_type: str) -> dict:
    """1개 SCR 에 대한 AI 분석.

    change_type: 'added' | 'modified' | 'removed'
    Returns: {ok, summary, scr_id, change_type, from_cache}
    """
    prev_text, new_text = _read_scr_pair(prev_folder, new_folder, scr_id)

    if change_type == "added":
        body = f"### {scr_id} (신규)\n\n```\n{new_text[:3500]}\n```"
    elif change_type == "removed":
        body = f"### {scr_id} (삭제)\n\n```\n{prev_text[:2500]}\n```"
    else:  # modified
        body = (
            f"### {scr_id} (수정)\n\n"
            f"#### 이전\n```\n{prev_text[:1800]}\n```\n\n"
            f"#### 신규\n```\n{new_text[:1800]}\n```"
        )

    type_label = {
        "added": "신규 SCR",
        "modified": "수정 SCR",
        "removed": "삭제 SCR",
    }.get(change_type, change_type)

    system_prompt = f"""당신은 QA 도메인 전문가입니다. {type_label} 1개의 변경 사항을
간결한 한국어 보고서로 정리합니다.

## 출력 형식 (마크다운)
```
### {scr_id}
- **변경 핵심**: 한 줄
- **TC 영향**: 어떤 종류의 TC 가 영향받는지 (수정/신규/삭제)
- **권장 액션**: 1-2 문장
```

원칙:
- 본문에서만 추출. 추측 금지.
- TC ID 자체는 모름 — "영향 종류" 만 추정.
- 디자인/문구만 변경된 경우 명시.
- 간결하게."""

    user_prompt = f"## {type_label} 본문\n\n{body}"

    try:
        summary = call_claude(system_prompt, user_prompt, max_tokens=1500)
    except Exception as e:
        return {"ok": False, "error": f"AI 호출 실패: {e}",
                  "scr_id": scr_id, "change_type": change_type}

    return {
        "ok": True,
        "summary": summary,
        "scr_id": scr_id,
        "change_type": change_type,
    }


def step_analyze_group(prev_folder: str, new_folder: str, group: dict) -> dict:
    """1개 그룹 (최대 5개 SCR) 에 대한 AI 요약 분석.

    Args:
      group: {index, type, scrs, label}

    Returns: {ok, summary, scr_details: [{scr_id, status, ...}], group}
    """
    gtype = group.get("type", "")
    scrs = group.get("scrs", [])

    # 각 SCR 의 본문 발췌
    blocks = []
    for scr_id in scrs:
        prev_text, new_text = _read_scr_pair(prev_folder, new_folder, scr_id)
        if gtype == "added":
            blocks.append(f"### {scr_id} (신규)\n\n```\n{new_text[:3500]}\n```")
        elif gtype == "removed":
            blocks.append(f"### {scr_id} (삭제)\n\n```\n{prev_text[:2500]}\n```")
        else:  # modified
            blocks.append(
                f"### {scr_id} (수정)\n\n"
                f"#### 이전\n```\n{prev_text[:1800]}\n```\n\n"
                f"#### 신규\n```\n{new_text[:1800]}\n```"
            )

    type_label = {
        "added": "신규 SCR",
        "modified": "수정 SCR",
        "removed": "삭제 SCR",
    }.get(gtype, gtype)

    system_prompt = f"""당신은 QA 도메인 전문가입니다. 다음 {type_label} 그룹의 변경 사항을
간결한 한국어 보고서로 정리합니다.

## 출력 형식 (마크다운)
```
## {type_label} ({len(scrs)}개)

### SCR-XXX
- **변경 핵심**: 한 줄
- **TC 영향**: 어떤 종류의 TC 가 영향받는지 (수정/신규/삭제)
- **권장 액션**: 1-2 문장

(각 SCR 마다 반복)
```

원칙:
- 본문에서만 추출. 추측 금지.
- TC ID 자체는 모름 — "영향 종류" 만 추정.
- 보수적 — 디자인/문구만 변경되었으면 그렇게 명시.
- 간결하게."""

    user_prompt = f"## {type_label} 본문 ({len(scrs)}개)\n\n" + "\n\n---\n\n".join(blocks)

    try:
        summary = call_claude(system_prompt, user_prompt, max_tokens=4000)
    except Exception as e:
        return {"ok": False, "error": f"AI 호출 실패: {e}", "group": group}

    return {
        "ok": True,
        "summary": summary,
        "group": group,
    }


def step_analyze_spec_diff(prev_folder: str, new_folder: str,
                            existing_tc_url: str = "") -> dict:
    """1단계 — spec 폴더 diff 분석 + AI 요약 보고서 생성.

    Args:
        prev_folder: 이전 기획서 spec 폴더 경로
        new_folder: 새 기획서 spec 폴더 경로
        existing_tc_url: 기존 TC Google Sheets URL (참고용 — 1단계는 아직 안 읽음)
    Returns:
        {
          "diff": {added, modified, removed, unchanged, common_changed},
          "summary": str (AI 가 요약한 변경 사항 마크다운),
          "scr_changes": [{scr_id, status, change_summary, impact_estimate}],
          "common_doc_changes": [...],  # 정책/디자인 변경 요약
          "meta": {prev_folder, new_folder, analyzed_at, ...},
        }
    """
    prev_p = Path(prev_folder).expanduser()
    new_p = Path(new_folder).expanduser()
    if not prev_p.exists() or not prev_p.is_dir():
        raise RuntimeError(f"이전 폴더 없음: {prev_folder}")
    if not new_p.exists() or not new_p.is_dir():
        raise RuntimeError(f"새 폴더 없음: {new_folder}")

    # 1) 파일 단위 diff (재사용)
    diff = diff_spec_folders(prev_p, new_p)

    # 2) 각 SCR 의 변경 내용 — modified 항목들의 본문 비교
    prev_cls = classify_spec_files(prev_p)
    new_cls = classify_spec_files(new_p)
    prev_scr_map = {p.stem.split(".")[0]: p for p in prev_cls["screens"]}
    new_scr_map = {p.stem.split(".")[0]: p for p in new_cls["screens"]}

    # 3) 각 변경 항목별 본문 발췌 (AI 요약용)
    scr_changes_raw = []  # AI 호출 전 raw 데이터
    for scr_id in diff["added"]:
        scr_path = new_scr_map.get(scr_id)
        if scr_path and scr_path.exists():
            scr_changes_raw.append({
                "scr_id": scr_id, "status": "added",
                "new_content": scr_path.read_text(encoding="utf-8")[:5000],
                "prev_content": "",
            })
    for scr_id in diff["modified"]:
        prev_path = prev_scr_map.get(scr_id)
        new_path = new_scr_map.get(scr_id)
        if prev_path and new_path:
            scr_changes_raw.append({
                "scr_id": scr_id, "status": "modified",
                "prev_content": prev_path.read_text(encoding="utf-8")[:5000],
                "new_content": new_path.read_text(encoding="utf-8")[:5000],
            })
    for scr_id in diff["removed"]:
        prev_path = prev_scr_map.get(scr_id)
        if prev_path and prev_path.exists():
            scr_changes_raw.append({
                "scr_id": scr_id, "status": "removed",
                "prev_content": prev_path.read_text(encoding="utf-8")[:3000],
                "new_content": "",
            })

    # 4) 공통 문서 변경 (정책/디자인/overview)
    common_doc_changes = []
    if diff["common_changed"]:
        # overview 비교
        if prev_cls.get("overview") and new_cls.get("overview"):
            if prev_cls["overview"].read_bytes() != new_cls["overview"].read_bytes():
                common_doc_changes.append({
                    "doc_type": "overview",
                    "file_name": new_cls["overview"].name,
                    "prev_size": prev_cls["overview"].stat().st_size,
                    "new_size": new_cls["overview"].stat().st_size,
                })
        # policy 변경
        prev_policy_set = {p.name for p in prev_cls["policy"]}
        new_policy_set = {p.name for p in new_cls["policy"]}
        for name in (new_policy_set - prev_policy_set):
            common_doc_changes.append({"doc_type": "policy", "file_name": name, "status": "added"})
        for name in (prev_policy_set - new_policy_set):
            common_doc_changes.append({"doc_type": "policy", "file_name": name, "status": "removed"})
        # design 변경
        prev_design_set = {p.name for p in prev_cls["design"]}
        new_design_set = {p.name for p in new_cls["design"]}
        for name in (new_design_set - prev_design_set):
            common_doc_changes.append({"doc_type": "design", "file_name": name, "status": "added"})
        for name in (prev_design_set - new_design_set):
            common_doc_changes.append({"doc_type": "design", "file_name": name, "status": "removed"})

    # 5) AI 요약 호출 — 변경 사항이 의미하는 바를 자연어로
    ai_summary = ""
    scr_changes = []  # AI 요약 포함
    if scr_changes_raw or common_doc_changes:
        # AI 프롬프트 — 각 SCR 변경의 핵심 의미와 TC 영향 추정 요청
        change_blocks = []
        for sc in scr_changes_raw[:15]:  # 토큰 제한 — 첫 15개만 자세히
            block = f"### {sc['scr_id']} ({sc['status']})\n"
            if sc["status"] == "added":
                block += f"신규 화면:\n```\n{sc['new_content'][:3000]}\n```"
            elif sc["status"] == "modified":
                block += f"이전 (요약):\n```\n{sc['prev_content'][:1500]}\n```\n\n신규 (요약):\n```\n{sc['new_content'][:1500]}\n```"
            elif sc["status"] == "removed":
                block += f"삭제 (이전 본문):\n```\n{sc['prev_content'][:2000]}\n```"
            change_blocks.append(block)

        common_block = ""
        if common_doc_changes:
            common_block = "\n## 공통 문서 변경\n" + "\n".join(
                f"- [{c['doc_type']}] {c['file_name']}: {c.get('status', 'modified')}"
                for c in common_doc_changes
            )

        system_prompt = """당신은 QA 도메인 전문가입니다. 기획서 spec 폴더의 변경 사항을 분석하여
변경의 의미와 TC 영향 범위를 한국어로 요약합니다.

## 출력 형식

```markdown
# 변경 분석 보고서

## 변경 요약 (개요)
(전체 변경의 한 문단 요약 — 어떤 기능이 어떻게 바뀌었는가)

## 화면별 변경 상세

### SCR-XXX (status)
- **변경 핵심**: 한 줄 요약
- **TC 영향 추정**: 어떤 종류의 TC 가 영향 받을 가능성 (수정/신규/삭제)
- **권장 액션**: 1-2 문장

(각 SCR 마다 반복)

## 공통 문서 변경 영향
(정책/디자인 변경이 TC 전반에 미치는 영향)

## 종합 권장사항
- 우선 검토할 SCR / TC 종류
- 주의 사항
```

원칙:
- spec 본문에서 추출. 임의 추측 금지.
- TC ID 자체는 모름 (아직 매핑 안 됨) — "영향 종류" 만 추정.
- 간결하게."""

        user_prompt = f"""다음은 기획서 두 버전의 변경 분석입니다.

## 변경 통계
- 신규 화면: {len(diff['added'])}개 — {', '.join(diff['added'][:10])}
- 수정 화면: {len(diff['modified'])}개 — {', '.join(diff['modified'][:10])}
- 삭제 화면: {len(diff['removed'])}개 — {', '.join(diff['removed'][:10])}
- 동일 화면: {len(diff['unchanged'])}개
- 공통 문서 변경: {diff['common_changed']}

## 변경 상세 (상위 15개)

{"\n\n---\n\n".join(change_blocks)}
{common_block}

위 변경 사항을 분석해 보고서를 작성하세요.
"""
        try:
            ai_summary = call_claude(system_prompt, user_prompt, max_tokens=8000)
        except Exception as e:
            ai_summary = f"# 변경 분석 보고서\n\nAI 분석 실패: {e}\n\n## 변경 통계\n- 신규: {len(diff['added'])}개\n- 수정: {len(diff['modified'])}개\n- 삭제: {len(diff['removed'])}개"

    # 6) 결과 구조 정리
    for sc in scr_changes_raw:
        scr_changes.append({
            "scr_id": sc["scr_id"],
            "status": sc["status"],
            "prev_size": len(sc.get("prev_content", "")),
            "new_size": len(sc.get("new_content", "")),
        })

    from datetime import datetime
    return {
        "diff": diff,
        "summary": ai_summary,
        "scr_changes": scr_changes,
        "common_doc_changes": common_doc_changes,
        "meta": {
            "prev_folder": str(prev_p),
            "new_folder": str(new_p),
            "existing_tc_url": existing_tc_url,
            "analyzed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        },
    }


# ── TC Update 2단계 — Sheets 읽기 + 사본 + 판정 로직 ─────────────────────────

def _extract_sheets_id(url: str) -> str:
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url or "")
    if not m:
        raise RuntimeError("Google Sheets URL 형식이 아닙니다.")
    return m.group(1)


# TC update 작업으로 생성되는 사본/백업의 보관 폴더 (Drive)
# 사용자별 설정 — ~/.tc-update-config.json 에 저장
# 각 사용자가 본인의 Drive 폴더를 지정 (개인/공유 드라이브 모두 가능)
TC_UPDATE_CONFIG_FILE = Path.home() / ".tc-update-config.json"


def _extract_folder_id(url_or_id: str) -> str:
    """폴더 URL 또는 ID 에서 폴더 ID 추출.
    예: 'https://drive.google.com/drive/folders/1oLsHJ...' → '1oLsHJ...'
        '1oLsHJ...' (이미 ID) → 그대로
    """
    s = (url_or_id or "").strip()
    if not s:
        return ""
    m = re.search(r"/folders/([a-zA-Z0-9_-]+)", s)
    if m:
        return m.group(1)
    # 그냥 ID 형태로 간주 (영숫자/대시/언더스코어만)
    if re.match(r"^[a-zA-Z0-9_-]+$", s):
        return s
    return ""


def load_tc_update_config() -> dict:
    """사용자 설정 로드. 없으면 빈 dict."""
    if not TC_UPDATE_CONFIG_FILE.exists():
        return {}
    try:
        return json.loads(TC_UPDATE_CONFIG_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_tc_update_config(data: dict) -> None:
    """사용자 설정 저장."""
    TC_UPDATE_CONFIG_FILE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def get_tc_update_folder_id() -> str:
    """현재 설정된 Drive 폴더 ID 반환. 없으면 빈 문자열 (= 폴더 지정 안 함, orphan)."""
    cfg = load_tc_update_config()
    return cfg.get("drive_folder_id", "") or ""


def copy_sheets_for_update(source_url: str, new_title: str = "") -> dict:
    """원본 Sheets 를 복사해 사본 ID/URL 반환. Drive API copy 사용.

    사본은 설정된 Drive 폴더 (load_tc_update_config) 아래에 자동 정리.
    폴더 미설정 또는 접근 불가 시 graceful fallback (원래대로 orphan).

    Returns: {ok, copy_id, copy_url, copy_title}
    """
    drive = get_drive_service()
    src_id = _extract_sheets_id(source_url)
    src_meta = drive.files().get(fileId=src_id, fields="id,name,mimeType").execute()
    if not new_title:
        from datetime import datetime
        ts = datetime.now().strftime("%Y-%m-%d %H:%M")
        new_title = f"{src_meta.get('name', 'TC')} (update-test {ts})"
    body = {"name": new_title}
    folder_id = get_tc_update_folder_id()
    if folder_id:
        body["parents"] = [folder_id]
    try:
        copied = drive.files().copy(fileId=src_id, body=body, fields="id,name,webViewLink,parents").execute()
    except Exception as e:
        # 폴더 접근 권한 없을 가능성 — 폴더 없이 재시도
        body.pop("parents", None)
        copied = drive.files().copy(fileId=src_id, body=body, fields="id,name,webViewLink").execute()
    return {
        "ok": True,
        "copy_id": copied["id"],
        "copy_url": copied.get("webViewLink") or f"https://docs.google.com/spreadsheets/d/{copied['id']}/edit",
        "copy_title": copied.get("name"),
        "source_id": src_id,
        "source_title": src_meta.get("name"),
    }


def backup_original_sheets(source_url: str) -> dict:
    """원본 Sheets 를 'backup before update YYYY-MM-DD HH:MM' 이름으로 복사 → 백업.

    원본을 그대로 두고 별도 사본으로 보관. 롤백 시 이 사본의 셀 값을 원본으로 복원.
    백업도 설정된 Drive 폴더 (load_tc_update_config) 아래에 자동 정리.

    Returns: {ok, backup_id, backup_url, backup_title, source_id, source_title}
    """
    drive = get_drive_service()
    src_id = _extract_sheets_id(source_url)
    src_meta = drive.files().get(fileId=src_id, fields="id,name").execute()
    from datetime import datetime
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    backup_title = f"{src_meta.get('name', 'TC')} (backup before update {ts})"
    body = {"name": backup_title}
    folder_id = get_tc_update_folder_id()
    if folder_id:
        body["parents"] = [folder_id]
    try:
        copied = drive.files().copy(
            fileId=src_id, body=body, fields="id,name,webViewLink,parents",
        ).execute()
    except Exception:
        body.pop("parents", None)
        copied = drive.files().copy(
            fileId=src_id, body=body, fields="id,name,webViewLink",
        ).execute()
    return {
        "ok": True,
        "backup_id": copied["id"],
        "backup_url": copied.get("webViewLink") or f"https://docs.google.com/spreadsheets/d/{copied['id']}/edit",
        "backup_title": copied.get("name"),
        "source_id": src_id,
        "source_title": src_meta.get("name"),
        "backed_up_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def _col_index_to_letter(idx: int) -> str:
    """0 → 'A', 25 → 'Z', 26 → 'AA' 변환."""
    s = ""
    n = idx
    while True:
        s = chr(ord("A") + (n % 26)) + s
        n = n // 26 - 1
        if n < 0:
            break
    return s


def read_tc_edit_log(target_sheets_id: str) -> dict:
    """사본의 'TC Edit Log' 시트를 읽어 적용 대상 셀 목록 반환.

    Returns: {
      ok, log_exists, entries: [{tc_id, sheet_title, row_index, field, new_value, ts, rationale}],
      count, sheets_seen: [list of sheet titles affected]
    }

    헤더 형식 (TC Edit Log 시트):
      ['시각', 'TC ID', '시트', '행', '필드', '이전 값(요약)', '새 값(요약)', '수정 사유']

    주의: '새 값(요약)' 은 200 자 잘림 — 실제 적용 시에는 사본의 해당 셀을 직접 읽어 사용.
    """
    svc = get_sheets_service()
    # 'TC Edit Log' 시트 존재 여부 확인
    meta = svc.spreadsheets().get(
        spreadsheetId=target_sheets_id,
        fields="sheets(properties(sheetId,title))",
    ).execute()
    titles = [s["properties"]["title"] for s in meta.get("sheets", [])]
    if "TC Edit Log" not in titles:
        return {"ok": True, "log_exists": False, "entries": [], "count": 0, "sheets_seen": []}

    resp = svc.spreadsheets().values().get(
        spreadsheetId=target_sheets_id,
        range="'TC Edit Log'!A1:Z2000",
    ).execute()
    values = resp.get("values", [])
    if not values or len(values) < 2:
        return {"ok": True, "log_exists": True, "entries": [], "count": 0, "sheets_seen": []}

    header = values[0]
    def col_idx(name):
        return header.index(name) if name in header else -1

    ci_ts = col_idx("시각")
    ci_tc = col_idx("TC ID")
    ci_sheet = col_idx("시트")
    ci_row = col_idx("행")
    ci_field = col_idx("필드")
    ci_old = col_idx("이전 값(요약)")
    ci_new = col_idx("새 값(요약)")
    ci_rationale = col_idx("rationale")

    entries = []
    sheets_seen = set()
    for r in values[1:]:
        if not any(str(c).strip() for c in r):
            continue
        def cell(ci):
            if ci < 0 or ci >= len(r):
                return ""
            return str(r[ci] or "").strip()
        tc_id = cell(ci_tc)
        sheet_title = cell(ci_sheet)
        row_str = cell(ci_row)
        field = cell(ci_field)
        if not (tc_id and sheet_title and row_str and field):
            continue
        try:
            row_idx = int(row_str)
        except Exception:
            continue
        entries.append({
            "tc_id": tc_id,
            "sheet_title": sheet_title,
            "row_index": row_idx,
            "field": field,
            "old_value": cell(ci_old),
            "new_value": cell(ci_new),
            "ts": cell(ci_ts),
            "rationale": cell(ci_rationale),
        })
        sheets_seen.add(sheet_title)

    return {
        "ok": True,
        "log_exists": True,
        "entries": entries,
        "count": len(entries),
        "sheets_seen": sorted(sheets_seen),
    }


# 헤더 캐시 — (sheets_id, sheet_title) → header list
# Sheets API quota 절약 (분당 60회 한도)
_HEADER_CACHE = {}


def _get_sheet_header(sheets_id: str, sheet_title: str) -> list:
    """시트의 헤더 행을 캐시 우선 반환. 캐시 miss 시 API 1회 호출."""
    key = (sheets_id, sheet_title)
    if key in _HEADER_CACHE:
        return _HEADER_CACHE[key]
    svc = get_sheets_service()
    resp = svc.spreadsheets().values().get(
        spreadsheetId=sheets_id,
        range=f"'{sheet_title}'!A1:Z3",
    ).execute()
    values = resp.get("values", [])
    header = []
    for row in values[:3]:
        if any(c in ("TC ID", "TC_ID") for c in row):
            header = row
            break
    if not header and values:
        header = values[0]
    _HEADER_CACHE[key] = header
    return header


def resolve_field_column(sheets_id: str, sheet_title: str, field: str) -> int:
    """주어진 시트에서 필드명(예: '테스트 스텝')에 해당하는 컬럼 인덱스 반환.

    헤더 별칭 매핑 (read_sheets_tcs 와 동일):
      precondition → '사전조건' / '사전 조건' / 'Precondition'
      steps → '테스트 스텝' / '테스트 단계' / 'Steps'
      expected → '기대결과' / '기대 결과' / '예상 결과' / 'Expected'

    헤더는 캐시됨 (분당 quota 절약 — 같은 시트 반복 조회 시 1회 API).

    Returns: 컬럼 0-based index, 못 찾으면 -1.
    """
    header = _get_sheet_header(sheets_id, sheet_title)

    # 필드 → 가능한 헤더 이름들
    field_aliases = {
        "precondition": ["사전조건", "사전 조건", "Precondition"],
        "사전조건":      ["사전조건", "사전 조건", "Precondition"],
        "steps":         ["테스트 스텝", "테스트 단계", "Steps"],
        "테스트 스텝":   ["테스트 스텝", "테스트 단계", "Steps"],
        "expected":      ["기대결과", "기대 결과", "예상 결과", "Expected"],
        "기대결과":      ["기대결과", "기대 결과", "예상 결과", "Expected"],
    }
    aliases = field_aliases.get(field, [field])
    for alias in aliases:
        if alias in header:
            return header.index(alias)
    return -1


# 메타/카탈로그 탭은 TC 시트가 아니므로 제외 — 제목으로 휴리스틱 매칭
TC_META_TAB_PATTERNS = [
    "화면 목록", "화면목록", "Screen List", "screens",
    "Update Log", "변경 이력", "Change Log",
    "변경 분석", "Diff", "diff",
]

# TC ID 패턴 — Generator 가 만드는 표준 형식 (예: SM-EML-001, SC-EXCH-002, SM-LITE-COMBO-005)
TC_ID_RE = re.compile(r"^[A-Z]{2,}-[A-Z0-9]+(?:-[A-Z0-9]+)*-\d{2,}$")


def _is_tc_meta_tab(title: str) -> bool:
    t = title.strip().lower()
    for pat in TC_META_TAB_PATTERNS:
        if pat.lower() in t:
            return True
    return False


def _parse_scr_field(value: str) -> list:
    """'SCR-104' / 'SCR-104, SCR-106' / 'SCR-104 / SCR-106' 등 모두 처리.
    빈 문자열, None 안전.
    """
    if not value:
        return []
    return re.findall(r"SCR-\d+[A-Z]?", str(value))


def read_sheets_tcs(sheets_id: str) -> dict:
    """Sheets API 로 모든 탭의 TC 행 읽어 TC ID ↔ SCR 매핑 + 메타 추출.

    헤더 별칭 매핑 (사용자 시트 형식 + Generator 표준 형식 모두 지원):
      TC ID:    "TC ID", "TC_ID", "ID"
      제목:     "소분류", "제목", "Title"     (사용자 시트는 소분류가 제목 역할)
      대분류:   "대분류"
      중분류:   "중분류"
      Steps:    "테스트 스텝", "테스트 단계", "Steps"
      Pre:      "사전조건", "사전 조건", "Precondition"
      Expected: "기대결과", "기대 결과", "예상 결과", "Expected"
      SCR:      "화면 코드", "화면코드", "SCR", "SCR ID", "관련 SCR", "화면 ID"
      Smoke:    "Smoke", "최소TC", "최소 TC"
      Priority: "중요도", "우선순위", "Priority"
      Type:     "분류", "Type"

    화면 목록 / Update Log 같은 메타 탭은 제외.
    TC ID 패턴 검증으로 빈 행/노이즈 제거.

    Returns: {
      "tabs": [{title, header, rows, skipped: bool, skip_reason}],
      "tcs": [{tc_id, scr_ids, title, type, sheet_title, row_index, category, sub_category}],
      "tc_by_scr": {scr_id: [tc_id, ...]},
    }
    """
    svc = get_sheets_service()
    meta = svc.spreadsheets().get(spreadsheetId=sheets_id,
                                     fields="sheets(properties(sheetId,title,index))").execute()
    sheet_titles = [s["properties"]["title"] for s in meta.get("sheets", [])]

    tabs = []
    tcs = []
    tc_by_scr: dict = {}

    if not sheet_titles:
        return {"tabs": [], "tcs": [], "tc_by_scr": {}}

    ranges = [f"'{t}'!A1:Z2000" for t in sheet_titles]
    batch = svc.spreadsheets().values().batchGet(spreadsheetId=sheets_id, ranges=ranges).execute()
    value_ranges = batch.get("valueRanges", [])

    for title, vr in zip(sheet_titles, value_ranges):
        values = vr.get("values", [])
        if not values:
            tabs.append({"title": title, "header": [], "rows": [], "skipped": True, "skip_reason": "빈 시트"})
            continue

        # 메타 탭 제외
        if _is_tc_meta_tab(title):
            tabs.append({"title": title, "header": values[0] if values else [],
                         "rows": [], "skipped": True, "skip_reason": "메타/카탈로그 탭"})
            continue

        # 헤더 행 탐색 — TC ID 컬럼 포함하는 첫 행 (상위 10행 안에서)
        header = []
        header_idx = -1
        for i, row in enumerate(values[:10]):
            cells = [str(c).strip() for c in row]
            if any(c in ("TC ID", "TC_ID") for c in cells):
                header = row
                header_idx = i
                break
        if header_idx < 0:
            # TC ID 컬럼이 없으면 TC 시트 아님
            tabs.append({"title": title, "header": values[0] if values else [],
                         "rows": [], "skipped": True, "skip_reason": "TC ID 헤더 없음"})
            continue

        def col(name_candidates):
            for nm in name_candidates:
                if nm in header:
                    return header.index(nm)
            return -1

        col_id = col(["TC ID", "TC_ID", "ID"])
        col_subc = col(["소분류"])  # 사용자 시트 — 소분류가 제목 역할
        col_title = col(["제목", "Title"])
        col_cat = col(["대분류"])
        col_mid = col(["중분류"])
        col_type = col(["분류", "Type"])
        col_scr = col(["화면 코드", "화면코드", "SCR", "SCR ID", "관련 SCR", "화면 ID"])
        col_steps = col(["테스트 스텝", "테스트 단계", "Steps"])
        col_pre = col(["사전조건", "사전 조건", "Precondition"])
        col_exp = col(["기대결과", "기대 결과", "예상 결과", "Expected"])
        col_smoke = col(["Smoke", "최소TC", "최소 TC"])
        col_prio = col(["중요도", "우선순위", "Priority"])

        def cell(row, ci):
            if ci < 0 or ci >= len(row):
                return ""
            return str(row[ci] or "").strip()

        rows_out = []
        for ri, row in enumerate(values[header_idx + 1:], start=header_idx + 2):
            if not any(str(c).strip() for c in row):
                continue
            tc_id = cell(row, col_id).replace("*", "").strip()
            # TC ID 패턴 검증 — Generator 표준 형식만 채택 (노이즈 차단)
            if not tc_id or not TC_ID_RE.match(tc_id):
                continue

            tc_title = cell(row, col_subc) or cell(row, col_title)
            tc_type = cell(row, col_type)
            steps_text = cell(row, col_steps)
            pre_text = cell(row, col_pre)
            exp_text = cell(row, col_exp)
            scr_field = cell(row, col_scr)
            cat_text = cell(row, col_cat)
            mid_text = cell(row, col_mid)
            smoke_text = cell(row, col_smoke)
            prio_text = cell(row, col_prio)

            # SCR ID 우선순위:
            # 1) 전용 컬럼 (화면 코드) — 가장 신뢰도 높음
            # 2) 없으면 본문 (steps/pre/exp/title) 에서 SCR-NNN 패턴 추출
            scr_ids = _parse_scr_field(scr_field)
            if not scr_ids:
                for blob in (tc_title, mid_text, steps_text, pre_text, exp_text):
                    scr_ids.extend(_parse_scr_field(blob))
            # 중복 제거 + 정렬 (숫자 부분 기준)
            scr_set = list(dict.fromkeys(scr_ids))

            def _scr_key(s):
                m = re.match(r"SCR-(\d+)", s)
                return int(m.group(1)) if m else 0
            scr_set.sort(key=_scr_key)

            tc_record = {
                "tc_id": tc_id,
                "scr_ids": scr_set,
                "title": tc_title,
                "type": tc_type,
                "category": cat_text,
                "sub_category": mid_text,
                "smoke": smoke_text,
                "priority": prio_text,
                "sheet_title": title,
                "row_index": ri,
            }
            tcs.append(tc_record)
            rows_out.append(tc_record)
            for s in scr_set:
                tc_by_scr.setdefault(s, []).append(tc_id)

        tabs.append({"title": title, "header": header, "rows": rows_out, "skipped": False})

    return {"tabs": tabs, "tcs": tcs, "tc_by_scr": tc_by_scr}


def classify_scr_change_kind(prev_text: str, new_text: str) -> str:
    """수정된 SCR 의 변경 종류를 휴리스틱으로 분류.

    Returns: 'policy_flow' | 'design_only' | 'mixed'
    """
    if not prev_text or not new_text:
        return "policy_flow"

    # 정책/플로우 키워드 — 이 단어들 주변이 바뀌면 policy_flow
    policy_keys = [
        "정책", "검증", "validation", "조건", "최소", "최대", "초과", "미만",
        "에러", "오류", "실패", "경고", "분기", "허용", "차단", "거부", "리다이렉트",
        "API", "endpoint", "status", "401", "403", "404", "500",
    ]
    design_keys = [
        "디자인", "색", "배경", "폰트", "여백", "위치", "정렬", "아이콘", "이미지",
        "텍스트", "문구", "라벨", "label", "placeholder", "글자",
    ]

    def diff_lines(a: str, b: str):
        import difflib
        a_lines = a.splitlines()
        b_lines = b.splitlines()
        d = list(difflib.unified_diff(a_lines, b_lines, lineterm=""))
        added = [ln[1:] for ln in d if ln.startswith("+") and not ln.startswith("+++")]
        removed = [ln[1:] for ln in d if ln.startswith("-") and not ln.startswith("---")]
        return added + removed

    changed = "\n".join(diff_lines(prev_text, new_text))
    if not changed.strip():
        return "design_only"

    has_policy = any(k.lower() in changed.lower() for k in policy_keys)
    has_design = any(k.lower() in changed.lower() for k in design_keys)
    if has_policy and has_design:
        return "mixed"
    if has_policy:
        return "policy_flow"
    if has_design:
        return "design_only"
    # 기본값: 모호한 변경은 정책으로 간주 (보수적)
    return "policy_flow"


def build_update_candidates(diff_result: dict, tc_data: dict,
                              prev_folder: str, new_folder: str,
                              scr_filter: list | None = None) -> list:
    """판정 로직 — SCR diff + TC↔SCR 매핑 → 4-way 후보 리스트.

    Args:
      scr_filter: 선택된 SCR ID 리스트 (None 이면 전체).
                    이 리스트에 있는 SCR 만 처리하고 나머지는 무시.

    Returns: list of {
      action: 'add'|'modify'|'delete'|'skip',
      default_checked: bool,
      scr_id, tc_id (optional), reason, sheet_title (optional),
      change_kind (optional), preview (optional),
    }
    """
    diff = diff_result.get("diff", {})
    tc_by_scr = tc_data.get("tc_by_scr", {})
    tcs_by_id = {t["tc_id"]: t for t in tc_data.get("tcs", [])}

    # SCR 필터링 — 선택된 SCR 만 남기기
    if scr_filter:
        filter_set = set(scr_filter)
        diff = {
            "added": [s for s in (diff.get("added") or []) if s in filter_set],
            "modified": [s for s in (diff.get("modified") or []) if s in filter_set],
            "removed": [s for s in (diff.get("removed") or []) if s in filter_set],
            "unchanged": diff.get("unchanged", []),
            "common_changed": diff.get("common_changed", False),
        }

    # SCR 본문 캐시 (수정 분류용)
    prev_p = Path(prev_folder).expanduser()
    new_p = Path(new_folder).expanduser()
    prev_cls = classify_spec_files(prev_p)
    new_cls = classify_spec_files(new_p)
    prev_map = {p.stem.split(".")[0]: p for p in prev_cls.get("screens", [])}
    new_map = {p.stem.split(".")[0]: p for p in new_cls.get("screens", [])}

    candidates = []

    # 1) 삭제된 SCR → 해당 TC 삭제 후보
    for scr_id in diff.get("removed", []):
        for tc_id in tc_by_scr.get(scr_id, []):
            tc = tcs_by_id.get(tc_id, {})
            candidates.append({
                "action": "delete",
                "default_checked": True,
                "scr_id": scr_id,
                "tc_id": tc_id,
                "tc_title": tc.get("title", ""),
                "sheet_title": tc.get("sheet_title", ""),
                "reason": f"SCR-{scr_id.split('-')[-1]} 삭제됨 — 검증 대상 없어짐",
            })

    # 2) 신규 SCR → 신규 TC 제안
    for scr_id in diff.get("added", []):
        candidates.append({
            "action": "add",
            "default_checked": True,
            "scr_id": scr_id,
            "tc_id": None,
            "reason": f"{scr_id} 신규 추가 — 해당 화면 TC 가 없음",
        })

    # 3) 수정된 SCR → 변경 종류 판정 + TC 매핑
    for scr_id in diff.get("modified", []):
        prev_path = prev_map.get(scr_id)
        new_path = new_map.get(scr_id)
        prev_text = prev_path.read_text(encoding="utf-8") if prev_path and prev_path.exists() else ""
        new_text = new_path.read_text(encoding="utf-8") if new_path and new_path.exists() else ""
        kind = classify_scr_change_kind(prev_text, new_text)

        mapped_tcs = tc_by_scr.get(scr_id, [])

        if not mapped_tcs:
            # 매핑된 TC 없음 → 신규 TC 제안 (수정인데 TC 가 없으니 새로 만들어야 할 수도)
            candidates.append({
                "action": "add",
                "default_checked": False,
                "scr_id": scr_id,
                "tc_id": None,
                "change_kind": kind,
                "reason": f"{scr_id} 수정됨 ({kind}) — 매핑된 TC 없음. 신규 TC 검토 필요",
            })
            continue

        for tc_id in mapped_tcs:
            tc = tcs_by_id.get(tc_id, {})
            # 디자인만 변경 → 기본 보류, 정책/플로우 → 기본 체크
            if kind == "design_only":
                default = False
                reason = f"{scr_id} 디자인/문구 변경 — TC step 표현 검토 (기본 보류)"
            elif kind == "policy_flow":
                default = True
                reason = f"{scr_id} 정책/플로우 변경 — step/expected 갱신 필요"
            else:  # mixed
                default = True
                reason = f"{scr_id} 정책 + 디자인 혼합 변경 — 검토 필요"
            candidates.append({
                "action": "modify",
                "default_checked": default,
                "scr_id": scr_id,
                "tc_id": tc_id,
                "tc_title": tc.get("title", ""),
                "sheet_title": tc.get("sheet_title", ""),
                "change_kind": kind,
                "reason": reason,
            })

    # 4) unchanged → 후보에 절대 포함 안 함 (하드 가드레일)

    return candidates


def step_propose_tc_update(scr_id: str, prev_scr_text: str, new_scr_text: str,
                              current_tc: dict) -> dict:
    """1건의 TC 에 대해 AI 수정안 제안.

    Args:
      scr_id: 'SCR-102' 등
      prev_scr_text: 이전 SCR 본문
      new_scr_text: 새 SCR 본문
      current_tc: {tc_id, title, steps, expected, precondition, ...} — 사본에서 읽은 현재 값

    Returns: {
      "ok": True,
      "proposal": {
        "steps": str | None,        # None = 변경 안 함
        "expected": str | None,
        "precondition": str | None,
        "rationale": str,            # 한 줄 이유 (왜 이렇게 바꾸는지)
        "no_change": bool,           # True 이면 변경 불필요
      }
    }

    원칙:
    - 보수적: SCR 변경이 명백하게 영향 주는 부분만 수정 제안
    - 디자인/문구 변경은 무시 (사용자 합의: design_only 는 기본 보류)
    - 확신 없으면 'no_change: true' 반환
    - JSON 으로 응답 강제 (파싱 안정성)
    """
    if not new_scr_text and not prev_scr_text:
        return {"ok": False, "error": "SCR 본문이 없습니다."}

    system_prompt = """당신은 QA 도메인 전문가입니다. 기획서(SCR) 변경에 따라
테스트 케이스를 어떻게 수정해야 할지 균형있게 제안합니다.

## 원칙 (반드시 지킬 것)

1. **TC 의 검증 범위 인식 — 가장 중요**:
   - **TC 제목이 "UI 확인", "화면 확인", "상태 표시", "Layout", "초기 화면" 등 광범위한 UI 검증** 이면,
     SCR 의 새 UI 요소 (배너, CTA, 안내 영역, 새 컨디셔널 블록) 도 **반드시 기대결과에 추가**.
   - **TC 가 특정 동작 1개만 검증** (예: 버튼 탭, 필터 전환, 특정 기능) 이면,
     그 동작과 무관한 새 요소는 무시.
   - 판단 기준: "이 TC 를 실행하면 새 UI 요소가 화면에 보일 텐데, TC 가 그걸 검증해야 하나?"

2. **새 컨디셔널/분기 추가 = 신규 검증 항목**:
   - SCR 에 `conditional [...]` 또는 새 상태 (permission, state) 분기가 추가되면 → **검증 누락**.
   - "디자인 변경" 으로 간주하지 말 것. 이건 새 **기능 분기**.

3. **정책/플로우 변경**: 검증 조건, 분기, API 응답 처리, 임계치 변경 등은 step/expected 갱신.

4. **디자인/문구 변경 — TC 본문이 인용했나로 판단**:

   핵심 기준: **"TC 의 step 이나 expected 가 그 변경된 요소를 인용·명시하고 있는가?"**

   인용 안 함 → ✅ 무시 OK:
   - 색상, 폰트, 폰트 크기, 여백, 그림자, 둥근 모서리 (시각 속성)
   - 애니메이션, 전환 효과 (fade-in → slide-up 등)
   - 의미 동등한 아이콘 교체 (TC 가 아이콘 명시 안 함)
   - 라벨 wording 변경 (TC 가 그 라벨 직접 안 씀)

   인용·명시함 → ❌ 갱신 필요:
   - TC step/expected 가 **정확한 라벨 인용**: '"Save" 버튼 탭' 인데 라벨이 "저장" 으로 바뀜
   - TC 가 **위치/정렬 검증**: "우측 상단의 X 버튼" 인데 위치 이동
   - TC 가 **아이콘 명시**: "🔔 아이콘 표시" 인데 다른 아이콘으로 교체
   - TC 가 **명시적 시각 검증**: "빨간색 배경 표시" 같은 색 검증
   - TC 가 **레이아웃 구조 검증**: "카드 리스트로 표시" → 구조 변경

   ### 예시 1 — 무시 OK
   ```
   SCR 변경: Submit 버튼 색 #1E40AF → #2563EB
   TC step: "저장하기를 시도한다"
   TC expected: "데이터가 저장된다"
   → 색 변경을 TC 가 인용 안 함 → no_change OK
   ```

   ### 예시 2 — 갱신 필요
   ```
   SCR 변경: "Save" 버튼 라벨 → "저장" 으로 변경
   TC step: '"Save" 버튼을 탭한다'
   TC expected: '"Saved successfully" 토스트 표시'
   → step 이 "Save" 직접 인용 → "저장" 으로 갱신
   ```

   ### 예시 3 — 신규 UI 요소 추가 (가장 중요)
   ```
   SCR 변경: 화면 상단에 새 CTA 배너 conditional 추가
   TC 제목: "UI 확인" / 기대결과: "상단 nav-bar, 필터 탭, 리스트 표시"
   → TC 가 "무엇이 표시되는지" 검증 → 새 CTA 배너 검증 항목 추가 필수
   (이건 디자인이 아니라 새 검증 항목 — 원칙 2 참조)
   ```

5. **부분 수정 가능**: 3개 필드 중 일부만 수정 필요하면 그것만. 변경 없는 필드는 `null`.

6. **애매하면 수정 제안 우선** (no_change 보수적):
   - 위 1-5 검토 후 **명확하게** 영향 없을 때만 `no_change: true`.
   - 50/50 애매한 경우 → 수정안 제시 (사람이 거부 가능).
   - 누락 위험이 잘못 제안 위험보다 큼 — 검증 공백을 만들지 말 것.

## 출력 형식 (JSON 만, 다른 텍스트 금지)
```json
{
  "no_change": false,
  "steps": "수정된 테스트 스텝 전문 또는 null (변경 없으면)",
  "expected": "수정된 기대결과 전문 또는 null",
  "precondition": "수정된 사전조건 전문 또는 null",
  "rationale": "한 줄로 — 왜 이렇게 수정하는가 (SCR 변경의 어느 부분이 어떻게 반영됐는지)"
}
```

변경 필요 없으면:
```json
{
  "no_change": true,
  "steps": null,
  "expected": null,
  "precondition": null,
  "rationale": "이 TC 는 SCR 변경의 영향을 받지 않음 (이유 한 줄)"
}
```
"""

    user_prompt = f"""## 대상 TC

- TC ID: {current_tc.get('tc_id', '')}
- 제목: {current_tc.get('title', '')}
- 매핑 SCR: {scr_id}

### 현재 사전조건
```
{current_tc.get('precondition', '') or '(없음)'}
```

### 현재 테스트 스텝
```
{current_tc.get('steps', '') or '(없음)'}
```

### 현재 기대결과
```
{current_tc.get('expected', '') or '(없음)'}
```

## {scr_id} 본문 (이전 버전)
```
{(prev_scr_text or '(없음)')[:4000]}
```

## {scr_id} 본문 (새 버전)
```
{(new_scr_text or '(없음)')[:4000]}
```

위 변경을 검토해 현재 TC 의 사전조건/테스트 스텝/기대결과를 수정해야 하는지
판단하고 JSON 으로만 응답하세요. 보수적으로 — 확신 없으면 no_change=true.
"""

    try:
        raw = call_claude(system_prompt, user_prompt, max_tokens=3000)
    except Exception as e:
        return {"ok": False, "error": f"AI 호출 실패: {e}"}

    # JSON 파싱 — 코드블록 포함 응답도 처리
    import json as _json
    text = raw.strip()
    # ```json ... ``` 코드블록 벗기기
    m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.DOTALL)
    if m:
        text = m.group(1)
    else:
        # 그냥 JSON 객체만 추출
        m = re.search(r"\{.*\}", text, re.DOTALL)
        if m:
            text = m.group(0)

    try:
        data = _json.loads(text)
    except Exception as e:
        return {"ok": False, "error": f"JSON 파싱 실패: {e}", "raw": raw[:500]}

    # 필드 정규화
    proposal = {
        "no_change": bool(data.get("no_change", False)),
        "steps": data.get("steps"),
        "expected": data.get("expected"),
        "precondition": data.get("precondition"),
        "rationale": str(data.get("rationale", "")).strip() or "(이유 없음)",
    }
    # no_change 일관성 — 셋 다 None 이면 자동으로 no_change=True
    if not any([proposal["steps"], proposal["expected"], proposal["precondition"]]):
        proposal["no_change"] = True

    return {"ok": True, "proposal": proposal}


def step_parse_structured_spec_with_diff(sess: dict, new_folder_path: str,
                                           prev_folder_path: str,
                                           include_unchanged: bool = False) -> dict:
    """diff 모드 — 신규 폴더를 파싱하되 변경된/추가된 화면만 screens 에 남기고
    unchanged 는 따로 표기. step_parse_structured_spec 의 결과 + diff 정보를 반환.
    include_unchanged=True 이면 모든 화면 포함(분류표 일관성 유지용).
    """
    new_folder = Path(new_folder_path).expanduser()
    prev_folder = Path(prev_folder_path).expanduser()
    if not prev_folder.exists():
        raise FileNotFoundError(f"이전 폴더 없음: {prev_folder}")

    push_log(sess, f"[diff] 비교 — 이전={prev_folder.name} / 신규={new_folder.name}")
    diff = diff_spec_folders(prev_folder, new_folder)
    push_log(sess,
        f"[diff] 추가 {len(diff['added'])}개 / 수정 {len(diff['modified'])}개 / "
        f"삭제 {len(diff['removed'])}개 / 동일 {len(diff['unchanged'])}개 / "
        f"공통 문서 변경={diff['common_changed']}")

    # 신규 폴더 정상 파싱
    spec = step_parse_structured_spec(sess, str(new_folder))

    # diff 결과를 screens 필터링에 반영 (재생성 대상만 남김)
    target_scrs = set(diff["added"]) | set(diff["modified"])
    if diff["common_changed"]:
        push_log(sess, "[diff] ⚠️ 공통 문서(정책/디자인/overview) 변경됨 — 모든 화면 재생성 권장")
    if include_unchanged or diff["common_changed"]:
        # 전체 유지 (분류표 일관성 + 공통 변경 시 안전 재생성)
        pass
    else:
        before_n = len(spec["screens"])
        spec["screens"] = [sc for sc in spec["screens"] if sc["id"] in target_scrs]
        push_log(sess, f"[diff] 재생성 대상으로 화면 필터링: {before_n}개 → {len(spec['screens'])}개")

    spec["_diff"] = diff
    spec["_prev_folder"] = str(prev_folder)
    return spec


def extract_pdf_text(pdf_path: Path) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(pdf_path))
        texts = []
        for i, page in enumerate(reader.pages):
            t = page.extract_text()
            if t:
                texts.append(f"--- 페이지 {i+1} ---\n{t}")
        return "\n\n".join(texts)
    except Exception as e:
        raise RuntimeError(f"PDF 파싱 실패: {e}")


def fetch_url_content(sess: dict, url: str, selected_files: list = None) -> str:
    try:
        import requests
    except ImportError:
        raise RuntimeError("requests 패키지가 필요합니다: pip install requests")

    push_log(sess, f"[파싱] URL 접속 중: {url}")

    # GitHub 리포지토리 URL 감지
    gh_match = re.match(r"https?://github\.com/([^/]+)/([^/\s]+?)(?:\.git)?/?$", url)
    if gh_match:
        owner, repo = gh_match.group(1), gh_match.group(2)
        return fetch_github_repo(sess, owner, repo, selected_files=selected_files)

    # github.io 또는 일반 URL
    try:
        headers = {"User-Agent": "Mozilla/5.0 (TC-Automation/2.0)"}
        resp = requests.get(url, headers=headers, timeout=30)
        resp.raise_for_status()
        content_type = resp.headers.get("content-type", "")

        if "html" in content_type:
            return html_to_text(resp.text)
        else:
            return resp.text
    except Exception as e:
        raise RuntimeError(f"URL 접속 실패: {e}")


def fetch_github_repo(sess: dict, owner: str, repo: str, selected_files: list = None) -> str:
    """GitHub 리포지토리에서 파일 내용을 가져온다.
    selected_files: 경로 문자열 리스트. None이면 .md 파일 최대 10개 자동 선택.
    """
    try:
        import requests
    except ImportError:
        raise RuntimeError("requests 패키지가 필요합니다")

    texts = []
    headers = {"User-Agent": "TC-Automation/2.0"}
    token = os.environ.get("GITHUB_TOKEN", "")
    if token:
        headers["Authorization"] = f"token {token}"

    # 기본 브랜치 확인
    branch = "main"
    try:
        r = requests.get(f"https://api.github.com/repos/{owner}/{repo}", headers=headers, timeout=10)
        if r.status_code == 200:
            branch = r.json().get("default_branch", "main")
    except Exception:
        pass

    if selected_files:
        # 사용자가 선택한 파일만 로드
        push_log(sess, f"[파싱] 선택된 파일 {len(selected_files)}개 로드 중...")
        for fpath in selected_files:
            raw_url = f"https://raw.githubusercontent.com/{owner}/{repo}/{branch}/{fpath}"
            try:
                r = requests.get(raw_url, headers=headers, timeout=15)
                if r.status_code == 200:
                    texts.append(f"## {fpath}\n\n{r.text}")
                    push_log(sess, f"[파싱] ✅ {fpath}")
                else:
                    push_log(sess, f"[파싱] ⚠️ {fpath} 로드 실패 (status {r.status_code})")
            except Exception as e:
                push_log(sess, f"[파싱] ⚠️ {fpath} 오류: {e}")
    else:
        # 선택 없으면 README + .md 최대 10개 자동
        for br in [branch, "main", "master"]:
            readme_url = f"https://raw.githubusercontent.com/{owner}/{repo}/{br}/README.md"
            try:
                resp = requests.get(readme_url, headers=headers, timeout=15)
                if resp.status_code == 200:
                    texts.append(f"## README.md\n\n{resp.text}")
                    push_log(sess, f"[파싱] README.md 로드 완료 (branch: {br})")
                    break
            except Exception:
                pass

        try:
            api_url = f"https://api.github.com/repos/{owner}/{repo}/git/trees/{branch}?recursive=1"
            resp = requests.get(api_url, headers=headers, timeout=15)
            if resp.status_code == 200:
                tree = resp.json().get("tree", [])
                md_files = [f["path"] for f in tree if f["path"].endswith(".md") and f["path"] != "README.md"][:10]
                push_log(sess, f"[파싱] GitHub 파일 트리: MD 파일 {len(md_files)}개 자동 선택")
                for fpath in md_files:
                    raw_url = f"https://raw.githubusercontent.com/{owner}/{repo}/{branch}/{fpath}"
                    try:
                        r = requests.get(raw_url, headers=headers, timeout=10)
                        if r.status_code == 200:
                            texts.append(f"## {fpath}\n\n{r.text}")
                    except Exception:
                        pass
        except Exception as e:
            push_log(sess, f"[파싱] GitHub API 오류 (무시): {e}")

    if not texts:
        raise RuntimeError(f"GitHub 리포지토리에서 콘텐츠를 가져올 수 없습니다: {owner}/{repo}")

    return "\n\n---\n\n".join(texts)


def fetch_web_page(sess: dict, url: str) -> str:
    """일반 웹 URL을 크롤링하여 텍스트 추출"""
    try:
        import requests
        import urllib3
    except ImportError:
        raise RuntimeError("requests 패키지가 필요합니다: pip install requests")
    push_log(sess, f"[파싱] 웹 크롤링 중: {url}")
    # TLS 검증: 기본 활성화. 사내 테스트 등 자가서명 인증서가 필요한 경우에만
    # TC_WEB_INSECURE=1 환경변수로 명시적으로 비활성화 가능.
    verify_tls = os.environ.get("TC_WEB_INSECURE", "").strip() not in ("1", "true", "yes")
    if not verify_tls:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        push_log(sess, f"[파싱] ⚠️ TLS 검증 비활성화됨 (TC_WEB_INSECURE=1)")
    try:
        headers = {"User-Agent": "Mozilla/5.0 (TC-Automation/2.0)"}
        resp = requests.get(url, headers=headers, timeout=30, verify=verify_tls)
        resp.raise_for_status()
        content_type = resp.headers.get("content-type", "")
        if "html" in content_type:
            return html_to_text(resp.text)
        else:
            return resp.text
    except requests.exceptions.SSLError as e:
        raise RuntimeError(
            f"TLS 인증서 검증 실패: {e}. "
            f"자가서명 인증서 환경이라면 TC_WEB_INSECURE=1 환경변수를 설정하세요."
        )
    except Exception as e:
        raise RuntimeError(f"웹 크롤링 실패: {e}")


def html_to_text(html: str) -> str:
    # script/style 제거
    html = re.sub(r"<script[^>]*>.*?</script>", "", html, flags=re.DOTALL | re.IGNORECASE)
    html = re.sub(r"<style[^>]*>.*?</style>", "", html, flags=re.DOTALL | re.IGNORECASE)
    # 태그 제거
    text = re.sub(r"<[^>]+>", " ", html)
    # 연속 공백 정리
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\n\s*\n\s*\n", "\n\n", text)
    return text.strip()


# ── 2단계: 정책 분석 ──────────────────────────────────────────────────────────
def step_policy(sess: dict, raw_text: str, project_name: str, focus_area: str = "") -> str:
    push_log(sess, "[정책] Claude AI로 정책 분석 중...")
    system = """당신은 소프트웨어 QA 전문가입니다.
주어진 문서에서 테스트 가능한 정책과 비즈니스 규칙을 추출합니다.
다음 형식으로 출력하세요:

# 정책 분석 결과

## 화면/페이지 인벤토리 (필수)
(문서에 등장하는 모든 화면/페이지를 빠짐없이 나열 — ID·이름·간단 설명)
(Splash, 진입 화면, 정적 화면, 브랜딩 화면, 빈 상태 화면 등 단순 화면도 반드시 포함)
(SCR-xxx, SCREEN-xxx 같은 코드가 있으면 그대로 사용)

## 대분류별 정책
(각 대분류별로 테스트 가능한 정책을 구체적으로 나열)

## 핵심 비즈니스 규칙
(서비스의 핵심 로직과 검증 포인트)

## 예외 처리 정책
(오류 케이스, 엣지 케이스, 유효성 검증 규칙)

⛔ 다음을 절대 드롭하지 마세요:
- 단순 UI 진입 화면 (Splash, 브랜딩 화면, Welcome, Landing)
- 정적 정보 화면 (About, Help, Empty State)
- 짧은 설명의 화면도 화면 인벤토리에 반드시 ID와 이름을 기록
- "테스트할 게 단순해 보여도" 생략하지 말 것 — UI 체크 TC 대상
"""
    focus_instruction = ""
    if focus_area:
        focus_instruction = f"""

⚠️ 중요: 사용자가 다음 기능에 대한 TC만 생성하려고 합니다. 해당 기능과 관련된 정책/규칙에 집중하세요:
→ {focus_area}

관련 없는 대분류나 기능의 정책은 간략하게만 다루고, 위 범위의 정책을 상세하게 분석하세요."""
    user = f"""프로젝트: {project_name}

다음 문서에서 TC 작성에 필요한 모든 정책과 규칙을 추출해주세요:

---
{raw_text[:200000]}
---

각 정책은 구체적이고 테스트 가능한 형태로 서술해주세요.{focus_instruction}

⚠️ 문서에 등장하는 모든 화면/기능/섹션을 빠짐없이 정리하세요. 주요한 것만 선별하지 말고 전체를 포괄하세요.
"""
    result = call_claude(system, user, max_tokens=16000)
    policy_path = sess["workspace"] / "02_policy.md"
    policy_path.write_text(result, encoding="utf-8")
    push_log(sess, f"[정책] 분석 완료 → {len(result):,}자")
    return result


# ── 3단계: 기능 목록 ──────────────────────────────────────────────────────────
def step_features(sess: dict, policy_text: str, project_name: str, focus_area: str = "") -> str:
    push_log(sess, "[기능] 기능 목록 생성 중...")
    system = """당신은 QA 엔지니어입니다. 정책 분석 결과를 바탕으로 테스트 가능한 기능 목록을 생성합니다.

다음 형식으로 feature_list.md를 작성하세요:

# Feature List — {프로젝트명}

## 대분류코드 — 대분류명

### FT-001 기능명
- 설명: (기능 설명)
- 테스트 포인트: (주요 검증 항목)
- 우선순위: High/Medium/Low

각 대분류의 모든 기능을 빠짐없이 나열하세요.

⛔ 필수: 정책 분석의 "화면/페이지 인벤토리"에 나열된 모든 화면을 각각 별도 기능으로 반드시 포함하세요.
- Splash 같은 단순 진입 화면도 "화면 진입/표시/버튼 동작" 같은 테스트 포인트로 기능화
- 정적 화면, Empty State, Welcome 화면도 모두 개별 기능으로 등록
- "테스트할 게 적어 보인다"는 이유로 생략 금지 — UI/UX 체크 TC 대상
"""
    focus_instruction = ""
    if focus_area:
        focus_instruction = f"""

⚠️ 중요: 다음 기능 범위에 해당하는 기능을 중심으로 목록을 작성하세요:
→ {focus_area}

해당 범위의 기능은 세부적으로 나열하고, 범위 밖의 기능은 포함하지 마세요."""
    user = f"""프로젝트: {project_name}

다음 정책 분석을 바탕으로 테스트해야 할 기능 목록을 작성해주세요:

---
{policy_text}
---

모든 주요 기능을 대분류별로 분류하여 나열하고, 각 기능의 테스트 포인트를 구체적으로 서술하세요.{focus_instruction}

⚠️ 정책 분석에 등장한 모든 화면/기능을 빠짐없이 나열하세요. 선별하지 말고 전체를 다루세요.
"""
    result = call_claude(system, user, max_tokens=16000)
    features_path = sess["workspace"] / "03_features.md"
    features_path.write_text(result, encoding="utf-8")
    push_log(sess, f"[기능] 기능 목록 완료 → {len(result):,}자")
    return result


def step_policy_features_combined(sess: dict, raw_text: str, project_name: str,
                                    focus_area: str = "") -> tuple[str, str]:
    """정책 반영 모드 최적화 — policy + features를 1회 호출로 통합 추출.
    반환: (policy_text, features_text).
    출력 형식은 기존 policy/features의 합집합 구조를 그대로 유지해 기존 코드 호환.
    """
    push_log(sess, "[정책+기능] 통합 추출 중... (1회 호출로 정책·기능 동시 생성)")
    system = """당신은 소프트웨어 QA 전문가입니다. 주어진 원문에서 테스트 준비에 필요한
정책·규칙과 기능 인벤토리를 **한 번에 통합 문서로** 추출합니다.

다음 정확한 형식으로 출력하세요:

# 통합 분석 결과 — {프로젝트명}

## [SECTION: POLICY] 화면/페이지 인벤토리 (필수)
(원문에 등장하는 모든 화면/페이지를 빠짐없이 나열 — ID · 이름 · 간단 설명 한 줄)
(Splash, 진입 화면, 정적 화면, 브랜딩 화면, 빈 상태 화면 등 단순 화면도 반드시 포함)
(SCR-xxx, SCREEN-xxx 같은 코드가 있으면 그대로 사용)

## [SECTION: POLICY] 대분류별 정책
(각 대분류별로 테스트 가능한 정책을 구체적으로 나열)

## [SECTION: POLICY] 핵심 비즈니스 규칙
(서비스의 핵심 로직과 검증 포인트)

## [SECTION: POLICY] 예외 처리 정책
(오류 케이스, 엣지 케이스, 유효성 검증 규칙)

---

# Feature List — {프로젝트명}

## [SECTION: FEATURES] 대분류코드 — 대분류명

### FT-001 기능명
- 설명: (기능 설명)
- 테스트 포인트: (주요 검증 항목)
- 우선순위: High/Medium/Low

(각 대분류·화면의 모든 기능을 빠짐없이 나열)

⛔ 다음을 절대 드롭하지 마세요:
- 단순 UI 진입 화면 (Splash, Welcome, Landing, 브랜딩)
- 정적 정보 화면 (About, Help, Empty State)
- 각 화면의 "진입·표시·버튼 동작" 같은 단순해 보이는 케이스
- 화면 인벤토리에 등록된 모든 화면은 Feature List에도 최소 1개 기능으로 반드시 등록
"""
    focus_instruction = ""
    if focus_area:
        focus_instruction = f"""

⚠️ 포커스 범위: 아래 범위에 해당하는 화면/기능만 포함하세요.
→ {focus_area}"""
    user = f"""프로젝트: {project_name}

아래 원문에서 (1) 정책·규칙 분석 + (2) 기능 인벤토리를 **한 문서로 통합** 추출하세요.

---
{raw_text[:200000]}
---
{focus_instruction}

⚠️ 문서에 등장하는 모든 화면/기능/섹션을 빠짐없이 다루세요. 선별하지 마세요.
⚠️ 반드시 위의 출력 형식(SECTION 마커 포함)을 정확히 따르세요. 후처리가 의존합니다.
"""
    result = call_claude(system, user, max_tokens=16000)

    # 결과를 policy_text / features_text 로 분리
    # `---` 또는 `# Feature List` 기준으로 나눔. 실패 시 전체를 둘 다에 넣어 안전하게 처리.
    policy_text = result
    features_text = result
    # 분리 마커: "# Feature List" (독립 # 헤더)
    m = re.search(r"^# Feature List\b", result, re.MULTILINE)
    if m:
        policy_text = result[:m.start()].rstrip()
        features_text = result[m.start():].strip()

    # 저장 (디버깅용)
    combined_path = sess["workspace"] / "02_combined.md"
    combined_path.write_text(result, encoding="utf-8")
    (sess["workspace"] / "02_policy.md").write_text(policy_text, encoding="utf-8")
    (sess["workspace"] / "03_features.md").write_text(features_text, encoding="utf-8")

    push_log(sess, f"[정책+기능] 통합 추출 완료 → 정책 {len(policy_text):,}자 · 기능 {len(features_text):,}자")
    return policy_text, features_text


# ── 4단계: 분류표 생성 ────────────────────────────────────────────────────────
def step_classify(sess: dict, features_text: str, project_name: str, focus_area: str = "") -> str:
    push_log(sess, "[분류] 계층 분류표 생성 중...")
    system = """당신은 TC 분류 전문가입니다. 기능 목록을 바탕으로 대분류/중분류/소분류 계층 분류표를 생성합니다.

다음 형식으로 출력하세요:

# TC 분류표 — {프로젝트명}

## 대분류: {대분류명}

### 중분류: {중분류명}

#### 소분류
- {소분류명}: {설명}

규칙:
- 대분류는 비즈니스 영역 단위 (예: Onboarding, Authentication, Trading)
- 중분류는 주요 기능 단위
- 소분류는 세부 케이스 단위
- ⛔ 대분류/중분류/소분류 이름에 괄호 코드를 붙이지 마세요. 예: "FOOTER (FOOT)" ❌ → "Footer" ✅
- ⛔ TC ID를 생성하지 마세요. TC ID는 이후 단계에서 검토자가 결정합니다.
- ⛔ TC ID 생성 규칙, TC ID 예시표, 기능-TC 매핑표를 포함하지 마세요.
- 분류표에는 대분류/중분류/소분류 구조만 출력하세요.

⛔ 필수 완전성:
- 기능 목록에 등장하는 모든 화면을 중분류로 반드시 포함하세요.
- Splash, Login Options, Welcome, Landing, Empty State 같은 단순 UI 진입 화면도 반드시 중분류로 포함
- 온보딩 흐름의 모든 화면(SCR-xxx 등 코드가 있으면 해당 화면명)을 누락 없이 분류
- "단순해 보여서 생략"은 금지 — UI/UX 체크 TC 대상이므로 반드시 분류표에 등록
"""
    focus_instruction = ""
    if focus_area:
        focus_instruction = f"""

⚠️ 중요: 다음 범위에 해당하는 기능의 분류표만 생성하세요:
→ {focus_area}

범위 밖의 기능은 분류표에 포함하지 마세요."""
    # 프로젝트별 카테고리 참조 데이터 로드
    project_policies = load_project_policies(project_name)
    category_ref = ""
    if project_policies:
        category_ref = f"""

⚠️ 기존 프로젝트 카테고리 참조: 아래 기존 대분류/중분류 이름과 동일한 용어를 사용하세요.
{project_policies[:4000]}"""
    user = f"""프로젝트: {project_name}

다음 기능 목록을 대분류/중분류/소분류 계층으로 분류해주세요:

---
{features_text}
---

분류 결과는 TC ID 생성의 기반이 되므로, 명확하고 일관성 있는 코드를 사용하세요.{focus_instruction}{category_ref}

⚠️ 문서 전체를 빠짐없이 분석하여 모든 대분류와 중분류를 포함하세요.
"""
    result = call_claude(system, user, max_tokens=16000)
    classify_path = sess["workspace"] / "04_classification_draft.md"
    classify_path.write_text(result, encoding="utf-8")
    push_log(sess, f"[분류] 분류표 생성 완료 → {len(result):,}자")
    return result


# ── Quick 모드 전용 3단계: 인벤토리 → 분류 → 섹션발췌 ──────────────────────
# 원문을 요약하지 않고 "항목 목록만" 먼저 추출한 뒤, 항목별로 원문의 해당 섹션만
# 발췌해 TC 작성에 전달한다. 총 토큰은 기존 Quick과 비슷하거나 적고, 누락이 줄어든다.

def step_quick_inventory(sess: dict, raw_text: str, project_name: str,
                          focus_area: str = "") -> str:
    """Quick 모드 1단계 — 원문에서 모든 화면/기능 인벤토리만 추출.
    요약이 아니라 '목록'이어서 AI가 축약할 여지가 적다."""
    push_log(sess, "[Quick 1/3] 인벤토리 추출 중...")
    system = """당신은 기획 문서 스캐너입니다. 주어진 원문에서 테스트 가능한 모든 항목(화면/페이지/기능)의 인벤토리만 추출합니다.

출력 형식 (정확히 이 형식만):
# 인벤토리

## 화면 목록
- {ID} | {이름} | {한 줄 설명}
- ...

## 독립 기능 목록 (화면에 속하지 않는 기능)
- {이름} | {한 줄 설명}
- ...

규칙:
- 화면 ID(SCR-xxx, SCREEN-xxx 등)가 원문에 있으면 그대로 쓰세요. 없으면 비워 두세요.
- 각 항목은 한 줄로 간결하게. 상세 설명·정책·절차는 절대 포함하지 마세요.
- ⛔ Splash, Empty State, Landing, Welcome, 진입 화면, 정적 화면도 **반드시** 포함하세요.
- ⛔ 원문에 등장하는 모든 화면을 빠짐없이 수집하세요. "단순해 보인다"고 생략 금지.
- ⛔ 요약·정리·설명 문구를 추가하지 마세요. 항목 목록만 출력.
- 일반적으로 원문 기획서 1개당 30~100개의 항목이 나옵니다. 10개 미만이면 뭔가 놓친 것입니다.
"""
    focus_instruction = ""
    if focus_area:
        focus_instruction = f"""

⚠️ 포커스 범위: 아래 범위에 해당하는 항목만 인벤토리에 포함하세요.
→ {focus_area}"""
    user = f"""프로젝트: {project_name}

아래 원문에서 화면/기능 인벤토리를 추출하세요.

---
{raw_text[:200000]}
---
{focus_instruction}

⚠️ 빠짐없이 모든 화면을 수집하세요. Splash, Login Options, 각 Onboarding 단계 화면, Empty State 등 **단순 화면도 반드시** 포함.
"""
    result = call_claude(system, user, max_tokens=8000)
    inv_path = sess["workspace"] / "02_inventory.md"
    inv_path.write_text(result, encoding="utf-8")
    # 간단한 카운트 로그
    item_count = len(re.findall(r"^-\s+", result, re.MULTILINE))
    push_log(sess, f"[Quick 1/3] 인벤토리 완료 — 항목 {item_count}개, {len(result):,}자")
    return result


def step_classify_from_inventory(sess: dict, inventory_text: str,
                                  project_name: str, focus_area: str = "") -> str:
    """Quick 모드 2단계 — 인벤토리 → 분류표.
    원문을 입력하지 않으므로 AI가 축약하지 않고 모든 항목을 분류한다."""
    push_log(sess, "[Quick 2/3] 분류표 생성 중...")
    system = """당신은 TC 분류 전문가입니다. 주어진 인벤토리의 모든 항목을 대분류/중분류/소분류 계층으로 분류합니다.

다음 형식으로 출력하세요:

# TC 분류표 — {프로젝트명}

## 대분류: {대분류명}

### 중분류: {중분류명}

#### 소분류
- {소분류명}: {설명}

규칙:
- 대분류는 비즈니스 영역 단위 (예: Onboarding, Authentication, Trading)
- 중분류는 화면/주요 기능 단위 (인벤토리의 각 화면/기능이 중분류가 됨)
- 소분류는 해당 중분류의 테스트 가능한 세부 케이스
- ⛔ 대분류/중분류/소분류 이름에 괄호 코드를 붙이지 마세요. 예: "FOOTER (FOOT)" ❌ → "Footer" ✅
- ⛔ TC ID를 생성하지 마세요.
- ⛔ TC ID 생성 규칙, TC ID 예시표, 기능-TC 매핑표를 포함하지 마세요.
- 분류표에는 대분류/중분류/소분류 구조만 출력하세요.

⛔ 필수 완전성 (매우 중요):
- 인벤토리의 **모든 항목을 반드시 중분류로** 포함하세요. 하나도 생략하지 마세요.
- Splash, Login Options, Welcome, Landing, Empty State 등 단순 화면도 반드시 중분류로 등록
- 인벤토리 항목 수 ≒ 중분류 수 (통상)
- 중분류가 10개 미만이면 뭔가 빠뜨린 것입니다.
"""
    focus_instruction = ""
    if focus_area:
        focus_instruction = f"""

⚠️ 포커스 범위: 인벤토리 중 아래 범위에 해당하는 항목만 분류하세요.
→ {focus_area}"""
    project_policies = load_project_policies(project_name)
    category_ref = ""
    if project_policies:
        category_ref = f"""

⚠️ 기존 프로젝트 카테고리 참조: 아래 기존 대분류/중분류 이름과 동일한 용어를 사용하세요.
{project_policies[:4000]}"""
    user = f"""프로젝트: {project_name}

아래 인벤토리의 모든 항목을 대분류/중분류/소분류 계층으로 분류해주세요:

---
{inventory_text}
---
{focus_instruction}{category_ref}

⚠️ 인벤토리의 **모든 항목**을 반드시 포함하세요. 빠뜨리면 안 됩니다.
"""
    result = call_claude(system, user, max_tokens=16000)
    classify_path = sess["workspace"] / "04_classification_draft.md"
    classify_path.write_text(result, encoding="utf-8")
    push_log(sess, f"[Quick 2/3] 분류표 생성 완료 → {len(result):,}자")
    return result


def extract_section_from_raw(raw_text: str, middle_name: str,
                              major_name: str = "", max_chars: int = 8000) -> str:
    """Quick 모드 3단계용 — 원문에서 middle_name과 관련된 섹션만 발췌한다.
    매칭 우선순위:
      1. 정확한 middle_name을 포함한 헤더(###/##)
      2. middle_name의 단어들을 포함한 헤더
      3. major_name 매칭 (백업)
    찾으면 해당 헤더부터 다음 같은 레벨 헤더까지 반환.
    못 찾으면 middle_name이 언급된 주변 ±500자 컨텍스트 반환.
    """
    if not raw_text or not middle_name:
        return ""
    lines = raw_text.splitlines()

    def header_level(line: str) -> int:
        m = re.match(r"^(#+)\s", line)
        return len(m.group(1)) if m else 0

    # 후보 검색어: 원문 middle_name + 공백 대체 + 단어 분할
    candidates = [middle_name.strip()]
    if "/" in middle_name:
        candidates.extend(p.strip() for p in middle_name.split("/") if p.strip())
    # 영단어 3자 이상만 개별 매칭 (너무 일반적인 단어 방어)
    words = [w for w in re.findall(r"[A-Za-z][A-Za-z0-9]{2,}", middle_name) if len(w) > 2]
    candidates.extend(words)

    best_span = None
    best_score = 0
    for i, line in enumerate(lines):
        lvl = header_level(line)
        if lvl == 0:
            continue
        # 헤더 텍스트 추출
        header_text = re.sub(r"^#+\s+", "", line).strip()
        # 점수 산정: middle_name 완전 매칭 > 부분 매칭 > 단어 매칭
        score = 0
        if middle_name.lower() in header_text.lower():
            score = 100
        else:
            for w in candidates:
                if w and w.lower() in header_text.lower():
                    score = max(score, 50)
        if score == 0 and major_name and major_name.lower() in header_text.lower():
            score = 10

        if score > best_score:
            # 같은 레벨 다음 헤더까지 발췌
            end = len(lines)
            for j in range(i + 1, len(lines)):
                jlvl = header_level(lines[j])
                if jlvl > 0 and jlvl <= lvl:
                    end = j
                    break
            best_span = (i, end)
            best_score = score
            if score >= 100:
                # 완전 매칭이면 더 찾지 않음
                break

    if best_span:
        i, end = best_span
        section = "\n".join(lines[i:end])
        return section[:max_chars]

    # Fallback: middle_name 언급 위치 ±500자
    pat = re.escape(middle_name)
    m = re.search(pat, raw_text, re.IGNORECASE)
    if m:
        start = max(0, m.start() - 500)
        end = min(len(raw_text), m.end() + 2000)
        return raw_text[start:end][:max_chars]
    return ""


# ── 5단계: Human Gate (블로킹) ────────────────────────────────────────────────
def step_gate(sess: dict, classification: str):
    push_log(sess, "[GATE] 분류표 검토 대기 중... 사용자 승인 필요")
    sess["status"] = "gate_waiting"
    # 입력 소스 SCR 매핑도 함께 전달 — 프론트의 Gate UI에서 안내문 노출
    _scr_map = sess.get("_source_scr_map") or {}
    # v0.9.22: Stage 1 분류표 중복 가능성 예측
    try:
        prediction = predict_classification_duplicates(classification)
    except Exception as _e:
        push_log(sess, f"[GATE] 중복 예측 스킵: {_e}")
        prediction = {"predictions": [], "high_risk_count": 0}
    if prediction.get("high_risk_count"):
        push_log(sess, f"[GATE] 분류표 중복 가능성 {prediction['high_risk_count']}개 그룹 탐지")
    push(sess, "gate", {
        "content": classification,
        "source_scr_map": _scr_map,
        "duplicate_prediction": prediction,
    })
    # 사용자 승인까지 블로킹
    sess["gate_event"].wait()
    approved_content = sess["approved"]
    push_log(sess, "[GATE] 분류표 승인됨. TC 작성 시작.")
    # 승인된 분류표 저장
    approved_path = sess["workspace"] / "classification_v1_APPROVED.md"
    approved_path.write_text(approved_content, encoding="utf-8")
    return approved_content


# ── 6단계: TC 작성 ────────────────────────────────────────────────────────────
def step_write_tc_per_screen(sess: dict, approved_classification: str,
                              spec_data: dict, project_name: str,
                              selected_domain_codes=None) -> tuple:
    """구조화 spec 모드 전용 — 화면(SCR) 1개당 LLM 1회 호출.

    Args:
        approved_classification: Human Gate 통과한 분류표 (build_classification_from_screen_list 결과).
        spec_data:               step_parse_structured_spec() 반환값.
        selected_domain_codes:   대분류 필터 (None=전체).
    Returns: (merged_tc_md, total_tc_count)

    핵심 차이점 (vs step_write_tc):
      - 정책/디자인을 시스템 프롬프트에 cache_control 로 1회만 전송 → 화면 N개 호출에 캐시 히트
      - 화면별 user prompt 는 해당 SCR 본문 전문 + 자동 추출한 cross-ref 인용만 포함
      - 화면 단위 매핑이 정확 (휴리스틱 발췌 X)
    """
    push_log(sess, "[화면별 TC 작성] 시작 — 구조화 spec 모드")
    tc_rules = load_tc_rules()
    fewshot = load_fewshot_examples()
    project_policies = load_project_policies(project_name)

    # 1) 시스템 프롬프트 블록 구성 — cache_control 적용 + 화면 격리 강화 (v0.10.x)
    #
    # 변경: 분류표는 시스템 블록에서 제거하고 user prompt 에서 화면별 축소 버전으로 주입.
    #       이유: SCR 단독 호출 vs 일괄 호출의 결과 갯수 편차 줄이기 (다른 화면 정보가
    #       AI 의 분배·중복 회피 동기를 만들어 갯수가 줄어드는 현상).
    #
    # ⚠️ Anthropic 제약: cache_control 블록은 최대 4개.
    #    캐시 #1: tc_rules + 카테고리 가이드 + 형식 규칙 (build_tc_system_prompt 의 분류표 제외 부분)
    #    캐시 #2: spec policy_text (전문)
    #    캐시 #3: spec design_text (전문)
    #    캐시 #4: 화면별 호출 모드 안내 (B-1 단독 호출 명시)
    sys_blocks = []
    # 분류표 자리에 placeholder 만 남기고 build_tc_system_prompt() 호출 → 화면별 축소 버전은 user 로 이동.
    base_system_prompt = build_tc_system_prompt(
        tc_rules=tc_rules,
        classification="(분류표는 user 메시지에 화면별 축소 버전으로 제공됩니다 — 그 화면만 처리하세요)",
        project_policies=project_policies,
        fewshot=fewshot,
    )
    sys_blocks.append({
        "type": "text",
        "text": base_system_prompt,
        "cache_control": {"type": "ephemeral"},  # 캐시 #1
    })
    if spec_data.get("policy_text"):
        sys_blocks.append({
            "type": "text",
            "text": f"\n## 공통 정책 문서 (구조화 spec 의 policy md 전문)\n{spec_data['policy_text'][:20000]}\n",
            "cache_control": {"type": "ephemeral"},  # 캐시 #2
        })
    if spec_data.get("design_text"):
        sys_blocks.append({
            "type": "text",
            "text": f"\n## 디자인 시스템 문서 (구조화 spec 의 design md 전문 — 토큰/컴포넌트 참조)\n{spec_data['design_text'][:20000]}\n",
            "cache_control": {"type": "ephemeral"},  # 캐시 #3
        })
    # B-1 단독 호출 명시 + 격리 강화 — AI 자기검열 차단
    sys_blocks.append({
        "type": "text",
        "text": (
            "\n## 화면별 단독 호출 모드 안내 (반드시 준수)\n\n"
            "**핵심 원칙**: 이 호출은 **하나의 화면(SCR)에 대한 단독 작업** 입니다. "
            "프로젝트에 다른 화면이 존재하더라도 이 호출에서는 **존재하지 않는 것처럼** 처리하세요. "
            "당신은 이 화면 전담 QA 엔지니어이고, 다른 화면은 다른 사람이 담당합니다.\n\n"
            "**금지 행동**:\n"
            "- 다른 화면에 만들 TC 와의 중복 회피를 의식하지 마세요 (중복은 후처리에서 자동 제거됨)\n"
            "- TC 갯수를 다른 화면들과 분배해서 줄이지 마세요\n"
            "- 'A 도 B 도 비슷하니 하나로 묶자' 같은 통합 시도 금지\n\n"
            "**지향**: user 메시지의 화면 명세에 등장하는 **모든** 상태/인터랙션/비고 마커를 "
            "각각 독립적인 TC 로 변환하세요. 카테고리 5가지(18/32/25/18/5~10%)와 카테고리 3·4 의무 포함은 그대로 유지. "
            "카테고리 5(연결 시나리오)는 이 화면이 시작점/종착지일 때만 작성, 중간 화면은 skip.\n"
        ),
        "cache_control": {"type": "ephemeral"},  # 캐시 #4
    })

    # 2) 처리 대상 화면 필터링 (selected_domain_codes)
    all_domains = extract_domains(approved_classification)
    if selected_domain_codes:
        domain_codes_set = set(selected_domain_codes)
        # 분류표의 대분류명 → 화면 매핑
    else:
        domain_codes_set = None

    # screen_rows 의 major(대분류) 와 domain 의 name 매칭으로 SuiteCode 결정 (대분류 단위 폴백용)
    suite_codes = sess.get("suite_codes", [])
    domain_to_suite: dict[str, str] = {}
    for i, dom in enumerate(all_domains):
        if i < len(suite_codes) and suite_codes[i]:
            domain_to_suite[dom["name"]] = suite_codes[i].strip().strip("-")
        else:
            domain_to_suite[dom["name"]] = dom["code"]

    # ── v0.10.x: 화면(SCR)별 고유 SuiteCode 결정 — 기존 인프라 재사용 ─────────
    # tc-rules.md §2-1 유형 B (Screen-based) 규칙 적용.
    # SM/SA 프로젝트는 SuiteCode 자리에 ScreenCode 를 사용하고 NNN 은 화면 내 001~.
    # 우선순위:
    #   1) projects/{project}/screen_code_map.md 에 등록된 ScreenCode (사용자 합의)
    #   2) 미등록 화면은 resolve_screen_code() 자동 파생 (기존 함수 재사용)
    # 두 단계 모두 used_codes 충돌 회피 내장.
    project_code_for_id = _detect_project_code(project_name)
    is_screen_mode = _is_screen_based(project_code_for_id)
    screen_map = load_screen_code_map(project_name) if is_screen_mode else {}

    rows_by_id_pre = {r["id"]: r for r in spec_data.get("screen_rows", [])}
    screen_to_suite: dict[str, str] = {}
    used_screen_codes: set[str] = set()
    # resume 시 같은 결과가 나오도록 ID 정렬 보장
    sorted_screens = sorted(spec_data["screens"], key=lambda sc: sc["id"])
    for sc in sorted_screens:
        meta_row = rows_by_id_pre.get(sc["id"], {})
        # screen_code_map.md 는 'Middle Name' 키로 매핑 — 화면명(title) 또는 중분류 사용.
        # 우선 title 로 시도 (Login, Email Input 등) → 없으면 SCR-ID 자체로 시도.
        middle_name = (sc.get("title") or meta_row.get("name") or sc["id"]).strip()
        if is_screen_mode:
            # 기존 resolve_screen_code() 가 1) screen_map 정확 일치 2) 자동 파생
            #   3) used_codes 충돌 회피 모두 처리
            code = resolve_screen_code(middle_name, screen_map, used_screen_codes)
        else:
            # Suite-based 프로젝트(SC, PC Web)는 대분류 단위 SuiteCode 사용
            major = meta_row.get("major", "")
            code = domain_to_suite.get(major, "FUNC") or "FUNC"
        used_screen_codes.add(code)
        screen_to_suite[sc["id"]] = code

    # screens 와 screen_rows 를 ID 로 조인
    rows_by_id = {r["id"]: r for r in spec_data.get("screen_rows", [])}
    target_screens = []
    for sc in spec_data["screens"]:
        check_stop(sess)
        meta_row = rows_by_id.get(sc["id"], {})
        major = meta_row.get("major") or ""
        if domain_codes_set is not None:
            # selected_domain_codes 는 코드 기반이므로 코드/이름 매칭 모두 시도
            matched_dom = next((d for d in all_domains if d["name"] == major or d["code"] == major), None)
            if not matched_dom or matched_dom["code"] not in domain_codes_set:
                continue
        target_screens.append((sc, meta_row, major))

    if not target_screens:
        raise RuntimeError("처리할 화면이 없습니다. 분류표/필터를 확인하세요.")

    push_log(sess, f"[화면별 TC 작성] 대상 화면 {len(target_screens)}개")
    project_code = _detect_project_code(project_name)

    # resume 지원 — 이미 처리된 화면은 건너뛴다.
    # 화면별 결과는 sess["workspace"]/per_screen/<SCR-ID>.md 에 저장된다.
    per_screen_dir = sess["workspace"] / "per_screen"
    per_screen_dir.mkdir(exist_ok=True)
    completed_ids = {p.stem for p in per_screen_dir.glob("SCR-*.md")}
    if completed_ids:
        push_log(sess, f"[화면별 TC 작성] resume — 이미 완료된 화면 {len(completed_ids)}개 건너뜀: {sorted(completed_ids)[:5]}{'...' if len(completed_ids) > 5 else ''}")

    tc_parts: list[str] = []
    seq_per_suite: dict[str, int] = {}
    total_tc = 0

    # 이미 완료된 화면의 결과를 먼저 메모리에 적재
    # 유형 B (Screen-based): 각 화면이 고유 SuiteCode + 001~ 라 누적 의미 없음. 그래도
    # 다음 화면 호출에서 seq_per_suite 가 의도치 않게 사용될 가능성 차단을 위해
    # 화면 단위로는 +1 만 표기하고 실제 starting 결정은 위쪽에서 is_screen_mode 분기.
    # 유형 A (Suite-based): 같은 SuiteCode 화면들이 누적되도록 + scr_tc_count.
    for sc, meta_row, major in target_screens:
        if sc["id"] not in completed_ids:
            continue
        existing_path = per_screen_dir / f"{sc['id']}.md"
        existing_text = existing_path.read_text(encoding="utf-8")
        suite_code = screen_to_suite.get(sc["id"]) or domain_to_suite.get(major, "FUNC") or "FUNC"
        scr_tc_count = count_unique_tc_ids(existing_text)
        if is_screen_mode:
            seq_per_suite[suite_code] = scr_tc_count + 1  # 참고용 (실제 reuse X)
        else:
            seq_per_suite[suite_code] = seq_per_suite.get(suite_code, 1) + scr_tc_count
        total_tc += scr_tc_count
        tc_parts.append(f"<!-- {sc['id']} — {sc['title']} (resumed) -->\n\n{existing_text.strip()}")

    pending = [(sc, mr, mj) for (sc, mr, mj) in target_screens if sc["id"] not in completed_ids]
    push_log(sess, f"[화면별 TC 작성] 대기 화면 {len(pending)}개 (총 {len(target_screens)}개 중)")

    for idx_pending, (sc, meta_row, major) in enumerate(pending, 1):
        check_stop(sess)
        idx = len(completed_ids) + idx_pending  # 전체 진행 인덱스
        suite_code = screen_to_suite.get(sc["id"]) or domain_to_suite.get(major, "FUNC") or "FUNC"
        # tc-rules.md §2-1:
        #   - 유형 B (Screen-based, SM/SA): NNN 은 화면 내 유니크 → 항상 1 부터
        #   - 유형 A (Suite-based, SC):     NNN 은 Suite 내 전역 유니크 → seq_per_suite 누적
        if is_screen_mode:
            starting = 1
        else:
            starting = seq_per_suite.get(suite_code, 1)

        # cross-ref 인용 추출
        policy_excerpts = extract_policy_excerpts(spec_data.get("policy_text", ""), sc["ref_keywords"])

        # A-1: 분류표 — 해당 화면 섹션만 추출. 못 찾으면 폴백으로 짧은 헤더만.
        screen_section = _extract_screen_classification_section(
            approved_classification, sc["id"], sc.get("title", "")
        )
        if not screen_section:
            screen_section = f"## 대분류: {major}\n### 중분류: {sc.get('title', sc['id'])}\n"

        user = build_screen_user_prompt(
            screen_meta=sc,
            project_name=project_name,
            project_code=project_code,
            suite_code=suite_code,
            starting_seq=starting,
            policy_excerpts=policy_excerpts,
            design_excerpts=None,  # 디자인은 system 캐시에 들어가 있어서 별도 발췌 불필요
            screen_classification_section=screen_section,
        )

        push(sess, "stage", {
            "stage": 4, "label": f"[{idx}/{len(target_screens)}] {sc['id']} TC 작성 중...",
            "pct": 55 + int(25 * (idx / max(len(target_screens), 1))),
        })

        # max_tokens=20000 — Anthropic SDK 의 'Streaming required >10min' 임계 회피.
        # 32000 까지 올렸더니 SDK 가 streaming 강제 → 21333 (≈ 8K tps × 약 9분) 이하로 유지.
        # 안전 마진 + 미래 모델 속도 변화 고려해 20000.
        # 현재 사용량: TC 56개 시 ~14K 토큰 → 20K 한도로도 충분히 여유.
        tc_draft = call_claude_cached(sys_blocks, user, max_tokens=20000)

        # 잘림 감지 + 자동 보충 호출 (기존 step_write_tc 와 동등 처리).
        # 증상 예시: 마지막 TC 가 '| 플랫' 같이 표 중간에서 끊기면 사전조건/스텝/기대결과
        # 컬럼이 빈 채로 Excel 에 들어감. 마지막 ### 이후 본문에 필수 섹션 없으면 보충.
        last_tc_match = list(re.finditer(r"^###\s", tc_draft, re.MULTILINE))
        if last_tc_match:
            last_tc_block = tc_draft[last_tc_match[-1].start():]
            if "**테스트 단계**" not in last_tc_block or "**예상 결과**" not in last_tc_block:
                push_log(sess, f"[화면별 TC 작성] {sc['id']} 마지막 TC 불완전 — 이어서 생성 중...")
                try:
                    continuation = call_claude_cached(
                        sys_blocks,
                        f"이전 응답이 max_tokens 한도로 잘렸습니다. 아래 미완성 TC 를 그대로 이어서 완성하세요. "
                        f"새 TC 추가 금지, 표/사전조건/테스트 단계/예상 결과/비고 누락 없이.\n\n---\n{last_tc_block}",
                        max_tokens=8000,  # 미완성 TC 1~2개 완성용 — 충분
                    )
                    tc_draft = tc_draft[:last_tc_match[-1].start()] + continuation
                    push_log(sess, f"[화면별 TC 작성] {sc['id']} 보충 완료")
                except Exception as e:
                    push_log(sess, f"[화면별 TC 작성] {sc['id']} 보충 실패: {e} — 잘린 채 진행")

        # C-2: 응답 말미의 '체크리스트 처리 결과' 섹션 분리 (TC 본문 영향 차단)
        tc_clean, checklist_report = _split_checklist_report(tc_draft)
        if checklist_report:
            (per_screen_dir / f"{sc['id']}_checklist.md").write_text(checklist_report, encoding="utf-8")

        # 화면별 즉시 저장 — resume 시 여기서 다시 시작
        (per_screen_dir / f"{sc['id']}.md").write_text(tc_clean, encoding="utf-8")

        # TC 개수 카운트 — 유형별 seq 누적 정책
        scr_tc_count = count_unique_tc_ids(tc_clean)
        if is_screen_mode:
            # 화면 단위 리셋 — 누적 의미 없으므로 참고용으로만
            seq_per_suite[suite_code] = scr_tc_count + 1
        else:
            # Suite-based: 같은 SuiteCode 끼리 NNN 전역 유니크 → 다음 화면 starting 누적
            seq_per_suite[suite_code] = starting + scr_tc_count
        total_tc += scr_tc_count

        # 화면 헤더 + 본문
        tc_parts.append(f"<!-- {sc['id']} — {sc['title']} -->\n\n{tc_clean.strip()}")
        # 체크리스트 처리 보고 요약 — 누락 항목 카운트
        if checklist_report:
            n_done = checklist_report.count("✓")
            n_miss = checklist_report.count("✗")
            push_log(sess,
                f"[화면별 TC 작성] [{idx}/{len(target_screens)}] {sc['id']} ({suite_code}) 완료 — "
                f"TC {scr_tc_count}개 (체크리스트 {n_done}✓ / {n_miss}✗, 누적 {total_tc})")
        else:
            push_log(sess,
                f"[화면별 TC 작성] [{idx}/{len(target_screens)}] {sc['id']} ({suite_code}) 완료 — "
                f"TC {scr_tc_count}개 (누적 {total_tc})")

    merged = "\n\n---\n\n".join(tc_parts)
    draft_path = sess["workspace"] / "tc_draft_per_screen.md"
    draft_path.write_text(merged, encoding="utf-8")
    push_log(sess, f"[화면별 TC 작성] 전체 완료 — 화면 {len(target_screens)}개, TC {total_tc}개")

    # 정책 A 후처리 검증: 같은 화면 안에서 예상 결과가 동일한 TC 의심 그룹 검출.
    # 자동 병합은 하지 않고 리포트로만 알림 (false positive 위험).
    try:
        dup_report = detect_same_result_duplicates(merged)
        if dup_report["total_suspect"] > 0:
            report_path = sess["workspace"] / "duplicate_suspects.md"
            lines = ["# 중복 의심 TC 그룹 (검토 필요)\n"]
            lines.append(f"동일 화면 내 같은 검증 결과를 갖는 TC 그룹 {len(dup_report['groups'])}개 / 통합 가능 TC {dup_report['total_suspect']}개\n")
            for i, g in enumerate(dup_report["groups"], 1):
                lines.append(f"\n## 그룹 {i} — {g['middle']}\n")
                lines.append(f"**공통 예상 결과**: {g['common_result']}\n")
                lines.append("**해당 TC**:")
                for tc in g["tcs"]:
                    lines.append(f"- `{tc['id']}` — {tc['title']}")
                lines.append("\n→ 검토 후 1개로 통합하고 테스트 단계에 다른 트리거를 추가하는 것을 권장")
            report_path.write_text("\n".join(lines), encoding="utf-8")
            push_log(sess,
                f"⚠️ [중복 의심] 같은 화면 내 동일 결과 TC {dup_report['total_suspect']}개 검출 — "
                f"{report_path.name} 참고")
        # 세션에 저장 — UI 알림용
        sess["_dup_same_result_report"] = dup_report
    except Exception as e:
        push_log(sess, f"[중복 검증] 스킵: {e}")

    return merged, total_tc


def step_write_tc(sess: dict, approved_classification: str, features_text: str,
                  policy_text: str, project_name: str,
                  selected_domain_codes=None) -> tuple:
    push_log(sess, "[TC 작성] TC 초안 작성 시작...")
    tc_rules = load_tc_rules()
    fewshot = load_fewshot_examples()
    project_policies = load_project_policies(project_name)
    if project_policies:
        push_log(sess, f"[TC 작성] 프로젝트 정책 로드됨: {project_name}")
    if fewshot:
        push_log(sess, f"[TC 작성] Few-shot 예시 로드됨")

    # 대분류 목록 추출
    all_domains = extract_domains(approved_classification)

    # 범위 필터 적용
    if selected_domain_codes:
        domains = [d for d in all_domains if d["code"] in selected_domain_codes]
        skipped = [d["code"] for d in all_domains if d["code"] not in selected_domain_codes]
        if skipped:
            push_log(sess, f"[TC 작성] 범위 제외 대분류: {', '.join(skipped)}")
    else:
        domains = all_domains

    if not domains:
        raise RuntimeError("선택된 대분류가 없습니다. 범위를 다시 확인하세요.")

    # SuiteCode 매핑 (Human Gate에서 검토자가 입력한 코드)
    suite_codes = sess.get("suite_codes", [])
    for i, domain in enumerate(domains):
        if i < len(suite_codes) and suite_codes[i]:
            domain["suite_code"] = suite_codes[i]
    if suite_codes:
        push_log(sess, f"[TC 작성] SuiteCode 매핑: {', '.join(d.get('suite_code', '?') for d in domains)}")

    push_log(sess, f"[TC 작성] 생성 대상 대분류 {len(domains)}개: {', '.join(d['code'] for d in domains)}")

    all_tc_parts = []
    total_tc = 0

    # 중분류 목록 추출 (분류표에서)
    def extract_middles_from_classification(classification_text: str, domain_code: str) -> list[str]:
        domain_section = extract_domain_section(classification_text, domain_code)
        middles = []
        for line in domain_section.splitlines():
            line = line.strip()
            if line.startswith("### "):
                mid_name = re.sub(r"^###\s+", "", line).strip()
                mid_name = re.sub(r"^중분류[:\s]*", "", mid_name).strip()
                mid_name = re.sub(r"\s*[\(\（][A-Z]{2,8}[\)\）]", "", mid_name).strip()
                if mid_name:
                    middles.append(mid_name)
        return middles

    for i, domain in enumerate(domains):
        domain_code = domain["code"]
        domain_name = domain["name"]

        # 중분류별 분할 호출
        middles = extract_middles_from_classification(approved_classification, domain_code)
        if not middles:
            middles = ["전체"]  # 중분류 추출 실패 시 전체로 1회 호출

        push_log(sess, f"[TC 작성] [{i+1}/{len(domains)}] {domain_code} — {domain_name} ({len(middles)}개 중분류)")

        domain_tc_parts = []
        # TC ID 스킴 분기 (tc-rules.md 섹션 2-1):
        #   - Suite-based (PC Web, SC): 중분류 전체에 걸쳐 cumulative seq
        #   - Screen-based (Mobile, SM): 화면(중분류)별 독립 seq, SuiteCode 자리에 ScreenCode 사용
        project_code = _detect_project_code(project_name)
        domain_suite_code = domain.get("suite_code", "").strip().strip("-")
        # 사용자 명시적 선택(auto_screen_code)이 있으면 우선, 없으면 프로젝트 코드 자동 판별
        user_auto_flag = sess.get("auto_screen_code")
        if user_auto_flag is True:
            screen_based = True
            scheme_reason = "사용자 선택: 시스템 규칙 자동 적용"
        elif user_auto_flag is False:
            screen_based = False
            scheme_reason = "사용자 선택: 수동 SuiteCode 입력"
        else:
            screen_based = _is_screen_based(project_code)
            scheme_reason = f"자동 판별 (project_code={project_code})"
        screen_map = load_screen_code_map(project_name) if screen_based else {}
        if screen_based:
            push_log(sess, f"[TC 작성] {domain_code} — Screen-based 스킴 ({scheme_reason}, map {len(screen_map)}개)")
        else:
            push_log(sess, f"[TC 작성] {domain_code} — Suite-based 스킴 ({scheme_reason})")

        cumulative_seq = 1  # Suite-based 전용 (Screen-based는 매 중분류마다 1로 리셋)
        # Screen-based 스킴: 이미 사용한 ScreenCode 추적 — 같은 코드가 여러 중분류에 매핑되면
        # 두 번째부터 접미사를 붙여 TC ID 충돌(SC-CON-001 3회 생성 등)을 방지한다.
        used_screen_codes = set()

        for mi, middle in enumerate(middles):
            label = f"{domain_code}/{middle}" if middle != "전체" else domain_code

            # 스킴별 effective code & starting seq 결정
            screen_character = ""
            screen_navigation = ""
            if screen_based and middle != "전체":
                # resolve_screen_code가 used_screen_codes를 참조해 충돌 시 마지막 단어를 1자씩 확장
                # 예: GOWE 충돌 → GOWEB → GOWEBV → 그래도 충돌이면 GOWE2 숫자 폴백
                effective_code = resolve_screen_code(middle, screen_map, used_screen_codes)
                # 이론적으로 resolve_screen_code가 유일한 코드를 반환하지만, 매핑 파일이 중복된 극단 케이스를 방어
                if effective_code in used_screen_codes:
                    for suffix in range(2, 100):
                        cand = f"{effective_code}{suffix}"
                        if cand not in used_screen_codes:
                            push_log(sess, f"[TC 작성] ⚠️ ScreenCode 최종 충돌 — {effective_code} → {cand} ({middle})")
                            effective_code = cand
                            break
                used_screen_codes.add(effective_code)

                screen_character = resolve_screen_character(middle, screen_map)
                screen_navigation = resolve_screen_navigation(middle, screen_map)
                starting_seq = 1  # 화면 단위 독립 네임스페이스
                id_scope_label = f"{effective_code} (screen)"
            else:
                effective_code = domain_suite_code
                starting_seq = cumulative_seq
                id_scope_label = f"{effective_code} (suite)" if effective_code else "(no-code)"

            meta_msg = ""
            if screen_character:
                meta_msg += f" [성격: {screen_character}]"
            if screen_navigation:
                meta_msg += f" [네비: {screen_navigation}]"
            push_log(sess, f"[TC 작성] [{i+1}/{len(domains)}] {label} 작성 중... "
                            f"→ SM-{effective_code}-{starting_seq:03d}~ [{id_scope_label}]{meta_msg}")
            push(sess, "stage", {
                "stage": 4,
                "label": f"TC 작성: {label} ({mi+1}/{len(middles)})",
                "pct": 55 + int(25 * ((i * len(middles) + mi) / max(len(domains) * len(middles), 1)))
            })

            system = build_tc_system_prompt(tc_rules, approved_classification, project_policies, fewshot)

            # 섹션 발췌 모드: Quick 모드 또는 정책 반영 모드(최적화) 모두에서 원문 발췌 활성
            # (`_section_extract` = 정책 반영 모드 TC 최적화, `_quick_mode` = Quick 전용 플래그)
            effective_features = features_text
            effective_policy = policy_text
            use_section = (sess.get("_quick_mode") or sess.get("_section_extract"))
            if use_section and middle != "전체":
                raw_full = sess.get("_raw_text", "")
                section = extract_section_from_raw(raw_full, middle, major_name=domain_name)
                if section:
                    effective_features = section
                    if sess.get("_quick_mode"):
                        effective_policy = section  # Quick은 원문 근거만
                        push_log(sess, f"[Quick 3/3] {label} — 원문 섹션 발췌 ({len(section):,}자)")
                    else:
                        # 정책 반영 모드: 원문 섹션 + 정책 요약 앞부분(≤4KB) 병행
                        effective_policy = (policy_text[:4000] + "\n\n---\n\n원문 섹션:\n" + section) if policy_text else section
                        push_log(sess, f"[정책+섹션] {label} — 원문 섹션 발췌 ({len(section):,}자)")
                # 섹션을 못 찾으면 features_text로 폴백

            # 중분류 단위 프롬프트 — 해당 중분류에 집중, effective_code & starting_seq & 성격 & 네비 전달
            # 입력 소스 SCR 매핑 (v0.9.13~) — AI 가 화면 코드 환각 안 하도록 명시적 주입
            _scr_map = sess.get("_source_scr_map") or {}
            if middle != "전체":
                domain_with_middle = dict(domain)
                domain_with_middle["_focus_middle"] = middle
                domain_with_middle["suite_code"] = effective_code  # screen-based면 ScreenCode로 교체
                user = build_tc_user_prompt(domain_with_middle, effective_features, effective_policy,
                                            project_name, approved_classification,
                                            starting_seq=starting_seq,
                                            screen_character=screen_character,
                                            screen_navigation=screen_navigation,
                                            source_scr_map=_scr_map)
            else:
                domain_copy = dict(domain)
                domain_copy["suite_code"] = effective_code
                user = build_tc_user_prompt(domain_copy, features_text, policy_text,
                                            project_name, approved_classification,
                                            starting_seq=starting_seq,
                                            screen_character=screen_character,
                                            screen_navigation=screen_navigation,
                                            source_scr_map=_scr_map)

            try:
                tc_draft = call_claude(system, user, max_tokens=16000)
            except Exception as e:
                push_log(sess, f"[TC 작성] {label} 오류: {e}, 재시도...")
                time.sleep(2)
                try:
                    tc_draft = call_claude(system, user, max_tokens=16000)
                except Exception as e2:
                    push_log(sess, f"[TC 작성] {label} 실패 (건너뜀): {e2}")
                    continue

            # 잘림 감지
            last_tc_match = list(re.finditer(r"^###\s", tc_draft, re.MULTILINE))
            if last_tc_match:
                last_tc_block = tc_draft[last_tc_match[-1].start():]
                if "**테스트 단계**" not in last_tc_block or "**예상 결과**" not in last_tc_block:
                    push_log(sess, f"[TC 작성] {label} 마지막 TC 불완전 — 이어서 생성 중...")
                    try:
                        continuation = call_claude(system, f"이전 응답이 잘렸습니다. 아래 TC를 완성하세요.\\n\\n---\\n{last_tc_block}", max_tokens=4096)
                        tc_draft = tc_draft[:last_tc_match[-1].start()] + continuation
                    except Exception:
                        pass

            # Post-processing: TC ID 재번호링 (유니크 강제)
            if effective_code:
                tc_draft, next_seq = renumber_tc_ids(tc_draft, project_code, effective_code, starting_seq)
                renumbered_count = next_seq - starting_seq
                if screen_based:
                    # 화면별 독립 카운트 — cumulative_seq는 건드리지 않음
                    push_log(sess, f"[TC 작성] {label} 완료 — TC {renumbered_count}개 "
                                    f"(SM-{effective_code}-001~{next_seq-1:03d})")
                else:
                    cumulative_seq = next_seq
                    push_log(sess, f"[TC 작성] {label} 완료 — TC {renumbered_count}개 (seq ~{cumulative_seq-1:03d})")
                total_tc += renumbered_count
            else:
                tc_count = len(re.findall(r"^###\s", tc_draft, re.MULTILINE))
                tc_count -= len(re.findall(r"^###\s*카테고리\s*\d", tc_draft, re.MULTILINE))
                push_log(sess, f"[TC 작성] {label} 완료 — TC {tc_count}개")
                total_tc += tc_count

            domain_tc_parts.append(tc_draft)
            check_stop(sess)

        # 대분류별 병합 저장
        merged = "\n\n---\n\n".join(domain_tc_parts)
        draft_path = sess["workspace"] / f"tc_draft_{domain_code}.md"
        draft_path.write_text(merged, encoding="utf-8")
        all_tc_parts.append(merged)
        push_log(sess, f"[TC 작성] {domain_code} 전체 완료 — TC {total_tc}개 (누적)")
        check_stop(sess)

    if not all_tc_parts:
        raise RuntimeError("TC 작성 결과가 없습니다.")

    min_tc = max(1, round(total_tc * 0.35))
    return "\n\n---\n\n".join(all_tc_parts), total_tc, min_tc


def extract_domains(classification: str) -> list[dict]:
    domains = []
    for line in classification.splitlines():
        line = line.strip()
        if not line.startswith("## "):
            continue
        text = line[3:].strip()
        # "대분류:" 접두사 제거
        text = re.sub(r"^대분류[:\s]*", "", text).strip()
        if not text or text.startswith("#"):
            continue
        # 괄호 코드 있으면 추출, 없으면 이름을 코드로 사용
        cm = re.search(r"[\(\（]([A-Z]{2,8})[\)\）]", text)
        if cm:
            code = cm.group(1)
            name = re.sub(r"\s*[\(\（][A-Z]{2,8}[\)\）]", "", text).strip()
        else:
            name = text
            # 이름에서 코드 자동 생성 (영문 대문자 약어)
            code = re.sub(r"[^A-Za-z]", "", name).upper()[:6] or "FUNC"
        domains.append({"name": name, "code": code})

    # 최소 1개 보장
    if not domains:
        domains = [{"name": "전체 기능", "code": "FUNC"}]

    return domains


def build_tc_system_prompt(tc_rules: str, classification: str, project_policies: str = "", fewshot: str = "") -> str:
    policy_section = ""
    if project_policies:
        policy_section = f"""

## 프로젝트별 정책 (반드시 준수)

{project_policies[:6000]}
"""
    fewshot_section = ""
    if fewshot:
        fewshot_section = f"""

## 참고 예시 (⚠️ 이 형식과 수준을 따라 작성하세요)

아래는 실제 사용 중인 TC 예시입니다. 형식, 상세도, 어투를 동일하게 따르세요.

{fewshot[:5000]}
"""
    return f"""당신은 전문 소프트웨어 QA 엔지니어입니다. 주어진 대분류의 테스트 케이스를 작성합니다.

## TC 작성 규칙

{tc_rules if tc_rules else "표준 TC 형식을 따릅니다."}
{policy_section}{fewshot_section}
## 분류표
{classification[:10000]}

## TC 생성 카테고리 (5가지 — 순서대로 작성, 비율 준수)

각 중분류에 대해 아래 5가지 카테고리 순서로 TC를 작성하세요.
해당 카테고리에 만들 TC가 없으면 skip 합니다.
⚠️ 카테고리 3(예외)과 4(에러)를 반드시 포함하세요. Positive만으로 구성하지 마세요.

### 카테고리 1: UI/UX 체크 (분류: Positive, 우선순위: Medium~High) — 약 18%
- 화면 레이아웃, 요소 배치, 텍스트 표시가 올바른지 확인
- 초기 진입 시 기본 상태 (기본값, placeholder, 비활성 버튼 등)
- 반응형 / 해상도별 레이아웃 깨짐 여부
- 로딩 상태, 빈 데이터 표시, 툴팁/안내 문구

### 카테고리 2: 주요 기능 (분류: Positive, 우선순위: High) — 약 32%
- 핵심 비즈니스 흐름 (Happy Path)
- 사용자가 가장 자주 수행하는 동작
- 데이터 입력 → 처리 → 결과 확인의 정상 흐름
- CRUD 동작, 상태 전환, 네비게이션

### 카테고리 3: 예외 기능 — 비즈니스 위험 중심 (분류: Negative, 우선순위: High~Medium) — 약 25%
- 잘못된 입력, 경계값, 허용 범위 초과
- 권한 없는 접근, 인증 만료 상태에서의 동작
- 동시 접근, 중복 요청 (더블 클릭 등)
- 데이터 정합성 위험 (잔고 부족, 수량 초과 등)
- 미연동 상태에서의 기능 접근

### 카테고리 4: 에러 처리 및 비기능 (분류: Edge, 우선순위: Medium~Low) — 약 18%
- 네트워크 오류, 서버 타임아웃 시 에러 메시지 표시
- 빈 응답, 잘못된 응답 형식에 대한 방어
- 성능 관련 (로딩 시간, 대량 데이터 표시)
- 접근성 (키보드 조작, 포커스 이동)

### 카테고리 5: 연결 시나리오 — 사용자 사용 순서 검증 (분류: Positive, 우선순위: High~Medium) — 약 5~10%
**조건부 카테고리** — 아래 두 경우에만 1~3개 TC 작성, 그 외 화면은 **skip**:
  · **시작점 화면**: 화면 명세에 "진입: (없음/온보딩/탭바/외부 진입)" 류이거나, 그룹 흐름의 첫 화면 (예: Onboarding 게이트, 탭 root, 외부 딥링크 진입점)
  · **종착지 화면**: 흐름의 결과를 보여주는 화면 (예: 성공/실패 결과 화면, 완료 화면, 인증 완료 후 일방향 화면)

작성 규칙:
- 화면 명세의 "**진입**" 메타 + 화면 안의 "navigateTo(...)" / "→ SCR-XXX" 표기를 단서로 흐름을 따라간다
- 다음 단계 화면 1~2개를 함께 거치는 **end-to-end happy path** 하나
- (시작점이면) 핵심 분기 1개 — 예: "Skip for now" 분기, "다른 옵션 선택" 분기
- (종착지면) round-trip 1개 — 예: 결과 → 다음 화면 진입 → 다시 그 결과로 돌아오지 못함을 확인 (one-way 검증)
- TC 단계는 화면 전환 사이의 **state 전달** 과 **로고/이름 일관성** 을 검증 (예: "선택한 거래소 OKX 가 다음 화면에서도 동일하게 표시")
- 중간 화면(통과만 하는 화면)은 카테고리 5를 skip — 자기 화면의 navigate 동작은 카테고리 2에서 이미 다룸
- 응답 길이가 부족하면 **카테고리 5를 가장 먼저 줄이세요** (필수 의무는 카테고리 3·4)

판단 가이드 (어떤 화면이 시작점/종착지인가):
- ✅ 시작점: 그룹 hub, 온보딩 게이트, 탭 root, 외부 진입 deeplink, 첫 인증 화면
- ✅ 종착지: 성공/실패 결과 화면, 완료(one-way) 화면, 그룹 흐름의 마지막 화면
- ❌ skip: 입력 화면 중간 단계, 단순 표시 화면, 모달/바텀시트(자체로 연결 시나리오 아님)

## ⚠️ 누락 방지 — 자주 통합되어 사라지는 TC 패턴 (반드시 분리 작성)

다음 3가지 패턴은 AI 가 "유사하니 묶자" 라고 판단해 통합하기 쉬운데, **반드시 별도 TC** 로 분리해야 합니다:

### 1) 유사 계산 공식의 분리 (카테고리 2)
같은 화면에 여러 계산 공식이 노출되면 **공식별로 별도 TC**:
- ⛔ 통합 금지: "Position size 와 Notional value 를 한 TC 에서 검증"
- ✅ 분리 작성:
  - `Position size 계산 정확성` — Margin × Leverage 공식 (qty × symbol 표기 검증)
  - `Notional value 계산 정확성` — 별도 TC, USD 환산값 검증
  - `Liquidation price 계산` / `Mark price 갱신` / `Δ% 산출` — 각각 별도
- 이유: 각 공식의 정확성이 독립적이며, 한 공식 실패가 다른 공식에 영향 없이 검출되어야 함

### 2) 구체 성능 임계값 검증 (카테고리 4)
일반 Loading TC 와 별도로 **구체 환경 + 정량 임계값** 케이스 분리:
- ⛔ 통합 금지: "Loading 상태가 표시된다" 만 검증하는 일반 TC 1개
- ✅ 분리 작성:
  - `Loading 상태 표시` — 일반 케이스 (카테고리 1 또는 2)
  - `3G/저사양 환경에서 5초 이내 로딩 완료` — 성능 임계값 케이스 (카테고리 4)
  - `대량 데이터(100+ 항목) 표시 시 렌더 성능` — 부하 케이스 (카테고리 4)
- 이유: 성능 회귀는 일반 Loading 검증으로 안 잡힘. 구체 환경 + 정량 SLA 가 명세에 있으면 별도 TC

### 3) 플랫폼별 분기 동작 (카테고리 1 또는 4)
iOS/Android, Mobile/Desktop 별로 동작 차이가 있는 trigger 는 **플랫폼별 분리**:
- ⛔ 통합 금지: "뒤로가기 시 시트 닫힘" 한 TC 로 묶음
- ✅ 분리 작성:
  - `Cancel 버튼 / 배경 오버레이 탭 시 시트 닫힘` — 일반 닫기
  - `iOS Safari 스와이프 백 제스처 시 시트 닫힘` — iOS 전용
  - `Android Chrome 시스템 뒤로가기 시 시트 닫힘` — Android 전용
- 다른 예: 키보드 입력 (iOS 가상 키보드 vs Android), 스크롤 관성, 햅틱 피드백
- 이유: iOS swipe back 은 brower 레벨 처리, Android back 은 시스템 이벤트 — 구현 경로 다름

## Smoke TC 선별
- 카테고리 1, 2에서 High 우선순위 TC를 bold 처리 (### **...**)
- 카테고리 3에서 High Negative TC도 bold 처리
- 카테고리 5의 happy path TC도 bold 처리 (E2E 핵심 흐름)

## 형식 규칙

### TC ID 규칙 (⚠️ 반드시 준수)
- 형식: `{{ProjectCode}}-{{SuiteCode}}-{{NNN}}`
- ProjectCode: `SC` (PC Web) 또는 `SM` (Mobile Web) — 프로젝트명에서 판별
- SuiteCode: 시트명의 약어 (대문자 영문, 하이픈 허용). 예: GNB&Footer→GNBF, Trade-Order→TRD-ORDR
- NNN: 001부터 순차 번호 (3자리 zero-padding)
- ⛔ `TC-`로 시작하면 안 됨. 반드시 ProjectCode(`SC` 또는 `SM`)로 시작
- 예시: SC-GNBF-001, SC-TRD-ACCT-001, SM-LITE-001, SM-TRD-ORDR-001

### TC 헤딩
- Smoke TC: `### **{{ProjectCode}}-{{SuiteCode}}-{{NNN}}** — {{제목}}`  (bold)
- 일반 TC: `### {{ProjectCode}}-{{SuiteCode}}-{{NNN}} — {{제목}}`  (plain)
- [★] 기호 사용 금지

### 소분류 (간결 라벨 원칙 — 매우 중요)
- 목적: **"무엇을 테스트하는지" 한눈에 식별하는 라벨**. 검증 내용 자체는 사전 조건/테스트 단계/예상 결과에 명시되므로 소분류는 **간결한 식별 라벨**이면 충분.
- **목표 길이**: 30자 이내 (한 줄), 가능하면 15~25자
- ✅ 권장 형태:
  - `Connect 버튼 탭 동작`
  - `OAuth 웹뷰 진입 확인`
  - `Disconnect 후 상태 전환`
  - `Empty 상태 표시`
  - `URL 바 정보 확인`
  - `재시도 가능 확인`
- ⛔ 금지 형태 (장황·중복):
  - 카테고리 헤딩 prefix: `에러 처리 / 연결 해제 실패 / Toast 3초 자동 닫힘 타이밍 정확성 확인` ❌
  - State 라벨 시드 prefix: `Normal 상태 표시 및 동작 — OAuth Connect 화면 UI 요소 표시 확인` ❌
  - 구체 URL/도메인 인용: `URL 바는 "https://www.okx.com/oauth/authorize?response_type=..." 표시 확인` ❌ → `URL 바 정보 확인` ✅
  - 메타 마커 prefix: `[보안] / [정책] / [통합 — 그룹 ...]` ❌ (메타는 비고나 우선순위로)
  - 불필요한 어미: `~ 동작 확인 (...)`, `~ 상태에서 ~ 시 ~ 확인` 처럼 trigger + 부연 모두 노출 ❌
- 정확한 검증 내용 (URL 정확성, 타이밍 ms, 색상 hex, 좌표 등)은 **테스트 단계와 예상 결과에 명시** — 소분류에는 키워드만.

### 사전 조건
- 번호 개조식: `1. 조건A / 2. 조건B`
- 어미: `~인 상태 / ~된 상태 / ~한 상태`
- 불릿(-) 금지

### TC 출력 형식
```
### {{ProjectCode}}-{{SuiteCode}}-{{NNN}} — [제목]

| 항목 | 내용 |
|------|------|
| 대분류 | {{대분류명}} |
| 중분류 | {{중분류명}} |
| 소분류 | {{소분류명}} |
| 분류 | Positive/Negative/Edge |
| 우선순위 | High/Medium/Low |
| 플랫폼 | Web(Mobile) |
| 연관 화면 | /경로 또는 화면명 |

**사전 조건**
1. 조건A인 상태
2. 조건B된 상태

**테스트 단계**
1. 단계 1
2. 단계 2

**예상 결과**
- 결과 1

**비고**
- (없으면 생략)
```

비개발자도 이해할 수 있도록 구체적으로 작성하세요. 모호한 표현("정상적으로", "제대로") 금지.
"""


def _detect_project_code(project_name: str) -> str:
    """프로젝트명에서 ProjectCode 추출"""
    name_lower = project_name.lower()
    if "mobile" in name_lower or "모바일" in name_lower:
        return "SM"
    return "SC"


def _is_screen_based(project_code: str) -> bool:
    """Screen-based TC ID 스킴을 사용하는 프로젝트 판정.

    - SM (Mobile Web), SA (Mobile App) → Screen-based (화면별 001~)
    - SC (PC Web) 등 → Suite-based (Suite 내 전역 001~)
    """
    return project_code.upper() in ("SM", "SA")


# Screen Code Map 로더 (프로젝트별 projects/{name}/screen_code_map.md 파싱)
# 매핑 구조: { middle_name: { "code": "SPL", "character": "entry" } }
_SCREEN_CODE_CACHE: dict[str, dict[str, dict[str, str]]] = {}

# 성격(character) → 필수 비기능 TC 가이드 (원칙 E, tc-rules.md §1-1)
SCREEN_CHARACTER_NFR_GUIDE = {
    "entry": "로딩 시간 3초 이내 (Google RAIL 기준). [미결] 임계치 PM 확정 필요.",
    "data-fetch": "초기 로딩 5초 이내 + 빈 상태 표시 + Pull-to-refresh. [미결] 임계치 확정 필요.",
    "realtime": "WebSocket 끊김 시 자동 재연결 + 지연 표시 인라인 배너. 재연결 10초 이내.",
    "form": "입력 반응 100ms 이내 (RAIL First-Input-Delay, 선택적).",
    "overlay": "외부 영역 탭 시 자동 닫힘 + 진입/퇴장 애니메이션 부드러움.",
    "static": "비기능 TC 불필요 (정적 컨텐츠, 스크롤만).",
}

# 네비게이션(navigation) → 필수 뒤로가기/스와이프 TC 가이드 (원칙 F, tc-rules.md §1-1)
SCREEN_NAVIGATION_TC_GUIDE = {
    "tab-root": "탭바 최상위 — 뒤로가기 시 앱 종료 확인 모달 또는 Splash 복귀. 브라우저 뒤로가기 특수 처리 확인.",
    "one-way": "⚠️ **이전 화면 복귀 금지** — 뒤로가기/iOS 스와이프 백 제스처 시 이전 화면(인증·주문 플로우)으로 돌아가면 안 됨. 차단 또는 홈 이동 또는 확인 모달. 재실행 방지 검증.",
    "sequential": "뒤로가기 → 이전 단계 화면 복귀 + 입력값(이메일·코드·약관 체크) 유지 확인.",
    "overlay": "뒤로가기 / 외부 영역 탭 / Escape → 오버레이만 닫힘, 배경 화면 상태 그대로 유지.",
    "detail": "뒤로가기 → 리스트 화면 복귀 + 이전 스크롤 위치 유지.",
    "static": "뒤로가기 → 상위 화면 자연 복귀 (depth 2+ 에서만 확인 TC 추가).",
}


def load_screen_code_map(project_name: str) -> dict[str, dict[str, str]]:
    """projects/{project}/screen_code_map.md를 파싱.

    Returns:
        {middle_name: {"code": str, "character": str, "navigation": str}}
    파일 테이블 형식:
        `| Middle | Code | 성격 | 네비 | SCR-ID | 설명 |`
        성격/네비 컬럼은 선택. 없으면 빈 문자열.
    """
    cache_key = project_name.lower()
    if cache_key in _SCREEN_CODE_CACHE:
        return _SCREEN_CODE_CACHE[cache_key]

    if not PROJECTS_RULES_DIR.exists():
        _SCREEN_CODE_CACHE[cache_key] = {}
        return {}

    pname_lower = project_name.lower().replace(" ", "").replace("-", "").replace("_", "")
    mapping: dict[str, dict[str, str]] = {}
    valid_chars = {"entry", "data-fetch", "realtime", "form", "overlay", "static"}
    valid_navs  = {"tab-root", "one-way", "sequential", "overlay", "detail", "static"}
    for folder in PROJECTS_RULES_DIR.iterdir():
        if not folder.is_dir():
            continue
        fname_lower = folder.name.lower().replace(" ", "").replace("-", "").replace("_", "")
        if fname_lower not in pname_lower and pname_lower not in fname_lower:
            continue
        map_file = folder / "screen_code_map.md"
        if not map_file.exists():
            continue
        text = map_file.read_text(encoding="utf-8")
        for line in text.splitlines():
            line = line.strip()
            if not line.startswith("|") or not line.endswith("|"):
                continue
            cells = [c.strip() for c in line.strip("|").split("|")]
            if len(cells) < 2:
                continue
            middle, code = cells[0], cells[1]
            if not middle or not code:
                continue
            if re.match(r"^-+$", middle) or middle.lower() in ("middle name", "이름"):
                continue
            if not re.fullmatch(r"[A-Z]{2,5}", code):
                continue
            # 3번째 셀: 성격 (선택)
            character = ""
            if len(cells) >= 3:
                cand = cells[2].lower().strip()
                if cand in valid_chars:
                    character = cand
            # 4번째 셀: 네비게이션 (선택)
            navigation = ""
            if len(cells) >= 4:
                cand = cells[3].lower().strip()
                if cand in valid_navs:
                    navigation = cand
            # navigation이 비어있으면 character 기반 기본 추론
            if not navigation:
                if character == "overlay":
                    navigation = "overlay"
                elif character == "static":
                    navigation = "static"
                else:
                    navigation = "sequential"  # 안전한 기본값
            if middle not in mapping:
                mapping[middle] = {
                    "code": code,
                    "character": character,
                    "navigation": navigation,
                }
        break

    _SCREEN_CODE_CACHE[cache_key] = mapping
    return mapping


# ── 옵션 조합 명세 (order_combinations.md) 파서 ──────────────────────────
def parse_order_combinations(project_name: str) -> dict:
    """projects/{project}/order_combinations.md 를 파싱.

    파일 구조 (가정):
      ## 3. 옵션 조합 매트릭스 (Decision Table)
        ### 3.1 정상 케이스 ...
          | OC-ID | ... | 기대 결과 |
          | OC-001 | ... |
      ## 4. 사용자 시나리오 (Use Case Sequence)
        ### S1: ...
          **페르소나:** ...
          **특성:** ...
          | Step | 동작 | 사용 OC | 검증 |
          | S1-1 | ... |

    Returns:
        {
          "raw_md": str (전문),
          "combos": [{id, category, raw_row, headers, cells}],  # 옵션 조합
          "scenarios": [{id, title, persona, traits, steps:[{step_id, action, oc_ref, verify}]}],
        }
    파일 없으면 None 반환.
    """
    if not PROJECTS_RULES_DIR.exists():
        return None
    pname_lower = project_name.lower().replace(" ", "").replace("-", "").replace("_", "")
    md_path = None
    for folder in PROJECTS_RULES_DIR.iterdir():
        if not folder.is_dir():
            continue
        fname_lower = folder.name.lower().replace(" ", "").replace("-", "").replace("_", "")
        if fname_lower not in pname_lower and pname_lower not in fname_lower:
            continue
        candidate = folder / "order_combinations.md"
        if candidate.exists():
            md_path = candidate
            break
    if not md_path:
        return None

    text = md_path.read_text(encoding="utf-8")
    result = {"raw_md": text, "combos": [], "scenarios": []}

    # 현재 카테고리 (### 3.1 정상 케이스 등) 추적
    current_category = None
    current_table_headers: list[str] = []
    in_section3 = False
    in_section4 = False
    current_scenario = None
    current_step_table_headers: list[str] = []

    lines = text.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i].rstrip()
        stripped = line.strip()

        # 섹션 진입 판단
        if stripped.startswith("## 3.") or stripped.startswith("## 3 "):
            in_section3 = True
            in_section4 = False
        elif stripped.startswith("## 4.") or stripped.startswith("## 4 "):
            in_section3 = False
            in_section4 = True
        elif stripped.startswith("## 5.") or stripped.startswith("## 5 "):
            in_section3 = False
            in_section4 = False

        if in_section3 and stripped.startswith("### 3."):
            current_category = re.sub(r"^###\s+3\.\d+\s+", "", stripped)
            current_table_headers = []
        if in_section4 and re.match(r"^###\s+(?:⭐\s*)?S\d+", stripped):
            # 신규 시나리오 시작
            m = re.match(r"^###\s+(?:⭐\s*)?(S\d+):\s*(.+)$", stripped)
            if m:
                current_scenario = {
                    "id": m.group(1),
                    "title": m.group(2).strip(),
                    "persona": "",
                    "traits": "",
                    "steps": [],
                }
                result["scenarios"].append(current_scenario)
                current_step_table_headers = []
            else:
                current_scenario = None

        # 페르소나/특성 라인 추출 (시나리오 안에서)
        if current_scenario and stripped.startswith("**페르소나:**"):
            current_scenario["persona"] = stripped.replace("**페르소나:**", "").strip()
        elif current_scenario and stripped.startswith("**특성:**"):
            current_scenario["traits"] = stripped.replace("**특성:**", "").strip()

        # 표 행 파싱
        if stripped.startswith("|") and stripped.endswith("|"):
            cells = [c.strip() for c in stripped.strip("|").split("|")]
            # 헤더 / 구분선 거르기
            if not cells or set(cells[0]) <= set("-"):
                i += 1
                continue
            first = cells[0]
            # 섹션 3: OC 행 (헤더 또는 데이터)
            if in_section3:
                if first.upper().startswith("OC-ID") or first.upper() == "OC-ID":
                    current_table_headers = cells
                elif re.match(r"^OC-\d+$", first, re.IGNORECASE) and current_table_headers:
                    combo = {
                        "id": first.upper(),
                        "category": current_category or "(unknown)",
                        "headers": current_table_headers,
                        "cells": cells,
                    }
                    result["combos"].append(combo)
            # 섹션 4: 시나리오 step
            if in_section4 and current_scenario is not None:
                if first.lower() == "step":
                    current_step_table_headers = cells
                elif re.match(r"^S\d+-\d+[a-z]?$", first, re.IGNORECASE) and current_step_table_headers:
                    # cells: [step_id, 동작, 사용 OC, 검증]
                    step = {
                        "step_id": first,
                        "action": cells[1] if len(cells) > 1 else "",
                        "oc_ref": cells[2] if len(cells) > 2 else "",
                        "verify": cells[3] if len(cells) > 3 else "",
                    }
                    current_scenario["steps"].append(step)
        i += 1

    return result
# ── /옵션 조합 명세 파서 ─────────────────────────────────────────────


def step_draft_combo_spec(spec_folder: str, domain: str,
                          target_scrs: list[str] | None = None,
                          existing_md: str = "") -> tuple[str, dict]:
    """spec 폴더 분석 → Combo 명세 (order_combinations.md 형식) 초안 작성.

    Args:
        spec_folder: 구조화 spec 폴더 경로 (overview/policy/design/scr 분리)
        domain: 'Trade' / 'Exchange' / 'Portfolio' / 'Markets' / 사용자 입력
        target_scrs: 분석 대상 SCR ID 리스트 (None=전체)
        existing_md: 기존 명세 있으면 → 갱신 모드 (참고용)
    Returns:
        markdown 형식 명세 초안 (사용자 검토용)
    """
    folder = Path(spec_folder).expanduser()
    if not folder.exists() or not folder.is_dir():
        raise RuntimeError(f"spec 폴더 없음: {spec_folder}")

    # 분석 대상 spec 수집
    cls = classify_spec_files(folder)
    overview_text = ""
    policy_text = ""
    if cls.get("overview") and cls["overview"].exists():
        overview_text = cls["overview"].read_text(encoding="utf-8")[:30000]
    if cls.get("policy"):
        policy_text = "\n\n".join(p.read_text(encoding="utf-8") for p in cls["policy"][:3])[:30000]

    # 화면 md 수집 — target_scrs 필터 (없으면 도메인 키워드 기반 추출)
    screens_md_parts = []
    domain_keywords = {
        "Trade": ["order", "trade", "lite", "position", "주문", "거래", "포지션"],
        "Exchange": ["exchange", "oauth", "api", "connect", "거래소", "연동", "연결"],
        "Portfolio": ["portfolio", "pnl", "balance", "포트폴리오"],
        "Markets": ["market", "마켓", "시세"],
    }
    keywords = domain_keywords.get(domain, [domain.lower()])
    for scr_md in sorted(folder.glob("scr/SCR-*.md")):
        scr_id = scr_md.stem  # SCR-104
        if target_scrs:
            if scr_id not in target_scrs:
                continue
        else:
            # 키워드 매칭 (파일명 또는 첫 1000자 본문)
            content = scr_md.read_text(encoding="utf-8")[:2000].lower()
            if not any(kw in content for kw in keywords):
                continue
        screens_md_parts.append(f"### {scr_id}\n{scr_md.read_text(encoding='utf-8')[:5000]}")
    screens_md = "\n\n---\n\n".join(screens_md_parts[:10])  # 최대 10개 (토큰 절약)

    # 갱신 모드 안내
    update_note = ""
    if existing_md:
        update_note = f"\n\n## 기존 명세 (갱신 대상)\n\n{existing_md[:15000]}\n\n→ 이 명세를 spec 변경 사항을 반영해 갱신하세요. 기존 OC/시나리오 ID 는 가능한 유지."

    # 참조 명세 자동 탐색 — 형식·구조 일관성 확보용
    # 우선순위: 1) spec 폴더 안 같은 도메인 / 2) spec 폴더 안 다른 도메인 / 3) projects/*/order_combinations.md
    reference_md = ""
    reference_source = ""
    if not existing_md:  # 갱신 모드 아닐 때만 (갱신 모드는 existing_md 가 더 정확한 참조)
        # 1) 현재 spec 폴더 안 *_combinations.md 들 찾기
        spec_combos = list(folder.glob("*_combinations.md"))
        # 같은 도메인 우선 (파일명 stem 이 domain 키워드 포함하면 같은 도메인으로 간주)
        domain_lower = domain.lower()
        same_domain = [p for p in spec_combos if domain_lower in p.stem.lower()]
        other_domain = [p for p in spec_combos if domain_lower not in p.stem.lower()]
        ref_candidates = same_domain + other_domain
        # 2) projects/{name}/order_combinations.md 도 후보
        if PROJECTS_RULES_DIR.exists():
            for proj_folder in PROJECTS_RULES_DIR.iterdir():
                if proj_folder.is_dir():
                    for p in proj_folder.glob("*_combinations.md"):
                        if p not in ref_candidates:
                            ref_candidates.append(p)
        # 첫 후보 사용 (없으면 빈 값)
        if ref_candidates:
            ref_path = ref_candidates[0]
            try:
                reference_md = ref_path.read_text(encoding="utf-8")[:20000]
                reference_source = ref_path.name
            except Exception:
                reference_md = ""
    reference_note = ""
    if reference_md:
        reference_note = (
            f"\n\n## 참조용 형식 예시 ({reference_source})\n\n"
            f"다른 도메인의 기존 명세입니다. **이 파일의 구조·형식·일관성을 그대로 따라 작성**하세요.\n"
            f"(내용은 다르지만 옵션 차원 표 / Decision Table 카테고리 / 시나리오 페르소나 / 위험 등급 형식 동일하게):\n\n"
            f"```markdown\n{reference_md}\n```\n"
        )

    system_prompt = f"""당신은 블랙박스 테스트 도메인 전문가입니다.
사용자가 제공한 기획서(spec)를 분석하여 **옵션 조합 + 사용자 시나리오 기반 그레이박스 테스트 명세** (`order_combinations.md` 형식) 의 초안을 작성합니다.

## 출력 형식 (반드시 준수)

```markdown
# {{프로젝트}} — 주문 옵션 조합 명세 (그레이박스 TC)

> ... (1-2 문장 도메인 설명) ...

## 1. {domain} 옵션 차원 (Reference)

| 차원 | 가능한 값 | 비고 |
|------|----------|------|
| (도메인에서 추출한 옵션 차원들 — 6~12개) | | |

## 2. 도메인 룰 (Reference)

**계산:** (있다면)
- (수식 또는 공식)

**검증 룰:**
- (입력 검증, 상태 전환 규칙)

**테스트 실행 위험 등급:**
| 등급 | 조건 | 권장 안내 메시지 템플릿 |
|------|------|---------------------|
| 🔴 고위험 | ... | "..." |
| 🟡 중위험 | ... | "..." |
| 🟢 저위험 | ... | "..." |
| ⚪ 위험 없음 | ... | (생략) |

## 3. 옵션 조합 매트릭스 (Decision Table)

### 3.1 정상 케이스 (Positive — 표준 흐름)

| OC-ID | (차원1) | (차원2) | ... | 검증 의도 | 기대 결과 |
|-------|---------|---------|-----|---------|---------|
| OC-001 | ... | ... |  | ... | ... |
(5~8개)

### 3.2 경계값 / 위험 조합 (Boundary)
| OC-ID | ... |
(4~6개)

### 3.3 입력 검증 실패 (Negative)
| OC-ID | ... |
(4~6개)

### 3.4 시스템 에러 케이스
| OC-ID | 상태 | 검증 의도 | 기대 결과 |
(3~5개)

## 4. 사용자 시나리오 (Use Case Sequence)

### S1: {{페르소나 이름}}

**페르소나:** ...
**특성:** ...
**총 {{N}} TC**

| Step | 동작 | 사용 OC | 검증 |
|------|------|---------|------|
| S1-1 | ... | OC-XXX | ... |
(3~6개)

### S2: ...
(2~4개 시나리오)

## 5. 운영 가이드
... (간략히)
```

## 작성 원칙

1. **spec 본문에서 추출** — 명세에 없는 임의의 룰 만들지 마세요.
2. **도메인 차원** — 사용자/시스템 입력 가능한 옵션을 모두 차원으로 정리.
3. **OC 매트릭스** — 도메인 폭발 방지를 위해 의미있는 조합만. 표준 6 + 경계 5 + Negative 5 + 에러 4 정도.
4. **시나리오** — 페르소나 기반 (실제 사용자 그룹 가정). 3~5개.
5. **위험 등급** — 실데이터 테스트 시 자금/계정 영향 등급. 도메인에 따라 적절히.
6. **블랙박스 표현** — 기대 결과에 DOM ID, state, 함수명 노출 금지.
7. **갱신 모드** — 기존 명세 주어지면 ID 유지하며 변경 사항만 반영.

⚠️ 응답은 markdown 명세 본문만. 설명/요약 텍스트 추가 금지.
"""

    user_prompt = f"""다음 기획서를 분석해 **{domain}** 도메인의 옵션 조합 명세 초안을 작성하세요.

## Overview (01_spec.md)
{overview_text}

## Policy (정책 문서)
{policy_text}

## 화면 명세 ({len(screens_md_parts)}개 SCR)

{screens_md}
{update_note}{reference_note}

위 정보를 바탕으로 **{domain}** 도메인의 옵션 조합 명세 markdown 을 작성하세요.
파일명은 `{domain.lower()}_combinations.md` 가정.
"""
    draft_md = call_claude(system_prompt, user_prompt, max_tokens=16000)
    meta = {
        "overview_size": len(overview_text),
        "policy_size": len(policy_text),
        "screens_count": len(screens_md_parts),
        "reference_source": reference_source,
        "reference_size": len(reference_md),
    }
    return draft_md, meta


def step_write_combo_tc(sess: dict | None, project_name: str,
                        combo_data: dict, project_code: str = "SM",
                        suite_combo: str = "COMBO",
                        suite_flow: str = "FLOW") -> tuple[str, int]:
    """옵션 조합 + 시나리오 → TC 변환 (그레이박스, 1회 LLM 호출).

    Args:
        sess: Flask 세션 (있으면 push_log; CLI 테스트는 None 가능)
        project_name: "supercycl" 등
        combo_data: parse_order_combinations() 반환값
        project_code: TC ID prefix (기본 SM)
        suite_combo: 옵션 조합 SuiteCode (예: 'ORDR-COMBO', 'LITE-COMBO')
        suite_flow: 시나리오 SuiteCode (예: 'ORDR-FLOW', 'LITE-FLOW')
    Returns:
        (tc_markdown, total_tc_count)
    """
    def _log(msg):
        if sess is not None:
            push_log(sess, msg)
        else:
            print(msg)

    tc_rules = load_tc_rules()
    project_policies = load_project_policies(project_name)

    total_combos = len(combo_data["combos"])
    total_steps = sum(len(s["steps"]) for s in combo_data["scenarios"])
    expected_total = total_combos + total_steps
    _log(f"[Combo TC] 옵션 조합 {total_combos}개 + 시나리오 step {total_steps}개 = 총 {expected_total}개 TC 변환 시작")

    # 시스템 프롬프트 — 그레이박스 정책 명시
    system_prompt = f"""당신은 선물 거래 도메인의 시니어 QA 엔지니어입니다.
주문 옵션 조합 명세 (order_combinations.md) 를 읽고 **그레이박스 테스트 케이스** 를 작성합니다.

## 핵심 원칙 (그레이박스 정책)

1. **기대 결과는 사용자 가시 동작/결과만** (블랙박스 표현)
   - ✅ "검증 실패 메시지가 표시된다. TP/SL 값이 저장되지 않는다. 시트는 닫히지 않고 유지된다."
   - ⛔ "#tpSlErr 표시, saveTpSl() 실패" (DOM/함수명 노출 금지)

2. **사전 조건의 마지막 줄에 (선택) 기술 힌트** — DOM ID/식별자 1~2개
   예: `3. (기술 참고) TP 입력 #tpSlTpInput, 에러 영역 #tpSlErr`
   힌트가 없는 OC 는 사전 조건에서 생략.

3. **테스트 단계** — 사용자 동작 위주 (탭, 입력, 슬라이드)

4. **소분류는 식별 라벨 (간결)** — 30자 내. 검증 내용은 단계/기대결과에.

## TC ID 규칙

- 옵션 조합 (Decision Table): `{project_code}-COMBO-NNN` (001부터 OC 순서대로)
- 시나리오 step: `{project_code}-FLOW-NNN` (001부터 S1-1, S1-2, ..., S6-N 순서대로)
- 대분류: "Trade(Combo)" — 화면 중심 'Trade' 시트와 분리하기 위한 전용 대분류명 (시트도 이 이름으로 생성됨)
- 중분류:
  - Combo TC: "Order Combo"
  - Flow TC: "User Flow"
- 소분류:
  - Combo: "OC-XXX 한 줄 검증 의도 요약" (예: "OC-001 Limit 5x Conservative")
  - Flow: "SX-N 단계 요약" (예: "S5-1 Limit 입력")

## 분류 규칙

- 정상 케이스 / 시나리오 정상 step → Positive, 우선순위 High~Medium
- 경계값 / 위험 조합 → Edge, Medium
- 입력 검증 실패 / 실수 시나리오 → Negative, High
- 시세/시스템 에러 → Edge, Medium
- 행위 가드 (더블 탭) → Negative, High

## TC 출력 형식 (반드시 준수)

```
### {project_code}-{suite_combo}-001 — [제목]

| 항목 | 내용 |
|------|------|
| 대분류 | Trade(Combo) |
| 중분류 | Order Combo |
| 소분류 | OC-001 Limit 5x Conservative |
| 분류 | Positive |
| 우선순위 | High |
| 플랫폼 | Web(Mobile) |
| 연관 화면 | SCR-102, SCR-104, SCR-601 |

**사전 조건**
1. 거래소가 1개 이상 연동된 상태
2. 충분한 잔고가 있는 상태
3. (기술 참고) Limit price 입력 #liteLimitPriceInput102, Δ% 표시 #liteLimitPriceDelta102
4. ⚠️ 🟢 저위험 — 5x, Limit 미체결 시 Cancel 환원

**테스트 단계**
1. SCR-102 에서 Long + Limit + Cross + 5x + Conservative 옵션을 설정한다.
2. Limit price 를 Mark price 보다 −1% 낮게 입력한다.
3. Long 버튼 → SCR-104 진입 → Confirm 탭한다.

**예상 결과**
- Open Order 가 등록된다.
- Open Orders 카드에 Δ% 가 −1% 로 표시된다.
- Mark price 가 Limit price 에 도달하면 자동 체결되어 Position 카드로 전환된다.
- 체결 시 TP +10% / SL −5% 가 자동 등록된다.
```

## Smoke TC 선별 (Combo 전용 — SCR Smoke 와 차별)

**원칙: SCR (화면 중심) Smoke 와 겹치지 않게 선별.**

SCR Smoke 가 이미 다루는 영역:
- 화면 진입 / UI 표시 / 버튼 클릭
- 단일 화면 안의 기본 동작 (예: "Long 버튼 탭 → SCR-104 진입")
- 단순 폼 입력

**Combo Smoke 는 SCR 가 못 다루는 영역만 bold (`### **{{ID}}**` 형식):**
- ✅ **여러 화면 걸친 상태 전환** (Limit 등록 → Mark 도달 → 체결 → Position 카드)
- ✅ **옵션 조합 정합성** (100x + Liquidation 임박 검증, Hedge 양방향 분리, 코인별 정밀도)
- ✅ **타겟 페르소나 핵심 흐름** (S5 Limit lifecycle 의 첫·마지막 step)
- ✅ **회복 경로** (실패 → Retry → 성공)
- ✅ **위험 가드** (TP 방향 위반, 더블 탭, 잔고 80% 경고)

**Combo Smoke 권장 갯수: 81 중 8~12개 (10~15%)**

**선별 우선순위 (점수 높은 것부터):**
1. ⭐ S5 Limit lifecycle (S5-1 입력 + S5-4 체결) — 2개
2. 표준 + 보수 조합 (OC-001 Limit 5x Conservative) — 1개
3. 100x 위험 (OC-011) — 1개
4. Hedge 분리 (OC-030) — 1개
5. TP 방향 위반 (OC-020) — 1개
6. 더블 탭 가드 (OC-061 빠른) — 1개
7. 코인 차원 대표 (OC-072 SOL 정밀도) — 1개
8. S7 다중 코인 카드 분리 (S7-5) — 1개

**Smoke 마킹: TC 헤딩에 `**` 추가**
- 일반: `### {project_code}-{suite_combo}-001 — Limit 5x Conservative 표준 주문`
- Smoke: `### **{project_code}-{suite_combo}-001** — Limit 5x Conservative 표준 주문`

## ⚠️ 테스트 실행 위험 안내 (사전 조건에 짧게 통합)

**위험 안내는 사전 조건의 마지막 한 줄로 작성**하세요. 별도 섹션 만들지 마세요.
자금 영향 없는 케이스 (입력 검증 실패, UI 표시 검증 등) 는 위험 줄도 생략.

**⚠️ 길이 제한: 60자 이내** (Excel 사전조건 컬럼 절단 방지)

**작성 형식:**
```
N. ⚠️ [등급] — [핵심 위험 — 짧게]
```

**4단계 등급:**
- 🔴 고위험: Leverage ≥ 50x 또는 잔고 사용 ≥ 80% 또는 다수 회전
- 🟡 중위험: Leverage 10~25x 또는 잔고 사용 50~79% 또는 Hedge 양방향
- 🟢 저위험: Leverage ≤ 5x 또는 작은 Margin
- ⚪ 위험 없음 (생략): 입력 검증 실패 / UI 표시 / Loading / 시세 끊김

**예시 (모두 60자 이내):**
- 🟢 `5. ⚠️ 🟢 저위험 — 5x, max 손실 ≈ Margin × 5%`
- 🟢 `5. ⚠️ 🟢 저위험 — 3x, 소액. Limit 미체결 시 Cancel 환원`
- 🟡 `5. ⚠️ 🟡 중위험 — 25x Isolated, max 손실 ≈ Margin 전액`
- 🟡 `5. ⚠️ 🟡 중위험 — Hedge 양방향, 자금 2배 노출`
- 🔴 `5. ⚠️ 🔴 고위험 — 100x, 1% 변동에 Margin 전액 손실. 테스트 계정 권장`
- 🔴 `5. ⚠️ 🔴 고위험 — 100x 스캘퍼, 누적 수수료 + 청산 위험`

**작성 원칙 (짧게 쓰기):**
- "Leverage" 같은 자명한 단어 생략 → "100x" 만
- 추정 손실은 "max 손실 ≈ Margin 전액" 정도로 (정확한 USDT 값 불필요)
- 권장사항은 핵심 한마디만 (예: "테스트 계정 권장")
- 형용사·접속사 최소화

**판단 가이드:**
- Limit 미체결: 위험 한 단계 낮춤
- Hedge 양방향: 한 단계 올림
- 입력 검증 실패: 줄 자체 생략

## 도메인 룰 (계산/검증 참조)

{tc_rules}

## 프로젝트 정책

{project_policies}
"""

    # 샘플 모드 표시 (CLI 테스트용)
    sample_note = ""
    if combo_data.get("_sample_mode"):
        # 처리할 OC ID 와 시나리오 ID 명시
        oc_ids = [c["id"] for c in combo_data["combos"]]
        scn_ids = [f"{s['id']}({len(s['steps'])}step)" for s in combo_data["scenarios"]]
        sample_note = (
            f"\n## ⚠️ 샘플 모드 (테스트용)\n\n"
            f"전체 명세 중 **다음 항목만** 변환하세요. 나머지는 무시:\n"
            f"- OC: {', '.join(oc_ids)}\n"
            f"- 시나리오: {', '.join(scn_ids)}\n"
        )

    # User 프롬프트 — order_combinations.md 전문 + 생성 지시
    user_prompt = f"""다음은 주문 옵션 조합 명세입니다. 이 명세를 바탕으로 그레이박스 TC 를 작성하세요.

명세 전문:
---
{combo_data['raw_md']}
---
{sample_note}
## 작성 지시

1. **옵션 조합 {total_combos}개**를 `{project_code}-{suite_combo}-001` ~ `{project_code}-{suite_combo}-{total_combos:03d}` 로 변환.
   - 표의 OC-NNN 순서대로 1:1 매핑.
   - 소분류에 원본 OC-ID 포함 (예: "OC-001 ...").

2. **시나리오 step {total_steps}개**를 `{project_code}-{suite_flow}-001` ~ `{project_code}-{suite_flow}-{total_steps:03d}` 로 변환.
   - S1-1, S1-2, ..., S2-1, S2-2, ..., S6-N 순서대로 1:1 매핑.
   - 소분류에 원본 step ID 포함 (예: "S1-2 SCR-102 입력").
   - 시나리오 페르소나/특성을 사전 조건에 반영 (예: "신규 사용자, 잔고 50 USDT 인 상태").

3. **총 {expected_total}개 TC** 를 만든다. 빠짐 없이.

4. 응답은 TC markdown 만. 설명/요약 텍스트는 제외.

5. 모든 TC 사이에는 `---` 구분선 1줄.

지금 작성을 시작하세요."""

    # 분할 호출 — COMBO 와 FLOW 를 따로 (max_tokens 20K 한도 회피)
    # 총 81 TC × 평균 ~350 토큰 ≈ 28K → 1회 호출은 안전하지 않음
    # 분할: (1) Combo 만, (2) Flow 만 — 각각 ~15K 안에 수렴
    _log("[Combo TC] LLM 분할 호출 시작 — 1/2: Order Combo")

    # 1단계: Order Combo 만
    user_prompt_combo = user_prompt + (
        f"\n\n## ⚠️ 이번 호출은 1/2: **Order Combo 전용**\n"
        f"- 옵션 조합 {total_combos}개 ({project_code}-{suite_combo}-001 ~ {total_combos:03d}) 만 작성.\n"
        f"- 시나리오 step ({project_code}-{suite_flow}-*) 은 다음 호출에서 처리. **이번 호출에서 작성 금지**.\n"
        f"- Smoke 선별: 이 호출에서 **최대 5개** (OC-001 표준, OC-011 100x, OC-020 TP 위반, "
        f"OC-030 Hedge, OC-061 더블 탭 또는 OC-072 코인 대표 — 가이드 참조).\n"
    )
    try:
        combo_md = call_claude(system_prompt, user_prompt_combo, max_tokens=16000)
    except Exception as e:
        _log(f"[Combo TC] LLM 호출 1/2 실패: {e}")
        raise
    combo_count = len(re.findall(rf"^###\s+\*?\*?{re.escape(project_code)}-{re.escape(suite_combo)}-\d+",
                                 combo_md, re.MULTILINE))
    _log(f"[Combo TC] 1/2 완료 — {suite_combo} {combo_count}개 (기대 {total_combos})")

    # 2단계: Flow 만 (Flow 가 0이면 skip)
    flow_md = ""
    flow_count = 0
    if total_steps > 0:
        _log("[Combo TC] LLM 분할 호출 2/2: User Flow")
        user_prompt_flow = user_prompt + (
            f"\n\n## ⚠️ 이번 호출은 2/2: **User Flow 전용**\n"
            f"- 시나리오 step {total_steps}개 ({project_code}-{suite_flow}-001 ~ {total_steps:03d}) 만 작성.\n"
            f"- Order Combo ({project_code}-{suite_combo}-*) 는 이전 호출에서 완료. **이번 호출에서 작성 금지**.\n"
            f"- Smoke 선별: 이 호출에서 **최대 5개** (S5-1 / S5-4 Limit lifecycle 핵심, "
            f"S4-1 near-liquidation, S7-5 다중 코인 카드, 그리고 페르소나 대표 1개 — 가이드 참조).\n"
        )
        try:
            flow_md = call_claude(system_prompt, user_prompt_flow, max_tokens=16000)
        except Exception as e:
            _log(f"[Combo TC] LLM 호출 2/2 실패: {e}")
            raise
        flow_count = len(re.findall(rf"^###\s+\*?\*?{re.escape(project_code)}-{re.escape(suite_flow)}-\d+",
                                    flow_md, re.MULTILINE))
        _log(f"[Combo TC] 2/2 완료 — {suite_flow} {flow_count}개 (기대 {total_steps})")

    # 합치기 — 두 블록 사이에 구분선
    if flow_md:
        result = combo_md.rstrip() + "\n\n---\n\n" + flow_md.lstrip()
    else:
        result = combo_md

    actual_count = combo_count + flow_count
    _log(f"[Combo TC] 총 {actual_count}개 TC 생성 완료 (예상 {expected_total}개)")

    return result, actual_count


def resolve_screen_code(middle_name: str, screen_map: dict,
                         used_codes: set | None = None) -> str:
    """중분류 이름을 ScreenCode로 변환 (v0.9.7b 규칙).

    ┌─ 규칙 ────────────────────────────────────────────────────
    │ 1. screen_map에 명시적 매핑이 있으면 그대로 사용 (최우선)
    │ 2. 없으면 자동 파생:
    │    - 1단어     → 앞 3자         (예: Splash → SPL)
    │    - 2단어     → 각 단어 앞 2자  (예: Email Input → EMIN)
    │    - 3단어 이상 → 앞 단어들 2자 + 마지막 단어 1자
    │                  (예: Google OAuth Start → GOOAS)
    │ 3. 최대 8자로 절단, 대문자화
    │ 4. used_codes와 충돌하면 마지막 단어를 1자씩 확장
    │    (예: GOOAC 충돌 → GOOACO → GOOACOM ...)
    │ 5. 그래도 충돌하면 숫자 접미사 폴백 (GOOAC2)
    │ 6. 영문자가 전혀 없으면 해시 기반 MID### 폴백
    └────────────────────────────────────────────────────────────
    """
    if not middle_name:
        return "SCR"

    # 1단계: screen_map 정확 일치
    if middle_name in screen_map:
        entry = screen_map[middle_name]
        if isinstance(entry, dict):
            return entry.get("code") or "SCR"
        return entry

    # 2단계: 단어 분해 — 공백/하이픈/언더스코어 단위 + 영문자만 추출
    words = [re.sub(r"[^A-Za-z]", "", w) for w in re.split(r"[\s\-_]+", middle_name) if w]
    words = [w for w in words if w]
    if not words:
        return f"MID{abs(hash(middle_name)) % 1000:03d}"

    # 2단계 계속: 단어 수별 기본 파생
    if len(words) == 1:
        base = words[0][:3].upper()
    elif len(words) == 2:
        base = (words[0][:2] + words[1][:2]).upper()
    else:  # 3단어 이상
        head = "".join(w[:2] for w in words[:-1])
        tail = words[-1][:1]
        base = (head + tail).upper()

    # 3단계: 최대 8자 절단
    base = base[:8]
    if not base:
        return f"MID{abs(hash(middle_name)) % 1000:03d}"

    # 4단계: 충돌 없으면 반환
    if used_codes is None or base not in used_codes:
        return base

    # 4단계 계속: 마지막 단어를 1자씩 확장 시도
    # base는 "앞 단어들의 앞 N자" + "마지막 단어의 앞 1~2자" 구조
    if len(words) >= 2:
        last = words[-1]
        # 마지막 단어 사용 자수: 2단어면 2자, 3단어 이상이면 1자
        last_used_n = 2 if len(words) == 2 else 1
        head_part = base[:-last_used_n]
        # extra_len을 키우며 시도
        for extra_len in range(last_used_n + 1, len(last) + 1):
            cand = (head_part + last[:extra_len]).upper()[:8]
            if cand not in used_codes:
                return cand
    else:
        # 단일 단어: 글자 수 늘려가며 시도
        w = words[0]
        for extra_len in range(4, len(w) + 1):
            cand = w[:extra_len].upper()[:8]
            if cand not in used_codes:
                return cand

    # 5단계: 숫자 접미사 폴백
    for n in range(2, 100):
        cand = f"{base}{n}"[:8]
        if cand not in used_codes:
            return cand
    return base  # 극단 케이스 — 호출부에서 추가 처리


def resolve_screen_character(middle_name: str, screen_map: dict) -> str:
    """Middle 이름의 성격(character) 반환. 맵에 없으면 빈 문자열.

    비기능 TC 조건부 생성 (원칙 E)에서 프롬프트 힌트로 사용.
    """
    if not middle_name or middle_name not in screen_map:
        return ""
    entry = screen_map[middle_name]
    if isinstance(entry, dict):
        return entry.get("character", "")
    return ""


def resolve_screen_navigation(middle_name: str, screen_map: dict) -> str:
    """Middle 이름의 네비게이션 분류 반환. 맵에 없으면 빈 문자열.

    뒤로가기/스와이프 TC 조건부 생성 (원칙 F)에서 프롬프트 힌트로 사용.
    """
    if not middle_name or middle_name not in screen_map:
        return ""
    entry = screen_map[middle_name]
    if isinstance(entry, dict):
        return entry.get("navigation", "")
    return ""


def detect_same_result_duplicates(tc_md: str) -> dict:
    """같은 화면 안에서 '예상 결과' 가 거의 동일한 TC 그룹을 검출.

    정책 A 의 후처리 안전망: AI 가 통합을 안 했을 경우 검토자가 확인할 수 있도록
    의심 그룹을 리포트로 만든다. 자동 병합은 하지 않음 (false positive 위험).

    매칭 기준:
      - 같은 대분류 + 같은 중분류 (= 같은 화면)
      - 사전 조건의 핵심 줄 1~3 개가 같음 (마지막 줄 "...상태" 정규화 비교)
      - 예상 결과의 핵심 결과 1~2개가 같음 (불릿 정규화 후 80% 이상 일치)

    Returns:
        {
          "groups": [
            {
              "middle": "Order Confirm",
              "common_result": "시트가 닫히고 SCR-102 로 복귀한다",
              "tcs": [
                {"id": "SM-TRAD-012", "title": "Cancel 버튼 탭으로 시트 닫기"},
                {"id": "SM-TRAD-013", "title": "배경 오버레이 탭으로 시트 닫기"},
              ],
            },
            ...
          ],
          "total_suspect": 4,    # 통합 대상이 될 수 있는 TC 수
        }
    """
    from collections import defaultdict

    # TC 블록 파싱 — 각 TC 의 (대분류, 중분류, 예상 결과 정규화) 추출
    tc_blocks = re.split(r"\n(?=###\s)", tc_md)
    tcs_by_screen: dict[tuple[str, str], list[dict]] = defaultdict(list)

    for block in tc_blocks:
        if not block.strip().startswith("### "):
            continue
        # TC ID 추출
        m_id = re.search(r"###\s+\*?\*?([A-Z]+-[A-Z]+-\d+)", block)
        if not m_id:
            continue
        tc_id = m_id.group(1)
        # 제목
        m_title = re.search(r"###\s+\*?\*?[A-Z]+-[A-Z]+-\d+\*?\*?\s*[—\-]\s*(.+)", block)
        title = m_title.group(1).strip() if m_title else ""
        # 대/중분류
        m_major = re.search(r"\|\s*대분류\s*\|\s*([^|]+?)\s*\|", block)
        m_middle = re.search(r"\|\s*중분류\s*\|\s*([^|]+?)\s*\|", block)
        major = m_major.group(1).strip() if m_major else ""
        middle = m_middle.group(1).strip() if m_middle else ""
        # 예상 결과 — 첫 2줄만 (핵심 검증 지표)
        m_then = re.search(r"\*\*예상\s*결과\*\*\s*\n((?:[^\n]+\n){1,5})", block)
        then_lines = []
        if m_then:
            for line in m_then.group(1).split("\n"):
                norm = re.sub(r"^[\-\*\d.\s]+", "", line).strip()
                norm = re.sub(r"\s+", " ", norm)
                # 의미 없는 짧은 줄 제외
                if len(norm) >= 8:
                    then_lines.append(norm)
        # 핵심 결과 키 (앞 2줄)
        result_key = " | ".join(sorted(then_lines[:2])) if then_lines else ""
        if not result_key or not middle:
            continue

        tcs_by_screen[(major, middle, result_key)].append({
            "id": tc_id, "title": title,
        })

    # 같은 키에 2개 이상 → 의심 그룹
    groups = []
    total_suspect = 0
    for (major, middle, result_key), tcs in tcs_by_screen.items():
        if len(tcs) >= 2:
            groups.append({
                "middle": middle,
                "common_result": result_key[:120] + ("..." if len(result_key) > 120 else ""),
                "tcs": tcs,
            })
            total_suspect += len(tcs) - 1  # 통합 시 줄어들 수

    return {"groups": groups, "total_suspect": total_suspect}


def detect_duplicate_error_tcs(tc_md: str) -> dict:
    """원칙 G — 같은 그룹 내 동일 비기능 패턴 TC 가 여러 화면에 반복되는 경우 검출.

    탐지 패턴 (제목 + 분류 기반):
      - 네트워크 끊김 / 연결 실패
      - 타임아웃
      - 3G / 저사양 / 로딩 시간
      - 백그라운드 / 포그라운드
      - 화면 회전
      - 메모리 부족 / 강제 종료

    Returns:
        {
          "patterns": [
            {
              "pattern": "네트워크 끊김",
              "major": "01.로그인_Gmail",
              "tcs": [
                {"id": "SM-LGI-007", "title": "...", "middle": "Google Sign-in"},
                {"id": "SM-LGI-013", "title": "...", "middle": "OAuth Confirm"},
              ],
              "count": 2,
            },
            ...
          ],
          "total_duplicates": 12,    # 통합 시 줄어들 TC 수
          "suggested_keep": [...],    # 유지 권장 TC IDs (대표 화면)
          "suggested_remove": [...],  # 제거 권장 TC IDs
        }

    중복 패턴이 없으면 patterns=[], total_duplicates=0.
    """
    # 패턴 키워드 — 한글/영문 모두 매칭
    patterns_def = [
        ("네트워크 끊김", [r"네트워크\s*(끊김|차단|단절|장애|불안정)", r"연결\s*실패", r"오프라인", r"network\s*(disconnect|fail)", r"connection\s*(lost|fail)"]),
        ("타임아웃", [r"타임\s*아웃", r"응답\s*지연", r"timeout"]),
        ("3G/저사양 로딩 시간", [r"3G\s*(환경|네트워크)?", r"저사양", r"로딩\s*시간", r"loading\s*time"]),
        ("백그라운드/포그라운드", [r"백그라운드", r"포그라운드", r"background", r"foreground"]),
        ("화면 회전", [r"화면\s*회전", r"orientation\s*change"]),
        ("메모리/강제 종료", [r"메모리\s*부족", r"앱\s*(강제\s*)?종료", r"out\s*of\s*memory", r"force\s*(close|quit|kill)"]),
    ]

    if not tc_md:
        return {"patterns": [], "total_duplicates": 0, "suggested_keep": [], "suggested_remove": []}

    # TC 블록 분할 — `### ` 또는 `### **` 헤더 기준
    blocks = re.split(r'(?=^###\s)', tc_md, flags=re.MULTILINE)

    detected = {}  # {(pattern, major): [tc_info, ...]}

    for block in blocks:
        if not block.strip().startswith("###"):
            continue
        # TC ID 추출
        m_id = re.match(r'###\s+\*?\*?([A-Z0-9-]+(?:-\d+)+)\*?\*?\s*(?:—|-)?\s*(.*?)(?:\n|$)', block)
        if not m_id:
            continue
        tc_id = m_id.group(1).strip()
        title = m_id.group(2).strip().rstrip('*').strip()
        # 대분류 / 중분류 추출 (테이블 형식)
        m_major = re.search(r'\|\s*대분류\s*\|\s*([^|]+?)\s*\|', block)
        m_middle = re.search(r'\|\s*중분류\s*\|\s*([^|]+?)\s*\|', block)
        major = m_major.group(1).strip() if m_major else ""
        middle = m_middle.group(1).strip() if m_middle else ""
        if not major:
            continue

        # 패턴 매칭 (title + middle 텍스트 합쳐서 검색)
        search_text = f"{title}\n{middle}".lower()
        for pattern_label, regexes in patterns_def:
            matched = any(re.search(rgx, search_text, re.IGNORECASE) for rgx in regexes)
            if matched:
                key = (pattern_label, major)
                detected.setdefault(key, []).append({
                    "id": tc_id,
                    "title": title,
                    "middle": middle,
                })
                break  # 한 TC 가 여러 패턴 매칭돼도 가장 먼저 잡힌 것만

    # 결과 가공: 같은 (패턴, 대분류) 에 2개 이상 있으면 중복으로 간주
    patterns_out = []
    suggested_keep = []
    suggested_remove = []
    total_duplicates = 0
    for (pattern_label, major), tcs in detected.items():
        if len(tcs) < 2:
            continue
        # TC ID 의 가장 작은 번호를 대표(유지) 로 — 일반적으로 entry 화면이 먼저 등장
        tcs_sorted = sorted(tcs, key=lambda t: t["id"])
        keep = tcs_sorted[0]
        remove_list = tcs_sorted[1:]
        suggested_keep.append(keep["id"])
        suggested_remove.extend([t["id"] for t in remove_list])
        total_duplicates += len(remove_list)
        patterns_out.append({
            "pattern": pattern_label,
            "major": major,
            "tcs": tcs_sorted,
            "count": len(tcs_sorted),
            "keep": keep["id"],
            "remove": [t["id"] for t in remove_list],
        })

    return {
        "patterns": patterns_out,
        "total_duplicates": total_duplicates,
        "suggested_keep": suggested_keep,
        "suggested_remove": suggested_remove,
    }


def predict_classification_duplicates(classification_md: str) -> dict:
    """분류표(TC 작성 전) 단계에서 중복 가능성 예측 — Stage 1.

    TC ID 가 아직 없으므로 휴리스틱 기반:
      1. 같은 대분류 내 중분류 5개 이상 → AI 가 비기능 TC 반복 작성 가능성 높음
      2. 같은 대분류 내 여러 중분류에 같은 소분류 키워드 등장 (예: 모두에 "로딩") → 통합 후보

    Returns:
      {
        "predictions": [
          {
            "type": "group_size" | "shared_minor",
            "major": "01.로그인_Gmail",
            "middle_count": 5,
            "middles": [...],
            "shared_keyword": "로딩 시간" (shared_minor 일 때만),
            "predicted_pattern": "비기능 TC (네트워크/타임아웃/로딩) 가 반복될 가능성 높음",
            "recommendation": "그룹 entry 화면 1개에만 비기능 TC 작성 권장"
          },
          ...
        ],
        "high_risk_count": int,
      }
    """
    if not classification_md:
        return {"predictions": [], "high_risk_count": 0}

    # 분류표 파싱: ## 대분류 / ### 중분류 / 소분류 불릿
    majors = {}  # {major_name: {middle_name: [minor1, minor2, ...]}}
    cur_major = None
    cur_middle = None
    for line in classification_md.split('\n'):
        line = line.rstrip()
        # 대분류 헤더: ## 또는 ## 대분류:
        m_major = re.match(r'^##\s+(?:대분류[:\s]*)?(.+?)\s*$', line)
        m_middle = re.match(r'^###\s+(?:중분류[:\s]*)?(.+?)\s*$', line)
        if m_major and not line.startswith('###') and not line.startswith('####'):
            name = m_major.group(1).strip()
            # '소분류' 이런 건 대분류가 아님
            if name and name not in ('소분류', '중분류'):
                cur_major = name
                cur_middle = None
                majors.setdefault(cur_major, {})
        elif m_middle and cur_major:
            mid_name = m_middle.group(1).strip()
            if mid_name and mid_name != '소분류':
                cur_middle = mid_name
                majors[cur_major].setdefault(cur_middle, [])
        elif cur_major and cur_middle:
            # 소분류 불릿: - 또는 *
            m_minor = re.match(r'^\s*[-*]\s+(.+?)\s*$', line)
            if m_minor:
                minor_text = m_minor.group(1).strip()
                # 콜론 앞부분만 추출 (예: "로딩 시간: 3G 환경..." → "로딩 시간")
                minor_label = minor_text.split(':', 1)[0].split('—', 1)[0].strip()
                if minor_label and minor_label != '소분류':
                    majors[cur_major][cur_middle].append(minor_label)

    predictions = []

    # 휴리스틱 1: 그룹 크기 (중분류 5개 이상)
    SIZE_THRESHOLD = 5
    for major_name, middles in majors.items():
        if len(middles) >= SIZE_THRESHOLD:
            predictions.append({
                "type": "group_size",
                "major": major_name,
                "middle_count": len(middles),
                "middles": list(middles.keys()),
                "predicted_pattern": "비기능 TC (네트워크/타임아웃/로딩 시간 등) 가 여러 화면에 반복 작성될 가능성 높음",
                "recommendation": "그룹의 entry 성격 화면 1개에만 비기능 TC 작성 권장 (원칙 G)"
            })

    # 휴리스틱 2: 같은 대분류 내 여러 중분류에 동일 소분류 키워드 등장
    DUPLICATE_KEYWORDS = [
        '로딩 시간', '네트워크', '타임아웃', '백그라운드', '포그라운드',
        '화면 회전', '메모리', '강제 종료', '오프라인',
    ]
    for major_name, middles in majors.items():
        # 각 키워드가 몇 개 중분류에 등장하는지 카운트
        keyword_hits = {kw: [] for kw in DUPLICATE_KEYWORDS}
        for mid_name, minors in middles.items():
            for minor_label in minors:
                for kw in DUPLICATE_KEYWORDS:
                    if kw in minor_label:
                        keyword_hits[kw].append(mid_name)
                        break
        for kw, mids_with_kw in keyword_hits.items():
            if len(mids_with_kw) >= 2:
                # 그룹 크기 휴리스틱과 중복 안 되게 — 같은 major 가 이미 group_size 로 들어가 있으면 키워드만 추가
                existing = next((p for p in predictions if p["major"] == major_name and p["type"] == "group_size"), None)
                if existing:
                    existing.setdefault("shared_keywords", []).append({"keyword": kw, "middles": mids_with_kw})
                else:
                    predictions.append({
                        "type": "shared_minor",
                        "major": major_name,
                        "middle_count": len(middles),
                        "middles": list(middles.keys()),
                        "shared_keyword": kw,
                        "shared_in_middles": mids_with_kw,
                        "predicted_pattern": f'"{kw}" 관련 소분류가 여러 중분류에 등장 — 통합 후보',
                        "recommendation": f'"{kw}" 검증은 그룹 대표 화면 1개에서만 권장'
                    })

    return {
        "predictions": predictions,
        "high_risk_count": len(predictions),
    }


def disambiguate_duplicate_minors(tc_md: str) -> str:
    """같은 (대분류, 중분류) 안에서 동일한 소분류 이름이 여러 TC에 쓰인 경우,
    각 TC의 제목(### 헤더 뒤 — 이후 부분)에서 핵심 키워드를 추출하여
    소분류 뒤에 ` — {키워드}` 형태로 자동 부여한다.

    예: 같은 중분류 'Splash'에 소분류 'Splash 화면 진입' × 4개
        TC들의 제목이 'UI 요소 표시', 'Get Started 버튼 탭', '뒤로가기 동작', '로딩 시간' 이라면
        → 각각 'Splash 화면 진입 — UI 요소 표시', '... — Get Started 버튼 탭' ...

    원칙:
    - 같은 (대,중,소) 그룹의 TC가 1개뿐이면 변경하지 않음 (불필요한 변형 회피)
    - 이미 다른 소분류 이름이면 변경하지 않음
    - 변형 결과의 길이는 60자 이내로 제한 (Excel 가독성)
    - title이 비어있거나 추출 실패 시 (1), (2) 같은 일련번호 fallback
    """
    if not tc_md:
        return tc_md

    # 1단계: TC 블록 단위로 분리 + 각 블록에서 (major, middle, minor, title) 추출
    blocks = re.split(r"(?m)^(?=###\s)", tc_md)
    parsed = []  # [(idx, block_text, major, middle, minor, title), ...]
    for idx, block in enumerate(blocks):
        if not block.lstrip().startswith("###"):
            parsed.append((idx, block, None, None, None, None))
            continue
        first_line = block.split("\n", 1)[0]
        # 카테고리/대중소분류 헤더 등 제외
        if re.match(r"^###\s*(카테고리|대분류|중분류|소분류|Category)\s*[:：\d]", first_line, re.IGNORECASE):
            parsed.append((idx, block, None, None, None, None))
            continue
        # 제목 추출 (— 뒤)
        cleaned = re.sub(r"\*\*", "", first_line)
        m = re.match(r"^###\s+\S+\s+—\s+(.+?)\s*$", cleaned)
        title = m.group(1).strip() if m else ""
        # 메타 테이블 필드
        major = _extract_md_field(block, "대분류")
        middle = _extract_md_field(block, "중분류")
        minor = _extract_md_field(block, "소분류")
        parsed.append((idx, block, major, middle, minor, title))

    # 2단계: (major, middle, minor) 그룹별 등장 인덱스 수집
    from collections import defaultdict as _dd
    groups = _dd(list)
    for idx, block, major, middle, minor, title in parsed:
        if major and middle and minor:
            groups[(major, middle, minor)].append((idx, title))

    # 3단계: 그룹 크기 >= 2인 경우만 변형 적용
    new_minors_per_idx: dict[int, str] = {}
    for (major, middle, minor), entries in groups.items():
        if len(entries) < 2:
            continue
        # 각 TC의 차별화 키워드 추출
        used_variants: set[str] = set()
        for ord_n, (idx, title) in enumerate(entries, 1):
            variant = _extract_minor_variant(title, minor)
            if not variant:
                # title에서 추출 불가 → 일련번호 폴백
                variant = f"({ord_n})"
            # 길이 제한 — 잘림으로 인한 'Empty S' 같은 어색한 표기 방지.
            # 합계 한도를 넉넉히 잡고(150자), 그래도 초과하면 잘리지 않고 일련번호 폴백.
            total_max = 150
            if len(minor) + 3 + len(variant) > total_max:
                variant = f"({ord_n})"
            # 같은 그룹 내 variant 충돌 방지
            base_variant = variant
            n = 2
            while variant in used_variants:
                variant = f"{base_variant} ({n})"
                n += 1
            used_variants.add(variant)
            new_minors_per_idx[idx] = f"{minor} — {variant}"

    if not new_minors_per_idx:
        return tc_md  # 변형 없음

    # 4단계: 각 블록의 `| 소분류 | ... |` 라인을 새 값으로 치환
    out_blocks = []
    for idx, block, major, middle, minor, title in parsed:
        if idx in new_minors_per_idx and minor:
            new_minor = new_minors_per_idx[idx]
            # `| 소분류 | <기존> |` 첫 번째 매칭만 치환
            block = re.sub(
                r"(\|\s*소분류\s*\|\s*)" + re.escape(minor) + r"(\s*\|)",
                r"\g<1>" + new_minor.replace("\\", "\\\\") + r"\g<2>",
                block, count=1
            )
        out_blocks.append(block)
    return "".join(out_blocks)


def _extract_md_field(block: str, field: str) -> str:
    """TC 블록의 마크다운 테이블에서 특정 필드 값 추출."""
    m = re.search(rf"\|\s*{re.escape(field)}\s*\|\s*([^|]+?)\s*\|", block)
    return m.group(1).strip() if m else ""


def _extract_minor_variant(title: str, minor: str) -> str:
    """TC title에서 소분류와 차별화되는 키워드 추출.

    전략:
    1. title이 minor를 prefix/suffix로 이미 포함하면 그 부분만 잘라낸 나머지 사용
    2. minor의 핵심 단어와 겹치는 부분 제거 후 남은 핵심 부분
    3. 결과가 너무 짧거나 어색하면 title 그대로 사용 (단, 길이 제한)
    """
    if not title:
        return ""
    cleaned = re.sub(r"\s+", " ", title).strip()
    norm_title = cleaned
    norm_minor = minor.strip()

    # 1) prefix/suffix 일치 — minor가 그대로 포함되면 잘라내기
    #    (잘림 한도 제거 — disambiguate_duplicate_minors 의 total_max(150)가 최종 가드)
    if norm_minor and norm_title.startswith(norm_minor):
        leftover = norm_title[len(norm_minor):].strip(" -—:")
        if len(leftover) >= 4:
            return _trim_leading_function_words(leftover)
    if norm_minor and norm_title.endswith(norm_minor):
        leftover = norm_title[:-len(norm_minor)].strip(" -—:")
        if len(leftover) >= 4:
            return _trim_leading_function_words(leftover)

    # 2) minor의 단어와 title이 부분 일치 — title 자체가 더 구체적인 표현이면 그대로 사용
    minor_set = set(re.split(r"[\s\-—:_/()]+", norm_minor))
    title_words = re.split(r"[\s\-—:_/()]+", norm_title)
    # title 단어가 minor 단어보다 많으면 (= 더 구체적) title 사용
    if len([w for w in title_words if w not in minor_set and len(w) >= 2]) >= 2:
        return norm_title

    # 3) 마지막 폴백 — title 그대로
    return norm_title


def _trim_leading_function_words(s: str) -> str:
    """문장 시작의 어색한 조사/접속어 제거 (예: '시 키보드 표시' → '키보드 표시')."""
    # 한국어 조사/접속어 + 영어 의미 없는 시작 단어
    pattern = r"^(시|에|의|에서|으로|로|가|이|는|은|을|를|와|과|the|a|an)\s+"
    while True:
        new_s = re.sub(pattern, "", s, flags=re.IGNORECASE)
        if new_s == s:
            break
        s = new_s
    return s.strip(" -—:")


def renumber_tc_ids(tc_md: str, project_code: str, suite_code: str, starting_seq: int) -> tuple[str, int]:
    """Markdown 내 TC ID(`{PC}-{SC}-NNN`)를 starting_seq부터 연속 재번호.

    같은 TC 블록(### ...) 안에서 여러 번 나오는 동일 ID는 하나로 취급.
    Suite 내 SeqNumber 전역 유니크 규칙 (tc-rules.md 섹션 2-1) 강제.

    Returns:
        (renumbered_md, next_seq_after_last)
    """
    if not suite_code:
        return tc_md, starting_seq

    pattern = re.compile(
        rf"({re.escape(project_code)}-{re.escape(suite_code)}-)(\d{{3,}})",
        re.IGNORECASE,
    )

    # 블록 분할: "### " 헤더 기준
    blocks = re.split(r"(?m)(?=^### )", tc_md)
    next_seq = starting_seq
    id_remap: dict[str, str] = {}

    for i, block in enumerate(blocks):
        header_match = re.match(r"^### .*", block)
        if not header_match:
            continue
        header = header_match.group(0)
        first = pattern.search(header)
        if not first:
            continue
        old_id = first.group(0)
        if old_id not in id_remap:
            new_id = f"{first.group(1)}{next_seq:03d}"
            id_remap[old_id] = new_id
            next_seq += 1

    # 모든 TC ID 치환 (블록 헤더 + 본문 참조 모두)
    def _sub(m: re.Match) -> str:
        old = m.group(0)
        return id_remap.get(old, old)

    return pattern.sub(_sub, tc_md), next_seq


def build_tc_user_prompt(domain: dict, features_text: str, policy_text: str,
                          project_name: str, classification: str,
                          starting_seq: int = 1,
                          screen_character: str = "",
                          screen_navigation: str = "",
                          source_scr_map: dict | None = None) -> str:
    # 해당 대분류의 분류 섹션 추출
    domain_section = extract_domain_section(classification, domain["code"])
    project_code = _detect_project_code(project_name)

    suite_code = domain.get("suite_code", "").strip().strip("-")  # trailing 하이픈 제거
    if suite_code:
        seq_str = f"{starting_seq:03d}"
        next_str = f"{starting_seq+1:03d}"
        example_id = f"{project_code}-{suite_code}-{seq_str}"
        tc_id_instruction = (
            f"⚠️ TC ID 형식: `{example_id}`, `{project_code}-{suite_code}-{next_str}`, ... "
            f"하이픈은 정확히 이 위치에만. `{project_code}-{suite_code}--001` 처럼 하이픈이 연속되면 안 됩니다.\n"
            f"⚠️ **Suite 내 SeqNumber 전역 유니크**: 이 호출의 시작 번호는 `{seq_str}`. "
            f"중분류가 바뀌어도 001로 리셋하지 말고 {seq_str}부터 연속 증가. "
            f"(tc-rules.md 섹션 2-1 SeqNumber 전역 유니크 규칙)"
        )
    else:
        tc_id_instruction = f"⚠️ TC ID의 ProjectCode: `{project_code}` (예: {project_code}-XXXX-001). `TC-`로 시작하면 안 됩니다."

    focus_middle = domain.get("_focus_middle", "")

    # 화면 성격별 비기능 TC 힌트 (tc-rules.md 원칙 E)
    nfr_instruction = ""
    if screen_character and screen_character != "static":
        guide = SCREEN_CHARACTER_NFR_GUIDE.get(screen_character, "")
        if guide:
            screen_label = focus_middle or "화면"
            nfr_instruction = (
                f"\n## ⚙️ 비기능 TC 필수 포함 (원칙 E — 화면 성격: `{screen_character}`)\n\n"
                f"이 화면은 `{screen_character}` 성격이므로 아래 비기능 TC를 **반드시 1개 이상** 포함하세요:\n"
                f"- {guide}\n\n"
                f"TC 작성 예시 (entry 성격 화면):\n"
                f"```\n"
                f"### {project_code}-{suite_code}-XXX — {screen_label} 로딩 시간\n"
                f"| 분류 | Edge |\n| 우선순위 | Medium |\n\n"
                f"**사전 조건**\n1. 3G 또는 저사양 기기 환경인 상태\n\n"
                f"**테스트 단계**\n화면 완전 표시까지의 시간을 측정한다.\n\n"
                f"**예상 결과**\n- (업계 표준 임계치) 이내에 표시된다\n\n"
                f"**비고**\n- [미결] 임계치는 업계 표준 기반 추정. PM 확정 필요.\n"
                f"```\n"
            )
    elif screen_character == "static":
        nfr_instruction = (
            f"\n## ⚙️ 비기능 TC (원칙 E — 화면 성격: `static`)\n\n"
            f"이 화면은 정적 컨텐츠이므로 **비기능 TC는 생성하지 마세요** (로딩·성능·실시간 TC 불필요).\n"
        )

    # 네비게이션 분류별 뒤로가기/스와이프 TC 힌트 (tc-rules.md 원칙 F)
    nav_instruction = ""
    if screen_navigation:
        nav_guide = SCREEN_NAVIGATION_TC_GUIDE.get(screen_navigation, "")
        screen_label = focus_middle or "화면"
        if screen_navigation == "one-way":
            nav_instruction = (
                f"\n## ⚠️ 뒤로가기/스와이프 TC 필수 (원칙 F — 네비게이션: `one-way`)\n\n"
                f"이 화면은 **완료 화면**이므로 뒤로가기 시 이전 화면 복귀를 차단해야 합니다:\n"
                f"- {nav_guide}\n\n"
                f"**반드시 아래 Negative TC를 1개 포함하세요** (우선순위 High):\n"
                f"```\n"
                f"### {project_code}-{suite_code}-XXX — {screen_label} 후 뒤로가기/스와이프 차단\n"
                f"| 분류 | Negative |\n| 우선순위 | High |\n\n"
                f"**사전 조건**\n1. 모바일 브라우저에서 접속한 상태\n"
                f"2. {screen_label} 화면에 진입 완료한 상태\n\n"
                f"**테스트 단계**\n브라우저 뒤로가기 버튼 또는 iOS 스와이프 백 제스처를 실행한다.\n\n"
                f"**예상 결과**\n- 이전 화면(인증·주문 등)으로 돌아가지 않는다\n"
                f"- 홈 이동 / 뒤로가기 차단 / 확인 모달 중 하나가 발생한다\n"
                f"- 이미 완료된 플로우를 재실행할 수 없는 상태가 유지된다\n\n"
                f"**비고**\n- [미결] 구체적 차단 방식은 PM 확정 필요\n"
                f"```\n"
            )
        elif screen_navigation in ("tab-root", "sequential", "overlay", "detail"):
            nav_instruction = (
                f"\n## ⚙️ 뒤로가기/스와이프 TC 포함 (원칙 F — 네비게이션: `{screen_navigation}`)\n\n"
                f"이 화면의 뒤로가기 동작을 검증하는 TC를 1개 포함하세요:\n"
                f"- {nav_guide}\n"
            )
        # static은 TC 불필요 (별도 instruction 없음)

    middle_instruction = ""
    if focus_middle:
        middle_instruction = f"""
⚠️ 이 호출에서는 중분류 "{focus_middle}"에 해당하는 TC만 상세하게 작성하세요.
다른 중분류의 TC는 작성하지 마세요. "{focus_middle}" 중분류의 모든 소분류에 대해 빠짐없이 작성하세요.
"""

    # 화면 코드는 입력 소스 자체에서 추출 (v0.9.13~)
    # 우선순위:
    #   1) source_scr_map: step_parse_sources 가 파일명/H1/본문에서 미리 뽑은 매핑 — 가장 정확
    #   2) _detect_screen_codes_for_middle: features_text 에서 중분류 이름 근처 SCR 패턴 탐색 (보조)
    #   3) 둘 다 없으면 → 빈 칸 명시 (AI 환각 금지)
    screen_code_hint = ""
    if focus_middle:
        # source_scr_map 에서 입력 파일별 매핑이 있으면 모두 안내 — AI 가 본문 컨텍스트로 해당 TC 의 SCR 결정
        scr_table_lines = []
        if source_scr_map:
            for fn, scr in source_scr_map.items():
                scr_table_lines.append(f"| {fn} | {scr} |")

        detected = _detect_screen_codes_for_middle(focus_middle, features_text)

        if scr_table_lines:
            scr_table = "\n".join(scr_table_lines)
            primary_example = list(source_scr_map.values())[0]
            allowed_scrs = list(source_scr_map.values())
            allowed_scrs_str = ", ".join(f"`{s}`" for s in allowed_scrs)
            screen_code_hint = f"""
## 📱 화면 코드 매핑 (입력 소스 기반 — 반드시 준수)

이 분류표는 다음 입력 소스에서 추출되었습니다. 각 TC 의 `| 연관 화면 |` 필드에는 **해당 TC 가 어느 입력 소스에서 나왔는지 기준으로** 아래 매핑 표의 화면 코드를 정확히 사용하세요:

| 입력 소스 파일 | 화면 코드 |
|---|---|
{scr_table}

## 🚨 입력 소스 범위 엄격 준수 (v0.9.26~ 절대 규칙)

**이 작업의 입력 화면 코드 = {allowed_scrs_str} — 이 외에는 절대 다루지 마세요.**

⚠️ 다음과 같은 상황에 주의하세요:

- 입력 파일 본문에 다른 SCR 코드(예: SCR-403, SCR-221, SCR-410)가 표/참고/링크로 언급될 수 있습니다
- 본문 끝의 부록 표("**에러 케이스 (프로필/알림/PnL)**" 같은 형식) 에 다른 화면들의 정보가 있을 수 있습니다
- 인터랙션 설명에서 navigateTo('scr-XXX') 같은 다른 화면 참조가 있을 수 있습니다

**위 모든 경우에 다른 화면들의 분류/TC 는 만들지 마세요.**

✅ **올바른 행동**:
- 위 매핑 표의 SCR 코드만 분류표/TC ID 에 사용
- 다른 SCR 이 본문에 등장해도 무시
- "Notifications", "PnL", "Profile" 등 무관한 키워드는 명시적으로 입력 파일에 없으면 분류 금지

❌ **금지 행동**:
- 본문 부록에 SCR-221 (알림) 정보가 있다고 "Notifications" 중분류 추가 금지
- navigateTo('scr-403') 참조 봤다고 "Profile" 분류 추가 금지
- 매핑 표에 없는 SCR 코드를 임의로 만들기 금지

규칙:
1. 각 TC 메타 테이블에 `| 연관 화면 | SCR-NNN |` 형식 정확히 사용 (예: `| 연관 화면 | {primary_example} |`)
2. 화면명 없이 코드만: "SCR-001" ✅ / "Splash (SCR-001)" ❌
3. 한 TC 가 여러 화면에 걸치면 쉼표 구분: `SCR-001, SCR-003` (단, 둘 다 위 매핑 표에 있어야 함)
4. **임의 코드 생성 금지**: 위 매핑 표에 없는 화면 코드(SCR-999, LGI, ABC, 또는 입력 외 다른 SCR)는 절대 만들지 마세요
5. 출처 입력 소스에 매핑이 없으면 `| 연관 화면 |  |` 빈 칸으로 두세요
"""
        elif detected:
            # source_scr_map 은 비어있지만 features_text 에서 SCR 패턴 발견 (예: 본문에 SCR-NNN 언급)
            codes_str = ", ".join(detected[:5])
            screen_code_hint = f"""
## 📱 연관 화면 코드 (입력 본문에서 탐지)

이 중분류 "{focus_middle}" 와 관련해 입력 본문에서 다음 화면 코드를 발견했습니다: **{codes_str}**

⚠️ **규칙**:
1. 각 TC 메타 테이블에 `| 연관 화면 | {detected[0]} |` 형식 사용
2. 위에 나열된 코드만 사용 — **임의 코드(LGI, ABC, SCR-999 등) 생성 금지**
3. 본문에 SCR 코드가 명시되지 않은 TC 는 `| 연관 화면 |  |` 필드를 비워두세요
"""
        else:
            # 입력 소스에 화면 식별자가 전혀 없음 — 빈 칸 명시
            screen_code_hint = f"""
## 📱 연관 화면 (입력 소스에 화면 식별자 없음)

⚠️ 이 입력 소스에서 화면 코드(SCR-NNN 형식)가 발견되지 않았습니다.

**규칙**: 각 TC 메타 테이블의 `| 연관 화면 |` 필드를 **반드시 빈 칸**으로 두세요:
```
| 연관 화면 |  |
```
- 임의 코드(SCR-001, LGI 등) 생성 금지
- 중분류 이름을 화면 코드 자리에 넣지 마세요
- 빈 칸이 정상입니다. 입력 소스에 식별자가 추가되면 다음 실행에서 자동 채워집니다.
"""

    # 분류표 우선 원칙 (Gate 검토 후 사용자 수정사항이 본문/features 보다 우선)
    classification_priority_hint = """
## 🎯 분류표 우선 원칙 (절대 규칙 — 반드시 준수)

⚠️ **사용자가 검토·승인한 분류표가 최종 진실 소스입니다.**

분류표는 단순 목차가 아닙니다. 사용자가 Gate 검토 단계에서 다음과 같은 수정을 했을 수 있습니다:
- 불필요한 기능 (예: 체크박스, 광고 영역 등) 제거
- 중분류/소분류 통합 또는 분리
- 시나리오 우선순위 조정
- 범위 외 항목 제외

**그러므로**:

1. **분류표에 있는 중분류/소분류만 TC 로 작성하세요.**
2. **본문(features_text, 정책)에 추가 기능이 보여도, 분류표에 없으면 TC 만들지 마세요.**
3. **분류표 ⊃ 본문이 아닙니다.** 본문이 더 자세할 수 있지만 분류표가 우선입니다.
4. 본문은 TC **상세 시나리오 작성을 위한 참고**일 뿐 (입력값/조건/결과 등). **항목 자체의 포함 여부 결정에는 사용 금지.**
5. 만약 본문에는 있는데 분류표에는 없는 기능이 있다면: **사용자가 의도적으로 제외한 것** — TC 만들지 마세요.

예시:
- 분류표에 "이메일 입력" 만 있고 본문에 "체크박스 동의" 가 있을 때
  → 체크박스 동의 TC 만들지 않음 (사용자가 의도적으로 제거한 것)
- 분류표에 "로그인" 이 있을 때
  → 본문의 로그인 관련 상세 시나리오를 참고하여 TC 작성 (정상)
"""

    # 그룹 단위 에러 패턴 통합 안내 (tc-rules.md 원칙 G 보강)
    dedup_hint = """
## 🔁 에러/예외 패턴 통합 규칙 (tc-rules.md 원칙 G — 반드시 준수)

⚠️ **같은 그룹/대분류의 여러 화면에 동일 패턴 비기능 TC 를 반복 작성하지 마세요.**

### 보존 (스펙 명시)
- 기획서의 "에러 케이스" 표에 **명시된** 에러 → 해당 화면 TC 로 보존
  (예: SCR-013 의 "권한 거부 Modal", "토큰 수신 실패 Modal")

### 통합 (AI 추가)
다음 패턴은 **그룹당 1개 대표 화면에만** 작성:
- 네트워크 끊김 / 연결 실패
- 타임아웃 처리
- 3G/저사양 환경 로딩 시간
- 백그라운드 → 포그라운드 복귀
- 화면 회전 / 메모리 부족

### 대표 화면 선택
1. entry 성격 화면 우선 (그룹의 진입 지점)
2. entry 가 없으면 그룹의 첫 번째 화면
3. 사용자가 입력한 화면 중에서만 선택

### 통합 TC 형식
```
**비고**
- [통합] 그룹 단위 비기능 검증 — 같은 패턴을 다른 화면별로 반복 작성하지 않음 (원칙 G).
```

### ❌ 금지 예시
같은 그룹 SCR-010, 011, 012, 013, 014 모두에 "네트워크 끊김" TC 5개 작성 → 1개로 통합 필수
"""

    # 마침표 종결 강화 안내 (tc-rules.md 9-2 절 보강)
    period_hint = """
## ✍️ 문장 종결 규칙 (tc-rules.md 9-2 — 반드시 준수)

**테스트 단계** 와 **예상 결과** 의 **모든 항목은 마침표(`.`)로 끝**나야 합니다.

✅ 올바른 예:
```
**테스트 단계**
1. Continue 버튼을 탭한다.

**예상 결과**
- 인증 코드 입력 화면으로 이동한다.
- AppState 의 email 값이 갱신된다.
```

❌ 잘못된 예:
```
**예상 결과**
- 인증 코드 입력 화면으로 이동한다  ← 마침표 누락
- email 값 갱신                     ← 단어형 + 마침표 누락
```

사전 조건은 명사형 (`~인 상태`, `~된 상태`) 이므로 마침표 불필요.
"""

    return f"""프로젝트: {project_name}
대분류: {domain['name']} ({domain['code']})
{tc_id_instruction}
{middle_instruction}
{nfr_instruction}
{nav_instruction}
{screen_code_hint}
{classification_priority_hint}
{dedup_hint}
{period_hint}
## 이 대분류의 분류 구조 (사용자 승인 — 최종 진실 소스)
{domain_section}

## 전체 기능 목록 (참고용 — 분류표가 우선)
⚠️ 아래 정보는 **TC 상세 작성을 위한 참고**입니다. 항목 포함 여부는 위 분류표 기준으로 결정하세요.
{features_text[:15000]}

## 관련 정책 (참고용 — 분류표가 우선)
⚠️ 아래 정책은 **TC 의 사전조건/예상결과 작성에 참고**합니다. 분류표에 없는 기능은 정책이 있어도 TC 만들지 마세요.
{policy_text[:15000]}

위 대분류({domain['name']}){' 중 "' + focus_middle + '" 중분류' if focus_middle else ''}에 속하는 TC를 아래 4가지 카테고리 순서로 상세하게 작성해주세요.
해당 카테고리에 만들 TC가 없으면 skip합니다.

1. UI/UX 체크: 화면 표시, 초기 상태, 레이아웃 (Positive, Medium~High) — 약 20%
2. 주요 기능: 핵심 정상 흐름 (Positive, High) — bold 처리 — 약 35%
3. 예외 기능: 잘못된 입력, 권한 오류, 잔고 부족, 중복 요청 등 (Negative, High~Medium) — High는 bold — 약 25%
4. 에러 처리 및 비기능: 네트워크 오류, 타임아웃, 빈 응답, 성능 (Edge, Medium~Low) — 약 20%

⚠️ 카테고리 3(예외)과 4(에러)를 반드시 포함하세요. Positive만으로 구성하지 마세요.
⚠️ 이미 다른 TC에서 검증되는 내용은 중복 작성하지 마세요.
⚠️ 중분류당 TC는 5~15개가 적정합니다. 유의미한 TC만 작성하고 양을 채우기 위한 TC는 만들지 마세요.

⚠️ **소분류 이름 유일성** (테스터 가독성 핵심):
- 같은 중분류 안에서 **여러 TC의 소분류 이름이 동일하면 안 됩니다**.
- 테스터는 TC ID보다 소분류 이름으로 케이스를 구분하므로, 소분류는 **그 자체로 무엇을 검증하는지 명확히** 표현해야 합니다.
- 같은 화면의 다른 시나리오라면 소분류에 **차별화 키워드**를 포함하세요.
  예시 ❌: "Splash 화면 진입" × 4개 (UI 표시/버튼 탭/뒤로가기/로딩)
  예시 ✅: "Splash 화면 UI 표시", "Splash Get Started 버튼 탭", "Splash 뒤로가기 차단", "Splash 로딩 시간"
- 소분류 길이는 30자 이내로 짧고 구체적으로.
- 사전 조건은 반드시 번호 개조식으로 작성
- **각 TC의 메타 테이블에 `| 연관 화면 |` 필드를 반드시 포함**하세요 (위 📱 섹션 참고)

⛔ **출력 형식 절대 규칙**:
- 모든 `### ...` 헤더는 반드시 TC ID로 시작해야 합니다 (예: `### SC-XXX-001 — 제목`)
- **TC가 아닌 설명·분석·메모·원칙 적용 결과**를 `### 헤더`로 작성하지 마세요.
- 이 중분류가 원칙 C(deferred/archived) 등에 의해 **TC 생성 제외 대상**이면, 아무것도 작성하지 말고 **완전히 빈 응답**을 반환하세요. 제외 이유 설명도 작성 금지.
- 부가 설명이 꼭 필요하면 TC 블록 안의 `**비고**` 섹션 또는 `---` 구분선 앞의 본문에만 쓰세요. `### 헤더`로는 절대 금지.
"""


def _detect_screen_codes_for_middle(middle: str, text: str) -> list[str]:
    """중분류 이름이 언급된 부근에서 화면 코드(SCR-NNN / SCREEN-NNN / PAGE-NNN)를 추출.
    우선순위:
      1. `| SCR-xxx | {middle} | ...` 인벤토리 표 행 (정확 매칭)
      2. middle 이름 주변 ±300자에 등장하는 SCR/SCREEN/PAGE 코드
    반환: 등장 순서 · 중복 제거한 코드 리스트
    """
    if not middle or not text:
        return []
    codes = []
    seen = set()

    # 패턴 1: 인벤토리 표 행 — `| CODE | NAME | ...`
    row_pat = re.compile(r"\|\s*((?:SCR|SCREEN|PAGE)-[A-Z0-9]+)\s*\|\s*([^|]+?)\s*\|")
    for m in row_pat.finditer(text):
        code, name = m.group(1), m.group(2).strip()
        if middle.lower() in name.lower() or name.lower() in middle.lower():
            if code not in seen:
                codes.append(code)
                seen.add(code)

    if codes:
        return codes

    # 패턴 2: middle 이름 주변에 등장하는 코드 (±300자 윈도우)
    code_pat = re.compile(r"(?:SCR|SCREEN|PAGE)-[A-Z0-9]+")
    for m in re.finditer(re.escape(middle), text, re.IGNORECASE):
        start = max(0, m.start() - 300)
        end = min(len(text), m.end() + 300)
        window = text[start:end]
        for c in code_pat.findall(window):
            if c not in seen:
                codes.append(c)
                seen.add(c)
    return codes


def extract_domain_section(classification: str, domain_code: str) -> str:
    """분류표에서 특정 대분류 섹션 추출.
    괄호 코드(`## 대분류: 이름 (CODE)`) 형식과 코드 없는 형식(`## 대분류: 이름`) 모두 지원.
    domain_code는 extract_domains에서 자동 생성된 코드(이름의 영문자 대문자 약어)."""
    lines = classification.splitlines()
    result = []
    in_section = False
    for line in lines:
        # 새 섹션 시작 감지
        m = re.match(r"^##\s+(.+)", line)
        if m and not line.startswith("###"):
            text = m.group(1).strip()
            text = re.sub(r"^대분류[:\s]*", "", text).strip()
            # 괄호 코드가 있으면 그걸로 매칭
            cm = re.search(r"[\(\（]([A-Z]{2,8})[\)\）]", text)
            if cm:
                matched = (cm.group(1) == domain_code)
            else:
                # 괄호 없음 → 이름에서 코드 자동 생성하여 매칭
                name_only = re.sub(r"\s*[\(\（][^\)\）]*[\)\）]", "", text).strip()
                auto_code = re.sub(r"[^A-Za-z]", "", name_only).upper()[:6] or "FUNC"
                matched = (auto_code == domain_code)
            if matched:
                in_section = True
                result.append(line)
                continue
            elif in_section:
                break
        if in_section:
            result.append(line)
    return "\n".join(result) if result else classification[:1000]


# ── 7단계: TC 검토 ────────────────────────────────────────────────────────────
def _extract_covered_middles(tc_content: str) -> set:
    """TC 초안에서 실제로 다뤄진 (대분류, 중분류) 페어 추출.
    세 가지 패턴 지원:
      1. `**대분류**: X` / `**중분류**: Y` 라인 (bold 라인 스타일)
      2. `## 대분류 / ### 중분류` 섹션 헤더 스타일
      3. `| 대분류 | X |` / `| 중분류 | Y |` 마크다운 테이블 스타일 (현재 TC 형식)
    """
    covered = set()

    # 패턴 1: bold 라인 스타일
    major = None
    middle = None
    for line in tc_content.splitlines():
        m_major = re.search(r"\*\*대분류\*\*\s*[:：]\s*(.+)", line)
        m_middle = re.search(r"\*\*중분류\*\*\s*[:：]\s*(.+)", line)
        if m_major:
            major = m_major.group(1).strip()
        if m_middle:
            middle = m_middle.group(1).strip()
            if major and middle:
                covered.add((major, middle))

    # 패턴 2: 섹션 헤더 스타일
    cur_major = None
    for line in tc_content.splitlines():
        line_s = line.strip()
        if line_s.startswith("## ") and not line_s.startswith("### "):
            txt = re.sub(r"^##\s+", "", line_s).strip()
            txt = re.sub(r"^대분류[:\s]*", "", txt).strip()
            txt = re.sub(r"\s*[\(\（][A-Z]{2,8}[\)\）]", "", txt).strip()
            if txt and not txt.startswith("카테고리"):
                cur_major = txt
        elif line_s.startswith("### ") and cur_major:
            txt = re.sub(r"^###\s+", "", line_s).strip()
            txt = re.sub(r"^중분류[:\s]*", "", txt).strip()
            txt = re.sub(r"\s*[\(\（][A-Z]{2,8}[\)\）]", "", txt).strip()
            if txt and not txt.lower().startswith("tc") and not txt.startswith("카테고리"):
                covered.add((cur_major, txt))

    # 패턴 3: 마크다운 테이블 — TC 블록별로 대분류/중분류 짝맞춤
    # 각 `### TC-ID` 블록 안에서 `| 대분류 | X |`과 `| 중분류 | Y |`를 찾는다
    blocks = re.split(r"\n(?=###\s)", tc_content)
    for block in blocks:
        if not block.strip().startswith("###"):
            continue
        major_match = re.search(r"\|\s*대분류\s*\|\s*([^|]+?)\s*\|", block)
        middle_match = re.search(r"\|\s*중분류\s*\|\s*([^|]+?)\s*\|", block)
        if major_match and middle_match:
            maj = major_match.group(1).strip()
            mid = middle_match.group(1).strip()
            # 괄호 코드 제거
            maj = re.sub(r"\s*[\(\（][A-Z]{2,8}[\)\）]", "", maj).strip()
            mid = re.sub(r"\s*[\(\（][A-Z]{2,8}[\)\）]", "", mid).strip()
            if maj and mid:
                covered.add((maj, mid))

    return covered


def _extract_classification_middles(classification: str) -> list:
    """분류표에서 (대분류, 중분류) 순서 리스트 반환."""
    pairs = []
    cur_major = None
    for line in classification.splitlines():
        line_s = line.strip()
        if line_s.startswith("## ") and not line_s.startswith("### "):
            txt = re.sub(r"^##\s+", "", line_s).strip()
            txt = re.sub(r"^대분류[:\s]*", "", txt).strip()
            txt = re.sub(r"\s*[\(\（][A-Z]{2,8}[\)\）]", "", txt).strip()
            if txt:
                cur_major = txt
        elif line_s.startswith("### ") and cur_major:
            txt = re.sub(r"^###\s+", "", line_s).strip()
            txt = re.sub(r"^중분류[:\s]*", "", txt).strip()
            txt = re.sub(r"\s*[\(\（][A-Z]{2,8}[\)\）]", "", txt).strip()
            if txt:
                pairs.append((cur_major, txt))
    return pairs


def _generate_missing_tc_for_middle(sess: dict, major: str, middle: str,
                                    classification: str, features_text: str,
                                    policy_text: str, project_name: str,
                                    tc_rules: str, fewshot: str,
                                    project_policies: str,
                                    screen_based: bool, screen_map: dict,
                                    project_code: str) -> str:
    """누락된 (대분류, 중분류) 하나에 대해 TC를 생성해 돌려준다."""
    # 대분류 코드 복원 (extract_domains 규칙과 동일)
    cm = re.search(r"[\(\（]([A-Z]{2,8})[\)\）]", major)
    if cm:
        domain_code = cm.group(1)
        domain_name = re.sub(r"\s*[\(\（][A-Z]{2,8}[\)\）]", "", major).strip()
    else:
        domain_name = major
        domain_code = re.sub(r"[^A-Za-z]", "", domain_name).upper()[:6] or "FUNC"

    # Screen-based면 screen code 매핑
    if screen_based:
        effective_code = resolve_screen_code(middle, screen_map)
        screen_character = resolve_screen_character(middle, screen_map)
        screen_navigation = resolve_screen_navigation(middle, screen_map)
        starting_seq = 1
    else:
        effective_code = domain_code
        screen_character = ""
        screen_navigation = ""
        starting_seq = 1

    domain_with_middle = {
        "name": domain_name,
        "code": domain_code,
        "suite_code": effective_code,
        "_focus_middle": middle,
    }
    # 섹션 발췌 (Quick / 정책 반영 최적화 공통)
    effective_features = features_text
    effective_policy = policy_text
    if sess.get("_quick_mode") or sess.get("_section_extract"):
        raw_full = sess.get("_raw_text", "")
        section = extract_section_from_raw(raw_full, middle, major_name=domain_name)
        if section:
            effective_features = section
            if sess.get("_quick_mode"):
                effective_policy = section
            else:
                effective_policy = (policy_text[:4000] + "\n\n---\n\n원문 섹션:\n" + section) if policy_text else section
    system = build_tc_system_prompt(tc_rules, classification, project_policies, fewshot)
    _scr_map = sess.get("_source_scr_map") or {}
    user = build_tc_user_prompt(domain_with_middle, effective_features, effective_policy,
                                project_name, classification,
                                starting_seq=starting_seq,
                                screen_character=screen_character,
                                screen_navigation=screen_navigation,
                                source_scr_map=_scr_map)
    try:
        tc_draft = call_claude(system, user, max_tokens=16000)
    except Exception as e:
        push_log(sess, f"[검토-보강] {domain_code}/{middle} 생성 실패: {e}")
        return ""

    # TC ID 재번호링
    if effective_code:
        tc_draft, _ = renumber_tc_ids(tc_draft, project_code, effective_code, starting_seq)
    return tc_draft


def detect_minor_quality_issues(tc_content: str) -> list:
    """TC 초안의 소분류 품질 이슈를 규칙 기반으로 검출.
    Returns: list of {tc_id, minor, issue_type, detail}
    검출 패턴 (build_excel 정규화로 잡히지 않는 케이스 포함):
      1. 키워드 중복: '계산 정확성' / '확인' 등이 2번 이상 등장
      2. 어색한 끝: 조사·trigger 단어로 끝 (... 시, 탭, 및, 또는, 자동, 별도)
      3. 메타마커 잔존: [정책]/[보안]/[제약] 등
      4. State 라벨 prefix 잔존: Error.X / Danger.X / Loading.X 시작
      5. 다중 라인 (카테고리 헤딩 잔존): '계산 정확성' / 'X' 형태
      6. DOM 식별자 단독: '#tpSlErr' 같은 hash 식별자만
      7. 너무 김 (40자 초과)
    """
    issues = []
    # TC 블록 추출 — '### {ID} — {title}' 헤딩 + 표
    tc_blocks = re.split(r"(?=^###\s+\*?\*?[A-Z]{2}-)", tc_content, flags=re.MULTILINE)
    awkward_endings = {"및", "또는", "에서", "시", "후", "자동", "별도", "경고", "탭", "선택"}
    meta_markers = ("[정책]", "[보안]", "[제약]", "[접근성]", "[세션]", "[타이밍]",
                    "[성능]", "[UX]", "[UI]", "[에러]", "[edge]", "[규약]", "[규칙]",
                    "[통합]", "[미결]", "[그룹]", "[메모]", "[TODO]", "[보류]")
    state_prefix_pattern = re.compile(
        r"^(Error|Loading|Normal|Hover|Empty|Active|Disabled|Success|Failure|Danger|Default)"
        r"\.[A-Za-z][A-Za-z0-9\-_]*"
    )
    dom_only_pattern = re.compile(r"^#[A-Za-z][A-Za-z0-9_\-]*$")

    for block in tc_blocks:
        m_id = re.match(r"###\s+\*?\*?([A-Z]{2}-[A-Z][A-Z0-9\-]*-\d+)", block)
        if not m_id:
            continue
        tc_id = m_id.group(1)
        m_minor = re.search(r"\|\s*소분류\s*\|\s*(.+?)\s*\|", block)
        if not m_minor:
            continue
        minor = m_minor.group(1).strip()
        if not minor:
            continue

        # 1) 키워드 중복
        for kw in ["계산 정확성", "정확성 확인", "확인 확인", "표시 표시", "동작 동작"]:
            count = len(re.findall(re.escape(kw), minor))
            if count >= 2:
                issues.append({
                    "tc_id": tc_id, "minor": minor,
                    "issue_type": "키워드 중복",
                    "detail": f"'{kw}' {count}회 등장",
                })
                break  # 한 항목당 1회만 보고

        # 2) 어색한 끝
        last_token = minor.rstrip(".·-—").split()[-1] if minor.split() else ""
        if last_token in awkward_endings:
            issues.append({
                "tc_id": tc_id, "minor": minor,
                "issue_type": "어색한 끝",
                "detail": f"'{last_token}' 으로 끝남 (절단 가능성)",
            })

        # 3) 메타마커 잔존
        for mk in meta_markers:
            if mk in minor:
                issues.append({
                    "tc_id": tc_id, "minor": minor,
                    "issue_type": "메타마커 잔존",
                    "detail": f"'{mk}' 포함 — 분류 메타이지 라벨 아님",
                })
                break

        # 4) State 라벨 prefix
        if state_prefix_pattern.match(minor):
            issues.append({
                "tc_id": tc_id, "minor": minor,
                "issue_type": "State prefix 잔존",
                "detail": "내부 상태 식별자 — 부연 라벨로 교체 권장",
            })

        # 5) 다중 라인 (카테고리 헤딩 잔존)
        if "\n" in minor:
            lines = [l.strip() for l in minor.split("\n") if l.strip()]
            if len(lines) >= 2:
                first = lines[0]
                if re.match(r"^(계산\s*정확성|에러\s*처리|UI[\s/]*UX|주요\s*기능|예외\s*기능)$", first):
                    issues.append({
                        "tc_id": tc_id, "minor": minor.replace("\n", " / "),
                        "issue_type": "카테고리 헤딩 잔존",
                        "detail": f"첫줄 '{first}' 카테고리 메타 — 1줄 압축 권장",
                    })

        # 6) DOM 식별자 단독
        if dom_only_pattern.match(minor.strip()):
            issues.append({
                "tc_id": tc_id, "minor": minor,
                "issue_type": "DOM 식별자 단독",
                "detail": "사람이 읽기 어려움 — 부연 추가 권장",
            })

        # 7) 너무 김
        max_line_len = max((len(l) for l in minor.split("\n")), default=0)
        if max_line_len > 40:
            issues.append({
                "tc_id": tc_id, "minor": minor[:50] + ("..." if len(minor) > 50 else ""),
                "issue_type": "너무 긴 라벨",
                "detail": f"{max_line_len}자 — 30자 내 식별 라벨 원칙 위반",
            })

    return issues


def step_review(sess: dict, tc_content: str, project_name: str,
                approved_classification: str = "",
                features_text: str = "", policy_text: str = "") -> str:
    """TC 품질 검토 + 누락 중분류 자동 보강.
    approved_classification이 주어지면 분류표 vs TC 초안 비교하여 누락 중분류의 TC를 자동 생성하고 tc_content에 추가한다.
    sess["_augmented_tc"]에 보강된 tc_content를 저장한다."""
    push_log(sess, "[검토] TC 품질 검토 중...")

    # ── 1) 누락 탐지 ─────────────────────────────────────────────
    augmented = tc_content
    missing_pairs = []
    if approved_classification:
        expected = _extract_classification_middles(approved_classification)
        covered = _extract_covered_middles(tc_content)
        # 대소문자·공백 관용 매칭
        def norm(s): return re.sub(r"\s+", "", s).lower()
        covered_norm = {(norm(a), norm(b)) for a, b in covered}
        for (maj, mid) in expected:
            if (norm(maj), norm(mid)) not in covered_norm:
                missing_pairs.append((maj, mid))
        if missing_pairs:
            push_log(sess, f"[검토] 누락된 중분류 {len(missing_pairs)}개 탐지: "
                             + ", ".join(f"{m}/{n}" for m, n in missing_pairs[:5])
                             + (" ..." if len(missing_pairs) > 5 else ""))
        else:
            push_log(sess, "[검토] 누락된 중분류 없음 (분류표 전체 커버)")

    # ── 2) 누락 보강 ─────────────────────────────────────────────
    if missing_pairs and features_text and policy_text:
        tc_rules = load_tc_rules()
        fewshot = load_fewshot_examples()
        project_policies = load_project_policies(project_name)
        project_code = _detect_project_code(project_name)
        user_auto_flag = sess.get("auto_screen_code")
        if user_auto_flag is True:
            screen_based = True
        elif user_auto_flag is False:
            screen_based = False
        else:
            screen_based = _is_screen_based(project_code)
        screen_map = load_screen_code_map(project_name) if screen_based else {}

        added_parts = []
        total_missing = len(missing_pairs)
        # 보강 단계: 82 → 87 구간을 누락 개수로 분할 + ETA 동적 갱신
        for idx, (maj, mid) in enumerate(missing_pairs):
            push_log(sess, f"[검토-보강] [{idx+1}/{total_missing}] {maj}/{mid} TC 생성 중...")
            remaining = total_missing - idx
            pct = 82 + int(5 * (idx / max(total_missing, 1)))
            push_stage(sess, 4, f"TC 보강 {idx+1}/{total_missing} — {mid}", pct,
                       eta_sec=remaining * 30 + 10)
            part = _generate_missing_tc_for_middle(
                sess, maj, mid, approved_classification, features_text, policy_text,
                project_name, tc_rules, fewshot, project_policies,
                screen_based, screen_map, project_code
            )
            if part.strip():
                added_parts.append(part)
            check_stop(sess)
        if added_parts:
            augmented = tc_content + "\n\n---\n\n## 검토 단계 보강 TC\n\n" + "\n\n---\n\n".join(added_parts)
            push_log(sess, f"[검토-보강] 보강 완료 — {len(added_parts)}개 중분류 TC 추가")

    sess["_augmented_tc"] = augmented

    # ── 3) 소분류 품질 규칙 검출 (NEW) ──────────────────────────
    quality_issues = detect_minor_quality_issues(augmented)
    quality_section = ""
    if quality_issues:
        # 이슈 타입별 그룹핑 + 로그 요약
        from collections import Counter
        type_counts = Counter(i["issue_type"] for i in quality_issues)
        summary = ", ".join(f"{t} {c}건" for t, c in type_counts.most_common())
        push_log(sess, f"[검토-품질] 소분류 품질 이슈 {len(quality_issues)}건 탐지: {summary}")
        # review_report 에 추가할 섹션 작성
        quality_lines = ["", "## 소분류 품질 이슈 (규칙 기반 검출)", ""]
        quality_lines.append(f"총 {len(quality_issues)}건 — {summary}")
        quality_lines.append("")
        quality_lines.append("| TC ID | 이슈 유형 | 소분류 | 상세 |")
        quality_lines.append("|-------|----------|--------|------|")
        for issue in quality_issues[:50]:  # 상위 50건만
            mid = (issue["minor"][:35] + "...") if len(issue["minor"]) > 35 else issue["minor"]
            quality_lines.append(
                f"| {issue['tc_id']} | {issue['issue_type']} | {mid} | {issue['detail']} |"
            )
        if len(quality_issues) > 50:
            quality_lines.append(f"\n_그 외 {len(quality_issues) - 50}건은 생략됨_")
        quality_section = "\n".join(quality_lines)
    else:
        push_log(sess, "[검토-품질] 소분류 품질 이슈 없음 ✓")

    # ── 4) AI 리뷰 리포트 작성 ──────────────────────────────────
    system = """당신은 시니어 QA 리뷰어입니다. TC 초안을 검토하고 개선 리포트를 작성합니다."""
    missing_section = ""
    if missing_pairs:
        missing_section = "\n누락 중분류 (자동 보강됨):\n" + "\n".join(f"- {m}/{n}" for m, n in missing_pairs[:30])
    user = f"""프로젝트: {project_name}

다음 TC 초안을 검토하고 review_report.md를 작성해주세요:

검토 항목:
1. TC ID 일관성 및 중복 여부
2. 사전 조건 작성 규칙 준수 여부
3. Positive/Negative/Edge 비율
4. 최소 TC 세트 선별 적절성
5. 예상 결과 측정 가능성
6. 누락된 케이스
{missing_section}

TC 초안 (처음 5000자):
---
{tc_content[:5000]}
---

리포트 형식:
# TC 검토 보고서
## 전체 요약
## 발견된 이슈
## 개선 권장 사항
## 최소 TC 세트 검증
## 누락 중분류 보강 결과
"""
    try:
        result = call_claude(system, user, max_tokens=3000)
    except Exception as e:
        result = f"# TC 검토 보고서\n\n검토 실행 중 오류: {e}\n"
    # 규칙 기반 품질 검출 결과를 리포트 끝에 첨부
    if quality_section:
        result = result.rstrip() + "\n\n---\n" + quality_section + "\n"
    review_path = sess["workspace"] / "07_review_report.md"
    review_path.write_text(result, encoding="utf-8")
    push_log(sess, "[검토] 검토 완료")
    return result


# ── 8단계: tc_final.md 생성 및 Excel 빌드 ─────────────────────────────────────
def step_build_excel(sess: dict, tc_content: str, project_name: str,
                     total_tc: int, min_tc: int) -> Path:
    # 후처리: 같은 (대,중,소) 그룹에 TC가 여러 개면 소분류에 차별화 키워드 자동 부여
    # MD ↔ Excel 일관성을 위해 tc_content 자체를 갱신 (이후 tc_files에도 변형 적용된 형태로 저장됨)
    sess.pop("_tc_content_disambiguated", None)  # 이전 호출 잔여 제거
    try:
        new_content = disambiguate_duplicate_minors(tc_content)
        if new_content != tc_content:
            tc_content = new_content
            # 세션에도 보관 — 호출부에서 tc_files에 저장 시 이 변형본 사용
            sess["_tc_content_disambiguated"] = tc_content
            push_log(sess, "[빌드] 소분류 중복 자동 구분 적용 — 같은 화면의 다른 시나리오에 차별화 키워드 부여")
    except Exception as e:
        push_log(sess, f"[빌드] 소분류 후처리 스킵: {e}")

    push_log(sess, "[빌드] tc_final.md 생성 중...")

    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    header = f"""# TC 최종본 — {project_name}

## 메타 정보
- 생성일: {now}
- Phase: P_WebApp
- 총 TC: {total_tc}개 / 최소 TC: {min_tc}개

---

"""
    tc_final_content = header + tc_content
    tc_final_path = sess["workspace"] / "tc_final.md"
    tc_final_path.write_text(tc_final_content, encoding="utf-8")
    push_log(sess, f"[빌드] tc_final.md 저장 완료 ({len(tc_final_content):,}자)")

    # 출력 디렉토리
    out_dir = OUTPUTS_DIR

    # ── 1순위: build_excel 모듈을 직접 import해서 run_build() 호출 ──
    # subprocess 우회 시 장점:
    #   - Windows cp949 인코딩 이슈 완전 제거 (stdout/stderr 디코드 경로 없음)
    #   - 가상환경 activation 의존성 없음 (Flask가 쓰는 같은 인터프리터)
    #   - 속도 ~1초 절감 (프로세스 생성/teardown 불필요)
    # 실패 시에만 subprocess → fallback 으로 강등
    if BUILD_EXCEL.exists():
        push_log(sess, f"[빌드] build_excel 모듈 호출 중... (대분류별 시트 생성 · 60초 내외)")
        try:
            build_excel_dir = str(BUILD_EXCEL.parent)
            if build_excel_dir not in sys.path:
                sys.path.insert(0, build_excel_dir)
            # 매 호출 시 fresh import (코드 수정 반영)
            import importlib
            import build_excel as _be
            importlib.reload(_be)
            # 상태/수정 사유 컬럼 + 🔄 변경 이력 시트는 「기존 TC 수정」 플로우(update-tc)에서만 포함.
            # 신규 TC 생성 시에는 깨끗한 Excel을 위해 생략.
            include_change = bool(sess.get("_include_change_columns"))
            # 사용자가 Human Gate에서 선택한 Excel 시트 옵션 (없으면 None=Full Set)
            excel_sheets = sess.get("_excel_sheets")
            result = _be.run_build("P_WebApp", str(tc_final_path), str(out_dir),
                                   verbose=False, include_change_columns=include_change,
                                   sheets=excel_sheets)
            if result and result.get("ok"):
                push_log(sess, f"[빌드] 모듈 호출 성공 — 총 {result['total_tc']} TC / Smoke {result['smoke_tc']} / 대분류 {result['major_count']}시트")
                sess["smoke_tc"] = result["smoke_tc"]
                out_path = result["out_path"]
                # v0.9.22: 통합 빌드인 경우 파일명에 _merged suffix 추가
                fname_suffix = sess.get("_excel_filename_suffix", "")
                if fname_suffix and out_path and out_path.exists():
                    new_name = out_path.stem + fname_suffix + out_path.suffix
                    new_path = out_path.parent / new_name
                    try:
                        out_path.rename(new_path)
                        out_path = new_path
                        push_log(sess, f"[빌드] 통합 결과 파일명 → {new_path.name}")
                    except Exception as _re:
                        push_log(sess, f"[빌드] 파일명 rename 실패 (원본 유지): {_re}")
                return out_path
        except Exception as e:
            push_log(sess, f"[빌드] 모듈 호출 실패 ({type(e).__name__}: {e}) — subprocess 폴백")

        # ── 2순위: subprocess (모듈 import 실패 시) ──
        try:
            child_env = os.environ.copy()
            child_env["PYTHONIOENCODING"] = "utf-8"
            child_env["PYTHONUTF8"] = "1"
            proc = subprocess.run(
                [sys.executable, str(BUILD_EXCEL),
                 "--phase", "P_WebApp",
                 "--tc",    str(tc_final_path),
                 "--output", str(out_dir)],
                capture_output=True, timeout=120, env=child_env,
            )
            stdout_str = (proc.stdout or b"").decode("utf-8", errors="replace")
            stderr_str = (proc.stderr or b"").decode("utf-8", errors="replace")
            if proc.returncode == 0:
                push_log(sess, "[빌드] subprocess 성공")
                m = re.search(r"Smoke Test\s*:\s*(\d+)개", stdout_str)
                if m:
                    sess["smoke_tc"] = int(m.group(1))
                excel_files = sorted(out_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
                if excel_files:
                    return excel_files[0]
                push_log(sess, "[빌드] ⚠️ subprocess 성공했지만 xlsx 파일 없음 — fallback")
            else:
                err_msg = (stderr_str or stdout_str or "").strip()[:500]
                push_log(sess, f"[빌드] subprocess 오류 (returncode={proc.returncode}): {err_msg}")
        except subprocess.TimeoutExpired:
            push_log(sess, "[빌드] subprocess 타임아웃, fallback으로 직접 생성")
        except FileNotFoundError as e:
            push_log(sess, f"[빌드] ❌ Python 실행 파일 못 찾음: {e} — 가상환경(.venv) activation 확인 필요")
        except Exception as e:
            push_log(sess, f"[빌드] subprocess 예외 ({type(e).__name__}): {e}")

    # ── 3순위(최후): 내부 간이 빌더 (구형 포맷, 시트 분할 없음 — 아무것도 못 만드는 것보다 낫다) ──
    push_log(sess, "[빌드] ⚠️ 최후 fallback: 내부 간이 빌더 — 대분류별 시트 분할/Traceability 없음")
    return build_excel_fallback(tc_final_content, out_dir, project_name, total_tc, min_tc)


def build_excel_fallback(tc_content: str, out_dir: Path, project_name: str,
                          total_tc: int, min_tc: int) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl 미설치: pip install openpyxl")

    NAVY, BLUE, TEAL = "1E2761", "3557A0", "028090"
    LIGHT, WHITE = "EEF1F8", "FFFFFF"

    def fill(c): return PatternFill("solid", fgColor=c)
    def hfont(bold=True, size=10, color=WHITE):
        return Font(name="Calibri", bold=bold, size=size, color=color)
    def bdr():
        s = Side(style="thin", color="D0D7DE")
        return Border(left=s, right=s, top=s, bottom=s)
    def center(w=False): return Alignment(horizontal="center", vertical="center", wrap_text=w)
    def left_align(w=True): return Alignment(horizontal="left", vertical="center", wrap_text=w)

    wb = Workbook()
    wb.remove(wb.active)

    # 표지 시트
    cov = wb.create_sheet("표지")
    cov.sheet_view.showGridLines = False
    cov.merge_cells("A1:J3")
    c = cov["A1"]
    c.value = f"TC Checklist — {project_name}"
    c.font = hfont(size=18)
    c.fill = fill(NAVY)
    c.alignment = center()
    cov.row_dimensions[1].height = 60

    meta = [
        ("작성일", datetime.now().strftime("%Y-%m-%d")),
        ("프로젝트", project_name),
        ("Phase", "P_WebApp"),
        ("총 TC 수", str(total_tc)),
        ("최소 TC 수", str(min_tc)),
        ("생성 도구", "TC 자동화 v2 (Claude AI)"),
    ]
    for ri, (lbl, val) in enumerate(meta, 5):
        c = cov.cell(ri, 1, lbl)
        c.font = hfont(); c.fill = fill(BLUE); c.alignment = center(); c.border = bdr()
        cov.merge_cells(f"A{ri}:C{ri}")
        c = cov.cell(ri, 4, val)
        c.font = hfont(bold=False, color=NAVY)
        c.fill = fill(LIGHT); c.alignment = left_align(w=False); c.border = bdr()
        cov.merge_cells(f"D{ri}:J{ri}")
        cov.row_dimensions[ri].height = 24
    for i in range(1, 11):
        cov.column_dimensions[get_column_letter(i)].width = 12
    cov.column_dimensions["A"].width = 16
    cov.column_dimensions["D"].width = 40

    # TC 전체목록 시트
    ws = wb.create_sheet("TC 전체목록")
    ws.freeze_panes = "A2"

    HEADERS = ["Smoke", "TC ID", "우선순위", "거래소", "대분류", "중분류", "소분류", "사전 조건", "스텝", "기대 결과"]
    COL_W   = [8,           18,     10,       10,      14,     14,     16,      50,       50,     50]

    def _priority_kr(p):
        return {"High": "높음", "Medium": "보통", "Low": "낮음"}.get(p, p or "보통")

    for ci, (h, w) in enumerate(zip(HEADERS, COL_W), 1):
        c = ws.cell(1, ci, h)
        c.font = hfont(); c.fill = fill(NAVY); c.alignment = center(); c.border = bdr()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    # TC 파싱 및 행 삽입
    tcs = parse_tc_markdown(tc_content)

    # Smoke 마킹: 중분류별 대표 Positive 1개 + High Negative 1개 + 대분류별 최소 1개
    from collections import defaultdict
    mid_groups = defaultdict(list)
    for tc in tcs:
        mid_groups[(tc.get("major",""), tc.get("middle",""))].append(tc)
    for key, group in mid_groups.items():
        pos_done = False
        for pri_t in [("high","높음"),("medium","보통"),("low","낮음")]:
            if pos_done: break
            for tc in group:
                if tc.get("priority","").lower() in pri_t and tc.get("type","").lower() == "positive":
                    tc["_smoke"] = True; pos_done = True; break
        neg_done = False
        for tc in group:
            if neg_done: break
            if tc.get("priority","").lower() in ("high","높음") and tc.get("type","").lower() == "negative":
                tc["_smoke"] = True; neg_done = True
    domain_has = {}
    for tc in tcs:
        if tc.get("_smoke"): domain_has[tc.get("major","")] = True
    for tc in tcs:
        d = tc.get("major","")
        if d not in domain_has: tc["_smoke"] = True; domain_has[d] = True

    FILL_MIN  = PatternFill("solid", fgColor="FFF9C4")
    FILL_NORM = PatternFill("solid", fgColor="FFFFFF")

    for ri, tc in enumerate(tcs, 2):
        is_min = tc.get("is_min", False)
        row_fill = FILL_MIN if is_min else FILL_NORM
        row_data = [
            "Y" if tc.get("_smoke") else "",
            tc.get("id", ""),
            _priority_kr(tc.get("priority", "")),
            "",  # 거래소
            tc.get("major", ""),
            tc.get("middle", ""),
            tc.get("minor", ""),
            tc.get("precondition", ""),
            tc.get("steps", ""),
            tc.get("expected", ""),
        ]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(ri, ci, val)
            c.fill = row_fill
            c.alignment = left_align() if ci >= 9 else center()
            c.border = bdr()
            c.font = Font(name="Calibri", size=9,
                          bold=(ci == 2 and is_min), color="1E2761" if is_min else "222222")
        ws.row_dimensions[ri].height = max(30, min(120, len(str(row_data[8])) // 3 + 20))

    # Smoke Test 시트 — Smoke 마킹된 TC만 모아 동일 구조로 렌더
    smoke_tcs = [t for t in tcs if t.get("_smoke")]
    ws_sm = wb.create_sheet("Smoke Test")
    ws_sm.freeze_panes = "A2"
    for ci, (h, w) in enumerate(zip(HEADERS, COL_W), 1):
        c = ws_sm.cell(1, ci, h)
        c.font = hfont(); c.fill = fill(NAVY); c.alignment = center(); c.border = bdr()
        ws_sm.column_dimensions[get_column_letter(ci)].width = w
    ws_sm.row_dimensions[1].height = 22
    for ri, tc in enumerate(smoke_tcs, 2):
        row_data = [
            "Y",
            tc.get("id", ""),
            _priority_kr(tc.get("priority", "")),
            "",
            tc.get("major", ""),
            tc.get("middle", ""),
            tc.get("minor", ""),
            tc.get("precondition", ""),
            tc.get("steps", ""),
            tc.get("expected", ""),
        ]
        for ci, val in enumerate(row_data, 1):
            c = ws_sm.cell(ri, ci, val)
            c.fill = PatternFill("solid", fgColor="E8F5E9")
            c.alignment = left_align() if ci >= 9 else center()
            c.border = bdr()
            c.font = Font(name="Calibri", size=9)
        ws_sm.row_dimensions[ri].height = max(30, min(120, len(str(row_data[8])) // 3 + 20))

    # 통계 시트
    stat = wb.create_sheet("통계")
    stat.sheet_view.showGridLines = False
    stat.merge_cells("A1:E1")
    c = stat["A1"]
    c.value = "TC 통계 요약"
    c.font = hfont(size=14); c.fill = fill(NAVY); c.alignment = center()
    stat.row_dimensions[1].height = 36

    type_counts = {}
    priority_counts = {}
    for tc in tcs:
        t = tc.get("type", "Unknown")
        p = tc.get("priority", "Unknown")
        type_counts[t] = type_counts.get(t, 0) + 1
        priority_counts[p] = priority_counts.get(p, 0) + 1

    stat_data = [
        ("총 TC 수", total_tc),
        ("최소 TC 수", min_tc),
        ("Positive", type_counts.get("Positive", 0)),
        ("Negative", type_counts.get("Negative", 0)),
        ("Edge", type_counts.get("Edge", 0)),
        ("High 우선순위", priority_counts.get("High", 0)),
        ("Medium 우선순위", priority_counts.get("Medium", 0)),
        ("Low 우선순위", priority_counts.get("Low", 0)),
    ]
    for ri, (lbl, val) in enumerate(stat_data, 3):
        c = stat.cell(ri, 1, lbl)
        c.font = hfont(); c.fill = fill(BLUE); c.alignment = center(); c.border = bdr()
        stat.merge_cells(f"A{ri}:C{ri}")
        c = stat.cell(ri, 4, val)
        c.font = hfont(bold=False, color=NAVY)
        c.fill = fill(LIGHT); c.alignment = center(); c.border = bdr()
        stat.column_dimensions["A"].width = 20
        stat.column_dimensions["D"].width = 15

    # 파일 저장
    today = datetime.now().strftime("%Y%m%d")
    safe_name = re.sub(r"[^\w\-_]", "_", project_name)[:20]
    version = 1
    while True:
        fname = f"SPCY_TC_P_WebApp_{safe_name}_{today}_v{version}.xlsx"
        fpath = out_dir / fname
        if not fpath.exists():
            break
        version += 1

    wb.save(str(fpath))
    return fpath


def parse_tc_excel(path: Path) -> list[dict]:
    """우리가 생성한 Excel 포맷의 TC를 파싱하여 딕셔너리 리스트로 변환.
    대분류별 시트(📑 prefix)를 순회하며 TC ID가 있는 행을 수집.
    헤더 컬럼명으로 매핑 (컬럼 순서 변경에 강건).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl 패키지가 필요합니다")

    wb = load_workbook(path, data_only=True)
    tcs = []
    seen_ids = set()

    # 한글 → 내부 키 매핑 (구·신 컬럼명 모두 인식 — 하위 호환)
    KEY_MAP = {
        "TC ID": "id",
        "Smoke": "smoke",
        # 신규 컬럼명 (v0.9.7b)
        "중요도": "priority",
        "대상 거래소": "exchange",
        "사전조건": "precondition",
        "테스트 스텝": "steps",
        "기대결과": "expected",
        # 구 컬럼명 (하위 호환)
        "우선순위": "priority",
        "관련 거래소": "exchange",
        "거래소": "exchange",
        "사전 조건": "precondition",
        "스텝": "steps",
        "기대 결과": "expected",
        # 공통
        "대분류": "major",
        "중분류": "middle",
        "소분류": "minor",
        "화면 코드": "screen_code",
        "연관 화면": "screen",
        "상태": "status",
        "수정 사유": "change_reason",
    }

    for sheet_name in wb.sheetnames:
        # 공통 시트는 제외 (📑로 시작하는 대분류 시트만 순회)
        if not sheet_name.startswith("📑"):
            continue
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue
        # 헤더 읽기 (1행)
        header_row = [c.value for c in ws[1]]
        col_idx = {}
        for i, h in enumerate(header_row):
            if not h:
                continue
            h_str = str(h).strip()
            key = KEY_MAP.get(h_str)
            if key:
                col_idx[key] = i

        if "id" not in col_idx:
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_id = row[col_idx["id"]] if col_idx["id"] < len(row) else None
            if not raw_id or not re.match(r"^[A-Z]{2,}-[A-Z0-9]+-\d+", str(raw_id).strip()):
                continue
            tid = str(raw_id).strip()
            if tid in seen_ids:
                continue
            seen_ids.add(tid)

            tc = {"id": tid, "title": ""}
            for key, idx in col_idx.items():
                if idx < len(row) and row[idx] is not None:
                    tc[key] = str(row[idx]).strip()

            # smoke은 "Y" 여부를 bool로
            tc["is_min"] = (tc.get("smoke", "").upper() == "Y")
            # 우선순위 한글 → 영문 역변환 (Medium 기본)
            pri_map = {"높음": "High", "보통": "Medium", "낮음": "Low"}
            if tc.get("priority") in pri_map:
                tc["priority"] = pri_map[tc["priority"]]
            tc.setdefault("major", "")
            tc.setdefault("middle", "")
            tc.setdefault("minor", "")
            tc.setdefault("priority", "Medium")
            tc.setdefault("screen", tc.get("screen_code", ""))
            tcs.append(tc)
    return tcs


def _extract_spec_sections(md_text: str) -> dict[str, dict]:
    """기획서 MD를 화면(SCR/SCREEN/PAGE) 단위로 분할하여
    {code: {"name": ..., "body": ..., "raw_header": ...}} 딕셔너리 반환.

    매칭 패턴 (우선순위):
      1. `### SCR-001: Splash` (또는 `SCREEN-001`, `PAGE-001`)
      2. `### SCR-001` (이름 없음)
      3. 인벤토리 표 행 `| SCR-001 | Splash | 설명 |` — 이것은 헤더가 아니지만
         화면 코드가 없는 헤더 아래에서 소속 화면을 식별할 때 보조로 사용

    헤더가 없으면 빈 dict 반환 (상위 레벨에서 fallback 처리).
    """
    sections = {}
    if not md_text:
        return sections

    lines = md_text.splitlines()
    current_code = None
    current_name = ""
    current_header = ""
    buffer = []

    # `### CODE: Name` 또는 `### CODE` 형식
    header_pat = re.compile(
        r"^(#{2,4})\s+((?:SCR|SCREEN|PAGE)-[A-Z0-9]+[A-Z]?)\s*[:\-]?\s*(.*?)\s*$"
    )

    def flush():
        if current_code:
            sections[current_code] = {
                "name": current_name,
                "body": "\n".join(buffer).strip(),
                "raw_header": current_header,
            }

    for line in lines:
        m = header_pat.match(line)
        if m:
            flush()
            current_code = m.group(2).strip()
            current_name = m.group(3).strip() or ""
            current_header = line
            buffer = []
        elif current_code is not None:
            buffer.append(line)
    flush()
    return sections


def _normalize_body(body: str) -> str:
    """비교용 정규화 — 공백/빈 줄 정리, 가독성 변경(줄바꿈 등)에 둔감하게."""
    if not body:
        return ""
    # 연속 공백 하나로, 줄 끝 공백 제거, 빈 줄 제거
    lines = [re.sub(r"\s+", " ", ln).strip() for ln in body.splitlines()]
    return "\n".join(ln for ln in lines if ln)


def diff_specs(old_md: str, new_md: str) -> dict:
    """두 기획서 MD를 화면 코드 단위로 비교하여 변경사항 리포트 생성.

    반환 구조:
    {
        "added":   [{"code": "SCR-020", "name": "...", "body": "..."}],
        "removed": [{"code": "SCR-008", "name": "...", "body": "..."}],
        "modified":[{"code": "SCR-001", "name": "...",
                     "old_body": "...", "new_body": "...",
                     "diff_summary": "(옵션)"}],
        "unchanged_codes": ["SCR-002", ...],
        "has_spec_codes": True/False,  # 화면 코드 체계 사용 여부
    }
    """
    old_sections = _extract_spec_sections(old_md)
    new_sections = _extract_spec_sections(new_md)
    has_codes = bool(old_sections or new_sections)

    result = {
        "added": [],
        "removed": [],
        "modified": [],
        "unchanged_codes": [],
        "has_spec_codes": has_codes,
    }

    if not has_codes:
        # 화면 코드 체계 없음 — fallback은 상위에서 처리 (예: 전체 raw diff)
        return result

    old_codes = set(old_sections.keys())
    new_codes = set(new_sections.keys())

    for code in sorted(new_codes - old_codes):
        s = new_sections[code]
        result["added"].append({
            "code": code, "name": s["name"], "body": s["body"][:2000]
        })

    for code in sorted(old_codes - new_codes):
        s = old_sections[code]
        result["removed"].append({
            "code": code, "name": s["name"], "body": s["body"][:2000]
        })

    for code in sorted(old_codes & new_codes):
        old_raw = old_sections[code]["body"]
        new_raw = new_sections[code]["body"]
        old_body = _normalize_body(old_raw)
        new_body = _normalize_body(new_raw)
        if old_body != new_body:
            # 라인 단위 diff — 변경된 라인만 시각적으로 표시 가능하게 함
            line_diff = _compute_line_diff(old_raw, new_raw)
            result["modified"].append({
                "code": code,
                "name": new_sections[code]["name"] or old_sections[code]["name"],
                "old_body": old_raw[:4000],
                "new_body": new_raw[:4000],
                "line_diff": line_diff,
            })
        else:
            result["unchanged_codes"].append(code)

    return result


def _compute_line_diff(old_text: str, new_text: str,
                       max_lines: int = 400) -> list[dict]:
    """difflib 기반 라인 단위 diff 계산.
    반환: [{"tag": "equal"|"insert"|"delete"|"replace",
            "old_lines": [...], "new_lines": [...]}] 형태의 블록 리스트.
    너무 크면 max_lines에서 잘라낸다.
    """
    import difflib
    old_lines = old_text.splitlines()
    new_lines = new_text.splitlines()
    # 공백만 다른 경우 노이즈 줄이기 — 동일하게 취급할지 여부는 difflib에 맡김
    matcher = difflib.SequenceMatcher(a=old_lines, b=new_lines, autojunk=False)
    blocks = []
    total_lines = 0
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        olines = old_lines[i1:i2]
        nlines = new_lines[j1:j2]
        # equal 블록이 너무 길면 앞뒤 2줄만 남겨 컨텍스트 제공
        if tag == "equal" and (i2 - i1) > 4:
            olines = old_lines[i1:i1+2] + ["... (중략)"] + old_lines[i2-2:i2]
            nlines = new_lines[j1:j1+2] + ["... (중략)"] + new_lines[j2-2:j2]
        blocks.append({"tag": tag, "old_lines": olines, "new_lines": nlines})
        total_lines += max(len(olines), len(nlines))
        if total_lines >= max_lines:
            blocks.append({"tag": "truncated", "old_lines": [], "new_lines": []})
            break
    return blocks


def group_tcs_by_scr(tcs: list[dict]) -> dict[str, list[dict]]:
    """TC 리스트를 화면 코드(SCR-xxx) 기준으로 그룹핑.
    화면 코드가 없거나 여러 개인 경우:
      - 여러 개면 각 코드에 중복 등록
      - 없으면 "(화면 코드 없음)" 그룹으로
    """
    groups = defaultdict(list) if False else {}  # noqa
    from collections import defaultdict as _dd
    groups = _dd(list)
    for tc in tcs:
        screen_field = tc.get("screen_code") or tc.get("screen") or ""
        codes = re.findall(r"(?:SCR|SCREEN|PAGE)-[A-Za-z0-9]+", screen_field)
        if not codes:
            groups["(화면 코드 없음)"].append(tc)
            continue
        for c in dict.fromkeys(codes):  # 중복 제거 · 순서 유지
            groups[c].append(tc)
    return dict(groups)


def next_tc_id(existing_tcs: list[dict], project_code: str, suite_code: str) -> int:
    """주어진 프로젝트/스위트 코드의 기존 TC 번호 중 최대값 + 1 반환."""
    max_seq = 0
    pat = re.compile(rf"^{re.escape(project_code)}-{re.escape(suite_code)}-(\d+)$")
    for tc in existing_tcs:
        m = pat.match(tc.get("id", ""))
        if m:
            seq = int(m.group(1))
            if seq > max_seq:
                max_seq = seq
    return max_seq + 1


def build_diff_report(old_md: str, new_md: str,
                       existing_tcs: list[dict]) -> dict:
    """기획서 diff + 기존 TC 매핑을 결합한 리포트 생성.

    반환 구조 (프론트엔드에 바로 사용 가능):
    {
      "has_spec_codes": bool,
      "added":   [{"code", "name", "body", "estimated_tc_count"}],
      "modified":[{"code", "name", "old_body", "new_body",
                   "affected_tc_ids": [...], "affected_count"}],
      "removed": [{"code", "name", "affected_tc_ids": [...]}],
      "summary": {"added_n", "modified_n", "removed_n", "unchanged_n"}
    }
    """
    base = diff_specs(old_md, new_md)
    tc_groups = group_tcs_by_scr(existing_tcs)

    report = {
        "has_spec_codes": base["has_spec_codes"],
        "added": [],
        "modified": [],
        "removed": [],
        "summary": {},
    }

    for a in base["added"]:
        report["added"].append({
            "code": a["code"],
            "name": a["name"],
            "body": a["body"],
            "estimated_tc_count": 8,  # 경험적 기본값 (5~15 중 중앙)
        })

    for m in base["modified"]:
        affected = tc_groups.get(m["code"], [])
        report["modified"].append({
            "code": m["code"],
            "name": m["name"],
            "old_body": m["old_body"],
            "new_body": m["new_body"],
            "line_diff": m.get("line_diff", []),
            "affected_tc_ids": [t.get("id", "") for t in affected if t.get("id")],
            "affected_count": len(affected),
        })

    for r in base["removed"]:
        affected = tc_groups.get(r["code"], [])
        report["removed"].append({
            "code": r["code"],
            "name": r["name"],
            "affected_tc_ids": [t.get("id", "") for t in affected if t.get("id")],
            "affected_count": len(affected),
        })

    report["summary"] = {
        "added_n":     len(report["added"]),
        "modified_n":  len(report["modified"]),
        "removed_n":   len(report["removed"]),
        "unchanged_n": len(base["unchanged_codes"]),
    }
    return report


def load_existing_tcs(file_path: str | Path) -> list[dict]:
    """기존 TC 파일(MD 또는 Excel)을 자동 판별하여 파싱."""
    p = Path(file_path)
    if not p.exists():
        raise RuntimeError(f"파일 없음: {file_path}")
    suf = p.suffix.lower()
    if suf in (".md", ".markdown", ".txt"):
        content = p.read_text(encoding="utf-8")
        return parse_tc_markdown(content)
    if suf in (".xlsx", ".xls"):
        return parse_tc_excel(p)
    raise RuntimeError(f"지원하지 않는 파일 형식: {suf} (.md, .xlsx만 지원)")


def parse_tc_markdown(content: str) -> list[dict]:
    """TC Markdown을 파싱하여 딕셔너리 리스트로 변환"""
    tcs = []
    # TC 블록 분할
    blocks = re.split(r"\n(?=###\s)", content)

    for block in blocks:
        if not block.strip().startswith("###"):
            continue
        # 카테고리 구분 헤더 스킵
        first_line = block.strip().split("\n")[0]
        if re.match(r"###\s*카테고리\s*\d", first_line):
            continue
        tc = parse_single_tc(block)
        if tc.get("id"):
            tcs.append(tc)
    return tcs


def parse_single_tc(block: str) -> dict:
    tc = {}
    lines = block.strip().splitlines()
    if not lines:
        return tc

    # 제목 줄 파싱
    heading = lines[0]
    is_min = bool(re.search(r"###\s+\*\*", heading))
    tc["is_min"] = is_min

    # TC ID 추출
    id_match = re.search(r"(SC-[A-Z]+-[A-Z]+-\d+|[A-Z]+-[A-Z]+-\d+)", heading)
    tc["id"] = id_match.group(1) if id_match else ""

    # 제목 추출
    title_match = re.search(r"—\s+(.+?)(?:\s*\*\*)?$", heading)
    if title_match:
        title = title_match.group(1).strip()
        title = re.sub(r"\*+", "", title).strip()
        tc["title"] = title
    else:
        tc["title"] = re.sub(r"^###\s+\*?\*?[A-Z\-0-9]+\*?\*?\s*—?\s*", "", heading).strip()

    # 테이블 파싱
    for line in lines:
        line = line.strip()
        if "|" in line:
            parts = [p.strip() for p in line.split("|") if p.strip()]
            if len(parts) >= 2:
                key, val = parts[0], parts[1]
                if "대분류" in key: tc["major"] = val
                elif "중분류" in key: tc["middle"] = val
                elif "소분류" in key: tc["minor"] = val
                elif "분류" in key and "대" not in key and "중" not in key and "소" not in key:
                    tc["type"] = val
                elif "우선순위" in key: tc["priority"] = val
                elif "플랫폼" in key: tc["platform"] = val
                elif "연관 화면" in key or "화면" in key: tc["screen"] = val

    # 섹션 파싱
    tc["precondition"] = extract_section(block, "사전 조건")
    tc["steps"]        = extract_section(block, "테스트 단계")
    tc["expected"]     = extract_section(block, "예상 결과")
    tc["note"]         = extract_section(block, "비고")

    # 기본값
    tc.setdefault("major", "")
    tc.setdefault("middle", "")
    tc.setdefault("minor", "")
    tc.setdefault("type", "Positive")
    tc.setdefault("priority", "Medium")
    tc.setdefault("platform", "")
    tc.setdefault("screen", "")

    return tc


def extract_section(block: str, section_name: str) -> str:
    pattern = re.compile(
        rf"\*\*{re.escape(section_name)}\*\*\n(.*?)(?=\n\*\*[가-힣\s]+\*\*|\Z)",
        re.DOTALL
    )
    m = pattern.search(block)
    if not m:
        return ""
    text = m.group(1).strip()
    # 사전 조건: 불릿(-) 형식이면 번호 개조식으로 변환
    if section_name == "사전 조건":
        lines = text.splitlines()
        # 불릿 라인이 하나라도 있으면 전체를 번호로 변환
        has_bullet = any(re.match(r"^[-*]\s+", l) for l in lines)
        if has_bullet:
            numbered, n = [], 1
            for l in lines:
                l = l.strip()
                if not l:
                    continue
                # 이미 번호가 있으면 그대로, 불릿이면 번호 부여
                if re.match(r"^\d+\.", l):
                    numbered.append(l)
                else:
                    body = re.sub(r"^[-*]\s+", "", l)
                    numbered.append(f"{n}. {body}")
                    n += 1
            text = "\n".join(numbered)
    return text


# ── 메인 파이프라인 ────────────────────────────────────────────────────────────
def run_pipeline(sess: dict, sources: list, project_name: str):
    focus_area = sess.get("focus_area", "")
    # 이어서 작업인지 확인
    resumed = sess.get("_resumed_state")
    # 소스 정보 보존 (이어서 작업 시에는 이전 상태에서 가져옴)
    if sources:
        _sources_info = [{"type": s.get("type",""), "content": s.get("content",""), "selected_files": s.get("selected_files")} for s in sources]
    elif resumed:
        _sources_info = resumed.get("data", {}).get("sources_info", [])
    else:
        _sources_info = []
    try:
        # ── 파싱 ──
        if resumed and resumed.get("stage") in ("policy", "features", "classifying", "gate_waiting", "tc_writing"):
            raw_text = resumed["data"].get("raw_text", "")
            push_log(sess, f"[이어서] 파싱 결과 복원됨 ({len(raw_text):,}자)")
            # v0.9.19: 입력 소스 SCR 매핑 복원 (체크포인트에 있으면 그것을, 없으면 raw_text 에서 자동 재추출)
            saved_scr_map = resumed["data"].get("source_scr_map") or {}
            if saved_scr_map:
                sess["_source_scr_map"] = dict(saved_scr_map)
                push_log(sess, f"[이어서] 화면 식별자 매핑 복원됨 ({len(saved_scr_map)}개)")
            else:
                # 폴백 — raw_text 에서 자동 재추출 (체크포인트가 v0.9.18 이하 버전이면 매핑 없음)
                fallback_map = {}
                # 입력 소스 정보가 있으면 파일별로 SCR 추출
                for src_info in (_sources_info or []):
                    if src_info.get("type") == "md":
                        fname = src_info.get("content", "")
                        if fname and raw_text:
                            # raw_text 안에서 해당 파일 섹션 찾기 — 헤더 패턴 사용
                            scr_id = extract_scr_from_source(fname, raw_text)
                            if scr_id:
                                fallback_map[fname] = scr_id
                # raw_text 전체에서 SCR 패턴 모두 수집 (파일 매핑 안 되도 표시)
                if not fallback_map and raw_text:
                    found = re.findall(r'SCR-\d+[A-Z]?', raw_text)
                    if found:
                        # unique 보존, 순서 유지
                        unique_scrs = list(dict.fromkeys(found))
                        # 가상 키로 보관 — UI 안내용
                        fallback_map = {f"(소스 #{i+1})": scr for i, scr in enumerate(unique_scrs[:10])}
                if fallback_map:
                    sess["_source_scr_map"] = fallback_map
                    push_log(sess, f"[이어서] 화면 식별자 자동 재추출 ({len(fallback_map)}개) — 옛 체크포인트 보강")
                else:
                    sess["_source_scr_map"] = {}
                    push_log(sess, "[이어서] 화면 식별자 매핑 없음 — Excel 화면 코드 컬럼 빈 칸 출력 (정상)")
        else:
            sess["status"] = "parsing"
            push_stage(sess, 1, "문서 파싱", 5)
            raw_text = step_parse_sources(sess, sources)
            check_stop(sess)
            sources_info = [{"type": s.get("type",""), "content": s.get("content",""), "selected_files": s.get("selected_files")} for s in sources]
            save_pipeline_state(project_name, "parsed", {"raw_text": raw_text[:20000], "focus_area": focus_area, "sources_info": sources_info, "source_scr_map": sess.get("_source_scr_map") or {}})
            save_project(project_name, last_sources=sources_info, last_focus_area=focus_area)

        if focus_area:
            push_log(sess, f"[포커스] TC 생성 범위: {focus_area}")

        # 생성 모드 결정 (resume 시에도 동일 로직 사용)
        gen_mode = sess.get("generation_mode", "summary")
        if gen_mode == "direct":
            push_log(sess, f"[모드] Quick 모드 — 3단계 (인벤토리 → 분류 → 섹션 발췌 TC)")
        else:
            push_log(sess, f"[모드] 정책 반영 모드 — 3단계 (통합 추출 → 분류 → 섹션 발췌 TC)")

        if gen_mode == "direct":
            # ── Quick 모드 3단계: 인벤토리 → 분류표 → 섹션발췌(TC 작성 시) ──
            # policy_text는 원문 전체 유지 (TC 작성 시 일반 정책 참조용으로도 사용됨)
            policy_text = raw_text
            if resumed and resumed.get("stage") in ("gate_waiting", "tc_writing"):
                # 복원 시 inventory도 복원 (있으면)
                inventory_text = resumed["data"].get("inventory_text", "") or resumed["data"].get("features_text", "")
                features_text = inventory_text or raw_text
                classification = resumed["data"].get("classification", "")
                push_log(sess, f"[이어서] 분류표 복원됨 ({len(classification):,}자)")
            else:
                # 1단계: 인벤토리 추출
                sess["status"] = "inventory"
                push_stage(sess, 2, "인벤토리 추출 (Quick 1/3)", 18)
                inventory_text = step_quick_inventory(sess, raw_text, project_name, focus_area)
                check_stop(sess)
                save_pipeline_state(project_name, "features", {"raw_text": raw_text[:20000], "policy_text": raw_text[:20000], "features_text": inventory_text, "inventory_text": inventory_text, "focus_area": focus_area, "sources_info": _sources_info, "source_scr_map": sess.get("_source_scr_map") or {}})

                # 2단계: 인벤토리 → 분류표
                sess["status"] = "classifying"
                push_stage(sess, 2, "분류표 생성 (Quick 2/3)", 35)
                classification = step_classify_from_inventory(sess, inventory_text, project_name, focus_area)
                check_stop(sess)
                # features_text 자리에 inventory_text를 세팅 — TC 작성 시 참고 목록으로 사용
                features_text = inventory_text
                save_pipeline_state(project_name, "gate_waiting", {"raw_text": raw_text[:20000], "policy_text": raw_text[:20000], "features_text": inventory_text, "inventory_text": inventory_text, "classification": classification, "focus_area": focus_area, "sources_info": _sources_info, "source_scr_map": sess.get("_source_scr_map") or {}})
            # sess에 flag 저장 — step_write_tc가 섹션 발췌 모드로 동작하도록
            sess["_quick_mode"] = True
            sess["_raw_text"] = raw_text
        else:
            # ── 정책 반영 모드 (최적화: 정책+기능 통합 1회 호출 + TC 작성 섹션 발췌) ──
            # 정책+기능 통합 추출 (기존 policy→features 2회 호출을 1회로)
            if resumed and resumed.get("stage") in ("classifying", "gate_waiting", "tc_writing"):
                policy_text = resumed["data"].get("policy_text", "")
                features_text = resumed["data"].get("features_text", "")
                push_log(sess, f"[이어서] 정책·기능 복원됨 (정책 {len(policy_text):,}자, 기능 {len(features_text):,}자)")
            else:
                sess["status"] = "policy_features"
                push_stage(sess, 2, "정책·기능 통합 추출", 22, eta_sec=90)
                policy_text, features_text = step_policy_features_combined(sess, raw_text, project_name, focus_area)
                check_stop(sess)
                save_pipeline_state(project_name, "features", {"raw_text": raw_text[:20000], "policy_text": policy_text, "features_text": features_text, "focus_area": focus_area, "sources_info": _sources_info, "source_scr_map": sess.get("_source_scr_map") or {}})

            # 분류표
            if resumed and resumed.get("stage") in ("gate_waiting", "tc_writing"):
                classification = resumed["data"].get("classification", "")
                push_log(sess, f"[이어서] 분류표 복원됨 ({len(classification):,}자)")
            else:
                sess["status"] = "classifying"
                push_stage(sess, 2, "분류표 생성", 38)
                classification = step_classify(sess, features_text, project_name, focus_area)
                check_stop(sess)
                save_pipeline_state(project_name, "gate_waiting", {"raw_text": raw_text[:20000], "policy_text": policy_text, "features_text": features_text, "classification": classification, "focus_area": focus_area, "sources_info": _sources_info, "source_scr_map": sess.get("_source_scr_map") or {}})

            # 정책 반영 모드도 TC 작성 시 원문 섹션 발췌 적용 (Quick 모드와 동일)
            sess["_raw_text"] = raw_text
            sess["_section_extract"] = True

        # Gate 단계에서 features/policy 도 sess 에 노출 — gate_chat 이 수정 시 함께 동기화하기 위함 (v0.9.15~)
        sess["_features_text"] = features_text
        sess["_policy_text"] = policy_text

        # Human Gate
        push_stage(sess, 3, "분류표 검토 대기", 45)
        approved = step_gate(sess, classification)
        check_stop(sess)

        # Gate 채팅에서 features/policy 도 수정됐을 수 있음 — 갱신본 사용 (v0.9.15~)
        synced_features = sess.get("_features_text", features_text)
        synced_policy = sess.get("_policy_text", policy_text)

        # TC 작성
        sess["status"] = "tc_writing"
        push_stage(sess, 4, "TC 작성 시작", 50)
        tc_content, total_tc, min_tc = step_write_tc(
            sess, approved, synced_features, synced_policy, project_name,
            selected_domain_codes=sess.get("selected_domains")
        )
        check_stop(sess)

        # TC 검토 (누락 중분류 자동 보강 포함)
        sess["status"] = "reviewing"
        # 검토 시작 — 누락 탐지 약 10초 + 보강(중분류당 ~30초) 가변
        push_stage(sess, 4, "TC 품질 검토 — 누락 탐지 중", 82, eta_sec=20)
        step_review(sess, tc_content, project_name,
                    approved_classification=approved,
                    features_text=synced_features,
                    policy_text=synced_policy)
        check_stop(sess)
        # 보강된 TC가 있으면 채택
        augmented_tc = sess.get("_augmented_tc")
        if augmented_tc and augmented_tc != tc_content:
            tc_content = augmented_tc
            # 보강된 TC의 개수 재계산
            new_tc_count = len(re.findall(r"^###\s+(?!카테고리)", tc_content, re.MULTILINE))
            if new_tc_count > total_tc:
                added = new_tc_count - total_tc
                push_log(sess, f"[검토-보강] TC 총 {new_tc_count}개 (원본 {total_tc} + 보강 {added})")
                total_tc = new_tc_count
                min_tc = max(1, round(total_tc * 0.35))

        push_stage(sess, 4, "TC 검토 완료 · Excel 준비", 88, eta_sec=10)

        # 원칙 G — 그룹 단위 에러 패턴 중복 탐지 (v0.9.17~)
        try:
            dup_report = detect_duplicate_error_tcs(tc_content)
            if dup_report["total_duplicates"] > 0:
                push_log(sess, f"[원칙 G] 중복 의심 패턴 {len(dup_report['patterns'])}개, 통합 시 약 {dup_report['total_duplicates']}개 TC 감소 예상")
                # SSE 로 프론트에 알림 — 사용자가 검토할 수 있게
                push(sess, "duplicate_warning", dup_report)
        except Exception as _e:
            push_log(sess, f"[원칙 G] 중복 탐지 스킵: {_e}")

        # Excel 빌드 — tc_final.md 생성(즉시) + build_excel.py 호출(60초 내외)
        sess["status"] = "building"
        push_stage(sess, 5, "Excel 빌드 — 시트 구성 중", 92, eta_sec=60)
        excel_path = step_build_excel(sess, tc_content, project_name, total_tc, min_tc)
        push_stage(sess, 5, "Excel 빌드 완료 · 마무리 중", 98, eta_sec=3)

        # 후처리(소분류 중복 차별화)가 적용된 tc_content를 마스터로 채택 (MD ↔ Excel 일치)
        tc_content = sess.get("_tc_content_disambiguated", tc_content)

        # TC 마크다운을 tc_files/에 저장 (수정 모드에서 재사용)
        today = datetime.now().strftime("%Y%m%d")
        safe_name = re.sub(r"[^\w\-_]", "_", project_name)[:30]
        tc_md_path = TC_FILES_DIR / f"{safe_name}_{today}.md"
        tc_md_path.write_text(tc_content, encoding="utf-8")
        save_project(project_name, str(tc_md_path), str(excel_path))
        clear_pipeline_state(project_name)  # 완료 시 상태 파일 정리
        push_log(sess, f"[저장] TC 파일 저장: {tc_md_path.name}")

        # 유니크 TC ID 수 — Excel 빌드 후 dedup 결과와 일치하도록 재계산
        unique_tc = count_unique_tc_ids(tc_content)
        if unique_tc and unique_tc != total_tc:
            push_log(sess, f"[완료] TC ID 중복 제거 — {total_tc}개 헤더 중 유니크 {unique_tc}개")
            total_tc = unique_tc
            min_tc = max(1, round(total_tc * 0.35))

        sess["result"] = excel_path
        sess["status"] = "done"
        push_stage(sess, 5, "완료", 100)
        push(sess, "done", {
            "filename":  excel_path.name,
            "size":      excel_path.stat().st_size,
            "sid":       sess["id"],
            "total_tc":  total_tc,
            "min_tc":    min_tc,
            "smoke_tc":  sess.get("smoke_tc"),
        })
        push_log(sess, f"[완료] Excel 파일 생성: {excel_path.name} ({excel_path.stat().st_size:,} bytes)")

        # 1시간 후 workspace 정리 (선택적)
        def cleanup():
            time.sleep(3600)
            import shutil
            ws = sess.get("workspace")
            if ws and Path(ws).exists():
                shutil.rmtree(ws, ignore_errors=True)
        threading.Thread(target=cleanup, daemon=True).start()

    except PipelineStopError:
        sess["status"] = "stopped"
        push_log(sess, "[중단] 사용자가 파이프라인을 중단했습니다.")
        push(sess, "stopped", {"msg": "파이프라인이 중단되었습니다.", "stage": sess.get("status", "")})

    except Exception as e:
        import traceback
        push_log(sess, f"[오류] {e}")
        push_log(sess, traceback.format_exc()[:1000])
        push_error(sess, str(e))


# ── 수정 파이프라인 ────────────────────────────────────────────────────────────
def run_modify_pipeline(sess: dict, project_name: str, existing_tc: str, change_desc: str, resume_stage: str = None):
    try:
        tc_rules = load_tc_rules()
        save_checkpoint(project_name, change_desc, "analyzing", {"existing_tc": existing_tc[:5000]})

        # Step M1: 영향도 분석
        sess["status"] = "analyzing"
        push_stage(sess, 2, "변경 영향 분석 중", 20)
        push_log(sess, "[영향 분석] 기존 TC와 변경사항을 비교 중...")

        tc_count_before = len(re.findall(r"^###\s", existing_tc, re.MULTILINE))

        system_impact = f"""당신은 QA 엔지니어입니다. 기존 TC 목록과 변경사항을 비교하여 영향받는 TC를 정확히 식별하세요.

TC 형식 규칙:
- bold heading (### **SC-XXX-YYY-NNN**) = 최소 TC
- plain heading (### SC-XXX-YYY-NNN) = 일반 TC

{tc_rules if tc_rules else ""}"""

        user_impact = f"""## 기존 TC ({tc_count_before}개)

{existing_tc[:12000]}

## 변경사항

{change_desc}

위 변경사항을 분석하여 다음 형식으로 영향도를 출력하세요.
각 항목에 TC ID와 구체적인 이유를 반드시 포함하세요.

## 삭제할 TC
| TC ID | 삭제 이유 |
|-------|---------|
| SC-XXX-YYY-001 | 해당 기능이 제거됨 |

## 수정할 TC
| TC ID | 변경 내용 |
|-------|---------|
| SC-XXX-YYY-002 | When 단계에서 OTP 인증 추가 |

## 새로 추가할 TC
| 기능/시나리오 | TC 개수 (예상) |
|------------|------------|
| OTP 인증 실패 처리 | 3 |

## 영향 요약
- 삭제: N개
- 수정: N개
- 신규 추가: N개
- 변경 후 예상 TC 수: N개"""

        impact_analysis = call_claude(system_impact, user_impact, max_tokens=4096)
        check_stop(sess)

        impact_path = sess["workspace"] / "impact_analysis.md"
        impact_path.write_text(impact_analysis, encoding="utf-8")
        push_log(sess, f"[영향 분석] 완료 — 결과 {len(impact_analysis):,}자")
        save_checkpoint(project_name, change_desc, "gate_waiting", {"impact_analysis": impact_analysis, "existing_tc": existing_tc[:5000]})

        # Step M2: Human Gate (영향도 검토)
        push_stage(sess, 3, "영향도 검토 대기", 40)
        push_log(sess, "[GATE] 영향도 분석 결과를 검토해주세요.")
        sess["status"] = "gate_waiting"
        push(sess, "gate", {"content": impact_analysis, "mode": "modify"})
        sess["gate_event"].wait()
        approved_plan = sess["approved"]
        check_stop(sess)

        approved_path = sess["workspace"] / "impact_APPROVED.md"
        approved_path.write_text(approved_plan, encoding="utf-8")
        push_log(sess, "[GATE] 영향도 승인됨. TC 수정 시작.")

        # Step M3: TC 수정 적용
        sess["status"] = "tc_writing"
        push_stage(sess, 4, "TC 수정 적용 중", 55)
        push_log(sess, "[TC 수정] 승인된 계획에 따라 TC를 수정 중...")

        system_modify = f"""당신은 QA 엔지니어입니다. 기존 TC를 수정 계획에 따라 정확히 수정하세요.

## 🎯 영향도 분석 우선 원칙 (절대 규칙 — 반드시 준수)

⚠️ **사용자가 검토·승인한 영향도 분석(approved_plan)이 최종 진실 소스입니다.**

영향도 분석은 단순 참고가 아닙니다. 사용자가 Gate 검토 단계에서 다음과 같은 수정을 했을 수 있습니다:
- 삭제 대상에서 일부 TC 제외 (보존)
- 수정 대상에 추가 TC 포함 또는 제외
- 신규 추가 항목 변경
- 변경 범위 조정

**그러므로**:

1. **approved_plan 에 명시된 변경(삭제/수정/추가)만 정확히 적용하세요.**
2. **approved_plan 에 없는 TC 는 변경하지 마세요** — 기존 TC 그대로 유지.
3. **기존 TC (existing_tc) 의 내용을 보고 영향이 있어 보여도, approved_plan 에 없으면 건드리지 마세요.**
4. **임의로 추가 TC 를 생성하지 마세요** — 영향도 분석의 "신규 추가" 항목만 추가.
5. 만약 approved_plan 의 지시가 모호하면 보수적으로 판단 (변경 최소화 우선).

예시:
- approved_plan: "SCR-003-001 삭제, SCR-005-002 수정 (Given 단계에 OTP 추가)"
- → 정확히 그 두 TC 만 변경, 나머지 모든 TC 는 그대로 유지
- 사용자가 "SCR-007-003 도 비슷해 보이니까 같이 수정" 식의 환상 금지

## TC 형식 규칙
- bold heading (### **SC-XXX-YYY-NNN**) = 최소 TC (smoke test 대상)
- plain heading (### SC-XXX-YYY-NNN) = 일반 TC
- Given(사전 조건)은 번호 매긴 목록
- 전체 TC의 약 35%가 최소 TC여야 함

{tc_rules if tc_rules else ""}"""

        user_modify = f"""## 수정 지시사항 (사용자 승인 — 최종 진실 소스)

{approved_plan}

## 변경 배경 (참고)

{change_desc}

## 기존 TC 전체 (참고용 — 승인된 수정 지시사항이 우선)

⚠️ 아래 기존 TC 들은 **참고용**입니다. 위 수정 지시사항에 명시된 변경만 적용하세요.
지시사항에 없는 TC 는 임의로 변경하지 마세요.

{existing_tc[:15000]}

---

위 수정 계획을 정확히 적용하여 **수정된 TC 전체**를 출력하세요.
- 삭제 대상 TC는 완전히 제거
- 수정 대상 TC는 내용 업데이트 (TC ID 유지)
- 신규 TC는 기존 ID 체계에 맞춰 추가 (마지막 번호 이어서)
- **수정 지시사항에 없는 TC는 그대로 유지** (임의 변경 금지)
- 기존 TC 형식(### heading, Given/When/Then)을 반드시 유지"""

        modified_tc = call_claude(system_modify, user_modify, max_tokens=16000)
        check_stop(sess)

        modified_path = sess["workspace"] / "tc_modified.md"
        modified_path.write_text(modified_tc, encoding="utf-8")

        total_tc = len(re.findall(r"^###\s", modified_tc, re.MULTILINE))
        min_tc   = max(1, round(total_tc * 0.35))
        push_log(sess, f"[TC 수정] 완료 — 총 {total_tc}개 TC (최소 TC: {min_tc}개)")

        # Step M4: TC 검토
        sess["status"] = "reviewing"
        push_stage(sess, 4, "TC 품질 검토", 80)
        step_review(sess, modified_tc, project_name)
        check_stop(sess)

        # Step M5: Excel 빌드
        sess["status"] = "building"
        push_stage(sess, 5, "Excel 빌드", 90)
        excel_path = step_build_excel(sess, modified_tc, project_name, total_tc, min_tc)
        # 후처리 적용본 채택 (MD ↔ Excel 일치)
        modified_tc = sess.get("_tc_content_disambiguated", modified_tc)

        # TC 파일 저장 (수정본으로 덮어쓰기)
        today = datetime.now().strftime("%Y%m%d")
        safe_name = re.sub(r"[^\w\-_]", "_", project_name)[:30]
        tc_md_path = TC_FILES_DIR / f"{safe_name}_{today}.md"
        tc_md_path.write_text(modified_tc, encoding="utf-8")
        save_project(project_name, str(tc_md_path), str(excel_path))

        # 유니크 TC ID 수 — Excel dedup 결과와 일치하도록 재계산
        unique_tc = count_unique_tc_ids(modified_tc)
        if unique_tc and unique_tc != total_tc:
            push_log(sess, f"[완료] TC ID 중복 제거 — {total_tc}개 헤더 중 유니크 {unique_tc}개")
            total_tc = unique_tc
            min_tc = max(1, round(total_tc * 0.35))

        sess["result"] = excel_path
        sess["status"] = "done"
        clear_checkpoint()
        push_stage(sess, 5, "완료", 100)
        push(sess, "done", {
            "filename":  excel_path.name,
            "size":      excel_path.stat().st_size,
            "sid":       sess["id"],
            "total_tc":  total_tc,
            "min_tc":    min_tc,
            "smoke_tc":  sess.get("smoke_tc"),
        })
        push_log(sess, f"[완료] 수정 Excel 저장: {excel_path.name}")

    except PipelineStopError:
        sess["status"] = "stopped"
        push_log(sess, "[중단] 사용자가 수정 파이프라인을 중단했습니다.")
        push(sess, "stopped", {"msg": "수정 파이프라인이 중단되었습니다."})

    except Exception as e:
        import traceback
        push_log(sess, f"[오류] {e}")
        push_log(sess, traceback.format_exc()[:1000])
        sess["status"] = "error"
        # 체크포인트가 있으면 복구 옵션 이벤트 발송
        cp = load_checkpoint()
        push(sess, "error_recovery", {
            "error": str(e),
            "has_checkpoint": bool(cp),
            "checkpoint_stage": cp.get("stage", ""),
            "checkpoint_project": cp.get("project_name", ""),
            "checkpoint_saved_at": cp.get("saved_at", ""),
        })
        push_error(sess, str(e))


# ── Flask 라우트 ───────────────────────────────────────────────────────────────

# ── JS 자가 검증 (계층 2) ──────────────────────────────────────────────────
# 매 / 요청마다 매우 가벼운 정규식 + node --check (있을 때) 로 검증.
# 결과를 페이지에 주입하면 계층 1 의 안전망이 받아서 사용자에게 배너로 노출.
_JS_CHECK_CACHE = {"html_id": None, "result": None}


def _quick_validate_served_html(html: str) -> dict:
    """서빙되는 HTML 안의 큰 <script> 블록을 추출해 node --check 로 검증.
    node 가 없거나 검증 실패 시 graceful — 응답에 문제 정보만 담는다.
    """
    import subprocess
    import re as _re
    import shutil as _shutil
    import tempfile as _tmp

    # 캐시 — 같은 HTML 본문이면 재검증 안 함 (id() 는 같은 문자열이면 같음 보장 안 됨,
    # 길이 + hash 로 키)
    key = (len(html), hash(html) & 0xFFFFFFFF)
    if _JS_CHECK_CACHE["html_id"] == key:
        return _JS_CHECK_CACHE["result"] or {"ok": True}

    result = {"ok": True}
    node_bin = _shutil.which("node")
    if not node_bin:
        result = {"ok": True, "skipped": "node not found in PATH"}
        _JS_CHECK_CACHE.update(html_id=key, result=result)
        return result

    scripts = _re.findall(r"<script[^>]*>(.*?)</script>", html, _re.DOTALL)
    if not scripts:
        result = {"ok": True, "skipped": "no script blocks"}
        _JS_CHECK_CACHE.update(html_id=key, result=result)
        return result

    # 가장 큰 스크립트 (보통 메인 inline script) 만 검증 — 안전망 inline script 는 작아서 무시
    main_script = max(scripts, key=len)
    try:
        with _tmp.NamedTemporaryFile(mode="w", suffix=".js", delete=False, encoding="utf-8") as f:
            f.write(main_script)
            tmp_path = f.name
        proc = subprocess.run([node_bin, "--check", tmp_path],
                                capture_output=True, text=True, timeout=10)
        try:
            os.unlink(tmp_path)
        except Exception:
            pass

        if proc.returncode != 0:
            err_lines = (proc.stderr or "").splitlines()
            # 너무 길면 잘라서 — 핵심 라인만
            first_err = next((ln for ln in err_lines if "Error" in ln or "SyntaxError" in ln), err_lines[0] if err_lines else "unknown")
            line_info = next((ln for ln in err_lines if _re.search(r":\d+", ln)), "")
            result = {
                "ok": False,
                "error": f"{first_err.strip()} | {line_info.strip()[:200]}",
                "full": "\n".join(err_lines[:20]),
            }
            # 서버 콘솔에도 출력 (개발자 알림)
            print(f"\n⚠️ [JS Self-Check 실패] {result['error']}", flush=True)
            print(result["full"][:2000], flush=True)
        else:
            result = {"ok": True}
    except subprocess.TimeoutExpired:
        result = {"ok": True, "skipped": "node check timeout"}
    except Exception as e:
        result = {"ok": True, "skipped": f"check error: {e}"}

    _JS_CHECK_CACHE.update(html_id=key, result=result)
    return result


@app.route("/")
def index():
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    api_warning = not api_key or api_key == "sk-ant-..."
    html = render_template_string(
        HTML_TEMPLATE,
        api_warning=api_warning,
        app_version=APP_VERSION,
        app_version_date=APP_VERSION_DATE,
        app_version_tagline=APP_VERSION_TAGLINE,
        app_version_highlights=APP_VERSION_HIGHLIGHTS,
    )
    # 계층 2: JS 자가 검증 결과를 페이지에 주입 — 계층 1 안전망이 받음
    check = _quick_validate_served_html(html)
    if not check.get("ok"):
        import json as _json
        injection = (
            "<script>window.__SERVER_JS_CHECK = "
            + _json.dumps(check, ensure_ascii=False)
            + ";</script>\n</body>"
        )
        html = html.replace("</body>", injection, 1)
    return html


# ── 관리자 엔드포인트 ─────────────────────────────────────────────────────────
# 활성 세션으로 간주되는 status (이 외에는 종료/대기 상태로 봄)
# v0.9.18: gate_waiting 도 활성 세션 — 사용자가 분류표 검토 중인데 재시작하면
#          승인 시점에 sid 가 사라져 "세션 없음" 404 발생함
_ACTIVE_STATUSES = {
    "parsing", "inventory", "classifying", "policy_features",
    "gate_waiting",  # ← v0.9.18 추가: 분류표 검토 대기 중인 사용자 보호
    "tc_writing", "reviewing", "building", "analyzing",
}

def _count_active_sessions() -> int:
    return sum(1 for s in SESSIONS.values() if s.get("status") in _ACTIVE_STATUSES)

def _is_localhost_request() -> bool:
    """요청이 로컬호스트(127.0.0.1 / ::1)에서 왔는지 검사."""
    addr = request.remote_addr or ""
    return addr in ("127.0.0.1", "::1", "localhost")


@app.route("/admin/status", methods=["GET"])
def admin_status():
    """서버 살아있음 + 버전 + 활성 세션 수 (재시작 폴링용)."""
    return jsonify({
        "ok": True,
        "version": APP_VERSION,
        "active_sessions": _count_active_sessions(),
        "localhost": _is_localhost_request(),
    })


@app.route("/admin/restart", methods=["POST"])
def admin_restart():
    """서버 자기 자신을 재시작 (os.execv).
    보안:
      - localhost 요청만 허용
      - 활성 세션 있으면 ?force=1 로 강제 가능
    """
    # 1) localhost 가드 (LAN 다른 PC 차단)
    if not _is_localhost_request():
        return jsonify({"ok": False, "error": "재시작은 로컬에서만 허용됩니다."}), 403

    # 2) 활성 세션 가드
    force = (request.args.get("force", "0") == "1") or \
            (request.get_json(silent=True) or {}).get("force") is True
    active = _count_active_sessions()
    if active > 0 and not force:
        return jsonify({
            "ok": False,
            "error": "active_sessions",
            "active_sessions": active,
            "message": f"진행 중인 세션이 {active}개 있습니다. 정말 재시작하려면 force=1 로 다시 요청하세요.",
        }), 409

    # 3) 응답을 먼저 보내고 백그라운드에서 실제 재시작 (1.5초 후 os.execv)
    def _do_restart():
        try:
            time.sleep(1.5)  # 응답 도달 시간 확보
        finally:
            # listen socket 명시적 close — 새 프로세스만 incoming connection 처리하도록
            try:
                sk = getattr(app, "_listen_sock", None)
                if sk is not None:
                    sk.close()
            except Exception:
                pass
            # 현재 인터프리터 + 동일 인자로 자기 자신 재실행
            # SO_REUSEADDR 가 활성화되어 있어 새 프로세스도 즉시 bind 가능
            os.execv(sys.executable, [sys.executable] + sys.argv)

    threading.Thread(target=_do_restart, daemon=True).start()
    return jsonify({
        "ok": True,
        "message": "서버 재시작 중... 약 3~5초 후 자동 재연결됩니다.",
        "delay_sec": 1.5,
    })


@app.route("/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "파일 없음"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"ok": False, "error": "파일명 없음"}), 400
    if not f.filename.lower().endswith(".pdf"):
        return jsonify({"ok": False, "error": "PDF 파일만 허용"}), 400
    save_path = SPECS_DIR / f.filename
    f.save(str(save_path))
    return jsonify({"ok": True, "filename": f.filename})


@app.route("/upload-md", methods=["POST"])
def upload_md():
    """마크다운 파일 업로드. 같은 이름 파일이 이미 있으면 _1/_2... 접미사로 구분 보존."""
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "파일 없음"}), 400
    f = request.files["file"]
    raw_name = (f.filename or "").strip()
    if not raw_name:
        return jsonify({"ok": False, "error": "파일명 없음"}), 400
    # webkitdirectory로 업로드 시 webkitRelativePath로 들어오는 경로 분리자 제거
    raw_name = raw_name.replace("\\", "/").split("/")[-1]
    # 안전한 파일명 (경로 조작 방지)
    safe_name = re.sub(r'[<>:"/\\|?*]', "_", raw_name).strip(". ")
    if not safe_name:
        return jsonify({"ok": False, "error": "유효하지 않은 파일명"}), 400
    ext = safe_name.rsplit(".", 1)[-1].lower() if "." in safe_name else ""
    if ext not in ("md", "markdown", "txt"):
        return jsonify({"ok": False, "error": ".md / .markdown / .txt 파일만 허용"}), 400

    # 중복 처리: 기존 파일 있으면 _1, _2 ... 접미사 자동 부여
    stem, suffix = safe_name.rsplit(".", 1)
    final_name = safe_name
    counter = 1
    while (SPECS_DIR / final_name).exists():
        final_name = f"{stem}_{counter}.{suffix}"
        counter += 1
        if counter > 999:
            return jsonify({"ok": False, "error": "같은 이름의 파일이 너무 많습니다"}), 400

    save_path = SPECS_DIR / final_name
    f.save(str(save_path))
    return jsonify({
        "ok": True,
        "filename": final_name,
        "original_filename": raw_name,
        "renamed": final_name != safe_name,
    })


def run_pipeline_structured(sess: dict, folder_path: str, project_name: str,
                              prev_folder_path: str = "",
                              include_unchanged: bool = False):
    """구조화 spec 폴더 파이프라인 — LLM 분류 단계 스킵, 화면별 1:1 호출.

    Args:
        prev_folder_path: 비어있지 않으면 diff 모드 — 이전 폴더 대비 추가/수정 SCR 만 재생성.
        include_unchanged: True면 diff 가 있어도 모든 화면 처리 (분류표 일관성용).
    """
    focus_area = sess.get("focus_area", "")
    diff_mode = bool(prev_folder_path)
    # Combo only 모드 — Stage 2~4 (분류표 / Human Gate / SCR TC) 모두 skip
    is_combo_only = sess.get("generation_mode") == "combo_only"

    try:
        # ── 0) Combo only 단축 경로 — 분류 단계 모두 건너뛰고 바로 Combo TC 생성
        if is_combo_only:
            push_log(sess, "[Combo Only 모드] SCR 분류 단계 skip — Combo TC 만 생성합니다.")
            sess["status"] = "tc_writing"
            push_stage(sess, 1, "Combo TC 생성 시작", 30)

            # workspace 준비
            if "workspace" not in sess:
                sess["workspace"] = Path(__file__).parent.parent / "workspace" / sess["id"]
                sess["workspace"].mkdir(parents=True, exist_ok=True)

            combo_opts = sess.get("combo_opts")
            if not combo_opts or not combo_opts.get("file_paths"):
                raise RuntimeError("Combo 모드인데 명세 파일이 지정되지 않았습니다.")

            merged_tc = ""
            total_tc = 0
            combo_md_parts = []
            combo_count = 0
            project_code = _detect_project_code(project_name)

            for fp in combo_opts["file_paths"]:
                p = Path(fp).expanduser()
                if not p.exists():
                    push_log(sess, f"[Combo] 파일 없음: {p}")
                    continue
                cdata = _parse_combo_file_direct(p)
                if not cdata:
                    push_log(sess, f"[Combo] 파싱 실패: {p.name}")
                    continue
                # 필터 적용
                combos_list = cdata["combos"]
                scenarios_list = cdata["scenarios"]
                oc_f = (combo_opts.get("oc_filter") or "").strip()
                scn_f = (combo_opts.get("scn_filter") or "").strip()
                is_filter = bool(oc_f or scn_f)
                if oc_f:
                    ids = [s.strip().upper() for s in oc_f.split(",")]
                    def _match(c_id):
                        for pat in ids:
                            if pat.endswith("*"):
                                if c_id.startswith(pat[:-1]):
                                    return True
                            elif c_id == pat:
                                return True
                        return False
                    combos_list = [c for c in combos_list if _match(c["id"])]
                elif is_filter:
                    combos_list = []
                if scn_f:
                    ids = [s.strip().upper() for s in scn_f.split(",")]
                    scenarios_list = [s for s in scenarios_list if s["id"].upper() in ids]
                elif is_filter:
                    scenarios_list = []

                domain = _parse_combo_md_lite(p)["domain"]
                if is_filter:
                    new_raw = _rebuild_raw_md_for_combo(
                        cdata["raw_md"],
                        [c["id"] for c in combos_list],
                        [s["id"] for s in scenarios_list],
                    )
                else:
                    new_raw = cdata["raw_md"]
                trimmed = {
                    "raw_md": new_raw, "combos": combos_list, "scenarios": scenarios_list,
                    "_sample_mode": is_filter,
                }
                # 파일별 SuiteCode 추출 — TC ID 충돌 방지
                # 예: order_combinations.md → ORDR / trade_lite__combinations.md → LITE
                suite_code = _derive_combo_suite_code(p.name)
                suite_combo = f"{suite_code}-COMBO"
                suite_flow = f"{suite_code}-FLOW"

                push_log(sess, f"[Combo] {p.name} 처리 — OC {len(combos_list)} + scenarios {len(scenarios_list)} ({sum(len(s['steps']) for s in scenarios_list)}step) → SuiteCode '{suite_code}'")
                try:
                    combo_tc_md, c_count = step_write_combo_tc(sess, project_name, trimmed,
                                                                 project_code=project_code,
                                                                 suite_combo=suite_combo,
                                                                 suite_flow=suite_flow)
                    if domain != "Trade(Combo)":
                        combo_tc_md = combo_tc_md.replace(
                            "| 대분류 | Trade(Combo) |",
                            f"| 대분류 | {domain} |",
                        )
                    combo_md_parts.append(combo_tc_md)
                    combo_count += c_count
                    push_log(sess, f"[Combo] {p.name} → {c_count} TC ({domain})")
                except Exception as ce:
                    push_log(sess, f"[Combo] LLM 호출 실패 ({p.name}): {ce}")
                check_stop(sess)

            if not combo_md_parts:
                raise RuntimeError("Combo TC 가 1개도 생성되지 않았습니다.")

            merged_tc = "\n\n---\n\n".join(combo_md_parts)
            total_tc = combo_count
            push_log(sess, f"[Combo] 통합 완료 — 총 {total_tc} TC")

            # ── Excel 빌드 ──
            sess["status"] = "building_excel"
            push_stage(sess, 2, "Excel 빌드", 90)
            min_tc = max(1, round(total_tc * 0.35))
            result_file = step_build_excel(sess, merged_tc, project_name, total_tc, min_tc)
            excel_path = Path(result_file)
            sess["result"] = str(excel_path)

            sess["status"] = "done"
            push_stage(sess, 3, "완료", 100)
            smoke_count = len(re.findall(r"^###\s+\*\*[A-Z]{2}-", merged_tc, re.MULTILINE))
            push(sess, "done", {
                "filename":  excel_path.name,
                "size":      excel_path.stat().st_size if excel_path.exists() else 0,
                "sid":       sess["id"],
                "total_tc":  total_tc,
                "min_tc":    min_tc,
                "smoke_tc":  smoke_count,
            })
            clear_pipeline_state(project_name)
            return

        # ── 1) 폴더 파싱 (LLM 호출 없음) ── [일반 SCR 모드]
        sess["status"] = "parsing"
        if diff_mode:
            push_stage(sess, 1, "구조화 spec 폴더 + diff 비교", 8)
            spec_data = step_parse_structured_spec_with_diff(
                sess, folder_path, prev_folder_path, include_unchanged=include_unchanged
            )
        else:
            push_stage(sess, 1, "구조화 spec 폴더 파싱", 8)
            spec_data = step_parse_structured_spec(sess, folder_path)
        check_stop(sess)

        if not spec_data["screens"]:
            if diff_mode and not spec_data.get("_diff", {}).get("common_changed"):
                # diff 모드에서 변경된 화면이 0개면 정상 종료
                push_log(sess, "[diff] 변경된 화면이 없습니다. 재생성할 TC 없음 — 종료.")
                sess["status"] = "done"
                push_stage(sess, 7, "완료 (변경 없음)", 100)
                push(sess, "done", {"result": None, "tc_count": 0, "diff": spec_data.get("_diff")})
                clear_pipeline_state(project_name)
                return
            raise RuntimeError(f"화면 md 가 없습니다: {folder_path}/scr/")
        if not spec_data["screen_rows"]:
            push_log(sess, "[경고] overview 의 화면 목록 표를 찾지 못함 — 화면 파일명/H1만으로 분류표 생성")
            # screen_rows fallback
            for sc in spec_data["screens"]:
                spec_data["screen_rows"].append({
                    "id": sc["id"], "name": sc["title"], "major": "공통",
                    "middle": sc["title"], "status": "active",
                    "desc": sc["description"], "entry": "",
                })

        # ── focus_area 에서 SCR ID 필터 추출 → 화면 한정 처리 ──
        # available_scrs 를 함께 넘겨야 "SCR-102, 104, 106" 같은 일괄 표기,
        # "SCR-102~116" 같은 범위 표기를 정확히 인식할 수 있다.
        available_scrs = {sc["id"] for sc in spec_data["screens"]}
        scr_filter = parse_scr_filter_from_focus(focus_area, available_scrs=available_scrs)
        if scr_filter:
            before_screens = len(spec_data["screens"])
            before_rows = len(spec_data["screen_rows"])
            spec_data["screens"] = [sc for sc in spec_data["screens"] if sc["id"] in scr_filter]
            spec_data["screen_rows"] = [r for r in spec_data["screen_rows"] if r["id"] in scr_filter]
            if not spec_data["screens"]:
                raise RuntimeError(
                    f"focus_area 에서 추출한 SCR ID({sorted(scr_filter)}) 가 폴더에서 매칭되지 않습니다. "
                    f"폴더에 있는 화면 md를 확인하세요."
                )
            push_log(sess,
                f"[focus] SCR 필터 적용 — {sorted(scr_filter)} → "
                f"화면 {before_screens}→{len(spec_data['screens'])}개, 행 {before_rows}→{len(spec_data['screen_rows'])}개")

        # raw_text 호환용 (resume/체크포인트용 — 전체 텍스트)
        raw_text = "\n\n".join([
            spec_data["overview_text"],
            spec_data["policy_text"],
            spec_data["design_text"],
        ])
        # SCR 매핑 — 화면 파일에서 ID 그대로
        sess["_source_scr_map"] = {sc["id"]: sc["id"] for sc in spec_data["screens"]}

        # ── 2) 분류표 자동 생성 (LLM 호출 없음) ──
        sess["status"] = "classifying"
        push_stage(sess, 2, "분류표 자동 생성 (LLM 호출 없음)", 25)
        # 필터링된 화면 기준으로 분류표 생성 (focus_area 가 있으면 그 화면들만).
        # screens 메타까지 넘겨야 SCR md 본문에서 소분류(세부 시나리오)를 자동 추출 가능.
        all_rows = spec_data["screen_rows"]
        classification = build_classification_from_screen_list(
            all_rows, project_name, screens_meta=spec_data.get("screens"),
        )
        classify_path = sess["workspace"] / "04_classification_draft.md"
        classify_path.write_text(classification, encoding="utf-8")
        push_log(sess, f"[분류] 화면 목록 표에서 분류표 자동 생성 — {len(classification):,}자")

        # 체크포인트 저장
        sources_info = [{"type": "spec_folder", "content": folder_path}]
        if diff_mode:
            sources_info.append({"type": "spec_folder_prev", "content": prev_folder_path})
        save_pipeline_state(project_name, "gate_waiting", {
            "raw_text": raw_text[:20000],
            "classification": classification,
            "focus_area": focus_area,
            "sources_info": sources_info,
            "source_scr_map": sess["_source_scr_map"],
            "structured_spec_folder": folder_path,
            "prev_folder": prev_folder_path,
            "include_unchanged": include_unchanged,
            "workspace": str(sess["workspace"]),
        })
        save_project(project_name, last_sources=sources_info, last_focus_area=focus_area)

        # spec_data 보관 — Human Gate 통과 후 step_write_tc_per_screen 에서 사용
        sess["_structured_spec_data"] = spec_data

        # ── 3) Human Gate ──
        sess["status"] = "gate_waiting"
        push_stage(sess, 3, "분류표 검토 대기 (Human Gate)", 50)
        approved = step_gate(sess, classification)
        check_stop(sess)

        # ── 4) 화면별 1:1 TC 작성 ──
        sess["status"] = "tc_writing"
        push_stage(sess, 4, "화면별 TC 작성 (cache 활용)", 55)
        # tc_writing 단계 진입 전 체크포인트 갱신 — resume 시 여기로 복귀
        save_pipeline_state(project_name, "tc_writing", {
            "raw_text": raw_text[:20000],
            "classification": classification,
            "approved_classification": approved,
            "focus_area": focus_area,
            "sources_info": sources_info,
            "source_scr_map": sess["_source_scr_map"],
            "structured_spec_folder": folder_path,
            "prev_folder": prev_folder_path,
            "include_unchanged": include_unchanged,
            "workspace": str(sess["workspace"]),
            "selected_domains": sess.get("selected_domains"),
            "suite_codes": sess.get("suite_codes"),
        })
        selected = sess.get("selected_domains")
        # Note: is_combo_only 는 함수 시작부에서 단축 경로로 분기 — 여기에 안 옴
        merged_tc, total_tc = step_write_tc_per_screen(
            sess, approved, spec_data, project_name,
            selected_domain_codes=selected,
        )

        # ── 5) Review (선택)
        sess["status"] = "reviewing"
        push_stage(sess, 5, "TC 검토 / 정리", 82)
        try:
            step_review(sess, merged_tc, project_name)
        except Exception as e:
            push_log(sess, f"[검토] 스킵 — {e}")

        # ── 5.5) Combo TC 생성 (combo_only 모드에서 단독, 또는 향후 확장 시 추가)
        combo_opts = sess.get("combo_opts")
        if combo_opts and combo_opts.get("file_paths"):
            push_stage(sess, 5, "Combo TC 생성", 88)
            push_log(sess, f"[Combo] {len(combo_opts['file_paths'])}개 명세 파일 처리 시작")
            combo_md_parts = []
            combo_total = 0
            for fp in combo_opts["file_paths"]:
                p = Path(fp).expanduser()
                if not p.exists():
                    push_log(sess, f"[Combo] 파일 없음: {p}")
                    continue
                cdata = _parse_combo_file_direct(p)
                if not cdata:
                    push_log(sess, f"[Combo] 파싱 실패: {p.name}")
                    continue
                # 필터 적용
                combos_list = cdata["combos"]
                scenarios_list = cdata["scenarios"]
                oc_f = (combo_opts.get("oc_filter") or "").strip()
                scn_f = (combo_opts.get("scn_filter") or "").strip()
                is_filter = bool(oc_f or scn_f)
                if oc_f:
                    ids = [s.strip().upper() for s in oc_f.split(",")]
                    def _match(c_id):
                        for pat in ids:
                            if pat.endswith("*"):
                                if c_id.startswith(pat[:-1]):
                                    return True
                            elif c_id == pat:
                                return True
                        return False
                    combos_list = [c for c in combos_list if _match(c["id"])]
                elif is_filter:
                    combos_list = []
                if scn_f:
                    ids = [s.strip().upper() for s in scn_f.split(",")]
                    scenarios_list = [s for s in scenarios_list if s["id"].upper() in ids]
                elif is_filter:
                    scenarios_list = []

                domain = _parse_combo_md_lite(p)["domain"]
                if is_filter:
                    new_raw = _rebuild_raw_md_for_combo(
                        cdata["raw_md"],
                        [c["id"] for c in combos_list],
                        [s["id"] for s in scenarios_list],
                    )
                else:
                    new_raw = cdata["raw_md"]
                trimmed = {
                    "raw_md": new_raw, "combos": combos_list, "scenarios": scenarios_list,
                    "_sample_mode": is_filter,
                }
                try:
                    project_code = _detect_project_code(project_name)
                    # 파일별 SuiteCode — TC ID 충돌 방지
                    suite_code = _derive_combo_suite_code(p.name)
                    combo_tc_md, c_count = step_write_combo_tc(
                        sess, project_name, trimmed,
                        project_code=project_code,
                        suite_combo=f"{suite_code}-COMBO",
                        suite_flow=f"{suite_code}-FLOW",
                    )
                    # 대분류 치환 (Trade(Combo) → 파일별 도메인)
                    if domain != "Trade(Combo)":
                        combo_tc_md = combo_tc_md.replace(
                            "| 대분류 | Trade(Combo) |",
                            f"| 대분류 | {domain} |",
                        )
                    combo_md_parts.append(combo_tc_md)
                    combo_total += c_count
                    push_log(sess, f"[Combo] {p.name} → {c_count} TC ({domain})")
                except Exception as ce:
                    push_log(sess, f"[Combo] LLM 호출 실패: {ce}")
            if combo_md_parts:
                merged_tc = merged_tc.rstrip() + "\n\n---\n\n" + "\n\n---\n\n".join(combo_md_parts)
                total_tc += combo_total
                push_log(sess, f"[Combo] 통합 완료 — Combo TC {combo_total}개 추가 (총 {total_tc})")

        # ── 6) Excel 빌드 — step_build_excel 이 내부에서 tc_final.md 도 직접 생성한다.
        sess["status"] = "building_excel"
        push_stage(sess, 6, "Excel 빌드", 92)
        # total_tc 는 step_write_tc_per_screen 에서 받은 값이 정확. min_tc 는 표준 35% 공식.
        min_tc = max(1, round(total_tc * 0.35))
        result_file = step_build_excel(sess, merged_tc, project_name, total_tc, min_tc)
        excel_path = Path(result_file)
        sess["result"] = str(excel_path)

        sess["status"] = "done"
        push_stage(sess, 7, "완료", 100)
        # 기존 파이프라인과 동일한 페이로드 — 프론트가 filename/size/total_tc/min_tc/smoke_tc 를 사용
        push(sess, "done", {
            "filename":  excel_path.name,
            "size":      excel_path.stat().st_size if excel_path.exists() else 0,
            "sid":       sess["id"],
            "total_tc":  total_tc,
            "min_tc":    min_tc,
            "smoke_tc":  sess.get("smoke_tc"),
        })
        push_log(sess, f"[완료] 화면 {len(spec_data['screens'])}개 → TC {total_tc}개 → {excel_path.name}")

        # 체크포인트 정리
        clear_pipeline_state(project_name)
    except Exception as e:
        sess["status"] = "error"
        push_error(sess, f"구조화 spec 파이프라인 오류: {e}")


@app.route("/start-spec-folder", methods=["POST"])
def start_spec_folder():
    """구조화 spec 폴더 1개를 받아 파이프라인 실행.
    POST body: {project_name, folder_path, focus_area?, prev_folder_path?, include_unchanged?, resume?}
    - prev_folder_path 가 있으면 diff 모드 (변경/추가 SCR 만 재생성)
    - resume=true 면 이전 체크포인트(workspace/per_screen/*.md)로부터 이어서 작업
    """
    data = request.get_json(force=True) or {}
    project_name = data.get("project_name", "프로젝트").strip() or "프로젝트"
    focus_area   = (data.get("focus_area") or "").strip()
    folder_path  = (data.get("folder_path") or "").strip()
    prev_folder  = (data.get("prev_folder_path") or "").strip()
    include_unchanged = bool(data.get("include_unchanged"))
    resume       = bool(data.get("resume"))
    generation_mode = (data.get("generation_mode") or "").strip()
    combo_opts   = data.get("combo") or None  # {file_paths, oc_filter, scn_filter}

    # resume 모드 — 저장된 체크포인트 복원
    if resume:
        state = load_pipeline_state(project_name)
        if not state:
            return jsonify({"ok": False, "error": "저장된 체크포인트가 없습니다."}), 400
        if not os.environ.get("ANTHROPIC_API_KEY"):
            return jsonify({"ok": False, "error": "ANTHROPIC_API_KEY가 설정되지 않았습니다."}), 500
        ckpt = state["data"]
        ckpt_folder = ckpt.get("structured_spec_folder")
        if not ckpt_folder:
            return jsonify({"ok": False, "error": "체크포인트에 spec 폴더 정보가 없습니다 (구조화 spec 모드 체크포인트가 아님)."}), 400

        sess = new_session()
        sess["project_name"] = project_name
        sess["focus_area"] = ckpt.get("focus_area", "")
        sess["generation_mode"] = "structured_spec"
        # workspace 복원 — per_screen/*.md 가 여기 있어야 resume 효력
        prev_workspace = ckpt.get("workspace")
        if prev_workspace and Path(prev_workspace).exists():
            sess["workspace"] = Path(prev_workspace)
            push_log(sess, f"[resume] workspace 복원: {prev_workspace}")
        sess["selected_domains"] = ckpt.get("selected_domains")
        sess["suite_codes"]      = ckpt.get("suite_codes") or []
        sess["_resumed_state"]   = state

        t = threading.Thread(
            target=run_pipeline_structured,
            args=(sess, ckpt_folder, project_name,
                  ckpt.get("prev_folder", ""),
                  bool(ckpt.get("include_unchanged"))),
            daemon=True,
        )
        sess["thread"] = t
        t.start()
        return jsonify({"ok": True, "sid": sess["id"], "resumed_stage": state["stage"]})

    if not folder_path:
        return jsonify({"ok": False, "error": "folder_path 가 비어 있습니다."}), 400
    folder = Path(folder_path).expanduser()
    if not folder.exists() or not folder.is_dir():
        return jsonify({"ok": False, "error": f"폴더가 없거나 디렉토리가 아닙니다: {folder}"}), 400
    if prev_folder:
        prev_p = Path(prev_folder).expanduser()
        if not prev_p.exists() or not prev_p.is_dir():
            return jsonify({"ok": False, "error": f"이전 폴더가 없거나 디렉토리가 아닙니다: {prev_p}"}), 400
    if not os.environ.get("ANTHROPIC_API_KEY"):
        return jsonify({"ok": False, "error": "ANTHROPIC_API_KEY가 설정되지 않았습니다."}), 500

    # 사전 검증 — 화면 md 가 있어야 함
    cls_check = classify_spec_files(folder)
    if not cls_check["screens"]:
        return jsonify({
            "ok": False,
            "error": f"화면 md 를 찾지 못했습니다. {folder}/scr/SCR-*.md 형식이 필요합니다.",
            "detail": {
                "overview": cls_check["overview"].name if cls_check["overview"] else None,
                "policy": [p.name for p in cls_check["policy"]],
                "design": [p.name for p in cls_check["design"]],
                "screens_count": 0,
            },
        }), 400

    # Combo 모드 검증 — 명세 파일이 실제 존재하는지
    if generation_mode == "combo_only":
        if not combo_opts or not combo_opts.get("file_paths"):
            return jsonify({
                "ok": False,
                "error": "Combo 모드 — *_combinations.md 명세 파일이 선택되지 않았습니다.\n"
                         "TC 생성 범위 위 'Combo 명세' 영역에서 파일을 선택하거나, 모드를 '정책 반영'으로 변경하세요.",
            }), 400
        missing = [fp for fp in combo_opts["file_paths"] if not Path(fp).expanduser().exists()]
        if missing:
            return jsonify({
                "ok": False,
                "error": "Combo 명세 파일을 찾을 수 없습니다:\n" + "\n".join(missing),
            }), 400

    sess = new_session()
    sess["project_name"] = project_name
    sess["focus_area"] = focus_area
    sess["generation_mode"] = generation_mode or "structured_spec"
    if combo_opts:
        sess["combo_opts"] = combo_opts

    sources_to_save = [{"type": "spec_folder", "content": str(folder)}]
    if prev_folder:
        sources_to_save.append({"type": "spec_folder_prev", "content": str(Path(prev_folder).expanduser())})
    # generation_mode + combo_opts 도 저장 — 다음에 같은 프로젝트 선택 시 모드 복원
    save_project(project_name, last_sources=sources_to_save, last_focus_area=focus_area,
                 last_generation_mode=(generation_mode or "summary"),
                 last_combo_opts=(combo_opts or None))

    t = threading.Thread(
        target=run_pipeline_structured,
        args=(sess, str(folder), project_name,
              str(Path(prev_folder).expanduser()) if prev_folder else "",
              include_unchanged),
        daemon=True,
    )
    sess["thread"] = t
    t.start()
    return jsonify({
        "ok": True,
        "sid": sess["id"],
        "summary": {
            "overview": cls_check["overview"].name if cls_check["overview"] else None,
            "policy_count": len(cls_check["policy"]),
            "design_count": len(cls_check["design"]),
            "screens_count": len(cls_check["screens"]),
        },
    })


@app.route("/recent-spec-folders", methods=["GET"])
def recent_spec_folders():
    """프로젝트별 최근 사용 spec 폴더 목록 반환.
    Query: project_name=... (없으면 전체 프로젝트의 최근 폴더 합집합)
    Returns: {ok, folders: [path, ...]} — 존재하는 경로만, 최신 순.
    """
    project_name = request.args.get("project_name", "").strip()
    projects = load_projects()
    candidates: list[str] = []
    if project_name:
        proj = next((p for p in projects if p["name"] == project_name), None)
        if proj:
            candidates = list(proj.get("recent_spec_folders") or [])
        # 폴백 — 프로젝트의 last_sources 에서 spec_folder 추출
        if not candidates and proj:
            for src in (proj.get("last_sources") or []):
                if src.get("type") in ("spec_folder", "spec_folder_prev") and src.get("content"):
                    if src["content"] not in candidates:
                        candidates.append(src["content"])
    else:
        # 프로젝트 미지정 — 모든 프로젝트의 recent_spec_folders 합집합 (최신 5개)
        seen = set()
        for p in projects:
            for f in (p.get("recent_spec_folders") or []):
                if f not in seen:
                    seen.add(f)
                    candidates.append(f)
                if len(candidates) >= 10:
                    break
            if len(candidates) >= 10:
                break

    # 실제 존재하는 폴더만 필터
    folders = [f for f in candidates if Path(f).expanduser().is_dir()]
    return jsonify({"ok": True, "folders": folders[:10]})


# ── Combo TC 엔드포인트 (메인 화면 통합 — 정책+Combo 모드에서 사용) ─────
# 단독 /combo 페이지는 메인 통합 후 제거됨. 아래 list-files / preview / generate /
# download 는 메인 화면 JS 와 향후 단독 사용 가능성 모두를 위해 유지.

@app.route("/combo/list-files", methods=["GET"])
def combo_list_files():
    """프로젝트 + (선택) spec 폴더의 *_combinations.md 파일 목록 반환.
    spec_folder 가 주어지면 그 폴더에도 검색.
    Returns: {ok, files: [{name, path, combos_count, scenarios_count, source}]}
    """
    project_name = (request.args.get("project_name") or "supercycl").strip()
    spec_folder = (request.args.get("spec_folder") or "").strip()
    pname_lower = project_name.lower().replace(" ", "").replace("-", "").replace("_", "")

    candidates: list[tuple[Path, str]] = []  # (folder, source label)

    # 1) spec 폴더 (있으면 우선)
    if spec_folder:
        sp = Path(spec_folder).expanduser()
        if sp.exists() and sp.is_dir():
            candidates.append((sp, "spec"))

    # 2) projects/{name}/ 폴더
    if PROJECTS_RULES_DIR.exists():
        for folder in PROJECTS_RULES_DIR.iterdir():
            if not folder.is_dir():
                continue
            fname_lower = folder.name.lower().replace(" ", "").replace("-", "").replace("_", "")
            if fname_lower in pname_lower or pname_lower in fname_lower:
                candidates.append((folder, "project"))
                break

    if not candidates:
        return jsonify({"ok": False,
                        "error": f"검색 위치 없음 — spec 폴더 또는 projects/{project_name}/ 가 존재해야 합니다."}), 404

    files = []
    seen_names: set[str] = set()
    for folder, source in candidates:
        for md_file in sorted(folder.glob("*_combinations.md")):
            if md_file.name in seen_names:
                continue
            seen_names.add(md_file.name)
            data = _parse_combo_md_lite(md_file)
            files.append({
                "name": md_file.name,
                "path": str(md_file),
                "source": source,
                "combos_count": data["combos_count"],
                "scenarios_count": data["scenarios_count"],
                "steps_count": data["steps_count"],
                "domain": data["domain"],
                "scenarios": data.get("scenarios", []),  # [{id, title, step_count, is_target}]
            })
    return jsonify({"ok": True, "files": files})


def _derive_combo_suite_code(filename: str) -> str:
    """Combo 명세 파일명 → 짧은 SuiteCode (TC ID prefix 충돌 방지).
    예: order_combinations.md → ORDR
        trade_lite__combinations.md → LITE
        exchange_combinations.md → EXCH
        portfolio_combinations.md → PORT
        markets_combinations.md → MKTS
        custom_xyz_combinations.md → XYZ (앞 토큰)
    """
    stem = filename.replace("_combinations.md", "").rstrip("_")
    # 알려진 도메인 매핑 (의미 명확)
    known = {
        "order": "ORDR", "trade": "TRD", "trade_lite": "LITE",
        "exchange": "EXCH", "portfolio": "PORT", "market": "MKTS",
        "markets": "MKTS",
    }
    if stem.lower() in known:
        return known[stem.lower()]
    # 폴백 — 첫 단어의 4자 uppercase
    first = stem.split("_")[0]
    code = re.sub(r"[^A-Za-z0-9]", "", first)[:4].upper()
    return code or "COMBO"


def _parse_combo_md_lite(md_path: Path) -> dict:
    """Combo md 파일에서 통계 + 시나리오 메타 추출 (가볍게).
    UI 의 시나리오 체크박스 표시용으로 시나리오 ID/title/step 수 포함.
    """
    try:
        text = md_path.read_text(encoding="utf-8")
    except Exception:
        return {"combos_count": 0, "scenarios_count": 0, "steps_count": 0,
                "domain": "", "scenarios": []}
    combos = len(re.findall(r"^\|\s*OC-\d+\s*\|", text, re.MULTILINE))

    # 시나리오 ID + title 추출 + 각 시나리오의 step 수
    scenarios_meta = []
    scenario_blocks = re.split(r"(?=^###\s+(?:⭐\s*)?S\d+:)", text, flags=re.MULTILINE)
    for block in scenario_blocks:
        m = re.match(r"^###\s+(⭐\s*)?(S\d+):\s*(.+)", block.strip())
        if not m:
            continue
        scenario_id = m.group(2)
        is_target = bool(m.group(1))
        # title — 첫 줄에서 'S1: 제목' 형태
        title = m.group(3).strip()
        # step 수
        step_count = len(re.findall(rf"^\|\s*{scenario_id}-\d+[a-z]?\s*\|", block, re.MULTILINE))
        scenarios_meta.append({
            "id": scenario_id,
            "title": title,
            "step_count": step_count,
            "is_target": is_target,
        })

    steps = sum(s["step_count"] for s in scenarios_meta)
    # 도메인 추정 (파일명 prefix)
    stem = md_path.stem.replace("_combinations", "")
    domain_map = {"order": "Trade(Combo)", "exchange": "Exchange(Combo)",
                  "portfolio": "Portfolio(Combo)", "market": "Markets(Combo)"}
    domain = domain_map.get(stem.lower(), f"{stem.title()}(Combo)")
    return {"combos_count": combos, "scenarios_count": len(scenarios_meta),
            "steps_count": steps, "domain": domain,
            "scenarios": scenarios_meta}


@app.route("/combo/preview", methods=["POST"])
def combo_preview():
    """선택된 파일의 OC/시나리오 상세 미리보기.
    Body: {file_paths: [...], oc_filter: '', scn_filter: ''}
    Returns: 각 파일별 trim 후 갯수.
    """
    body = request.get_json(force=True) or {}
    file_paths = body.get("file_paths") or []
    oc_filter = (body.get("oc_filter") or "").strip()
    scn_filter = (body.get("scn_filter") or "").strip()

    summary = []
    for fp in file_paths:
        p = Path(fp)
        if not p.exists():
            summary.append({"file": fp, "error": "파일 없음"})
            continue
        # 파일 단독 파싱 (project_name 우회)
        data = _parse_combo_file_direct(p)
        if not data:
            summary.append({"file": p.name, "error": "파싱 실패"})
            continue
        # 필터 적용 (정책: 한쪽만 지정되면 다른 쪽은 0)
        combos = data["combos"]
        scenarios = data["scenarios"]
        is_filter = bool(oc_filter or scn_filter)
        if oc_filter:
            ids = [s.strip().upper() for s in oc_filter.split(",")]
            def _match(c_id):
                for pat in ids:
                    if pat.endswith("*"):
                        if c_id.startswith(pat[:-1]):
                            return True
                    elif c_id == pat:
                        return True
                return False
            combos = [c for c in combos if _match(c["id"])]
        elif is_filter:
            combos = []
        if scn_filter:
            ids = [s.strip().upper() for s in scn_filter.split(",")]
            scenarios = [s for s in scenarios if s["id"].upper() in ids]
        elif is_filter:
            scenarios = []

        steps = sum(len(s["steps"]) for s in scenarios)
        summary.append({
            "file": p.name,
            "domain": _parse_combo_md_lite(p)["domain"],
            "combos_count": len(combos),
            "scenarios_count": len(scenarios),
            "steps_count": steps,
            "total_tc": len(combos) + steps,
        })
    return jsonify({"ok": True, "summary": summary})


def _parse_combo_file_direct(md_path: Path) -> dict | None:
    """파일 경로 직접 받아 OC/시나리오 파싱.
    parse_order_combinations 의 프로젝트 폴더 검색을 건너뛰기 위한 헬퍼.
    """
    if not md_path.exists():
        return None
    text = md_path.read_text(encoding="utf-8")
    result = {"raw_md": text, "combos": [], "scenarios": []}
    current_category = None
    current_table_headers: list[str] = []
    in_section3 = False
    in_section4 = False
    current_scenario = None
    current_step_table_headers: list[str] = []
    for line in text.splitlines():
        stripped = line.strip()
        if stripped.startswith("## 3.") or stripped.startswith("## 3 "):
            in_section3 = True; in_section4 = False
        elif stripped.startswith("## 4.") or stripped.startswith("## 4 "):
            in_section3 = False; in_section4 = True
        elif stripped.startswith("## 5.") or stripped.startswith("## 5 "):
            in_section3 = False; in_section4 = False
        if in_section3 and stripped.startswith("### 3."):
            current_category = re.sub(r"^###\s+3\.\d+\s+", "", stripped)
            current_table_headers = []
        if in_section4 and re.match(r"^###\s+(?:⭐\s*)?S\d+", stripped):
            m = re.match(r"^###\s+(?:⭐\s*)?(S\d+):\s*(.+)$", stripped)
            if m:
                current_scenario = {"id": m.group(1), "title": m.group(2).strip(),
                                    "persona": "", "traits": "", "steps": []}
                result["scenarios"].append(current_scenario)
                current_step_table_headers = []
        if current_scenario and stripped.startswith("**페르소나:**"):
            current_scenario["persona"] = stripped.replace("**페르소나:**", "").strip()
        elif current_scenario and stripped.startswith("**특성:**"):
            current_scenario["traits"] = stripped.replace("**특성:**", "").strip()
        if stripped.startswith("|") and stripped.endswith("|"):
            cells = [c.strip() for c in stripped.strip("|").split("|")]
            if not cells or set(cells[0]) <= set("-"):
                continue
            first = cells[0]
            if in_section3:
                if first.upper().startswith("OC-ID") or first.upper() == "OC-ID":
                    current_table_headers = cells
                elif re.match(r"^OC-\d+$", first, re.IGNORECASE) and current_table_headers:
                    result["combos"].append({"id": first.upper(),
                                             "category": current_category or "(unknown)",
                                             "headers": current_table_headers,
                                             "cells": cells})
            if in_section4 and current_scenario is not None:
                if first.lower() == "step":
                    current_step_table_headers = cells
                elif re.match(r"^S\d+-\d+[a-z]?$", first, re.IGNORECASE) and current_step_table_headers:
                    current_scenario["steps"].append({
                        "step_id": first, "action": cells[1] if len(cells) > 1 else "",
                        "oc_ref": cells[2] if len(cells) > 2 else "",
                        "verify": cells[3] if len(cells) > 3 else ""})
    return result


@app.route("/combo/generate", methods=["POST"])
def combo_generate():
    """LLM 호출 → TC md + Excel 빌드.
    Body: {file_paths, oc_filter, scn_filter, project_name, project_code}
    Returns: {ok, tc_md_path, excel_path, total_tc, smoke_count}
    """
    body = request.get_json(force=True) or {}
    file_paths = body.get("file_paths") or []
    oc_filter = (body.get("oc_filter") or "").strip()
    scn_filter = (body.get("scn_filter") or "").strip()
    project_name = (body.get("project_name") or "supercycl").strip()
    project_code = (body.get("project_code") or "SM").strip().upper()

    if not file_paths:
        return jsonify({"ok": False, "error": "파일을 선택하세요"}), 400
    if not os.environ.get("ANTHROPIC_API_KEY"):
        return jsonify({"ok": False, "error": "ANTHROPIC_API_KEY 미설정"}), 500

    # 작업 폴더
    import uuid
    workspace = Path(__file__).parent.parent / "workspace" / f"combo_{uuid.uuid4().hex[:8]}"
    workspace.mkdir(parents=True, exist_ok=True)

    all_tc_md_parts = []
    all_results = []
    try:
        for fp in file_paths:
            p = Path(fp)
            if not p.exists():
                continue
            data = _parse_combo_file_direct(p)
            if not data:
                continue
            # 필터 적용 + raw_md trim
            combos_orig = data["combos"]
            scenarios_orig = data["scenarios"]
            combos = combos_orig
            scenarios = scenarios_orig
            if oc_filter:
                ids = [s.strip().upper() for s in oc_filter.split(",")]
                def _match(c_id):
                    for pat in ids:
                        if pat.endswith("*"):
                            if c_id.startswith(pat[:-1]):
                                return True
                        elif c_id == pat:
                            return True
                    return False
                combos = [c for c in combos if _match(c["id"])]
            elif scn_filter:
                # scn 만 지정되면 OC 0
                combos = []
            if scn_filter:
                ids = [s.strip().upper() for s in scn_filter.split(",")]
                scenarios = [s for s in scenarios if s["id"].upper() in ids]
            elif oc_filter:
                scenarios = []

            # raw_md trim — Combo 전용 도메인 적용
            domain = _parse_combo_md_lite(p)["domain"]
            is_filter = bool(oc_filter or scn_filter)
            if is_filter:
                new_raw = _rebuild_raw_md_for_combo(data["raw_md"],
                                                    [c["id"] for c in combos],
                                                    [s["id"] for s in scenarios])
            else:
                new_raw = data["raw_md"]

            trimmed = {
                "raw_md": new_raw, "combos": combos, "scenarios": scenarios,
                "_sample_mode": is_filter, "_domain_override": domain,
            }
            # LLM 호출
            tc_md, count = step_write_combo_tc(None, project_name, trimmed,
                                               project_code=project_code)
            # 대분류 일괄 치환 (프롬프트 결과가 'Trade(Combo)' 로 나오는데 도메인이 다르면 교체)
            if domain != "Trade(Combo)":
                tc_md = tc_md.replace("| 대분류 | Trade(Combo) |",
                                       f"| 대분류 | {domain} |")
            all_tc_md_parts.append(tc_md)
            all_results.append({"file": p.name, "domain": domain, "tc_count": count})

        if not all_tc_md_parts:
            return jsonify({"ok": False, "error": "생성된 TC 없음"}), 500

        # 합본 md
        merged_md = "\n\n---\n\n".join(all_tc_md_parts)
        tc_md_path = workspace / "tc_final.md"
        tc_md_path.write_text(merged_md, encoding="utf-8")

        # Excel 빌드
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "build_excel",
            Path(__file__).parent.parent.parent / "tc-agent" / "scripts" / "build_excel.py"
        )
        build_excel_mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(build_excel_mod)
        excel_dir = workspace
        build_excel_mod.run_build("Combo", tc_md_path, excel_dir, verbose=False)
        excel_files = list(excel_dir.glob("*.xlsx"))
        excel_path = excel_files[0] if excel_files else None

        # 결과 통계
        total_tc = sum(r["tc_count"] for r in all_results)
        smoke_count = len(re.findall(r"^###\s+\*\*[A-Z]{2}-", merged_md, re.MULTILINE))

        return jsonify({
            "ok": True,
            "workspace": str(workspace),
            "tc_md_path": str(tc_md_path),
            "excel_path": str(excel_path) if excel_path else None,
            "excel_filename": excel_path.name if excel_path else None,
            "total_tc": total_tc,
            "smoke_count": smoke_count,
            "results": all_results,
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e),
                        "trace": traceback.format_exc()[-2000:]}), 500


def _rebuild_raw_md_for_combo(raw_md: str, kept_oc_ids: list[str],
                              kept_scn_ids: list[str]) -> str:
    """raw_md 를 재구성 — 필터된 OC 행과 시나리오 블록만 남김."""
    kept_oc_set = set(kept_oc_ids)
    kept_scn_set = set(kept_scn_ids)
    out_lines = []
    lines = raw_md.splitlines()
    in_section4 = False
    current_scn_keep = True
    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()
        if stripped.startswith("## 4.") or stripped.startswith("## 4 "):
            in_section4 = True
            out_lines.append(line); i += 1; continue
        if stripped.startswith("## 5.") or stripped.startswith("## 5 "):
            in_section4 = False
            current_scn_keep = True
        if in_section4:
            m_scn = re.match(r"^###\s+(?:⭐\s*)?(S\d+):", stripped)
            if m_scn:
                current_scn_keep = m_scn.group(1) in kept_scn_set
            elif stripped.startswith("### "):
                current_scn_keep = True
            if not current_scn_keep:
                i += 1; continue
        if stripped.startswith("|") and stripped.endswith("|"):
            cells = [c.strip() for c in stripped.strip("|").split("|")]
            if cells:
                first = cells[0].upper()
                if re.match(r"^OC-\d+$", first):
                    if first not in kept_oc_set:
                        i += 1; continue
        out_lines.append(line); i += 1
    return "\n".join(out_lines)


@app.route("/combo/draft-spec", methods=["POST"])
def combo_draft_spec():
    """AI 가 spec 폴더를 분석해 Combo 명세 초안 작성.
    Body: {spec_folder, domain, target_scrs?, existing_md?}
    Returns: {ok, draft_md}
    """
    body = request.get_json(force=True) or {}
    spec_folder = (body.get("spec_folder") or "").strip()
    domain = (body.get("domain") or "Trade").strip()
    target_scrs = body.get("target_scrs") or None
    existing_md = body.get("existing_md") or ""

    if not spec_folder:
        return jsonify({"ok": False, "error": "spec_folder 가 필요합니다."}), 400
    if not Path(spec_folder).expanduser().exists():
        return jsonify({"ok": False, "error": f"spec 폴더 없음: {spec_folder}"}), 400
    if not os.environ.get("ANTHROPIC_API_KEY"):
        return jsonify({"ok": False, "error": "ANTHROPIC_API_KEY 미설정"}), 500
    try:
        draft, meta = step_draft_combo_spec(spec_folder, domain, target_scrs, existing_md)
        return jsonify({"ok": True, "draft_md": draft, "domain": domain, "meta": meta})
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e),
                        "trace": traceback.format_exc()[-1500:]}), 500


@app.route("/combo/save-spec", methods=["POST"])
def combo_save_spec():
    """사용자가 검토/수정한 명세를 파일로 저장.
    Body: {save_path, content}
    save_path 는 안전상 spec 폴더 또는 projects/{name}/ 안만 허용.
    Returns: {ok, saved_path}
    """
    body = request.get_json(force=True) or {}
    save_path = (body.get("save_path") or "").strip()
    content = body.get("content") or ""

    if not save_path or not content:
        return jsonify({"ok": False, "error": "save_path 와 content 필요"}), 400
    p = Path(save_path).expanduser().resolve()

    # 보안 — 허용 위치 검증
    allowed_roots = []
    if PROJECTS_RULES_DIR.exists():
        allowed_roots.append(PROJECTS_RULES_DIR.resolve())
    # spec 폴더 — projects.json 의 last_sources 에서 spec_folder 추출
    try:
        for proj in load_projects():
            for src in (proj.get("last_sources") or []):
                if src.get("type") in ("spec_folder", "spec_folder_prev"):
                    sp = Path(src["content"]).expanduser().resolve()
                    if sp.exists():
                        allowed_roots.append(sp)
    except Exception:
        pass
    if not any(str(p).startswith(str(r)) for r in allowed_roots):
        return jsonify({"ok": False,
                        "error": "저장 위치는 spec 폴더 또는 projects/ 안만 허용됩니다.",
                        "allowed": [str(r) for r in allowed_roots]}), 403

    # 파일명 검증
    if not p.name.endswith("_combinations.md"):
        return jsonify({"ok": False, "error": "파일명은 *_combinations.md 형식이어야 합니다."}), 400

    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(content, encoding="utf-8")
    return jsonify({"ok": True, "saved_path": str(p), "size": len(content)})


@app.route("/combo/read-spec", methods=["GET"])
def combo_read_spec():
    """기존 명세 파일 내용 읽기 (편집 모드 시작용).
    Query: path
    Returns: {ok, content}
    """
    fp = (request.args.get("path") or "").strip()
    if not fp:
        return jsonify({"ok": False, "error": "path 필요"}), 400
    p = Path(fp).expanduser()
    if not p.exists() or not p.is_file():
        return jsonify({"ok": False, "error": f"파일 없음: {fp}"}), 404
    return jsonify({"ok": True, "content": p.read_text(encoding="utf-8")})


@app.route("/combo/delete-spec", methods=["POST"])
def combo_delete_spec():
    """Combo 명세 파일 삭제. 보안 — spec 폴더 또는 projects/ 안만 허용.
    Body: {path}
    Returns: {ok, deleted_path}
    """
    body = request.get_json(force=True) or {}
    target_path = (body.get("path") or "").strip()
    if not target_path:
        return jsonify({"ok": False, "error": "path 필요"}), 400
    p = Path(target_path).expanduser().resolve()
    if not p.exists() or not p.is_file():
        return jsonify({"ok": False, "error": f"파일 없음: {target_path}"}), 404
    if not p.name.endswith("_combinations.md"):
        return jsonify({"ok": False, "error": "*_combinations.md 파일만 삭제 가능합니다."}), 400

    # 허용 위치 검증 (save-spec 과 동일 로직)
    allowed_roots = []
    if PROJECTS_RULES_DIR.exists():
        allowed_roots.append(PROJECTS_RULES_DIR.resolve())
    try:
        for proj in load_projects():
            for src in (proj.get("last_sources") or []):
                if src.get("type") in ("spec_folder", "spec_folder_prev"):
                    sp = Path(src["content"]).expanduser().resolve()
                    if sp.exists():
                        allowed_roots.append(sp)
    except Exception:
        pass
    if not any(str(p).startswith(str(r)) for r in allowed_roots):
        return jsonify({"ok": False,
                        "error": "삭제 위치는 spec 폴더 또는 projects/ 안만 허용됩니다."}), 403

    # 안전 — 휴지통 효과: 같은 위치에 .deleted_<timestamp> 백업 후 삭제
    import time
    backup = p.with_suffix(p.suffix + f".deleted_{int(time.time())}")
    try:
        p.rename(backup)
        return jsonify({"ok": True, "deleted_path": str(p), "backup": str(backup),
                        "message": "파일이 백업되어 삭제되었습니다 (." + backup.name.split('.', 1)[-1] + ")."})
    except Exception as e:
        return jsonify({"ok": False, "error": f"삭제 실패: {e}"}), 500


# ── 기존 TC 업데이트 endpoint (v0.11.x — 1단계: spec diff 분석만) ─────────
@app.route("/update/config", methods=["GET"])
def update_config_get():
    """현재 사용자의 TC update 설정 조회 (Drive 폴더 ID 등).
    Returns: {ok, drive_folder_id, drive_folder_url, drive_folder_name?}
    """
    cfg = load_tc_update_config()
    folder_id = cfg.get("drive_folder_id", "")
    folder_url = cfg.get("drive_folder_url", "")
    folder_name = ""
    # 폴더 이름도 같이 조회 (UI 에 표시용) — 권한 있으면
    if folder_id:
        try:
            drive = get_drive_service()
            meta = drive.files().get(fileId=folder_id, fields="id,name").execute()
            folder_name = meta.get("name", "")
        except Exception:
            folder_name = "(접근 권한 없음 또는 폴더 없음)"
    return jsonify({
        "ok": True,
        "drive_folder_id": folder_id,
        "drive_folder_url": folder_url,
        "drive_folder_name": folder_name,
    })


@app.route("/update/config", methods=["POST"])
def update_config_set():
    """TC update 설정 저장.
    Body: {drive_folder_url}  (URL 또는 폴더 ID — 자동 추출)
    Returns: {ok, drive_folder_id, drive_folder_url, drive_folder_name?, ...}
    """
    body = request.get_json(force=True) or {}
    raw = (body.get("drive_folder_url") or "").strip()
    if not raw:
        # 빈 값 = 설정 제거 (orphan 으로 돌아감)
        save_tc_update_config({})
        return jsonify({"ok": True, "cleared": True})

    folder_id = _extract_folder_id(raw)
    if not folder_id:
        return jsonify({"ok": False,
                          "error": "Drive 폴더 URL 형식이 아닙니다. 예: https://drive.google.com/drive/folders/..."}), 400

    # 폴더 권한 점검
    folder_name = ""
    try:
        drive = get_drive_service()
        meta = drive.files().get(fileId=folder_id, fields="id,name,mimeType").execute()
        folder_name = meta.get("name", "")
        if meta.get("mimeType") != "application/vnd.google-apps.folder":
            return jsonify({"ok": False,
                              "error": "지정한 ID 가 폴더가 아닙니다."}), 400
    except Exception as e:
        return jsonify({"ok": False,
                          "error": f"폴더 접근 실패 — 권한 확인 필요: {e}"}), 400

    cfg = {
        "drive_folder_id": folder_id,
        "drive_folder_url": raw,
        "drive_folder_name": folder_name,
    }
    save_tc_update_config(cfg)
    return jsonify({"ok": True, **cfg})


@app.route("/update/history", methods=["GET"])
def update_history_list():
    """프로젝트의 TC update 작업 이력 조회.
    Query: ?project=<name>
    """
    project_name = (request.args.get("project") or "").strip()
    if not project_name:
        return jsonify({"ok": True, "history": []})
    return jsonify({"ok": True, "history": get_update_history(project_name)})


@app.route("/update/history/delete", methods=["POST"])
def update_history_delete():
    """프로젝트의 update_history 에서 특정 항목 삭제.
    Body: {project, index}
    """
    body = request.get_json(force=True) or {}
    project_name = (body.get("project") or "").strip()
    index = body.get("index")
    if not project_name:
        return jsonify({"ok": False, "error": "project 필요"}), 400
    if not isinstance(index, int):
        return jsonify({"ok": False, "error": "index (정수) 필요"}), 400
    ok = delete_update_history(project_name, index)
    if not ok:
        return jsonify({"ok": False, "error": "프로젝트 없음 또는 index 범위 초과"}), 404
    return jsonify({"ok": True, "history": get_update_history(project_name)})


@app.route("/update/analyze", methods=["POST"])
def update_analyze():
    """기획서 두 버전을 비교해 변경 분석 보고서 생성.
    Body: {prev_folder, new_folder, existing_tc_url?, project_name?}
    Returns: {ok, diff, summary, scr_changes, common_doc_changes, meta}
    """
    body = request.get_json(force=True) or {}
    prev_folder = (body.get("prev_folder") or "").strip()
    new_folder = (body.get("new_folder") or "").strip()
    existing_tc_url = (body.get("existing_tc_url") or "").strip()
    project_name = (body.get("project_name") or "").strip()

    if not prev_folder or not new_folder:
        return jsonify({"ok": False,
                        "error": "이전 폴더와 새 폴더 경로 모두 필요합니다."}), 400
    if not os.environ.get("ANTHROPIC_API_KEY"):
        return jsonify({"ok": False, "error": "ANTHROPIC_API_KEY 미설정"}), 500

    try:
        result = step_analyze_spec_diff(prev_folder, new_folder, existing_tc_url)
        # 분석이 성공했으면 history 저장 (Sheets URL 까지 있어야 함)
        if project_name and existing_tc_url:
            try:
                save_update_history(project_name, prev_folder, new_folder, existing_tc_url)
            except Exception:
                pass  # history 저장 실패는 분석 결과 반환을 막지 않음
        return jsonify({"ok": True, **result})
    except Exception as e:
        import traceback
        return jsonify({
            "ok": False, "error": str(e),
            "trace": traceback.format_exc()[-2000:],
        }), 500


@app.route("/update/scr-tcs", methods=["POST"])
def update_scr_tcs():
    """진단용 — 특정 Sheets 의 특정 SCR 에 매핑된 모든 TC 본문 반환.
    Body: {sheets_url, scr_id, sheet_title? (선택)}
    """
    body = request.get_json(force=True) or {}
    url = (body.get("sheets_url") or "").strip()
    scr_id = (body.get("scr_id") or "").strip()
    filter_sheet = (body.get("sheet_title") or "").strip()
    if not url or not scr_id:
        return jsonify({"ok": False, "error": "sheets_url 과 scr_id 필요"}), 400
    try:
        sid = _extract_sheets_id(url)
        data = read_sheets_tcs(sid)
        matched = []
        for t in data.get("tcs", []):
            if scr_id not in t.get("scr_ids", []):
                continue
            if filter_sheet and t.get("sheet_title") != filter_sheet:
                continue
            # 본문 읽기
            try:
                body_row = _read_tc_row_from_sheets(sid, t["sheet_title"], t["row_index"])
                matched.append({
                    "tc_id": t["tc_id"],
                    "title": t["title"],
                    "sheet_title": t["sheet_title"],
                    "row_index": t["row_index"],
                    "scr_ids": t["scr_ids"],
                    "precondition": body_row.get("precondition", ""),
                    "steps": body_row.get("steps", ""),
                    "expected": body_row.get("expected", ""),
                })
            except Exception:
                pass
        return jsonify({"ok": True, "scr_id": scr_id, "count": len(matched), "tcs": matched})
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e),
                          "trace": traceback.format_exc()[-1500:]}), 500


@app.route("/update/analyze-plan", methods=["POST"])
def update_analyze_plan():
    """사전 견적 — AI 호출 없이 diff + 그룹화 + 캐시 hit 상태 반환.

    Body: {prev_folder, new_folder}
    Returns: {
      ok, scr_diff_stats, groups, total_groups, cached_indices,
      estimated_minutes_min, estimated_minutes_max
    }
    """
    body = request.get_json(force=True) or {}
    prev_folder = (body.get("prev_folder") or "").strip()
    new_folder = (body.get("new_folder") or "").strip()
    if not prev_folder or not new_folder:
        return jsonify({"ok": False,
                          "error": "이전 폴더와 새 폴더 경로 모두 필요합니다."}), 400

    try:
        prev_p = Path(prev_folder).expanduser()
        new_p = Path(new_folder).expanduser()
        if not prev_p.exists() or not new_p.exists():
            return jsonify({"ok": False, "error": "폴더 경로가 존재하지 않습니다."}), 400

        # diff 만 (AI 호출 없음 — 빠름)
        diff = diff_spec_folders(prev_p, new_p)
        groups = _group_scr_changes(diff)

        # 캐시 hit 상태 점검
        cached_indices = []
        for g in groups:
            cached = _load_diff_cache(prev_folder, new_folder, g["index"])
            if cached:
                cached_indices.append(g["index"])

        # SCR 평면 리스트 — UI 가 체크리스트로 활용 (각 SCR 캐시 상태 포함)
        scrs_flat = []
        for s in (diff.get("added") or []):
            cached = _load_scr_cache(prev_folder, new_folder, s)
            scrs_flat.append({
                "scr_id": s, "type": "added",
                "from_cache": bool(cached),
            })
        for s in (diff.get("modified") or []):
            cached = _load_scr_cache(prev_folder, new_folder, s)
            scrs_flat.append({
                "scr_id": s, "type": "modified",
                "from_cache": bool(cached),
            })
        for s in (diff.get("removed") or []):
            cached = _load_scr_cache(prev_folder, new_folder, s)
            scrs_flat.append({
                "scr_id": s, "type": "removed",
                "from_cache": bool(cached),
            })

        total = len(groups)
        new_count = total - len(cached_indices)
        return jsonify({
            "ok": True,
            "scr_diff_stats": {
                "added": len(diff.get("added", []) or []),
                "modified": len(diff.get("modified", []) or []),
                "removed": len(diff.get("removed", []) or []),
                "unchanged": len(diff.get("unchanged", []) or []),
                "common_changed": diff.get("common_changed", False),
            },
            "groups": groups,
            "total_groups": total,
            "cached_indices": cached_indices,
            "new_count": new_count,
            "estimated_minutes_min": max(1, new_count),       # 그룹당 약 60초
            "estimated_minutes_max": max(1, new_count * 2),    # 그룹당 약 120초
            # SCR 단위 체크리스트용
            "scrs": scrs_flat,
            "total_scrs": len(scrs_flat),
            "cached_scrs": [s["scr_id"] for s in scrs_flat if s["from_cache"]],
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e),
                          "trace": traceback.format_exc()[-2000:]}), 500


@app.route("/update/analyze-scr", methods=["POST"])
def update_analyze_scr():
    """1개 SCR 의 AI 분석. SCR 단위 캐시 사용.

    Body: {prev_folder, new_folder, scr_id, force_refresh?}
    Returns: {ok, scr_id, change_type, summary, from_cache, cached_at?}
    """
    body = request.get_json(force=True) or {}
    prev_folder = (body.get("prev_folder") or "").strip()
    new_folder = (body.get("new_folder") or "").strip()
    scr_id = (body.get("scr_id") or "").strip()
    force_refresh = bool(body.get("force_refresh", False))

    if not prev_folder or not new_folder or not scr_id:
        return jsonify({"ok": False,
                          "error": "prev_folder, new_folder, scr_id 모두 필요"}), 400

    try:
        # SCR 의 change_type 판별 — diff 기준
        diff = diff_spec_folders(Path(prev_folder).expanduser(),
                                    Path(new_folder).expanduser())
        if scr_id in (diff.get("added") or []):
            change_type = "added"
        elif scr_id in (diff.get("modified") or []):
            change_type = "modified"
        elif scr_id in (diff.get("removed") or []):
            change_type = "removed"
        else:
            return jsonify({"ok": False,
                              "error": f"{scr_id} 가 변경 SCR 목록에 없음 (unchanged 또는 존재하지 않음)"}), 400

        # 캐시 확인
        if not force_refresh:
            cached = _load_scr_cache(prev_folder, new_folder, scr_id)
            if cached and cached.get("summary"):
                return jsonify({
                    "ok": True,
                    "scr_id": scr_id,
                    "change_type": change_type,
                    "summary": cached["summary"],
                    "from_cache": True,
                    "cached_at": cached.get("_cached_at", ""),
                })

        if not os.environ.get("ANTHROPIC_API_KEY"):
            return jsonify({"ok": False, "error": "ANTHROPIC_API_KEY 미설정"}), 500

        result = step_analyze_scr_one(prev_folder, new_folder, scr_id, change_type)
        if not result.get("ok"):
            return jsonify(result), 500

        _save_scr_cache(prev_folder, new_folder, scr_id, {
            "summary": result["summary"],
            "change_type": change_type,
        })

        return jsonify({
            "ok": True,
            "scr_id": scr_id,
            "change_type": change_type,
            "summary": result["summary"],
            "from_cache": False,
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e),
                          "trace": traceback.format_exc()[-2000:]}), 500


@app.route("/update/analyze-group", methods=["POST"])
def update_analyze_group():
    """1개 그룹 AI 분석. 캐시 hit 면 즉시 반환.

    Body: {prev_folder, new_folder, group_index, force_refresh?}
    Returns: {ok, group, summary, from_cache, next_group_index, total_groups, is_last}
    """
    body = request.get_json(force=True) or {}
    prev_folder = (body.get("prev_folder") or "").strip()
    new_folder = (body.get("new_folder") or "").strip()
    group_index = body.get("group_index")
    force_refresh = bool(body.get("force_refresh", False))

    if not prev_folder or not new_folder:
        return jsonify({"ok": False,
                          "error": "이전 폴더와 새 폴더 경로 모두 필요합니다."}), 400
    if group_index is None:
        return jsonify({"ok": False, "error": "group_index 필요"}), 400

    try:
        prev_p = Path(prev_folder).expanduser()
        new_p = Path(new_folder).expanduser()
        diff = diff_spec_folders(prev_p, new_p)
        groups = _group_scr_changes(diff)
        if group_index < 0 or group_index >= len(groups):
            return jsonify({"ok": False,
                              "error": f"group_index 범위 초과 (0~{len(groups)-1})"}), 400

        target = groups[group_index]

        # 캐시 점검
        from_cache = False
        if not force_refresh:
            cached = _load_diff_cache(prev_folder, new_folder, group_index)
            if cached and cached.get("summary"):
                from_cache = True
                return jsonify({
                    "ok": True,
                    "group": target,
                    "summary": cached["summary"],
                    "from_cache": True,
                    "cached_at": cached.get("_cached_at", ""),
                    "next_group_index": group_index + 1 if group_index + 1 < len(groups) else None,
                    "total_groups": len(groups),
                    "is_last": (group_index + 1) >= len(groups),
                })

        # AI 호출 (캐시 miss 또는 force_refresh)
        if not os.environ.get("ANTHROPIC_API_KEY"):
            return jsonify({"ok": False, "error": "ANTHROPIC_API_KEY 미설정"}), 500

        result = step_analyze_group(prev_folder, new_folder, target)
        if not result.get("ok"):
            return jsonify(result), 500

        # 캐시 저장
        _save_diff_cache(prev_folder, new_folder, group_index, {
            "summary": result["summary"],
            "group": target,
        })

        return jsonify({
            "ok": True,
            "group": target,
            "summary": result["summary"],
            "from_cache": False,
            "next_group_index": group_index + 1 if group_index + 1 < len(groups) else None,
            "total_groups": len(groups),
            "is_last": (group_index + 1) >= len(groups),
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e),
                          "trace": traceback.format_exc()[-2000:]}), 500


# 세션 캐시 — plan 결과를 apply 에서 재사용
_UPDATE_SESSIONS = {}


@app.route("/update/inspect", methods=["POST"])
def update_inspect():
    """사용자가 가져온 Sheets 의 실제 구조를 진단.
    Body: {existing_tc_url}
    Returns: {ok, file_meta, tabs:[{title, header, sample_rows, detected_tc_count, ...}], parser_result}
    """
    body = request.get_json(force=True) or {}
    url = (body.get("existing_tc_url") or "").strip()
    if not url:
        return jsonify({"ok": False, "error": "existing_tc_url 필요"}), 400

    try:
        sheets_id = _extract_sheets_id(url)
        drive = get_drive_service()
        meta = drive.files().get(fileId=sheets_id,
                                   fields="id,name,mimeType,modifiedTime,owners(displayName,emailAddress)").execute()

        svc = get_sheets_service()
        ss = svc.spreadsheets().get(spreadsheetId=sheets_id,
                                       fields="sheets(properties(sheetId,title,index,gridProperties))").execute()
        tab_titles = [s["properties"]["title"] for s in ss.get("sheets", [])]

        ranges = [f"'{t}'!A1:Z2000" for t in tab_titles]
        batch = svc.spreadsheets().values().batchGet(spreadsheetId=sheets_id, ranges=ranges).execute()
        value_ranges = batch.get("valueRanges", [])

        tabs_report = []
        for title, vr in zip(tab_titles, value_ranges):
            values = vr.get("values", [])
            # 헤더 후보 탐색 (TC ID 또는 ID 컬럼 포함하는 첫 행)
            header_idx = -1
            header = []
            for i, row in enumerate(values[:10]):
                joined = " ".join(str(c) for c in row)
                if "TC ID" in joined or "TC_ID" in joined or "ID" in joined:
                    header = row
                    header_idx = i
                    break
            if header_idx < 0 and values:
                header = values[0]
                header_idx = 0

            sample_rows = values[header_idx + 1:header_idx + 6] if header_idx >= 0 else []

            # 본문에서 SCR 패턴 탐지
            all_text = "\n".join(" ".join(str(c) for c in row) for row in values[header_idx + 1:] if any(str(c).strip() for c in row))
            scrs = sorted(set(re.findall(r"SCR-\d+", all_text)), key=lambda s: int(s.split("-")[1]))
            tc_ids = sorted(set(re.findall(r"\b[A-Z]{2,}-[A-Z0-9]+(?:-[A-Z0-9]+)*-\d{2,}\b", all_text)))

            # 데이터 행 수 (헤더 이후 비어있지 않은 행)
            data_rows = [r for r in values[header_idx + 1:] if any(str(c).strip() for c in r)]

            tabs_report.append({
                "title": title,
                "header_row_index": header_idx,
                "header": header,
                "header_count": len(header),
                "sample_rows": sample_rows,
                "data_row_count": len(data_rows),
                "detected_scr_ids": scrs[:30],
                "detected_scr_count": len(scrs),
                "detected_tc_ids_sample": tc_ids[:10],
                "detected_tc_count": len(tc_ids),
            })

        # 현재 read_sheets_tcs 가 어떻게 파싱하는지 — 실제 호출 결과
        try:
            parsed = read_sheets_tcs(sheets_id)
            parser_result = {
                "tcs_total": len(parsed.get("tcs", [])),
                "scr_keys_total": len(parsed.get("tc_by_scr", {})),
                "tabs_with_tcs": [{"title": t["title"], "tc_count": len(t.get("rows", []))} for t in parsed.get("tabs", [])],
                "sample_tcs": parsed.get("tcs", [])[:5],
                "sample_tc_by_scr": dict(list(parsed.get("tc_by_scr", {}).items())[:5]),
            }
        except Exception as ex:
            parser_result = {"error": str(ex)}

        return jsonify({
            "ok": True,
            "file_meta": meta,
            "tabs": tabs_report,
            "parser_result": parser_result,
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e),
                          "trace": traceback.format_exc()[-2500:]}), 500


@app.route("/update/plan", methods=["POST"])
def update_plan():
    """2단계 — 원본 Sheets 사본 생성 + TC↔SCR 매핑 + 4-way 후보 리스트.
    Body: {prev_folder, new_folder, existing_tc_url, make_copy?(default True), project_name?}
    Returns: {ok, session_id, copy_url, copy_id, candidates, stats, diff}
    """
    body = request.get_json(force=True) or {}
    prev_folder = (body.get("prev_folder") or "").strip()
    new_folder = (body.get("new_folder") or "").strip()
    existing_tc_url = (body.get("existing_tc_url") or "").strip()
    make_copy = bool(body.get("make_copy", True))
    project_name = (body.get("project_name") or "").strip()
    # 선택된 SCR 만 후보로 (None 이면 전체)
    scr_filter = body.get("scr_filter")
    if scr_filter is not None and not isinstance(scr_filter, list):
        scr_filter = None
    if scr_filter:
        scr_filter = [str(s).strip() for s in scr_filter if str(s).strip()]
        if not scr_filter:
            scr_filter = None
    # 제외할 action 들 — 'add' 등. 신규 SCR 은 신규 TC 생성 모드에서 처리하므로 기본 제외 권장
    exclude_actions = body.get("exclude_actions") or []
    if not isinstance(exclude_actions, list):
        exclude_actions = []
    exclude_actions = set(str(a).strip() for a in exclude_actions if str(a).strip())

    if not prev_folder or not new_folder or not existing_tc_url:
        return jsonify({"ok": False,
                        "error": "이전 폴더, 새 폴더, 기존 TC Sheets URL 모두 필요합니다."}), 400

    # 입력이 다 갖춰진 시점에 history 저장 (Plan 호출 자체가 실제 사용 의도)
    if project_name:
        try:
            save_update_history(project_name, prev_folder, new_folder, existing_tc_url)
        except Exception:
            pass

    try:
        # 1) Spec diff (1단계 재사용 — AI 요약은 생략, diff 만 필요)
        diff_only = step_analyze_spec_diff(prev_folder, new_folder, existing_tc_url)

        # 2) 사본 생성 (실험 단계 — 원본 보호)
        copy_info = {}
        target_sheets_id = _extract_sheets_id(existing_tc_url)
        if make_copy:
            copy_info = copy_sheets_for_update(existing_tc_url)
            target_sheets_id = copy_info["copy_id"]

        # 3) TC↔SCR 매핑 추출 (사본 기준 — 같은 내용이지만 일관성)
        tc_data = read_sheets_tcs(target_sheets_id)

        # 4) 판정 로직 → 4-way 후보 (선택된 SCR 만)
        candidates = build_update_candidates(
            diff_only, tc_data, prev_folder, new_folder,
            scr_filter=scr_filter,
        )
        # 4-1) 제외 action 적용 — 'add' 등 (신규는 신규 TC 생성 모드에서 별도 처리)
        if exclude_actions:
            candidates = [c for c in candidates if c.get("action") not in exclude_actions]

        # 5) 세션 저장 (apply 단계에서 재사용)
        import uuid
        session_id = f"upd_{uuid.uuid4().hex[:12]}"
        _UPDATE_SESSIONS[session_id] = {
            "candidates": candidates,
            "target_sheets_id": target_sheets_id,
            "copy_info": copy_info,
            "diff": diff_only.get("diff", {}),
            "prev_folder": prev_folder,
            "new_folder": new_folder,
            "scr_filter": scr_filter,
        }

        # 통계
        stats = {
            "total": len(candidates),
            "add": sum(1 for c in candidates if c["action"] == "add"),
            "modify": sum(1 for c in candidates if c["action"] == "modify"),
            "delete": sum(1 for c in candidates if c["action"] == "delete"),
            "default_checked": sum(1 for c in candidates if c.get("default_checked")),
            "scr_added": len(diff_only["diff"].get("added", [])),
            "scr_modified": len(diff_only["diff"].get("modified", [])),
            "scr_removed": len(diff_only["diff"].get("removed", [])),
            "scr_unchanged": len(diff_only["diff"].get("unchanged", [])),
            "tcs_total": len(tc_data.get("tcs", [])),
            "tabs": [t["title"] for t in tc_data.get("tabs", [])],
        }

        return jsonify({
            "ok": True,
            "session_id": session_id,
            "copy_url": copy_info.get("copy_url", ""),
            "copy_title": copy_info.get("copy_title", ""),
            "target_sheets_id": target_sheets_id,
            "candidates": candidates,
            "stats": stats,
        })
    except Exception as e:
        import traceback
        return jsonify({
            "ok": False, "error": str(e),
            "trace": traceback.format_exc()[-2500:],
        }), 500


@app.route("/update/preview", methods=["POST"])
def update_preview():
    """후보 1건의 실제 SCR 본문 + diff 반환.

    Body: {session_id, idx}
    Returns: {
      ok, candidate, prev_text, new_text, unified_diff,
      changed_lines: [{type:'add'|'del', text}],
    }
    """
    body = request.get_json(force=True) or {}
    session_id = (body.get("session_id") or "").strip()
    idx = body.get("idx")
    sess = _UPDATE_SESSIONS.get(session_id)
    if not sess:
        return jsonify({"ok": False, "error": "세션이 없습니다. /update/plan 을 다시 호출하세요."}), 404
    candidates = sess.get("candidates", [])
    if idx is None or idx < 0 or idx >= len(candidates):
        return jsonify({"ok": False, "error": f"잘못된 idx: {idx}"}), 400

    c = candidates[idx]
    scr_id = c.get("scr_id") or ""

    try:
        prev_folder = sess.get("prev_folder", "")
        new_folder = sess.get("new_folder", "")
        prev_cls = classify_spec_files(Path(prev_folder).expanduser()) if prev_folder else {"screens": []}
        new_cls = classify_spec_files(Path(new_folder).expanduser()) if new_folder else {"screens": []}
        prev_map = {p.stem.split(".")[0]: p for p in prev_cls.get("screens", [])}
        new_map = {p.stem.split(".")[0]: p for p in new_cls.get("screens", [])}
        prev_path = prev_map.get(scr_id)
        new_path = new_map.get(scr_id)

        prev_text = prev_path.read_text(encoding="utf-8") if prev_path and prev_path.exists() else ""
        new_text = new_path.read_text(encoding="utf-8") if new_path and new_path.exists() else ""

        # unified diff
        import difflib
        prev_lines = prev_text.splitlines()
        new_lines = new_text.splitlines()
        unified = list(difflib.unified_diff(prev_lines, new_lines,
                                                fromfile=f"prev/{scr_id}",
                                                tofile=f"new/{scr_id}",
                                                n=3, lineterm=""))
        # 변경 라인만 따로 (휴리스틱이 본 줄들)
        changed_lines = []
        for ln in unified:
            if ln.startswith("+++") or ln.startswith("---") or ln.startswith("@@"):
                continue
            if ln.startswith("+"):
                changed_lines.append({"type": "add", "text": ln[1:]})
            elif ln.startswith("-"):
                changed_lines.append({"type": "del", "text": ln[1:]})

        return jsonify({
            "ok": True,
            "candidate": c,
            "scr_id": scr_id,
            "has_prev": bool(prev_text),
            "has_new": bool(new_text),
            "prev_text": prev_text[:8000],
            "new_text": new_text[:8000],
            "unified_diff": "\n".join(unified)[:10000],
            "changed_lines": changed_lines[:200],
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e),
                          "trace": traceback.format_exc()[-2000:]}), 500


def _read_tc_row_from_sheets(sheets_id: str, sheet_title: str, row_index: int) -> dict:
    """사본 Sheets 에서 특정 행을 읽어 {tc_id, title, steps, expected, precondition, ...} 반환.

    헤더 매핑은 read_sheets_tcs 와 동일한 별칭 규칙 사용.
    """
    svc = get_sheets_service()
    # 헤더 + 해당 행만
    header_range = f"'{sheet_title}'!A1:Z3"
    row_range = f"'{sheet_title}'!A{row_index}:Z{row_index}"
    batch = svc.spreadsheets().values().batchGet(
        spreadsheetId=sheets_id, ranges=[header_range, row_range]
    ).execute()
    vrs = batch.get("valueRanges", [])
    if len(vrs) < 2:
        raise RuntimeError("시트에서 행을 읽지 못함")

    header_values = vrs[0].get("values", [])
    row_values = vrs[1].get("values", [[]])
    row = row_values[0] if row_values else []

    # 헤더 행 (TC ID 컬럼 포함) 찾기
    header = []
    for r in header_values[:3]:
        if any(c in ("TC ID", "TC_ID") for c in r):
            header = r
            break
    if not header and header_values:
        header = header_values[0]

    def col(names):
        for nm in names:
            if nm in header:
                return header.index(nm)
        return -1

    def cell(ci):
        if ci < 0 or ci >= len(row):
            return ""
        return str(row[ci] or "").strip()

    return {
        "tc_id": cell(col(["TC ID", "TC_ID", "ID"])).replace("*", "").strip(),
        "title": cell(col(["소분류", "제목", "Title"])),
        "category": cell(col(["대분류"])),
        "sub_category": cell(col(["중분류"])),
        "precondition": cell(col(["사전조건", "사전 조건", "Precondition"])),
        "steps": cell(col(["테스트 스텝", "테스트 단계", "Steps"])),
        "expected": cell(col(["기대결과", "기대 결과", "예상 결과", "Expected"])),
        "scr_field": cell(col(["화면 코드", "화면코드", "SCR", "SCR ID"])),
        "_header": header,
        "_col_pre": col(["사전조건", "사전 조건", "Precondition"]),
        "_col_steps": col(["테스트 스텝", "테스트 단계", "Steps"]),
        "_col_exp": col(["기대결과", "기대 결과", "예상 결과", "Expected"]),
    }


@app.route("/update/propose", methods=["POST"])
def update_propose():
    """1건의 TC 에 대한 AI 수정안 제안.

    Body: {session_id, idx}
    Returns: {ok, current_tc, proposal, scr_id}
    """
    body = request.get_json(force=True) or {}
    session_id = (body.get("session_id") or "").strip()
    idx = body.get("idx")
    sess = _UPDATE_SESSIONS.get(session_id)
    if not sess:
        return jsonify({"ok": False, "error": "세션이 없습니다. /update/plan 을 다시 호출하세요."}), 404
    candidates = sess.get("candidates", [])
    if idx is None or idx < 0 or idx >= len(candidates):
        return jsonify({"ok": False, "error": f"잘못된 idx: {idx}"}), 400

    c = candidates[idx]
    if c.get("action") != "modify":
        return jsonify({"ok": False,
                          "error": f"수정안 제안은 modify 액션만 지원 (현재: {c.get('action')})"}), 400
    if not c.get("tc_id"):
        return jsonify({"ok": False, "error": "TC ID 가 없는 후보 — 수정 불가"}), 400
    if not c.get("sheet_title"):
        return jsonify({"ok": False, "error": "시트 정보 없음"}), 400

    scr_id = c.get("scr_id") or ""
    target_id = sess.get("target_sheets_id")

    try:
        # 1) 사본에서 현재 TC 행 읽기 (TC ID 로 row_index 찾기)
        #    Quota 절약: 세션에 tc_data 캐시 (한 세션 내 propose 여러 번 호출되어도 1회만 읽음)
        tc_data = sess.get("_tc_data_cache")
        if not tc_data:
            tc_data = read_sheets_tcs(target_id)
            sess["_tc_data_cache"] = tc_data
            # tc_id+sheet → row_index 빠른 조회 인덱스
            sess["_tc_locator"] = {(t["tc_id"], t["sheet_title"]): t["row_index"]
                                      for t in tc_data.get("tcs", [])}

        locator = sess.get("_tc_locator") or {}
        row_index = locator.get((c["tc_id"], c["sheet_title"]))
        if row_index is None:
            return jsonify({"ok": False,
                              "error": f"사본에서 {c['tc_id']} 를 찾지 못함"}), 404

        # 본문까지 읽기 (이건 TC 마다 한 번 — 같은 TC 반복 propose 면 캐시)
        body_cache_key = (c["tc_id"], c["sheet_title"], row_index)
        body_cache = sess.setdefault("_tc_body_cache", {})
        if body_cache_key in body_cache:
            current_tc = body_cache[body_cache_key]
        else:
            current_tc = _read_tc_row_from_sheets(target_id, c["sheet_title"], row_index)
            current_tc["row_index"] = row_index
            body_cache[body_cache_key] = current_tc

        # 2) SCR 본문 (prev/new) 가져오기
        prev_folder = sess.get("prev_folder", "")
        new_folder = sess.get("new_folder", "")
        prev_cls = classify_spec_files(Path(prev_folder).expanduser()) if prev_folder else {"screens": []}
        new_cls = classify_spec_files(Path(new_folder).expanduser()) if new_folder else {"screens": []}
        prev_map = {p.stem.split(".")[0]: p for p in prev_cls.get("screens", [])}
        new_map = {p.stem.split(".")[0]: p for p in new_cls.get("screens", [])}
        prev_path = prev_map.get(scr_id)
        new_path = new_map.get(scr_id)
        prev_text = prev_path.read_text(encoding="utf-8") if prev_path and prev_path.exists() else ""
        new_text = new_path.read_text(encoding="utf-8") if new_path and new_path.exists() else ""

        # 3) AI 호출
        result = step_propose_tc_update(scr_id, prev_text, new_text, current_tc)
        if not result.get("ok"):
            return jsonify(result), 500

        # 4) 응답에 현재값도 같이 (diff UI 용)
        # _col_* 키는 내부용이라 응답에서 제거
        clean_tc = {k: v for k, v in current_tc.items() if not k.startswith("_")}

        return jsonify({
            "ok": True,
            "scr_id": scr_id,
            "current_tc": clean_tc,
            "proposal": result["proposal"],
        })
    except Exception as e:
        import traceback
        err_msg = str(e)
        # 429 quota 초과면 명확한 한글 메시지
        if "429" in err_msg or "Quota exceeded" in err_msg or "RATE_LIMIT" in err_msg:
            return jsonify({
                "ok": False,
                "error": "Google Sheets API 한도 초과 (분당 60회) — 1-2분 기다린 후 다시 시도하세요. 일괄 적용 직후나 여러 카드를 빠르게 열면 자주 발생합니다.",
                "rate_limited": True,
            }), 429
        return jsonify({"ok": False, "error": err_msg,
                          "trace": traceback.format_exc()[-2000:]}), 500


@app.route("/update/commit", methods=["POST"])
def update_commit():
    """승인된 수정안을 사본 Sheets 의 행 셀에 직접 쓰기.

    Body: {session_id, idx, accepted: {steps?, expected?, precondition?}}
    Returns: {ok, written_cells, log_entry}
    """
    body = request.get_json(force=True) or {}
    session_id = (body.get("session_id") or "").strip()
    idx = body.get("idx")
    accepted = body.get("accepted") or {}
    sess = _UPDATE_SESSIONS.get(session_id)
    if not sess:
        return jsonify({"ok": False, "error": "세션이 없습니다."}), 404
    candidates = sess.get("candidates", [])
    if idx is None or idx < 0 or idx >= len(candidates):
        return jsonify({"ok": False, "error": f"잘못된 idx: {idx}"}), 400

    c = candidates[idx]
    if c.get("action") != "modify" or not c.get("tc_id"):
        return jsonify({"ok": False, "error": "modify + tc_id 필요"}), 400

    target_id = sess.get("target_sheets_id")
    sheet_title = c.get("sheet_title")

    try:
        # 1) row_index + 컬럼 인덱스 다시 확인
        tc_data = read_sheets_tcs(target_id)
        target_tc = next((t for t in tc_data.get("tcs", [])
                              if t["tc_id"] == c["tc_id"]
                              and t["sheet_title"] == c["sheet_title"]), None)
        if not target_tc:
            return jsonify({"ok": False, "error": f"{c['tc_id']} 를 사본에서 못 찾음"}), 404

        current_tc = _read_tc_row_from_sheets(target_id, sheet_title, target_tc["row_index"])

        # 2) 적용할 필드 + 컬럼 인덱스 매핑
        field_to_col = {
            "precondition": current_tc["_col_pre"],
            "steps": current_tc["_col_steps"],
            "expected": current_tc["_col_exp"],
        }

        # 3) 각 필드별로 단일 셀 update
        svc = get_sheets_service()
        written = []
        for field, new_value in accepted.items():
            if new_value is None:
                continue
            ci = field_to_col.get(field, -1)
            if ci < 0:
                continue
            # A1 표기 — 컬럼 letter
            col_letter = ""
            n = ci
            while True:
                col_letter = chr(ord('A') + (n % 26)) + col_letter
                n = n // 26 - 1
                if n < 0:
                    break
            cell_range = f"'{sheet_title}'!{col_letter}{target_tc['row_index']}"
            svc.spreadsheets().values().update(
                spreadsheetId=target_id,
                range=cell_range,
                valueInputOption="RAW",
                body={"values": [[str(new_value)]]},
            ).execute()
            written.append({
                "field": field,
                "cell": cell_range,
                "before": current_tc.get(field, ""),
                "after": new_value,
            })

        # 4) Update Log 시트에 기록 (있으면 append, 없으면 생성)
        from datetime import datetime
        log_title = "TC Edit Log"
        try:
            svc.spreadsheets().batchUpdate(
                spreadsheetId=target_id,
                body={"requests": [{"addSheet": {"properties": {"title": log_title}}}]}
            ).execute()
            # 헤더 추가
            svc.spreadsheets().values().update(
                spreadsheetId=target_id,
                range=f"'{log_title}'!A1",
                valueInputOption="RAW",
                body={"values": [["시각", "TC ID", "시트", "행", "필드", "이전 값(요약)", "새 값(요약)", "수정 사유"]]},
            ).execute()
        except Exception:
            pass  # 이미 있으면 무시

        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        rationale = (accepted.get("_rationale") or "").strip()
        log_rows = []
        for w in written:
            log_rows.append([
                ts, c["tc_id"], sheet_title, str(target_tc["row_index"]),
                w["field"],
                (w["before"] or "")[:200],
                (w["after"] or "")[:200],
                rationale[:300],
            ])
        if log_rows:
            svc.spreadsheets().values().append(
                spreadsheetId=target_id,
                range=f"'{log_title}'!A1",
                valueInputOption="RAW",
                insertDataOption="INSERT_ROWS",
                body={"values": log_rows},
            ).execute()

        return jsonify({
            "ok": True,
            "written_cells": written,
            "log_sheet": log_title,
            "tc_id": c["tc_id"],
            "sheet_title": sheet_title,
            "row_index": target_tc["row_index"],
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e),
                          "trace": traceback.format_exc()[-2000:]}), 500


@app.route("/update/promote", methods=["POST"])
def update_promote():
    """사본(copy)의 TC Edit Log 에 기록된 변경을 원본(source) Sheets 에 반영.

    Body: {
      copy_url: str,         # 변경이 적용된 사본 URL
      source_url: str,       # 갱신할 원본 URL
      dry_run: bool          # True 면 미리보기만 (셀 좌표 + 값 반환, 쓰기 X)
    }

    Returns (dry_run=True):
      {ok, dry_run: True, plan: [{tc_id, sheet, row, field, cell, new_value}],
       count, sheets_seen, log_count, backup_needed: bool}

    Returns (dry_run=False):
      {ok, applied_count, backup: {backup_url, backup_title}, written_cells: [...]}
    """
    body = request.get_json(force=True) or {}
    copy_url = (body.get("copy_url") or "").strip()
    source_url = (body.get("source_url") or "").strip()
    dry_run = bool(body.get("dry_run", True))

    if not copy_url or not source_url:
        return jsonify({"ok": False,
                          "error": "copy_url 과 source_url 모두 필요합니다."}), 400

    try:
        copy_id = _extract_sheets_id(copy_url)
        source_id = _extract_sheets_id(source_url)

        # 1) 사본의 TC Edit Log 읽기
        log = read_tc_edit_log(copy_id)
        if not log["log_exists"]:
            return jsonify({"ok": False,
                              "error": "사본에 'TC Edit Log' 시트가 없습니다. 먼저 수정안을 적용하세요."}), 400
        if log["count"] == 0:
            return jsonify({"ok": False,
                              "error": "TC Edit Log 가 비어있습니다. 적용할 변경이 없습니다."}), 400

        # 2) 원본에서 같은 TC ID 의 행 위치 찾기 (사본의 행 번호는 원본과 같다는 보장 없음)
        source_tc_data = read_sheets_tcs(source_id)
        # tc_id + sheet_title → row_index 매핑
        source_locator = {}
        for t in source_tc_data.get("tcs", []):
            key = (t["tc_id"], t["sheet_title"])
            source_locator[key] = t["row_index"]

        # 3) 각 log entry → 원본 행/컬럼 찾고 사본의 실제 셀 값 읽기
        svc = get_sheets_service()
        plan = []
        missing = []  # 원본에서 못 찾은 TC

        # 사본의 실제 셀 값을 batch 로 효율 읽기
        copy_ranges = []
        copy_range_keys = []  # plan 매칭용
        for e in log["entries"]:
            col_idx = resolve_field_column(copy_id, e["sheet_title"], e["field"])
            if col_idx < 0:
                continue
            col_letter = _col_index_to_letter(col_idx)
            cell_addr = f"'{e['sheet_title']}'!{col_letter}{e['row_index']}"
            copy_ranges.append(cell_addr)
            copy_range_keys.append((e, col_idx, col_letter))

        # 사본의 실제 셀 값 batchGet
        if copy_ranges:
            batch = svc.spreadsheets().values().batchGet(
                spreadsheetId=copy_id, ranges=copy_ranges,
            ).execute()
            value_ranges = batch.get("valueRanges", [])
        else:
            value_ranges = []

        for (e, col_idx, col_letter), vr in zip(copy_range_keys, value_ranges):
            # 사본 셀의 실제 값 (요약본 아닌 전체)
            v = vr.get("values", [[""]])
            actual_value = (v[0][0] if v and v[0] else "") or ""

            # 원본에서 같은 TC 의 행 찾기
            key = (e["tc_id"], e["sheet_title"])
            source_row = source_locator.get(key)
            if source_row is None:
                missing.append({
                    "tc_id": e["tc_id"],
                    "sheet_title": e["sheet_title"],
                    "reason": "원본에서 TC ID 를 찾지 못함",
                })
                continue

            # 원본 시트의 컬럼 인덱스 (사본과 다를 수 있으니 별도 조회)
            source_col = resolve_field_column(source_id, e["sheet_title"], e["field"])
            if source_col < 0:
                missing.append({
                    "tc_id": e["tc_id"],
                    "sheet_title": e["sheet_title"],
                    "reason": f"원본에 '{e['field']}' 컬럼 없음",
                })
                continue
            source_col_letter = _col_index_to_letter(source_col)
            source_cell_addr = f"'{e['sheet_title']}'!{source_col_letter}{source_row}"

            plan.append({
                "tc_id": e["tc_id"],
                "sheet_title": e["sheet_title"],
                "field": e["field"],
                "source_row": source_row,
                "source_cell": source_cell_addr,
                "new_value": actual_value,
                "old_value_summary": e.get("old_value", ""),
                "rationale": e.get("rationale", ""),
            })

        if dry_run:
            return jsonify({
                "ok": True,
                "dry_run": True,
                "plan": plan,
                "count": len(plan),
                "missing": missing,
                "log_count": log["count"],
                "sheets_seen": log["sheets_seen"],
                "backup_needed": True,
            })

        # 실제 적용 흐름
        # 4) 원본 백업 (필수)
        backup_info = backup_original_sheets(source_url)

        # 5) 원본에 batchUpdate
        data = []
        for p in plan:
            data.append({
                "range": p["source_cell"],
                "values": [[p["new_value"]]],
            })
        if data:
            svc.spreadsheets().values().batchUpdate(
                spreadsheetId=source_id,
                body={"valueInputOption": "RAW", "data": data},
            ).execute()

        # 6) 원본에 'Promote Log' 시트 추가 (적용 이력)
        try:
            svc.spreadsheets().batchUpdate(
                spreadsheetId=source_id,
                body={"requests": [{"addSheet": {"properties": {"title": "Promote Log"}}}]},
            ).execute()
            svc.spreadsheets().values().update(
                spreadsheetId=source_id,
                range="'Promote Log'!A1",
                valueInputOption="RAW",
                body={"values": [[
                    "적용 시각", "TC ID", "시트", "필드", "원본 셀",
                    "새 값(요약)", "사본 URL", "백업 URL"
                ]]},
            ).execute()
        except Exception:
            pass  # 이미 있으면 무시

        from datetime import datetime
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_rows = []
        for p in plan:
            log_rows.append([
                ts, p["tc_id"], p["sheet_title"], p["field"], p["source_cell"],
                (p["new_value"] or "")[:200],
                copy_url, backup_info["backup_url"],
            ])
        if log_rows:
            svc.spreadsheets().values().append(
                spreadsheetId=source_id,
                range="'Promote Log'!A1",
                valueInputOption="RAW",
                insertDataOption="INSERT_ROWS",
                body={"values": log_rows},
            ).execute()

        return jsonify({
            "ok": True,
            "dry_run": False,
            "applied_count": len(plan),
            "missing": missing,
            "backup": {
                "backup_id": backup_info["backup_id"],
                "backup_url": backup_info["backup_url"],
                "backup_title": backup_info["backup_title"],
                "backed_up_at": backup_info["backed_up_at"],
            },
            "source_url": source_url,
            "applied_at": ts,
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e),
                          "trace": traceback.format_exc()[-2500:]}), 500


@app.route("/update/rollback", methods=["POST"])
def update_rollback():
    """백업 사본의 셀 값을 원본 Sheets 로 복원 (롤백).

    Body: {source_url, backup_url, scope: 'promote_log' | 'all'}

    scope='promote_log' (권장):
      - 원본의 'Promote Log' 시트를 읽어 어떤 셀을 promote 로 갱신했는지 파악
      - 그 셀들만 백업의 같은 셀 값으로 복원 (가장 안전)

    scope='all':
      - 백업의 모든 탭/셀을 원본에 통째로 덮어쓰기 (강력하지만 위험)
      - 만약 백업 이후 원본에 다른 수동 수정이 있었다면 그것도 사라짐

    Returns: {ok, scope, restored_count, message}
    """
    body = request.get_json(force=True) or {}
    source_url = (body.get("source_url") or "").strip()
    backup_url = (body.get("backup_url") or "").strip()
    scope = (body.get("scope") or "promote_log").strip()

    if not source_url or not backup_url:
        return jsonify({"ok": False,
                          "error": "source_url 과 backup_url 모두 필요합니다."}), 400

    try:
        source_id = _extract_sheets_id(source_url)
        backup_id = _extract_sheets_id(backup_url)
        svc = get_sheets_service()

        if scope == "promote_log":
            # 원본의 'Promote Log' 읽기 → 적용된 셀 좌표 목록 추출
            meta = svc.spreadsheets().get(
                spreadsheetId=source_id,
                fields="sheets(properties(sheetId,title))",
            ).execute()
            titles = [s["properties"]["title"] for s in meta.get("sheets", [])]
            if "Promote Log" not in titles:
                return jsonify({"ok": False,
                                  "error": "원본에 'Promote Log' 시트가 없습니다. 이전에 promote 한 기록이 없어 롤백 대상이 없습니다."}), 400

            resp = svc.spreadsheets().values().get(
                spreadsheetId=source_id,
                range="'Promote Log'!A1:Z2000",
            ).execute()
            values = resp.get("values", [])
            if not values or len(values) < 2:
                return jsonify({"ok": True, "scope": "promote_log",
                                  "restored_count": 0,
                                  "message": "Promote Log 가 비어있습니다."})

            header = values[0]
            def col_idx(name):
                return header.index(name) if name in header else -1
            ci_cell = col_idx("원본 셀")
            if ci_cell < 0:
                return jsonify({"ok": False, "error": "'원본 셀' 컬럼을 못 찾음"}), 500

            # 적용된 셀 좌표 수집 (중복 제거 — 가장 최근 백업 시점 기준으로 복원)
            cell_addrs = []
            for row in values[1:]:
                if ci_cell < len(row) and row[ci_cell]:
                    addr = str(row[ci_cell]).strip()
                    if addr and addr not in cell_addrs:
                        cell_addrs.append(addr)

            if not cell_addrs:
                return jsonify({"ok": True, "scope": "promote_log",
                                  "restored_count": 0,
                                  "message": "복원할 셀이 없습니다."})

            # 백업에서 같은 셀들의 값 읽기
            backup_batch = svc.spreadsheets().values().batchGet(
                spreadsheetId=backup_id, ranges=cell_addrs,
            ).execute()
            backup_vrs = backup_batch.get("valueRanges", [])

            # 원본에 batchUpdate
            data = []
            for addr, vr in zip(cell_addrs, backup_vrs):
                vals = vr.get("values", [[""]])
                v = (vals[0][0] if vals and vals[0] else "") or ""
                data.append({"range": addr, "values": [[v]]})

            if data:
                svc.spreadsheets().values().batchUpdate(
                    spreadsheetId=source_id,
                    body={"valueInputOption": "RAW", "data": data},
                ).execute()

            # 'Rollback Log' 에 기록
            try:
                svc.spreadsheets().batchUpdate(
                    spreadsheetId=source_id,
                    body={"requests": [{"addSheet": {"properties": {"title": "Rollback Log"}}}]},
                ).execute()
                svc.spreadsheets().values().update(
                    spreadsheetId=source_id,
                    range="'Rollback Log'!A1",
                    valueInputOption="RAW",
                    body={"values": [["롤백 시각", "복원 셀 수", "백업 URL", "scope"]]},
                ).execute()
            except Exception:
                pass

            from datetime import datetime
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            svc.spreadsheets().values().append(
                spreadsheetId=source_id,
                range="'Rollback Log'!A1",
                valueInputOption="RAW",
                insertDataOption="INSERT_ROWS",
                body={"values": [[ts, str(len(data)), backup_url, scope]]},
            ).execute()

            return jsonify({
                "ok": True,
                "scope": "promote_log",
                "restored_count": len(data),
                "message": f"{len(data)}개 셀을 백업 시점으로 복원했습니다.",
                "restored_at": ts,
            })

        elif scope == "all":
            # 모든 탭 전체 복원 — 위험하므로 일단 미지원 (필요 시 추후)
            return jsonify({"ok": False,
                              "error": "scope='all' 은 아직 지원하지 않습니다. scope='promote_log' 사용"}), 400
        else:
            return jsonify({"ok": False,
                              "error": f"알 수 없는 scope: {scope}"}), 400

    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e),
                          "trace": traceback.format_exc()[-2500:]}), 500


@app.route("/update/apply-bulk", methods=["POST"])
def update_apply_bulk():
    """선택된 후보들에 대해 AI 수정안 제안 + 사본 셀 갱신 + TC Edit Log 기록을 순차 실행.

    Body: {session_id, selections: [{idx, accept: bool}], skip_existing?: bool}

    동작:
      체크된 modify 후보들에 대해:
        1. propose 호출 (AI) — no_change 면 skip
        2. 사본 셀 갱신 (3개 필드: precondition, steps, expected)
        3. TC Edit Log 시트에 기록

    skip_existing=True (default): 이미 TC Edit Log 에 있는 TC 는 skip (중복 호출 방지)

    Returns: {
      ok, processed, applied, no_change, errors,
      details: [{tc_id, status, fields_changed?, error?}],
      log_sheet: 'TC Edit Log', copy_url
    }
    """
    body = request.get_json(force=True) or {}
    session_id = (body.get("session_id") or "").strip()
    selections = body.get("selections") or []
    skip_existing = bool(body.get("skip_existing", True))

    sess = _UPDATE_SESSIONS.get(session_id)
    if not sess:
        return jsonify({"ok": False, "error": "세션이 없습니다. /update/plan 을 다시 호출하세요."}), 404

    try:
        target_id = sess["target_sheets_id"]
        candidates = sess["candidates"]
        prev_folder = sess.get("prev_folder", "")
        new_folder = sess.get("new_folder", "")

        # 수정 대상 수집 (modify + tc_id 있는 것만)
        targets = []
        for sel in selections:
            idx = sel.get("idx")
            if idx is None or idx < 0 or idx >= len(candidates):
                continue
            if not sel.get("accept"):
                continue
            c = candidates[idx]
            if c.get("action") != "modify" or not c.get("tc_id"):
                continue
            targets.append({**c, "idx": idx})

        if not targets:
            return jsonify({"ok": True, "processed": 0,
                            "message": "체크된 수정 후보가 없습니다.",
                            "details": []})

        svc = get_sheets_service()

        # TC Edit Log 시트 확인/생성 + 헤더 + 기존 (TC ID, sheet, field) 조합 수집 (중복 방지)
        existing_keys = set()  # (tc_id, sheet_title, field)
        log_title = "TC Edit Log"
        log_header = ["시각", "TC ID", "시트", "행", "필드", "이전 값(요약)", "새 값(요약)", "수정 사유"]
        try:
            svc.spreadsheets().batchUpdate(
                spreadsheetId=target_id,
                body={"requests": [{"addSheet": {"properties": {"title": log_title}}}]},
            ).execute()
            svc.spreadsheets().values().update(
                spreadsheetId=target_id,
                range=f"'{log_title}'!A1",
                valueInputOption="RAW",
                body={"values": [log_header]},
            ).execute()
        except Exception:
            # 이미 있으면 기존 (tc_id, sheet, field) 조합 읽기
            try:
                resp = svc.spreadsheets().values().get(
                    spreadsheetId=target_id,
                    range=f"'{log_title}'!A1:Z2000",
                ).execute()
                vals = resp.get("values", [])
                if vals:
                    h = vals[0]
                    ci_tc = h.index("TC ID") if "TC ID" in h else -1
                    ci_sh = h.index("시트") if "시트" in h else -1
                    ci_fd = h.index("필드") if "필드" in h else -1
                    for r in vals[1:]:
                        def c(i):
                            return str(r[i]).strip() if i >= 0 and i < len(r) else ""
                        k = (c(ci_tc), c(ci_sh), c(ci_fd))
                        if all(k):
                            existing_keys.add(k)
            except Exception:
                pass

        # SCR 별 본문 캐시 (같은 SCR 여러번 안 읽도록)
        scr_text_cache = {}
        def get_scr_pair(scr_id):
            if scr_id not in scr_text_cache:
                scr_text_cache[scr_id] = _read_scr_pair(prev_folder, new_folder, scr_id)
            return scr_text_cache[scr_id]

        # 사본의 TC 매핑 (한 번만 읽기)
        tc_data = read_sheets_tcs(target_id)
        tc_index = {(t["tc_id"], t["sheet_title"]): t for t in tc_data.get("tcs", [])}

        # 처리 시작
        details = []
        applied_count = 0
        no_change_count = 0
        error_count = 0
        skipped_count = 0
        new_log_rows = []
        from datetime import datetime

        for tgt in targets:
            tc_id = tgt["tc_id"]
            sheet_title = tgt.get("sheet_title", "")
            scr_id = tgt.get("scr_id", "")

            # 사본에서 행 찾기
            target_tc = tc_index.get((tc_id, sheet_title))
            if not target_tc:
                error_count += 1
                details.append({"tc_id": tc_id, "sheet": sheet_title,
                                  "status": "error",
                                  "error": f"사본에서 {tc_id} 행을 찾지 못함"})
                continue
            row_index = target_tc["row_index"]

            # 현재 TC 행 본문 읽기
            try:
                current_tc = _read_tc_row_from_sheets(target_id, sheet_title, row_index)
            except Exception as e:
                error_count += 1
                details.append({"tc_id": tc_id, "sheet": sheet_title,
                                  "status": "error",
                                  "error": f"행 읽기 실패: {e}"})
                continue

            # AI 호출
            prev_text, new_text = get_scr_pair(scr_id)
            try:
                result = step_propose_tc_update(scr_id, prev_text, new_text, current_tc)
            except Exception as e:
                error_count += 1
                details.append({"tc_id": tc_id, "sheet": sheet_title,
                                  "status": "error",
                                  "error": f"AI 호출 실패: {e}"})
                continue

            if not result.get("ok"):
                error_count += 1
                details.append({"tc_id": tc_id, "sheet": sheet_title,
                                  "status": "error",
                                  "error": result.get("error", "AI 응답 실패")})
                continue

            proposal = result["proposal"]
            if proposal.get("no_change"):
                no_change_count += 1
                details.append({"tc_id": tc_id, "sheet": sheet_title,
                                  "status": "no_change",
                                  "rationale": proposal.get("rationale", "")})
                continue

            # 필드별 갱신 (사본 셀)
            field_to_col = {
                "precondition": current_tc["_col_pre"],
                "steps": current_tc["_col_steps"],
                "expected": current_tc["_col_exp"],
            }
            fields_changed = []
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for field, new_value in [
                ("precondition", proposal.get("precondition")),
                ("steps", proposal.get("steps")),
                ("expected", proposal.get("expected")),
            ]:
                if new_value is None:
                    continue
                # skip_existing 체크
                if skip_existing and (tc_id, sheet_title, field) in existing_keys:
                    continue
                ci = field_to_col.get(field, -1)
                if ci < 0:
                    continue
                col_letter = _col_index_to_letter(ci)
                cell_range = f"'{sheet_title}'!{col_letter}{row_index}"
                try:
                    svc.spreadsheets().values().update(
                        spreadsheetId=target_id,
                        range=cell_range,
                        valueInputOption="RAW",
                        body={"values": [[str(new_value)]]},
                    ).execute()
                    fields_changed.append(field)
                    new_log_rows.append([
                        ts, tc_id, sheet_title, str(row_index), field,
                        (current_tc.get(field, "") or "")[:200],
                        (new_value or "")[:200],
                        (proposal.get("rationale") or "")[:300],
                    ])
                    existing_keys.add((tc_id, sheet_title, field))
                except Exception as e:
                    details.append({"tc_id": tc_id, "sheet": sheet_title,
                                      "field": field, "status": "error",
                                      "error": f"셀 쓰기 실패: {e}"})

            if fields_changed:
                applied_count += 1
                details.append({"tc_id": tc_id, "sheet": sheet_title,
                                  "status": "applied",
                                  "fields_changed": fields_changed,
                                  "rationale": proposal.get("rationale", "")})
            else:
                # AI 가 변경 제안했는데 skip_existing 이 다 막은 경우 — 사실상 no_change
                skipped_count += 1
                details.append({"tc_id": tc_id, "sheet": sheet_title,
                                  "status": "skipped",
                                  "reason": "이미 TC Edit Log 에 있어 skip"})

        # TC Edit Log 에 일괄 append
        if new_log_rows:
            try:
                svc.spreadsheets().values().append(
                    spreadsheetId=target_id,
                    range=f"'{log_title}'!A1",
                    valueInputOption="RAW",
                    insertDataOption="INSERT_ROWS",
                    body={"values": new_log_rows},
                ).execute()
            except Exception:
                pass  # 핵심 셀은 이미 박힘 — 로그 실패는 치명적 X

        return jsonify({
            "ok": True,
            "processed": len(targets),
            "applied": applied_count,
            "no_change": no_change_count,
            "skipped": skipped_count,
            "errors": error_count,
            "details": details,
            "log_sheet": log_title,
            "copy_url": sess.get("copy_info", {}).get("copy_url", ""),
        })
    except Exception as e:
        import traceback
        return jsonify({
            "ok": False, "error": str(e),
            "trace": traceback.format_exc()[-2500:],
        }), 500


# 레거시 — 기존 /update/apply 는 그대로 두고 (호환성), 새 UI 는 /update/apply-bulk 사용
@app.route("/update/apply", methods=["POST"])
def update_apply():
    """3단계 (레거시) — Update Log 메모만 기록.

    ⚠️ 사본 셀은 갱신하지 않음. 실제 갱신은 /update/apply-bulk 사용.
    이전 UI 호환성 위해 유지.
    """
    body = request.get_json(force=True) or {}
    session_id = (body.get("session_id") or "").strip()
    selections = body.get("selections") or []
    sess = _UPDATE_SESSIONS.get(session_id)
    if not sess:
        return jsonify({"ok": False, "error": "세션이 없습니다."}), 404

    try:
        target_id = sess["target_sheets_id"]
        candidates = sess["candidates"]

        applied = []
        for sel in selections:
            idx = sel.get("idx")
            if idx is None or idx < 0 or idx >= len(candidates):
                continue
            if not sel.get("accept"):
                continue
            applied.append({**candidates[idx], "idx": idx})

        if not applied:
            return jsonify({"ok": True, "applied_count": 0,
                            "message": "선택된 항목이 없습니다."})

        svc = get_sheets_service()
        from datetime import datetime
        log_title = f"Update Log {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        svc.spreadsheets().batchUpdate(
            spreadsheetId=target_id,
            body={"requests": [{"addSheet": {"properties": {"title": log_title}}}]}
        ).execute()

        header = ["#", "액션", "SCR ID", "TC ID", "TC 제목", "시트", "변경 종류", "사유", "추출 시각"]
        rows = [header]
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for i, a in enumerate(applied, start=1):
            rows.append([
                str(i), a["action"], a.get("scr_id", ""), a.get("tc_id", "") or "(신규)",
                a.get("tc_title", ""), a.get("sheet_title", ""),
                a.get("change_kind", ""), a.get("reason", ""), ts,
            ])

        svc.spreadsheets().values().update(
            spreadsheetId=target_id,
            range=f"'{log_title}'!A1",
            valueInputOption="RAW",
            body={"values": rows},
        ).execute()

        return jsonify({
            "ok": True,
            "applied_count": len(applied),
            "log_sheet": log_title,
            "copy_url": sess.get("copy_info", {}).get("copy_url", ""),
            "warning": "레거시 모드 — 사본 셀은 갱신 안 됨. /update/apply-bulk 사용 권장",
        })
    except Exception as e:
        import traceback
        return jsonify({
            "ok": False, "error": str(e),
            "trace": traceback.format_exc()[-2500:],
        }), 500


@app.route("/combo/download/<workspace_id>/<filename>", methods=["GET"])
def combo_download(workspace_id, filename):
    """생성된 Excel/TC md 다운로드."""
    # 보안 — workspace_id 검증 (alphanum + dash + underscore + combo_ prefix)
    if not re.match(r"^combo_[a-f0-9]+$", workspace_id):
        return jsonify({"ok": False, "error": "잘못된 workspace id"}), 400
    if not re.match(r"^[A-Za-z0-9_\-\.]+$", filename):
        return jsonify({"ok": False, "error": "잘못된 파일명"}), 400
    file_path = Path(__file__).parent.parent / "workspace" / workspace_id / filename
    if not file_path.exists() or not file_path.is_file():
        return jsonify({"ok": False, "error": "파일 없음"}), 404
    return send_file(str(file_path), as_attachment=True)




@app.route("/preview-spec-folder", methods=["POST"])
def preview_spec_folder():
    """폴더 경로만 받아 미리 분류 결과 반환 (시작 전 검증용)."""
    data = request.get_json(force=True) or {}
    folder_path = (data.get("folder_path") or "").strip()
    if not folder_path:
        return jsonify({"ok": False, "error": "folder_path 가 비어 있습니다."}), 400
    folder = Path(folder_path).expanduser()
    if not folder.exists() or not folder.is_dir():
        return jsonify({"ok": False, "error": f"폴더 없음: {folder}"}), 400

    cls = classify_spec_files(folder)
    overview_text = cls["overview"].read_text(encoding="utf-8") if cls["overview"] else ""
    rows = parse_screen_list_table(overview_text) if overview_text else []
    return jsonify({
        "ok": True,
        "folder": str(folder),
        "overview": cls["overview"].name if cls["overview"] else None,
        "policy": [p.name for p in cls["policy"]],
        "design": [p.name for p in cls["design"]],
        "screens": [p.name for p in cls["screens"]],
        "screen_rows_count": len(rows),
        "majors": sorted(set(r["major"] for r in rows if r["major"])),
    })


@app.route("/diff-spec-folders", methods=["POST"])
def diff_spec_folders_route():
    """두 폴더 비교 미리보기 — 변경 SCR 목록 반환."""
    data = request.get_json(force=True) or {}
    new_path = (data.get("folder_path") or "").strip()
    prev_path = (data.get("prev_folder_path") or "").strip()
    if not new_path or not prev_path:
        return jsonify({"ok": False, "error": "folder_path 와 prev_folder_path 모두 필요합니다."}), 400
    new_p = Path(new_path).expanduser()
    prev_p = Path(prev_path).expanduser()
    if not new_p.exists() or not new_p.is_dir():
        return jsonify({"ok": False, "error": f"신규 폴더 없음: {new_p}"}), 400
    if not prev_p.exists() or not prev_p.is_dir():
        return jsonify({"ok": False, "error": f"이전 폴더 없음: {prev_p}"}), 400
    try:
        diff = diff_spec_folders(prev_p, new_p)
    except Exception as e:
        return jsonify({"ok": False, "error": f"diff 실패: {e}"}), 500
    return jsonify({
        "ok": True,
        "new_folder": str(new_p),
        "prev_folder": str(prev_p),
        "added": diff["added"],
        "modified": diff["modified"],
        "removed": diff["removed"],
        "unchanged_count": len(diff["unchanged"]),
        "common_changed": diff["common_changed"],
        "regenerate_count": len(diff["added"]) + len(diff["modified"]),
    })


@app.route("/check-resume-spec", methods=["GET"])
def check_resume_spec():
    """프로젝트의 구조화 spec 모드 체크포인트가 있는지 확인.
    Query: project_name=...
    Returns: {ok, resumable, stage, completed_screens, total_screens, ...}
    """
    project_name = request.args.get("project_name", "").strip()
    if not project_name:
        return jsonify({"ok": False, "error": "project_name 필요"}), 400
    state = load_pipeline_state(project_name)
    if not state:
        return jsonify({"ok": True, "resumable": False})
    ckpt = state["data"]
    folder = ckpt.get("structured_spec_folder")
    if not folder:
        return jsonify({"ok": True, "resumable": False, "reason": "구조화 spec 모드 체크포인트 아님"})
    workspace = ckpt.get("workspace")
    completed = []
    if workspace and Path(workspace).exists():
        per_screen_dir = Path(workspace) / "per_screen"
        if per_screen_dir.exists():
            completed = sorted([p.stem for p in per_screen_dir.glob("SCR-*.md")])
    # 신규 폴더에서 화면 개수 다시 계산 (참고용)
    folder_p = Path(folder)
    total = 0
    if folder_p.exists():
        cls = classify_spec_files(folder_p)
        total = len(cls["screens"])
    return jsonify({
        "ok": True,
        "resumable": True,
        "stage": state["stage"],
        "folder": folder,
        "prev_folder": ckpt.get("prev_folder", ""),
        "completed_screens": completed,
        "completed_count": len(completed),
        "total_screens": total,
        "remaining": max(total - len(completed), 0) if total else None,
    })


@app.route("/start", methods=["POST"])
def start():
    data = request.get_json(force=True) or {}
    project_name = data.get("project_name", "프로젝트").strip() or "프로젝트"
    focus_area   = (data.get("focus_area") or "").strip()
    resume       = data.get("resume", False)  # 이어서 작업 여부

    # 이어서 작업 모드
    if resume:
        state = load_pipeline_state(project_name)
        if not state:
            return jsonify({"ok": False, "error": "저장된 작업 상태가 없습니다."}), 400
        if not os.environ.get("ANTHROPIC_API_KEY"):
            return jsonify({"ok": False, "error": "ANTHROPIC_API_KEY가 설정되지 않았습니다."}), 500
        sess = new_session()
        sess["project_name"] = project_name
        sess["focus_area"] = state["data"].get("focus_area", "")
        sess["_resumed_state"] = state
        # 이어서 작업 시 소스는 빈 배열 (이미 파싱 결과가 state에 있음)
        t = threading.Thread(
            target=run_pipeline,
            args=(sess, [], project_name),
            daemon=True
        )
        sess["thread"] = t
        t.start()
        return jsonify({"ok": True, "sid": sess["id"], "resumed_stage": state["stage"]})

    # 새 형식: sources 배열 / 구버전 호환: input_type + content
    sources = data.get("sources")
    if not sources:
        input_type = data.get("input_type", "text")
        content    = data.get("content", "").strip()
        sources    = [{"type": input_type, "content": content}]

    if not sources:
        return jsonify({"ok": False, "error": "소스가 없습니다."}), 400
    if not os.environ.get("ANTHROPIC_API_KEY"):
        return jsonify({"ok": False, "error": "ANTHROPIC_API_KEY가 설정되지 않았습니다."}), 500

    for src in sources:
        t = src.get("type", "text")
        c = src.get("content", "").strip()
        if not c:
            return jsonify({"ok": False, "error": f"{t} 소스의 내용이 비어 있습니다."}), 400
        if t not in ("pdf", "url", "web", "md", "text"):
            return jsonify({"ok": False, "error": f"소스 유형 오류: {t}"}), 400
        if t == "pdf" and not (SPECS_DIR / c).exists():
            return jsonify({"ok": False, "error": f"PDF 파일 없음: {c}"}), 400
        if t == "md" and not (SPECS_DIR / c).exists():
            return jsonify({"ok": False, "error": f"마크다운 파일 없음: {c}"}), 400

    sess = new_session()
    sess["project_name"] = project_name
    sess["focus_area"] = focus_area
    # 생성 모드: "summary" (기본, policy/features/classify 3단계) / "direct" (원문 직접 분류)
    gen_mode = (data.get("generation_mode") or "summary").strip().lower()
    if gen_mode not in ("summary", "direct"):
        gen_mode = "summary"
    sess["generation_mode"] = gen_mode
    # 소스 정보를 프로젝트에 저장 (다음에 불러오기용)
    sources_to_save = []
    for src in sources:
        sources_to_save.append({
            "type": src.get("type", "text"),
            "content": src.get("content", ""),
            "selected_files": src.get("selected_files"),
        })
    save_project(project_name, last_sources=sources_to_save, last_focus_area=focus_area)
    t = threading.Thread(
        target=run_pipeline,
        args=(sess, sources, project_name),
        daemon=True
    )
    sess["thread"] = t
    t.start()
    return jsonify({"ok": True, "sid": sess["id"]})


@app.route("/projects/<name>/sources")
def get_project_sources(name):
    """프로젝트에 저장된 이전 소스 정보 반환"""
    projects = load_projects()
    proj = next((p for p in projects if p["name"] == name), None)
    if not proj:
        return jsonify({"ok": True, "has_sources": False})
    last_sources = proj.get("last_sources")
    last_focus = proj.get("last_focus_area", "")
    last_generation_mode = proj.get("last_generation_mode", "")
    last_combo_opts = proj.get("last_combo_opts") or None
    # last_sources가 없으면 pipeline_state에서 가져옴
    if not last_sources:
        state = load_pipeline_state(name)
        if state and state.get("data", {}).get("sources_info"):
            last_sources = state["data"]["sources_info"]
            last_focus = state["data"].get("focus_area", "")
    if not last_sources:
        return jsonify({"ok": True, "has_sources": False})
    return jsonify({
        "ok": True,
        "has_sources": True,
        "sources": last_sources,
        "focus_area": last_focus,
        "generation_mode": last_generation_mode,
        "combo_opts": last_combo_opts,
    })


@app.route("/projects/<name>/state")
def get_project_state(name):
    """프로젝트의 저장된 파이프라인 상태 조회"""
    state = load_pipeline_state(name)
    if not state:
        return jsonify({"ok": True, "has_state": False})
    return jsonify({
        "ok": True,
        "has_state": True,
        "stage": state["stage"],
        "saved_at": state["saved_at"],
        "focus_area": state["data"].get("focus_area", ""),
    })


@app.route("/stream/<sid>")
def stream(sid):
    sess = SESSIONS.get(sid)
    if not sess:
        return jsonify({"error": "세션 없음"}), 404

    def generate():
        while True:
            try:
                evt = sess["events"].get(timeout=30)
                payload = json.dumps(evt, ensure_ascii=False)
                yield f"data: {payload}\n\n"
                if evt["type"] in ("done", "error"):
                    break
            except queue.Empty:
                # heartbeat
                yield "data: {\"type\":\"ping\"}\n\n"
                # 세션이 완료/오류면 종료
                if sess["status"] in ("done", "error"):
                    break

    return Response(
        generate(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        }
    )


@app.route("/projects", methods=["GET"])
def list_projects():
    projects = load_projects()
    return jsonify(projects)


@app.route("/projects", methods=["POST"])
def create_project():
    """새 프로젝트 생성 (빈 TC로 등록)"""
    data = request.get_json(force=True) or {}
    project_name = data.get("name", "").strip()
    if not project_name:
        return jsonify({"ok": False, "error": "프로젝트명을 입력하세요."}), 400
    projects = load_projects()
    if any(p["name"] == project_name for p in projects):
        return jsonify({"ok": False, "error": "이미 존재하는 프로젝트명입니다."}), 400
    save_project(project_name, "", "")
    return jsonify({"ok": True, "project_name": project_name})


@app.route("/projects/<path:project_name>", methods=["DELETE"])
def delete_project(project_name):
    """프로젝트 삭제. 빈 이름 유령 레코드도 함께 정리."""
    # 유령 레코드(name=빈값)를 포함한 잘못된 레코드까지 정리
    raw = []
    if PROJECTS_FILE.exists():
        try:
            raw = json.loads(PROJECTS_FILE.read_text(encoding="utf-8"))
        except Exception:
            raw = []
    target = (project_name or "").strip()
    cleaned = [
        p for p in raw
        if isinstance(p, dict) and (p.get("name") or "").strip() and p.get("name") != target
    ]
    PROJECTS_FILE.write_text(json.dumps(cleaned, ensure_ascii=False, indent=2), encoding="utf-8")
    return jsonify({"ok": True})


def excel_to_md(file_path: Path) -> str:
    """Excel TC 파일을 md 텍스트로 변환 (헤더 행 기준 자동 파싱)"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl 패키지가 필요합니다: pip install openpyxl")
    wb = load_workbook(str(file_path), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError("Excel 파일이 비어 있습니다.")

    # 헤더 행 찾기 (TC ID 또는 ID 컬럼 포함)
    header_row, header_idx = None, 0
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c else "" for c in row]
        if any(k in " ".join(cells) for k in ["TC ID", "tc_id", "ID", "제목", "Title"]):
            header_row = cells
            header_idx = i
            break
    if header_row is None:
        header_row = [str(c) if c else f"col{j}" for j, c in enumerate(rows[0])]

    lines = []
    for row in rows[header_idx + 1:]:
        if not any(c for c in row):
            continue
        vals = {header_row[j]: (str(row[j]).strip() if row[j] is not None else "") for j in range(min(len(header_row), len(row)))}
        tc_id    = vals.get("TC ID") or vals.get("ID") or vals.get("tc_id") or ""
        title    = vals.get("제목") or vals.get("Title") or vals.get("title") or ""
        major    = vals.get("대분류") or ""
        middle   = vals.get("중분류") or ""
        minor    = vals.get("소분류") or ""
        tc_type  = vals.get("분류") or vals.get("Type") or "Positive"
        priority = vals.get("우선순위") or vals.get("Priority") or "Medium"
        platform = vals.get("플랫폼") or vals.get("Platform") or ""
        screen   = vals.get("연관 화면") or vals.get("화면") or vals.get("Screen") or ""
        precond  = vals.get("사전 조건") or vals.get("Precondition") or vals.get("Given") or ""
        steps    = vals.get("테스트 단계") or vals.get("Steps") or vals.get("When") or ""
        expected = vals.get("예상 결과") or vals.get("Expected") or vals.get("Then") or ""
        note     = vals.get("비고") or vals.get("Note") or ""
        is_min   = vals.get("최소TC") or vals.get("최소 TC") or ""

        heading = f"### {'**' + tc_id + '**' if is_min and is_min.upper() == 'Y' else tc_id} — {title}"
        lines.append(heading)
        lines.append(f"\n| 항목 | 내용 |\n|------|------|")
        if major:    lines.append(f"| 대분류 | {major} |")
        if middle:   lines.append(f"| 중분류 | {middle} |")
        if minor:    lines.append(f"| 소분류 | {minor} |")
        lines.append(f"| 분류 | {tc_type} |")
        lines.append(f"| 우선순위 | {priority} |")
        if platform: lines.append(f"| 플랫폼 | {platform} |")
        if screen:   lines.append(f"| 연관 화면 | {screen} |")
        lines.append(f"\n**사전 조건**\n{precond}")
        lines.append(f"\n**테스트 단계**\n{steps}")
        lines.append(f"\n**예상 결과**\n{expected}")
        if note:     lines.append(f"\n**비고**\n{note}")
        lines.append("")
    return "\n".join(lines)


@app.route("/upload-tc", methods=["POST"])
def upload_tc():
    """TC 파일 업로드 — .md / .xlsx / .xls 지원"""
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "파일 없음"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"ok": False, "error": "파일명 없음"}), 400

    ext = Path(f.filename).suffix.lower()
    if ext not in (".md", ".xlsx", ".xls"):
        return jsonify({"ok": False, "error": ".md / .xlsx / .xls 파일만 허용"}), 400

    project_name = request.form.get("project_name", Path(f.filename).stem)
    safe_name = re.sub(r"[^\w\-_]", "_", project_name)[:30]
    today = datetime.now().strftime("%Y%m%d")

    if ext == ".md":
        save_path = TC_FILES_DIR / f"{safe_name}_{today}_upload.md"
        f.save(str(save_path))
    else:
        # Excel → 임시 저장 후 md 변환
        tmp_path = TC_FILES_DIR / f"{safe_name}_{today}_tmp{ext}"
        f.save(str(tmp_path))
        try:
            md_text = excel_to_md(tmp_path)
        except Exception as e:
            tmp_path.unlink(missing_ok=True)
            return jsonify({"ok": False, "error": f"Excel 변환 실패: {e}"}), 400
        tmp_path.unlink(missing_ok=True)
        save_path = TC_FILES_DIR / f"{safe_name}_{today}_upload.md"
        save_path.write_text(md_text, encoding="utf-8")

    save_project(project_name, str(save_path), "")
    return jsonify({"ok": True, "project_name": project_name, "tc_file": save_path.name})


@app.route("/import-sheets", methods=["POST"])
def import_sheets():
    """Google Sheets URL에서 TC 데이터를 가져와 md로 변환"""
    try:
        import requests as req_lib
    except ImportError:
        return jsonify({"ok": False, "error": "requests 패키지가 필요합니다"}), 500

    data = request.get_json(force=True) or {}
    url = data.get("url", "").strip()
    project_name = data.get("project_name", "").strip()

    if not url:
        return jsonify({"ok": False, "error": "URL을 입력하세요."}), 400
    if not project_name:
        return jsonify({"ok": False, "error": "프로젝트명을 입력하세요."}), 400

    # Google Sheets ID 추출 및 CSV export URL 생성
    sheets_match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if not sheets_match:
        return jsonify({"ok": False, "error": "Google Sheets URL 형식이 올바르지 않습니다."}), 400

    sheet_id = sheets_match.group(1)
    # gid(시트 탭 ID) 추출
    gid_match = re.search(r"[#&?]gid=(\d+)", url)
    gid = gid_match.group(1) if gid_match else "0"
    csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"

    try:
        resp = req_lib.get(csv_url, timeout=20)
        if resp.status_code == 403:
            return jsonify({"ok": False, "error": "Sheets 접근 권한이 없습니다. '링크가 있는 사용자 공개' 설정 필요"}), 403
        resp.raise_for_status()
        csv_text = resp.text
    except Exception as e:
        return jsonify({"ok": False, "error": f"Sheets 다운로드 실패: {e}"}), 400

    # CSV → md 변환
    import csv, io
    reader = csv.reader(io.StringIO(csv_text))
    rows = list(reader)
    if not rows:
        return jsonify({"ok": False, "error": "시트가 비어 있습니다."}), 400

    # 헤더 탐색
    header_row, header_idx = None, 0
    for i, row in enumerate(rows):
        if any(k in " ".join(row) for k in ["TC ID", "ID", "제목", "Title"]):
            header_row = row
            header_idx = i
            break
    if header_row is None:
        header_row = rows[0]

    lines = []
    for row in rows[header_idx + 1:]:
        if not any(c.strip() for c in row):
            continue
        vals = {header_row[j]: row[j].strip() if j < len(row) else "" for j in range(len(header_row))}
        tc_id    = vals.get("TC ID") or vals.get("ID") or ""
        title    = vals.get("제목") or vals.get("Title") or ""
        tc_type  = vals.get("분류") or vals.get("Type") or "Positive"
        priority = vals.get("우선순위") or vals.get("Priority") or "Medium"
        precond  = vals.get("사전 조건") or vals.get("Precondition") or ""
        steps    = vals.get("테스트 단계") or vals.get("Steps") or ""
        expected = vals.get("예상 결과") or vals.get("Expected") or ""
        is_min   = vals.get("최소TC") or vals.get("최소 TC") or ""
        heading  = f"### {'**' + tc_id + '**' if is_min.upper() == 'Y' else tc_id} — {title}"
        lines += [heading, f"\n| 항목 | 내용 |\n|------|------|",
                  f"| 분류 | {tc_type} |", f"| 우선순위 | {priority} |",
                  f"\n**사전 조건**\n{precond}", f"\n**테스트 단계**\n{steps}",
                  f"\n**예상 결과**\n{expected}", ""]

    md_text = "\n".join(lines)
    safe_name = re.sub(r"[^\w\-_]", "_", project_name)[:30]
    today = datetime.now().strftime("%Y%m%d")
    save_path = TC_FILES_DIR / f"{safe_name}_{today}_sheets.md"
    save_path.write_text(md_text, encoding="utf-8")
    save_project(project_name, str(save_path), "")
    return jsonify({"ok": True, "project_name": project_name, "tc_file": save_path.name, "tc_count": len(lines) // 8})


@app.route("/upload-existing-tc", methods=["POST"])
def upload_existing_tc():
    """기존 TC 파일(MD/Excel)을 업로드하여 specs 또는 tc_files에 저장.
    반환: 저장 경로 + 감지된 TC 개수.
    """
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "파일 없음"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"ok": False, "error": "파일명 없음"}), 400
    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
    if ext not in ("md", "markdown", "xlsx", "xls"):
        return jsonify({"ok": False, "error": ".md / .xlsx 파일만 허용"}), 400
    # 업로드 저장 — tc_files 디렉토리 (Excel 포함)
    save_path = TC_FILES_DIR / f.filename
    f.save(str(save_path))
    try:
        tcs = load_existing_tcs(save_path)
    except Exception as e:
        return jsonify({"ok": False, "error": f"TC 파일 파싱 실패: {e}"}), 400
    return jsonify({
        "ok": True,
        "filename": f.filename,
        "tc_count": len(tcs),
        "tc_ids_sample": [t["id"] for t in tcs[:5]],
    })


def _tc_to_markdown_block(tc: dict) -> str:
    """파싱된 TC dict를 markdown TC 블록 형식으로 직렬화 (build_excel가 파싱 가능한 포맷).
    이미 raw_text가 있으면 그걸 사용 (업데이트 전 원본 보존용)."""
    if tc.get("raw_block"):
        return tc["raw_block"]
    tid = tc.get("id", "")
    title = tc.get("title", "") or tid
    bold = "**" if tc.get("is_min") else ""
    header = f"### {bold}{tid}{bold} — {title}"
    table = [
        f"| 대분류 | {tc.get('major', '')} |",
        f"| 중분류 | {tc.get('middle', '')} |",
        f"| 소분류 | {tc.get('minor', '')} |",
        f"| 분류 | {tc.get('type', 'Positive')} |",
        f"| 우선순위 | {tc.get('priority', 'Medium')} |",
    ]
    if tc.get("screen") or tc.get("screen_code"):
        table.append(f"| 연관 화면 | {tc.get('screen') or tc.get('screen_code')} |")
    sections = [
        f"**사전 조건**\n{tc.get('precondition', '')}",
        f"**테스트 단계**\n{tc.get('steps', '')}",
        f"**예상 결과**\n{tc.get('expected', '')}",
    ]
    if tc.get("note"):
        sections.append(f"**비고**\n{tc['note']}")
    return "\n".join([header, "", "| 항목 | 내용 |", "|------|------|", *table, "", *sections])


def _save_tc_history_snapshot(project_name: str, existing_tcs: list[dict]) -> Path:
    """업데이트 실행 직전의 TC 상태를 tc_history/에 스냅샷 저장."""
    history_dir = BASE_DIR / "tc_history" / (re.sub(r"[^\w\-_]", "_", project_name or "unnamed")[:30])
    history_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    snap_path = history_dir / f"snapshot_{ts}.md"
    parts = [f"# TC 스냅샷 — {project_name or '(unnamed)'}\n\n생성: {datetime.now().isoformat()}\n\n---\n"]
    for tc in existing_tcs:
        parts.append(_tc_to_markdown_block(tc))
        parts.append("\n")
    snap_path.write_text("\n\n".join(parts), encoding="utf-8")
    return snap_path


def _update_tc_for_modified_scr(tc: dict, new_section: str,
                                  tc_rules: str) -> tuple[dict, str]:
    """수정된 SCR에 대해 기존 TC 하나를 새 사양 기반으로 재작성.
    반환: (업데이트된 TC dict, AI가 서술한 수정 사유)
    """
    system = """당신은 시니어 QA 엔지니어입니다. 기존 TC를 새 기획서 섹션 기반으로 **업데이트**합니다.

규칙:
- TC ID는 **절대 변경하지 마세요** — 그대로 유지
- 사전 조건 / 테스트 단계 / 예상 결과를 새 사양에 맞게 재작성
- 기존 대분류/중분류/소분류 유지 (사양에서 변경된 경우만 수정)
- 이미 존재하는 Given/When/Then 구조와 어투를 따르세요
- 수정 사유를 한 문장(50자 이내)으로 요약하여 마지막에 `수정 사유: ...` 형식으로 포함
"""
    user = f"""기존 TC (업데이트 대상):
---
{_tc_to_markdown_block(tc)[:3000]}
---

새 기획서 섹션 (이 TC가 커버하는 화면의 변경된 내용):
---
{new_section[:4000]}
---

위 기존 TC를 새 기획서에 맞게 업데이트해주세요. TC ID는 그대로 유지.
출력 형식: 기존 TC와 동일한 markdown 블록 + 마지막에 `수정 사유: ...` 한 줄 추가.
"""
    try:
        result = call_claude(system, user, max_tokens=3000)
    except Exception as e:
        return tc, f"AI 오류로 원본 유지: {e}"

    # 수정 사유 추출
    reason = ""
    m = re.search(r"수정\s*사유\s*[:：]\s*(.+)", result)
    if m:
        reason = m.group(1).strip().split("\n")[0][:200]
        result = re.sub(r"수정\s*사유\s*[:：].*$", "", result, flags=re.MULTILINE).rstrip()

    # 결과 파싱하여 새 TC dict 생성
    try:
        parsed = parse_tc_markdown(result)
        if parsed:
            updated = parsed[0]
            updated["id"] = tc["id"]  # ID 강제 유지
            updated["status"] = "🔄 수정됨"
            updated["change_reason"] = reason
            updated["raw_block"] = result.strip()
            # 기존에 있던 필드 중 새 파싱에 없는 것 보존 (screen_code 등)
            for k, v in tc.items():
                if k not in updated and v:
                    updated[k] = v
            return updated, reason
    except Exception as e:
        return tc, f"파싱 실패로 원본 유지: {e}"
    return tc, reason


@app.route("/analyze-diff", methods=["POST"])
def analyze_diff():
    """기획서 변경 기반 TC 수정 — 1단계: 두 기획서 + 기존 TC 파일 분석.

    입력 JSON:
      {
        "project_name": (optional) — 저장된 프로젝트 이름. 로그/저장용.
        "old_spec_path":  (필수) SPECS_DIR 내 파일명
        "new_spec_path":  (필수) SPECS_DIR 내 파일명
        "existing_tc_path": (선택) TC_FILES_DIR 내 파일명. 없으면 "diff만 분석".
      }

    반환: build_diff_report() 결과 + 추가 메타.
    """
    data = request.get_json(force=True) or {}
    project_name = (data.get("project_name") or "").strip()
    old_path_name = (data.get("old_spec_path") or "").strip()
    new_path_name = (data.get("new_spec_path") or "").strip()
    tc_path_name = (data.get("existing_tc_path") or "").strip()

    if not new_path_name:
        return jsonify({"ok": False, "error": "새 기획서가 필요합니다."}), 400
    if not old_path_name:
        return jsonify({"ok": False, "error": "이전 기획서가 필요합니다."}), 400

    old_path = SPECS_DIR / old_path_name
    new_path = SPECS_DIR / new_path_name
    if not old_path.exists():
        return jsonify({"ok": False, "error": f"이전 기획서 파일 없음: {old_path_name}"}), 400
    if not new_path.exists():
        return jsonify({"ok": False, "error": f"새 기획서 파일 없음: {new_path_name}"}), 400

    try:
        old_md = old_path.read_text(encoding="utf-8")
        new_md = new_path.read_text(encoding="utf-8")
    except Exception as e:
        return jsonify({"ok": False, "error": f"기획서 읽기 실패: {e}"}), 400

    # 기존 TC 로드 (선택)
    existing_tcs = []
    tc_load_error = None
    if tc_path_name:
        tc_path = TC_FILES_DIR / tc_path_name
        if not tc_path.exists():
            # SPECS_DIR도 확인 (업로드 경로가 달랐을 경우 대비)
            alt = SPECS_DIR / tc_path_name
            if alt.exists():
                tc_path = alt
            else:
                return jsonify({"ok": False, "error": f"기존 TC 파일 없음: {tc_path_name}"}), 400
        try:
            existing_tcs = load_existing_tcs(tc_path)
        except Exception as e:
            tc_load_error = str(e)

    # 리포트 생성
    report = build_diff_report(old_md, new_md, existing_tcs)

    # 변경 없음 확인
    summary = report["summary"]
    no_changes = (summary["added_n"] == 0 and summary["modified_n"] == 0
                  and summary["removed_n"] == 0)

    # 세션 생성 (이후 /update-tc 호출 시 재사용)
    sess = new_session()
    sess["project_name"] = project_name
    sess["_diff_context"] = {
        "old_md": old_md,
        "new_md": new_md,
        "existing_tcs": existing_tcs,
        "report": report,
        "old_spec_name": old_path_name,
        "new_spec_name": new_path_name,
        "tc_path_name": tc_path_name,
    }

    return jsonify({
        "ok": True,
        "sid": sess["id"],
        "report": report,
        "no_changes": no_changes,
        "existing_tc_count": len(existing_tcs),
        "tc_load_error": tc_load_error,
        "has_spec_codes": report.get("has_spec_codes", False),
    })


@app.route("/update-tc", methods=["POST"])
def update_tc():
    """analyze-diff로 생성된 리포트 승인 후 실제 TC 갱신 실행.

    입력 JSON:
      {
        "sid": (필수) — analyze-diff에서 받은 세션 ID
        "approved": {
          "added":   ["SCR-020", "SCR-021"],      # 승인된 신규 SCR 코드 목록
          "modified":["SCR-001"],                 # 승인된 수정 SCR 코드 목록
          "removed": ["SCR-008"],                 # 승인된 삭제 SCR 코드 목록 (Deprecated 처리)
        }
      }

    흐름:
      1. 스냅샷 저장 (tc_history/)
      2. 승인된 신규 SCR → 신규 TC 생성 (새 파이프라인 트리거)
      3. 승인된 수정 SCR → 기존 TC 재작성 (ID 유지)
      4. 승인된 삭제 SCR → 해당 TC에 status=Deprecated
      5. 영향 없는 TC → 상태 빈 값 (유지)
      6. 결과 TC 리스트 → Excel 빌드
    """
    data = request.get_json(force=True) or {}
    sid = (data.get("sid") or "").strip()
    approved = data.get("approved") or {}

    sess = SESSIONS.get(sid)
    if not sess:
        return jsonify({"ok": False, "error": "세션 없음 — /analyze-diff를 먼저 호출하세요."}), 404
    ctx = sess.get("_diff_context")
    if not ctx:
        return jsonify({"ok": False, "error": "diff 컨텍스트 없음"}), 400
    if not os.environ.get("ANTHROPIC_API_KEY"):
        return jsonify({"ok": False, "error": "ANTHROPIC_API_KEY 미설정"}), 500

    approved_added    = set(approved.get("added") or [])
    approved_modified = set(approved.get("modified") or [])
    approved_removed  = set(approved.get("removed") or [])

    # 백엔드 가드 — 승인 항목이 하나도 없으면 빈 Excel 생성 방지
    if not (approved_added or approved_modified or approved_removed):
        return jsonify({
            "ok": False,
            "error": "승인된 변경사항이 없습니다. 최소 1개 이상의 항목을 체크한 뒤 승인하세요."
        }), 400

    def worker():
        try:
            existing_tcs = list(ctx["existing_tcs"])  # copy
            project_name = sess.get("project_name", "")

            # 1. 스냅샷
            push_log(sess, "[갱신] TC 스냅샷 저장 중...")
            snap = _save_tc_history_snapshot(project_name, existing_tcs)
            push_log(sess, f"[갱신] 스냅샷 저장됨: {snap.name}")
            push_stage(sess, 2, "기존 TC 스냅샷 저장 완료", 10)

            # 2. 새 기획서의 SCR별 섹션 추출 (TC 재작성용 근거)
            new_sections = _extract_spec_sections(ctx["new_md"])
            report = ctx["report"]

            # 공통: tc_rules / fewshot 로드
            tc_rules = load_tc_rules()
            fewshot = load_fewshot_examples()
            project_policies = load_project_policies(project_name)
            project_code = _detect_project_code(project_name)

            # 3. 수정된 SCR별 TC 재작성
            tc_by_scr = group_tcs_by_scr(existing_tcs)
            modified_count = 0
            total_modified_tcs = sum(
                len(tc_by_scr.get(code, []))
                for code in approved_modified
            )
            done_mod_tcs = 0
            for code in approved_modified:
                affected = tc_by_scr.get(code, [])
                new_section = new_sections.get(code, {}).get("body", "")
                if not new_section:
                    push_log(sess, f"[갱신] ⚠️ {code} 새 섹션 없음 — 건너뜀")
                    continue
                for tc in affected:
                    push_log(sess, f"[갱신] 수정 ({code}) → {tc.get('id')}")
                    updated_tc, reason = _update_tc_for_modified_scr(tc, new_section, tc_rules)
                    # existing_tcs에서 ID로 찾아 교체
                    for i, t in enumerate(existing_tcs):
                        if t.get("id") == updated_tc.get("id"):
                            existing_tcs[i] = updated_tc
                            break
                    modified_count += 1
                    done_mod_tcs += 1
                    pct = 10 + int(40 * done_mod_tcs / max(total_modified_tcs, 1))
                    push_stage(sess, 2, f"수정 TC 재작성 {done_mod_tcs}/{total_modified_tcs}", pct, eta_sec=(total_modified_tcs - done_mod_tcs) * 30)

            # 4. 삭제된 SCR → Deprecated 태그
            deprecated_count = 0
            for code in approved_removed:
                affected = tc_by_scr.get(code, [])
                for tc in affected:
                    for i, t in enumerate(existing_tcs):
                        if t.get("id") == tc.get("id"):
                            existing_tcs[i]["status"] = "🗑️ Deprecated"
                            existing_tcs[i]["change_reason"] = f"{code} 삭제로 폐기"
                            deprecated_count += 1
                            break
            push_log(sess, f"[갱신] Deprecated 처리 {deprecated_count}건")

            # 5. 신규 SCR → 신규 TC 생성
            added_tc_count = 0
            if approved_added:
                push_stage(sess, 3, f"신규 SCR {len(approved_added)}개 TC 생성 중", 55, eta_sec=len(approved_added) * 40)
                # 분류표 임시 생성 — 각 신규 SCR을 별도 중분류로
                temp_classification_parts = [f"# TC 분류표 — 신규 추가"]
                for code in approved_added:
                    new_s = new_sections.get(code, {})
                    name = new_s.get("name") or code
                    temp_classification_parts.append(f"\n## 대분류: 신규 화면\n\n### 중분류: {name}\n\n#### 소분류\n- 화면 진입/표시/동작: {code} 기본 TC")
                temp_classification = "\n".join(temp_classification_parts)

                # 각 신규 SCR마다 단일 호출로 TC 생성
                for idx, code in enumerate(approved_added):
                    new_s = new_sections.get(code, {})
                    name = new_s.get("name") or code
                    body = new_s.get("body", "")
                    # build_tc_user_prompt 사용을 위한 domain 객체 구성
                    domain = {"name": "신규 화면", "code": "NEWSCR",
                              "suite_code": re.sub(r"[^A-Za-z]", "", name).upper()[:6] or "NEW",
                              "_focus_middle": name}
                    # 연관 화면 코드 hint 주입을 위해 body에 code 포함
                    prompt_context = f"{code}: {name}\n\n{body}"
                    starting_seq = next_tc_id(existing_tcs, project_code, domain["suite_code"])
                    user_prompt = build_tc_user_prompt(
                        domain, prompt_context, prompt_context, project_name,
                        temp_classification, starting_seq=starting_seq,
                        source_scr_map=(sess.get("_source_scr_map") or {}),
                    )
                    system_prompt = build_tc_system_prompt(
                        tc_rules, temp_classification, project_policies, fewshot
                    )
                    push_log(sess, f"[갱신] 신규 ({code}) TC 생성 중...")
                    try:
                        tc_md = call_claude(system_prompt, user_prompt, max_tokens=8000)
                    except Exception as e:
                        push_log(sess, f"[갱신] {code} 생성 실패: {e}")
                        continue
                    # TC ID 재번호
                    tc_md, _ = renumber_tc_ids(tc_md, project_code, domain["suite_code"], starting_seq)
                    new_tcs = parse_tc_markdown(tc_md)
                    for nt in new_tcs:
                        nt["status"] = "🆕 신규"
                        nt["change_reason"] = f"{code} 추가로 신규 생성"
                        nt["screen_code"] = code
                        nt["screen"] = code
                        existing_tcs.append(nt)
                        added_tc_count += 1

            # 6. 결과 마크다운 생성 → Excel 빌드
            push_stage(sess, 4, "tc_final.md 저장 중", 85, eta_sec=5)
            final_md = "\n\n---\n\n".join(_tc_to_markdown_block(t) for t in existing_tcs)
            workspace = sess["workspace"]
            tc_final_path = workspace / "tc_final.md"
            tc_final_path.write_text(final_md, encoding="utf-8")

            # Excel 빌드 — 기존 TC 수정 플로우이므로 상태/수정 사유 컬럼 + 🔄 변경 이력 시트 포함
            sess["_include_change_columns"] = True
            push_stage(sess, 5, "Excel 빌드 중", 92, eta_sec=60)
            excel_path = step_build_excel(sess, final_md, project_name,
                                           total_tc=len(existing_tcs),
                                           min_tc=max(1, round(len(existing_tcs) * 0.35)))
            # 후처리 적용본 채택 (MD ↔ Excel 일치)
            final_md = sess.get("_tc_content_disambiguated", final_md)

            # tc_files에 저장
            today = datetime.now().strftime("%Y%m%d")
            safe_name = re.sub(r"[^\w\-_]", "_", project_name or "Updated")[:30]
            tc_md_path = TC_FILES_DIR / f"{safe_name}_{today}.md"
            tc_md_path.write_text(final_md, encoding="utf-8")
            save_project(project_name, str(tc_md_path), str(excel_path))

            # 변경 이력 요약 저장
            change_log_path = workspace / "change_log.md"
            change_log_path.write_text(
                f"# 변경 이력\n\n"
                f"- 수정된 TC: {modified_count}건\n"
                f"- Deprecated TC: {deprecated_count}건\n"
                f"- 신규 TC: {added_tc_count}건\n"
                f"- 스냅샷: {snap}\n",
                encoding="utf-8"
            )

            # 유니크 TC ID 수 — Excel dedup 결과와 일치하도록 재계산
            unique_tc = count_unique_tc_ids(final_md) or len(existing_tcs)

            sess["result"] = excel_path
            sess["status"] = "done"
            push_stage(sess, 5, "완료", 100)
            push(sess, "done", {
                "filename": excel_path.name,
                "size": excel_path.stat().st_size,
                "sid": sess["id"],
                "total_tc": unique_tc,
                "smoke_tc": sess.get("smoke_tc"),
                "min_tc": max(1, round(unique_tc * 0.35)),
                "update_summary": {
                    "modified": modified_count,
                    "deprecated": deprecated_count,
                    "added": added_tc_count,
                },
            })
            push_log(sess, f"[완료] 업데이트 반영 — 수정 {modified_count} / 추가 {added_tc_count} / Deprecated {deprecated_count}")
        except Exception as e:
            push_log(sess, f"[오류] {e}")
            push_error(sess, str(e))

    t = threading.Thread(target=worker, daemon=True)
    sess["thread"] = t
    t.start()
    return jsonify({"ok": True, "sid": sess["id"]})


@app.route("/start-modify", methods=["POST"])
def start_modify():
    data = request.get_json(force=True) or {}
    project_name  = data.get("project_name", "").strip()
    change_desc   = data.get("change_desc", "").strip()

    if not project_name:
        return jsonify({"ok": False, "error": "프로젝트명이 없습니다."}), 400
    if not change_desc:
        return jsonify({"ok": False, "error": "변경사항 내용이 없습니다."}), 400
    if not os.environ.get("ANTHROPIC_API_KEY"):
        return jsonify({"ok": False, "error": "ANTHROPIC_API_KEY가 설정되지 않았습니다."}), 500

    # 기존 TC 파일 로드
    projects = load_projects()
    proj = next((p for p in projects if p["name"] == project_name), None)
    if not proj or not Path(proj["tc_file"]).exists():
        return jsonify({"ok": False, "error": f"'{project_name}' 프로젝트의 TC 파일을 찾을 수 없습니다."}), 404

    existing_tc = Path(proj["tc_file"]).read_text(encoding="utf-8")

    sess = new_session()
    sess["project_name"] = project_name
    sess["mode"] = "modify"
    t = threading.Thread(
        target=run_modify_pipeline,
        args=(sess, project_name, existing_tc, change_desc),
        daemon=True
    )
    sess["thread"] = t
    t.start()
    return jsonify({"ok": True, "sid": sess["id"]})


@app.route("/checkpoint", methods=["GET"])
def get_checkpoint():
    cp = load_checkpoint()
    return jsonify({"ok": True, "checkpoint": cp})

@app.route("/checkpoint", methods=["DELETE"])
def del_checkpoint():
    clear_checkpoint()
    return jsonify({"ok": True})

@app.route("/restart-modify", methods=["POST"])
def restart_modify():
    """체크포인트 기반 재시작 또는 처음부터 재시작"""
    data = request.get_json(force=True) or {}
    mode = data.get("mode", "fresh")  # "resume" | "fresh"

    cp = load_checkpoint()
    if mode == "resume" and not cp:
        return jsonify({"ok": False, "error": "저장된 체크포인트가 없습니다."}), 400

    project_name = cp.get("project_name") if mode == "resume" else data.get("project_name", "")
    change_desc  = cp.get("change_desc")  if mode == "resume" else data.get("change_desc", "")

    if not project_name:
        return jsonify({"ok": False, "error": "프로젝트명이 없습니다."}), 400

    projects = load_projects()
    proj = next((p for p in projects if p["name"] == project_name), None)
    if not proj or not proj.get("tc_file") or not Path(proj["tc_file"]).exists():
        return jsonify({"ok": False, "error": f"'{project_name}' TC 파일을 찾을 수 없습니다."}), 404

    existing_tc = Path(proj["tc_file"]).read_text(encoding="utf-8")
    sess = new_session()
    sess["project_name"] = project_name
    sess["mode"] = "modify"
    resume_stage = cp.get("stage") if mode == "resume" else None
    t = threading.Thread(
        target=run_modify_pipeline,
        args=(sess, project_name, existing_tc, change_desc, resume_stage),
        daemon=True
    )
    sess["thread"] = t
    t.start()
    return jsonify({"ok": True, "sid": sess["id"], "mode": mode})


@app.route("/stop/<sid>", methods=["POST"])
def stop_pipeline(sid):
    sess = SESSIONS.get(sid)
    if not sess:
        return jsonify({"ok": False, "error": "세션 없음"}), 404
    sess["stop_requested"] = True
    # Human Gate에서 대기 중이면 즉시 해제 후 중단
    if sess["status"] == "gate_waiting":
        sess["approved"] = ""
        sess["gate_event"].set()
    return jsonify({"ok": True})


@app.route("/screen-code-map", methods=["GET"])
def get_screen_code_map():
    """프로젝트별 Screen Code 매핑 테이블 조회.

    프론트엔드의 Human Gate에서 "시스템 규칙 자동 적용" 모드 ON 시,
    각 중분류가 어떤 ScreenCode로 매핑될지 미리보기하는 데 사용.
    """
    project_name = request.args.get("project", "").strip()
    project_code = _detect_project_code(project_name)
    screen_based_default = _is_screen_based(project_code)
    mapping = load_screen_code_map(project_name) if project_name else {}
    return jsonify({
        "ok": True,
        "project_code": project_code,
        "screen_based_default": screen_based_default,
        "map": mapping,
    })


@app.route("/regenerate-classification", methods=["POST"])
def regenerate_classification():
    """분류표 재생성 — 정책/기능 목록은 유지하고 분류표부터 다시 생성"""
    data = request.get_json(force=True) or {}
    project_name = (data.get("project_name") or "").strip() or "프로젝트"
    focus_area   = (data.get("focus_area") or "").strip()

    if not os.environ.get("ANTHROPIC_API_KEY"):
        return jsonify({"ok": False, "error": "ANTHROPIC_API_KEY가 설정되지 않았습니다."}), 500

    # 저장된 파이프라인 상태에서 이전 결과 로드
    state = load_pipeline_state(project_name)
    if not state or not state["data"].get("features_text"):
        return jsonify({"ok": False, "error": "이전 분석 결과가 없습니다. 처음부터 시작해주세요."}), 400

    sess = new_session()
    sess["project_name"] = project_name
    sess["focus_area"] = focus_area
    # 분류표 직전 단계(features)에서 재시작하도록 상태 설정
    sess["_resumed_state"] = {
        "stage": "features",
        "data": {
            "raw_text": state["data"].get("raw_text", ""),
            "policy_text": state["data"].get("policy_text", ""),
            "features_text": state["data"].get("features_text", ""),
            "focus_area": focus_area,
        }
    }
    t = threading.Thread(
        target=run_pipeline,
        args=(sess, [], project_name),
        daemon=True
    )
    sess["thread"] = t
    t.start()
    return jsonify({"ok": True, "sid": sess["id"]})


@app.route("/approve/<sid>", methods=["POST"])
def approve(sid):
    sess = SESSIONS.get(sid)
    if not sess:
        # v0.9.18: 사용자 친화적 메시지 — 서버 재시작 등으로 세션 소실 시
        return jsonify({
            "ok": False,
            "error": "세션이 만료되었습니다. 서버가 재시작됐을 수 있어요.\n페이지를 새로고침하고 '이어서 작업'을 클릭하면 분류표 검토 단계부터 복원됩니다."
        }), 404
    data = request.get_json(force=True) or {}
    content = data.get("content", "").strip()
    if not content:
        return jsonify({"ok": False, "error": "승인할 내용이 없습니다."}), 400
    # 선택된 대분류 코드 저장 (None이면 전체 생성)
    selected = data.get("selected_domains")  # list of codes or null
    sess["selected_domains"] = selected if selected else None
    # SuiteCode 저장 (검토자가 Human Gate에서 입력)
    suite_codes = data.get("suite_codes")  # list of codes
    sess["suite_codes"] = suite_codes if suite_codes else []
    # 시스템 규칙 자동 적용 (screen_code_map 기반 Screen-based 스킴) 여부.
    # None이면 프로젝트 코드 기반 자동 판별 (_is_screen_based).
    # True/False면 사용자 명시적 선택 우선.
    auto_flag = data.get("auto_screen_code")
    sess["auto_screen_code"] = auto_flag  # None / True / False
    # ── Excel 시트 옵션 (Full/Light/Custom 프리셋 → 시트별 dict) ──
    # 프론트가 보내는 형식:
    #   {"cover": bool, "stats": bool, "smoke": bool,
    #    "traceability": bool, "tc_list": bool, "change_history": bool}
    # 누락/None 이면 step_build_excel 에서 Full Set 으로 폴백
    excel_sheets = data.get("excel_sheets")
    if isinstance(excel_sheets, dict):
        sess["_excel_sheets"] = {k: bool(v) for k, v in excel_sheets.items()}
    sess["approved"] = content
    sess["gate_event"].set()
    return jsonify({"ok": True})


@app.route("/gate-chat/<sid>", methods=["POST"])
def gate_chat(sid):
    """Human Gate 채팅 — AI와 대화하며 분류표/영향도 문서를 수정"""
    import anthropic
    sess = SESSIONS.get(sid)
    if not sess:
        # v0.9.18: 사용자 친화적 메시지
        return jsonify({
            "ok": False,
            "error": "세션이 만료되었습니다. 서버가 재시작됐을 수 있어요.\n페이지를 새로고침하고 '이어서 작업'을 클릭하면 분류표 검토 단계부터 복원됩니다."
        }), 404
    if sess.get("status") != "gate_waiting":
        return jsonify({"ok": False, "error": "Gate 대기 상태가 아닙니다."}), 400

    data = request.get_json(force=True) or {}
    user_msg = data.get("message", "").strip()
    current_doc = data.get("current_doc", "").strip()
    gate_mode = data.get("gate_mode", "new")   # "new" | "modify"
    history = data.get("history", [])           # [{role, content}, ...]

    if not user_msg:
        return jsonify({"ok": False, "error": "메시지를 입력해주세요."}), 400

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return jsonify({"ok": False, "error": "ANTHROPIC_API_KEY 미설정"}), 500

    client = anthropic.Anthropic(api_key=api_key)

    if gate_mode == "modify":
        system_prompt = """당신은 TC(Test Case) 수정 전문가입니다.
현재 AI가 생성한 '변경 영향도 분석' 문서를 사용자의 요청에 따라 수정합니다.

규칙:
- 사용자의 수정 요청을 정확히 반영하여 문서를 업데이트하세요.
- 응답은 반드시 두 부분으로 구성하세요:
  1. [REPLY] 태그: 수정 내용에 대한 간단한 설명 (1-3문장)
  2. [DOCUMENT] 태그: 수정된 전체 문서 내용 (마크다운 형식 유지)
- 수정 요청이 없거나 질문인 경우, [DOCUMENT] 태그에는 기존 문서를 그대로 포함하세요.
- 문서 형식(마크다운, 헤딩, 목록 등)을 유지하세요."""
    else:
        system_prompt = """당신은 TC(Test Case) 분류 전문가입니다.
현재 AI가 생성한 '분류표' 문서를 사용자의 요청에 따라 적극적으로 수정합니다.

## 절대 규칙

1. **짧은 명령형 요청도 명확한 수정 의도로 해석하세요** (한국어 특성).
   - "splash 케이스 삭제" → Splash 관련 중분류/소분류를 모두 제거
   - "이메일 입력 빼줘" → Email Input 중분류 제거
   - "AUTH 대분류 3번 지워" → AUTH 대분류의 3번째 항목 제거
   - "체크박스 기능 빼" → 체크박스 관련 항목 모두 제거

2. **삭제 요청을 절대 거부하지 마세요**. 사용자가 "X 삭제/제거/빼줘" 라고 하면 X 와 관련된 항목을 분류표에서 모두 찾아 제거하세요.

3. **모호한 경우에도 일단 수정 시도하고, [REPLY] 에 어떤 해석을 했는지 명시**하세요. 거부 대신 "Splash 관련 항목을 모두 제거했습니다. 다른 의도였으면 알려주세요." 식으로.

4. **수정이 필요 없는 진짜 질문(단순 문의)** 일 때만 기존 문서 그대로 반환:
   - "Splash 가 뭐야?" — 질문이므로 [DOCUMENT] 그대로
   - "이 분류표 잘 만든 거 같아?" — 평가 질문이므로 그대로
   - "splash 삭제" — 명령이므로 반드시 수정

## 응답 형식 (반드시 준수)

응답은 정확히 두 부분으로 구성:
1. [REPLY] 태그: 수정 내용 명확히 (예: "Onboarding 대분류의 Splash 중분류 1개를 제거했습니다.")
2. [DOCUMENT] 태그: 수정된 **전체** 분류표 (마크다운 형식 유지, 잘리지 않게 끝까지)

분류표 형식(## 대분류, ### 중분류, 표 컬럼, TC ID 패턴) 유지."""

    # 메시지 구성 — 히스토리 + 현재 문서 컨텍스트 + 사용자 메시지
    messages = []
    for h in history:
        messages.append({"role": h["role"], "content": h["content"]})

    # 현재 문서를 컨텍스트로 포함
    full_user_msg = f"""현재 문서:\n```\n{current_doc}\n```\n\n사용자 요청: {user_msg}"""
    messages.append({"role": "user", "content": full_user_msg})

    # v0.9.20: max_tokens 4096 → 16384 (긴 분류표 출력 잘림 방지)
    try:
        resp = client.messages.create(
            model=MODEL,
            max_tokens=16384,
            temperature=0,
            system=system_prompt,
            messages=messages
        )
        ai_text = resp.content[0].text
        stop_reason = getattr(resp, "stop_reason", None)

        # [REPLY]와 [DOCUMENT] 파싱
        reply_text = ""
        updated_doc = current_doc  # 기본값: 변경 없음
        truncated = False

        reply_match = re.search(r'\[REPLY\](.*?)(?=\[DOCUMENT\]|$)', ai_text, re.DOTALL)
        doc_match   = re.search(r'\[DOCUMENT\](.*?)$', ai_text, re.DOTALL)

        if reply_match:
            reply_text = reply_match.group(1).strip()
        if doc_match:
            updated_doc = doc_match.group(1).strip()
            # 코드블록 래핑 제거
            if updated_doc.startswith("```"):
                lines = updated_doc.split("\n")
                updated_doc = "\n".join(lines[1:-1] if lines[-1].strip() == "```" else lines[1:]).strip()
            # 잘림 감지: stop_reason == "max_tokens" 또는 출력이 비정상적으로 짧음
            if stop_reason == "max_tokens":
                truncated = True
                push_log(sess, "[Gate] AI 응답이 max_tokens 도달 — 분류표가 잘렸을 수 있음")

        if not reply_text:
            reply_text = ai_text.strip()

        # 잘림 감지 시 사용자 알림 + 원본 유지 (잘린 문서로 덮어쓰기 방지)
        if truncated:
            updated_doc = current_doc
            reply_text = "⚠️ 응답이 너무 길어 잘렸습니다. 더 작은 단위로 나눠서 요청해 주세요. (예: 'Splash 한 번에' 대신 'Splash 중분류 1개만 삭제')\n\n" + reply_text

        # ── v0.9.15: 분류표 모드에서 변경이 감지되면 features/policy 도 동기화 ──
        # 사용자가 "체크박스 기능 제거" 등 의미 있는 수정을 했을 때, 본문(features/policy)에서도
        # 해당 항목을 제거하여 TC 작성 시 일관된 컨텍스트가 전달되도록 함.
        sync_status = None  # 'synced' | 'unchanged' | 'failed'
        if gate_mode == "new" and updated_doc and updated_doc != current_doc:
            try:
                sync_status = _sync_features_policy_with_classification(
                    sess, client, updated_doc, current_doc, user_msg
                )
                if sync_status == "synced":
                    reply_text += "\n\n📌 본문 정책·기능 정보도 분류표와 일관되도록 자동 동기화했습니다."
            except Exception as _sync_e:
                sync_status = "failed"
                # 동기화 실패해도 분류표 수정은 그대로 반영 (최소 보장)
                push_log(sess, f"[Gate] features/policy 동기화 실패 (분류표만 반영): {_sync_e}")

        return jsonify({
            "ok": True,
            "reply": reply_text,
            "updated_doc": updated_doc,
            "sync_status": sync_status,
            "changed": updated_doc != current_doc,  # v0.9.20: 변경 여부 명시
            "truncated": truncated,                  # v0.9.20: 잘림 여부
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


def _sync_features_policy_with_classification(sess, client, new_classification, old_classification, user_msg):
    """v0.9.15: Gate 채팅에서 분류표가 수정되면 features/policy 도 함께 동기화.

    사용자의 의도(예: "체크박스 기능 제거")를 본문 문서에도 반영하여
    TC 작성 시 분류표·본문 사이의 모순을 제거.

    Returns:
        "synced" — 동기화 성공
        "unchanged" — features/policy 가 sess 에 없거나 변경 사항이 사소함
        "failed" — AI 호출 실패 (예외 raise 됨)
    """
    features_text = sess.get("_features_text", "")
    policy_text = sess.get("_policy_text", "")
    if not features_text and not policy_text:
        return "unchanged"  # gate 단계 진입 전이거나 복원 케이스 — 동기화 불가

    sync_system = """당신은 TC 분류 전문가입니다. 사용자가 분류표를 수정했을 때, 본문 정책·기능 문서를 분류표와 일관되게 동기화합니다.

규칙:
- 사용자의 분류표 변경(추가/제거/수정)을 본문 문서에 반영하세요.
- 분류표에서 **제거된 항목** 은 본문 정책·기능 문서에서도 제거하세요.
- 분류표에서 **수정된 항목** 은 본문 문서에서도 동일하게 수정하세요.
- 분류표에 그대로 있는 항목은 본문 문서에서도 유지하세요.
- 분류표 변경과 무관한 본문 내용은 절대 변경하지 마세요.
- 응답은 반드시 다음 두 부분으로 구성:
  1. [POLICY] 태그: 동기화된 정책 문서 전체
  2. [FEATURES] 태그: 동기화된 기능 문서 전체
- 변경할 내용이 없으면 두 태그 안에 원본을 그대로 포함하세요."""

    sync_user = f"""## 사용자 요청
{user_msg}

## 변경 전 분류표
```
{old_classification[:8000]}
```

## 변경 후 분류표 (사용자 승인 — 최종)
```
{new_classification[:8000]}
```

## 현재 정책 문서 (동기화 대상)
```
{policy_text[:10000]}
```

## 현재 기능 문서 (동기화 대상)
```
{features_text[:10000]}
```

위 분류표 변경에 맞춰 정책·기능 문서를 동기화한 결과를 [POLICY] 와 [FEATURES] 태그로 반환하세요."""

    resp = client.messages.create(
        model=MODEL,
        max_tokens=8192,
        temperature=0,
        system=sync_system,
        messages=[{"role": "user", "content": sync_user}],
    )
    sync_text = resp.content[0].text

    # [POLICY] / [FEATURES] 파싱
    policy_match = re.search(r'\[POLICY\](.*?)(?=\[FEATURES\]|$)', sync_text, re.DOTALL)
    features_match = re.search(r'\[FEATURES\](.*?)$', sync_text, re.DOTALL)

    new_policy = policy_text
    new_features = features_text

    if policy_match:
        candidate = policy_match.group(1).strip()
        # 코드블록 래핑 제거
        if candidate.startswith("```"):
            lines = candidate.split("\n")
            candidate = "\n".join(lines[1:-1] if lines[-1].strip() == "```" else lines[1:]).strip()
        if candidate:
            new_policy = candidate

    if features_match:
        candidate = features_match.group(1).strip()
        if candidate.startswith("```"):
            lines = candidate.split("\n")
            candidate = "\n".join(lines[1:-1] if lines[-1].strip() == "```" else lines[1:]).strip()
        if candidate:
            new_features = candidate

    # 세션 갱신
    sess["_policy_text"] = new_policy
    sess["_features_text"] = new_features
    return "synced"


@app.route("/merge-classification/<sid>", methods=["POST"])
def merge_classification(sid):
    """Stage 1 — 분류표 자동 정리 (TC 작성 전).

    예측된 중복 가능성에 따라 AI 가 분류표 재구성:
    - 같은 그룹 내 비기능 TC 후보를 entry 화면에 통합 (비고에 [통합] 명시)
    - 중분류 자체는 보존 (사용자 의도 명시 없으면 함부로 통합 X)

    응답: { "ok": true, "updated_doc": str, "predictions_resolved": int }
    """
    import anthropic
    sess = SESSIONS.get(sid)
    if not sess:
        return jsonify({
            "ok": False,
            "error": "세션이 만료되었습니다. 페이지를 새로고침하고 '이어서 작업'을 클릭하세요."
        }), 404
    if sess.get("status") != "gate_waiting":
        return jsonify({"ok": False, "error": "Gate 대기 상태가 아닙니다."}), 400

    data = request.get_json(force=True) or {}
    current_doc = (data.get("current_doc") or "").strip()
    predictions = data.get("predictions") or []
    if not current_doc:
        return jsonify({"ok": False, "error": "분류표가 비어있습니다."}), 400
    if not predictions:
        return jsonify({"ok": False, "error": "통합할 예측 정보가 없습니다."}), 400

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return jsonify({"ok": False, "error": "ANTHROPIC_API_KEY 미설정"}), 500
    client = anthropic.Anthropic(api_key=api_key)

    # 예측 결과 텍스트화
    pred_text_lines = []
    for i, p in enumerate(predictions, 1):
        major = p.get("major", "")
        rec = p.get("recommendation", "")
        if p.get("type") == "shared_minor":
            kw = p.get("shared_keyword", "")
            mids = p.get("shared_in_middles", [])
            pred_text_lines.append(f"{i}. [{major}] '{kw}' 키워드가 {len(mids)}개 중분류에 등장 ({', '.join(mids)}). {rec}")
        else:
            mc = p.get("middle_count", 0)
            mids = p.get("middles", [])
            pred_text_lines.append(f"{i}. [{major}] 중분류 {mc}개 → {rec}")
    pred_text = "\n".join(pred_text_lines)

    system_prompt = """당신은 TC 분류 전문가입니다. 분류표에서 중복 가능성을 줄이도록 정리합니다.

## 절대 규칙

1. **중분류 자체를 함부로 삭제하지 마세요** — 사용자가 명시적으로 요청한 경우만 삭제.
2. **소분류는 정리 가능** — 같은 그룹 내 여러 중분류에 같은 비기능 소분류가 반복되면, 그룹 entry 화면(첫 번째 중분류 또는 'Splash'/'Sign-in Start' 같은 진입 성격) 1개로 통합하고 다른 중분류에선 제거.
3. **통합된 소분류 옆에 [통합] 표시** — 예: "- 로딩 시간 [통합 — 그룹 대표 화면 검증]"
4. **응답은 두 부분 형식 유지**:
   1. [REPLY] 태그: 무엇을 정리했는지 1-3문장 요약
   2. [DOCUMENT] 태그: 정리된 전체 분류표 (마크다운)
5. **분류표 형식 유지** (## 대분류, ### 중분류, 소분류 불릿)
"""

    user_msg = f"""## 현재 분류표
```
{current_doc[:12000]}
```

## 발견된 중복 가능성 ({len(predictions)}개)
{pred_text}

## 통합 지시
위 예측에 따라 분류표의 비기능 소분류 (네트워크/타임아웃/로딩 시간 등) 를 그룹 대표 화면 1개에만 남기고 다른 중분류에선 제거하세요.
스펙 명시 케이스(에러 처리 등 고유 시나리오) 는 유지하세요.
"""

    try:
        resp = client.messages.create(
            model=MODEL,
            max_tokens=16384,
            temperature=0,
            system=system_prompt,
            messages=[{"role": "user", "content": user_msg}],
        )
        ai_text = resp.content[0].text
        stop_reason = getattr(resp, "stop_reason", None)

        reply_match = re.search(r'\[REPLY\](.*?)(?=\[DOCUMENT\]|$)', ai_text, re.DOTALL)
        doc_match = re.search(r'\[DOCUMENT\](.*?)$', ai_text, re.DOTALL)
        reply_text = reply_match.group(1).strip() if reply_match else ai_text.strip()
        updated_doc = current_doc
        if doc_match:
            updated_doc = doc_match.group(1).strip()
            if updated_doc.startswith("```"):
                lines = updated_doc.split("\n")
                updated_doc = "\n".join(lines[1:-1] if lines[-1].strip() == "```" else lines[1:]).strip()

        truncated = (stop_reason == "max_tokens")
        if truncated:
            return jsonify({
                "ok": False,
                "error": "AI 응답이 너무 길어 잘렸습니다. 분류표가 매우 큰 경우 발생할 수 있어요. 수동으로 미니 채팅에서 통합 요청해 주세요.",
            }), 500

        push_log(sess, f"[분류표 정리] 자동 통합 완료 — {len(predictions)}개 패턴 처리")
        return jsonify({
            "ok": True,
            "updated_doc": updated_doc,
            "reply": reply_text,
            "predictions_resolved": len(predictions),
        })

    except Exception as e:
        return jsonify({"ok": False, "error": f"분류표 정리 실패: {e}"}), 500


@app.route("/merge-tcs/<sid>", methods=["POST"])
def merge_tcs(sid):
    """Stage 2 — TC 자동 통합 + Excel 재빌드 (Step 5).

    1. tc_final.md 로드
    2. AI 호출 — duplicate_report 의 통합 후보 TC 제거 + 대표 TC 비고에 [통합] 태그
    3. tc_final.md 새 버전 저장 (원본 보존)
    4. step_build_excel() 다시 호출 → 새 Excel 생성 (_merged suffix)
    5. 응답: { ok, filename, removed_count, new_total_tc }
    """
    import anthropic
    sess = SESSIONS.get(sid)
    if not sess:
        return jsonify({
            "ok": False,
            "error": "세션이 만료되었습니다. 페이지를 새로고침하세요."
        }), 404

    data = request.get_json(force=True) or {}
    duplicate_report = data.get("duplicate_report") or {}
    patterns = duplicate_report.get("patterns") or []
    if not patterns:
        return jsonify({"ok": False, "error": "통합할 중복 패턴이 없습니다."}), 400

    # tc_final.md 로드
    tc_final_path = sess.get("workspace", Path()) / "tc_final.md"
    if not tc_final_path.exists():
        return jsonify({"ok": False, "error": f"TC 파일을 찾을 수 없습니다: {tc_final_path}"}), 404
    tc_md = tc_final_path.read_text(encoding="utf-8")

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return jsonify({"ok": False, "error": "ANTHROPIC_API_KEY 미설정"}), 500
    client = anthropic.Anthropic(api_key=api_key)

    # 통합 지시 텍스트 생성
    merge_instructions = []
    total_to_remove = 0
    for p in patterns:
        keep = p.get("keep", "")
        remove = p.get("remove") or []
        pattern_name = p.get("pattern", "")
        major = p.get("major", "")
        if keep and remove:
            merge_instructions.append(
                f'- [{major}] "{pattern_name}" 패턴: '
                f'유지 = `{keep}`, 제거 = {", ".join(f"`{r}`" for r in remove)}'
            )
            total_to_remove += len(remove)
    if not merge_instructions:
        return jsonify({"ok": False, "error": "통합할 TC 가 없습니다."}), 400

    instructions_text = "\n".join(merge_instructions)

    system_prompt = """당신은 TC 통합 전문가입니다. 기존 TC 마크다운에서 중복 패턴을 통합 정리합니다.

## 절대 규칙

1. **통합 후보 TC 들을 마크다운에서 완전히 제거** (`### TC-ID — title` 부터 다음 `### ` 또는 `---` 직전까지).
2. **유지 권장 TC 의 비고에 [통합] 태그 추가** — 예: `**비고**\\n- [통합] 그룹 단위 비기능 검증 — 같은 패턴을 다른 화면별로 반복 작성하지 않음 (원칙 G)`
3. **다른 모든 TC 는 그대로 유지** — 명시되지 않은 TC 는 절대 변경 금지.
4. **마크다운 형식 보존** — `### TC-ID`, `| 분류 |`, `**테스트 단계**` 등 모든 형식 유지.
5. **응답은 통합된 마크다운 전체** — 메타 헤더(`# TC 최종본 — ...`) 도 포함하여 그대로 반환.

응답 형식:
[REPLY] 1-2문장 요약
[DOCUMENT] 통합된 전체 마크다운
"""

    user_msg = f"""## 통합 지시
{instructions_text}

## 현재 TC 마크다운 (전체 보존, 위 지시만 적용)
```
{tc_md[:18000]}
```

위 통합 지시를 정확히 적용하여 **수정된 전체 TC 마크다운**을 반환하세요."""

    try:
        resp = client.messages.create(
            model=MODEL,
            max_tokens=16384,
            temperature=0,
            system=system_prompt,
            messages=[{"role": "user", "content": user_msg}],
        )
        ai_text = resp.content[0].text
        stop_reason = getattr(resp, "stop_reason", None)

        doc_match = re.search(r'\[DOCUMENT\](.*?)$', ai_text, re.DOTALL)
        reply_match = re.search(r'\[REPLY\](.*?)(?=\[DOCUMENT\]|$)', ai_text, re.DOTALL)
        if not doc_match:
            return jsonify({"ok": False, "error": "AI 응답 파싱 실패 — DOCUMENT 태그 누락"}), 500

        new_tc_md = doc_match.group(1).strip()
        if new_tc_md.startswith("```"):
            lines = new_tc_md.split("\n")
            new_tc_md = "\n".join(lines[1:-1] if lines[-1].strip() == "```" else lines[1:]).strip()

        if stop_reason == "max_tokens":
            return jsonify({"ok": False, "error": "AI 응답이 잘렸습니다. TC 가 많은 경우 발생 — 수동 정리 필요"}), 500

        # 새 tc_final.md 저장 (원본은 그대로 둠 — 백업용으로 _v1 보존)
        merged_path = sess["workspace"] / "tc_final_merged.md"
        merged_path.write_text(new_tc_md, encoding="utf-8")

        # TC 개수 재계산
        new_tc_count = len(re.findall(r"^###\s", new_tc_md, re.MULTILINE))
        old_tc_count = len(re.findall(r"^###\s", tc_md, re.MULTILINE))
        actual_removed = max(0, old_tc_count - new_tc_count)
        new_smoke = max(1, round(new_tc_count * 0.35))

        # Excel 재빌드 — 임시로 tc_final.md 를 새 내용으로 교체했다가 빌드 후 복원
        original_tc_md = tc_md
        try:
            tc_final_path.write_text(new_tc_md, encoding="utf-8")
            # build_excel 호출 시 _merged suffix 가 붙도록 sess 에 마킹
            sess["_excel_filename_suffix"] = "_merged"
            project_name = sess.get("project_name", "프로젝트")
            new_excel_path = step_build_excel(sess, new_tc_md, project_name, new_tc_count, new_smoke)
        finally:
            # 원본 tc_final.md 복원 (재시도 등을 위해)
            tc_final_path.write_text(original_tc_md, encoding="utf-8")
            sess.pop("_excel_filename_suffix", None)

        # sess["result"] 갱신 (병합본을 새 결과로)
        sess["result"] = new_excel_path

        push_log(sess, f"[TC 통합] 자동 통합 완료 — {actual_removed}개 TC 제거 (총 {old_tc_count} → {new_tc_count}). 새 Excel: {new_excel_path.name}")

        reply_text = reply_match.group(1).strip() if reply_match else f"{actual_removed}개 TC 통합 완료"

        return jsonify({
            "ok": True,
            "filename": new_excel_path.name,
            "removed_count": actual_removed,
            "new_total_tc": new_tc_count,
            "new_smoke_tc": new_smoke,
            "reply": reply_text,
            "size": new_excel_path.stat().st_size if new_excel_path.exists() else 0,
        })

    except Exception as e:
        import traceback
        push_log(sess, f"[TC 통합] 실패: {e}")
        push_log(sess, traceback.format_exc()[:500])
        return jsonify({"ok": False, "error": f"TC 통합 실패: {e}"}), 500


@app.route("/export-gate", methods=["POST"])
def export_gate():
    """Human Gate 분류표를 Excel로 내보내기 (범위 선택 포함)"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = request.get_json() or {}
    content = data.get("content", "")
    mode    = data.get("mode", "new")   # "new" | "modify"
    domains = extract_domains(content)

    wb = Workbook()

    # ── Sheet 1: 범위 선택 ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "TC 생성 범위 선택"
    ws1.sheet_view.showGridLines = False

    NAVY  = "1E2761"; TEAL  = "028090"; LIGHT = "EBF5FB"; YELLOW = "FFF9C4"
    def hfont(bold=True, color="FFFFFF", size=10):
        return Font(name="Calibri", bold=bold, color=color, size=size)
    def fill(c): return PatternFill("solid", fgColor=c)
    def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def left(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
    def bdr():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    headers = ["대분류 코드", "대분류명", "TC 생성 포함 (Y/N)", "비고"]
    col_w   = [16, 30, 20, 40]
    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        c = ws1.cell(1, ci, h)
        c.font = hfont(); c.fill = fill(NAVY); c.alignment = center(); c.border = bdr()
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[1].height = 28

    # 안내 메모 행
    ws1.merge_cells("A2:D2")
    note = ws1["A2"]
    msg = "TC 생성 포함 열에 Y(포함) 또는 N(제외)를 입력하세요. 기본값은 Y(전체 생성)입니다."
    if mode == "modify":
        msg = "수정 영향도 분석 결과입니다. 실제 TC를 수정할 범위에 Y를 표시하세요."
    note.value = msg
    note.font = Font(name="Calibri", size=9, italic=True, color="555555")
    note.alignment = left()
    note.fill = fill("FFF3CD")
    ws1.row_dimensions[2].height = 20

    for ri, d in enumerate(domains, 3):
        ws1.cell(ri, 1, d["code"]).alignment = center()
        ws1.cell(ri, 2, d["name"]).alignment = left()
        c = ws1.cell(ri, 3, "Y")
        c.alignment = center()
        c.fill = fill(YELLOW)
        c.font = Font(name="Calibri", bold=True, color="1E6F50")
        ws1.cell(ri, 4, "").alignment = left()
        for ci in range(1, 5):
            ws1.cell(ri, ci).border = bdr()
        ws1.row_dimensions[ri].height = 22

    # ── Sheet 2: 원문 (분류표 마크다운) ───────────────────────────────────────
    ws2 = wb.create_sheet("분류표 원문")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 120
    title_cell = ws2["A1"]
    title_cell.value = "분류표 / 영향도 분석 원문"
    title_cell.font = hfont(size=12)
    title_cell.fill = fill(TEAL)
    title_cell.alignment = left()
    ws2.row_dimensions[1].height = 26

    for ri, line in enumerate(content.splitlines(), 2):
        c = ws2.cell(ri, 1, line)
        c.font = Font(name="Consolas", size=9)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
        ws2.row_dimensions[ri].height = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = "gate_review.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# ── Google Drive 헬퍼 ─────────────────────────────────────────────────────────
def load_config():
    if CONFIG_FILE.exists():
        return json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
    return {}


GOOGLE_SCOPES = [
    "https://www.googleapis.com/auth/drive.file",
    "https://www.googleapis.com/auth/drive.readonly",
    "https://www.googleapis.com/auth/spreadsheets",
]


def _get_google_creds():
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request

    if not DRIVE_CREDS_FILE.exists():
        raise FileNotFoundError(
            "credentials.json이 없습니다.\n"
            "Google Cloud Console → API 및 서비스 → 사용자 인증 정보 →\n"
            "OAuth 2.0 클라이언트 ID (데스크톱 앱)를 만들고\n"
            f"다운로드한 파일을 '{DRIVE_CREDS_FILE}' 으로 저장하세요."
        )
    creds = None
    if DRIVE_TOKEN_FILE.exists():
        try:
            creds = Credentials.from_authorized_user_file(str(DRIVE_TOKEN_FILE), GOOGLE_SCOPES)
        except Exception:
            creds = None
        if creds and set(creds.scopes or []) != set(GOOGLE_SCOPES):
            # scope 가 늘어났으면 재인증 필요
            creds = None
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception:
                creds = None
        if not creds or not creds.valid:
            flow = InstalledAppFlow.from_client_secrets_file(str(DRIVE_CREDS_FILE), GOOGLE_SCOPES)
            creds = flow.run_local_server(port=0)
        DRIVE_TOKEN_FILE.write_text(creds.to_json(), encoding="utf-8")
    return creds


def get_drive_service():
    from googleapiclient.discovery import build
    return build("drive", "v3", credentials=_get_google_creds())


def get_sheets_service():
    from googleapiclient.discovery import build
    return build("sheets", "v4", credentials=_get_google_creds())


@app.route("/open-folder", methods=["POST"])
def open_folder():
    import subprocess as _sp
    import platform
    folder = str(OUTPUTS_DIR)
    if platform.system() == "Windows":
        _sp.Popen(["explorer", folder])
    elif platform.system() == "Darwin":
        _sp.Popen(["open", folder])
    else:
        _sp.Popen(["xdg-open", folder])
    return jsonify({"ok": True})


@app.route("/drive/status")
def drive_status():
    """Drive 인증 상태 + 현재 계정 이메일 확인"""
    try:
        service = get_drive_service()
        about = service.about().get(fields="user(emailAddress,displayName)").execute()
        user = about.get("user", {})
        return jsonify({
            "ok": True,
            "authenticated": True,
            "email": user.get("emailAddress", ""),
            "name": user.get("displayName", ""),
        })
    except FileNotFoundError as e:
        return jsonify({"ok": False, "authenticated": False, "error": str(e), "need_credentials": True})
    except Exception as e:
        return jsonify({"ok": False, "authenticated": False, "error": str(e)})


@app.route("/drive/folders")
def drive_folders():
    """Drive 폴더 목록 조회 (검색어 지원)"""
    q = (request.args.get("q") or "").strip()
    try:
        service = get_drive_service()
        query = "mimeType='application/vnd.google-apps.folder' and trashed=false"
        if q:
            # 이름에 검색어 포함된 폴더
            query += f" and name contains '{q}'"
        result = service.files().list(
            q=query,
            fields="files(id,name,parents,modifiedTime)",
            orderBy="modifiedTime desc",
            pageSize=50
        ).execute()
        return jsonify({"ok": True, "folders": result.get("files", [])})
    except FileNotFoundError as e:
        return jsonify({"ok": False, "error": str(e), "need_credentials": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/upload-to-drive", methods=["POST"])
def upload_to_drive():
    data = request.get_json() or {}
    filename = data.get("filename", "")
    sid = data.get("sid", "")
    folder_id = data.get("folder_id", "")  # 사용자가 선택한 폴더 (우선)

    # 세션에서 결과 파일 경로 확인, 없으면 outputs 폴더에서 탐색
    file_path = None
    sess = SESSIONS.get(sid)
    if sess and sess.get("result") and Path(sess["result"]).exists():
        file_path = Path(sess["result"])
    else:
        candidate = (OUTPUTS_DIR / filename).resolve()
        if str(candidate).startswith(str(OUTPUTS_DIR.resolve())) and candidate.exists():
            file_path = candidate

    if not file_path:
        return jsonify({"ok": False, "error": "파일을 찾을 수 없습니다."})

    # 폴더 ID: 요청 파라미터 > config 기본값
    if not folder_id:
        config = load_config()
        folder_id = config.get("google_drive", {}).get("upload_folder_id")
    if not folder_id:
        return jsonify({"ok": False, "error": "업로드할 폴더를 선택해주세요.", "need_folder": True})

    try:
        from googleapiclient.http import MediaFileUpload
        service = get_drive_service()
        file_metadata = {"name": filename, "parents": [folder_id]}
        media = MediaFileUpload(
            str(file_path),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        uploaded = service.files().create(
            body=file_metadata, media_body=media, fields="id,webViewLink,parents"
        ).execute()
        # 업로드된 폴더 URL 구성
        parent_id = uploaded.get("parents", [folder_id])[0]
        folder_url = f"https://drive.google.com/drive/folders/{parent_id}"
        return jsonify({
            "ok": True,
            "file_id": uploaded.get("id"),
            "link": folder_url,
            "file_link": uploaded.get("webViewLink"),
        })
    except FileNotFoundError as e:
        return jsonify({"ok": False, "error": str(e), "need_credentials": True})
    except Exception as e:
        err_str = str(e)
        if any(k in err_str for k in ["insufficient_scope", "accessNotConfigured",
                                       "forbidden", "403", "Request had insufficient"]):
            return jsonify({"ok": False, "error": "Google Drive 접근 권한이 없습니다.", "need_credentials": True})
        return jsonify({"ok": False, "error": err_str})


@app.route("/download/<sid>/<filename>")
def download(sid, filename):
    sess = SESSIONS.get(sid)
    if not sess:
        return jsonify({"error": "세션 없음"}), 404
    result = sess.get("result")
    if not result or not Path(result).exists():
        # outputs 폴더에서 파일명으로 검색
        fpath = OUTPUTS_DIR / filename
        if not fpath.exists():
            return jsonify({"error": "파일 없음"}), 404
        result = fpath
    return send_file(str(result), as_attachment=True, download_name=filename)


@app.route("/github-tree", methods=["POST"])
def github_tree():
    """GitHub 리포지토리 파일 트리를 반환 (파일 선택 미리보기용)"""
    try:
        import requests as req_lib
    except ImportError:
        return jsonify({"ok": False, "error": "requests 패키지가 필요합니다"})

    data = request.get_json(force=True)
    url = (data.get("url") or "").strip()

    gh_match = re.match(r"https?://github\.com/([^/]+)/([^/\s]+?)(?:\.git)?/?$", url)
    if not gh_match:
        return jsonify({"ok": False, "error": "GitHub 리포지토리 URL이 아닙니다"})

    owner, repo = gh_match.group(1), gh_match.group(2)
    headers = {"User-Agent": "TC-Automation/2.0"}
    token = os.environ.get("GITHUB_TOKEN", "")
    if token:
        headers["Authorization"] = f"token {token}"

    # 기본 브랜치 확인
    branch = "main"
    try:
        r = req_lib.get(f"https://api.github.com/repos/{owner}/{repo}", headers=headers, timeout=10)
        if r.status_code == 200:
            branch = r.json().get("default_branch", "main")
        elif r.status_code == 404:
            return jsonify({"ok": False, "error": f"리포지토리를 찾을 수 없습니다: {owner}/{repo} (private이면 GITHUB_TOKEN 필요)"})
        elif r.status_code == 401:
            return jsonify({"ok": False, "error": "GitHub 인증 실패: GITHUB_TOKEN을 확인해주세요"})
    except Exception as e:
        return jsonify({"ok": False, "error": f"GitHub API 오류: {e}"})

    # 파일 트리 가져오기
    try:
        r = req_lib.get(f"https://api.github.com/repos/{owner}/{repo}/git/trees/{branch}?recursive=1", headers=headers, timeout=15)
        if r.status_code != 200:
            return jsonify({"ok": False, "error": f"파일 트리 로드 실패 (status {r.status_code})"})
        tree = r.json().get("tree", [])
    except Exception as e:
        return jsonify({"ok": False, "error": f"파일 트리 오류: {e}"})

    # 파일 목록 (blob만, 디렉토리 제외)
    files = [
        {"path": f["path"], "size": f.get("size", 0)}
        for f in tree
        if f["type"] == "blob"
    ]

    return jsonify({"ok": True, "owner": owner, "repo": repo, "branch": branch, "files": files})


@app.route("/shutdown", methods=["POST"])
def shutdown():
    func = request.environ.get("werkzeug.server.shutdown")
    if func:
        func()
    else:
        os._exit(0)
    return jsonify({"ok": True})


# ── HTML 템플릿 ────────────────────────────────────────────────────────────────
HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>TC 자동화 v2</title>
<style>
  :root {
    --bg: #EEF1F8;
    --navy: #1E2761;
    --blue: #3557A0;
    --teal: #028090;
    --white: #FFFFFF;
    --text: #222222;
    --muted: #6B7A99;
    --border: #D0D7DE;
    --success: #1C6E38;
    --warn: #B45309;
    --danger: #B91C1C;
    --radius: 12px;
    --shadow: 0 2px 12px rgba(30,39,97,0.10);
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { background: var(--bg); color: var(--text); font-family: 'Segoe UI', system-ui, sans-serif; min-height: 100vh; }

  /* 경고 배너 */
  .api-warning {
    background: #FEF3C7; border-bottom: 2px solid #D97706;
    color: #92400E; padding: 10px 24px; font-size: 14px; text-align: center;
    font-weight: 500;
  }

  /* 헤더 */
  header {
    background: var(--navy); color: var(--white);
    padding: 20px 32px; display: flex; align-items: center; gap: 16px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.18);
  }
  header > div { display: flex; flex-direction: column; gap: 2px; }
  header h1 { font-size: 22px; font-weight: 700; letter-spacing: -0.3px; }
  header .header-sub { font-size: 13px; opacity: 0.7; margin-left: auto; }
  .version-badge { font-size: 11px; font-weight: 600; color: #FF8C00; letter-spacing: 0.3px; }
  /* TC Update 직전 작업 불러오기 — 드롭다운 행 hover */
  .recent-row:hover { background: #F9FAFB; }
  /* 헤더 우측 서버 재시작 버튼 */
  .btn-restart-server {
    padding: 6px 12px; border-radius: 6px;
    background: rgba(255,255,255,0.10); color: #FFFFFF;
    border: 1px solid rgba(255,255,255,0.30);
    font-size: 12px; font-weight: 600; cursor: pointer;
    transition: all 0.15s; white-space: nowrap;
    display: inline-flex; align-items: center; gap: 5px;
  }
  .btn-restart-server:hover {
    background: rgba(239, 68, 68, 0.85); border-color: #EF4444;
  }
  .btn-restart-server:disabled { opacity: 0.5; cursor: not-allowed; }
  /* 재시작 진행 오버레이 */
  .restart-overlay {
    position: fixed; inset: 0; z-index: 10000;
    background: rgba(15, 23, 42, 0.85); color: #FFFFFF;
    display: none; align-items: center; justify-content: center;
    flex-direction: column; gap: 16px; text-align: center; padding: 24px;
  }
  .restart-overlay.visible { display: flex; }
  .restart-overlay-spinner {
    width: 56px; height: 56px; border: 4px solid rgba(255,255,255,0.25);
    border-top-color: #FFFFFF; border-radius: 50%;
    animation: restartSpin 0.9s linear infinite;
  }
  @keyframes restartSpin { to { transform: rotate(360deg); } }
  .restart-overlay-title { font-size: 20px; font-weight: 700; letter-spacing: -0.3px; }
  .restart-overlay-msg { font-size: 14px; opacity: 0.85; max-width: 480px; line-height: 1.6; }
  .restart-overlay-status { font-size: 12px; opacity: 0.7; font-family: monospace; }

  /* 진행 스텝 바 */
  .steps-bar {
    background: var(--white); border-bottom: 1px solid var(--border);
    padding: 14px 32px; display: flex; gap: 0; overflow-x: auto;
  }
  .step-item {
    display: flex; align-items: center; gap: 8px;
    font-size: 13px; color: var(--muted); white-space: nowrap; flex: 1;
  }
  .step-item.active { color: var(--blue); font-weight: 600; }
  .step-item.done   { color: var(--success); }
  .step-num {
    width: 26px; height: 26px; border-radius: 50%;
    background: var(--border); color: var(--muted);
    display: flex; align-items: center; justify-content: center;
    font-size: 12px; font-weight: 700; flex-shrink: 0;
  }
  .step-item.active .step-num { background: var(--blue); color: var(--white); }
  .step-item.done   .step-num { background: var(--success); color: var(--white); }
  .step-arrow { color: var(--border); margin: 0 4px; font-size: 16px; }

  /* 메인 */
  main { max-width: 860px; margin: 32px auto; padding: 0 20px 60px; }

  /* 카드 */
  .card {
    background: var(--white); border-radius: var(--radius);
    box-shadow: var(--shadow); padding: 28px 32px; margin-bottom: 20px;
  }
  .card-title {
    font-size: 17px; font-weight: 700; color: var(--navy);
    margin-bottom: 18px; display: flex; align-items: center; gap: 8px;
  }
  .card-title .badge {
    font-size: 11px; font-weight: 600; padding: 2px 8px;
    border-radius: 99px; background: var(--bg); color: var(--muted);
  }

  /* 폼 */
  .form-label { font-size: 13px; font-weight: 600; color: var(--navy); margin-bottom: 6px; display: block; }
  .form-input {
    width: 100%; padding: 10px 14px; border: 1.5px solid var(--border);
    border-radius: 8px; font-size: 14px; color: var(--text); background: var(--bg);
    transition: border-color 0.15s;
  }
  .form-input:focus { outline: none; border-color: var(--blue); background: var(--white); }
  textarea.form-input { resize: vertical; min-height: 120px; }
  .form-group { margin-bottom: 16px; }

  /* 프로젝트 드롭다운 */
  .project-select-row {
    display: flex; gap: 8px; align-items: center; margin-bottom: 16px;
  }
  .project-select-row label {
    font-size: 13px; font-weight: 600; color: var(--navy); white-space: nowrap;
  }
  .project-dropdown {
    flex: 1; padding: 8px 12px; border: 1.5px solid var(--border);
    border-radius: 8px; font-size: 13px; color: var(--text);
    background: #fff; cursor: pointer; appearance: none;
    background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='6'%3E%3Cpath d='M0 0l5 6 5-6z' fill='%23666'/%3E%3C/svg%3E");
    background-repeat: no-repeat; background-position: right 12px center;
    padding-right: 30px;
  }
  .project-dropdown:focus { border-color: var(--blue); outline: none; }
  .btn-new-project {
    padding: 7px 14px; background: var(--teal); color: #fff;
    border: none; border-radius: 8px; font-size: 12px; font-weight: 600;
    cursor: pointer; white-space: nowrap;
  }
  .btn-new-project:hover { background: #026D75; }
  .project-new-inline {
    display: none; gap: 8px; align-items: center; margin-bottom: 12px;
  }
  .project-new-inline input {
    flex: 1; padding: 7px 10px; border: 1.5px solid var(--border);
    border-radius: 8px; font-size: 13px;
  }
  .project-new-inline input:focus { border-color: var(--blue); outline: none; }
  .project-resume-bar {
    padding: 10px 12px; margin-bottom: 16px;
    border: 1.5px solid #60A5FA; border-radius: 8px;
    background: #EFF6FF; display: none; align-items: center; gap: 10px;
    font-size: 12px; color: #1E40AF;
  }
  .project-resume-bar strong { font-weight: 700; }
  .btn-resume-pipeline {
    padding: 5px 14px; background: #2563EB; color: #fff; border: none;
    border-radius: 6px; font-size: 12px; font-weight: 600; cursor: pointer;
    white-space: nowrap;
  }
  .btn-resume-pipeline:hover { background: #1D4ED8; }
  .pj-status-tag {
    font-size: 10px; padding: 1px 6px; border-radius: 8px;
    font-weight: 600; margin-left: 4px; vertical-align: middle;
  }
  .pj-status-tag.in-progress { background: #FEF3C7; color: #92400E; }
  .pj-status-tag.completed { background: #D1FAE5; color: #065F46; }
  .pj-status-tag.sample { background: #E5E7EB; color: #6B7280; }

  /* 모드 스위처 */
  .mode-switcher {
    display: flex; background: var(--bg); border-radius: 10px;
    padding: 4px; gap: 4px; margin-bottom: 20px;
  }
  .mode-btn {
    flex: 1; padding: 10px 16px; border: none; border-radius: 8px;
    font-size: 14px; font-weight: 600; cursor: pointer;
    background: none; color: var(--muted); transition: all 0.2s;
  }
  .mode-btn.active {
    background: var(--white); color: var(--navy);
    box-shadow: 0 1px 4px rgba(0,0,0,.12);
  }

  /* 입력 유형 탭 버튼 */
  .radio-group { display: flex; gap: 10px; flex-wrap: wrap; }
  .type-btn {
    display: flex; align-items: center; gap: 6px; cursor: pointer;
    padding: 8px 16px; border: 1.5px solid var(--border); border-radius: 8px;
    font-size: 13px; font-weight: 500; transition: all 0.15s;
    background: var(--white); color: var(--text);
  }
  .type-btn:hover { border-color: var(--blue); color: var(--blue); }
  .type-btn.active { border-color: var(--blue); background: #EBF2FF; color: var(--blue); }

  /* 프로젝트 선택 드롭다운 */
  .project-select {
    width: 100%; padding: 11px 14px; border: 1.5px solid var(--border);
    border-radius: 8px; font-size: 14px; color: var(--text);
    background: var(--bg); cursor: pointer; appearance: none;
    background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 12 12'%3E%3Cpath fill='%23888' d='M6 8L1 3h10z'/%3E%3C/svg%3E");
    background-repeat: no-repeat; background-position: right 14px center;
  }
  .project-select:focus { outline: none; border-color: var(--blue); }

  /* 프로젝트 카드 (선택된 프로젝트 정보) */
  .project-card {
    background: #F0FBF8; border: 1.5px solid #B2E8DC; border-radius: 10px;
    padding: 14px 16px; margin-top: 10px; display: none;
  }
  .project-card.visible { display: block; }
  .project-card-name { font-size: 15px; font-weight: 700; color: var(--navy); }
  .project-card-meta { font-size: 12px; color: var(--muted); margin-top: 3px; }

  /* TC 파일 업로드 영역 */
  .tc-upload-area {
    border: 2px dashed #B2E8DC; border-radius: 10px; padding: 20px;
    text-align: center; cursor: pointer; color: var(--muted); font-size: 13px;
    transition: all 0.2s; margin-top: 10px;
  }
  .tc-upload-area:hover { border-color: var(--teal); color: var(--teal); }

  /* 직접 입력 textarea */
  .text-input-area {
    min-height: 260px; resize: vertical; font-family: inherit;
    line-height: 1.7; font-size: 14px;
  }
  .input-hint {
    font-size: 12px; color: var(--muted); margin-top: 6px;
  }

  /* 드롭존 */
  .dropzone {
    border: 2px dashed var(--border); border-radius: 10px;
    padding: 32px; text-align: center; cursor: pointer;
    transition: all 0.2s; color: var(--muted); font-size: 14px;
  }
  .dropzone:hover, .dropzone.drag-over { border-color: var(--blue); background: #EBF2FF; color: var(--blue); }
  .dropzone .icon { font-size: 36px; margin-bottom: 8px; }
  .dropzone .hint { font-size: 12px; margin-top: 4px; }
  #fileInput { display: none; }

  /* 버튼 */
  .btn {
    padding: 11px 24px; border-radius: 8px; font-size: 14px; font-weight: 600;
    cursor: pointer; border: none; transition: all 0.15s; display: inline-flex;
    align-items: center; gap: 8px;
  }
  .btn-primary { background: var(--blue); color: var(--white); }
  .btn-primary:hover { background: var(--navy); }
  .btn-primary:disabled { opacity: 0.5; cursor: not-allowed; }
  .btn-success { background: var(--success); color: var(--white); }
  .btn-success:hover { background: #155228; }
  .btn-teal { background: var(--teal); color: var(--white); }
  .btn-teal:hover { background: #016070; }
  .btn-stop {
    background: none; color: #e53e3e; border: 1.5px solid #e53e3e;
    padding: 7px 18px; border-radius: 8px; font-size: 13px; font-weight: 600;
    cursor: pointer; transition: all 0.15s; display: inline-flex; align-items: center; gap: 6px;
  }
  .btn-stop:hover { background: #fff5f5; }
  .btn-stop:disabled { opacity: 0.4; cursor: not-allowed; }

  /* 중단 배너 */
  .stopped-banner {
    background: #FFF5F5; border: 1.5px solid #FEB2B2; border-radius: 10px;
    padding: 18px 22px; margin-top: 16px; display: flex; align-items: center; gap: 12px;
  }
  .stopped-banner .stopped-icon { font-size: 28px; flex-shrink: 0; }
  .stopped-banner .stopped-msg  { font-size: 14px; font-weight: 600; color: #c53030; }
  .stopped-banner .stopped-sub  { font-size: 12px; color: #888; margin-top: 2px; }

  /* 다중 소스 입력 */
  .source-add-bar { display: flex; gap: 8px; margin-bottom: 12px; flex-wrap: wrap; }
  .btn-add-source {
    padding: 7px 14px; border-radius: 8px; font-size: 12px; font-weight: 600;
    background: var(--white); color: var(--text); border: 1.5px solid var(--border);
    cursor: pointer; transition: all 0.15s;
  }
  .btn-add-source:hover { border-color: var(--blue); color: var(--blue); background: #EBF2FF; }
  .btn-clear-sources {
    padding: 7px 14px; border-radius: 8px; font-size: 12px; font-weight: 600;
    background: #FEF2F2; color: #B91C1C; border: 1.5px solid #FCA5A5;
    cursor: pointer; transition: all 0.15s; margin-left: auto;
  }
  .btn-clear-sources:hover { background: #FEE2E2; border-color: #EF4444; color: #991B1B; }
  .source-card {
    border: 1.5px solid var(--border); border-radius: 10px;
    margin-bottom: 10px; overflow: hidden; background: var(--white);
  }
  .source-card-header {
    display: flex; align-items: center; justify-content: space-between;
    padding: 8px 12px; background: var(--bg); border-bottom: 1px solid var(--border);
  }
  .source-type-badge {
    font-size: 12px; font-weight: 700; padding: 3px 10px; border-radius: 12px;
  }
  .source-type-badge.pdf  { background: #FEE2E2; color: #C53030; }
  .source-type-badge.url  { background: #EBF2FF; color: var(--blue); }
  .source-type-badge.text { background: #D1FAE5; color: #276749; }
  .btn-remove-source {
    background: none; border: none; cursor: pointer; font-size: 14px;
    color: var(--muted); padding: 2px 8px; border-radius: 4px;
  }
  .btn-remove-source:hover { background: #FEE2E2; color: #C53030; }
  .source-card-body { padding: 12px; }
  .src-dropzone {
    border: 2px dashed var(--border); border-radius: 8px; padding: 18px;
    text-align: center; cursor: pointer; font-size: 13px; color: var(--muted);
    transition: all 0.15s;
  }
  .src-dropzone:hover { border-color: var(--blue); color: var(--blue); }
  .src-file-name { font-size: 13px; font-weight: 600; color: var(--success); padding: 4px 0; }
  .source-empty {
    border: 2px dashed var(--border); border-radius: 10px; padding: 28px 16px;
    text-align: center; font-size: 13px; color: var(--muted); line-height: 1.8;
  }

  /* TC 가져오기 탭 */
  .tc-import-tabs { display: flex; gap: 4px; margin-bottom: 10px; }
  .tc-import-tab { flex: 1; padding: 7px; font-size: 12px; font-weight: 600; border: 1px solid var(--border); border-radius: 6px; background: var(--surface); cursor: pointer; color: var(--muted); }
  .tc-import-tab.active { background: var(--navy); color: #fff; border-color: var(--navy); }

  /* 새 프로젝트 생성 */
  .btn-new-project { font-size: 11px; padding: 3px 10px; border: 1px solid var(--teal); color: var(--teal); background: none; border-radius: 6px; cursor: pointer; font-weight: 600; }
  .btn-new-project:hover { background: var(--teal); color: #fff; }
  .new-project-form { margin-top: 10px; padding: 12px; background: var(--surface); border: 1px solid var(--border); border-radius: 8px; }
  .btn-delete-project { margin-top: 6px; font-size: 11px; padding: 2px 8px; border: 1px solid #e53e3e; color: #e53e3e; background: none; border-radius: 6px; cursor: pointer; }
  .btn-delete-project:hover { background: #e53e3e; color: #fff; }

  /* 진행률 바 */
  .progress-wrap { background: var(--bg); border-radius: 99px; height: 10px; overflow: hidden; margin: 12px 0; }
  .progress-bar  { height: 100%; border-radius: 99px; background: linear-gradient(90deg, var(--blue), var(--teal)); transition: width 0.4s ease; }

  /* 로그 박스 */
  .log-box {
    background: #1a1d2e; color: #a8d8a8; font-family: 'Consolas', 'Monaco', monospace;
    font-size: 12px; padding: 14px 16px; border-radius: 8px; height: 200px;
    overflow-y: auto; line-height: 1.6; margin-top: 12px;
  }
  .log-box .log-line { padding: 1px 0; }
  .log-box .log-line.error { color: #f87171; }

  /* 서브스텝 아이콘 — 3가지 명확한 상태: 대기(회색) · 진행 중(파랑 테두리) · 완료(초록 + ✓) */
  .substeps { display: flex; gap: 8px; margin: 14px 0; flex-wrap: wrap; }
  .substep {
    display: flex; align-items: center; gap: 6px;
    padding: 6px 12px; border-radius: 20px; font-size: 12px; font-weight: 600;
    background: #F3F4F6; color: #9CA3AF; border: 1.5px solid #E5E7EB;
    transition: all 0.25s;
  }
  .substep.active {
    background: #DBEAFE; color: #1E40AF; border-color: #3B82F6;
    box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.2);
    animation: pulse-active 1.6s ease-in-out infinite;
  }
  .substep.done {
    background: #D1FAE5; color: #065F46; border-color: #10B981;
  }
  .substep.done::after {
    content: ' ✓'; margin-left: 2px; font-weight: 800;
  }
  @keyframes pulse-active {
    0%, 100% { box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.2); }
    50%      { box-shadow: 0 0 0 5px rgba(59, 130, 246, 0.35); }
  }

  /* (제거됨) .domain-chip — 대분류 체크리스트 UI 삭제로 미사용 */

  /* Gate 채팅 UI */
  .gate-layout {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 16px;
    min-height: 480px;
  }
  @media (max-width: 860px) {
    .gate-layout { grid-template-columns: 1fr; }
  }
  .gate-chat-panel {
    display: flex; flex-direction: column;
    border: 1.5px solid var(--border); border-radius: 10px; overflow: hidden;
    background: #fff;
  }
  .gate-chat-header {
    padding: 10px 14px; background: var(--navy); color: #fff;
    font-size: 13px; font-weight: 600; display: flex; align-items: center; gap: 6px;
  }
  .gate-chat-messages {
    flex: 1; overflow-y: auto; padding: 12px; display: flex;
    flex-direction: column; gap: 10px; min-height: 140px; max-height: 280px;
    background: #F8FAFC;
  }
  .gate-msg { max-width: 88%; padding: 9px 13px; border-radius: 10px; font-size: 13px; line-height: 1.55; }
  .gate-msg.assistant {
    align-self: flex-start; background: #EDF2FF; color: var(--navy);
    border-bottom-left-radius: 3px;
  }
  .gate-msg.user {
    align-self: flex-end; background: var(--teal); color: #fff;
    border-bottom-right-radius: 3px;
  }
  .gate-msg.system {
    align-self: center; background: #F0FDF4; color: #166534;
    font-size: 12px; border: 1px solid #BBF7D0; border-radius: 8px; max-width: 96%;
  }
  .gate-chat-input-row {
    display: flex; gap: 8px; padding: 10px;
    border-top: 1px solid var(--border); background: #fff;
  }
  .gate-chat-input {
    flex: 1; border: 1.5px solid var(--border); border-radius: 8px;
    padding: 8px 12px; font-size: 13px; resize: none; outline: none;
    font-family: inherit; min-height: 40px; max-height: 100px;
    line-height: 1.5; color: var(--text);
  }
  .gate-chat-input:focus { border-color: var(--blue); }
  .gate-chat-send {
    padding: 8px 16px; background: var(--teal); color: #fff;
    border: none; border-radius: 8px; font-size: 13px; font-weight: 600;
    cursor: pointer; white-space: nowrap; align-self: flex-end;
  }
  .gate-chat-send:disabled { opacity: 0.5; cursor: not-allowed; }
  /* ── Sticky Mini AI 채팅 패널 (Step 3 분류 검토 전용) ── */
  /* 3가지 상태: collapsed(입력만, 약 60px) / expanded(메시지+입력, ~280px) / modal(전체화면) */
  .floating-ai-bar {
    position: fixed; left: 0; right: 0; bottom: 0;
    z-index: 9000;
    background: linear-gradient(135deg, rgba(30, 58, 95, 0.97) 0%, rgba(45, 91, 110, 0.97) 100%);
    backdrop-filter: blur(8px); -webkit-backdrop-filter: blur(8px);
    border-top: 3px solid #14B8A6;
    box-shadow: 0 -8px 24px rgba(0, 0, 0, 0.18);
    color: #FFFFFF;
    display: none;
    transform: translateY(100%);
    transition: transform 0.32s cubic-bezier(0.22, 1, 0.36, 1);
  }
  .floating-ai-bar.visible {
    display: flex; flex-direction: column;
    transform: translateY(0);
    animation: floatingAiPulse 1.4s ease-out 0.32s 1;
  }
  @keyframes floatingAiPulse {
    0%, 100% { box-shadow: 0 -8px 24px rgba(0,0,0,0.18); }
    50% { box-shadow: 0 -8px 24px rgba(0,0,0,0.18), 0 0 0 4px rgba(20, 184, 166, 0.55); }
  }
  /* 안전망: card3 hidden / stepBar3 비활성 시 절대 숨김 */
  body:has(#card3.hidden) .floating-ai-bar { display: none !important; }
  body:has(#stepBar3:not(.active)) .floating-ai-bar { display: none !important; }

  /* 헤더 — 라벨 + 메시지 개수 + 컨트롤 버튼 */
  .floating-ai-header {
    display: flex; align-items: center; gap: 10px;
    padding: 8px 20px; border-bottom: 1px solid rgba(255,255,255,0.12);
    background: rgba(0,0,0,0.10);
    cursor: pointer; user-select: none;
  }
  .floating-ai-bar.collapsed .floating-ai-header { border-bottom: none; }
  .floating-ai-header-label {
    flex: 1; font-size: 13px; font-weight: 700; color: #FFFFFF;
    display: flex; align-items: center; gap: 8px; white-space: nowrap;
  }
  .floating-ai-header-label small {
    font-weight: 400; opacity: 0.7; font-size: 11px;
  }
  .floating-ai-msg-badge {
    font-size: 10px; font-weight: 700;
    padding: 2px 8px; border-radius: 999px;
    background: rgba(20, 184, 166, 0.32); color: #A7F3D0;
  }
  .floating-ai-ctrl {
    background: rgba(255,255,255,0.10);
    color: #FFFFFF; border: 1px solid rgba(255,255,255,0.22);
    border-radius: 6px;
    padding: 4px 10px; font-size: 11px; font-weight: 600;
    cursor: pointer; white-space: nowrap;
    transition: background 0.15s ease;
  }
  .floating-ai-ctrl:hover { background: rgba(255,255,255,0.20); }

  /* 메시지 영역 (collapsed 상태에서 숨김) */
  .floating-ai-messages {
    overflow-y: auto; padding: 12px 20px;
    display: flex; flex-direction: column; gap: 8px;
    max-height: 220px; min-height: 80px;
    background: rgba(255,255,255,0.04);
    transition: max-height 0.25s ease, padding 0.25s ease, opacity 0.2s ease;
  }
  .floating-ai-bar.collapsed .floating-ai-messages {
    max-height: 0; min-height: 0; padding: 0 20px;
    opacity: 0; pointer-events: none; overflow: hidden;
  }
  .floating-ai-msg {
    max-width: 88%; padding: 8px 12px; border-radius: 10px;
    font-size: 12.5px; line-height: 1.55;
    word-wrap: break-word; word-break: break-word;
  }
  .floating-ai-msg.assistant {
    align-self: flex-start;
    background: rgba(255,255,255,0.94); color: #1E3A5F;
    border-bottom-left-radius: 3px;
  }
  .floating-ai-msg.user {
    align-self: flex-end;
    background: linear-gradient(135deg, #14B8A6 0%, #0D9488 100%);
    color: #FFFFFF;
    border-bottom-right-radius: 3px;
  }
  .floating-ai-msg.system {
    align-self: center;
    background: rgba(34, 197, 94, 0.18); color: #BBF7D0;
    font-size: 11px;
    border: 1px solid rgba(34, 197, 94, 0.32);
    border-radius: 8px; max-width: 96%;
  }
  .floating-ai-empty {
    align-self: center; color: rgba(255,255,255,0.55);
    font-size: 11.5px; padding: 14px 0;
  }

  /* 입력 영역 */
  .floating-ai-inner {
    max-width: 1100px; margin: 0 auto; width: 100%;
    display: flex; align-items: flex-end; gap: 10px;
    padding: 10px 20px 12px 20px;
  }
  .floating-ai-input {
    flex: 1; border: 1.5px solid rgba(255, 255, 255, 0.25); border-radius: 10px;
    background: rgba(255, 255, 255, 0.96);
    padding: 10px 14px; font-size: 13px; outline: none;
    font-family: inherit; min-height: 38px; max-height: 100px;
    line-height: 1.5; color: var(--text); resize: none;
    box-shadow: 0 2px 6px rgba(0, 0, 0, 0.12);
    transition: box-shadow 0.18s ease, border-color 0.18s ease;
  }
  .floating-ai-input:focus {
    border-color: #14B8A6;
    box-shadow: 0 2px 6px rgba(0, 0, 0, 0.12), 0 0 0 3px rgba(20, 184, 166, 0.32);
  }
  .floating-ai-send {
    padding: 9px 20px;
    background: linear-gradient(135deg, #14B8A6 0%, #0D9488 100%);
    color: #fff; border: none; border-radius: 10px;
    font-size: 13px; font-weight: 700;
    cursor: pointer; white-space: nowrap;
    box-shadow: 0 2px 8px rgba(20, 184, 166, 0.42);
    transition: transform 0.15s ease, box-shadow 0.15s ease;
  }
  .floating-ai-send:hover {
    transform: translateY(-1px);
    box-shadow: 0 4px 12px rgba(20, 184, 166, 0.55);
  }
  .floating-ai-send:disabled { opacity: 0.5; cursor: not-allowed; transform: none; }

  /* body padding — collapsed 56px / expanded 약 340px (헤더+메시지+입력) */
  body.has-floating-ai { padding-bottom: 340px; }
  body.has-floating-ai.has-floating-ai-collapsed { padding-bottom: 64px; }

  /* 좁은 화면 */
  @media (max-width: 640px) {
    .floating-ai-header { padding: 8px 14px; }
    .floating-ai-messages { padding: 10px 14px; max-height: 180px; }
    .floating-ai-inner { padding: 8px 14px 10px 14px; gap: 8px; }
    .floating-ai-send { padding: 8px 14px; }
    body.has-floating-ai { padding-bottom: 300px; }
  }

  /* ── Mini Chat 확대 모달 (full-size view) ── */
  .floating-ai-modal {
    position: fixed; inset: 0; z-index: 10000;
    background: rgba(15, 23, 42, 0.72);
    display: none; align-items: center; justify-content: center;
    padding: 20px;
  }
  .floating-ai-modal.open { display: flex; }
  .floating-ai-modal-box {
    background: #FFFFFF; color: var(--text);
    border-radius: 14px;
    width: 100%; max-width: 720px; height: 78vh;
    min-width: 360px; min-height: 320px;
    max-height: 96vh;
    display: flex; flex-direction: column;
    overflow: hidden; position: relative;
    box-shadow: 0 20px 60px rgba(0,0,0,0.3);
    resize: both;  /* native 리사이즈 — 우측 하단 모서리 드래그 */
  }
  /* native resize 핸들 시각 강화 (브라우저 기본 핸들이 작아서 보이게) */
  .floating-ai-modal-resize {
    position: absolute; bottom: 0; right: 0;
    width: 18px; height: 18px;
    background:
      linear-gradient(135deg, transparent 50%, var(--muted) 50%, var(--muted) 60%, transparent 60%, transparent 70%, var(--muted) 70%, var(--muted) 80%, transparent 80%);
    pointer-events: none;  /* native resize 가 잡음 */
    border-bottom-right-radius: 14px;
    opacity: 0.5;
  }
  .floating-ai-modal-header {
    padding: 14px 20px; background: var(--navy); color: #FFFFFF;
    display: flex; align-items: center; gap: 10px;
    border-bottom: 1px solid rgba(255,255,255,0.10);
  }
  .floating-ai-modal-title {
    flex: 1; font-size: 15px; font-weight: 700;
  }
  .floating-ai-modal-close {
    background: rgba(255,255,255,0.12); color: #FFFFFF;
    border: 1px solid rgba(255,255,255,0.22);
    border-radius: 6px; padding: 5px 12px;
    font-size: 12px; cursor: pointer;
  }
  .floating-ai-modal-close:hover { background: rgba(255,255,255,0.22); }
  .floating-ai-modal-messages {
    flex: 1; overflow-y: auto; padding: 18px 22px;
    display: flex; flex-direction: column; gap: 10px;
    background: #F8FAFC;
  }
  .floating-ai-modal-input-row {
    display: flex; gap: 10px; padding: 14px 20px;
    border-top: 1px solid var(--border); background: #FFFFFF;
  }
  /* details/summary 기본 마커 제거 (Safari/WebKit 포함) */
  #tcSummaryDetails > summary::-webkit-details-marker { display: none; }
  #tcSummaryDetails > summary { list-style: none; }
  #tcSummaryDetails > summary:hover { background: rgba(59, 130, 246, 0.06); border-radius: 10px; }
  .gate-viewer-panel {
    display: flex; flex-direction: column;
    border: 1.5px solid var(--border); border-radius: 10px; overflow: hidden;
    background: #fff;
  }
  .gate-viewer-header {
    padding: 10px 14px; background: #2D3748; color: #fff;
    font-size: 13px; font-weight: 600; display: flex; align-items: center; gap: 6px;
  }
  .gate-viewer-content {
    flex: 1; overflow-y: auto; padding: 16px;
    font-size: 13px; line-height: 1.75; color: var(--text);
    min-height: 320px; max-height: 420px; background: #fff;
  }
  .gate-viewer-content h2 { font-size: 15px; font-weight: 700; color: var(--navy); margin: 14px 0 6px; border-bottom: 1.5px solid #E2E8F0; padding-bottom: 4px; }
  .gate-viewer-content h3 { font-size: 13px; font-weight: 600; color: var(--blue); margin: 10px 0 4px; }
  .gate-viewer-content ul, .gate-viewer-content ol { padding-left: 18px; margin: 4px 0; }
  .gate-viewer-content li { margin: 2px 0; }
  .gate-viewer-content strong { color: var(--navy); }
  .gate-viewer-content code { background: #EDF2FF; padding: 1px 5px; border-radius: 3px; font-size: 12px; }
  .gate-viewer-content blockquote { border-left: 3px solid var(--teal); padding-left: 10px; color: var(--muted); margin: 6px 0; }
  .gate-doc-updated { animation: flashUpdate 0.6s ease; }
  @keyframes flashUpdate { 0%,100%{background:#fff} 50%{background:#E0F2FE} }
  /* 기존 textarea (숨김) */
  .gate-textarea { display: none; }

  /* 완료 카드 */
  .result-box {
    background: linear-gradient(135deg, #EBF2FF, #D1FAE5);
    border: 1.5px solid #A7F3D0; border-radius: 10px;
    padding: 20px 24px; margin-bottom: 16px;
  }
  .result-filename { font-size: 18px; font-weight: 700; color: var(--navy); }
  .result-meta { font-size: 13px; color: var(--muted); margin-top: 4px; }
  .tc-stats { display: flex; gap: 24px; margin-top: 14px; }
  .tc-stat { text-align: center; }
  .tc-stat-num { font-size: 28px; font-weight: 800; color: var(--blue); }
  .tc-stat-label { font-size: 12px; color: var(--muted); }

  /* 숨김 */
  .hidden { display: none !important; }

  /* TC 분류 요약 표 — 옵션 E (메인 행 + 종속 행) */
  .tc-summary-table .tc-summary-row-main:hover { background: #EFF6FF !important; }
  .tc-summary-table .tc-summary-row-main td:last-child:hover { background: #DBEAFE; }
  .tc-summary-table .tc-summary-row-nested { transition: opacity 0.15s ease; }
  .tc-summary-table .tc-summary-row-nested td { border-top: 0 !important; }

  /* 안내 */
  .info-box {
    background: #EBF2FF; border-left: 4px solid var(--blue);
    padding: 10px 14px; border-radius: 0 6px 6px 0; font-size: 13px; color: var(--navy);
    margin-bottom: 14px;
  }

  /* 완료 카드 - 파일 행 */
  .result-file {
    display: flex; align-items: center; gap: 14px;
    padding: 16px; background: #F0FBF8; border: 1px solid #B2E8DC;
    border-radius: 12px; margin-bottom: 16px;
  }
  .ricon { font-size: 32px; flex-shrink: 0; }
  .rinfo { flex: 1; overflow: hidden; }
  .rname { font-size: 14px; font-weight: 700; color: var(--navy);
    overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
  .rsize { font-size: 12px; color: var(--muted); margin-top: 2px; }

  /* 액션 버튼 행 */
  .action-row { display: flex; gap: 12px; }
  .btn-dl {
    flex: 1; padding: 12px; background: var(--teal); color: #fff; border: none;
    border-radius: 10px; font-size: 14px; font-weight: 600; cursor: pointer;
    transition: background .2s; display: flex; align-items: center;
    justify-content: center; gap: 6px;
  }
  .btn-dl:hover { background: #016570; }
  .btn-drive {
    flex: 1; padding: 12px; background: #fff; color: var(--navy);
    border: 2px solid #D1D9EE; border-radius: 10px; font-size: 14px;
    font-weight: 600; cursor: pointer; transition: all .2s;
    display: flex; align-items: center; justify-content: center; gap: 6px;
  }
  .btn-drive:hover { border-color: #3557A0; background: #F7F9FF; }

  /* Drive 모달 */
  .modal-bg {
    position: fixed; inset: 0; background: rgba(0,0,0,.45);
    display: none; align-items: center; justify-content: center; z-index: 999;
  }
  .modal-bg.open { display: flex; }
  .modal {
    background: #fff; border-radius: 16px; padding: 32px;
    max-width: 440px; width: 90%; box-shadow: 0 8px 32px rgba(0,0,0,.18);
  }
  .modal h2 { font-size: 17px; font-weight: 700; color: var(--navy); margin-bottom: 16px; }
  .modal ol { padding-left: 20px; color: #444; font-size: 13px; line-height: 2.2; }
  .modal-btns { display: flex; gap: 10px; margin-top: 22px; }
  .btn-open-drive {
    flex: 1; padding: 11px; background: #4285F4; color: #fff; border: none;
    border-radius: 8px; font-size: 13px; font-weight: 600; cursor: pointer;
  }
  .btn-modal-close {
    flex: 1; padding: 11px; background: #F4F6FB; color: #444; border: none;
    border-radius: 8px; font-size: 13px; font-weight: 600; cursor: pointer;
  }

  /* Drive 폴더 선택 리스트 */
  .drive-folder-item {
    display: flex; align-items: center; gap: 10px;
    padding: 10px 14px; border-bottom: 1px solid #E2E8F0;
    cursor: pointer; transition: background 0.1s;
    font-size: 13px;
  }
  .drive-folder-item:hover { background: #EBF2FF; }
  .drive-folder-item.selected { background: #DBEAFE; border-left: 3px solid #2563EB; }
  .drive-folder-item:last-child { border-bottom: none; }

  /* 토스트 */
  #toast {
    position: fixed; bottom: 28px; left: 50%; transform: translateX(-50%);
    padding: 12px 24px; border-radius: 10px; font-size: 14px; font-weight: 600;
    display: none; z-index: 1000; box-shadow: 0 4px 16px rgba(0,0,0,.2); white-space: nowrap;
  }
  #toast.success { background: var(--teal); color: #fff; }
  #toast.error   { background: #e53e3e; color: #fff; }

  @media (max-width: 480px) { .action-row { flex-direction: column; } }

  /* GitHub 파일 트리 */
  .btn-preview-tree { margin-top: 8px; padding: 5px 12px; font-size: 12px; border: 1px solid var(--border); border-radius: 6px; background: var(--surface); cursor: pointer; }
  .btn-preview-tree:hover { background: var(--hover); }
  .tree-panel { margin-top: 10px; border: 1px solid var(--border); border-radius: 8px; overflow: hidden; max-height: 360px; overflow-y: auto; font-size: 12px; }
  .tree-loading, .tree-error { padding: 16px; color: var(--muted); }
  .tree-error { color: #e53e3e; }
  .tree-header { padding: 10px 14px; background: var(--surface); border-bottom: 1px solid var(--border); display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 8px; position: sticky; top: 0; z-index: 1; }
  .tree-actions button { padding: 3px 8px; font-size: 11px; border: 1px solid var(--border); border-radius: 4px; background: var(--surface); cursor: pointer; margin-left: 4px; }
  .btn-apply-tree { background: var(--surface); font-weight: 600; }
  .tree-list { padding: 8px 0; }
  .tree-folder { padding: 4px 14px; }
  .tree-folder-label { display: block; font-weight: 600; color: var(--text); padding: 3px 0; cursor: pointer; }
  .tree-file-label { display: block; padding: 2px 0 2px 20px; color: var(--muted); cursor: pointer; }
  .tree-file-label:hover, .tree-folder-label:hover { color: var(--text); }
  .file-size { color: var(--muted); font-size: 10px; }
</style>
</head>
<body>

<!-- ── JS 에러 안전망 (계층 1) ─────────────────────────────────────────────
   이 inline script 는 다른 모든 script 보다 먼저 로드/실행되어 SyntaxError 가
   메인 스크립트에서 발생해도 살아남아 사용자에게 에러를 보여준다.
   증상: 어제 큰 script 의 SyntaxError 로 restartServer 등 다수 함수가
        정의되지 않아 버튼이 침묵하던 사고가 있었음.
-->
<div id="jsErrorBanner" style="display:none; position:sticky; top:0; z-index:99999;
     background:#FEE2E2; color:#991B1B; padding:10px 14px; border-bottom:2px solid #DC2626;
     font-family:system-ui, sans-serif; font-size:13px; line-height:1.5;">
  <strong>⚠️ JavaScript 오류 감지 — 일부 버튼이 동작하지 않을 수 있습니다.</strong>
  <span id="jsErrorBannerMsg" style="display:block; margin-top:4px; font-family:monospace; font-size:11px; white-space:pre-wrap;"></span>
  <button onclick="document.getElementById('jsErrorBanner').style.display='none'"
          style="float:right; background:none; border:1px solid #991B1B; color:#991B1B;
                 padding:2px 8px; border-radius:4px; cursor:pointer; font-size:11px;">닫기</button>
</div>
<script>
(function() {
  window.__jsErrors = [];
  function showError(msg) {
    var b = document.getElementById('jsErrorBanner');
    var m = document.getElementById('jsErrorBannerMsg');
    if (!b || !m) return;
    b.style.display = 'block';
    var prev = m.textContent || '';
    m.textContent = (prev ? prev + ' | ' : '') + msg;
    window.__jsErrors.push(msg);
  }
  // 1) 런타임 에러 잡기
  window.addEventListener('error', function(e) {
    var msg = '[' + (e.filename || 'inline').split('/').pop() + ':' + (e.lineno || '?') + '] ' + (e.message || 'unknown');
    showError(msg);
  });
  // 2) Promise rejection 잡기
  window.addEventListener('unhandledrejection', function(e) {
    showError('[Promise] ' + (e.reason && (e.reason.message || e.reason) || 'unknown'));
  });
  // 3) DOM 로드 후 핵심 함수 정의 점검 (큰 script 의 SyntaxError 감지)
  window.addEventListener('DOMContentLoaded', function() {
    var critical = ['restartServer', 'startUpdateAnalyze', 'startUpdatePlan', 'loadProjects'];
    var missing = critical.filter(function(n) { return typeof window[n] !== 'function'; });
    if (missing.length) {
      showError('핵심 함수 미정의: ' + missing.join(', ') + ' — 메인 스크립트가 SyntaxError 로 중단됐을 가능성. 브라우저 콘솔(F12) 을 열어 첫 번째 빨간 에러를 확인하세요.');
    }
    // 서버 측 JS 검증 결과도 표시
    if (window.__SERVER_JS_CHECK && window.__SERVER_JS_CHECK.ok === false) {
      showError('서버 JS 검증 실패: ' + window.__SERVER_JS_CHECK.error);
    }
  });
})();
</script>

{% if api_warning %}
<div class="api-warning">
  ⚠️ <strong>ANTHROPIC_API_KEY 미설정</strong> — .env 파일에 API 키를 설정해주세요.
  예: <code>ANTHROPIC_API_KEY=sk-ant-...</code>
</div>
{% endif %}

<header>
  <div>
    <h1>🤖 TC 자동화 v2</h1>
    <span class="version-badge" style="cursor:pointer;" onclick="showWhatsNew()" title="{{ app_version }} 릴리즈 노트 보기">{{ app_version }}</span>
  </div>
  <span class="header-sub">Claude AI · PDF / URL / 텍스트 → Excel</span>
  <button type="button" class="btn-restart-server" id="btnRestartServer" onclick="restartServer()" title="코드 변경 반영을 위해 서버를 재시작합니다 (로컬에서만)">
    🔄 서버 재시작
  </button>
</header>

<!-- 서버 재시작 진행 오버레이 -->
<div id="restartOverlay" class="restart-overlay">
  <div class="restart-overlay-spinner"></div>
  <div class="restart-overlay-title">서버 재시작 중...</div>
  <div class="restart-overlay-msg">
    잠시만 기다려주세요. 서버가 다시 살아나면 자동으로 새로고침됩니다.<br>
    <span style="font-size:12px;opacity:0.8;">5~10초가 지나도 새로고침되지 않으면 직접 페이지를 새로고침해 주세요.</span>
  </div>
  <div class="restart-overlay-status" id="restartOverlayStatus">서버 응답 대기 중...</div>
</div>

<!-- What's new 배너 ({{ app_version }} 첫 방문 시 자동 표시, localStorage로 dismiss 기억) -->
<div id="whatsNewBanner" style="display:none; margin:12px 0; padding:12px 16px; background:linear-gradient(135deg, #EFF6FF 0%, #F0FDF4 100%); border:1px solid #93C5FD; border-radius:10px;">
  <div style="display:flex; justify-content:space-between; align-items:flex-start; gap:12px;">
    <div style="flex:1; font-size:13px; color:#1E40AF;">
      🎨 <strong>{{ app_version }} — {{ app_version_tagline }}</strong>
      <div style="margin-top:6px; font-size:12px; color:#374151; line-height:1.6;">
        {% for line in app_version_highlights %}• {{ line | safe }}<br>{% endfor %}
      </div>
    </div>
    <div style="display:flex; flex-direction:column; gap:6px;">
      <button onclick="showWhatsNew()" style="padding:4px 10px; font-size:11px; background:#2563EB; color:#FFFFFF; border:none; border-radius:6px; cursor:pointer; white-space:nowrap;">📖 자세히 보기</button>
      <button onclick="dismissWhatsNew()" style="padding:4px 10px; font-size:11px; background:#FFFFFF; color:#6B7280; border:1px solid #D1D5DB; border-radius:6px; cursor:pointer; white-space:nowrap;">✕ 닫기</button>
    </div>
  </div>
</div>

<!-- What's new 상세 모달 -->
<div id="whatsNewModal" style="display:none; position:fixed; inset:0; background:rgba(0,0,0,0.5); z-index:1000; align-items:center; justify-content:center;">
  <div style="background:#FFFFFF; border-radius:12px; max-width:720px; width:92%; max-height:84vh; overflow:hidden; display:flex; flex-direction:column;">
    <div style="padding:16px 20px; border-bottom:1px solid #E5E7EB; display:flex; justify-content:space-between; align-items:center;">
      <div><strong style="font-size:15px; color:#1E40AF;">📖 {{ app_version }} 릴리즈 노트</strong></div>
      <button onclick="document.getElementById('whatsNewModal').style.display='none'" style="border:none; background:none; font-size:20px; cursor:pointer; color:#6B7280;">✕</button>
    </div>
    <div style="padding:16px 20px; overflow:auto; font-size:13px; line-height:1.7; color:#374151;">
      <h3 style="margin:0 0 8px; color:#1E40AF;">🎨 {{ app_version }} — {{ app_version_date }} ({{ app_version_tagline }})</h3>
      <h4 style="color:#065F46; margin-top:16px;">✨ 이번 변경 요약</h4>
      <ul>
        {% for line in app_version_highlights %}<li>{{ line | safe }}</li>{% endfor %}
      </ul>
      <h4 style="color:#6B7280; margin-top:16px;">💡 전체 변경 이력</h4>
      <ul>
        <li>버전별 상세 내역은 프로젝트 루트의 <code>CHANGELOG.md</code> 참고</li>
        <li>이전 버전 기능 모두 유지 — 실행 방법 변경 없음</li>
      </ul>
    </div>
    <div style="padding:12px 20px; border-top:1px solid #E5E7EB; text-align:right;">
      <button onclick="document.getElementById('whatsNewModal').style.display='none'" style="padding:6px 16px; background:#2563EB; color:#FFFFFF; border:none; border-radius:6px; cursor:pointer;">확인</button>
    </div>
  </div>
</div>

<!-- 진행 스텝 바 -->
<div class="steps-bar">
  <div class="step-item active" id="stepBar1">
    <div class="step-num">1</div>📥 입력
  </div>
  <span class="step-arrow">›</span>
  <div class="step-item" id="stepBar2">
    <div class="step-num">2</div>⚙️ 분석
  </div>
  <span class="step-arrow">›</span>
  <div class="step-item" id="stepBar3">
    <div class="step-num">3</div>🔍 검토
  </div>
  <span class="step-arrow">›</span>
  <div class="step-item" id="stepBar4">
    <div class="step-num">4</div>✍️ TC 생성
  </div>
  <span class="step-arrow">›</span>
  <div class="step-item" id="stepBar5">
    <div class="step-num">5</div>📊 완료
  </div>
</div>

<main>

  <!-- Step 1: 입력 카드 -->
  <div class="card" id="card1">
    <div class="card-title">📥 입력 설정 <span class="badge">Step 1</span></div>

    <!-- 프로젝트 선택 -->
    <div class="project-select-row">
      <label>📁 프로젝트</label>
      <select class="project-dropdown" id="projectDropdown" onchange="onProjectDropdownChange()">
        <option value="">— 프로젝트를 선택하세요 —</option>
      </select>
      <button class="btn-new-project" onclick="toggleNewProjectInline()">+ 새 프로젝트</button>
      <button id="btnDeleteProject" style="display:none;padding:7px 10px;background:none;border:1px solid #E5E7EB;border-radius:8px;font-size:12px;cursor:pointer;color:#DC2626;" onclick="deleteDashProject()">삭제</button>
    </div>
    <div class="project-new-inline" id="newProjectInline">
      <input type="text" id="newDashProjectName" placeholder="새 프로젝트명 입력"
        onkeydown="if(event.isComposing||event.keyCode===229)return;if(event.key==='Enter')createDashProject();">
      <button class="btn-new-project" onclick="createDashProject()">생성</button>
      <button style="padding:7px 10px;background:none;border:1px solid var(--border);border-radius:8px;font-size:12px;cursor:pointer;" onclick="toggleNewProjectInline()">취소</button>
    </div>
    <div class="project-resume-bar" id="resumeBar">
      <span>💾 이전 작업이 있습니다 — <strong id="resumeStage"></strong> 단계까지 완료</span>
      <span id="resumeSavedAt" style="font-size:11px;opacity:0.7;margin-left:auto;"></span>
      <button class="btn-resume-pipeline" onclick="resumePipeline()">이어서 작업</button>
      <button style="padding:5px 10px;background:none;border:1px solid #93C5FD;border-radius:6px;font-size:11px;cursor:pointer;color:#1E40AF;" onclick="discardResume()">처음부터</button>
    </div>

    <!-- 모드 스위처 -->
    <div class="mode-switcher">
      <button type="button" class="mode-btn active" id="modeNew" onclick="switchMode('new')">
        🆕 신규 TC 생성
      </button>
      <button type="button" class="mode-btn" id="modeModify" onclick="switchMode('modify')">
        ✏️ 기존 TC 수정
      </button>
    </div>

    <!-- ── 신규 생성 모드 ── -->
    <div id="panelNew">


      <div class="form-group">
        <label class="form-label">프로젝트명</label>
        <input type="text" id="projectName" class="form-input"
               placeholder="예: Supercycl 모바일" value=""
               oninput="onProjectNameInputForResume()">
      </div>

      <!-- ═════════════════════════════════════════════════════════
           입력 소스 — 우선순위 재배치 (v0.10.x)
           1) 구조화 spec 폴더 = 기본 펼침 (권장 / 기획팀 표준)
           2) 개별 소스 = 기본 접힘 (임시 작업용)
           ═════════════════════════════════════════════════════════ -->

      <!-- ① 구조화 spec 폴더 (권장) ────────────────────────────── -->
      <div class="form-group">
        <div id="specFolderSection" style="border:1.5px solid #3B82F6;border-radius:10px;background:linear-gradient(135deg,#EFF6FF 0%,#DBEAFE 100%);overflow:hidden;">
          <button type="button" id="specFolderHeader" onclick="toggleAccordion('spec')"
                  style="width:100%;padding:14px 16px;background:transparent;border:none;cursor:pointer;display:flex;align-items:center;justify-content:space-between;text-align:left;">
            <span style="display:flex;align-items:center;gap:10px;">
              <span style="font-size:16px;">📁</span>
              <span style="font-weight:700;font-size:14px;color:#1E3A5F;">구조화 spec 폴더</span>
              <span style="font-size:11px;background:#3B82F6;color:#FFF;padding:2px 8px;border-radius:10px;font-weight:600;">권장</span>
              <span style="font-size:11.5px;color:#475569;font-weight:400;">기획팀 표준 — overview/policy/design/scr 분리</span>
            </span>
            <span id="specFolderChevron" style="font-size:12px;color:#3B82F6;font-weight:700;">▼</span>
          </button>
          <div id="specFolderBody" style="padding:0 16px 16px 16px;">

            <!-- 신규 spec 폴더 경로 + 최근 사용 드롭다운 -->
            <label style="display:flex;align-items:center;justify-content:space-between;margin-bottom:5px;">
              <span style="font-size:12px;color:#1E3A5F;font-weight:600;">신규 spec 폴더 경로 <span style="color:#DC2626;">*</span></span>
              <button type="button" id="btnRecentFolders" onclick="toggleRecentFolders()"
                      style="font-size:11px;color:#3B82F6;background:#FFFFFF;border:1px solid #93C5FD;border-radius:4px;padding:3px 9px;cursor:pointer;">
                최근 사용 ▾
              </button>
            </label>
            <input type="text" id="specFolderPath" class="form-input"
                   style="font-size:13px;font-family:monospace;background:#FFFFFF;"
                   placeholder="예: /Users/me/projects/specs/v0.47.2-2026-05-07"
                   oninput="onSpecFolderChanged()"/>
            <!-- 최근 사용 폴더 드롭다운 (프로젝트별) -->
            <div id="recentFoldersPanel" style="display:none;margin-top:4px;background:#FFFFFF;border:1px solid #93C5FD;border-radius:6px;max-height:200px;overflow-y:auto;font-size:12px;"></div>

            <!-- 버전 diff 모드 -->
            <div style="margin-top:10px;padding:10px 12px;background:#FFFBEB;border:1px solid #FCD34D;border-radius:6px;">
              <div style="display:flex;align-items:center;justify-content:space-between;">
                <label style="font-size:12px;color:#78350F;font-weight:600;display:flex;align-items:center;gap:5px;">
                  <span>🔄</span> 버전 diff 모드
                  <span style="font-weight:400;color:#92400E;">— 변경된 SCR 만 <strong>신규 TC 생성</strong></span>
                </label>
                <button type="button" id="btnDiffToggle" onclick="toggleDiffMode()"
                        style="padding:3px 10px;font-size:11px;border:1px solid #FCD34D;background:#FFFFFF;border-radius:4px;cursor:pointer;color:#78350F;">사용</button>
              </div>
              <div style="margin-top:6px; padding:6px 8px; background:#FFFFFF; border:1px dashed #FCD34D; border-radius:4px; font-size:10px; color:#92400E; line-height:1.6;">
                💡 <strong>언제 쓰나요?</strong> TC 가 아직 없거나 처음부터 새로 만들 때. 결과는 <strong>Excel 파일</strong>.<br>
                ⚠️ <strong>이미 운영 중인 Google Sheets TC 를 수정하고 싶다면</strong> → 상단의 <strong>"✏️ 기존 TC 수정"</strong> 모드 사용 (결과: Sheets 셀 직접 갱신 + 자동 백업).
              </div>
              <div id="diffBox" style="display:none;margin-top:8px;">
                <label style="display:flex;align-items:center;justify-content:space-between;margin-bottom:4px;">
                  <span style="font-size:11px;color:#78350F;font-weight:600;">이전 버전 폴더 경로</span>
                  <button type="button" id="btnRecentPrevFolders" onclick="toggleRecentPrevFolders()"
                          style="font-size:10px;color:#78350F;background:#FFFFFF;border:1px solid #FCD34D;border-radius:4px;padding:2px 7px;cursor:pointer;">
                    최근 사용 ▾
                  </button>
                </label>
                <input type="text" id="prevSpecFolderPath" class="form-input"
                       style="font-size:12px;font-family:monospace;background:#FFFFFF;"
                       placeholder="비교할 이전 버전 폴더 경로"
                       oninput="onSpecFolderChanged()"/>
                <div id="recentPrevFoldersPanel" style="display:none;margin-top:4px;background:#FFFFFF;border:1px solid #FCD34D;border-radius:6px;max-height:160px;overflow-y:auto;font-size:11px;"></div>
                <div style="margin-top:8px;display:flex;gap:8px;align-items:center;flex-wrap:wrap;">
                  <button type="button" onclick="previewDiff()" style="padding:5px 11px;font-size:11px;background:#F59E0B;color:#FFF;border:none;border-radius:5px;cursor:pointer;font-weight:500;">📊 변경 화면 미리보기</button>
                  <label style="font-size:11px;color:#78350F;display:flex;align-items:center;gap:5px;cursor:pointer;">
                    <input type="checkbox" id="includeUnchanged" style="margin:0;"/>
                    동일 화면도 재생성
                  </label>
                </div>
                <div id="diffPreview" style="display:none;margin-top:6px;padding:8px 10px;background:#FFFFFF;border:1px solid #FCD34D;border-radius:5px;font-size:11px;"></div>
              </div>
            </div>

            <!-- 액션 -->
            <div style="display:flex;gap:6px;margin-top:12px;flex-wrap:wrap;align-items:center;">
              <button type="button" onclick="previewSpecFolder()"
                      style="padding:7px 14px;font-size:12.5px;background:#3B82F6;color:#FFF;border:none;border-radius:6px;cursor:pointer;font-weight:600;">
                🔍 미리보기
              </button>
              <button type="button" id="btnResumeSpec" onclick="resumeSpecFolder()"
                      style="display:none;padding:7px 14px;font-size:12.5px;background:#7C3AED;color:#FFF;border:none;border-radius:6px;cursor:pointer;font-weight:600;">
                ▶ 이어서 작업
              </button>
              <span id="specFolderHint" style="font-size:11px;color:#475569;align-self:center;">폴더 경로 입력 후 미리보기로 분류 결과를 확인하세요.</span>
            </div>
            <div id="specFolderPreview" style="display:none;margin-top:10px;padding:10px 12px;background:#FFFFFF;border:1px solid #DBEAFE;border-radius:6px;font-size:12px;"></div>
          </div>
        </div>
      </div>

      <!-- ② 개별 소스 (임시 작업용 — 기본 접힘) ──────────────────── -->
      <div class="form-group">
        <div id="legacySourceSection" style="border:1px solid #CBD5E1;border-radius:10px;background:#F8FAFC;overflow:hidden;">
          <button type="button" id="legacySourceHeader" onclick="toggleAccordion('legacy')"
                  style="width:100%;padding:12px 16px;background:transparent;border:none;cursor:pointer;display:flex;align-items:center;justify-content:space-between;text-align:left;">
            <span style="display:flex;align-items:center;gap:10px;">
              <span style="font-size:14px;">📝</span>
              <span style="font-weight:600;font-size:13px;color:#475569;">개별 소스 입력</span>
              <span style="font-size:10.5px;background:#CBD5E1;color:#1F2937;padding:2px 7px;border-radius:8px;font-weight:500;">임시 작업용</span>
              <span style="font-size:11.5px;color:#94A3B8;font-weight:400;">PDF · GitHub URL · 웹페이지 · 마크다운 · 텍스트</span>
            </span>
            <span id="legacySourceChevron" style="font-size:12px;color:#94A3B8;font-weight:700;">▶</span>
          </button>
          <div id="legacySourceBody" style="display:none;padding:0 16px 14px 16px;">
            <div class="source-add-bar" style="margin-bottom:10px;">
              <button type="button" class="btn-add-source" onclick="addSource('pdf')">📄 PDF 추가</button>
              <button type="button" class="btn-add-source" onclick="addSource('url')">🔗 GitHub URL 추가</button>
              <button type="button" class="btn-add-source" onclick="addSource('web')">🌐 웹 URL 추가</button>
              <button type="button" class="btn-add-source" onclick="addSource('md')">📝 마크다운 파일 추가</button>
              <button type="button" class="btn-add-source" onclick="addSource('text')">✏️ 텍스트 추가</button>
              <button type="button" class="btn-clear-sources" id="btnClearAllSources" onclick="clearAllSources()" style="display:none;">🗑 전체 삭제</button>
            </div>
            <div id="sourceList"></div>
            <div id="sourceEmpty" class="source-empty">
              소스를 추가하세요.<br>
              <span style="font-size:12px">PDF · GitHub URL · 웹페이지 · 마크다운 파일 · 텍스트 등을 자유롭게 조합할 수 있습니다.</span>
            </div>
          </div>
        </div>
      </div>

      <div class="form-group">
        <label class="form-label" style="display:flex;align-items:center;justify-content:space-between">
          <span>TC 생성 범위</span>
          <span style="font-size:11px;color:var(--muted);font-weight:400">선택 — 비우면 전체 TC 생성</span>
        </label>
        <textarea id="focusArea" class="form-input" rows="9"
          style="resize:vertical; min-height:200px; font-size:13px; line-height:1.5;"
          oninput="onInputsChanged()"
          placeholder="특정 기능에 대해서만 TC를 만들려면 여기에 입력하세요.&#10;예) import 기능 / 로그인 및 회원가입 / 결제 모듈의 환불 처리&#10;💡 구조화 spec 모드 (관대한 인식):&#10;   • SCR-104 또는 SCR-102, SCR-104, SCR-106 (개별)&#10;   • SCR-102, 104, 106, 116 (일괄 — SCR 한 번만 적어도 OK)&#10;   • SCR-102~116 (범위)"></textarea>
        <div style="font-size:11px; color:var(--muted); margin-top:3px;">
          입력하면 해당 기능에 집중하여 TC를 생성합니다. 여러 기능은 쉼표 또는 줄바꿈으로 구분하세요.
        </div>
        <!-- 이전 범위 힌트 배너 (프로젝트 선택 시 동적으로 표시) -->
        <div id="previousFocusHint" style="display:none; margin-top:8px; padding:8px 12px; background:#FFFBEB; border:1px solid #FCD34D; border-radius:6px; font-size:12px; color:#92400E; display:none;">
          <div style="display:flex;align-items:center;justify-content:space-between;gap:8px;flex-wrap:wrap;">
            <div>
              💡 이전 작업 범위:
              <code id="previousFocusValue" style="background:#FFFFFF;padding:2px 6px;border-radius:3px;border:1px solid #FCD34D;font-family:monospace;margin-left:4px;"></code>
            </div>
            <div style="display:flex;gap:6px;">
              <button type="button" onclick="reusePreviousFocus()" style="padding:4px 10px;background:#F59E0B;color:#FFFFFF;border:none;border-radius:4px;font-size:11px;font-weight:600;cursor:pointer;">↪ 재사용</button>
              <button type="button" onclick="dismissPreviousFocusHint()" style="padding:4px 10px;background:#FFFFFF;color:#92400E;border:1px solid #FCD34D;border-radius:4px;font-size:11px;cursor:pointer;">✕ 무시</button>
            </div>
          </div>
        </div>
      </div>

      <!-- ▼ Combo 카드 (생성 모드 = 정책+Combo 일 때만 표시) ▼ -->
      <div class="form-group" id="comboGroup" style="display:none; padding:12px 14px; background:#FEF3C7; border:1px solid #FCD34D; border-radius:8px;">
        <label class="form-label" style="display:flex;align-items:center;gap:6px;margin-bottom:8px;">
          <span>🧩 Combo 명세 (옵션 조합 TC)</span>
          <span style="font-size:11px;color:var(--muted);font-weight:400;">spec 폴더 또는 프로젝트의 *_combinations.md</span>
        </label>
        <div id="comboFileList" style="font-size:12px; color:#6B7280; padding:8px; background:#FFFFFF; border-radius:4px;">
          <em>프로젝트 또는 spec 폴더가 설정되면 파일 목록이 표시됩니다.</em>
        </div>
        <!-- 시나리오 체크박스 (선택) — 명세에서 동적 로드 -->
        <div id="comboScenarioArea" style="display:none; margin-top:10px;">
          <div style="font-size:12px; font-weight:600; color:#374151; margin-bottom:6px;">생성할 시나리오 선택</div>
          <div id="comboScenarioList" style="background:#FFFFFF; padding:8px 10px; border-radius:4px; max-height:460px; overflow:auto;">
            <em style="color:#6B7280; font-size:12px;">파일 선택 시 자동 로드됩니다.</em>
          </div>
          <div style="font-size:11px;color:#6B7280; margin-top:4px;">
            전체 체크 시 → OC 매트릭스 전부 + 선택된 시나리오들이 생성됩니다 (체크 안 한 시나리오는 skip).
          </div>
        </div>

        <!-- + 새 명세 작성 / 갱신 버튼 -->
        <div style="margin-top:12px;display:flex;gap:8px;flex-wrap:wrap;align-items:center;">
          <button type="button" onclick="openComboSpecModal('new')"
                  style="padding:6px 14px;background:#10B981;color:white;border:0;border-radius:6px;font-size:12px;cursor:pointer;font-weight:600;">
            ➕ 새 명세 작성
          </button>
          <button type="button" id="btnEditComboSpec" onclick="openComboSpecModal('edit')"
                  style="padding:6px 14px;background:#F59E0B;color:white;border:0;border-radius:6px;font-size:12px;cursor:pointer;font-weight:600;">
            ✏️ 기존 명세 갱신
          </button>
          <span style="font-size:11px;color:#6B7280;margin-left:4px;">
            AI 가 spec 폴더를 분석해 *_combinations.md 초안을 작성합니다 (사용자 검토 후 저장)
          </span>
        </div>

        <div id="comboValidationMsg" style="margin-top:8px;font-size:12px;display:none;"></div>

        <!-- 고급 옵션: OC 필터 (펼침, 숨겨진 상태) -->
        <details style="margin-top:10px;background:#FFFFFF;border:1px dashed #D1D5DB;border-radius:6px;padding:6px 10px;">
          <summary style="cursor:pointer;font-size:11px;color:#6B7280;">⚙️ 고급 — OC 필터 (특정 조합만 생성)</summary>
          <div style="margin-top:8px;font-size:11px;color:#4B5563;">
            <label>OC ID 필터 (예: <code>OC-001,OC-070</code> 또는 <code>OC-07*</code> — 와일드카드)</label>
            <input type="text" id="comboOcFilter" class="form-input" style="font-size:12px;padding:6px 10px;margin-top:4px;"
                   placeholder="비우면 OC 매트릭스 전체 (시나리오 체크박스와 별개)">
            <div style="margin-top:4px;font-size:11px;color:#6B7280;">
              💡 OC ID 는 명세를 열어보면 확인 가능. 입력 시 시나리오 체크박스는 무시되고 OC 만 생성.
            </div>
          </div>
        </details>
      </div>
      <!-- ▲ /Combo 카드 ▲ -->

      <div class="form-group" id="generationModeGroup" style="padding:12px 14px;background:#F8FAFC;border:1px solid #E5E7EB;border-radius:8px;">
        <label class="form-label" style="display:flex;align-items:center;gap:6px;margin-bottom:8px;">
          <span>⚙️ 생성 모드</span>
          <span style="font-size:11px;color:var(--muted);font-weight:400;">선택</span>
        </label>
        <div style="display:flex;flex-direction:column;gap:8px;">
          <label style="display:flex;align-items:flex-start;gap:8px;padding:8px 10px;border:1px solid #D1D5DB;border-radius:6px;background:#FFFFFF;cursor:pointer;" id="modeStdLabel">
            <input type="radio" name="genMode" value="summary" checked onchange="onGenModeChanged()" style="margin-top:3px;">
            <div style="flex:1;">
              <div style="font-size:13px;font-weight:600;color:#111827;">정책 반영 모드 <span style="font-size:11px;color:#6B7280;font-weight:400;">(기본)</span></div>
              <div style="font-size:12px;color:#4B5563;margin-top:2px;">정책·기능·분류 3단계 분석으로 규칙을 체계적으로 반영합니다. 대용량·복잡한 문서에 적합.</div>
            </div>
          </label>
          <label style="display:flex;align-items:flex-start;gap:8px;padding:8px 10px;border:1px solid #D1D5DB;border-radius:6px;background:#FFFFFF;cursor:pointer;" id="modeQuickLabel">
            <input type="radio" name="genMode" value="direct" onchange="onGenModeChanged()" style="margin-top:3px;">
            <div style="flex:1;">
              <div style="font-size:13px;font-weight:600;color:#111827;">Quick 모드 <span style="font-size:11px;color:#6B7280;font-weight:400;">(원문 직접)</span></div>
              <div style="font-size:12px;color:#4B5563;margin-top:2px;">요약 단계 없이 원문을 직접 분류합니다. 누락 최소화 · 200KB 이하 권장.</div>
            </div>
          </label>
          <label style="display:flex;align-items:flex-start;gap:8px;padding:8px 10px;border:1px solid #D1D5DB;border-radius:6px;background:#FFFFFF;cursor:pointer;" id="modeComboLabel">
            <input type="radio" name="genMode" value="combo_only" onchange="onGenModeChanged()" style="margin-top:3px;">
            <div style="flex:1;">
              <div style="font-size:13px;font-weight:600;color:#111827;">Combo 모드 <span style="font-size:11px;color:#6B7280;font-weight:400;">(옵션 조합 TC 만 — SCR 기반 TC 생략)</span></div>
              <div style="font-size:12px;color:#4B5563;margin-top:2px;"><code>*_combinations.md</code> 명세 파일로 <strong>Combo TC 만</strong> 생성합니다. SCR 기반 TC 가 이미 만들어져 있을 때 추가 작업용. 비용 절감.</div>
            </div>
          </label>
        </div>
        <div id="modeSourceHint" style="font-size:11px;color:#6B7280;margin-top:8px;">
          ▸ 소스 크기: <span id="modeSourceSize">-</span> · <span id="modeSourceStatus">소스를 추가하면 권장 모드가 안내됩니다.</span>
        </div>
      </div>

      <button class="btn btn-primary" id="startBtn" onclick="startPipeline()">
        🚀 파이프라인 시작
      </button>
    </div>

    <!-- ── 수정 모드 (기획서 diff 기반 TC 갱신) ── -->
    <div id="panelModify" class="hidden">

      <!-- 직전 작업 불러오기 + ⚙ 설정 (1순위 — info-box 보다 먼저) -->
      <div style="display:flex; gap:10px; align-items:center; margin-bottom:14px;">
        <div id="updateRecentArea" style="display:none; position:relative;">
          <button type="button" id="updateRecentBtn" onclick="toggleUpdateRecent()"
                  style="background:#EEF2FF; border:1px solid #C7D2FE;
                         padding:10px 16px; border-radius:8px; cursor:pointer;
                         font-size:14px; color:#3730A3; font-weight:600;
                         display:inline-flex; align-items:center; gap:8px;">
            <span>📂 직전 작업 불러오기</span>
            <span id="updateRecentCount" style="background:#C7D2FE; color:#1E1B4B; font-size:11px; padding:2px 8px; border-radius:10px; font-weight:600;">0</span>
            <span id="updateRecentChevron" style="font-size:10px; margin-left:2px;">▾</span>
          </button>
          <div id="updateRecentList" style="display:none; position:absolute; top:100%; left:0; right:auto; min-width:420px; max-width:640px; margin-top:4px; z-index:50;
                  background:#FFF; border:1px solid #C7D2FE; border-radius:6px;
                  box-shadow:0 4px 12px rgba(0,0,0,0.08);
                  max-height:280px; overflow-y:auto;"></div>
        </div>

        <button type="button" id="updateConfigBtn" onclick="openConfigModal()"
                style="background:#F3F4F6; border:1px solid #D1D5DB;
                       padding:10px 14px; border-radius:8px; cursor:pointer;
                       font-size:13px; color:#374151; font-weight:500;
                       display:inline-flex; align-items:center; gap:6px;"
                title="사본/백업이 저장될 Drive 폴더 설정">
          <span>⚙</span>
          <span>설정</span>
          <span id="configFolderName" style="font-size:11px; color:#6B7280; max-width:200px; overflow:hidden; text-overflow:ellipsis; white-space:nowrap;"></span>
        </button>
      </div>

      <!-- 설정 모달 -->
      <div id="configModal" class="hidden" style="position:fixed; inset:0; z-index:10000; background:rgba(0,0,0,0.6); display:flex; align-items:center; justify-content:center; padding:20px;">
        <div style="background:#FFF; border-radius:12px; max-width:600px; width:100%; max-height:88vh; display:flex; flex-direction:column; overflow:hidden;">
          <div style="padding:18px 22px; border-bottom:1px solid #E5E7EB; display:flex; align-items:center; justify-content:space-between;">
            <div style="font-size:16px; font-weight:700; color:#1E3A5F;">⚙ TC Update 설정</div>
            <button onclick="closeConfigModal()" style="background:none; border:none; cursor:pointer; font-size:22px; color:#6B7280;">×</button>
          </div>
          <div style="padding:20px 22px; overflow-y:auto; flex:1;">
            <div style="font-size:13px; font-weight:600; color:#111827; margin-bottom:8px;">📁 사본·백업 보관 Drive 폴더</div>
            <div style="font-size:11px; color:#6B7280; margin-bottom:10px; line-height:1.6;">
              TC Update 작업으로 생성되는 사본/백업 Sheets 들이 이 폴더에 자동 저장됩니다.<br>
              • 개인 드라이브 폴더 (본인만 접근) 또는 공유 드라이브 폴더 (팀 공유) 가능<br>
              • 비워두면 폴더 지정 없이 Drive 루트에 만들어짐 (orphan)<br>
              • 동료가 도구 사용 시 본인의 Drive 폴더 ID 를 따로 설정해야 함 (각자 본인 계정으로 작업)
            </div>
            <input type="text" id="configFolderUrl" class="form-input"
                   placeholder="https://drive.google.com/drive/folders/..."
                   style="font-family:monospace; font-size:12px;">
            <div id="configCurrentInfo" style="margin-top:10px; padding:10px 12px; background:#F0F9FF; border:1px solid #BAE6FD; border-radius:6px; font-size:12px; color:#075985;"></div>
            <div id="configMsg" style="margin-top:10px; font-size:12px;"></div>
          </div>
          <div style="padding:14px 22px; border-top:1px solid #E5E7EB; display:flex; gap:10px; align-items:center; justify-content:flex-end;">
            <button class="btn" onclick="clearConfig()" style="background:#FEE2E2; color:#991B1B; border:1px solid #FCA5A5;">🗑 폴더 지정 제거</button>
            <button class="btn" onclick="closeConfigModal()" style="background:#F3F4F6;">취소</button>
            <button class="btn btn-primary" onclick="saveConfig()" id="configSaveBtn">💾 저장</button>
          </div>
        </div>
      </div>

      <div class="info-box" style="margin-bottom:12px;background:#EEF2FF;border-color:#C7D2FE;">
        📝 <strong>기획서 변경 기반 TC 갱신</strong> <span style="font-size:11px;color:#6B7280;">(v0.11.x)</span><br>
        <span style="font-size:12px;color:#4B5563;">
          이전 기획서 폴더와 새 기획서 폴더를 비교해 운영 중인 Google Sheets TC 를 자동 갱신합니다.<br>
          <strong>1단계 — 변경 분석</strong>: AI 가 SCR 별 변경 사항 요약<br>
          <strong>2단계 — TC 매핑</strong>: 사본 자동 생성 + 영향 받는 TC 후보 산출 + 개별/일괄 적용<br>
          <strong>3단계 — 원본 반영</strong>: 사본 검토 후 원본 Sheets 에 셀 단위 갱신 (자동 백업 + 롤백 가능)
        </span>
      </div>

      <!-- Slot 1: 이전 기획서 폴더 -->
      <div class="form-group">
        <label class="form-label">📁 이전 기획서 폴더 경로 <span style="color:#DC2626;">*</span></label>
        <input type="text" id="updatePrevFolder" class="form-input"
               placeholder="예: /Users/me/projects/spec-v0.51.0-2026-05-01"
               style="font-family:monospace;font-size:13px;">
        <div style="font-size:11px;color:var(--muted);margin-top:4px;">구조화 spec 폴더 (overview/policy/design/scr 분리)</div>
      </div>

      <!-- Slot 2: 새 기획서 폴더 -->
      <div class="form-group">
        <label class="form-label">📁 새 기획서 폴더 경로 <span style="color:#DC2626;">*</span></label>
        <input type="text" id="updateNewFolder" class="form-input"
               placeholder="예: /Users/me/projects/spec-v0.52.0-2026-05-08"
               style="font-family:monospace;font-size:13px;">
      </div>

      <!-- Slot 3: 갱신 대상 Google Sheets URL -->
      <div class="form-group">
        <label class="form-label">📑 갱신 대상 Google Sheets URL <span style="font-size:11px;color:#DC2626;font-weight:600">(Promote 시 이 시트가 변경됩니다)</span></label>
        <input type="text" id="updateTcUrl" class="form-input"
               placeholder="https://docs.google.com/spreadsheets/d/..."
               style="font-family:monospace;font-size:13px;">
        <div style="font-size:11px;color:var(--muted);margin-top:4px;">
          작업 흐름: 입력한 시트 → 자동 사본 생성 → 사본에서 검토 → Promote 시 입력 시트 갱신 (+백업).<br>
          💡 시험만 하고 싶으면 → Promote 단계 누르지 않으면 입력 시트는 안 변경됨 (사본만 만들어졌다 사라짐).
        </div>
      </div>

      <div style="display:flex; gap:8px; flex-wrap:wrap;">
        <button class="btn btn-primary" id="startDiffBtn" onclick="startSmartAnalyze()">
          🔍 변경사항 분석 (스마트)
        </button>
        <button class="btn" id="legacyDiffBtn" onclick="startUpdateAnalyze()" style="background:#F3F4F6;">
          ⚡ 빠른 분석 (한 번에)
        </button>
      </div>

      <!-- 스마트 분석 — 사전 견적 화면 -->
      <div id="smartPlanArea" class="hidden" style="margin-top:16px; padding:16px; background:#FFFFFF; border:1px solid #E5E7EB; border-radius:10px;">
        <div style="font-size:14px; font-weight:700; color:#1E3A5F; margin-bottom:10px;">
          📊 사전 견적 <span style="font-size:11px; color:#6B7280; font-weight:400;">(AI 호출 전 — 즉시)</span>
        </div>
        <div id="smartPlanStats" style="font-size:13px; color:#4B5563; margin-bottom:10px;"></div>
        <div id="smartPlanGroups" style="margin:12px 0; max-height:340px; overflow-y:auto;"></div>
        <div style="display:flex; gap:10px; margin-top:14px; align-items:center; flex-wrap:wrap;">
          <button class="btn" onclick="cancelSmartPlan()" style="background:#F3F4F6;">← 취소</button>
          <button class="btn btn-primary" id="startGroupAnalysisBtn" onclick="startGroupAnalysis()">
            ▶ 1번째 그룹부터 분석 시작
          </button>
          <span style="flex:1; font-size:12px; color:#6B7280;">
            각 그룹이 끝날 때마다 <strong>이어서/건너뛰기/종료</strong> 를 선택할 수 있어요.
          </span>
        </div>
      </div>

      <!-- 스마트 분석 — 그룹별 진행 화면 -->
      <div id="smartProgressArea" class="hidden" style="margin-top:16px; padding:16px; background:#FFFFFF; border:1px solid #E5E7EB; border-radius:10px;">
        <div style="font-size:14px; font-weight:700; color:#1E3A5F; margin-bottom:12px;">
          🧭 그룹별 분석 진행
        </div>
        <div id="smartProgressList" style="margin-bottom:12px;"></div>
        <div id="smartCurrentResult" style="margin:14px 0; padding:12px; background:#F9FAFB; border:1px solid #E5E7EB; border-radius:6px; font-size:13px; color:#111827; line-height:1.7; max-height:420px; overflow-y:auto;"></div>
        <div id="smartActionRow" style="display:flex; gap:10px; align-items:center; flex-wrap:wrap;">
          <!-- 동적으로 채워짐 -->
        </div>
      </div>

      <!-- 스마트 분석 — 최종 합본 -->
      <div id="smartFinalArea" class="hidden" style="margin-top:16px; padding:16px; background:#F0FDF4; border:1px solid #BBF7D0; border-radius:10px;">
        <div style="font-size:14px; font-weight:700; color:#166534; margin-bottom:10px;">
          ✅ 분석 완료 — 합본 보고서
        </div>
        <div id="smartFinalSummary" style="font-size:13px; color:#111827; line-height:1.7; max-height:540px; overflow-y:auto;"></div>
        <div style="display:flex; gap:10px; margin-top:14px;">
          <button class="btn" onclick="resetSmartAnalyze()" style="background:#F3F4F6;">← 새로 분석</button>
          <button class="btn btn-primary" id="planBtnFromSmart" onclick="startUpdatePlan()" style="flex:0 0 auto">
            🧭 2단계 — TC 매핑 + 영향 분석 (사본 생성)
          </button>
        </div>
      </div>

      <!-- 변경사항 리포트 영역 (분석 후 노출) -->
      <div id="diffReportArea" class="hidden" style="margin-top:16px; padding:14px; background:#FFFFFF; border:1px solid #E5E7EB; border-radius:10px;">
        <div style="font-size:14px; font-weight:700; color:#1E3A5F; margin-bottom:10px;">📊 변경사항 분석 결과 (1단계)</div>
        <div id="diffReportStats" style="font-size:13px; color:#4B5563; margin-bottom:10px;"></div>
        <div id="diffReportSummary" style="font-size:13px; color:#111827; line-height:1.7; margin-top:12px;"></div>
        <div style="display:flex; gap:10px; margin-top:14px; align-items:center;">
          <button class="btn" style="flex:0 0 auto" onclick="cancelUpdate()">← 입력 다시</button>
          <button class="btn btn-primary" id="planBtn" onclick="startUpdatePlan()" style="flex:0 0 auto">
            🧭 2단계 — TC 매핑 + 영향 분석 (사본 생성)
          </button>
          <span style="flex:1; font-size:12px; color:#6B7280;">
            기존 TC Sheets 의 사본을 만들고, 변경 SCR 에 영향받는 TC 후보를 보여줍니다.
          </span>
        </div>
      </div>

      <!-- 2단계 — 영향 분석 (Plan) 결과 -->
      <div id="planReportArea" class="hidden" style="margin-top:16px; padding:14px; background:#FFFFFF; border:1px solid #E5E7EB; border-radius:10px;">
        <div style="font-size:14px; font-weight:700; color:#1E3A5F; margin-bottom:10px;">🧭 영향 분석 (2단계)</div>
        <div id="planStats" style="font-size:13px; color:#4B5563; margin-bottom:10px;"></div>
        <div id="planCopyLink" style="font-size:13px; margin-bottom:10px;"></div>

        <!-- 필터 -->
        <div style="display:flex; gap:8px; margin-bottom:10px; flex-wrap:wrap;">
          <label style="font-size:12px;"><input type="checkbox" id="filterAdd" checked onchange="renderCandidates()"> 🆕 신규</label>
          <label style="font-size:12px;"><input type="checkbox" id="filterModify" checked onchange="renderCandidates()"> ✏️ 수정</label>
          <label style="font-size:12px;"><input type="checkbox" id="filterDelete" checked onchange="renderCandidates()"> 🗑️ 삭제</label>
          <span style="flex:1"></span>
          <button class="btn" style="font-size:12px; padding:4px 10px;" onclick="checkAll(true)">모두 체크</button>
          <button class="btn" style="font-size:12px; padding:4px 10px;" onclick="checkAll(false)">모두 해제</button>
        </div>

        <div id="candidatesList" style="max-height:480px; overflow-y:auto; border:1px solid #E5E7EB; border-radius:6px;"></div>

        <!-- 일괄 적용 영역 — 사본의 실제 셀을 AI 제안으로 갱신 -->
        <div style="margin-top:14px; padding:12px 14px; background:#EEF2FF; border:1px solid #C7D2FE; border-radius:8px;">
          <div style="font-size:13px; font-weight:600; color:#3730A3; margin-bottom:6px;">
            📦 일괄 적용 — 체크된 수정 후보 전부에 AI 제안 + 사본 셀 갱신
          </div>
          <div style="font-size:11px; color:#4B5563; margin-bottom:10px; line-height:1.6;">
            각 후보마다 AI 가 수정안을 만들고 <strong>사본 시트의 셀을 실제로 갱신</strong>합니다.<br>
            건당 약 5초 소요. 이미 TC Edit Log 에 있는 항목은 자동 skip 됩니다.<br>
            <em>개별로 검토하면서 "이 TC 만 사본에 적용" 으로도 하나씩 적용 가능합니다.</em>
          </div>
          <div style="display:flex; gap:10px; align-items:center; flex-wrap:wrap;">
            <button class="btn btn-primary" id="applyBulkBtn" onclick="applyBulk()" style="flex:0 0 auto;">
              📦 선택 항목 일괄 적용 (AI 호출 + 사본 셀 갱신)
            </button>
            <span id="applyStatus" style="flex:1; font-size:12px; color:#4B5563;">
              적용 후 사본 'TC Edit Log' 시트에 이력 기록 + 사본 셀에 실제 반영됩니다.
            </span>
          </div>
          <div id="applyBulkProgress" class="hidden" style="margin-top:10px;"></div>
        </div>

        <!-- 3단계: 사본 → 원본 적용 (별도 영역, 시각적 구분) -->
        <div style="margin-top:24px; padding:14px 16px; background:linear-gradient(135deg, #FEF2F2 0%, #FEE2E2 100%); border:2px solid #DC2626; border-radius:10px;">
          <div style="font-size:14px; font-weight:700; color:#991B1B; margin-bottom:6px;">
            ⚠️ 3단계 — 사본 → 원본 적용 (실제 갱신)
          </div>
          <div style="font-size:12px; color:#7F1D1D; margin-bottom:12px; line-height:1.6;">
            지금까지는 모두 <strong>사본</strong> 에서 작업했습니다. 검토가 끝났다면 사본의 변경을 <strong>원본 Sheets 에 반영</strong> 합니다.<br>
            원본은 자동으로 <strong>백업</strong> 되며, 적용 후 <strong>롤백</strong> 가능합니다.
          </div>
          <div style="display:flex; gap:10px; align-items:center; flex-wrap:wrap;">
            <button class="btn" id="promoteBtn" onclick="startPromote()" style="background:#DC2626; color:#FFF; border:none; font-weight:600;">
              ✏️ 원본에 반영 (미리보기 → 확인 → 적용)
            </button>
            <button class="btn" id="rollbackBtn" onclick="startRollback()" style="background:#FFF; color:#991B1B; border:1px solid #DC2626; display:none;">
              🔄 마지막 적용 롤백
            </button>
            <span id="promoteStatus" style="flex:1; font-size:11px; color:#7F1D1D;"></span>
          </div>
        </div>

      </div>
    </div>
  </div>

  <!-- Promote 미리보기 모달 -->
  <div id="promoteModal" class="hidden" style="position:fixed; inset:0; z-index:10000; background:rgba(0,0,0,0.6); display:flex; align-items:center; justify-content:center; padding:20px;">
    <div style="background:#FFF; border-radius:12px; max-width:880px; width:100%; max-height:88vh; display:flex; flex-direction:column; overflow:hidden;">
      <div style="padding:18px 22px; border-bottom:1px solid #E5E7EB; display:flex; align-items:center; justify-content:space-between;">
        <div style="font-size:16px; font-weight:700; color:#991B1B;">⚠️ 원본 Sheets 갱신 — 미리보기</div>
        <button onclick="closePromoteModal()" style="background:none; border:none; cursor:pointer; font-size:22px; color:#6B7280;">×</button>
      </div>
      <div id="promoteModalBody" style="padding:18px 22px; overflow-y:auto; flex:1;"></div>
      <div style="padding:14px 22px; border-top:1px solid #E5E7EB; display:flex; gap:10px; align-items:center; justify-content:flex-end;">
        <button class="btn" onclick="closePromoteModal()" style="background:#F3F4F6;">취소</button>
        <button class="btn btn-primary" id="promoteConfirmBtn" onclick="confirmPromote()" style="background:#DC2626; color:#FFF; border:none; font-weight:600;" disabled>
          ✏️ 원본에 적용 (백업 + 갱신)
        </button>
      </div>
    </div>
  </div>

  <!-- Step 2~4: 파이프라인 통합 카드 (Gate는 별도 card3에서 사용자 개입) -->
  <div class="card hidden" id="card2">
    <div class="card-title" style="display:flex;align-items:center;justify-content:space-between;gap:10px;flex-wrap:wrap;">
      <span>
        <span id="pipelineCardTitle">⚙️ 파이프라인 실행</span>
        <span class="badge" id="pipelineCardBadge">Step 2</span>
      </span>
      <button type="button" onclick="restartFromScratch()"
              style="padding:6px 12px;font-size:12px;background:#F1F5F9;color:#1F2937;border:1px solid #94A3B8;border-radius:6px;cursor:pointer;font-weight:500;">
        ← 처음으로 (입력 변경)
      </button>
    </div>

    <div class="substeps">
      <div class="substep" id="sub1">📄 파싱</div>
      <div class="substep" id="sub2">🔎 정책·기능</div>
      <div class="substep" id="sub3">🗂 분류</div>
      <div class="substep" id="sub4">🔍 검토</div>
      <div class="substep" id="sub5">✍️ TC 생성</div>
      <div class="substep" id="sub6">📊 Excel</div>
    </div>

    <div style="font-size:14px; font-weight:600; color:var(--blue); margin-bottom:6px;">
      <span id="stageLabel">시작 중...</span><span id="countdownLabel" style="font-weight:400; font-size:12px; color:var(--muted);"></span>
    </div>
    <div class="progress-wrap">
      <div class="progress-bar" id="progressBar" style="width:0%"></div>
    </div>

    <div class="log-box" id="logBox" style="max-height:465px; min-height:425px;"></div>
    <div style="margin-top:12px; text-align:right;">
      <button class="btn-stop" id="stopBtn2" onclick="stopPipeline()">⏹ 파이프라인 중단</button>
    </div>
  </div>

  <!-- Step 3: Human Gate 카드 -->
  <div class="card hidden" id="card3">
    <div class="card-title" id="gateTitle" style="display:flex;align-items:center;justify-content:space-between;gap:10px;flex-wrap:wrap;">
      <span>🔍 분류표 검토 <span class="badge">Step 3 · Human Gate</span></span>
      <span style="display:flex;gap:6px;flex-wrap:wrap;">
        <button type="button" onclick="restartFromScratch()"
                title="입력 필드 값은 그대로 유지 + Step 1 화면으로 복귀 (TC 생성 범위 등 변경 가능)"
                style="padding:6px 12px;font-size:12px;background:#FEF3C7;color:#78350F;border:1px solid #FBBF24;border-radius:6px;cursor:pointer;font-weight:600;">
          ← 입력 화면으로 (값 유지)
        </button>
      </span>
    </div>
    <!-- 통합 네비게이션 — 평시: 입력 변경 안내 / 중단 후: 처음부터/이어서 -->
    <div id="gateNavBox" style="margin:0 0 10px 0; padding:10px 14px; background:#FEF3C7; border:1px dashed #FBBF24; border-radius:6px; font-size:12px; color:#78350F; display:flex; gap:10px; align-items:center; flex-wrap:wrap;">
      <div id="gateNavMsg" style="flex:1; min-width:240px;">
        💡 <strong>분류표를 다시 만들고 싶나요?</strong><br>
        <span style="font-size:11px;">예: TC 생성 범위 추가/수정 — 입력값은 그대로 보존됩니다.</span>
      </div>
      <div id="gateNavButtons" style="flex:0 0 auto; display:flex; gap:8px; flex-wrap:wrap;">
        <button type="button" onclick="restartFromScratch()"
                style="padding:8px 16px; font-size:13px; background:#F59E0B; color:#FFF; border:none; border-radius:6px; cursor:pointer; font-weight:600;">
          ← 입력 화면으로 (값 유지)
        </button>
      </div>
    </div>
    <div class="info-box" id="gateInfoBox">
      AI가 생성한 분류표입니다. 하단 AI 도우미로 수정을 요청하고, 아래 표에서 결과를 확인한 뒤 승인하세요.
    </div>

    <!-- 입력 소스 화면 식별자 안내 (v0.9.13~) — gate 이벤트 도착 시 채워짐 -->
    <div id="sourceScrNotice" style="display:none;margin:10px 0;padding:10px 14px;border-radius:8px;font-size:12.5px;line-height:1.55;"></div>

    <!-- v0.9.22 Stage 1 — 분류표 단계 중복 가능성 예측 알림 -->
    <div id="classifyDuplicateNotice" style="display:none;margin:10px 0;padding:14px 16px;background:#FFFBEB;border:1.5px solid #FCD34D;border-radius:10px;"></div>

    <!-- TC ID 생성 방식 선택 패널 (독립 영역, 항상 표시) -->
    <div id="tcIdModePanel" style="background:linear-gradient(135deg, #EFF6FF 0%, #DBEAFE 100%);border:2px solid #3B82F6;border-radius:12px;padding:16px 20px;margin:14px 0 18px 0;box-shadow:0 2px 8px rgba(59, 130, 246, 0.15);">
      <div style="display:flex;align-items:center;justify-content:space-between;gap:16px;flex-wrap:wrap;">
        <div style="flex:1;min-width:300px;">
          <div style="font-size:15px;font-weight:700;color:#1E3A5F;margin-bottom:6px;display:flex;align-items:center;gap:8px;">
            <span style="font-size:18px;">🪪</span>
            <span>TC ID 생성 방식</span>
          </div>
          <div id="tcIdModeDesc" style="font-size:12.5px;color:#1E3A5F;line-height:1.6;">
            체크하면 시스템이 화면별 코드(<code style="background:#FFFFFF;padding:1px 5px;border-radius:3px;border:1px solid #93C5FD;font-family:monospace;">SM-SPL-001</code>, <code style="background:#FFFFFF;padding:1px 5px;border-radius:3px;border:1px solid #93C5FD;font-family:monospace;">SM-LGI-001</code>...)로 자동 매핑합니다.<br>
            해제하면 아래 <strong>TC 분류 요약</strong> 표에서 SuiteCode를 직접 입력할 수 있습니다.
          </div>
        </div>
        <label id="tcIdModeLabel" style="display:inline-flex;align-items:center;gap:10px;background:#FFFFFF;padding:12px 16px;border-radius:10px;border:2px solid #3B82F6;cursor:pointer;user-select:none;font-weight:700;color:#1E3A5F;box-shadow:0 1px 3px rgba(0,0,0,0.08);white-space:nowrap;transition:all 0.2s;">
          <input type="checkbox" id="tcIdModeToggle" checked onchange="setTcIdMode(this.checked)" style="width:20px;height:20px;cursor:pointer;accent-color:#3B82F6;">
          <span style="font-size:13.5px;">System Generated TC IDs</span>
        </label>
      </div>
    </div>

    <!-- AI 채팅 영역 (단독 폭) — v0.9.12: 시각적으로 숨김. 미니 채팅 패널이 UI 담당.
         단, sendGateChat() / addGateChatMsg() 가 #gateChatInput / #gateChatMessages 를
         참조하므로 DOM 자체는 유지 (단일 진실 소스). 미니/모달 채팅이 미러링하여 사용. -->
    <div class="gate-chat-panel" style="display:none;">
      <div class="gate-chat-header">
        💬 AI와 대화하여 수정
      </div>
      <div class="gate-chat-messages" id="gateChatMessages">
        <!-- 메시지가 여기에 추가됨 -->
      </div>
      <div class="gate-chat-input-row">
        <textarea class="gate-chat-input" id="gateChatInput"
          placeholder="수정 요청을 입력하세요. 예) AUTH 대분류 케이스 3번 삭제해줘"
          onkeydown="if((event.isComposing||event.keyCode===229))return;if(event.key==='Enter'&&!event.shiftKey){event.preventDefault();sendGateChat();}"></textarea>
        <button class="gate-chat-send" id="gateChatSend" onclick="sendGateChat()">전송</button>
      </div>
    </div>

    <!-- 분류표 & TC 매핑 표 영역 — 메인 (renderGateViewer가 채워넣음) -->
    <div id="gateViewer">
      <!-- TC 분류 요약 표가 여기에 들어감 -->
    </div>

    <!-- 원본 마크다운 보기 — 토글 가능 (기본 접힘) -->
    <details style="margin:14px 0;">
      <summary style="cursor:pointer; padding:8px 14px; background:#F3F4F6; border:1px solid #D1D5DB; border-radius:8px; font-size:13px; font-weight:600; color:#374151; user-select:none; list-style:none; display:flex; align-items:center; gap:6px;">
        <span id="rawDocToggleIcon">▶</span>
        <span>📄 분류표 원본 마크다운 보기 (펼치기)</span>
      </summary>
      <div id="rawDocContainer" style="margin-top:8px; padding:14px; background:#FAFAFA; border:1px solid #E5E7EB; border-radius:8px;">
        <div id="rawDocBadge" style="font-size:11px; color:var(--muted); margin-bottom:8px;"></div>
        <pre id="rawDocContent" style="margin:0; padding:12px; background:#FFFFFF; border:1px solid #E5E7EB; border-radius:6px; font-size:12px; line-height:1.7; max-height:400px; overflow:auto; white-space:pre-wrap; word-break:break-word; font-family:ui-monospace,SFMono-Regular,Menlo,monospace; color:#111827;"></pre>
      </div>
    </details>
    <script>
      // details 펼치기/접기 시 ▶ ↔ ▼ 화살표 + 라벨 전환
      // (분류표 원본 + TC 분류 요약 모두 처리)
      (function() {
        document.addEventListener('toggle', function(e) {
          if (!e.target || e.target.tagName !== 'DETAILS') return;
          // 1) 원본 마크다운 보기 토글
          var rawIcon = e.target.querySelector('#rawDocToggleIcon');
          if (rawIcon) {
            rawIcon.textContent = e.target.open ? '▼' : '▶';
            var sumText = e.target.querySelector('summary span:last-child');
            if (sumText) sumText.textContent = e.target.open ? '📄 분류표 원본 마크다운 보기 (접기)' : '📄 분류표 원본 마크다운 보기 (펼치기)';
            return;
          }
          // 2) TC 분류 요약 토글 (renderGateViewer 가 동적 생성)
          if (e.target.id === 'tcSummaryDetails') {
            var sIcon = e.target.querySelector('#tcSummaryToggleIcon');
            if (sIcon) sIcon.textContent = e.target.open ? '▼' : '▶';
            var hint = e.target.querySelector('#tcSummaryFoldHint');
            if (hint) hint.textContent = e.target.open ? '(클릭하여 접기)' : '(클릭하여 펼치기)';
          }
        }, true);
      })();
    </script>

    <!-- 숨겨진 원본 데이터 저장용 (gate-chat 등에서 참조) -->
    <textarea class="gate-textarea" id="gateContent" style="display:none;"></textarea>

    <!-- SuiteCode 입력 테이블 (Viewer 밖 — 독립 영역) -->
    <div id="suiteCodeSection" style="display:none; margin-top:14px; margin-bottom:14px;"></div>

    <!-- Excel 출력 옵션 (Full/Light/Custom) — 승인 시 step_build_excel로 전달됨 -->
    <div id="excelOptionPanel" style="margin-top:18px;background:#F0F9FF;border:1.5px solid #93C5FD;border-radius:10px;padding:14px 16px;">
      <div style="font-size:14px;font-weight:700;color:#1E3A5F;margin-bottom:10px;display:flex;align-items:center;gap:8px;">
        <span>⚙</span><span>Excel 출력 옵션</span>
        <span style="font-size:11px;color:#6B7280;font-weight:400;">— 승인 시 적용됩니다</span>
      </div>
      <div style="display:flex;flex-direction:column;gap:6px;">
        <label class="excel-preset-row" style="display:flex;align-items:flex-start;gap:10px;cursor:pointer;padding:8px 10px;border-radius:8px;background:#FFFFFF;border:1.5px solid #DBEAFE;">
          <input type="radio" name="excelPreset" value="full" checked onchange="onExcelPresetChange()" style="margin-top:3px;">
          <div style="flex:1;">
            <div style="font-size:13px;font-weight:700;color:#1E3A5F;">📦 Full Set <span style="font-size:11px;color:#6B7280;font-weight:400;">(정식 산출물)</span></div>
            <div style="font-size:11.5px;color:#4B5563;margin-top:2px;">표지 · 통계 · Smoke · Traceability · 변경 이력 · TC 전체 — 정식 배포·공유용</div>
          </div>
        </label>
        <label class="excel-preset-row" style="display:flex;align-items:flex-start;gap:10px;cursor:pointer;padding:8px 10px;border-radius:8px;background:#FFFFFF;border:1.5px solid #DBEAFE;">
          <input type="radio" name="excelPreset" value="light" onchange="onExcelPresetChange()" style="margin-top:3px;">
          <div style="flex:1;">
            <div style="font-size:13px;font-weight:700;color:#1E3A5F;">🪶 Light <span style="font-size:11px;color:#6B7280;font-weight:400;">(중간 산출물 · 반복 작업용)</span></div>
            <div style="font-size:11.5px;color:#4B5563;margin-top:2px;">TC 전체 목록만 — 표지·통계 등 제외 (빠른 확인용)</div>
          </div>
        </label>
        <label class="excel-preset-row" style="display:flex;align-items:flex-start;gap:10px;cursor:pointer;padding:8px 10px;border-radius:8px;background:#FFFFFF;border:1.5px solid #DBEAFE;">
          <input type="radio" name="excelPreset" value="custom" onchange="onExcelPresetChange()" style="margin-top:3px;">
          <div style="flex:1;">
            <div style="font-size:13px;font-weight:700;color:#1E3A5F;">🛠 Custom <span style="font-size:11px;color:#6B7280;font-weight:400;">(시트별 직접 선택)</span></div>
            <div style="font-size:11.5px;color:#4B5563;margin-top:2px;">아래에서 포함할 시트를 직접 체크하세요</div>
          </div>
        </label>
      </div>
      <!-- Custom 모드 — 시트별 체크박스 (기본 숨김) -->
      <div id="excelCustomPanel" style="display:none;margin-top:10px;padding:12px;background:#FFFFFF;border:1px solid #DBEAFE;border-radius:8px;">
        <div style="font-size:11.5px;color:#1E3A5F;font-weight:600;margin-bottom:8px;">포함할 시트 선택</div>
        <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:6px 14px;">
          <label style="display:flex;align-items:center;gap:6px;font-size:12.5px;color:#374151;cursor:pointer;">
            <input type="checkbox" class="excel-sheet-cb" data-sheet="cover" onchange="onExcelSheetCheckChange()" checked> 📋 표지
          </label>
          <label style="display:flex;align-items:center;gap:6px;font-size:12.5px;color:#374151;cursor:pointer;">
            <input type="checkbox" class="excel-sheet-cb" data-sheet="stats" onchange="onExcelSheetCheckChange()" checked> 📊 TC 통계
          </label>
          <label style="display:flex;align-items:center;gap:6px;font-size:12.5px;color:#374151;cursor:pointer;">
            <input type="checkbox" class="excel-sheet-cb" data-sheet="smoke" onchange="onExcelSheetCheckChange()" checked> 🔥 Smoke Test
          </label>
          <label style="display:flex;align-items:center;gap:6px;font-size:12.5px;color:#374151;cursor:pointer;">
            <input type="checkbox" class="excel-sheet-cb" data-sheet="traceability" onchange="onExcelSheetCheckChange()" checked> 🔗 Traceability
          </label>
          <label style="display:flex;align-items:center;gap:6px;font-size:12.5px;color:#9CA3AF;cursor:not-allowed;" title="필수 시트 — 항상 포함">
            <input type="checkbox" class="excel-sheet-cb" data-sheet="tc_list" checked disabled> 📌 TC 전체 목록 <span style="font-size:10px;">(필수)</span>
          </label>
          <label style="display:flex;align-items:center;gap:6px;font-size:12.5px;color:#374151;cursor:pointer;" id="excelChangeHistoryLabel">
            <input type="checkbox" class="excel-sheet-cb" data-sheet="change_history" onchange="onExcelSheetCheckChange()" checked> 🔄 변경 이력 <span style="font-size:10px;color:#6B7280;">(수정 모드 전용)</span>
          </label>
        </div>
        <div id="excelSheetSummary" style="margin-top:10px;font-size:11.5px;color:#1E3A5F;background:#EFF6FF;padding:6px 10px;border-radius:6px;">
          💡 요약: <strong>전체 시트 (6개)</strong>
        </div>
      </div>
    </div>

    <div style="display:flex; gap:10px; align-items:center; flex-wrap:wrap; margin-top:18px;">
      <button class="btn btn-success" onclick="approveGate()">
        ✅ 승인 및 TC 생성 시작
      </button>
      <button style="padding:8px 16px;background:#EFF6FF;color:#1D4ED8;border:1.5px solid #93C5FD;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;" onclick="regenerateClassification()">
        🔄 분류표 다시 생성
      </button>
      <button type="button" onclick="exportGateExcel()"
        style="padding:8px 16px; border:1.5px solid var(--teal); border-radius:8px;
          background:none; cursor:pointer; color:var(--teal); font-weight:600; font-size:13px;
          display:flex; align-items:center; gap:5px;">
        📥 Excel로 내보내기
      </button>
      <button class="btn-stop" id="stopBtn3" onclick="stopPipeline()">⏹ 여기서 중단</button>
    </div>
  </div>

  <!-- Step 4는 card2로 통합됨 (진행 상태·로그를 한곳에서 표시) -->

  <!-- Step 5: 완료 카드 -->
  <div class="card hidden" id="card5">
    <div class="card-title">📊 완료 <span class="badge">Step 5</span></div>

    <!-- 파일 정보 행 -->
    <div class="result-file">
      <span class="ricon">📊</span>
      <div class="rinfo">
        <div class="rname" id="resultFilename">—</div>
        <div class="rsize" id="resultMeta">—</div>
      </div>
      <button onclick="openFolder()"
        style="margin-left:auto;background:none;border:1px solid #CBD5E0;
          color:#555;font-size:12px;padding:5px 12px;border-radius:8px;
          cursor:pointer;white-space:nowrap">
        📁 폴더 열기
      </button>
    </div>

    <!-- 원칙 G — 중복 의심 TC 알림 (v0.9.17~). duplicate_warning SSE 이벤트로 채워짐 -->
    <div id="duplicateNotice" style="display:none;margin:12px 0;padding:14px 16px;background:#FFFBEB;border:1.5px solid #FCD34D;border-radius:10px;"></div>

    <!-- TC 통계 -->
    <div class="tc-stats" style="margin-bottom:16px">
      <div class="tc-stat">
        <div class="tc-stat-num" id="statTotal">—</div>
        <div class="tc-stat-label">총 TC</div>
      </div>
      <div class="tc-stat">
        <div class="tc-stat-num" id="statSmoke">—</div>
        <div class="tc-stat-label">🔥 Smoke TC</div>
      </div>
    </div>

    <!-- 추가 작업 -->
    <div style="display:flex; gap:10px; margin-bottom:16px; flex-wrap:wrap;">
      <button onclick="startNextIteration()" style="padding:10px 20px;background:#2563EB;color:#fff;border:none;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;">
        🔄 추가 TC 생성 (소스/범위 수정 후 재시작)
      </button>
      <button onclick="startNextIteration(true)" style="padding:10px 16px;background:#fff;color:#1D4ED8;border:1.5px solid #93C5FD;border-radius:8px;font-size:13px;cursor:pointer;">
        📝 범위만 변경하여 재시작
      </button>
    </div>

    <!-- 액션 버튼 행 -->
    <div class="action-row">
      <button class="btn-drive" id="driveBtn" onclick="uploadToDrive()">
        <svg width="16" height="16" viewBox="0 0 87.3 78" xmlns="http://www.w3.org/2000/svg">
          <path d="m6.6 66.85 3.85 6.65c.8 1.4 1.95 2.5 3.3 3.3l13.75-23.8h-27.5c0 1.55.4 3.1 1.2 4.5z" fill="#0066da"/>
          <path d="m43.65 25-13.75-23.8c-1.35.8-2.5 1.9-3.3 3.3l-25.4 44a9.06 9.06 0 0 0-1.2 4.5h27.5z" fill="#00ac47"/>
          <path d="m73.55 76.8c1.35-.8 2.5-1.9 3.3-3.3l1.6-2.75 7.65-13.25c.8-1.4 1.2-2.95 1.2-4.5h-27.502l5.852 11.5z" fill="#ea4335"/>
          <path d="m43.65 25 13.75-23.8c-1.35-.8-2.9-1.2-4.5-1.2h-18.5c-1.6 0-3.15.45-4.5 1.2z" fill="#00832d"/>
          <path d="m59.8 53h-32.3l-13.75 23.8c1.35.8 2.9 1.2 4.5 1.2h50.8c1.6 0 3.15-.45 4.5-1.2z" fill="#2684fc"/>
          <path d="m73.4 26.5-12.7-22c-.8-1.4-1.95-2.5-3.3-3.3l-13.75 23.8 16.15 27h27.45c0-1.55-.4-3.1-1.2-4.5z" fill="#ffba00"/>
        </svg>
        Google Drive에 올리기
      </button>
    </div>
  </div>

</main>

<!-- 오류 복구 모달 -->
<div class="modal-bg" id="recovery-modal">
  <div class="modal">
    <h2>⚠️ 파이프라인 오류 발생</h2>
    <p id="recoveryErrorMsg" style="font-size:13px;color:#e53e3e;margin-bottom:12px"></p>
    <div id="recoveryCheckpointInfo" class="hidden" style="font-size:12px;color:#555;background:#f7f7f7;padding:10px;border-radius:8px;margin-bottom:14px"></div>
    <div style="display:flex;flex-direction:column;gap:8px">
      <button class="btn btn-primary" id="btnResume" onclick="restartModify('resume')" style="display:none">
        🔄 이어서 재시작 (체크포인트 복원)
      </button>
      <button class="btn btn-primary" onclick="restartModify('fresh')">
        ▶ 처음부터 재시작
      </button>
      <button class="btn" onclick="closeRecoveryModal()">
        ✕ 종료
      </button>
    </div>
  </div>
</div>

<!-- Google Drive 연동 안내 모달 -->
<div class="modal-bg" id="drive-modal">
  <div class="modal">
    <h2>🔑 Google Drive 연동 설정</h2>
    <p style="font-size:13px;color:#444;margin-bottom:14px">
      최초 1회 <strong>credentials.json</strong> 파일 설정이 필요합니다.
    </p>
    <ol>
      <li><a href="https://console.cloud.google.com/apis/credentials" target="_blank"
          style="color:#3557A0">Google Cloud Console</a> 접속</li>
      <li><strong>+ 사용자 인증 정보 만들기 → OAuth 클라이언트 ID</strong></li>
      <li>애플리케이션 유형: <strong>데스크톱 앱</strong> 선택 후 만들기</li>
      <li>JSON 다운로드 → <strong>tc-ui 폴더에 <code>credentials.json</code>으로 저장</strong></li>
      <li>저장 후 다시 <strong>Drive에 올리기</strong> 클릭 → 브라우저에서 Google 계정 인증</li>
    </ol>
    <p style="font-size:12px;color:#888;margin-top:10px">
      또한 <code>config.json</code>에 <code>google_drive.upload_folder_id</code>를 설정해야 합니다.
    </p>
    <div class="modal-btns">
      <button class="btn-open-drive"
        onclick="window.open('https://console.cloud.google.com/apis/credentials','_blank')">
        Cloud Console 열기
      </button>
      <button class="btn-modal-close" onclick="closeDriveModal()">닫기</button>
    </div>
  </div>
</div>

<!-- Google Drive 폴더 선택 모달 -->
<div class="modal-bg" id="drive-folder-modal">
  <div class="modal" style="max-width:540px;">
    <h2 style="display:flex;align-items:center;gap:8px;">📁 업로드할 Drive 폴더 선택</h2>
    <div id="driveFolderEmail" style="font-size:12px;color:#666;margin-bottom:10px;"></div>
    <input type="text" id="driveFolderSearch" placeholder="폴더명 검색..."
      oninput="loadDriveFolders(this.value)"
      style="width:100%;padding:8px 12px;border:1.5px solid #D0D7DE;border-radius:8px;font-size:13px;margin-bottom:10px;">
    <div id="driveFolderList" style="max-height:320px;overflow-y:auto;border:1px solid #E2E8F0;border-radius:8px;background:#F9FAFB;"></div>
    <div id="driveSelectedFolder" style="font-size:12px;color:#166534;margin-top:8px;min-height:18px;"></div>
    <div class="modal-btns" style="margin-top:14px;">
      <button id="driveUploadBtn" onclick="confirmDriveUpload()" disabled
        style="padding:8px 18px;background:#2563EB;color:#fff;border:none;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;">
        ✅ 이 폴더에 업로드
      </button>
      <button class="btn-modal-close" onclick="closeDriveFolderModal()">취소</button>
    </div>
  </div>
</div>

<div id="toast"></div>

<script>
let currentSid = null;
let currentFilename = null;
let eventSource = null;
let currentMode = 'new';          // 'new' | 'modify'
let projects = [];                // 프로젝트 목록 캐시
let selectedTcFile = null;        // 업로드된 TC 파일명

// ── 프로젝트 드롭다운 ─────────────────────────────────────────────────────────
let selectedProject = null;
let projectResumeState = null;

async function loadDashProjects() {
  try {
    const r = await fetch('/projects');
    const list = await r.json();
    projects = list;
    const dd = document.getElementById('projectDropdown');
    dd.innerHTML = '<option value="">-- 프로젝트를 선택하세요 --</option>';
    // 샘플(is_sample)은 하단에 배치
    const normal = list.filter(p => !p.is_sample);
    const samples = list.filter(p => p.is_sample);
    const sorted = [...normal, ...samples];
    sorted.forEach(p => {
      const opt = document.createElement('option');
      opt.value = p.name;
      let label = p.name + '  [진행 중]';
      if (p.is_sample) label += '  (샘플)';
      opt.textContent = label;
      dd.appendChild(opt);
    });
    if (selectedProject) dd.value = selectedProject;
  } catch(e) {
    console.error('프로젝트 로드 실패', e);
  }
}

async function onProjectDropdownChange() {
  const name = document.getElementById('projectDropdown').value;
  projectResumeState = null;
  // TC Update history 도 새 프로젝트 기준으로 다시 로드
  _updateRecentLoaded = '';
  if (typeof loadUpdateRecent === 'function') loadUpdateRecent();

  // ── UI 전체 초기화 ──
  // SSE 연결 종료
  if (eventSource) { eventSource.close(); eventSource = null; if (typeof updateRestartButtonState === 'function') updateRestartButtonState(); }
  currentSid = null;
  currentFilename = null;
  stopCountdown();
  // 카드 숨김 (card2~card5) + Step 1 입력 카드 복원
  ['card2','card3','card5'].forEach(function(id) {
    var card = document.getElementById(id);
    if (card) card.classList.add('hidden');
  });
  document.getElementById('card1').classList.remove('hidden');
  // 중단/오류 배너 제거
  document.querySelectorAll('.stopped-banner, .error-banner').forEach(function(el) { el.remove(); });
  // 버튼 상태 복원
  document.getElementById('startBtn').disabled = false;
  var startModifyBtn = document.getElementById('startModifyBtn');
  if (startModifyBtn) startModifyBtn.disabled = false;
  setStepBar(1);
  // 상단 card2 내부 요소 복원 (TC 단계에서 숨겼던 것들)
  var logBox = document.getElementById('logBox');
  if (logBox) { logBox.style.display = ''; logBox.innerHTML = ''; }
  var stopBtn2 = document.getElementById('stopBtn2');
  if (stopBtn2) stopBtn2.style.display = '';
  var progressWrap = document.getElementById('progressBar');
  if (progressWrap && progressWrap.parentElement) progressWrap.parentElement.style.display = '';
  var substeps = document.querySelector('#card2 .substeps');
  if (substeps) substeps.style.display = '';
  document.getElementById('stageLabel').textContent = '시작 중...';
  // 소스/포커스 초기화
  sources = [];
  sourceCounter = 0;
  renderSources();
  document.getElementById('focusArea').value = '';
  // SuiteCode 섹션 리셋
  var suiteSection = document.getElementById('suiteCodeSection');
  if (suiteSection) { suiteSection.innerHTML = ''; suiteSection.style.display = 'none'; }
  // 이어서 작업 바 / 삭제 버튼
  document.getElementById('resumeBar').style.display = 'none';
  document.getElementById('btnDeleteProject').style.display = name ? 'inline-block' : 'none';

  if (!name) { selectedProject = null; document.getElementById('projectName').value = ''; return; }
  selectedProject = name;
  document.getElementById('projectName').value = name;
  // 수정 모드 드롭다운도 연동
  var sel = document.getElementById('projectSelect');
  if (sel) {
    for (var i = 0; i < sel.options.length; i++) {
      if (sel.options[i].value === name) { sel.selectedIndex = i; break; }
    }
  }

  // 프로젝트 변경 시 이전 세션의 UI 상태 완전 초기화
  // (중요: 이전 focus_area가 input에 남아 새 파이프라인에 전달되는 버그 방지
  //  + startBtn 등 버튼 상태가 이전 세션에 묶여 비활성으로 남는 문제 방지)
  document.getElementById('focusArea').value = '';
  document.getElementById('previousFocusHint').style.display = 'none';
  window._previousFocusArea = '';

  // SSE 연결 끊기 + 세션 참조 리셋
  if (eventSource) { eventSource.close(); eventSource = null; if (typeof updateRestartButtonState === 'function') updateRestartButtonState(); }
  currentSid = null;
  currentFilename = null;
  stopCountdown();

  // 진행/결과 카드 숨김 + Step 1 입력 카드 복원
  ['card2','card3','card5'].forEach(function(id) {
    var card = document.getElementById(id);
    if (card) card.classList.add('hidden');
  });
  document.getElementById('card1').classList.remove('hidden');

  // 중단/오류 배너 제거
  document.querySelectorAll('.stopped-banner, .error-banner').forEach(function(el) { el.remove(); });

  // 모든 버튼 재활성화 (이전 세션에서 disable된 상태 승계 방지)
  document.getElementById('startBtn').disabled = false;
  var startModifyBtn = document.getElementById('startModifyBtn');
  if (startModifyBtn) startModifyBtn.disabled = false;
  setStopButtonsDisabled(false);

  // TC 모드 토글 상태 리셋 (새 프로젝트는 다시 판별)
  window._autoScreenCode = undefined;

  // Step bar를 1단계로
  setStepBar(1);

  // ── 해당 프로젝트 데이터 로드 ──
  // 이전 작업 상태 확인
  var prevFocus = '';
  try {
    var r = await fetch('/projects/' + encodeURIComponent(name) + '/state');
    var d = await r.json();
    if (d.ok && d.has_state) {
      projectResumeState = d;
      var stageNames = {parsed:'문서 파싱', policy:'정책 분석', features:'기능 목록', classifying:'분류표 생성', gate_waiting:'분류표 검토 대기', tc_writing:'TC 작성'};
      document.getElementById('resumeStage').textContent = stageNames[d.stage] || d.stage;
      document.getElementById('resumeSavedAt').textContent = d.saved_at || '';
      document.getElementById('resumeBar').style.display = 'flex';
      // focus_area는 자동 채우지 않고 힌트로만 제공 — 의도치 않은 범위 제한 방지
      if (d.focus_area) prevFocus = d.focus_area;
    }
  } catch(e) {}
  // 이전 소스 복원
  try {
    var r2 = await fetch('/projects/' + encodeURIComponent(name) + '/sources');
    var d2 = await r2.json();
    if (d2.ok && d2.has_sources && d2.sources.length > 0) {
      sources = [];
      sourceCounter = 0;
      // spec_folder / spec_folder_prev 타입은 일반 입력 소스가 아니라 구조화 모드 박스로 분리
      var specFolderEntry = null;
      var specFolderPrevEntry = null;
      d2.sources.forEach(function(s) {
        if (s.type === 'spec_folder') { specFolderEntry = s; return; }
        if (s.type === 'spec_folder_prev') { specFolderPrevEntry = s; return; }
        var id = ++sourceCounter;
        sources.push({ id: id, type: s.type, content: s.content || '', selected_files: s.selected_files || null });
      });
      renderSources();

      // 구조화 spec 폴더 자동 복원 — 경로 채움 (spec 섹션은 기본 펼침 상태라 토글 불필요)
      if (specFolderEntry) {
        var pathEl = document.getElementById('specFolderPath');
        if (pathEl) pathEl.value = specFolderEntry.content || '';
        // spec 섹션이 접혀있으면 자동으로 펼침 (기본은 펼침이지만 사용자가 닫았을 수 있음)
        var specBody = document.getElementById('specFolderBody');
        if (specBody && specBody.style.display === 'none') {
          toggleAccordion('spec');
        }
        // 이전 폴더가 있으면 diff 박스도 자동 복원
        if (specFolderPrevEntry) {
          var diffBox = document.getElementById('diffBox');
          var prevEl = document.getElementById('prevSpecFolderPath');
          if (prevEl) prevEl.value = specFolderPrevEntry.content || '';
          if (diffBox && diffBox.style.display === 'none' && typeof toggleDiffMode === 'function') {
            toggleDiffMode();
          }
        }
      }

      if (d2.focus_area) prevFocus = d2.focus_area;  // state의 값과 sources의 값 중 존재하는 것

      // 생성 모드 복원 — Combo 였으면 Combo 라디오 선택 + 카드 자동 펼침
      if (d2.generation_mode) {
        const modeRadio = document.querySelector('input[name="genMode"][value="' + d2.generation_mode + '"]');
        if (modeRadio) {
          modeRadio.checked = true;
          if (typeof onGenModeChanged === 'function') onGenModeChanged();
          // Combo 옵션도 복원 (잠시 후 파일 목록 로드된 다음 적용)
          if (d2.generation_mode === 'combo_only' && d2.combo_opts) {
            setTimeout(function() { restoreComboOpts(d2.combo_opts); }, 800);
          }
        }
      }

      var loadedCount = sources.length + (specFolderEntry ? 1 : 0) + (specFolderPrevEntry ? 1 : 0);
      var msg = '이전 소스 ' + loadedCount + '개를 불러왔습니다.';
      if (specFolderEntry) msg += ' (📁 구조화 spec 폴더 모드)';
      if (d2.generation_mode === 'combo_only') msg += ' · 🧩 Combo 모드 복원';
      showToast(msg);
    }
  } catch(e) {}

  // 이전 범위가 있으면 힌트 배너로 표시 (자동 채움 X)
  if (prevFocus) {
    window._previousFocusArea = prevFocus;
    document.getElementById('previousFocusValue').textContent = prevFocus;
    document.getElementById('previousFocusHint').style.display = 'block';
  }
}

// 이전 범위 재사용 (사용자 명시적 클릭 시에만)
function reusePreviousFocus() {
  if (!window._previousFocusArea) return;
  document.getElementById('focusArea').value = window._previousFocusArea;
  document.getElementById('previousFocusHint').style.display = 'none';
  showToast('이전 범위를 불러왔습니다: ' + window._previousFocusArea.substring(0, 40) + (window._previousFocusArea.length > 40 ? '...' : ''));
}

// 힌트 무시
function dismissPreviousFocusHint() {
  document.getElementById('previousFocusHint').style.display = 'none';
  window._previousFocusArea = '';
}

function toggleNewProjectInline() {
  const row = document.getElementById('newProjectInline');
  const visible = row.style.display === 'flex';
  row.style.display = visible ? 'none' : 'flex';
  if (!visible) document.getElementById('newDashProjectName').focus();
}

async function retryPipeline() {
  // 오류 배너 제거
  document.querySelectorAll('.error-banner').forEach(el => el.remove());
  var name = selectedProject || document.getElementById('projectName').value.trim();
  if (!name) { alert('프로젝트를 선택해주세요.'); return; }
  document.getElementById('startBtn').disabled = true;
  document.getElementById('stageLabel').textContent = '이어서 재시작 중...';
  try {
    var r = await fetch('/start', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ project_name: name, resume: true })
    });
    var d = await r.json();
    if (!d.ok) { alert(d.error); document.getElementById('startBtn').disabled = false; return; }
    currentSid = d.sid;
    showToast('이전 단계에서 이어서 재시작합니다.');
    connectStream(d.sid);
  } catch(e) {
    alert('오류: ' + e.message);
    document.getElementById('startBtn').disabled = false;
  }
}

function isPipelineRunning() {
  // 파이프라인이 진행 중인지 (SSE 연결 살아있고 + currentSid 있음)
  return !!(eventSource && currentSid);
}

function restartFromScratch() {
  // 진행 중 가드 — 사용자에게 명확히 알림 + 강제 종료 방지
  if (isPipelineRunning()) {
    alert('⚠️ 파이프라인이 진행 중입니다.\\n\\n'
      + '먼저 진행 중인 작업을 완료하거나 [■ 파이프라인 중단] 버튼으로 중단하세요.\\n\\n'
      + '입력 화면으로 돌아가면 화면은 바뀌지만 백엔드 작업은 계속 진행되어 결과 충돌이 발생할 수 있습니다.');
    return;
  }
  document.querySelectorAll('.error-banner, .stopped-banner').forEach(el => el.remove());
  document.getElementById('card1').classList.remove('hidden');
  document.getElementById('card2').classList.add('hidden');
  document.getElementById('card3').classList.add('hidden');
  document.getElementById('startBtn').disabled = false;
  setStepBar(1);
  // gateNavBox 평시 상태로 reset (다음에 Step 3 진입 시 깨끗하게)
  resetGateNavBox();
  showToast('소스를 확인하고 파이프라인을 다시 시작하세요.');
}

function resetGateNavBox() {
  const box = document.getElementById('gateNavBox');
  const msg = document.getElementById('gateNavMsg');
  const btns = document.getElementById('gateNavButtons');
  if (!box || !msg || !btns) return;
  box.style.background = '#FEF3C7';
  box.style.borderColor = '#FBBF24';
  box.style.borderStyle = 'dashed';
  msg.style.color = '#78350F';
  msg.innerHTML = '💡 <strong>분류표를 다시 만들고 싶나요?</strong><br>'
    + '<span style="font-size:11px;">예: TC 생성 범위 추가/수정 — 입력값은 그대로 보존됩니다.</span>';
  btns.innerHTML = ''
    + '<button type="button" onclick="restartFromScratch()" '
    +   'style="padding:8px 16px; font-size:13px; background:#F59E0B; color:#FFF; border:none; border-radius:6px; cursor:pointer; font-weight:600;">'
    +   '← 입력 화면으로 (값 유지)'
    + '</button>';
}

// 진행 중 상태에 따라 '입력 화면으로' 버튼 시각 갱신 — 회색 + 안내 툴팁
function updateRestartButtonState() {
  const buttons = document.querySelectorAll('button[onclick="restartFromScratch()"]');
  const running = isPipelineRunning();
  buttons.forEach(function(btn) {
    if (running) {
      btn.style.opacity = '0.4';
      btn.style.cursor = 'not-allowed';
      btn.title = '⚠️ 파이프라인 진행 중 — 중단 후 사용 가능';
    } else {
      btn.style.opacity = '';
      btn.style.cursor = '';
      btn.title = '입력 필드 값은 그대로 유지 + Step 1 화면으로 복귀';
    }
  });
}

function startNextIteration(focusOnly) {
  // 완료/진행 카드 숨김 + Step 1 입력 카드 복원
  ['card2','card3','card5'].forEach(function(id) {
    document.getElementById(id).classList.add('hidden');
  });
  document.getElementById('card1').classList.remove('hidden');
  document.querySelectorAll('.stopped-banner, .error-banner').forEach(function(el) { el.remove(); });
  // Drive 버튼 라벨 reset — 이전 업로드 완료 상태 잔존 방지
  if (typeof resetDriveBtn === 'function') resetDriveBtn();
  // 상단 card2 요소 복원
  var logBox = document.getElementById('logBox');
  if (logBox) { logBox.style.display = ''; logBox.innerHTML = ''; }
  var stopBtn2 = document.getElementById('stopBtn2');
  if (stopBtn2) stopBtn2.style.display = '';
  var progressWrap = document.getElementById('progressBar');
  if (progressWrap && progressWrap.parentElement) progressWrap.parentElement.style.display = '';
  var substeps = document.querySelector('#card2 .substeps');
  if (substeps) substeps.style.display = '';
  // SuiteCode 리셋
  var suiteSection = document.getElementById('suiteCodeSection');
  if (suiteSection) { suiteSection.innerHTML = ''; suiteSection.style.display = 'none'; }
  // 버튼 활성화
  document.getElementById('startBtn').disabled = false;
  setStepBar(1);
  stopCountdown();
  if (eventSource) { eventSource.close(); eventSource = null; if (typeof updateRestartButtonState === 'function') updateRestartButtonState(); }
  // 범위만 변경 모드: 포커스 입력란으로 스크롤
  if (focusOnly) {
    document.getElementById('focusArea').focus();
    document.getElementById('focusArea').scrollIntoView({ behavior: 'smooth', block: 'center' });
    showToast('TC 생성 범위를 수정하고 파이프라인을 시작하세요.');
  } else {
    document.getElementById('card1').scrollIntoView({ behavior: 'smooth', block: 'start' });
    showToast('소스와 범위를 수정하고 파이프라인을 시작하세요.');
  }
}

async function resumePipeline() {
  if (!selectedProject) return;
  document.getElementById('startBtn').disabled = true;
  document.getElementById('card1').classList.add('hidden');
  document.getElementById('card2').classList.remove('hidden');
  setStepBar(2);
  document.getElementById('card2').scrollIntoView({ behavior: 'smooth', block: 'start' });
  try {
    const r = await fetch('/start', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ project_name: selectedProject, resume: true })
    });
    const d = await r.json();
    if (!d.ok) { alert(d.error); document.getElementById('startBtn').disabled = false; return; }
    currentSid = d.sid;
    showToast('이전 작업에서 이어서 진행합니다.');
    connectStream(d.sid);
  } catch(e) {
    alert('오류: ' + e.message);
    document.getElementById('startBtn').disabled = false;
  }
}

function discardResume() {
  document.getElementById('resumeBar').style.display = 'none';
  projectResumeState = null;
  showToast('처음부터 새로 시작합니다.');
}

async function deleteDashProject() {
  if (!selectedProject) return;
  if (!confirm(selectedProject + ' 프로젝트를 목록에서 숨길까요?')) return;
  try {
    const r = await fetch('/projects/' + encodeURIComponent(selectedProject), { method: 'DELETE' });
    const d = await r.json();
    if (!d.ok) { alert(d.error); return; }
    selectedProject = null;
    document.getElementById('projectName').value = '';
    document.getElementById('resumeBar').style.display = 'none';
    document.getElementById('btnDeleteProject').style.display = 'none';
    await loadDashProjects();
    showToast('프로젝트가 목록에서 제거되었습니다.');
  } catch(e) {
    alert('오류: ' + e.message);
  }
}

async function createDashProject() {
  const input = document.getElementById('newDashProjectName');
  const name = input.value.trim();
  if (!name) { input.focus(); return; }
  try {
    const r = await fetch('/projects', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ name })
    });
    const d = await r.json();
    if (!d.ok) { alert(d.error); return; }
    input.value = '';
    document.getElementById('newProjectInline').style.display = 'none';
    await loadDashProjects();
    document.getElementById('projectDropdown').value = name;
    onProjectDropdownChange();
    showToast('프로젝트 생성: ' + name);
  } catch(e) {
    alert('오류: ' + e.message);
  }
}

// 페이지 로드 시 프로젝트 목록 불러오기
setTimeout(loadDashProjects, 100);

// 페이지 로드 시 modify 패널이 이미 보이는 상태면 history 도 같이 로드
// (프로젝트 목록 로딩 끝나기까지 약간 여유를 줘서 500ms 후)
setTimeout(function() {
  const panelMod = document.getElementById('panelModify');
  if (panelMod && !panelMod.classList.contains('hidden')) {
    _updateRecentLoaded = '';
    loadUpdateRecent();
  }
}, 500);

// ── 모드 전환 ──────────────────────────────────────────────────────────────────
function switchMode(mode) {
  currentMode = mode;
  document.getElementById('modeNew').classList.toggle('active', mode === 'new');
  document.getElementById('modeModify').classList.toggle('active', mode === 'modify');
  document.getElementById('panelNew').classList.toggle('hidden', mode !== 'new');
  document.getElementById('panelModify').classList.toggle('hidden', mode !== 'modify');
  if (mode === 'modify') {
    loadProjects();
    initTcDropzone();
    _updateRecentLoaded = '';  // 모드 진입 시 캐시 무시하고 강제 재로드
    loadUpdateConfig();  // 설정 라벨 갱신 (폴더 이름 표시)
    loadUpdateRecent();
  }
}

// ── TC Update 직전 작업 불러오기 ──────────────────────────────────────────
let _updateRecentLoaded = '';

// ── TC Update 설정 (Drive 폴더 ID 등) ────────────────────────────────────
async function loadUpdateConfig() {
  try {
    const r = await fetch('/update/config');
    const d = await r.json();
    const labelEl = document.getElementById('configFolderName');
    if (!labelEl) return;
    if (d.drive_folder_id && d.drive_folder_name) {
      labelEl.textContent = '· ' + d.drive_folder_name;
      labelEl.style.color = '#166534';
    } else if (d.drive_folder_id) {
      labelEl.textContent = '· (폴더 지정됨)';
      labelEl.style.color = '#6B7280';
    } else {
      labelEl.textContent = '· 폴더 미지정';
      labelEl.style.color = '#DC2626';
    }
  } catch (e) {
    // 무시
  }
}

async function openConfigModal() {
  document.getElementById('configModal').classList.remove('hidden');
  document.getElementById('configMsg').textContent = '';
  const infoEl = document.getElementById('configCurrentInfo');
  const urlInput = document.getElementById('configFolderUrl');
  infoEl.innerHTML = '⏳ 현재 설정 확인 중...';
  try {
    const r = await fetch('/update/config');
    const d = await r.json();
    if (d.drive_folder_id) {
      urlInput.value = d.drive_folder_url || '';
      infoEl.innerHTML = '✅ 현재 폴더: <strong>' + escapeHtml(d.drive_folder_name || '(이름 조회 불가)') + '</strong><br>'
        + '<span style="font-family:monospace; font-size:11px;">ID: ' + escapeHtml(d.drive_folder_id) + '</span>';
    } else {
      urlInput.value = '';
      infoEl.innerHTML = '⚠️ 폴더 미지정 — 사본이 Drive 루트에 만들어집니다 (orphan).';
      infoEl.style.background = '#FEF3C7';
      infoEl.style.borderColor = '#FBBF24';
      infoEl.style.color = '#78350F';
    }
  } catch (e) {
    infoEl.innerHTML = '<span style="color:#DC2626;">설정 조회 실패: ' + escapeHtml(e.message) + '</span>';
  }
}

function closeConfigModal() {
  document.getElementById('configModal').classList.add('hidden');
}

async function saveConfig() {
  const url = document.getElementById('configFolderUrl').value.trim();
  const msgEl = document.getElementById('configMsg');
  const btn = document.getElementById('configSaveBtn');
  btn.disabled = true;
  btn.textContent = '⏳ 저장 중...';
  msgEl.innerHTML = '';
  try {
    const r = await fetch('/update/config', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ drive_folder_url: url }),
    });
    const d = await r.json();
    if (!d.ok) {
      msgEl.innerHTML = '<span style="color:#DC2626;">❌ ' + escapeHtml(d.error || '오류') + '</span>';
      return;
    }
    if (d.cleared) {
      msgEl.innerHTML = '<span style="color:#166534;">✅ 폴더 지정 제거됨 — 다음 사본부터 Drive 루트에 생성됩니다.</span>';
    } else {
      msgEl.innerHTML = '<span style="color:#166534;">✅ 저장 완료 — 폴더: <strong>' + escapeHtml(d.drive_folder_name || '(이름 없음)') + '</strong></span>';
    }
    // 버튼 옆 라벨 갱신
    await loadUpdateConfig();
    // 잠깐 후 자동 닫기
    setTimeout(closeConfigModal, 1200);
  } catch (e) {
    msgEl.innerHTML = '<span style="color:#DC2626;">❌ ' + escapeHtml(e.message) + '</span>';
  } finally {
    btn.disabled = false;
    btn.textContent = '💾 저장';
  }
}

async function clearConfig() {
  if (!confirm('Drive 폴더 지정을 제거할까요?\\n\\n이후 만들어지는 사본은 Drive 루트 (orphan) 에 생성됩니다.')) {
    return;
  }
  document.getElementById('configFolderUrl').value = '';
  await saveConfig();
}

async function loadUpdateRecent() {
  // 활성 select 찾기 — projectDropdown (현재) 또는 projectSelect (레거시)
  const projSelect = document.getElementById('projectDropdown')
                       || document.getElementById('projectSelect');
  const project = projSelect ? (projSelect.value || '').trim() : '';
  const area = document.getElementById('updateRecentArea');
  const countEl = document.getElementById('updateRecentCount');
  if (!area) return;

  if (!project) {
    area.style.display = 'none';
    _updateRecentLoaded = '';
    window._updateRecentCache = [];
    return;
  }
  if (_updateRecentLoaded === project) return;  // 같은 프로젝트면 재로드 skip
  _updateRecentLoaded = project;

  try {
    const r = await fetch('/update/history?project=' + encodeURIComponent(project));
    const d = await r.json();
    const hist = (d && d.history) || [];
    window._updateRecentCache = hist;
    if (!hist.length) {
      area.style.display = 'none';
      return;
    }
    area.style.display = 'inline-block';
    if (countEl) countEl.textContent = String(hist.length);
  } catch (e) {
    area.style.display = 'none';
  }
}

function toggleUpdateRecent() {
  const list = document.getElementById('updateRecentList');
  const chev = document.getElementById('updateRecentChevron');
  if (!list) return;
  const open = list.style.display !== 'none';
  if (open) {
    list.style.display = 'none';
    if (chev) chev.textContent = '▾';
    return;
  }
  // 펼치기 — 그 시점에 항상 신선한 데이터로 렌더
  const hist = window._updateRecentCache || [];
  if (!hist.length) {
    list.innerHTML = '<div style="padding:14px; font-size:12px; color:#6B7280; text-align:center;">저장된 작업 이력이 없습니다.</div>';
  } else {
    list.innerHTML = hist.map(function(h, i) {
      const prevShort = escapeHtml((h.prev_folder || '').split('/').slice(-2).join('/'));
      const newShort = escapeHtml((h.new_folder || '').split('/').slice(-2).join('/'));
      const sheetShort = (h.existing_tc_url || '').replace(/^https?:\/\/docs\.google\.com\/spreadsheets\/d\//, '').slice(0, 24) + '…';
      return ''
        + '<div class="recent-row" style="padding:10px 14px; border-bottom:1px solid #F3F4F6; display:flex; align-items:center; gap:10px;">'
        +   '<div style="flex:1; min-width:0; cursor:pointer;" onclick="applyUpdateRecent(' + i + ')">'
        +     '<div style="font-size:13px; font-weight:600; color:#111827;">' + escapeHtml(h.label || '?') + '</div>'
        +     '<div style="font-size:10px; color:#6B7280; margin-top:3px; font-family:monospace; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;">'
        +       '📁 …/' + prevShort + ' → …/' + newShort
        +     '</div>'
        +     '<div style="font-size:10px; color:#9CA3AF; margin-top:2px;">📑 ' + escapeHtml(sheetShort) + ' · ' + escapeHtml(h.saved_at || '') + '</div>'
        +   '</div>'
        +   '<button type="button" onclick="applyUpdateRecent(' + i + ')" '
        +     'style="flex:0 0 auto; font-size:11px; padding:5px 10px; background:#EEF2FF; color:#3730A3; border:1px solid #C7D2FE; border-radius:4px; cursor:pointer; font-weight:600;">'
        +     '↻ 불러오기</button>'
        +   '<button type="button" onclick="deleteUpdateRecent(event, ' + i + ')" '
        +     'title="이 이력 삭제" '
        +     'style="flex:0 0 auto; font-size:13px; padding:5px 9px; background:#FFF; color:#991B1B; border:1px solid #FCA5A5; border-radius:4px; cursor:pointer; line-height:1;">'
        +     '✕</button>'
        + '</div>';
    }).join('');
  }
  list.style.display = 'block';
  if (chev) chev.textContent = '▴';
}

function applyUpdateRecent(idx) {
  const hist = window._updateRecentCache || [];
  const h = hist[idx];
  if (!h) return;
  document.getElementById('updatePrevFolder').value = h.prev_folder || '';
  document.getElementById('updateNewFolder').value = h.new_folder || '';
  document.getElementById('updateTcUrl').value = h.existing_tc_url || '';
  // 드롭다운 닫기
  document.getElementById('updateRecentList').style.display = 'none';
  const chev = document.getElementById('updateRecentChevron');
  if (chev) chev.textContent = '▾';
  if (typeof showToast === 'function') {
    showToast('✅ 불러옴: ' + (h.label || '?'));
  }
}

async function deleteUpdateRecent(ev, idx) {
  if (ev && ev.stopPropagation) ev.stopPropagation();
  const hist = window._updateRecentCache || [];
  const h = hist[idx];
  if (!h) return;
  if (!confirm('이 이력을 삭제할까요?\\n\\n' + (h.label || '?') + ' (' + (h.saved_at || '') + ')\\n\\n원본 Sheets 와 spec 폴더는 영향 없음.')) {
    return;
  }
  const projSelect = document.getElementById('projectDropdown')
                       || document.getElementById('projectSelect');
  const project = projSelect ? (projSelect.value || '').trim() : '';
  if (!project) {
    alert('프로젝트를 먼저 선택하세요.');
    return;
  }
  try {
    const r = await fetch('/update/history/delete', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ project: project, index: idx }),
    });
    const d = await r.json();
    if (!d.ok) {
      alert('삭제 실패: ' + (d.error || '오류'));
      return;
    }
    // 캐시 갱신 + 드롭다운 다시 그림
    window._updateRecentCache = d.history || [];
    _updateRecentLoaded = '';  // 다음 toggle 에서 강제 재로드
    const countEl = document.getElementById('updateRecentCount');
    if (countEl) countEl.textContent = String((d.history || []).length);
    // 0개면 드롭다운 + 버튼 자체 숨김
    if (!(d.history || []).length) {
      document.getElementById('updateRecentList').style.display = 'none';
      const area = document.getElementById('updateRecentArea');
      if (area) area.style.display = 'none';
    } else {
      // 다시 렌더 — toggleUpdateRecent 를 강제로 다시 호출 (열린 상태 유지)
      const list = document.getElementById('updateRecentList');
      list.style.display = 'none';
      toggleUpdateRecent();
    }
    if (typeof showToast === 'function') {
      showToast('🗑 삭제됨: ' + (h.label || '?'));
    }
  } catch (e) {
    alert('네트워크 오류: ' + e.message);
  }
}

// 외부 클릭 시 드롭다운 닫기
document.addEventListener('click', function(e) {
  const area = document.getElementById('updateRecentArea');
  const list = document.getElementById('updateRecentList');
  if (!area || !list || list.style.display === 'none') return;
  if (!area.contains(e.target)) {
    list.style.display = 'none';
    const chev = document.getElementById('updateRecentChevron');
    if (chev) chev.textContent = '▾';
  }
});

// 샘플 기획서 불러오기 — 직접 입력 모드로 전환 후 textarea 채우기
// 간단한 토스트 메시지
function showToast(msg) {
  let t = document.getElementById('toast');
  if (!t) {
    t = document.createElement('div');
    t.id = 'toast';
    t.style.cssText = 'position:fixed;bottom:28px;left:50%;transform:translateX(-50%);' +
      'background:#2D3748;color:#fff;padding:10px 22px;border-radius:24px;' +
      'font-size:13px;font-weight:600;z-index:9999;opacity:0;transition:opacity 0.3s;pointer-events:none;';
    document.body.appendChild(t);
  }
  t.textContent = msg;
  t.style.opacity = '1';
  setTimeout(() => { t.style.opacity = '0'; }, 3000);
}

// 프로젝트 목록 불러오기
async function loadProjects() {
  try {
    const r = await fetch('/projects');
    const raw = await r.json();
    // 빈 이름 유령 레코드 제거 (최종 프론트엔드 방어선)
    projects = Array.isArray(raw) ? raw.filter(p => p && (p.name || '').trim()) : [];
    const sel = document.getElementById('projectSelect');
    sel.innerHTML = '<option value="">— 프로젝트를 선택하세요 —</option>';
    projects.forEach(p => {
      const label = p.name + ' (' + (p.updated_at || '신규') + ')';
      const opt = document.createElement('option');
      opt.value = p.name; opt.textContent = label;
      sel.appendChild(opt);
    });
    if (projects.length === 0) {
      const opt = document.createElement('option');
      opt.value = '';
      opt.textContent = '저장된 프로젝트 없음 — 새 프로젝트를 만드세요';
      opt.disabled = true;
      sel.appendChild(opt);
    }
  } catch(e) {
    console.error('프로젝트 목록 로드 실패', e);
  }
}

function onProjectSelect() {
  const name = document.getElementById('projectSelect').value;
  const card = document.getElementById('projectCard');
  const deleteBtn = document.getElementById('btnDeleteProject');
  // TC 수정 모드면 history 다시 로드 (선택 변경 시)
  if (currentMode === 'modify') {
    _updateRecentLoaded = '';  // 강제 재로드
    loadUpdateRecent();
  }
  if (!name) { card.classList.remove('visible'); deleteBtn.style.display = 'none'; return; }
  const proj = projects.find(p => p.name === name);
  if (!proj) return;
  document.getElementById('projectCardName').textContent = proj.name;
  document.getElementById('projectCardMeta').textContent =
    '최근 업데이트: ' + (proj.updated_at || '—') + (proj.excel_file ? '  |  Excel: ' + proj.excel_file.split('/').pop() : '');
  card.classList.add('visible');
  deleteBtn.style.display = 'inline-block';
  selectedTcFile = null;
  document.getElementById('tcUploadStatus').classList.add('hidden');
  document.getElementById('uploadProjectNameGroup').style.display = 'none';
}

function toggleNewProjectForm() {
  const form = document.getElementById('newProjectForm');
  form.classList.toggle('hidden');
  if (!form.classList.contains('hidden')) document.getElementById('newProjectName').focus();
}

async function createProject() {
  const name = document.getElementById('newProjectName').value.trim();
  if (!name) { alert('프로젝트명을 입력해주세요.'); return; }
  try {
    const r = await fetch('/projects', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ name })
    });
    const d = await r.json();
    if (!d.ok) { alert('❌ ' + d.error); return; }
    document.getElementById('newProjectName').value = '';
    document.getElementById('newProjectForm').classList.add('hidden');
    await loadProjects();
    // 방금 만든 프로젝트 자동 선택
    document.getElementById('projectSelect').value = name;
    onProjectSelect();
    showToast('✅ 프로젝트 생성 완료: ' + name);
  } catch(e) {
    alert('오류: ' + e.message);
  }
}

async function deleteProject() {
  const name = document.getElementById('projectSelect').value;
  if (!name) return;
  if (!confirm('"' + name + '" 프로젝트를 삭제할까요?\\n(TC 파일은 삭제되지 않습니다)')) return;
  try {
    const r = await fetch('/projects/' + encodeURIComponent(name), { method: 'DELETE' });
    const d = await r.json();
    if (!d.ok) { alert('❌ ' + d.error); return; }
    await loadProjects();
    document.getElementById('projectCard').classList.remove('visible');
    document.getElementById('btnDeleteProject').style.display = 'none';
    showToast('🗑 프로젝트 삭제 완료');
  } catch(e) {
    alert('오류: ' + e.message);
  }
}

// TC 가져오기 탭 전환
function switchImportTab(tab) {
  document.getElementById('tabExcel').classList.toggle('active', tab === 'excel');
  document.getElementById('tabSheets').classList.toggle('active', tab === 'sheets');
  document.getElementById('panelExcel').classList.toggle('hidden', tab !== 'excel');
  document.getElementById('panelSheets').classList.toggle('hidden', tab !== 'sheets');
}

// Excel / md 파일 업로드 — 이벤트는 switchMode('modify') 이후 패널 표시 시점에 등록
function initTcDropzone() {
  const tcDropzone = document.getElementById('tcDropzone');
  const tcFileInput = document.getElementById('tcFileInput');
  if (!tcDropzone || tcDropzone._initialized) return;
  tcDropzone._initialized = true;
  tcDropzone.addEventListener('dragover', e => { e.preventDefault(); tcDropzone.style.borderColor = 'var(--teal)'; });
  tcDropzone.addEventListener('dragleave', () => { tcDropzone.style.borderColor = ''; });
  tcDropzone.addEventListener('drop', e => {
    e.preventDefault(); tcDropzone.style.borderColor = '';
    if (e.dataTransfer.files.length > 0) uploadTcFile(e.dataTransfer.files[0]);
  });
  if (tcFileInput) {
    tcFileInput.addEventListener('change', () => {
      if (tcFileInput.files.length > 0) uploadTcFile(tcFileInput.files[0]);
    });
  }
}

async function uploadTcFile(file) {
  const ext = file.name.split('.').pop().toLowerCase();
  if (!['md', 'xlsx', 'xls'].includes(ext)) {
    alert('.xlsx / .xls / .md 파일만 업로드 가능합니다.'); return;
  }
  document.getElementById('uploadProjectNameGroup').style.display = 'block';
  const stem = file.name.replace(/\.[^.]+$/, '');
  const projName = document.getElementById('uploadProjectName').value.trim() ||
                   document.getElementById('projectSelect').value || stem;

  tcDropzone.textContent = '⏳ 변환 중...';
  const fd = new FormData();
  fd.append('file', file);
  fd.append('project_name', projName);
  try {
    const r = await fetch('/upload-tc', { method: 'POST', body: fd });
    const d = await r.json();
    if (d.ok) {
      selectedTcFile = d.tc_file;
      tcDropzone.textContent = '✅ ' + file.name + ' 업로드 완료';
      const status = document.getElementById('tcUploadStatus');
      status.textContent = '✅ ' + file.name + ' 등록됨 (프로젝트: ' + d.project_name + ')';
      status.classList.remove('hidden');
      await loadProjects();
      document.getElementById('projectSelect').value = d.project_name;
      onProjectSelect();
      showToast('✅ TC 파일 등록 완료');
    } else {
      tcDropzone.textContent = '❌ ' + d.error;
    }
  } catch(e) {
    tcDropzone.textContent = '❌ 업로드 실패: ' + e.message;
  }
}

// Google Sheets 가져오기
async function importSheets() {
  const url = document.getElementById('sheetsUrl').value.trim();
  const projName = document.getElementById('uploadProjectName').value.trim() ||
                   document.getElementById('projectSelect').value;
  if (!url) { alert('Sheets URL을 입력해주세요.'); return; }
  if (!projName) {
    document.getElementById('uploadProjectNameGroup').style.display = 'block';
    document.getElementById('uploadProjectName').focus();
    alert('프로젝트명을 입력해주세요.'); return;
  }
  document.getElementById('uploadProjectNameGroup').style.display = 'block';
  const status = document.getElementById('tcUploadStatus');
  status.textContent = '⏳ Sheets 가져오는 중...';
  status.classList.remove('hidden');
  try {
    const r = await fetch('/import-sheets', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ url, project_name: projName })
    });
    const d = await r.json();
    if (d.ok) {
      selectedTcFile = d.tc_file;
      status.textContent = '✅ Sheets 가져오기 완료 (TC ' + (d.tc_count || '?') + '개, 프로젝트: ' + d.project_name + ')';
      await loadProjects();
      document.getElementById('projectSelect').value = d.project_name;
      onProjectSelect();
      showToast('✅ Google Sheets 연동 완료');
    } else {
      status.textContent = '❌ ' + d.error;
    }
  } catch(e) {
    status.textContent = '❌ 오류: ' + e.message;
  }
}

// ── 기획서 diff 기반 TC 갱신 ─────────────────────────────────
let _diffSid = null;
let _diffReport = null;
let _diffFiles = { old: null, new: null, tc: null };  // {old_spec_path, new_spec_path, existing_tc_path}

async function handleDiffFileUpload(evt, slot) {
  const file = evt.target.files[0];
  if (!file) return;
  const ext = (file.name.split('.').pop() || '').toLowerCase();
  const statusEl = document.getElementById(slot === 'old' ? 'oldSpecStatus' :
                                            slot === 'new' ? 'newSpecStatus' : 'existingTcStatus');
  const dropzone = document.getElementById(slot === 'old' ? 'oldSpecDropzone' :
                                            slot === 'new' ? 'newSpecDropzone' : 'existingTcDropzone');

  // 업로드: MD/TXT는 /upload-md, TC(xlsx/md)는 /upload-existing-tc
  const fd = new FormData();
  fd.append('file', file);
  let endpoint, savedKind;
  if (slot === 'tc') {
    endpoint = '/upload-existing-tc';
    savedKind = 'tc';
  } else {
    endpoint = '/upload-md';
    savedKind = 'spec';
  }
  statusEl.style.display = 'block';
  statusEl.style.color = '#2563EB';
  statusEl.textContent = '⏳ 업로드 중...';
  try {
    const r = await fetch(endpoint, { method: 'POST', body: fd });
    const d = await r.json();
    if (!d.ok) {
      statusEl.style.color = '#DC2626';
      statusEl.textContent = '❌ ' + (d.error || '업로드 실패');
      return;
    }
    _diffFiles[slot] = d.filename;
    statusEl.style.color = 'var(--success)';
    let msg = '✓ ' + d.filename;
    if (d.tc_count !== undefined) msg += ` (TC ${d.tc_count}개 감지)`;
    statusEl.textContent = msg;
    dropzone.style.borderColor = 'var(--success)';
    dropzone.style.background = '#F0FDF4';
  } catch(e) {
    statusEl.style.color = '#DC2626';
    statusEl.textContent = '❌ 네트워크 오류: ' + e.message;
  }
}

async function startDiffAnalyze() {
  if (!_diffFiles.old || !_diffFiles.new) {
    alert('이전 기획서와 새 기획서를 모두 업로드하세요.');
    return;
  }
  const projectName = (document.getElementById('projectDropdown') || document.getElementById('projectSelect') || {}).value || '';
  const btn = document.getElementById('startDiffBtn');
  btn.disabled = true;
  btn.textContent = '⏳ 분석 중...';
  try {
    const r = await fetch('/analyze-diff', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({
        project_name: projectName,
        old_spec_path: _diffFiles.old,
        new_spec_path: _diffFiles.new,
        existing_tc_path: _diffFiles.tc || '',
      })
    });
    const d = await r.json();
    if (!d.ok) {
      alert('분석 오류: ' + d.error);
      btn.disabled = false; btn.textContent = '🔍 변경사항 분석';
      return;
    }
    _diffSid = d.sid;
    _diffReport = d.report;
    if (d.no_changes) {
      renderDiffReport(d.report, { noChanges: true, existingCount: d.existing_tc_count });
    } else {
      renderDiffReport(d.report, { noChanges: false, existingCount: d.existing_tc_count });
    }
    btn.disabled = false; btn.textContent = '🔍 변경사항 분석';
  } catch(e) {
    alert('네트워크 오류: ' + e.message);
    btn.disabled = false; btn.textContent = '🔍 변경사항 분석';
  }
}

function _escapeHtml(s) {
  return (s || '').replace(/[<>&]/g, c => ({'<':'&lt;','>':'&gt;','&':'&amp;'}[c]));
}

function renderDiffReport(report, meta) {
  const area = document.getElementById('diffReportArea');
  const summaryEl = document.getElementById('diffReportSummary');
  const sectionsEl = document.getElementById('diffReportSections');
  area.classList.remove('hidden');

  if (meta.noChanges) {
    summaryEl.innerHTML = '🎉 <strong>변경사항이 없습니다.</strong> 두 기획서의 구조적 차이가 탐지되지 않았습니다.';
    sectionsEl.innerHTML = '';
    return;
  }

  if (!report.has_spec_codes) {
    summaryEl.innerHTML = '⚠️ <strong>화면 코드(SCR/SCREEN/PAGE)가 탐지되지 않아 구조 비교가 어렵습니다.</strong><br>기획서에 <code>### SCR-001: 이름</code> 같은 헤더를 추가하거나, 「신규 TC 생성」 탭을 이용하세요.';
    sectionsEl.innerHTML = '';
    return;
  }

  const s = report.summary;
  summaryEl.innerHTML = `🆕 신규 <strong>${s.added_n}</strong>개 · 🔄 수정 <strong>${s.modified_n}</strong>개 · 🗑️ 삭제 <strong>${s.removed_n}</strong>개 · ✅ 유지 <strong>${s.unchanged_n}</strong>개` +
    (meta.existingCount ? ` <span style="color:var(--muted);">(기존 TC ${meta.existingCount}개)</span>` : '');

  let html = '';

  // 툴바: 전체 선택/해제 + 모두 펼치기/접기
  const totalItems = report.added.length + report.modified.length + report.removed.length;
  if (totalItems > 0) {
    html += `<div style="display:flex;gap:8px;margin-top:8px;flex-wrap:wrap;">
      <button type="button" onclick="toggleSectionSelection('all', true)" style="padding:4px 10px;font-size:11px;background:#EFF6FF;border:1px solid #BFDBFE;border-radius:4px;cursor:pointer;font-weight:600;color:#1E40AF;">✓ 전체 선택 (${totalItems}개)</button>
      <button type="button" onclick="toggleSectionSelection('all', false)" style="padding:4px 10px;font-size:11px;background:#FFFFFF;border:1px solid #D1D5DB;border-radius:4px;cursor:pointer;">☐ 전체 해제</button>
      ${report.modified.length > 0 ? `
        <span style="width:1px;background:#E5E7EB;margin:0 4px;"></span>
        <button type="button" onclick="toggleAllDiffDetails(true)" style="padding:4px 10px;font-size:11px;background:#FFFFFF;border:1px solid #D1D5DB;border-radius:4px;cursor:pointer;">▼ 모두 펼치기</button>
        <button type="button" onclick="toggleAllDiffDetails(false)" style="padding:4px 10px;font-size:11px;background:#FFFFFF;border:1px solid #D1D5DB;border-radius:4px;cursor:pointer;">▲ 모두 접기</button>
      ` : ''}
    </div>`;
  }

  if (report.added.length) {
    html += `<div style="margin-top:10px;display:flex;align-items:center;justify-content:space-between;">
      <strong style="color:#047857;">🆕 신규 화면 (승인 시 TC 생성) <span style="font-size:11px;color:var(--muted);font-weight:normal;">${report.added.length}개</span></strong>
      <div style="display:flex;gap:4px;">
        <button type="button" onclick="toggleSectionSelection('add', true)" style="padding:2px 8px;font-size:11px;background:#F0FDF4;border:1px solid #BBF7D0;border-radius:4px;cursor:pointer;">✓ 전체 선택</button>
        <button type="button" onclick="toggleSectionSelection('add', false)" style="padding:2px 8px;font-size:11px;background:#FFFFFF;border:1px solid #D1D5DB;border-radius:4px;cursor:pointer;">☐ 전체 해제</button>
      </div>
    </div>`;
    html += '<div style="display:flex;flex-direction:column;gap:4px;margin-top:6px;">';
    for (const [idx, a] of report.added.entries()) {
      const detailsId = `diff-add-${idx}`;
      html += `<div style="background:#F0FDF4;border:1px solid #BBF7D0;border-radius:6px;overflow:hidden;">
        <label style="display:flex;align-items:flex-start;gap:8px;padding:6px 10px;cursor:pointer;">
          <input type="checkbox" class="diff-add-chk" value="${_escapeHtml(a.code)}" checked style="margin-top:3px;">
          <div style="flex:1;"><strong>${_escapeHtml(a.code)}</strong> — ${_escapeHtml(a.name)} <span style="font-size:11px;color:var(--muted);">(예상 TC ~${a.estimated_tc_count}개)</span></div>
          ${a.body ? `<button type="button" onclick="toggleDiffDetail('${detailsId}', this)" style="margin-left:auto;padding:2px 8px;font-size:11px;background:#FFFFFF;border:1px solid #86EFAC;border-radius:4px;cursor:pointer;white-space:nowrap;">▸ 내용</button>` : ''}
        </label>
        ${a.body ? `<div id="${detailsId}" style="display:none;padding:8px 12px;border-top:1px solid #BBF7D0;background:#FFFFFF;">
          <div style="font-size:11px;color:var(--muted);margin-bottom:4px;">신규 화면 내용 (미리보기)</div>
          <pre style="font-size:11px;background:#F9FAFB;padding:8px;border-radius:4px;max-height:240px;overflow:auto;white-space:pre-wrap;word-break:break-word;">${_escapeHtml(a.body)}</pre>
        </div>` : ''}
      </div>`;
    }
    html += '</div>';
  }
  if (report.modified.length) {
    html += `<div style="margin-top:10px;display:flex;align-items:center;justify-content:space-between;">
      <strong style="color:#B45309;">🔄 수정 화면 (승인 시 해당 TC ID 유지하며 재작성) <span style="font-size:11px;color:var(--muted);font-weight:normal;">${report.modified.length}개</span></strong>
      <div style="display:flex;gap:4px;">
        <button type="button" onclick="toggleSectionSelection('mod', true)" style="padding:2px 8px;font-size:11px;background:#FEF3C7;border:1px solid #FDE68A;border-radius:4px;cursor:pointer;">✓ 전체 선택</button>
        <button type="button" onclick="toggleSectionSelection('mod', false)" style="padding:2px 8px;font-size:11px;background:#FFFFFF;border:1px solid #D1D5DB;border-radius:4px;cursor:pointer;">☐ 전체 해제</button>
      </div>
    </div>`;
    html += '<div style="display:flex;flex-direction:column;gap:4px;margin-top:6px;">';
    for (const [idx, m] of report.modified.entries()) {
      const tcIds = (m.affected_tc_ids || []).slice(0, 5).join(', ') + (m.affected_count > 5 ? ` 외 ${m.affected_count - 5}` : '');
      const detailsId = `diff-mod-${idx}`;
      html += `<div style="background:#FEF3C7;border:1px solid #FDE68A;border-radius:6px;overflow:hidden;">
        <label style="display:flex;align-items:flex-start;gap:8px;padding:6px 10px;cursor:pointer;">
          <input type="checkbox" class="diff-mod-chk" value="${_escapeHtml(m.code)}" ${m.affected_count>0?'checked':''} style="margin-top:3px;">
          <div style="flex:1;"><strong>${_escapeHtml(m.code)}</strong> — ${_escapeHtml(m.name)}<br>
            <span style="font-size:11px;color:#92400E;">영향 TC: ${m.affected_count}개 ${tcIds ? '(' + _escapeHtml(tcIds) + ')' : '(매핑 없음)'}</span>
          </div>
          <button type="button" onclick="toggleDiffDetail('${detailsId}', this)" style="margin-left:auto;padding:2px 8px;font-size:11px;background:#FFFFFF;border:1px solid #FCD34D;border-radius:4px;cursor:pointer;white-space:nowrap;">▸ 변경 보기</button>
        </label>
        <div id="${detailsId}" style="display:none;padding:8px 12px;border-top:1px solid #FDE68A;background:#FFFFFF;">
          <div style="display:flex;gap:8px;margin-bottom:6px;font-size:11px;">
            <button type="button" onclick="switchDiffMode('${detailsId}', 'unified')" class="diff-mode-btn diff-mode-unified-${idx}" style="padding:3px 10px;background:#FDE68A;border:1px solid #F59E0B;border-radius:4px;cursor:pointer;font-weight:600;">🔀 변경점만 (기본)</button>
            <button type="button" onclick="switchDiffMode('${detailsId}', 'sidebyside')" class="diff-mode-btn diff-mode-sidebyside-${idx}" style="padding:3px 10px;background:#FFFFFF;border:1px solid #D1D5DB;border-radius:4px;cursor:pointer;">📄 전체 2단 보기</button>
          </div>
          <div class="diff-view-unified-${idx}">${_renderUnifiedDiff(m.line_diff)}</div>
          <div class="diff-view-sidebyside-${idx}" style="display:none;">
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;">
              <div>
                <div style="font-size:11px;color:#991B1B;font-weight:600;margin-bottom:4px;">— 이전</div>
                <pre style="font-size:11px;background:#FEF2F2;padding:8px;border-radius:4px;max-height:260px;overflow:auto;white-space:pre-wrap;word-break:break-word;">${_escapeHtml(m.old_body || '(비어있음)')}</pre>
              </div>
              <div>
                <div style="font-size:11px;color:#065F46;font-weight:600;margin-bottom:4px;">+ 새</div>
                <pre style="font-size:11px;background:#F0FDF4;padding:8px;border-radius:4px;max-height:260px;overflow:auto;white-space:pre-wrap;">${_escapeHtml(m.new_body || '(비어있음)')}</pre>
              </div>
            </div>
          </div>
        </div>
      </div>`;
    }
    html += '</div>';
  }
  if (report.removed.length) {
    html += `<div style="margin-top:10px;display:flex;align-items:center;justify-content:space-between;">
      <strong style="color:#6B7280;">🗑️ 삭제 화면 (승인 시 기존 TC에 Deprecated 표시) <span style="font-size:11px;color:var(--muted);font-weight:normal;">${report.removed.length}개</span></strong>
      <div style="display:flex;gap:4px;">
        <button type="button" onclick="toggleSectionSelection('rem', true)" style="padding:2px 8px;font-size:11px;background:#F3F4F6;border:1px solid #D1D5DB;border-radius:4px;cursor:pointer;">✓ 전체 선택</button>
        <button type="button" onclick="toggleSectionSelection('rem', false)" style="padding:2px 8px;font-size:11px;background:#FFFFFF;border:1px solid #D1D5DB;border-radius:4px;cursor:pointer;">☐ 전체 해제</button>
      </div>
    </div>`;
    html += '<div style="display:flex;flex-direction:column;gap:4px;margin-top:6px;">';
    for (const [idx, r] of report.removed.entries()) {
      const tcIds = (r.affected_tc_ids || []).slice(0, 5).join(', ') + (r.affected_count > 5 ? ` 외 ${r.affected_count - 5}` : '');
      html += `<label style="display:flex;align-items:flex-start;gap:8px;padding:6px 10px;background:#F3F4F6;border:1px solid #D1D5DB;border-radius:6px;cursor:pointer;">
        <input type="checkbox" class="diff-rem-chk" value="${_escapeHtml(r.code)}" ${r.affected_count>0?'checked':''} style="margin-top:3px;">
        <div style="flex:1;"><strong>${_escapeHtml(r.code)}</strong> — ${_escapeHtml(r.name)}<br>
          <span style="font-size:11px;color:var(--muted);">영향 TC: ${r.affected_count}개 ${tcIds ? '(' + _escapeHtml(tcIds) + ')' : '(매핑 없음)'}</span>
        </div>
      </label>`;
    }
    html += '</div>';
  }
  sectionsEl.innerHTML = html;

  // 체크박스 변경 시 승인 버튼 상태 동기화
  document.querySelectorAll('.diff-add-chk, .diff-mod-chk, .diff-rem-chk').forEach(cb => {
    cb.addEventListener('change', _updateApproveButtonState);
  });
  _updateApproveButtonState();
}

function cancelDiff() {
  document.getElementById('diffReportArea').classList.add('hidden');
  _diffSid = null;
  _diffReport = null;
}

// ──────────────────────────────────────────────────────────
// v0.11.x — 스마트 분석 (그룹별 진행 + 캐시 + 사용자 통제)
// ──────────────────────────────────────────────────────────
let _smartPlan = null;            // 사전 견적 결과
let _smartCurrentIndex = 0;        // 현재 진행 그룹 index
let _smartResults = [];            // 누적된 summary 배열
let _smartStopped = false;

async function startSmartAnalyze() {
  const prevFolder = document.getElementById('updatePrevFolder').value.trim();
  const newFolder = document.getElementById('updateNewFolder').value.trim();
  if (!prevFolder || !newFolder) {
    alert('이전 기획서 폴더와 새 기획서 폴더 경로 모두 입력하세요.');
    return;
  }

  // 다른 영역 숨기기
  document.getElementById('diffReportArea').classList.add('hidden');
  document.getElementById('smartProgressArea').classList.add('hidden');
  document.getElementById('smartFinalArea').classList.add('hidden');

  const planArea = document.getElementById('smartPlanArea');
  const statsEl = document.getElementById('smartPlanStats');
  const groupsEl = document.getElementById('smartPlanGroups');
  planArea.classList.remove('hidden');
  statsEl.innerHTML = '<em style="color:#6B7280;">⏳ 견적 산출 중... (AI 없음, 즉시)</em>';
  groupsEl.innerHTML = '';

  try {
    const r = await fetch('/update/analyze-plan', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ prev_folder: prevFolder, new_folder: newFolder }),
    });
    const d = await r.json();
    if (!d.ok) {
      statsEl.innerHTML = '<span style="color:#DC2626;">❌ ' + (d.error || '오류') + '</span>';
      return;
    }
    _smartPlan = d;
    _smartCurrentIndex = 0;
    _smartResults = [];
    _smartStopped = false;
    renderSmartPlan(d);
  } catch (e) {
    statsEl.innerHTML = '<span style="color:#DC2626;">❌ 네트워크 오류: ' + e.message + '</span>';
  }
}

function renderSmartPlan(d) {
  const statsEl = document.getElementById('smartPlanStats');
  const groupsEl = document.getElementById('smartPlanGroups');
  const s = d.scr_diff_stats || {};
  const scrs = d.scrs || [];
  const cachedScrs = d.cached_scrs || [];

  statsEl.innerHTML =
    '<div style="display:flex; gap:10px; flex-wrap:wrap; font-size:12px;">' +
    '<span style="background:#DCFCE7; color:#166534; padding:4px 10px; border-radius:4px;">🆕 신규 SCR ' + s.added + '</span>' +
    '<span style="background:#FEF3C7; color:#92400E; padding:4px 10px; border-radius:4px;">✏️ 수정 SCR ' + s.modified + '</span>' +
    '<span style="background:#FEE2E2; color:#991B1B; padding:4px 10px; border-radius:4px;">🗑️ 삭제 SCR ' + s.removed + '</span>' +
    '<span style="background:#F3F4F6; color:#374151; padding:4px 10px; border-radius:4px;">⚪ 동일 ' + s.unchanged + '</span>' +
    '</div>' +
    '<div style="margin-top:8px; font-size:12px; color:#374151;">' +
    '📋 <strong>' + scrs.length + '개 SCR 변경</strong> · 💾 캐시 ' + cachedScrs.length +
    ' · ⏱ 예상 ' + Math.max(1, Math.ceil((scrs.length - cachedScrs.length) * 5 / 60)) + '~' +
    Math.max(1, Math.ceil((scrs.length - cachedScrs.length) * 10 / 60)) + '분 (선택한 SCR 기준)' +
    '</div>';

  // 종류별 정렬 — added → modified → removed (백엔드가 이미 그 순서로 보냄)
  // 타입별 일괄 토글 + 전체 체크박스
  const byType = { added: [], modified: [], removed: [] };
  scrs.forEach(function(s) { (byType[s.type] || []).push(s); });

  const icon = { added: '🆕', modified: '✏️', removed: '🗑️' };
  const typeLabel = { added: '신규 SCR', modified: '수정 SCR', removed: '삭제 SCR' };
  const bg = { added: '#F0FDF4', modified: '#FFFBEB', removed: '#FEF2F2' };
  const border = { added: '#BBF7D0', modified: '#FDE68A', removed: '#FCA5A5' };

  let html = '';
  // 전체 토글
  html += '<div style="display:flex; gap:10px; align-items:center; margin-bottom:10px; padding-bottom:8px; border-bottom:1px solid #E5E7EB;">'
    +   '<label style="font-size:12px; font-weight:600; cursor:pointer;"><input type="checkbox" id="scrSelectAll" onchange="toggleAllScrs(this.checked)" checked> 전체 선택</label>'
    +   '<span style="font-size:11px; color:#6B7280;">— 분석할 SCR 만 체크하세요 (캐시된 SCR 은 즉시 결과)</span>'
    + '</div>';

  ['added', 'modified', 'removed'].forEach(function(t) {
    const list = byType[t];
    if (!list.length) return;
    // 신규 (added) 타입은 헤더에 "신규 TC 생성 모드로 이동" 버튼 추가
    const newTcBtn = t === 'added'
      ? '<button type="button" onclick="goToNewTcForAddedScrs()" '
        + 'style="margin-left:auto; font-size:11px; padding:4px 10px; background:#F0FDF4; color:#166534; border:1px solid #86EFAC; border-radius:4px; cursor:pointer; font-weight:600;">'
        + '✨ 신규 TC 생성 모드로 이동 →'
        + '</button>'
      : '';
    html += '<div style="margin-bottom:14px;">'
      + '<div style="display:flex; gap:8px; align-items:center; margin-bottom:6px;">'
      +   '<label style="font-size:12px; font-weight:600; cursor:pointer;">'
      +     '<input type="checkbox" class="scr-type-all" data-type="' + t + '" checked>'
      +     ' ' + icon[t] + ' ' + typeLabel[t] + ' (' + list.length + '개)'
      +   '</label>'
      +   newTcBtn
      + '</div>'
      + '<div style="display:grid; grid-template-columns:repeat(auto-fill, minmax(160px, 1fr)); gap:6px;">';
    list.forEach(function(s) {
      const cacheTag = s.from_cache
        ? '<span style="background:#DBEAFE; color:#1E40AF; padding:0 4px; border-radius:3px; font-size:9px; margin-left:4px;">💾</span>'
        : '';
      html += '<label class="scr-check-row" style="display:flex; align-items:center; gap:6px; padding:5px 8px; background:' + bg[t] + '; border:1px solid ' + border[t] + '; border-radius:4px; cursor:pointer; font-size:12px;">'
        + '<input type="checkbox" class="scr-chk" data-scr="' + escapeHtml(s.scr_id) + '" data-type="' + t + '" checked>'
        + '<span style="font-family:monospace;">' + escapeHtml(s.scr_id) + '</span>'
        + cacheTag
        + '</label>';
    });
    html += '</div></div>';
  });

  groupsEl.innerHTML = html;
  updateStartButtonLabel();

  // 체크박스 변화 시 시작 버튼 라벨 갱신
  document.querySelectorAll('.scr-chk').forEach(function(cb) {
    cb.addEventListener('change', function() {
      updateStartButtonLabel();
      syncTypeCheckbox(cb.dataset.type);
      syncSelectAllCheckbox();
    });
  });
  // 타입별 일괄 토글
  document.querySelectorAll('.scr-type-all').forEach(function(cb) {
    cb.addEventListener('change', function() {
      toggleTypeScrs(cb.dataset.type, cb.checked);
    });
  });
}

function toggleAllScrs(checked) {
  document.querySelectorAll('.scr-chk').forEach(function(cb) { cb.checked = checked; });
  document.querySelectorAll('.scr-type-all').forEach(function(cb) { cb.checked = checked; });
  updateStartButtonLabel();
}

function toggleTypeScrs(type, checked) {
  document.querySelectorAll('.scr-chk[data-type="' + type + '"]').forEach(function(cb) {
    cb.checked = checked;
  });
  updateStartButtonLabel();
  syncSelectAllCheckbox();
}

function syncTypeCheckbox(type) {
  const all = document.querySelectorAll('.scr-chk[data-type="' + type + '"]');
  const checked = document.querySelectorAll('.scr-chk[data-type="' + type + '"]:checked');
  const typeCb = document.querySelector('.scr-type-all[data-type="' + type + '"]');
  if (typeCb) typeCb.checked = all.length === checked.length;
}

function syncSelectAllCheckbox() {
  const all = document.querySelectorAll('.scr-chk');
  const checked = document.querySelectorAll('.scr-chk:checked');
  const allCb = document.getElementById('scrSelectAll');
  if (allCb) allCb.checked = all.length === checked.length;
}

function getSelectedScrs() {
  return Array.from(document.querySelectorAll('.scr-chk:checked'))
    .map(function(cb) { return cb.dataset.scr; });
}

function updateStartButtonLabel() {
  const btn = document.getElementById('startGroupAnalysisBtn');
  if (!btn) return;
  const n = getSelectedScrs().length;
  if (n === 0) {
    btn.textContent = '⚠ SCR 을 선택하세요';
    btn.disabled = true;
  } else {
    btn.textContent = '▶ 선택한 ' + n + '개 SCR 분석 시작';
    btn.disabled = false;
  }
}

function cancelSmartPlan() {
  document.getElementById('smartPlanArea').classList.add('hidden');
  _smartPlan = null;
}

// 선택된 SCR 리스트 (사용자가 체크박스로 고른 것)
let _selectedScrs = [];

async function startGroupAnalysis() {
  if (!_smartPlan) return;
  const selected = getSelectedScrs();
  if (selected.length === 0) {
    alert('분석할 SCR 을 하나 이상 선택하세요.');
    return;
  }
  _selectedScrs = selected;
  _smartResults = [];
  _smartCurrentIndex = 0;
  _smartStopped = false;
  document.getElementById('smartPlanArea').classList.add('hidden');
  document.getElementById('smartProgressArea').classList.remove('hidden');
  document.getElementById('smartCurrentResult').innerHTML = '';
  await runScr(0);
}

async function runScr(idx) {
  if (!_selectedScrs || idx >= _selectedScrs.length) {
    showFinalSummary();
    return;
  }
  _smartCurrentIndex = idx;
  _smartStopped = false;

  renderProgressList();
  const resultEl = document.getElementById('smartCurrentResult');
  const actionEl = document.getElementById('smartActionRow');
  const scrId = _selectedScrs[idx];
  resultEl.innerHTML = '<em style="color:#6B7280;">⏳ ' + escapeHtml(scrId) + ' 분석 중 ('
    + (idx + 1) + '/' + _selectedScrs.length + ')...</em>';
  actionEl.innerHTML = '';

  const prevFolder = document.getElementById('updatePrevFolder').value.trim();
  const newFolder = document.getElementById('updateNewFolder').value.trim();

  try {
    const r = await fetch('/update/analyze-scr', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({
        prev_folder: prevFolder,
        new_folder: newFolder,
        scr_id: scrId,
      }),
    });
    const d = await r.json();
    if (!d.ok) {
      resultEl.innerHTML = '<span style="color:#DC2626;">❌ ' + (d.error || '오류') + '</span>';
      renderActionRow(idx, idx + 1 >= _selectedScrs.length);
      return;
    }
    _smartResults[idx] = {
      scr_id: d.scr_id,
      change_type: d.change_type,
      summary: d.summary,
      from_cache: d.from_cache,
      cached_at: d.cached_at,
      skipped: false,
    };
    const cacheTag = d.from_cache
      ? '<span style="background:#DBEAFE; color:#1E40AF; padding:1px 8px; border-radius:4px; font-size:10px; margin-left:6px;">💾 캐시 ' + escapeHtml(d.cached_at || '') + '</span>'
      : '<span style="background:#DCFCE7; color:#166534; padding:1px 8px; border-radius:4px; font-size:10px; margin-left:6px;">✨ 새 분석</span>';
    const typeIcon = { added: '🆕', modified: '✏️', removed: '🗑️' }[d.change_type] || '·';
    resultEl.innerHTML = '<div style="margin-bottom:8px;">'
      + '<strong>' + (idx + 1) + ' / ' + _selectedScrs.length + ' · ' + typeIcon + ' '
      + escapeHtml(d.scr_id) + '</strong>' + cacheTag + '</div>'
      + renderMarkdownBasic(d.summary || '(요약 없음)');
    renderProgressList();
    const isLast = (idx + 1) >= _selectedScrs.length;
    renderActionRow(idx, isLast);
  } catch (e) {
    resultEl.innerHTML = '<span style="color:#DC2626;">❌ 네트워크 오류: ' + e.message + '</span>';
    renderActionRow(idx, idx + 1 >= _selectedScrs.length);
  }
}

function renderProgressList() {
  const listEl = document.getElementById('smartProgressList');
  if (!_selectedScrs || !_selectedScrs.length) { listEl.innerHTML = ''; return; }
  const total = _selectedScrs.length;
  let html = '<div style="display:flex; gap:4px; flex-wrap:wrap;">';
  for (let i = 0; i < total; i++) {
    const r = _smartResults[i];
    let state, bg, color;
    if (i < _smartCurrentIndex) {
      if (r && r.skipped) { state = '⏭'; bg = '#F3F4F6'; color = '#6B7280'; }
      else if (r && r.from_cache) { state = '💾'; bg = '#DBEAFE'; color = '#1E40AF'; }
      else if (r) { state = '✅'; bg = '#DCFCE7'; color = '#166534'; }
      else { state = '?'; bg = '#F3F4F6'; color = '#6B7280'; }
    } else if (i === _smartCurrentIndex) {
      if (r) {
        if (r.skipped) { state = '⏭'; bg = '#F3F4F6'; color = '#6B7280'; }
        else if (r.from_cache) { state = '💾'; bg = '#DBEAFE'; color = '#1E40AF'; }
        else { state = '✅'; bg = '#DCFCE7'; color = '#166534'; }
      } else {
        state = '⏳'; bg = '#FEF3C7'; color = '#92400E';
      }
    } else {
      state = '⏸'; bg = '#F9FAFB'; color = '#9CA3AF';
    }
    const scr = _selectedScrs[i];
    html += '<span style="background:' + bg + '; color:' + color + '; padding:3px 8px; border-radius:4px; font-size:10px; font-family:monospace;">'
      + state + ' ' + escapeHtml(scr) + '</span>';
  }
  html += '</div>';
  listEl.innerHTML = html;
}

function renderActionRow(idx, isLast) {
  const actionEl = document.getElementById('smartActionRow');
  if (isLast) {
    actionEl.innerHTML = ''
      + '<button class="btn btn-primary" onclick="showFinalSummary()">'
      +   '✅ 완료 — 합본 보기'
      + '</button>';
  } else {
    const nextScr = _selectedScrs[idx + 1] || '';
    actionEl.innerHTML = ''
      + '<button class="btn btn-primary" onclick="continueScr(' + (idx + 1) + ')">'
      +   '▶ 이어서 ' + escapeHtml(nextScr) + ' 분석'
      + '</button>'
      + '<button class="btn" onclick="skipScr(' + (idx + 1) + ')" style="background:#F3F4F6;">'
      +   '⏭ ' + escapeHtml(nextScr) + ' 건너뛰기'
      + '</button>'
      + '<button class="btn" onclick="stopAnalysis()" style="background:#FEE2E2; color:#991B1B;">'
      +   '🛑 여기서 종료'
      + '</button>';
  }
}

async function continueScr(idx) {
  await runScr(idx);
}

function skipScr(idx) {
  if (!_selectedScrs) return;
  _smartResults[idx] = {
    scr_id: _selectedScrs[idx],
    summary: '_(건너뛰었습니다 — 사용자 선택)_',
    from_cache: false,
    skipped: true,
  };
  _smartCurrentIndex = idx;
  renderProgressList();
  const resultEl = document.getElementById('smartCurrentResult');
  resultEl.innerHTML = '<div style="color:#6B7280;">⏭ ' + escapeHtml(_selectedScrs[idx]) + ' 건너뜀.</div>';
  renderActionRow(idx, idx + 1 >= _selectedScrs.length);
}

function stopAnalysis() {
  _smartStopped = true;
  showFinalSummary();
}

function showFinalSummary() {
  document.getElementById('smartProgressArea').classList.add('hidden');
  const finalArea = document.getElementById('smartFinalArea');
  finalArea.classList.remove('hidden');
  const sumEl = document.getElementById('smartFinalSummary');

  const analyzed = _smartResults.filter(function(r) { return r && !r.skipped; }).length;
  const skipped = _smartResults.filter(function(r) { return r && r.skipped; }).length;
  const stopped = _smartStopped;
  const total = _selectedScrs ? _selectedScrs.length : 0;

  let html = '<div style="font-size:12px; color:#374151; margin-bottom:10px;">';
  html += '📊 총 ' + total + '개 SCR — ✅ 분석 ' + analyzed + ' · ⏭ 건너뜀 ' + skipped;
  if (stopped) html += ' · 🛑 중단됨';
  html += '</div>';

  const typeIcon = { added: '🆕', modified: '✏️', removed: '🗑️' };
  for (let i = 0; i < _smartResults.length; i++) {
    const r = _smartResults[i];
    if (!r) continue;
    html += '<div style="margin-bottom:14px; padding-bottom:12px; border-bottom:1px solid #E5E7EB;">';
    html += '<div style="font-size:12px; font-weight:700; color:#1E3A5F; margin-bottom:6px;">';
    html += (typeIcon[r.change_type] || '·') + ' ' + escapeHtml(r.scr_id || '?');
    if (r.from_cache) html += ' <span style="background:#DBEAFE; color:#1E40AF; padding:1px 6px; border-radius:3px; font-size:10px;">💾 캐시</span>';
    if (r.skipped) html += ' <span style="background:#F3F4F6; color:#6B7280; padding:1px 6px; border-radius:3px; font-size:10px;">⏭ 건너뜀</span>';
    html += '</div>';
    html += '<div style="font-size:12px; color:#111827;">' + renderMarkdownBasic(r.summary || '') + '</div>';
    html += '</div>';
  }
  sumEl.innerHTML = html;
}

function resetSmartAnalyze() {
  document.getElementById('smartFinalArea').classList.add('hidden');
  document.getElementById('smartPlanArea').classList.add('hidden');
  document.getElementById('smartProgressArea').classList.add('hidden');
  _smartPlan = null;
  _smartResults = [];
  _smartCurrentIndex = 0;
}

// ──────────────────────────────────────────────────────────
// v0.11.x 신규 — 폴더 기반 spec diff 분석 (1단계: 분석만, 레거시)
// ──────────────────────────────────────────────────────────
function cancelUpdate() {
  document.getElementById('diffReportArea').classList.add('hidden');
}

async function startUpdateAnalyze() {
  const prevFolder = document.getElementById('updatePrevFolder').value.trim();
  const newFolder = document.getElementById('updateNewFolder').value.trim();
  const tcUrl = document.getElementById('updateTcUrl').value.trim();
  const btn = document.getElementById('startDiffBtn');
  const reportArea = document.getElementById('diffReportArea');
  const statsEl = document.getElementById('diffReportStats');
  const summaryEl = document.getElementById('diffReportSummary');

  if (!prevFolder || !newFolder) {
    alert('이전 기획서 폴더와 새 기획서 폴더 경로 모두 입력하세요.');
    return;
  }

  btn.disabled = true;
  btn.textContent = '⏳ 분석 중... (1~3분)';
  reportArea.classList.remove('hidden');
  statsEl.innerHTML = '<em style="color:#6B7280;">AI 가 spec 폴더 비교 + 요약 작성 중...</em>';
  summaryEl.innerHTML = '';

  const projectName = (document.getElementById('projectDropdown') || document.getElementById('projectSelect') || {}).value || '';

  try {
    const resp = await fetch('/update/analyze', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({
        prev_folder: prevFolder,
        new_folder: newFolder,
        existing_tc_url: tcUrl,
        project_name: projectName,
      })
    });
    const data = await resp.json();
    if (!data.ok) {
      statsEl.innerHTML = '<span style="color:#DC2626;">❌ ' + (data.error || '오류') + '</span>';
      if (data.trace) {
        summaryEl.innerHTML = '<pre style="font-size:11px;color:#991B1B;background:#FEE2E2;padding:8px;border-radius:4px;overflow:auto;">' + escapeHtml(data.trace) + '</pre>';
      }
      return;
    }

    // 통계 렌더
    const d = data.diff || {};
    statsEl.innerHTML =
      '<div style="display:flex;gap:14px;flex-wrap:wrap;font-size:12px;">' +
      '<span style="background:#DCFCE7;color:#166534;padding:4px 10px;border-radius:4px;">➕ 신규 ' + (d.added || []).length + '개</span>' +
      '<span style="background:#FEF3C7;color:#92400E;padding:4px 10px;border-radius:4px;">✏️ 수정 ' + (d.modified || []).length + '개</span>' +
      '<span style="background:#FEE2E2;color:#991B1B;padding:4px 10px;border-radius:4px;">🗑️ 삭제 ' + (d.removed || []).length + '개</span>' +
      '<span style="background:#F3F4F6;color:#374151;padding:4px 10px;border-radius:4px;">⚪ 동일 ' + (d.unchanged || []).length + '개</span>' +
      (d.common_changed ? '<span style="background:#FED7AA;color:#9A3412;padding:4px 10px;border-radius:4px;">📋 공통 문서 변경</span>' : '') +
      '</div>' +
      '<div style="margin-top:8px;font-size:11px;color:#6B7280;">분석 시각: ' + (data.meta && data.meta.analyzed_at || '-') + '</div>';

    // AI 요약 마크다운 렌더 (간단 HTML 변환 — 마크다운 라이브러리 없으니 기본만)
    const summaryMd = data.summary || '(요약 없음)';
    summaryEl.innerHTML = renderMarkdownBasic(summaryMd);

    if (typeof showToast === 'function') {
      showToast('✅ 변경 분석 완료 — 신규 ' + (d.added || []).length + ' / 수정 ' + (d.modified || []).length + ' / 삭제 ' + (d.removed || []).length, 'success');
    }
  } catch (e) {
    statsEl.innerHTML = '<span style="color:#DC2626;">❌ 네트워크 오류: ' + e.message + '</span>';
  } finally {
    btn.disabled = false;
    btn.textContent = '🔍 변경사항 분석';
  }
}

// 매우 기본적인 markdown → HTML 변환 (h1/h2/h3/bold/code/list/줄바꿈)
function renderMarkdownBasic(md) {
  if (!md) return '';
  let html = escapeHtml(md);
  // 코드 블록 — \\n 으로 escape (Python f-string 안에 있어서)
  html = html.replace(/```(\\w*)\\n([\\s\\S]*?)```/g, function(_, lang, code) {
    return '<pre style="background:#F3F4F6;padding:10px;border-radius:4px;overflow:auto;font-size:11px;">' + code + '</pre>';
  });
  // 헤딩
  html = html.replace(/^### (.+)$/gm, '<h3 style="font-size:14px;color:#1E40AF;margin:14px 0 6px 0;font-weight:700;">$1</h3>');
  html = html.replace(/^## (.+)$/gm, '<h2 style="font-size:16px;color:#1E3A8A;margin:18px 0 8px 0;font-weight:700;border-bottom:1px solid #E5E7EB;padding-bottom:4px;">$1</h2>');
  html = html.replace(/^# (.+)$/gm, '<h1 style="font-size:18px;color:#1E3A8A;margin:20px 0 10px 0;font-weight:700;">$1</h1>');
  // 인라인 코드
  html = html.replace(/`([^`]+)`/g, '<code style="background:#F3F4F6;padding:1px 5px;border-radius:3px;font-size:12px;">$1</code>');
  // 굵게
  html = html.replace(/\\*\\*([^*]+)\\*\\*/g, '<strong>$1</strong>');
  // 리스트
  html = html.replace(/^- (.+)$/gm, '<li style="margin-left:18px;">$1</li>');
  // 줄바꿈
  html = html.replace(/\\n\\n/g, '<br><br>').replace(/\\n/g, '<br>');
  return html;
}

function escapeHtml(s) {
  return String(s)
    .replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;')
    .replace(/"/g, '&quot;').replace(/'/g, '&#39;');
}

// ── TC Update 2단계 — Plan + 4-way UI ──────────────────────────────────────
let _updateSessionId = null;
let _updateCandidates = [];

// 검증용 — 자동 점검 (Playwright) 에서 모듈 스코프 변수에 접근하기 위해
window.__debugSetCandidates = function(items) {
  _updateCandidates = Array.isArray(items) ? items : [];
  if (!_updateSessionId) _updateSessionId = 'debug-session';
  renderCandidates();
};

async function startUpdatePlan() {
  const prevFolder = document.getElementById('updatePrevFolder').value.trim();
  const newFolder = document.getElementById('updateNewFolder').value.trim();
  const tcUrl = document.getElementById('updateTcUrl').value.trim();
  const btn = document.getElementById('planBtn');
  const planArea = document.getElementById('planReportArea');
  const statsEl = document.getElementById('planStats');
  const linkEl = document.getElementById('planCopyLink');
  const listEl = document.getElementById('candidatesList');

  if (!tcUrl) {
    alert('기존 TC Google Sheets URL 을 입력하세요.');
    return;
  }

  btn.disabled = true;
  btn.textContent = '⏳ 사본 생성 + 매핑 중...';
  planArea.classList.remove('hidden');
  // 큰 진행 배너 — 노란 배경 + 스피너 + 강조된 글자
  statsEl.innerHTML =
    '<div style="display:flex; align-items:center; gap:12px; padding:14px 16px; '
    + 'background:linear-gradient(135deg, #FEF3C7 0%, #FDE68A 100%); '
    + 'border:1px solid #F59E0B; border-radius:8px;">'
    +   '<div class="plan-spinner" style="width:24px; height:24px; '
    +     'border:3px solid #FCD34D; border-top-color:#92400E; '
    +     'border-radius:50%; animation:plan-spin 0.8s linear infinite; flex-shrink:0;"></div>'
    +   '<div style="flex:1;">'
    +     '<div style="font-size:14px; font-weight:700; color:#78350F;">⏳ 후보 산출 중 — 잠시 기다려주세요</div>'
    +     '<div style="font-size:12px; color:#92400E; margin-top:4px;">'
    +       'Sheets 사본 생성 → TC↔SCR 매핑 → 4-way 후보 리스트 산출'
    +     '</div>'
    +   '</div>'
    + '</div>'
    + '<style>@keyframes plan-spin { to { transform:rotate(360deg); } }</style>';
  linkEl.innerHTML = '';
  listEl.innerHTML = '';
  // 적용 버튼이 이미 화면에 있을 수 있으므로 비활성화 (이전 plan 의 잔존)
  const applyBtn = document.getElementById('applyBtn');
  if (applyBtn) {
    applyBtn.disabled = true;
    applyBtn.style.opacity = '0.4';
    applyBtn.style.cursor = 'not-allowed';
    applyBtn.dataset.loadingMsg = '⏳ 후보 산출 중 — 잠시만요';
  }

  const projectName = (document.getElementById('projectDropdown') || document.getElementById('projectSelect') || {}).value || '';

  try {
    const resp = await fetch('/update/plan', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({
        prev_folder: prevFolder,
        new_folder: newFolder,
        existing_tc_url: tcUrl,
        make_copy: true,
        project_name: projectName,
        // 1단계에서 사용자가 선택한 SCR 들 (없으면 백엔드가 전체 처리)
        scr_filter: (typeof _selectedScrs !== 'undefined' && _selectedScrs && _selectedScrs.length)
                      ? _selectedScrs : null,
        // 신규(add) 는 별도 '신규 TC 생성 모드' 에서 처리 → 2단계에서는 수정/삭제만
        exclude_actions: ['add'],
      })
    });
    const data = await resp.json();
    if (!data.ok) {
      statsEl.innerHTML = '<span style="color:#DC2626;">❌ ' + (data.error || '오류') + '</span>';
      if (data.trace) {
        listEl.innerHTML = '<pre style="font-size:11px;color:#991B1B;background:#FEE2E2;padding:8px;overflow:auto;">' + escapeHtml(data.trace) + '</pre>';
      }
      return;
    }

    _updateSessionId = data.session_id;
    _updateCandidates = data.candidates || [];
    const s = data.stats || {};

    // 필터 적용 여부 안내
    const filterNote = (typeof _selectedScrs !== 'undefined' && _selectedScrs && _selectedScrs.length)
      ? '<div style="margin-bottom:8px; padding:6px 10px; background:#EEF2FF; border:1px solid #C7D2FE; border-radius:4px; font-size:12px; color:#3730A3;">'
          + '🎯 <strong>필터 적용</strong> — 1단계 선택 ' + _selectedScrs.length + '개 SCR 중 <strong>수정/삭제</strong>만 (신규는 별도 모드)<br>'
          + '<span style="font-size:11px;">' + _selectedScrs.map(escapeHtml).join(', ') + '</span>'
          + '</div>'
      : '<div style="margin-bottom:8px; padding:6px 10px; background:#F3F4F6; border:1px solid #D1D5DB; border-radius:4px; font-size:11px; color:#6B7280;">'
          + 'ℹ️ 신규(add) 후보는 제외됨 — 신규 SCR 은 <strong>신규 TC 생성 모드</strong> 에서 처리하세요.'
          + '</div>';

    statsEl.innerHTML = filterNote +
      '<div style="display:flex;gap:10px;flex-wrap:wrap;font-size:12px;">' +
      '<span style="background:#DCFCE7;color:#166534;padding:4px 10px;border-radius:4px;">🆕 신규 ' + s.add + '</span>' +
      '<span style="background:#FEF3C7;color:#92400E;padding:4px 10px;border-radius:4px;">✏️ 수정 ' + s.modify + '</span>' +
      '<span style="background:#FEE2E2;color:#991B1B;padding:4px 10px;border-radius:4px;">🗑️ 삭제 ' + s.delete + '</span>' +
      '<span style="background:#E0E7FF;color:#3730A3;padding:4px 10px;border-radius:4px;">기본 체크 ' + s.default_checked + '/' + s.total + '</span>' +
      '<span style="background:#F3F4F6;color:#374151;padding:4px 10px;border-radius:4px;">기존 TC ' + s.tcs_total + '개</span>' +
      '</div>' +
      '<div style="margin-top:6px;font-size:11px;color:#6B7280;">SCR diff — 신규 ' + s.scr_added + ' · 수정 ' + s.scr_modified + ' · 삭제 ' + s.scr_removed + ' · 동일 ' + s.scr_unchanged + ' | 시트: ' + (s.tabs || []).join(', ') + '</div>';

    if (data.copy_url) {
      linkEl.innerHTML = '📋 사본 생성됨: <a href="' + data.copy_url + '" target="_blank" style="color:#1E40AF;">' + escapeHtml(data.copy_title || '사본 Sheets') + '</a>';
      window._lastPlanCopyUrl = data.copy_url;  // promote 에서 사용
    }

    renderCandidates();

    if (typeof showToast === 'function') {
      showToast('✅ 영향 분석 완료 — 후보 ' + s.total + '건', 'success');
    }
  } catch (e) {
    statsEl.innerHTML = '<span style="color:#DC2626;">❌ 네트워크 오류: ' + e.message + '</span>';
  } finally {
    btn.disabled = false;
    btn.textContent = '🧭 2단계 — TC 매핑 + 영향 분석 (사본 생성)';
    // 적용 버튼 다시 활성화 (로딩 끝)
    const applyBtn2 = document.getElementById('applyBtn');
    if (applyBtn2) {
      applyBtn2.disabled = false;
      applyBtn2.style.opacity = '';
      applyBtn2.style.cursor = '';
      delete applyBtn2.dataset.loadingMsg;
    }
  }
}

function renderCandidates() {
  const listEl = document.getElementById('candidatesList');
  const filters = {
    add: document.getElementById('filterAdd').checked,
    modify: document.getElementById('filterModify').checked,
    delete: document.getElementById('filterDelete').checked,
  };
  const visible = _updateCandidates
    .map((c, idx) => ({ ...c, idx }))
    .filter(c => filters[c.action]);

  if (visible.length === 0) {
    listEl.innerHTML = '<div style="padding:14px;color:#6B7280;text-align:center;">후보 없음 (필터 또는 unchanged-only)</div>';
    return;
  }

  const icon = { add: '🆕', modify: '✏️', delete: '🗑️', skip: '⏸️' };
  const bg = { add: '#F0FDF4', modify: '#FFFBEB', delete: '#FEF2F2', skip: '#F9FAFB' };

  // SCR 별 그룹화 — Map 으로 순서 보존
  const groups = new Map();
  visible.forEach(function(c) {
    const key = c.scr_id || '(공통)';
    if (!groups.has(key)) {
      groups.set(key, { scr_id: key, sheet_title: c.sheet_title || '', items: [] });
    }
    groups.get(key).items.push(c);
  });

  function renderRow(c) {
    const kindTag = c.change_kind
      ? '<span style="background:#E0E7FF;color:#3730A3;padding:1px 6px;border-radius:3px;font-size:10px;margin-left:6px;">' + c.change_kind + '</span>'
      : '';
    const tcInfo = c.tc_id
      ? '<code style="background:#F3F4F6;padding:1px 5px;border-radius:3px;font-size:11px;">' + escapeHtml(c.tc_id) + '</code>'
        + (c.tc_title ? ' <span style="color:#374151;">' + escapeHtml(c.tc_title) + '</span>' : '')
      : '<em style="color:#6B7280;">(신규 — TC ID 미정)</em>';
    return ''
      + '<div style="border-bottom:1px solid #E5E7EB; background:' + bg[c.action] + ';">'
      +   '<div style="display:flex; align-items:flex-start; gap:8px; padding:8px 10px;">'
      +     '<input type="checkbox" class="cand-chk" data-idx="' + c.idx + '" data-scr="' + escapeHtml(c.scr_id || '(공통)') + '" data-action="' + c.action + '"'
      +       (c.default_checked ? ' checked' : '') + ' style="margin-top:3px;" onchange="updateGroupCheckCounts()">'
      +     '<div style="flex:1;">'
      +       '<div style="font-size:12px; font-weight:600; color:#111827;">'
      +         icon[c.action] + ' ' + c.action.toUpperCase() + kindTag
      +       '</div>'
      +       '<div style="font-size:12px; margin-top:3px;">' + tcInfo + '</div>'
      +       '<div style="font-size:11px; color:#6B7280; margin-top:3px;">' + escapeHtml(c.reason || '') + '</div>'
      +     '</div>'
      +     '<div style="display:flex; flex-direction:column; gap:4px; flex:0 0 auto; align-self:flex-start;">'
      +       '<button type="button" class="btn-preview" data-idx="' + c.idx + '" onclick="togglePreview(' + c.idx + ')" '
      +         'style="font-size:11px; padding:4px 8px; background:#FFF; border:1px solid #CBD5E1; border-radius:4px; cursor:pointer; color:#1E40AF; white-space:nowrap;">'
      +         '▼ diff 보기'
      +       '</button>'
      +       (c.action === 'modify' && c.tc_id
        ? '<button type="button" class="btn-propose" data-idx="' + c.idx + '" onclick="toggleProposal(' + c.idx + ')" '
          + 'style="font-size:11px; padding:4px 8px; background:#EEF2FF; border:1px solid #C7D2FE; border-radius:4px; cursor:pointer; color:#3730A3; white-space:nowrap;">'
          + '💡 수정안 보기'
          + '</button>'
        : '')
      +       (c.action === 'add'
        ? '<button type="button" class="btn-new-tc" data-scr="' + escapeHtml(c.scr_id || '') + '" onclick="goToNewTcMode(\\'' + (c.scr_id || '') + '\\')" '
          + 'style="font-size:11px; padding:4px 8px; background:#F0FDF4; border:1px solid #86EFAC; border-radius:4px; cursor:pointer; color:#166534; white-space:nowrap; font-weight:600;">'
          + '✨ 신규 TC 생성 →'
          + '</button>'
        : '')
      +     '</div>'
      +   '</div>'
      +   '<div id="preview-' + c.idx + '" class="cand-preview" style="display:none; padding:8px 14px 12px 38px; background:#FFFFFF; border-top:1px dashed #E5E7EB;"></div>'
      +   '<div id="proposal-' + c.idx + '" class="cand-proposal" style="display:none; padding:10px 14px 14px 38px; background:#FAFBFF; border-top:1px dashed #C7D2FE;"></div>'
      + '</div>';
  }

  // 그룹 헤더의 change_kind 요약: 그룹 안 action 모음
  function summarizeKinds(items) {
    const counts = { add: 0, modify: 0, delete: 0, skip: 0 };
    items.forEach(function(c) { counts[c.action] = (counts[c.action] || 0) + 1; });
    const parts = [];
    if (counts.add) parts.push('<span style="background:#DCFCE7;color:#166534;padding:1px 6px;border-radius:3px;font-size:10px;">🆕 ' + counts.add + '</span>');
    if (counts.modify) parts.push('<span style="background:#FEF3C7;color:#92400E;padding:1px 6px;border-radius:3px;font-size:10px;">✏️ ' + counts.modify + '</span>');
    if (counts.delete) parts.push('<span style="background:#FEE2E2;color:#991B1B;padding:1px 6px;border-radius:3px;font-size:10px;">🗑️ ' + counts.delete + '</span>');
    return parts.join(' ');
  }

  // 그룹 헤더 + 그룹 본문 렌더
  const groupHtml = Array.from(groups.values()).map(function(g) {
    const scrId = g.scr_id;
    const sheetTitle = g.sheet_title || '';
    const total = g.items.length;
    const checkedInit = g.items.filter(function(c) { return c.default_checked; }).length;
    const kindBadges = summarizeKinds(g.items);
    const sheetTag = sheetTitle
      ? '<span style="font-size:11px; color:#6B7280;">[' + escapeHtml(sheetTitle) + ']</span>'
      : '';
    // modify + tc_id 가 있어야 일괄 적용 가능
    const groupHasApplicable = g.items.some(function(c) { return c.action === 'modify' && c.tc_id; });

    const headerHtml = ''
      + '<summary style="cursor:pointer; padding:10px 12px; background:#F8FAFC; border-bottom:1px solid #E5E7EB; '
      +   'display:flex; align-items:center; gap:10px; list-style:none;" '
      +   'data-scr-header="' + escapeHtml(scrId) + '">'
      +   '<span class="grp-caret" style="font-size:10px; color:#6B7280; width:10px;">▶</span>'
      +   '<strong style="font-size:13px; color:#111827; font-family:monospace;">' + escapeHtml(scrId) + '</strong>'
      +   sheetTag
      +   '<span style="font-size:11px; color:#374151;">' + total + ' TC</span>'
      +   '<span style="display:flex; gap:4px;">' + kindBadges + '</span>'
      +   '<span class="grp-checked-count" data-scr="' + escapeHtml(scrId) + '" '
      +     'style="font-size:11px; color:#3730A3; background:#E0E7FF; padding:2px 8px; border-radius:10px; margin-left:auto;">'
      +     '체크 <strong>' + checkedInit + '</strong>/' + total
      +   '</span>'
      +   (groupHasApplicable
        ? '<button type="button" class="btn-grp-apply" data-scr="' + escapeHtml(scrId) + '" '
          + 'onclick="event.preventDefault(); event.stopPropagation(); applyBulkForScr(\\'' + escapeJsString(scrId) + '\\');" '
          + 'style="font-size:11px; padding:5px 10px; background:#FEF3C7; border:1px solid #F59E0B; border-radius:4px; cursor:pointer; color:#78350F; font-weight:600; white-space:nowrap;">'
          + '📦 SCR 일괄 적용'
          + '</button>'
        : '')
      + '</summary>';

    const bodyHtml = g.items.map(renderRow).join('');

    return ''
      + '<details class="scr-group" data-scr="' + escapeHtml(scrId) + '" style="border:1px solid #E5E7EB; border-radius:6px; margin-bottom:8px; overflow:hidden; background:#FFF;">'
      +   headerHtml
      +   '<div>' + bodyHtml + '</div>'
      + '</details>';
  }).join('');

  listEl.innerHTML = ''
    + '<div style="display:flex; gap:8px; margin-bottom:8px; font-size:11px;">'
    +   '<button type="button" onclick="expandAllScrGroups(true)" '
    +     'style="padding:4px 10px; background:#FFF; border:1px solid #CBD5E1; border-radius:4px; cursor:pointer; color:#374151;">▼ 모두 펼치기</button>'
    +   '<button type="button" onclick="expandAllScrGroups(false)" '
    +     'style="padding:4px 10px; background:#FFF; border:1px solid #CBD5E1; border-radius:4px; cursor:pointer; color:#374151;">▶ 모두 접기</button>'
    +   '<span style="color:#6B7280; align-self:center; margin-left:4px;">SCR ' + groups.size + '개 — 그룹 헤더 클릭으로 펼침/접기</span>'
    + '</div>'
    + groupHtml;

  // caret 회전 — details 의 toggle 이벤트 캡쳐
  listEl.querySelectorAll('details.scr-group').forEach(function(d) {
    d.addEventListener('toggle', function() {
      const caret = d.querySelector('.grp-caret');
      if (caret) caret.textContent = d.open ? '▼' : '▶';
    });
  });
}

// JS 문자열 안에 단일 따옴표 안전하게 끼우기
function escapeJsString(s) {
  return String(s || '').replace(/\\\\/g, '\\\\\\\\').replace(/'/g, "\\\\'");
}

function expandAllScrGroups(open) {
  document.querySelectorAll('details.scr-group').forEach(function(d) {
    d.open = !!open;
  });
}

// 체크박스 변경 시 각 SCR 그룹의 체크 수 라이브 업데이트
function updateGroupCheckCounts() {
  document.querySelectorAll('.grp-checked-count').forEach(function(el) {
    const scr = el.dataset.scr;
    const all = document.querySelectorAll('.cand-chk[data-scr="' + (scr || '').replace(/"/g, '\\\\"') + '"]');
    const checked = Array.from(all).filter(function(cb) { return cb.checked; }).length;
    el.innerHTML = '체크 <strong>' + checked + '</strong>/' + all.length;
  });
}

async function togglePreview(idx) {
  const area = document.getElementById('preview-' + idx);
  const btn = document.querySelector('.btn-preview[data-idx="' + idx + '"]');
  if (!area) return;
  if (area.style.display !== 'none' && area.dataset.loaded === '1') {
    area.style.display = 'none';
    if (btn) btn.textContent = '▼ diff 보기';
    return;
  }
  area.style.display = 'block';
  if (btn) btn.textContent = '▲ 닫기';
  if (area.dataset.loaded === '1') return;

  area.innerHTML = '<em style="font-size:11px;color:#6B7280;">⏳ SCR 본문 불러오는 중...</em>';
  try {
    const r = await fetch('/update/preview', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ session_id: _updateSessionId, idx: idx }),
    });
    const d = await r.json();
    if (!d.ok) {
      area.innerHTML = '<span style="color:#DC2626;font-size:11px;">❌ ' + (d.error || '오류') + '</span>';
      return;
    }
    area.innerHTML = renderPreview(d);
    area.dataset.loaded = '1';
  } catch (e) {
    area.innerHTML = '<span style="color:#DC2626;font-size:11px;">❌ ' + e.message + '</span>';
  }
}

function renderPreview(d) {
  const c = d.candidate || {};
  let html = '';

  // 메타 정보
  html += '<div style="font-size:11px; color:#374151; margin-bottom:8px;">';
  html += '<strong>SCR:</strong> ' + escapeHtml(d.scr_id || '-') + ' · ';
  html += '<strong>판정:</strong> ' + escapeHtml(c.change_kind || c.action || '-');
  if (c.action === 'add') {
    html += ' · <em>(신규 SCR — 이전 본문 없음, 새 본문만 표시)</em>';
  } else if (c.action === 'delete') {
    html += ' · <em>(삭제된 SCR — 새 본문 없음, 이전 본문만 표시)</em>';
  }
  html += '</div>';

  // unified diff (가장 유용)
  if (d.unified_diff && d.unified_diff.trim()) {
    html += '<div style="font-size:10px; color:#6B7280; margin-bottom:4px;">📋 spec 본문 diff (이전 → 새 버전):</div>';
    html += '<pre style="font-size:11px; line-height:1.5; background:#0F172A; color:#E2E8F0; padding:10px 12px; border-radius:6px; overflow-x:auto; max-height:340px; overflow-y:auto;">';
    html += d.unified_diff.split('\\n').map(function(ln) {
      const esc = escapeHtml(ln);
      if (ln.startsWith('+++') || ln.startsWith('---')) {
        return '<span style="color:#94A3B8;">' + esc + '</span>';
      } else if (ln.startsWith('@@')) {
        return '<span style="color:#A78BFA;">' + esc + '</span>';
      } else if (ln.startsWith('+')) {
        return '<span style="color:#86EFAC; background:rgba(34,197,94,0.15);">' + esc + '</span>';
      } else if (ln.startsWith('-')) {
        return '<span style="color:#FCA5A5; background:rgba(239,68,68,0.15);">' + esc + '</span>';
      }
      return '<span style="color:#CBD5E1;">' + esc + '</span>';
    }).join('\\n');
    html += '</pre>';
  } else if (c.action === 'add') {
    // 신규 SCR — 새 본문만 표시
    html += '<div style="font-size:10px; color:#6B7280; margin-bottom:4px;">📄 신규 SCR 본문:</div>';
    html += '<pre style="font-size:11px; line-height:1.5; background:#F9FAFB; padding:10px 12px; border-radius:6px; max-height:340px; overflow-y:auto; white-space:pre-wrap;">';
    html += escapeHtml(d.new_text || '(본문 없음)');
    html += '</pre>';
  } else if (c.action === 'delete') {
    html += '<div style="font-size:10px; color:#6B7280; margin-bottom:4px;">📄 삭제된 SCR 본문 (이전):</div>';
    html += '<pre style="font-size:11px; line-height:1.5; background:#FEF2F2; padding:10px 12px; border-radius:6px; max-height:340px; overflow-y:auto; white-space:pre-wrap;">';
    html += escapeHtml(d.prev_text || '(본문 없음)');
    html += '</pre>';
  } else {
    html += '<div style="font-size:11px; color:#6B7280;">변경된 내용이 없거나 본문을 찾지 못했습니다.</div>';
  }

  return html;
}

// ── TC Update — AI 수정안 제안/적용 ──────────────────────────────────────
async function toggleProposal(idx) {
  const area = document.getElementById('proposal-' + idx);
  const btn = document.querySelector('.btn-propose[data-idx="' + idx + '"]');
  if (!area) return;
  if (area.style.display !== 'none' && area.dataset.loaded === '1') {
    area.style.display = 'none';
    if (btn) btn.textContent = '💡 수정안 보기';
    return;
  }
  area.style.display = 'block';
  if (btn) btn.textContent = '▲ 닫기';
  if (area.dataset.loaded === '1') return;

  area.innerHTML = '<em style="font-size:11px;color:#6B7280;">⏳ AI 가 수정안 작성 중... (10-20초)</em>';
  try {
    const r = await fetch('/update/propose', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ session_id: _updateSessionId, idx: idx }),
    });
    // HTTP 자체 오류 (429 등) 우선 처리
    if (!r.ok) {
      let text = '';
      try { text = await r.text(); } catch(_) {}
      const msg = 'HTTP ' + r.status + (text ? ' — ' + text.slice(0, 300) : '');
      area.innerHTML = '<div style="padding:8px 10px; background:#FEE2E2; border:1px solid #FCA5A5; border-radius:4px; font-size:11px; color:#991B1B;">'
        + '❌ ' + escapeHtml(msg)
        + (r.status === 429 ? '<br><br>⚠️ Google Sheets API 분당 한도 초과 (60회). 1-2분 기다린 후 다시 시도하세요.' : '')
        + '<br><button onclick="retryProposal(' + idx + ')" style="margin-top:6px; padding:3px 8px; font-size:10px; background:#FFF; border:1px solid #DC2626; border-radius:3px; cursor:pointer;">🔁 재시도</button>'
        + '</div>';
      return;
    }
    const d = await r.json();
    if (!d.ok) {
      const errMsg = d.error || d.trace || '오류 (응답에 메시지 없음)';
      area.innerHTML = '<div style="padding:8px 10px; background:#FEE2E2; border:1px solid #FCA5A5; border-radius:4px; font-size:11px; color:#991B1B;">'
        + '❌ ' + escapeHtml(String(errMsg).slice(0, 500))
        + '<br><button onclick="retryProposal(' + idx + ')" style="margin-top:6px; padding:3px 8px; font-size:10px; background:#FFF; border:1px solid #DC2626; border-radius:3px; cursor:pointer;">🔁 재시도</button>'
        + '</div>';
      return;
    }
    area.innerHTML = renderProposal(d, idx);
    area.dataset.loaded = '1';
  } catch (e) {
    area.innerHTML = '<div style="padding:8px 10px; background:#FEE2E2; border:1px solid #FCA5A5; border-radius:4px; font-size:11px; color:#991B1B;">'
      + '❌ 네트워크 오류: ' + escapeHtml(e.message || '(메시지 없음)')
      + '<br><button onclick="retryProposal(' + idx + ')" style="margin-top:6px; padding:3px 8px; font-size:10px; background:#FFF; border:1px solid #DC2626; border-radius:3px; cursor:pointer;">🔁 재시도</button>'
      + '</div>';
  }
}

function retryProposal(idx) {
  // 캐시 상태 reset 하고 다시 호출
  const area = document.getElementById('proposal-' + idx);
  if (area) {
    delete area.dataset.loaded;
    area.style.display = 'none';
  }
  toggleProposal(idx);
}

function renderProposal(d, idx) {
  const p = d.proposal || {};
  const cur = d.current_tc || {};

  if (p.no_change) {
    return ''
      + '<div style="font-size:12px; color:#374151; padding:8px 10px; background:#F0FDF4; border:1px solid #BBF7D0; border-radius:4px;">'
      +   '<strong style="color:#166534;">✅ 변경 불필요</strong> '
      +   '<span style="color:#6B7280; margin-left:6px;">' + escapeHtml(p.rationale || '') + '</span>'
      + '</div>';
  }

  function fieldDiff(label, before, after) {
    if (after === null || after === undefined) {
      return ''
        + '<div style="margin-bottom:8px;">'
        +   '<div style="font-size:11px; color:#6B7280; margin-bottom:3px;">' + escapeHtml(label) + ' — <em>변경 없음</em></div>'
        + '</div>';
    }
    return ''
      + '<div style="margin-bottom:10px;">'
      +   '<div style="font-size:11px; font-weight:600; color:#3730A3; margin-bottom:4px;">' + escapeHtml(label) + '</div>'
      +   '<div style="display:grid; grid-template-columns:1fr 1fr; gap:6px;">'
      +     '<div style="background:#FEF2F2; border:1px solid #FCA5A5; border-radius:4px; padding:6px 8px;">'
      +       '<div style="font-size:9px; color:#991B1B; margin-bottom:2px; text-transform:uppercase;">현재</div>'
      +       '<pre style="font-size:11px; color:#7F1D1D; white-space:pre-wrap; word-break:break-word; margin:0; font-family:inherit; line-height:1.5;">' + escapeHtml(before || '(없음)') + '</pre>'
      +     '</div>'
      +     '<div style="background:#F0FDF4; border:1px solid #86EFAC; border-radius:4px; padding:6px 8px;">'
      +       '<div style="font-size:9px; color:#166534; margin-bottom:2px; text-transform:uppercase;">제안</div>'
      +       '<pre style="font-size:11px; color:#14532D; white-space:pre-wrap; word-break:break-word; margin:0; font-family:inherit; line-height:1.5;">' + escapeHtml(after || '') + '</pre>'
      +     '</div>'
      +   '</div>'
      + '</div>';
  }

  let html = ''
    + '<div style="font-size:11px; color:#3730A3; padding:6px 10px; background:#EEF2FF; border:1px solid #C7D2FE; border-radius:4px; margin-bottom:10px;">'
    +   '💡 <strong>AI 수정안</strong> — ' + escapeHtml(p.rationale || '')
    + '</div>';

  html += fieldDiff('사전조건', cur.precondition, p.precondition);
  html += fieldDiff('테스트 스텝', cur.steps, p.steps);
  html += fieldDiff('기대결과', cur.expected, p.expected);

  // 적용 버튼 — 각 필드별 체크박스 + 일괄 적용
  const hasAny = (p.steps !== null) || (p.expected !== null) || (p.precondition !== null);
  if (hasAny) {
    html += '<div style="display:flex; gap:8px; align-items:center; margin-top:10px; padding-top:8px; border-top:1px solid #E5E7EB;">';
    if (p.precondition !== null) html += '<label style="font-size:11px;"><input type="checkbox" class="prop-field" data-idx="' + idx + '" data-field="precondition" checked> 사전조건</label>';
    if (p.steps !== null) html += '<label style="font-size:11px;"><input type="checkbox" class="prop-field" data-idx="' + idx + '" data-field="steps" checked> 테스트 스텝</label>';
    if (p.expected !== null) html += '<label style="font-size:11px;"><input type="checkbox" class="prop-field" data-idx="' + idx + '" data-field="expected" checked> 기대결과</label>';
    html += '<span style="flex:1;"></span>';
    html += '<button type="button" onclick="applyProposal(' + idx + ')" '
      + 'style="font-size:11px; padding:5px 12px; background:#1E40AF; color:#FFF; border:none; border-radius:4px; cursor:pointer; font-weight:600;">'
      + '✏️ 이 TC 만 사본에 적용</button>';
    html += '</div>';
    html += '<div id="proposal-status-' + idx + '" style="font-size:11px; color:#6B7280; margin-top:6px;"></div>';
  }

  return html;
}

async function applyProposal(idx) {
  // 어느 필드를 적용할지 — 체크박스 상태 수집
  const checkboxes = document.querySelectorAll('.prop-field[data-idx="' + idx + '"]');
  const fields = {};
  checkboxes.forEach(cb => {
    if (cb.checked) {
      // proposal 카드에서 텍스트 직접 추출하기보다 다시 propose 호출해 가져오는 게 깨끗
      // 다만 비용 절약을 위해 cache 활용 — 현재는 단순히 dataset 으로 처리
      fields[cb.dataset.field] = '__USE_LAST_PROPOSAL__';
    }
  });

  if (Object.keys(fields).length === 0) {
    alert('적용할 필드를 하나 이상 선택하세요.');
    return;
  }

  // 마지막 propose 결과를 다시 가져옴 (서버 캐시 X — 항상 fresh 한 게 안전)
  const statusEl = document.getElementById('proposal-status-' + idx);
  if (statusEl) statusEl.textContent = '⏳ 적용 준비 중 — 수정안 재확인...';

  let proposal;
  try {
    const r = await fetch('/update/propose', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ session_id: _updateSessionId, idx: idx }),
    });
    const d = await r.json();
    if (!d.ok) {
      if (statusEl) statusEl.innerHTML = '<span style="color:#DC2626;">❌ ' + (d.error || '오류') + '</span>';
      return;
    }
    proposal = d.proposal;
  } catch (e) {
    if (statusEl) statusEl.innerHTML = '<span style="color:#DC2626;">❌ ' + e.message + '</span>';
    return;
  }

  // 선택된 필드만 추출
  const accepted = { _rationale: proposal.rationale || '' };
  for (const f of Object.keys(fields)) {
    accepted[f] = proposal[f];
  }

  if (!confirm('선택한 필드를 사본 시트에 적용합니다. 진행할까요?\\n\\n사본만 수정하며 원본은 그대로입니다.')) {
    if (statusEl) statusEl.textContent = '취소됨.';
    return;
  }

  if (statusEl) statusEl.textContent = '⏳ 사본에 적용 중...';

  try {
    const r = await fetch('/update/commit', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ session_id: _updateSessionId, idx: idx, accepted: accepted }),
    });
    const d = await r.json();
    if (!d.ok) {
      if (statusEl) statusEl.innerHTML = '<span style="color:#DC2626;">❌ ' + (d.error || '오류') + '</span>';
      return;
    }
    const n = (d.written_cells || []).length;
    if (statusEl) statusEl.innerHTML = '<span style="color:#166534;">✅ 적용 완료 — ' + n + '개 셀 갱신 ('
      + escapeHtml(d.sheet_title) + ' 행 ' + d.row_index + '). TC Edit Log 에 기록됨.</span>';
  } catch (e) {
    if (statusEl) statusEl.innerHTML = '<span style="color:#DC2626;">❌ ' + e.message + '</span>';
  }
}

function checkAll(checked) {
  document.querySelectorAll('.cand-chk').forEach(cb => { cb.checked = checked; });
}

// 신규 TC 생성 모드로 이동 — TC 수정 모드의 add 후보에서 호출
// 1단계 (스마트 분석) 결과의 신규 SCR 그룹에서 호출 — 여러 신규 SCR 처리
function goToNewTcForAddedScrs() {
  const added = (_smartPlan?.scrs || []).filter(s => s.type === 'added').map(s => s.scr_id);
  if (!added.length) {
    alert('신규 SCR 이 없습니다.');
    return;
  }
  // 첫 번째 SCR ID 로 진입 (안내 메시지에 표시) + 전체 목록은 안내에 풀어줌
  goToNewTcMode(added[0], added);
}

function goToNewTcMode(scrId, addedList) {
  // 1. 새 기획서 폴더 (modify 모드의 입력) → spec 폴더 (new 모드의 입력) 으로 복사
  const newFolder = document.getElementById('updateNewFolder')?.value?.trim() || '';
  const specPathEl = document.getElementById('specFolderPath');
  if (specPathEl && newFolder) {
    specPathEl.value = newFolder;
  }

  // 2. 안내 토스트 (모드 전환 전에 한 번)
  if (typeof showToast === 'function') {
    showToast('✨ 신규 TC 생성 모드로 이동 — ' + scrId + ' (spec 폴더 자동 채워짐)');
  }

  // 3. 모드 전환 (existing TC 수정 → 신규 TC 생성)
  if (typeof switchMode === 'function') {
    switchMode('new');
  }

  // 3-1. 생성 모드를 '정책 반영 (summary)' 로 강제 리셋 — Combo 카드 같은 잔존 상태 정리
  //   (의도 = SCR 의 신규 TC 작성. Combo 가 아님)
  const summaryRadio = document.querySelector('input[name="genMode"][value="summary"]');
  if (summaryRadio) {
    summaryRadio.checked = true;
    if (typeof onGenModeChanged === 'function') onGenModeChanged();
  }

  // 4. 화면 상단으로 스크롤
  window.scrollTo({ top: 0, behavior: 'smooth' });

  // 5. 사용자에게 다음 액션 안내 (alert 대신 안내 박스 사용)
  setTimeout(function() {
    const panelNew = document.getElementById('panelNew');
    if (panelNew) {
      // 안내 박스 — 이미 있으면 갱신, 없으면 추가
      let helpBox = document.getElementById('newTcGuideForScr');
      if (!helpBox) {
        helpBox = document.createElement('div');
        helpBox.id = 'newTcGuideForScr';
        helpBox.style.cssText = 'margin:10px 0; padding:10px 14px; background:#F0FDF4; border:2px solid #86EFAC; border-radius:8px; font-size:12px; color:#166534; line-height:1.6;';
        panelNew.insertBefore(helpBox, panelNew.firstChild);
      }
      // 여러 SCR 처리 모드 (addedList 가 있으면) 또는 단일 SCR 모드
      const scrListHtml = (addedList && addedList.length)
        ? '<div style="margin:6px 0; padding:6px 10px; background:#FFF; border:1px solid #D1FAE5; border-radius:4px; font-family:monospace; font-size:11px;">'
          + addedList.map(escapeHtml).join(', ')
          + '</div>'
        : '';
      const titleHtml = (addedList && addedList.length > 1)
        ? '✨ <strong>' + addedList.length + '개 신규 SCR</strong> 의 TC 를 생성하려고 왔어요.'
        : '✨ <strong>' + escapeHtml(scrId) + '</strong> 의 신규 TC 를 생성하려고 왔어요.';
      helpBox.innerHTML = titleHtml + '<br>'
        + scrListHtml
        + '• spec 폴더는 자동 채워졌습니다.<br>'
        + '• 아래 진행 흐름대로 분석을 시작하세요.<br>'
        + '• Step 3 (검토) 단계에서 <strong>위 SCR 들만 체크</strong> → TC 생성으로 진행.<br>'
        + '<button onclick="document.getElementById(\\'newTcGuideForScr\\').remove()" style="margin-top:6px; padding:2px 8px; font-size:10px; background:#FFF; border:1px solid #166534; border-radius:3px; cursor:pointer;">알겠어요 — 안내 닫기</button>';
    }
  }, 300);
}

// SCR 단위 일괄 적용 — 해당 SCR 그룹의 체크된 modify 만 처리
async function applyBulkForScr(scrId) {
  if (!_updateSessionId) {
    alert('먼저 2단계 영향 분석을 실행하세요.');
    return;
  }
  if (!scrId) return;

  // 해당 SCR 그룹의 체크된 modify 후보 idx 만 수집
  const escSel = String(scrId).replace(/"/g, '\\\\"');
  const checkboxes = Array.from(document.querySelectorAll('.cand-chk[data-scr="' + escSel + '"]'));
  const selections = checkboxes.map(function(cb) {
    return { idx: parseInt(cb.dataset.idx, 10), accept: cb.checked };
  });
  const checkedModifyCount = selections.filter(function(s) {
    if (!s.accept) return false;
    const c = (_updateCandidates || [])[s.idx];
    return c && c.action === 'modify' && c.tc_id;
  }).length;

  if (checkedModifyCount === 0) {
    alert('이 SCR 그룹에 체크된 수정 후보가 없습니다.\\n\\nSCR: ' + scrId + '\\n수정 후보 (✏️ MODIFY) 항목을 하나 이상 체크하세요.');
    return;
  }

  const estimateSec = checkedModifyCount * 5;
  const estimateMin = Math.max(1, Math.ceil(estimateSec / 60));
  if (!confirm(
    '📦 SCR 일괄 적용 — ' + scrId + '\\n\\n'
    + checkedModifyCount + '개 수정 후보에 대해:\\n'
    + '  1. AI 가 각 TC 의 수정안 생성\\n'
    + '  2. 사본 시트의 셀을 실제로 갱신\\n'
    + '  3. TC Edit Log 에 이력 기록\\n\\n'
    + '⏱ 예상 시간: 약 ' + estimateMin + '분 (건당 ~5초)\\n'
    + '💰 AI 호출 ' + checkedModifyCount + '회 발생\\n\\n'
    + '진행할까요?'
  )) {
    return;
  }

  // 그룹 헤더의 버튼 + 그룹 헤더 옆에 진행 표시
  const groupBtn = document.querySelector('.btn-grp-apply[data-scr="' + escSel + '"]');
  if (groupBtn) {
    groupBtn.disabled = true;
    groupBtn.textContent = '⏳ 적용 중...';
    groupBtn.style.opacity = '0.6';
    groupBtn.style.cursor = 'not-allowed';
  }
  // 전체 일괄 적용 영역에도 SCR 단위 진행 표시 — 사용자가 결과를 한 곳에서 보도록
  const progressEl = document.getElementById('applyBulkProgress');
  const statusEl = document.getElementById('applyStatus');
  if (progressEl) {
    progressEl.classList.remove('hidden');
    progressEl.innerHTML = ''
      + '<div style="display:flex; align-items:center; gap:10px; padding:10px 12px; background:#FEF3C7; border:1px solid #FBBF24; border-radius:6px;">'
      +   '<div style="width:20px; height:20px; border:3px solid #FCD34D; border-top-color:#92400E; border-radius:50%; animation:plan-spin 0.8s linear infinite; flex-shrink:0;"></div>'
      +   '<div style="flex:1; font-size:12px; color:#78350F;">'
      +     '<strong>SCR 일괄 적용 중 — ' + escapeHtml(scrId) + '</strong> · ' + checkedModifyCount + '건 (약 ' + estimateMin + '분).<br>'
      +     '<span style="font-size:11px;">⚠ 페이지를 닫지 마세요. 완료 후 결과가 표시됩니다.</span>'
      +   '</div>'
      + '</div>';
  }
  if (statusEl) statusEl.innerHTML = '';

  try {
    const r = await fetch('/update/apply-bulk', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({
        session_id: _updateSessionId,
        selections: selections,
        skip_existing: true,
      }),
    });
    const d = await r.json();
    if (!d.ok) {
      if (progressEl) {
        progressEl.innerHTML = '<div style="padding:12px; background:#FEE2E2; color:#991B1B; border-radius:6px;">❌ ' + escapeHtml(d.error || '오류') + '</div>';
      }
      return;
    }
    // 결과 카드 — SCR 표기 추가
    d._scope_label = 'SCR ' + scrId;
    renderBulkResult(d);
  } catch (e) {
    if (progressEl) {
      progressEl.innerHTML = '<div style="padding:12px; background:#FEE2E2; color:#991B1B; border-radius:6px;">❌ 네트워크 오류: ' + escapeHtml(e.message) + '</div>';
    }
  } finally {
    if (groupBtn) {
      groupBtn.disabled = false;
      groupBtn.textContent = '📦 SCR 일괄 적용';
      groupBtn.style.opacity = '';
      groupBtn.style.cursor = 'pointer';
    }
  }
}

// 일괄 적용 — 체크된 modify 후보 전부에 AI 호출 + 사본 셀 갱신
async function applyBulk() {
  if (!_updateSessionId) {
    alert('먼저 2단계 영향 분석을 실행하세요.');
    return;
  }

  const selections = Array.from(document.querySelectorAll('.cand-chk')).map(function(cb) {
    return { idx: parseInt(cb.dataset.idx, 10), accept: cb.checked };
  });
  // 체크된 modify 후보만 카운트
  const checkedModifyCount = selections.filter(function(s) {
    if (!s.accept) return false;
    const c = (_updateCandidates || [])[s.idx];
    return c && c.action === 'modify' && c.tc_id;
  }).length;

  if (checkedModifyCount === 0) {
    alert('체크된 수정 후보가 없습니다.\\n\\n수정 후보 (✏️ MODIFY) 항목을 하나 이상 체크하세요.');
    return;
  }

  const estimateSec = checkedModifyCount * 5;
  const estimateMin = Math.ceil(estimateSec / 60);
  if (!confirm(
    '📦 일괄 적용\\n\\n'
    + checkedModifyCount + '개 수정 후보에 대해:\\n'
    + '  1. AI 가 각 TC 의 수정안 생성\\n'
    + '  2. 사본 시트의 셀을 실제로 갱신\\n'
    + '  3. TC Edit Log 에 이력 기록\\n\\n'
    + '⏱ 예상 시간: 약 ' + estimateMin + '분 (건당 ~5초)\\n'
    + '💰 AI 호출 ' + checkedModifyCount + '회 발생\\n\\n'
    + '진행할까요?'
  )) {
    return;
  }

  const btn = document.getElementById('applyBulkBtn');
  const progressEl = document.getElementById('applyBulkProgress');
  const statusEl = document.getElementById('applyStatus');
  btn.disabled = true;
  btn.textContent = '⏳ 일괄 적용 중...';
  progressEl.classList.remove('hidden');
  progressEl.innerHTML = ''
    + '<div style="display:flex; align-items:center; gap:10px; padding:10px 12px; background:#FEF3C7; border:1px solid #FBBF24; border-radius:6px;">'
    +   '<div style="width:20px; height:20px; border:3px solid #FCD34D; border-top-color:#92400E; border-radius:50%; animation:plan-spin 0.8s linear infinite; flex-shrink:0;"></div>'
    +   '<div style="flex:1; font-size:12px; color:#78350F;">'
    +     '<strong>일괄 적용 중...</strong> AI 호출 + 사본 셀 갱신 — ' + checkedModifyCount + '건 처리 (약 ' + estimateMin + '분 소요).<br>'
    +     '<span style="font-size:11px;">⚠ 페이지를 닫지 마세요. 완료 후 자동으로 결과가 표시됩니다.</span>'
    +   '</div>'
    + '</div>';
  statusEl.innerHTML = '';

  try {
    const r = await fetch('/update/apply-bulk', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({
        session_id: _updateSessionId,
        selections: selections,
        skip_existing: true,
      }),
    });
    const d = await r.json();
    if (!d.ok) {
      progressEl.innerHTML = '<div style="padding:12px; background:#FEE2E2; color:#991B1B; border-radius:6px;">❌ ' + escapeHtml(d.error || '오류') + '</div>';
      return;
    }
    // 결과 카드 렌더
    renderBulkResult(d);
  } catch (e) {
    progressEl.innerHTML = '<div style="padding:12px; background:#FEE2E2; color:#991B1B; border-radius:6px;">❌ 네트워크 오류: ' + escapeHtml(e.message) + '</div>';
  } finally {
    btn.disabled = false;
    btn.textContent = '📦 선택 항목 일괄 적용 (AI 호출 + 사본 셀 갱신)';
  }
}

function renderBulkResult(d) {
  const progressEl = document.getElementById('applyBulkProgress');
  const scopeText = d._scope_label ? ' (' + d._scope_label + ')' : '';
  let html = ''
    + '<div style="padding:14px 16px; background:linear-gradient(135deg, #F0FDF4 0%, #DCFCE7 100%); border:2px solid #16A34A; border-radius:8px;">'
    +   '<div style="font-size:14px; font-weight:700; color:#166534; margin-bottom:10px;">'
    +     '✅ 일괄 적용 완료' + escapeHtml(scopeText) + ' — ' + d.processed + '건 처리'
    +   '</div>'
    +   '<div style="display:flex; gap:10px; flex-wrap:wrap; font-size:12px; margin-bottom:12px;">'
    +     '<span style="background:#FFF; color:#166534; padding:4px 10px; border-radius:4px; border:1px solid #BBF7D0;">✏️ 적용 ' + d.applied + '건</span>'
    +     '<span style="background:#FFF; color:#1E40AF; padding:4px 10px; border-radius:4px; border:1px solid #BFDBFE;">✅ 변경 불필요 ' + d.no_change + '건</span>';
  if (d.skipped) html += '<span style="background:#FFF; color:#6B7280; padding:4px 10px; border-radius:4px; border:1px solid #E5E7EB;">⏭ 중복 skip ' + d.skipped + '건</span>';
  if (d.errors) html += '<span style="background:#FEE2E2; color:#991B1B; padding:4px 10px; border-radius:4px; border:1px solid #FCA5A5;">⚠ 오류 ' + d.errors + '건</span>';
  html += '</div>';

  if (d.copy_url) {
    html += '<div style="font-size:12px; color:#374151; margin-bottom:10px;">'
      + '📋 사본에서 확인: <a href="' + d.copy_url + '" target="_blank" style="color:#1E40AF;">사본 Sheets 열기 →</a>'
      + ' · 이력은 <strong>TC Edit Log</strong> 시트에 기록됨'
      + '</div>';
  }

  // 상세 토글
  html += '<details style="margin-top:8px;"><summary style="cursor:pointer; font-size:11px; color:#6B7280; padding:4px 0;">▸ 상세 결과 보기 (' + (d.details || []).length + '건)</summary>';
  html += '<div style="margin-top:6px; max-height:240px; overflow-y:auto; border:1px solid #E5E7EB; border-radius:6px; background:#FFF;">';
  html += '<table style="width:100%; font-size:11px; border-collapse:collapse;">';
  html += '<thead style="background:#F9FAFB; position:sticky; top:0;"><tr>'
    + '<th style="padding:5px 8px; text-align:left; border-bottom:1px solid #E5E7EB; width:130px;">TC ID</th>'
    + '<th style="padding:5px 8px; text-align:left; border-bottom:1px solid #E5E7EB; width:90px; white-space:nowrap;">상태</th>'
    + '<th style="padding:5px 8px; text-align:left; border-bottom:1px solid #E5E7EB;">필드/이유</th>'
    + '</tr></thead><tbody>';
  (d.details || []).forEach(function(item) {
    const stateLabel = {
      applied: '<span style="color:#166534;">✏️ 적용</span>',
      no_change: '<span style="color:#1E40AF;">✅ 불필요</span>',
      skipped: '<span style="color:#6B7280;">⏭ skip</span>',
      error: '<span style="color:#991B1B;">⚠ 오류</span>',
    }[item.status] || escapeHtml(item.status);
    const detail = item.fields_changed
      ? item.fields_changed.join(', ')
      : (item.rationale || item.reason || item.error || '');
    html += '<tr style="border-bottom:1px solid #F3F4F6;">'
      + '<td style="padding:5px 8px; font-family:monospace;">' + escapeHtml(item.tc_id || '') + '</td>'
      + '<td style="padding:5px 8px; white-space:nowrap;">' + stateLabel + '</td>'
      + '<td style="padding:5px 8px; color:#374151;">' + escapeHtml(detail).slice(0, 200) + '</td>'
      + '</tr>';
  });
  html += '</tbody></table></div></details>';
  html += '</div>';

  progressEl.innerHTML = html;

  if (typeof showToast === 'function') {
    showToast('✅ 일괄 적용 완료 — 적용 ' + d.applied + ' / 변경불필요 ' + d.no_change + ' / 오류 ' + d.errors);
  }
}

async function applyUpdates() {
  if (!_updateSessionId) {
    alert('먼저 2단계 영향 분석을 실행하세요.');
    return;
  }
  // 분석 진행 중이면 차단 (사용자가 로딩 못 봤어도 클릭 안 통하게)
  const applyBtn = document.getElementById('applyBtn');
  if (applyBtn && applyBtn.disabled) {
    alert(applyBtn.dataset.loadingMsg || '분석 진행 중입니다. 잠시 기다려주세요.');
    return;
  }
  const selections = Array.from(document.querySelectorAll('.cand-chk')).map(cb => ({
    idx: parseInt(cb.dataset.idx, 10),
    accept: cb.checked,
  }));
  const accepted = selections.filter(s => s.accept).length;
  if (accepted === 0) {
    if (!confirm('선택된 항목이 없습니다. 그래도 진행할까요?')) return;
  } else if (!confirm(accepted + '건을 사본 Sheets 의 Update Log 시트에 기록합니다. 진행할까요?')) {
    return;
  }

  const btn = document.getElementById('applyBtn');
  const statusEl = document.getElementById('applyStatus');
  btn.disabled = true;
  btn.textContent = '⏳ 적용 중...';
  statusEl.textContent = '사본 Sheets 에 기록 중...';

  try {
    const resp = await fetch('/update/apply', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ session_id: _updateSessionId, selections: selections }),
    });
    const data = await resp.json();
    if (!data.ok) {
      statusEl.innerHTML = '<span style="color:#DC2626;">❌ ' + (data.error || '오류') + '</span>';
      return;
    }
    statusEl.innerHTML = '✅ 적용 완료 — ' + data.applied_count + '건이 [' + escapeHtml(data.log_sheet || '') + '] 시트에 기록됨. '
      + (data.copy_url ? '<a href="' + data.copy_url + '" target="_blank" style="color:#1E40AF;">사본 열기 →</a>' : '');
  } catch (e) {
    statusEl.innerHTML = '<span style="color:#DC2626;">❌ ' + e.message + '</span>';
  } finally {
    btn.disabled = false;
    btn.textContent = '✅ 선택 항목 사본에 적용';
  }
}

// ── 사본 → 원본 적용 (Promote) + 롤백 ────────────────────────────────────
let _promotePlan = null;           // dry_run 결과 보관
let _lastBackupUrl = '';           // promote 완료 후 롤백 위해

async function startPromote() {
  // 중복 클릭 가드 — 이미 한 번 promote 했으면 명시 확인
  if (_lastBackupUrl) {
    if (!confirm(
      '⚠️ 이미 원본에 반영했습니다.\\n\\n'
      + '다시 누르면 다음이 발생합니다:\\n'
      + '  • 새 백업 사본이 또 만들어짐 (Drive 용량 낭비)\\n'
      + '  • Promote Log 시트에 같은 행이 중복 추가\\n'
      + '  • 롤백 기준이 새 백업으로 바뀜 (이전 백업 추적 어려움)\\n\\n'
      + '💡 권장: 사본 셀을 더 수정한 뒤에 다시 promote 하세요.\\n'
      + '그래도 진행할까요?'
    )) {
      return;
    }
  }

  const copyUrl = (window._lastPlanCopyUrl || '').trim() ||
                    (document.querySelector('#planCopyLink a')?.href || '').trim();
  const sourceUrl = document.getElementById('updateTcUrl').value.trim();
  if (!copyUrl) {
    alert('사본 URL 을 찾을 수 없어요. 2단계 분석을 먼저 실행하세요.');
    return;
  }
  if (!sourceUrl) {
    alert('원본 Sheets URL 이 비어있습니다.');
    return;
  }

  const modal = document.getElementById('promoteModal');
  const body = document.getElementById('promoteModalBody');
  const confirmBtn = document.getElementById('promoteConfirmBtn');
  modal.classList.remove('hidden');
  confirmBtn.disabled = true;
  body.innerHTML = ''
    + '<div style="display:flex; align-items:center; gap:12px; padding:14px;">'
    +   '<div style="width:20px; height:20px; border:3px solid #FDE68A; border-top-color:#92400E; border-radius:50%; animation:plan-spin 0.8s linear infinite;"></div>'
    +   '<div>'
    +     '<div style="font-size:13px; font-weight:600; color:#78350F;">⏳ 미리보기 산출 중</div>'
    +     '<div style="font-size:11px; color:#92400E; margin-top:3px;">사본의 TC Edit Log → 원본 셀 좌표 매칭</div>'
    +   '</div>'
    + '</div>';

  try {
    const r = await fetch('/update/promote', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({
        copy_url: copyUrl,
        source_url: sourceUrl,
        dry_run: true,
      }),
    });
    const d = await r.json();
    if (!d.ok) {
      body.innerHTML = '<div style="padding:12px; background:#FEE2E2; color:#991B1B; border-radius:6px;">❌ ' + escapeHtml(d.error || '오류') + '</div>';
      return;
    }
    _promotePlan = d;
    renderPromotePreview(d, copyUrl, sourceUrl);
    confirmBtn.disabled = false;
  } catch (e) {
    body.innerHTML = '<div style="padding:12px; background:#FEE2E2; color:#991B1B; border-radius:6px;">❌ ' + escapeHtml(e.message) + '</div>';
  }
}

function renderPromotePreview(d, copyUrl, sourceUrl) {
  const body = document.getElementById('promoteModalBody');
  const plan = d.plan || [];
  const missing = d.missing || [];

  let html = '';

  // 갱신 대상 URL 강조 — 사용자가 무엇을 변경하는지 명확히
  html += '<div style="padding:10px 14px; background:#FEE2E2; border:2px solid #DC2626; border-radius:6px; margin-bottom:14px; font-size:13px; color:#7F1D1D;">'
    + '🎯 <strong>갱신 대상 Sheets:</strong><br>'
    + '<a href="' + (sourceUrl || '#') + '" target="_blank" style="color:#991B1B; font-family:monospace; font-size:11px; word-break:break-all;">' + escapeHtml(sourceUrl || '(없음)') + '</a>'
    + '<br><span style="font-size:11px;">이 시트의 셀이 실제로 변경됩니다. 백업이 자동 생성되니 롤백 가능합니다.</span>'
    + '</div>';

  // 요약
  html += '<div style="display:flex; gap:10px; flex-wrap:wrap; margin-bottom:14px; font-size:12px;">'
    + '<span style="background:#DCFCE7; color:#166534; padding:4px 10px; border-radius:4px;">📋 적용 대상 ' + plan.length + '개 셀</span>'
    + '<span style="background:#F3F4F6; color:#374151; padding:4px 10px; border-radius:4px;">📄 TC Edit Log 총 ' + d.log_count + '건</span>'
    + (missing.length ? '<span style="background:#FEE2E2; color:#991B1B; padding:4px 10px; border-radius:4px;">⚠️ 매칭 실패 ' + missing.length + '건</span>' : '')
    + '<span style="background:#FEF3C7; color:#92400E; padding:4px 10px; border-radius:4px;">📁 영향 시트: ' + (d.sheets_seen || []).join(', ') + '</span>'
    + '</div>';

  // 안내
  html += '<div style="padding:10px 12px; background:#FEF3C7; border:1px solid #FBBF24; border-radius:6px; margin-bottom:14px; font-size:12px; color:#78350F; line-height:1.6;">'
    + '⚠️ <strong>적용 시 자동 진행:</strong><br>'
    + '&nbsp;&nbsp;1. 원본 Sheets 를 <code>backup before update YYYY-MM-DD HH:MM</code> 이름으로 사본 복사 (자동 백업)<br>'
    + '&nbsp;&nbsp;2. ' + plan.length + '개 셀을 원본에 갱신 (batchUpdate)<br>'
    + '&nbsp;&nbsp;3. 원본에 <code>Promote Log</code> 시트 생성/추가 (적용 이력)<br>'
    + '&nbsp;&nbsp;4. 적용 후 <strong>🔄 마지막 적용 롤백</strong> 버튼 활성화'
    + '</div>';

  // 적용 셀 리스트
  if (plan.length) {
    html += '<div style="font-size:13px; font-weight:600; color:#1E3A5F; margin-bottom:6px;">📝 적용 셀 (앞 20개)</div>';
    html += '<div style="max-height:280px; overflow-y:auto; border:1px solid #E5E7EB; border-radius:6px; margin-bottom:14px;">';
    html += '<table style="width:100%; font-size:11px; border-collapse:collapse;">';
    html += '<thead style="background:#F9FAFB; position:sticky; top:0;"><tr>'
      + '<th style="padding:6px 8px; text-align:left; border-bottom:1px solid #E5E7EB;">TC ID</th>'
      + '<th style="padding:6px 8px; text-align:left; border-bottom:1px solid #E5E7EB;">필드</th>'
      + '<th style="padding:6px 8px; text-align:left; border-bottom:1px solid #E5E7EB;">원본 셀</th>'
      + '<th style="padding:6px 8px; text-align:left; border-bottom:1px solid #E5E7EB;">새 값 (앞 100자)</th>'
      + '</tr></thead><tbody>';
    plan.slice(0, 20).forEach(function(p) {
      html += '<tr style="border-bottom:1px solid #F3F4F6;">'
        + '<td style="padding:5px 8px; font-family:monospace;">' + escapeHtml(p.tc_id) + '</td>'
        + '<td style="padding:5px 8px;">' + escapeHtml(p.field) + '</td>'
        + '<td style="padding:5px 8px; font-family:monospace; color:#1E40AF;">' + escapeHtml(p.source_cell) + '</td>'
        + '<td style="padding:5px 8px; color:#374151;">' + escapeHtml((p.new_value || '').slice(0, 100)) + (p.new_value && p.new_value.length > 100 ? '…' : '') + '</td>'
        + '</tr>';
    });
    html += '</tbody></table>';
    if (plan.length > 20) {
      html += '<div style="padding:8px; background:#F9FAFB; font-size:11px; color:#6B7280; text-align:center;">… 외 ' + (plan.length - 20) + '건</div>';
    }
    html += '</div>';
  }

  // 매칭 실패
  if (missing.length) {
    html += '<div style="font-size:13px; font-weight:600; color:#991B1B; margin-bottom:6px;">⚠️ 매칭 실패 (' + missing.length + '건)</div>';
    html += '<div style="background:#FEF2F2; border:1px solid #FCA5A5; border-radius:6px; padding:8px 12px; font-size:11px; color:#7F1D1D; max-height:150px; overflow-y:auto;">';
    missing.forEach(function(m) {
      html += '<div>• ' + escapeHtml(m.tc_id) + ' [' + escapeHtml(m.sheet_title) + '] — ' + escapeHtml(m.reason) + '</div>';
    });
    html += '</div>';
  }

  body.innerHTML = html;
}

function closePromoteModal() {
  document.getElementById('promoteModal').classList.add('hidden');
}

async function confirmPromote() {
  if (!_promotePlan) return;
  const copyUrl = (window._lastPlanCopyUrl || '').trim() ||
                    (document.querySelector('#planCopyLink a')?.href || '').trim();
  const sourceUrl = document.getElementById('updateTcUrl').value.trim();
  const confirmBtn = document.getElementById('promoteConfirmBtn');
  const body = document.getElementById('promoteModalBody');

  if (!confirm('원본 Sheets 에 ' + _promotePlan.count + '개 셀을 갱신합니다.\\n\\n원본은 자동 백업되며, 적용 후 롤백 가능합니다.\\n\\n진행할까요?')) {
    return;
  }

  confirmBtn.disabled = true;
  confirmBtn.textContent = '⏳ 백업 + 적용 중...';

  try {
    const r = await fetch('/update/promote', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({
        copy_url: copyUrl,
        source_url: sourceUrl,
        dry_run: false,
      }),
    });
    const d = await r.json();
    if (!d.ok) {
      body.innerHTML = '<div style="padding:12px; background:#FEE2E2; color:#991B1B; border-radius:6px;">❌ ' + escapeHtml(d.error || '오류') + '</div>';
      confirmBtn.disabled = false;
      confirmBtn.textContent = '✏️ 원본에 적용 (백업 + 갱신)';
      return;
    }
    _lastBackupUrl = d.backup?.backup_url || '';
    // 모달 닫고 상태 갱신
    closePromoteModal();
    const statusEl = document.getElementById('promoteStatus');
    // 갱신된 원본 (= source_url) 링크도 함께 표시 — 사용자가 결과물 바로 확인 가능
    const updatedSourceUrl = d.source_url || sourceUrl || '';
    statusEl.innerHTML = '✅ 적용 완료 — <strong>' + d.applied_count + '개 셀</strong> 갱신됨'
      + ' · 적용 시각 ' + escapeHtml(d.applied_at)
      + '<br><span style="font-size:11px;">'
      + (updatedSourceUrl
          ? '📄 <a href="' + updatedSourceUrl + '" target="_blank" style="color:#166534; font-weight:600;">갱신된 원본 열기 →</a> · '
          : '')
      + '<a href="' + d.backup.backup_url + '" target="_blank" style="color:#1E40AF;">💾 백업 사본 열기</a>'
      + '</span>';
    statusEl.style.color = '#166534';

    // promote 버튼 시각 갱신 — '이미 적용됨' 표시 (회색) + 라벨 변경
    const promoteBtn = document.getElementById('promoteBtn');
    if (promoteBtn) {
      promoteBtn.textContent = '✓ 이미 적용됨 — 다시 적용하려면 클릭';
      promoteBtn.style.background = '#9CA3AF';
      promoteBtn.style.opacity = '0.7';
      promoteBtn.title = '주의: 다시 누르면 새 백업 + Promote Log 중복';
    }

    // 롤백 버튼 노출
    const rbBtn = document.getElementById('rollbackBtn');
    rbBtn.style.display = 'inline-block';
    rbBtn.dataset.backupUrl = d.backup.backup_url;

    if (typeof showToast === 'function') {
      showToast('✅ 원본 갱신 완료 — ' + d.applied_count + '개 셀');
    }
  } catch (e) {
    body.innerHTML = '<div style="padding:12px; background:#FEE2E2; color:#991B1B; border-radius:6px;">❌ ' + escapeHtml(e.message) + '</div>';
    confirmBtn.disabled = false;
    confirmBtn.textContent = '✏️ 원본에 적용 (백업 + 갱신)';
  }
}

async function startRollback() {
  const rbBtn = document.getElementById('rollbackBtn');
  const backupUrl = rbBtn.dataset.backupUrl || _lastBackupUrl;
  const sourceUrl = document.getElementById('updateTcUrl').value.trim();
  if (!backupUrl) {
    alert('백업 URL 이 없습니다. 먼저 promote 를 실행하세요.');
    return;
  }
  // 중복 클릭 가드 — 이미 한 번 롤백했으면 명시 확인
  if (rbBtn.dataset.rolledBack === '1') {
    if (!confirm(
      '⚠️ 이미 롤백했습니다.\\n\\n'
      + '같은 백업 기준으로 다시 롤백을 시도하면 같은 셀에 같은 값이 한 번 더 쓰입니다 (idempotent).\\n'
      + '대부분 의미 없습니다.\\n\\n'
      + '그래도 진행할까요?'
    )) {
      return;
    }
  }
  if (!confirm('마지막 적용을 롤백합니다.\\n\\n원본의 Promote Log 에 기록된 셀들이 백업 시점의 값으로 복원됩니다.\\n진행할까요?')) {
    return;
  }

  rbBtn.disabled = true;
  rbBtn.textContent = '⏳ 롤백 중...';

  try {
    const r = await fetch('/update/rollback', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({
        source_url: sourceUrl,
        backup_url: backupUrl,
        scope: 'promote_log',
      }),
    });
    const d = await r.json();
    if (!d.ok) {
      alert('롤백 실패: ' + (d.error || '오류'));
      rbBtn.disabled = false;
      rbBtn.textContent = '🔄 마지막 적용 롤백';
      return;
    }
    const statusEl = document.getElementById('promoteStatus');
    statusEl.innerHTML = '🔄 롤백 완료 — <strong>' + d.restored_count + '개 셀</strong> 백업 시점으로 복원됨 · ' + escapeHtml(d.restored_at || '');
    statusEl.style.color = '#1E40AF';

    // 롤백 버튼 시각 갱신 — '이미 롤백됨' 표시 (회색)
    rbBtn.dataset.rolledBack = '1';
    rbBtn.style.background = '#F3F4F6';
    rbBtn.style.color = '#6B7280';
    rbBtn.style.borderColor = '#D1D5DB';
    rbBtn.title = '이미 롤백됨 — 다시 누르면 같은 값으로 idempotent 복원';

    if (typeof showToast === 'function') {
      showToast('🔄 롤백 완료 — ' + d.restored_count + '개 셀');
    }
  } catch (e) {
    alert('롤백 네트워크 오류: ' + e.message);
  } finally {
    rbBtn.disabled = false;
    rbBtn.textContent = '🔄 마지막 적용 롤백';
  }
}

// 섹션(add/mod/rem/all)별 체크박스 일괄 선택/해제
function toggleSectionSelection(section, checked) {
  const selectors = {
    add: '.diff-add-chk',
    mod: '.diff-mod-chk',
    rem: '.diff-rem-chk',
    all: '.diff-add-chk, .diff-mod-chk, .diff-rem-chk',
  };
  const sel = selectors[section];
  if (!sel) return;
  document.querySelectorAll(sel).forEach(cb => {
    cb.checked = checked;
  });
  _updateApproveButtonState();
}

// 라인 단위 diff를 GitHub 스타일 unified diff로 렌더
function _renderUnifiedDiff(blocks) {
  if (!blocks || !blocks.length) {
    return '<div style="font-size:12px;color:var(--muted);padding:8px;">변경 정보가 없습니다.</div>';
  }
  const lineStyle = 'font-family:ui-monospace,SFMono-Regular,Menlo,monospace;font-size:11px;padding:2px 8px;white-space:pre-wrap;word-break:break-word;border-left:3px solid transparent;';
  let html = '<div style="max-height:360px;overflow:auto;background:#F9FAFB;border:1px solid #E5E7EB;border-radius:4px;">';
  for (const blk of blocks) {
    if (blk.tag === 'truncated') {
      html += `<div style="${lineStyle}color:var(--muted);font-style:italic;padding:4px 8px;">... (이후 내용은 생략)</div>`;
      continue;
    }
    if (blk.tag === 'equal') {
      for (const ln of blk.old_lines) {
        const txt = ln === '... (중략)' ? '  …' : '  ' + ln;
        html += `<div style="${lineStyle}color:#6B7280;">${_escapeHtml(txt)}</div>`;
      }
      continue;
    }
    if (blk.tag === 'delete' || blk.tag === 'replace') {
      for (const ln of blk.old_lines) {
        html += `<div style="${lineStyle}background:#FEE2E2;border-left-color:#DC2626;color:#991B1B;">− ${_escapeHtml(ln)}</div>`;
      }
    }
    if (blk.tag === 'insert' || blk.tag === 'replace') {
      for (const ln of blk.new_lines) {
        html += `<div style="${lineStyle}background:#DCFCE7;border-left-color:#16A34A;color:#166534;">+ ${_escapeHtml(ln)}</div>`;
      }
    }
  }
  html += '</div>';
  return html;
}

function switchDiffMode(detailsId, mode) {
  const container = document.getElementById(detailsId);
  if (!container) return;
  const idxMatch = detailsId.match(/\d+$/);
  if (!idxMatch) return;
  const idx = idxMatch[0];
  const unified = container.querySelector('.diff-view-unified-' + idx);
  const side    = container.querySelector('.diff-view-sidebyside-' + idx);
  const btnU    = container.querySelector('.diff-mode-unified-' + idx);
  const btnS    = container.querySelector('.diff-mode-sidebyside-' + idx);
  if (mode === 'unified') {
    if (unified) unified.style.display = 'block';
    if (side)    side.style.display = 'none';
    if (btnU) { btnU.style.background = '#FDE68A'; btnU.style.borderColor = '#F59E0B'; btnU.style.fontWeight = '600'; }
    if (btnS) { btnS.style.background = '#FFFFFF'; btnS.style.borderColor = '#D1D5DB'; btnS.style.fontWeight = 'normal'; }
  } else {
    if (unified) unified.style.display = 'none';
    if (side)    side.style.display = 'block';
    if (btnU) { btnU.style.background = '#FFFFFF'; btnU.style.borderColor = '#D1D5DB'; btnU.style.fontWeight = 'normal'; }
    if (btnS) { btnS.style.background = '#FDE68A'; btnS.style.borderColor = '#F59E0B'; btnS.style.fontWeight = '600'; }
  }
}

// 수정 화면 카드의 변경 내용 펼침/접힘
function toggleDiffDetail(detailsId, btn) {
  const el = document.getElementById(detailsId);
  if (!el) return;
  const opened = el.style.display !== 'none';
  el.style.display = opened ? 'none' : 'block';
  if (btn) {
    const base = btn.textContent.replace(/^[▸▾]\s*/, '');
    btn.textContent = (opened ? '▸ ' : '▾ ') + base;
  }
}

function toggleAllDiffDetails(open) {
  document.querySelectorAll('[id^="diff-mod-"], [id^="diff-add-"]').forEach(el => {
    if (!el.id.match(/^diff-(mod|add)-\d+$/)) return;
    el.style.display = open ? 'block' : 'none';
  });
  // 버튼 라벨 동기화 — 각 카드의 "▸/▾" 표시
  document.querySelectorAll('button').forEach(btn => {
    const t = btn.textContent || '';
    if (t.includes('변경 보기') || t.endsWith('내용')) {
      btn.textContent = (open ? '▾ ' : '▸ ') + t.replace(/^[▸▾]\s*/, '');
    }
  });
}

function _getApprovedSelection() {
  return {
    added:    [...document.querySelectorAll('.diff-add-chk:checked')].map(el => el.value),
    modified: [...document.querySelectorAll('.diff-mod-chk:checked')].map(el => el.value),
    removed:  [...document.querySelectorAll('.diff-rem-chk:checked')].map(el => el.value),
  };
}

function _updateApproveButtonState() {
  const btn = document.querySelector('#diffReportArea .btn-primary');
  if (!btn) return;
  const sel = _getApprovedSelection();
  const total = sel.added.length + sel.modified.length + sel.removed.length;
  if (total === 0) {
    btn.disabled = true;
    btn.style.opacity = '0.5';
    btn.style.cursor = 'not-allowed';
    btn.textContent = '⚠️ 승인할 항목을 1개 이상 선택하세요';
  } else {
    btn.disabled = false;
    btn.style.opacity = '1';
    btn.style.cursor = 'pointer';
    btn.textContent = `✅ 승인 후 TC 갱신 시작 (${total}개)`;
  }
}

async function approveAndUpdate() {
  if (!_diffSid) { alert('분석 세션이 없습니다. 먼저 분석을 실행하세요.'); return; }
  const approved = _getApprovedSelection();
  if (approved.added.length + approved.modified.length + approved.removed.length === 0) {
    alert('승인할 항목을 최소 1개 이상 선택하세요.');
    return;
  }
  document.getElementById('card1').classList.add('hidden');
  document.getElementById('card2').classList.remove('hidden');
  setStepBar(2);
  document.getElementById('card2').scrollIntoView({ behavior: 'smooth', block: 'start' });
  document.getElementById('stageLabel').textContent = 'TC 갱신 시작 중...';

  try {
    const r = await fetch('/update-tc', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ sid: _diffSid, approved: approved })
    });
    const d = await r.json();
    if (!d.ok) { alert('오류: ' + d.error); return; }
    currentSid = d.sid;
    connectStream(d.sid);
  } catch(e) {
    alert('네트워크 오류: ' + e.message);
  }
}

// TC 수정 시작 (기존 — 레거시, 현재 UI에서는 사용 안 함)
async function startModify() {
  const projectName = document.getElementById('projectSelect').value;
  const changeDesc  = document.getElementById('changeDesc').value.trim();

  if (!projectName) { alert('수정할 프로젝트를 선택하세요.'); return; }
  if (!changeDesc)   { alert('변경사항 내용을 입력하세요.'); return; }

  document.getElementById('startModifyBtn').disabled = true;
  document.getElementById('card1').classList.add('hidden');
  document.getElementById('card2').classList.remove('hidden');
  setStepBar(2);
  document.getElementById('card2').scrollIntoView({ behavior: 'smooth', block: 'start' });
  document.getElementById('stageLabel').textContent = '수정 파이프라인 시작 중...';

  try {
    const r = await fetch('/start-modify', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ project_name: projectName, change_desc: changeDesc })
    });
    const d = await r.json();
    if (!d.ok) {
      alert('오류: ' + d.error);
      document.getElementById('startModifyBtn').disabled = false;
      return;
    }
    currentSid = d.sid;
    connectStream(d.sid);
  } catch(e) {
    alert('오류: ' + e.message);
    document.getElementById('startModifyBtn').disabled = false;
  }
}

// ── 입력 유형 전환 ──────────────────────────────────────────────────────────────
// ── GitHub 파일 트리 미리보기 ─────────────────────────────────────────────────
async function previewGithubTree(srcId) {
  const urlInput = document.getElementById('srcUrl_' + srcId);
  const panel = document.getElementById('treePanel_' + srcId);
  const url = urlInput ? urlInput.value.trim() : '';
  if (!url) { alert('GitHub URL을 먼저 입력해주세요.'); return; }

  panel.classList.remove('hidden');
  panel.innerHTML = '<div class="tree-loading">⏳ 파일 목록 불러오는 중...</div>';

  try {
    const resp = await fetch('/github-tree', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ url })
    });
    const data = await resp.json();
    if (!data.ok) {
      panel.innerHTML = '<div class="tree-error">❌ ' + data.error + '</div>';
      return;
    }

    const files = data.files;
    // 폴더별 그룹핑
    const groups = {};
    files.forEach(f => {
      const parts = f.path.split('/');
      const folder = parts.length > 1 ? parts[0] : '(루트)';
      if (!groups[folder]) groups[folder] = [];
      groups[folder].push(f);
    });

    let html = '<div class="tree-header">';
    html += '<strong>' + data.owner + '/' + data.repo + '</strong> (branch: ' + data.branch + ') — 총 ' + files.length + '개 파일';
    html += '<div class="tree-actions">';
    html += '<button type="button" onclick="selectAllTree(' + srcId + ', true)">전체 선택</button> ';
    html += '<button type="button" onclick="selectAllTree(' + srcId + ', false)">전체 해제</button> ';
    html += '<button type="button" class="btn-apply-tree" onclick="applyTreeSelection(' + srcId + ')">✅ 선택 적용</button>';
    html += '</div></div>';
    html += '<div class="tree-list">';

    Object.entries(groups).sort().forEach(([folder, flist]) => {
      html += '<div class="tree-folder">';
      html += '<label class="tree-folder-label"><input type="checkbox" class="folder-cb" data-srcid="' + srcId + '" data-folder="' + folder + '" onchange="toggleFolder(this)"> 📁 ' + folder + ' (' + flist.length + ')</label>';
      flist.forEach(f => {
        const sizeKb = f.size > 0 ? ' <span class="file-size">(' + Math.ceil(f.size/1024) + 'KB)</span>' : '';
        const ext = f.path.split('.').pop().toLowerCase();
        const icon = ext === 'md' ? '📝' : ext === 'pdf' ? '📄' : ext === 'json' ? '{}' : '📄';
        html += '<label class="tree-file-label"><input type="checkbox" class="file-cb" data-srcid="' + srcId + '" data-folder="' + folder + '" value="' + f.path + '"> ' + icon + ' ' + f.path.split('/').pop() + sizeKb + '</label>';
      });
      html += '</div>';
    });

    html += '</div>';
    panel.innerHTML = html;
  } catch(e) {
    panel.innerHTML = '<div class="tree-error">❌ 오류: ' + e.message + '</div>';
  }
}

function toggleFolder(cb) {
  const srcId = cb.dataset.srcid;
  const folder = cb.dataset.folder;
  document.querySelectorAll('.file-cb[data-srcid="' + srcId + '"][data-folder="' + folder + '"]')
    .forEach(fc => fc.checked = cb.checked);
}

function selectAllTree(srcId, checked) {
  document.querySelectorAll('.file-cb[data-srcid="' + srcId + '"], .folder-cb[data-srcid="' + srcId + '"]')
    .forEach(cb => cb.checked = checked);
}

function applyTreeSelection(srcId) {
  const selected = [];
  document.querySelectorAll('.file-cb[data-srcid="' + srcId + '"]:checked')
    .forEach(cb => selected.push(cb.value));

  if (selected.length === 0) { alert('최소 1개 파일을 선택해주세요.'); return; }

  const src = sources.find(s => s.id === srcId);
  if (src) src.selected_files = selected;

  const panel = document.getElementById('treePanel_' + srcId);
  const applyBtn = panel.querySelector('.btn-apply-tree');
  if (applyBtn) applyBtn.textContent = '✅ ' + selected.length + '개 파일 선택됨';
  applyBtn.style.background = 'var(--teal)';
  applyBtn.style.color = '#fff';
}

// ── 다중 소스 관리 ──────────────────────────────────────────────────────────
let sources = [];
let sourceCounter = 0;

// ── 입력 변경 감지 ──────────────────────────────────────────────
// 사용자가 소스(추가/삭제/수정) 또는 TC 생성 범위를 변경하면 호출.
// 이전 세션의 결과 카드를 숨기고 startBtn을 활성화해 "다시 시작 가능" 상태로.
// 진행 중 세션(currentSid)이 있으면 해당 카드는 유지 (사용자 실수로 실행 중단 방지).
function onInputsChanged() {
  // 1. 새 파이프라인 시작 가능하도록 버튼 활성
  var startBtn = document.getElementById('startBtn');
  if (startBtn) startBtn.disabled = false;
  var startModifyBtn = document.getElementById('startModifyBtn');
  if (startModifyBtn) startModifyBtn.disabled = false;

  // 2. 이전 결과 카드는 항상 무효화 (입력이 바뀌면 이전 결과는 stale)
  var card5 = document.getElementById('card5');
  if (card5) card5.classList.add('hidden');

  // 3. 실행 중이 아니면 진행/Gate/작성 카드도 숨김 + Step 1 입력 카드 복원
  if (!currentSid) {
    ['card2','card3'].forEach(function(id) {
      var card = document.getElementById(id);
      if (card) card.classList.add('hidden');
    });
    document.getElementById('card1').classList.remove('hidden');
    document.querySelectorAll('.stopped-banner, .error-banner').forEach(function(el) { el.remove(); });
    setStepBar(1);
  }

  // 4. 생성 모드 안내 갱신 (소스 크기 기반 권장 모드 제시)
  refreshGenerationModeHint();
}

// ── 생성 모드 관련 ──────────────────────────────────────────────
// 정책 반영 모드(summary): policy → features → classify 3단계
// Quick 모드(direct): 원문을 직접 분류 — 누락 최소화, 200KB 이하 권장
const DIRECT_MODE_SIZE_LIMIT = 200 * 1024; // 200KB

function estimateSourceBytes() {
  // 각 소스의 content 크기 합산 (UTF-8 기준 대략)
  let total = 0;
  for (const s of sources) {
    if (!s || !s.content) continue;
    try { total += new Blob([s.content]).size; } catch(e) { total += (s.content || '').length; }
  }
  return total;
}

function formatBytes(n) {
  if (n <= 0) return '0 B';
  if (n < 1024) return n + ' B';
  if (n < 1024 * 1024) return (n / 1024).toFixed(1) + ' KB';
  return (n / 1024 / 1024).toFixed(2) + ' MB';
}

function refreshGenerationModeHint() {
  const sizeEl = document.getElementById('modeSourceSize');
  const statusEl = document.getElementById('modeSourceStatus');
  const quickLabel = document.getElementById('modeQuickLabel');
  const quickRadio = document.querySelector('input[name="genMode"][value="direct"]');
  if (!sizeEl || !statusEl) return;

  const bytes = estimateSourceBytes();
  sizeEl.textContent = bytes > 0 ? formatBytes(bytes) : '-';

  if (bytes === 0) {
    statusEl.textContent = '소스를 추가하면 권장 모드가 안내됩니다.';
    statusEl.style.color = '#6B7280';
    if (quickLabel) quickLabel.style.opacity = '1';
    if (quickRadio) quickRadio.disabled = false;
    return;
  }

  if (bytes <= DIRECT_MODE_SIZE_LIMIT) {
    statusEl.innerHTML = '🟢 Quick 모드 사용 가능 (원문 직접 분류 — 누락 최소화)';
    statusEl.style.color = '#047857';
    if (quickLabel) quickLabel.style.opacity = '1';
    if (quickRadio) quickRadio.disabled = false;
  } else {
    statusEl.innerHTML = '🟡 소스가 200KB를 초과합니다 — 정책 반영 모드 권장';
    statusEl.style.color = '#92400E';
    if (quickLabel) quickLabel.style.opacity = '0.55';
    if (quickRadio) {
      quickRadio.disabled = true;
      // 비활성 시 기본 모드로 복귀
      if (quickRadio.checked) {
        const stdRadio = document.querySelector('input[name="genMode"][value="summary"]');
        if (stdRadio) stdRadio.checked = true;
      }
    }
  }
}

function onGenModeChanged() {
  // 선택된 라디오 카드에 outline 강조
  document.querySelectorAll('input[name="genMode"]').forEach(function(r) {
    const box = r.closest('label');
    if (!box) return;
    if (r.checked) {
      box.style.borderColor = '#2563EB';
      box.style.background = '#EFF6FF';
    } else {
      box.style.borderColor = '#D1D5DB';
      box.style.background = '#FFFFFF';
    }
  });

  // Combo 카드 토글 — 'combo_only' 모드일 때만 표시
  const isCombo = getSelectedGenerationMode() === 'combo_only';
  const comboGroup = document.getElementById('comboGroup');
  if (comboGroup) {
    comboGroup.style.display = isCombo ? 'block' : 'none';
    if (isCombo) loadComboFiles();
  }
}

// Combo 명세 파일 목록 로드
async function loadComboFiles() {
  const listEl = document.getElementById('comboFileList');
  if (!listEl) return;
  const projectName = (document.getElementById('projectSelect')?.value || 'supercycl').trim();
  const specPath = document.getElementById('specFolderPath')?.value?.trim() || '';

  listEl.innerHTML = '<em style="color:#6B7280;">파일 목록 조회 중...</em>';
  try {
    const url = '/combo/list-files?project_name=' + encodeURIComponent(projectName) +
                (specPath ? '&spec_folder=' + encodeURIComponent(specPath) : '');
    const resp = await fetch(url);
    const data = await resp.json();
    if (!data.ok) {
      listEl.innerHTML = '<span style="color:#DC2626;">⚠ ' + (data.error || '오류') + '</span>';
      return;
    }
    if (!data.files || !data.files.length) {
      listEl.innerHTML = '<span style="color:#DC2626;">⚠ <strong>*_combinations.md</strong> 파일이 없습니다.' +
        '<br>projects/' + projectName + '/ 또는 spec 폴더에 추가하세요.</span>';
      return;
    }
    // 중복 도메인 감지 (같은 domain 라벨이 2개 이상) — 행에 강조 표시용
    const domainCounts = {};
    data.files.forEach(function(f) { domainCounts[f.domain] = (domainCounts[f.domain] || 0) + 1; });

    listEl.innerHTML = data.files.map(function(f, i) {
      const isDup = domainCounts[f.domain] > 1;
      const rowStyle = isDup
        ? 'background:#FEF3C7;border-left:3px solid #F59E0B;padding-left:6px;'
        : '';
      const dupTag = isDup
        ? '<span style="font-size:10px;color:#92400E;background:#FCD34D;padding:1px 5px;border-radius:3px;margin-left:4px;">중복</span>'
        : '';
      return '<div style="display:flex;align-items:center;gap:10px;padding:6px 4px;border-radius:3px;' + rowStyle + '">' +
        '<label style="display:flex;align-items:center;gap:10px;cursor:pointer;flex:1;">' +
          '<input type="checkbox" class="combo-file-cb" value="' + f.path + '" data-scenarios="' + encodeURIComponent(JSON.stringify(f.scenarios || [])) + '" data-domain="' + f.domain + '" checked onchange="renderComboScenarios()">' +
          '<span><strong>' + f.name + '</strong> → <code style="background:#F3F4F6;padding:1px 6px;border-radius:3px;">' + f.domain + '</code>' + dupTag + '</span>' +
        '</label>' +
        '<span style="color:#6B7280;font-size:11px;">OC ' + f.combos_count + ' · 시나리오 ' + f.scenarios_count + ' (' + f.steps_count + 'step)</span>' +
        '<button type="button" class="combo-file-delete-btn" data-path="' + f.path + '" data-name="' + f.name + '" ' +
                'style="background:transparent;border:1px solid #DC2626;color:#DC2626;padding:2px 8px;border-radius:4px;font-size:11px;cursor:pointer;" ' +
                'title="이 명세 파일 삭제 (백업됨)">🗑️</button>' +
      '</div>';
    }).join('');
    // 시나리오 체크박스도 같이 렌더
    renderComboScenarios();
    // 삭제 버튼 이벤트 위임
    listEl.querySelectorAll('.combo-file-delete-btn').forEach(function(btn) {
      btn.addEventListener('click', function() {
        deleteComboSpecFile(btn.dataset.path, btn.dataset.name);
      });
    });
  } catch (e) {
    listEl.innerHTML = '<span style="color:#DC2626;">⚠ 네트워크 오류: ' + e.message + '</span>';
  }
}

// Combo 명세 파일 삭제 (백업 후 — .deleted_<timestamp> 로 이동)
async function deleteComboSpecFile(filePath, fileName) {
  if (!confirm('이 명세 파일을 삭제하시겠습니까?\\n\\n' + fileName + '\\n\\n파일은 같은 위치에 .deleted_<timestamp> 로 백업됩니다 (필요하면 수동 복원 가능).')) {
    return;
  }
  try {
    const resp = await fetch('/combo/delete-spec', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ path: filePath })
    });
    const data = await resp.json();
    if (!data.ok) {
      alert('❌ 삭제 실패: ' + (data.error || ''));
      return;
    }
    if (typeof showToast === 'function') {
      showToast('🗑️ 삭제됨: ' + fileName, 'success');
    }
    // 목록 새로고침
    await loadComboFiles();
  } catch (e) {
    alert('네트워크 오류: ' + e.message);
  }
}

// 선택된 파일들의 시나리오를 모아 체크박스로 렌더
function renderComboScenarios() {
  const area = document.getElementById('comboScenarioArea');
  const listEl = document.getElementById('comboScenarioList');
  if (!area || !listEl) return;

  // 체크된 파일들의 scenarios 만 모음 — 파일별로 그룹핑
  const checked = document.querySelectorAll('.combo-file-cb:checked');
  const fileGroups = [];  // [{ fileName, filePath, domain, scenarios: [...] }]
  let totalScenarios = 0;
  checked.forEach(function(cb) {
    try {
      const scns = JSON.parse(decodeURIComponent(cb.dataset.scenarios || '[]'));
      const filePath = cb.value;
      const fileName = filePath.split('/').pop();
      // 같은 파일 안에 있는 시나리오 묶음
      fileGroups.push({
        fileName: fileName,
        filePath: filePath,
        domain: cb.dataset.domain || fileName.replace('_combinations.md', ''),
        scenarios: scns,
      });
      totalScenarios += scns.length;
    } catch (e) {}
  });

  if (!totalScenarios) {
    area.style.display = 'none';
    return;
  }
  area.style.display = 'block';

  // "전체 선택" + 파일별 그룹 헤더 + 시나리오 체크박스
  let html =
    '<label style="display:flex;align-items:center;gap:8px;padding:4px;border-bottom:1px solid #E5E7EB;cursor:pointer;font-weight:600;font-size:12px;">' +
    '  <input type="checkbox" id="comboScnAll" checked onchange="toggleAllScenarios(this.checked)">' +
    '  <span>전체 시나리오 (' + totalScenarios + '개, ' + fileGroups.length + ' 파일)</span>' +
    '</label>';

  fileGroups.forEach(function(group) {
    // 파일 그룹 헤더 — 다중 파일일 때만 표시 (단일이면 불필요)
    if (fileGroups.length > 1) {
      html += '<div style="margin:8px 0 4px 16px;padding:4px 8px;background:#F9FAFB;border-left:3px solid #FCD34D;font-size:11px;color:#374151;font-weight:600;">' +
        '📄 ' + group.fileName +
        '</div>';
    }
    group.scenarios.forEach(function(s) {
      const target = s.is_target ? '<span style="color:#F59E0B;margin-right:4px;">⭐</span>' : '';
      // 파일 출처 라벨 (다중 파일이면 시나리오 ID 옆에)
      const fileLabel = fileGroups.length > 1
        ? '<code style="background:#FEF3C7;color:#92400E;padding:1px 5px;border-radius:3px;font-size:10px;margin-right:4px;">' + group.fileName.replace('_combinations.md', '') + '</code>'
        : '';
      html += '<label style="display:flex;align-items:center;gap:8px;padding:4px 4px 4px ' + (fileGroups.length > 1 ? '32' : '16') + 'px;cursor:pointer;font-size:12px;">' +
        '<input type="checkbox" class="combo-scn-cb" value="' + s.id + '" data-file="' + group.filePath + '" checked>' +
        target + fileLabel +
        '<code style="background:#F3F4F6;padding:1px 5px;border-radius:3px;font-size:11px;">' + s.id + '</code>' +
        '<span>' + s.title + '</span>' +
        '<span style="margin-left:auto;color:#6B7280;font-size:11px;">' + s.step_count + ' step</span>' +
        '</label>';
    });
  });

  listEl.innerHTML = html;
}

function toggleAllScenarios(checked) {
  document.querySelectorAll('.combo-scn-cb').forEach(function(c) { c.checked = checked; });
}

// 사용자가 선택한 Combo 파일 경로 (체크된 것)
function getSelectedComboFiles() {
  return Array.from(document.querySelectorAll('.combo-file-cb:checked')).map(function(c) { return c.value; });
}

// 사용자가 선택한 시나리오 ID 들 (콤마 구분 문자열로 반환 — 백엔드 scn_filter 형식)
function getSelectedComboScenarios() {
  // 전체 체크되어 있으면 "" (필터 없음 = 전체) 반환
  const allCb = document.getElementById('comboScnAll');
  if (allCb && allCb.checked) {
    // 그 안의 모든 시나리오가 다 체크되어 있는지 한 번 더 확인
    const all = document.querySelectorAll('.combo-scn-cb');
    const checked = document.querySelectorAll('.combo-scn-cb:checked');
    if (all.length === checked.length) return '';  // 전체 = 필터 없음
  }
  const checked = Array.from(document.querySelectorAll('.combo-scn-cb:checked'));
  return checked.map(function(c) { return c.value; }).join(',');
}

// 이전 작업의 Combo 옵션 복원 (프로젝트 선택 후 자동 호출)
function restoreComboOpts(opts) {
  if (!opts) return;
  // 파일 경로 — 이미 loadComboFiles 가 체크박스 렌더했으니, 일치하는 것만 체크 유지
  if (opts.file_paths && opts.file_paths.length) {
    const checkboxes = document.querySelectorAll('.combo-file-cb');
    checkboxes.forEach(function(cb) {
      cb.checked = opts.file_paths.indexOf(cb.value) !== -1;
    });
    // 파일 체크 변경에 따라 시나리오 영역 다시 렌더
    if (typeof renderComboScenarios === 'function') renderComboScenarios();
  }
  // OC 필터
  if (opts.oc_filter) {
    const ocEl = document.getElementById('comboOcFilter');
    if (ocEl) ocEl.value = opts.oc_filter;
  }
  // 시나리오 필터 — 체크박스에 반영
  if (opts.scn_filter) {
    setTimeout(function() {  // 시나리오 체크박스 렌더 끝난 후 적용
      const wantedIds = opts.scn_filter.split(',').map(function(s) { return s.trim(); });
      const allCb = document.getElementById('comboScnAll');
      if (allCb) allCb.checked = false;  // 전체 해제 후 일부만 체크
      document.querySelectorAll('.combo-scn-cb').forEach(function(cb) {
        cb.checked = wantedIds.indexOf(cb.value) !== -1;
      });
    }, 300);
  }
}

// Combo 모드 시작 전 검증
function validateComboBeforeStart() {
  if (getSelectedGenerationMode() !== 'combo_only') return { ok: true };
  const files = getSelectedComboFiles();
  if (!files.length) {
    return { ok: false,
             msg: 'Combo 명세 파일이 선택되지 않았습니다. Combo 모드는 *_combinations.md 파일이 필요합니다.\\n대안: 모드를 "정책 반영" 으로 변경하거나, projects/{프로젝트}/ 에 *_combinations.md 추가.' };
  }
  return { ok: true };
}

// 초기 선택 강조 + What's New 배너 체크
if (typeof document !== 'undefined') {
  document.addEventListener('DOMContentLoaded', function() {
    try { onGenModeChanged(); refreshGenerationModeHint(); } catch(e) {}
    try { checkWhatsNewBanner(); } catch(e) {}
  });
}

// v0.9.6 What's New 배너 표시 — localStorage에 dismiss 기록이 없으면 자동 표시
const _WHATS_NEW_VERSION = '{{ app_version }}';
const _WHATS_NEW_KEY = 'tc_whatsnew_dismissed_' + _WHATS_NEW_VERSION;
// 서버 재시작 후 버전 비교용
window._INITIAL_APP_VERSION = '{{ app_version }}';

function checkWhatsNewBanner() {
  try {
    if (localStorage.getItem(_WHATS_NEW_KEY) === '1') return;
  } catch(e) {}
  const banner = document.getElementById('whatsNewBanner');
  if (banner) banner.style.display = 'block';
}

function dismissWhatsNew() {
  const banner = document.getElementById('whatsNewBanner');
  if (banner) banner.style.display = 'none';
  try { localStorage.setItem(_WHATS_NEW_KEY, '1'); } catch(e) {}
}

function showWhatsNew() {
  const modal = document.getElementById('whatsNewModal');
  if (modal) modal.style.display = 'flex';
}

function getSelectedGenerationMode() {
  const sel = document.querySelector('input[name="genMode"]:checked');
  return sel ? sel.value : 'summary';
}

function addSource(type) {
  const id = ++sourceCounter;
  sources.push({ id, type, content: '' });
  renderSources();
  onInputsChanged();  // 소스 추가 감지
  if (type === 'url' || type === 'web') {
    setTimeout(() => document.getElementById('srcUrl_' + id)?.focus(), 50);
  } else if (type === 'text') {
    setTimeout(() => document.getElementById('srcText_' + id)?.focus(), 50);
  } else if (type === 'pdf' || type === 'md') {
    setTimeout(() => {
      const zone = document.getElementById('srcZone_' + id);
      if (zone) zone.scrollIntoView({ behavior: 'smooth', block: 'center' });
    }, 50);
  }
}

function removeSource(id) {
  sources = sources.filter(s => s.id !== id);
  renderSources();
  onInputsChanged();  // 소스 삭제 감지
}

function clearAllSources() {
  if (sources.length === 0) return;
  const n = sources.length;
  if (!confirm('입력 소스 ' + n + '개를 모두 삭제하시겠습니까?\\n\\n이 작업은 되돌릴 수 없습니다.')) return;
  sources = [];
  renderSources();
  onInputsChanged();  // 전체 삭제 감지
}

function updateSourceContent(id, value) {
  const s = sources.find(s => s.id === id);
  if (s) s.content = value;
  onInputsChanged();  // 소스 내용 편집 감지 (URL/텍스트 입력 등)
}

async function onSrcFileChange(id, input) {
  if (!input.files.length) return;
  const file = input.files[0];
  if (!file.name.toLowerCase().endsWith('.pdf')) { alert('PDF 파일만 허용됩니다.'); return; }
  const zone = document.getElementById('srcZone_' + id);
  if (zone) zone.innerHTML = '⏳ 업로드 중...';
  const fd = new FormData();
  fd.append('file', file);
  try {
    const resp = await fetch('/upload', { method: 'POST', body: fd });
    const data = await resp.json();
    if (data.ok) {
      const s = sources.find(s => s.id === id);
      if (s) s.content = data.filename;
      if (zone) zone.innerHTML = '<span class="src-file-name">✅ ' + data.filename + '</span>';
      onInputsChanged();  // PDF 업로드 완료 감지
    } else {
      if (zone) zone.innerHTML = '❌ ' + data.error;
    }
  } catch(e) {
    if (zone) zone.innerHTML = '❌ 업로드 실패';
  }
}

// 공통 업로드 헬퍼 — File[] 배열 → /upload-md 순차 업로드 → sources 배열 갱신
async function _uploadMdFilesToSources(id, files) {
  const zone = document.getElementById('srcZone_' + id);
  const uploaded = [];
  const failed = [];

  if (zone) zone.innerHTML = `⏳ ${files.length}개 파일 업로드 중...`;

  for (let i = 0; i < files.length; i++) {
    const f = files[i];
    if (zone) zone.innerHTML = `⏳ 업로드 중 ${i+1}/${files.length}: ${_escapeHtml(f.name)}`;
    const fd = new FormData();
    // webkitdirectory로 들어온 File은 webkitRelativePath 있음 — 서버는 파일명만 쓰므로 파일 객체 그대로 전달
    fd.append('file', f);
    try {
      const resp = await fetch('/upload-md', { method: 'POST', body: fd });
      const data = await resp.json();
      if (data.ok) {
        uploaded.push(data.filename);
      } else {
        failed.push(f.name + ': ' + (data.error || '업로드 실패'));
      }
    } catch(e) {
      failed.push(f.name + ': 네트워크 오류');
    }
  }

  if (uploaded.length === 0) {
    if (zone) zone.innerHTML = '❌ 모든 파일 업로드 실패<br><span style="font-size:11px;">' + failed.map(_escapeHtml).join('<br>') + '</span>';
    return { uploaded: [], failed };
  }

  // 첫 파일 → 현재 카드
  const firstSource = sources.find(s => s.id === id);
  if (firstSource) firstSource.content = uploaded[0];

  // 나머지 파일 → 새 md 소스로 추가
  for (let i = 1; i < uploaded.length; i++) {
    const newId = ++sourceCounter;
    sources.push({ id: newId, type: 'md', content: uploaded[i] });
  }

  renderSources();
  onInputsChanged();

  if (failed.length > 0) {
    showToast('⚠️ ' + failed.length + '개 파일 업로드 실패 — 콘솔 확인');
    console.warn('마크다운 업로드 실패 목록:', failed);
  } else if (uploaded.length > 1) {
    showToast('✅ ' + uploaded.length + '개 마크다운 파일 추가됨');
  }
  return { uploaded, failed };
}

async function onMdFileChange(id, input) {
  if (!input.files.length) return;
  const files = Array.from(input.files);
  await _uploadMdFilesToSources(id, files);
}

async function onMdFolderChange(id, input) {
  if (!input.files.length) return;
  const all = Array.from(input.files);

  // .md / .markdown / .txt 만 필터 (숨김 파일 제외)
  const mdExt = /\.(md|markdown|txt)$/i;
  const filtered = all.filter(f => {
    const name = (f.name || '').toLowerCase();
    if (!name || name.startsWith('.')) return false;
    return mdExt.test(name);
  });

  if (filtered.length === 0) {
    const zone = document.getElementById('srcZone_' + id);
    if (zone) zone.innerHTML = '❌ 선택한 폴더에 .md / .markdown / .txt 파일이 없습니다 (' + all.length + '개 파일 중 0개 매칭)';
    return;
  }

  // 파일명 기준 정렬 (쪽대본 번호 순서 유지 도움 — 01_spec, 02_spec ...)
  filtered.sort((a, b) => {
    const an = (a.webkitRelativePath || a.name).toLowerCase();
    const bn = (b.webkitRelativePath || b.name).toLowerCase();
    return an.localeCompare(bn);
  });

  // 50개 초과 시 경고 확인
  if (filtered.length > 50) {
    const msg = `📁 폴더에서 ${filtered.length}개의 마크다운 파일을 발견했습니다.\n\n모두 업로드하시겠습니까?\n\n(권장: 50개 이하 · 파일이 너무 많으면 TC 생성 시간이 오래 걸리거나 실패할 수 있습니다)`;
    if (!confirm(msg)) {
      const zone = document.getElementById('srcZone_' + id);
      if (zone) zone.innerHTML = '❌ 사용자가 업로드를 취소했습니다 (' + filtered.length + '개 파일).';
      return;
    }
  }

  // 폴더 구조 정보 — 상위 경로 보여주기
  const folderName = (filtered[0].webkitRelativePath || '').split('/')[0] || '(폴더)';
  showToast('📁 "' + folderName + '" 폴더에서 ' + filtered.length + '개 .md 파일 발견 — 업로드 시작');

  await _uploadMdFilesToSources(id, filtered);
}

function renderSources() {
  const list  = document.getElementById('sourceList');
  const empty = document.getElementById('sourceEmpty');
  const clearBtn = document.getElementById('btnClearAllSources');
  if (!list) return;
  // 전체 삭제 버튼: 2개 이상일 때만 노출
  if (clearBtn) clearBtn.style.display = (sources.length >= 2) ? '' : 'none';
  if (sources.length === 0) {
    list.innerHTML = '';
    empty && empty.classList.remove('hidden');
    return;
  }
  empty && empty.classList.add('hidden');
  list.innerHTML = sources.map(src => {
    const badgeClass = src.type;
    const badges = {
      pdf: '📄 PDF', url: '🔗 GitHub URL', web: '🌐 웹 URL', md: '📝 마크다운', text: '✏️ 텍스트',
      spec_folder: '📁 구조화 spec 폴더', spec_folder_prev: '📁 이전 spec 폴더'
    };
    let body = '';
    if (src.type === 'pdf') {
      body = src.content
        ? '<span class="src-file-name">✅ ' + src.content + '</span>'
        : '<div class="src-dropzone" id="srcZone_' + src.id + '" onclick="document.getElementById(&#39;srcFile_' + src.id + '&#39;).click()">클릭하여 PDF 파일 선택<br><span style="font-size:11px;color:var(--muted)">최대 50MB · PDF만 허용</span></div>';
      body += '<input type="file" id="srcFile_' + src.id + '" accept=".pdf" style="display:none" onchange="onSrcFileChange(' + src.id + ', this)">';
    } else if (src.type === 'url') {
      body = '<input type="text" id="srcUrl_' + src.id + '" class="form-input" placeholder="https://github.com/user/repo" value="' + src.content + '" oninput="updateSourceContent(' + src.id + ', this.value)">';
      body += '<div class="input-hint">GitHub 저장소 URL (private repo는 GITHUB_TOKEN 필요)</div>';
      body += '<button type="button" class="btn-preview-tree" onclick="previewGithubTree(' + src.id + ')">🗂 파일 목록 보기</button>';
      body += '<div id="treePanel_' + src.id + '" class="tree-panel hidden"></div>';
    } else if (src.type === 'web') {
      body = '<input type="text" id="srcUrl_' + src.id + '" class="form-input" placeholder="https://example.com 또는 https://app.vercel.app" value="' + src.content + '" oninput="updateSourceContent(' + src.id + ', this.value)">';
      body += '<div class="input-hint">웹 서비스, 랜딩 페이지, Vercel 배포 URL 등 — HTML을 크롤링하여 분석합니다</div>';
    } else if (src.type === 'md') {
      if (src.content) {
        body = '<span class="src-file-name">✅ ' + src.content + '</span>';
      } else {
        body = '<div class="src-dropzone" id="srcZone_' + src.id + '" style="padding:16px;">'
             + '<div style="font-size:13px;color:var(--text);margin-bottom:8px;">📝 마크다운 파일 또는 폴더 업로드</div>'
             + '<div style="font-size:11px;color:var(--muted);margin-bottom:10px;">여러 .md 파일 선택 가능 · 폴더 선택 시 내부 .md만 자동 필터</div>'
             + '<div style="display:flex;gap:8px;justify-content:center;flex-wrap:wrap;">'
             + '<button type="button" class="btn" style="padding:6px 14px;font-size:12px;" onclick="document.getElementById(&#39;srcFile_' + src.id + '&#39;).click()">📄 파일 선택</button>'
             + '<button type="button" class="btn" style="padding:6px 14px;font-size:12px;" onclick="document.getElementById(&#39;srcDir_' + src.id + '&#39;).click()">📁 폴더 선택</button>'
             + '</div></div>';
      }
      body += '<input type="file" id="srcFile_' + src.id + '" accept=".md,.markdown,.txt" multiple style="display:none" onchange="onMdFileChange(' + src.id + ', this)">';
      body += '<input type="file" id="srcDir_' + src.id + '" webkitdirectory directory multiple style="display:none" onchange="onMdFolderChange(' + src.id + ', this)">';
    } else {
      body = '<textarea id="srcText_' + src.id + '" class="form-input text-input-area" placeholder="기획서 내용, 슬랙 메시지 등 자유롭게 붙여넣으세요..." oninput="updateSourceContent(' + src.id + ', this.value)">' + src.content + '</textarea>';
    }
    return '<div class="source-card">' +
      '<div class="source-card-header">' +
        '<span class="source-type-badge ' + badgeClass + '">' + badges[src.type] + '</span>' +
        '<button type="button" class="btn-remove-source" onclick="removeSource(' + src.id + ')">✕ 삭제</button>' +
      '</div>' +
      '<div class="source-card-body">' + body + '</div>' +
      '</div>';
  }).join('');
}

async function startPipeline() {
  const projectName = document.getElementById('projectName').value.trim() || '프로젝트';
  const focusArea = document.getElementById('focusArea').value.trim();

  // Combo 모드 검증 (모드가 combo_only 인데 파일 없으면 차단)
  const comboCheck = validateComboBeforeStart();
  if (!comboCheck.ok) {
    alert('⚠️ Combo 모드 검증 실패\\n\\n' + comboCheck.msg);
    return;
  }

  // 구조화 spec 폴더 모드 분기 — v0.10.x: spec 폴더 경로가 비어있지 않으면 spec 모드 우선
  const specPathEl = document.getElementById('specFolderPath');
  const specFolderActive = specPathEl && specPathEl.value.trim();
  if (specFolderActive) {
    return startPipelineSpecFolder(projectName, focusArea, specPathEl.value.trim());
  }

  if (sources.length === 0) { alert('구조화 spec 폴더 경로를 입력하거나, 개별 소스를 하나 이상 추가하세요.'); return; }

  const typeNames = { pdf: 'PDF', url: 'GitHub URL', web: '웹 URL', md: '마크다운', text: '텍스트' };
  for (const s of sources) {
    if (!s.content.trim()) {
      alert(typeNames[s.type] + ' 소스의 내용을 입력해주세요.');
      return;
    }
  }

  const payload = sources.map(s => ({ type: s.type, content: s.content, selected_files: s.selected_files || null }));

  document.getElementById('startBtn').disabled = true;
  document.getElementById('card1').classList.add('hidden');
  document.getElementById('card2').classList.remove('hidden');
  setStepBar(2);
  document.getElementById('card2').scrollIntoView({ behavior: 'smooth', block: 'start' });
  // 새 파이프라인 시작 시 중단 버튼 재활성화 (이전 세션 상태 승계 방지)
  setStopButtonsDisabled(false);
  var stopBtn2Init = document.getElementById('stopBtn2');
  if (stopBtn2Init) stopBtn2Init.style.display = '';

  try {
    const resp = await fetch('/start', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ sources: payload, project_name: projectName, focus_area: focusArea || null, generation_mode: getSelectedGenerationMode() })
    });
    const data = await resp.json();
    if (!data.ok) {
      alert('오류: ' + data.error);
      document.getElementById('startBtn').disabled = false;
      return;
    }
    currentSid = data.sid;
    connectStream(data.sid);
  } catch(e) {
    alert('서버 오류: ' + e.message);
    document.getElementById('startBtn').disabled = false;
  }
}

async function startPipelineSpecFolder(projectName, focusArea, folderPath) {
  document.getElementById('startBtn').disabled = true;
  document.getElementById('card1').classList.add('hidden');
  document.getElementById('card2').classList.remove('hidden');
  setStepBar(2);
  document.getElementById('card2').scrollIntoView({ behavior: 'smooth', block: 'start' });
  setStopButtonsDisabled(false);

  // diff 모드 옵션 수집
  const diffBox = document.getElementById('diffBox');
  const prevPath = (document.getElementById('prevSpecFolderPath') && document.getElementById('prevSpecFolderPath').value || '').trim();
  const diffActive = diffBox && diffBox.style.display !== 'none' && prevPath;
  const includeUnchanged = document.getElementById('includeUnchanged') ? document.getElementById('includeUnchanged').checked : false;

  try {
    const isComboMode = getSelectedGenerationMode() === 'combo_only';
    const comboPayload = isComboMode ? {
      file_paths: getSelectedComboFiles(),
      oc_filter: (document.getElementById('comboOcFilter')?.value || '').trim(),
      scn_filter: getSelectedComboScenarios(),  // 체크박스 기반 — 전체면 빈 문자열
    } : null;

    const resp = await fetch('/start-spec-folder', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({
        project_name: projectName,
        focus_area: focusArea || null,
        folder_path: folderPath,
        prev_folder_path: diffActive ? prevPath : '',
        include_unchanged: includeUnchanged,
        generation_mode: getSelectedGenerationMode(),
        combo: comboPayload,
      })
    });
    const data = await resp.json();
    if (!data.ok) {
      alert('오류: ' + data.error);
      document.getElementById('startBtn').disabled = false;
      document.getElementById('card1').classList.remove('hidden');
      document.getElementById('card2').classList.add('hidden');
      return;
    }
    currentSid = data.sid;
    if (data.summary) {
      showToast(`📁 폴더 분류: overview ${data.summary.overview ? '✓' : '✗'} · policy ${data.summary.policy_count} · design ${data.summary.design_count} · 화면 ${data.summary.screens_count}개${diffActive ? ' (diff 모드)' : ''}`, 'success');
    }
    connectStream(data.sid);
  } catch(e) {
    alert('서버 오류: ' + e.message);
    document.getElementById('startBtn').disabled = false;
  }
}

function toggleDiffMode() {
  const box = document.getElementById('diffBox');
  const btn = document.getElementById('btnDiffToggle');
  if (!box || !btn) return;
  const enabled = box.style.display === 'none';
  box.style.display = enabled ? '' : 'none';
  btn.textContent = enabled ? '사용 중 ✓' : '사용';
  btn.style.background = enabled ? '#FCD34D' : '#FFFFFF';
}

async function previewDiff() {
  const newPath = document.getElementById('specFolderPath').value.trim();
  const prevPath = document.getElementById('prevSpecFolderPath').value.trim();
  if (!newPath || !prevPath) { alert('두 폴더 경로를 모두 입력하세요.'); return; }
  const prev = document.getElementById('diffPreview');
  prev.style.display = '';
  prev.innerHTML = '⏳ diff 계산 중...';
  try {
    const r = await fetch('/diff-spec-folders', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ folder_path: newPath, prev_folder_path: prevPath })
    });
    const d = await r.json();
    if (!d.ok) { prev.innerHTML = '<span style="color:#DC2626;">❌ ' + d.error + '</span>'; return; }
    let html = '<div style="line-height:1.6;">';
    html += `<div><strong>🆕 추가:</strong> ${d.added.length}개${d.added.length ? ' — <code>' + d.added.slice(0, 8).join(', ') + (d.added.length > 8 ? ', ...' : '') + '</code>' : ''}</div>`;
    html += `<div><strong>📝 수정:</strong> ${d.modified.length}개${d.modified.length ? ' — <code>' + d.modified.slice(0, 8).join(', ') + (d.modified.length > 8 ? ', ...' : '') + '</code>' : ''}</div>`;
    html += `<div><strong>🗑 삭제:</strong> ${d.removed.length}개${d.removed.length ? ' — <code>' + d.removed.slice(0, 8).join(', ') + (d.removed.length > 8 ? ', ...' : '') + '</code>' : ''}</div>`;
    html += `<div><strong>✅ 동일:</strong> ${d.unchanged_count}개</div>`;
    if (d.common_changed) {
      html += '<div style="margin-top:4px;color:#B45309;"><strong>⚠️ 공통 문서(정책/디자인/overview) 변경됨</strong> — 모든 화면 재생성을 권장합니다 (전체 일관성 위해).</div>';
    }
    html += `<div style="margin-top:6px;color:#059669;"><strong>재생성 대상:</strong> ${d.regenerate_count}개${d.common_changed ? ' (공통 변경 시 전체 처리됨)' : ''}</div>`;
    html += '</div>';
    prev.innerHTML = html;
  } catch(e) {
    prev.innerHTML = '<span style="color:#DC2626;">❌ 서버 오류: ' + e.message + '</span>';
  }
}

// 입력 디바운스 — 타이핑 중 매번 호출하지 않게
let _resumeCheckTimer = null;
function onProjectNameInputForResume() {
  if (_resumeCheckTimer) clearTimeout(_resumeCheckTimer);
  _resumeCheckTimer = setTimeout(() => {
    // 새 구조: spec 아코디언 body 가 펼쳐져 있을 때만 resume 체크
    const body = document.getElementById('specFolderBody');
    if (body && body.style.display !== 'none') checkResumeSpec();
  }, 500);
}

async function checkResumeSpec() {
  const projectName = document.getElementById('projectName').value.trim();
  if (!projectName) return;
  try {
    const r = await fetch('/check-resume-spec?project_name=' + encodeURIComponent(projectName));
    const d = await r.json();
    const btn = document.getElementById('btnResumeSpec');
    if (!btn) return;
    if (d.ok && d.resumable) {
      btn.style.display = '';
      btn.title = `완료된 화면 ${d.completed_count}개 / 전체 ${d.total_screens || '?'}개 (stage=${d.stage})`;
      btn.textContent = `▶ 이어서 작업 (${d.completed_count}/${d.total_screens || '?'} 완료)`;
    } else {
      btn.style.display = 'none';
    }
  } catch(e) {
    // 무시
  }
}

async function resumeSpecFolder() {
  const projectName = document.getElementById('projectName').value.trim() || '프로젝트';
  if (!confirm(`"${projectName}" 의 이전 작업을 이어서 실행하시겠습니까?\\n(이미 완료된 화면은 건너뛰고 미완료 화면만 처리합니다)`)) return;

  document.getElementById('startBtn').disabled = true;
  document.getElementById('card1').classList.add('hidden');
  document.getElementById('card2').classList.remove('hidden');
  setStepBar(2);
  setStopButtonsDisabled(false);

  try {
    const r = await fetch('/start-spec-folder', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ project_name: projectName, resume: true })
    });
    const d = await r.json();
    if (!d.ok) {
      alert('이어서 작업 실패: ' + d.error);
      document.getElementById('startBtn').disabled = false;
      document.getElementById('card1').classList.remove('hidden');
      document.getElementById('card2').classList.add('hidden');
      return;
    }
    currentSid = d.sid;
    showToast(`▶ 이어서 작업 — stage=${d.resumed_stage}`, 'success');
    connectStream(d.sid);
  } catch(e) {
    alert('서버 오류: ' + e.message);
    document.getElementById('startBtn').disabled = false;
  }
}

// ── 신규 아코디언 시스템 (v0.10.x UI 정식화) ────────────────────────
// 이전: toggleSpecFolderMode() 가 별도 박스 표시/숨김
// 현재: spec 섹션은 기본 펼침, legacy 소스는 기본 접힘. toggleAccordion('spec'|'legacy')
function toggleAccordion(which) {
  const map = {
    spec:   { body: 'specFolderBody',   chev: 'specFolderChevron' },
    legacy: { body: 'legacySourceBody', chev: 'legacySourceChevron' },
  };
  const cfg = map[which];
  if (!cfg) return;
  const body = document.getElementById(cfg.body);
  const chev = document.getElementById(cfg.chev);
  if (!body || !chev) return;
  const opening = body.style.display === 'none';
  body.style.display = opening ? '' : 'none';
  chev.textContent = opening ? '▼' : '▶';
  if (which === 'spec' && opening) checkResumeSpec();
}

// 하위 호환 — 이전 toggleSpecFolderMode() 호출하는 곳 보호
function toggleSpecFolderMode() {
  // 이전 코드 호환: 호출되면 spec 아코디언 토글
  toggleAccordion('spec');
}

// ── 최근 사용 폴더 드롭다운 ─────────────────────────────────────────
async function toggleRecentFolders() {
  const panel = document.getElementById('recentFoldersPanel');
  if (!panel) return;
  if (panel.style.display !== 'none' && panel.dataset.loaded === '1') {
    panel.style.display = 'none';
    return;
  }
  await loadRecentFolders('specFolderPath', 'recentFoldersPanel');
}

async function toggleRecentPrevFolders() {
  const panel = document.getElementById('recentPrevFoldersPanel');
  if (!panel) return;
  if (panel.style.display !== 'none' && panel.dataset.loaded === '1') {
    panel.style.display = 'none';
    return;
  }
  await loadRecentFolders('prevSpecFolderPath', 'recentPrevFoldersPanel');
}

async function loadRecentFolders(targetInputId, panelId) {
  const projectName = document.getElementById('projectName').value.trim();
  const panel = document.getElementById(panelId);
  panel.style.display = '';
  panel.innerHTML = '<div style="padding:10px;color:#94A3B8;">⏳ 불러오는 중...</div>';
  try {
    const url = '/recent-spec-folders' + (projectName ? '?project_name=' + encodeURIComponent(projectName) : '');
    const r = await fetch(url);
    const d = await r.json();
    if (!d.ok || !d.folders || d.folders.length === 0) {
      panel.innerHTML = '<div style="padding:10px;color:#94A3B8;">최근 사용한 폴더가 없습니다.</div>';
      panel.dataset.loaded = '1';
      return;
    }
    // 안전한 렌더 — DOM API 사용해 onclick handler 직접 부여 (escape 함정 회피)
    panel.innerHTML = '';
    d.folders.forEach(function(f) {
      var div = document.createElement('div');
      div.style.cssText = 'padding:8px 12px;border-bottom:1px solid #F1F5F9;cursor:pointer;font-family:monospace;color:#1E293B;';
      div.textContent = f;
      div.onclick = function() { pickRecentFolder(targetInputId, panelId, f); };
      div.onmouseover = function() { this.style.background = '#F1F5F9'; };
      div.onmouseout = function() { this.style.background = ''; };
      panel.appendChild(div);
    });
    panel.dataset.loaded = '1';
  } catch (e) {
    panel.innerHTML = '<div style="padding:10px;color:#DC2626;">오류: ' + e.message + '</div>';
  }
}

function pickRecentFolder(targetInputId, panelId, path) {
  const input = document.getElementById(targetInputId);
  if (input) {
    input.value = path;
    input.dispatchEvent(new Event('input'));
  }
  const panel = document.getElementById(panelId);
  if (panel) panel.style.display = 'none';
}

function onSpecFolderChanged() {
  const prev = document.getElementById('specFolderPreview');
  if (prev) { prev.style.display = 'none'; prev.innerHTML = ''; }
  const hint = document.getElementById('specFolderHint');
  if (hint) hint.textContent = '경로가 변경됐습니다. 미리보기를 다시 실행하세요.';
}

async function previewSpecFolder() {
  const path = document.getElementById('specFolderPath').value.trim();
  if (!path) { alert('폴더 경로를 입력하세요.'); return; }
  const prev = document.getElementById('specFolderPreview');
  prev.style.display = '';
  prev.innerHTML = '⏳ 폴더 스캔 중...';
  try {
    const r = await fetch('/preview-spec-folder', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ folder_path: path })
    });
    const d = await r.json();
    if (!d.ok) { prev.innerHTML = '<span style="color:#DC2626;">❌ ' + d.error + '</span>'; return; }
    let html = '<div style="line-height:1.7;">';
    html += `<div><strong>📂 폴더:</strong> <code>${d.folder}</code></div>`;
    html += `<div><strong>📑 overview:</strong> ${d.overview || '<span style="color:#DC2626;">없음 (분류표 자동 생성 불가)</span>'}</div>`;
    html += `<div><strong>📋 policy:</strong> ${d.policy.length === 0 ? '<span style="color:#9CA3AF;">없음</span>' : d.policy.join(', ')}</div>`;
    html += `<div><strong>🎨 design:</strong> ${d.design.length === 0 ? '<span style="color:#9CA3AF;">없음</span>' : d.design.join(', ')}</div>`;
    html += `<div><strong>🖥️ 화면 md:</strong> ${d.screens.length}개`;
    if (d.screens.length > 0) html += ` <span style="color:#64748B;">(${d.screens.slice(0, 5).join(', ')}${d.screens.length > 5 ? ', ...' : ''})</span>`;
    html += '</div>';
    html += `<div><strong>📊 화면 목록 표 행:</strong> ${d.screen_rows_count}행`;
    if (d.majors.length > 0) html += ` · 대분류: ${d.majors.join(', ')}`;
    html += '</div>';
    if (d.screens.length === 0) {
      html += '<div style="margin-top:6px;color:#DC2626;">⚠️ scr/SCR-*.md 형식의 화면 파일이 필요합니다.</div>';
    } else {
      html += '<div style="margin-top:6px;color:#059669;">✅ 시작 준비 완료. 화면 ' + d.screens.length + '개에 대해 LLM 호출이 발생합니다.</div>';
    }
    html += '</div>';
    prev.innerHTML = html;
  } catch(e) {
    prev.innerHTML = '<span style="color:#DC2626;">❌ 서버 오류: ' + e.message + '</span>';
  }
}

let _sseReconnectCount = 0;
const _SSE_MAX_RECONNECT = 3;

function connectStream(sid) {
  // 이전 중단 배너 제거 + 재연결 카운트 리셋
  document.querySelectorAll('.stopped-banner, .error-banner, .session-lost-banner').forEach(el => el.remove());
  _sseReconnectCount = 0;

  eventSource = new EventSource('/stream/' + sid);
  eventSource.onopen = () => {
    // 성공 연결 → 재연결 카운트 리셋, 중단 버튼 활성
    _sseReconnectCount = 0;
    setStopButtonsDisabled(false);
    if (typeof updateRestartButtonState === 'function') updateRestartButtonState();
  };
  eventSource.onmessage = (e) => {
    try {
      const evt = JSON.parse(e.data);
      handleEvent(evt);
    } catch(err) {}
  };
  eventSource.onerror = async () => {
    // 먼저 세션이 서버에 살아있는지 HEAD로 확인
    // (/stream/<sid> 가 404면 세션이 사라진 것 — 영원히 재시도해도 소용없음)
    let sessionAlive = false;
    try {
      const head = await fetch('/stream/' + sid, { method: 'HEAD' });
      sessionAlive = (head.status !== 404);
    } catch(e) {
      // 네트워크 자체 문제 — 재시도 의미 있음
      sessionAlive = true;
    }

    if (!sessionAlive) {
      // 세션 소멸 확정 → 재연결 중단 + 명확한 안내
      if (eventSource) { eventSource.close(); eventSource = null; if (typeof updateRestartButtonState === 'function') updateRestartButtonState(); }
      showSessionLostBanner();
      setStopButtonsDisabled(true);
      stopCountdown();
      return;
    }

    // 세션은 살아있음 — 일반 네트워크 지연 등. 제한된 횟수만 재시도 안내
    _sseReconnectCount++;
    if (_sseReconnectCount === 1) {
      addLog('⚠️ SSE 연결 오류. 재연결 시도 중...', true);
    } else if (_sseReconnectCount >= _SSE_MAX_RECONNECT) {
      addLog(`⚠️ 재연결 실패(${_sseReconnectCount}회). 새로고침이나 세션 확인 필요.`, true);
      if (eventSource) { eventSource.close(); eventSource = null; if (typeof updateRestartButtonState === 'function') updateRestartButtonState(); }
      showSessionLostBanner(true);
      return;
    }
    setStopButtonsDisabled(false);
  };
}

function showSessionLostBanner(afterRetries) {
  // 이미 표시되어 있으면 중복 방지
  if (document.querySelector('.session-lost-banner')) return;
  const banner = document.createElement('div');
  banner.className = 'session-lost-banner';
  banner.style.cssText = 'margin:12px 0;padding:14px 18px;background:#FEF2F2;border:1.5px solid #DC2626;border-radius:10px;color:#991B1B;';
  banner.innerHTML = `
    <div style="font-weight:700;font-size:14px;margin-bottom:6px;">🔌 세션이 종료되었습니다</div>
    <div style="font-size:12.5px;line-height:1.6;color:#7F1D1D;">
      서버와의 연결이 끊어졌습니다${afterRetries ? ' (재연결 시도 실패)' : ''}. 다음 중 하나를 선택하세요:<br>
      • <strong>새로 시작</strong>: 브라우저 새로고침 후 처음부터<br>
      • <strong>이어서 작업</strong>: 프로젝트 선택 후 아래 버튼 (승인 전 단계만 지원)
    </div>
    <div style="margin-top:10px;display:flex;gap:8px;">
      <button onclick="location.reload()" style="padding:6px 14px;background:#DC2626;color:#FFFFFF;border:none;border-radius:6px;font-size:12px;font-weight:600;cursor:pointer;">🔄 새로고침</button>
      <button onclick="document.querySelector('.session-lost-banner').remove()" style="padding:6px 14px;background:#FFFFFF;color:#6B7280;border:1px solid #D1D5DB;border-radius:6px;font-size:12px;cursor:pointer;">✕ 닫기</button>
    </div>
  `;
  // 진행 중 카드 위에 삽입
  const card2 = document.getElementById('card2');
  if (card2 && !card2.classList.contains('hidden')) {
    card2.insertBefore(banner, card2.firstChild);
  } else {
    const wrap = document.querySelector('.app') || document.body;
    wrap.insertBefore(banner, wrap.firstChild);
  }
}

function handleEvent(evt) {
  if (evt.type === 'ping') return;

  if (evt.type === 'stage') {
    const { stage, label, pct, eta_sec } = evt.data;
    updateProgress(label, pct, eta_sec);
    updateSubsteps(stage, label);
    // 통합된 card2 제목/배지를 현 단계에 맞게 갱신
    const titleEl = document.getElementById('pipelineCardTitle');
    const badgeEl = document.getElementById('pipelineCardBadge');
    if (titleEl && badgeEl) {
      if (stage <= 2) {
        titleEl.textContent = '⚙️ 분석 및 분류';
        badgeEl.textContent = 'Step 2';
      } else if (stage === 3) {
        titleEl.textContent = '🔍 분류표 검토 대기';
        badgeEl.textContent = 'Step 3';
      } else if (stage === 4) {
        titleEl.textContent = '✍️ TC 생성';
        badgeEl.textContent = 'Step 4';
      } else if (stage >= 5) {
        titleEl.textContent = '📊 Excel 빌드';
        badgeEl.textContent = 'Step 5';
      }
    }
    setStepBar(stage);
  }

  if (evt.type === 'log') {
    // 모든 로그는 card2의 단일 logBox에 표시 (통합 뷰)
    addLog(evt.data.msg);
  }

  if (evt.type === 'gate') {
    const isModify = evt.data.mode === 'modify';
    window._gateMode = isModify ? 'modify' : 'new';
    // 제목/안내문 전환
    document.getElementById('gateTitle').innerHTML = isModify
      ? '📋 영향도 검토 <span class="badge">Step 3 · Human Gate</span>'
      : '🔍 분류표 검토 <span class="badge">Step 3 · Human Gate</span>';
    document.getElementById('gateInfoBox').innerHTML = isModify
      ? 'AI가 분석한 변경 영향도입니다. 하단 AI 도우미로 수정을 요청하고, 아래 내용에서 확인한 뒤 승인하세요.'
      : 'AI가 생성한 분류표입니다. 하단 AI 도우미로 수정을 요청하고, 아래 표에서 확인한 뒤 승인하세요.';
    initGateChat(evt.data.content);
    // 입력 소스 SCR 매핑 — 프론트 전역에 저장 (미니 채팅/뷰어 등에서 참조 가능)
    window._sourceScrMap = evt.data.source_scr_map || {};
    // 사용자 안내: 입력 소스에서 화면 식별자 발견 여부
    renderSourceScrNotice(window._sourceScrMap);
    // v0.9.22 Stage 1 — 분류표 중복 가능성 예측 알림
    window._classifyPrediction = evt.data.duplicate_prediction || { predictions: [], high_risk_count: 0 };
    try { renderClassifyDuplicateNotice(window._classifyPrediction); } catch (_) {}
    // Gate 진입 시 입력/파이프라인 카드는 숨기고 Gate 패널에 집중
    document.getElementById('card1').classList.add('hidden');
    document.getElementById('card2').classList.add('hidden');
    document.getElementById('card3').classList.remove('hidden');
    setStepBar(3);
    // Excel 옵션 — '🔄 변경 이력' 체크박스: 신규 모드면 disabled, 수정 모드면 enabled
    var changeCb = document.querySelector('.excel-sheet-cb[data-sheet="change_history"]');
    var changeLabel = document.getElementById('excelChangeHistoryLabel');
    if (changeCb && changeLabel) {
      if (isModify) {
        changeCb.disabled = false;
        changeLabel.style.color = '#374151';
        changeLabel.style.cursor = 'pointer';
      } else {
        changeCb.disabled = true;
        changeCb.checked = false;
        changeLabel.style.color = '#9CA3AF';
        changeLabel.style.cursor = 'not-allowed';
      }
      try { updateExcelSheetSummary(); } catch(_) {}
    }

    // TC ID 모드 기본값 결정: 프로젝트가 SM/SA면 System Generated ON, 아니면 OFF
    // (screen_code_map 서버 응답으로 screen_based_default 확인)
    var _projName = document.getElementById('projectName').value || '';
    loadScreenCodeMap(_projName).then(function(res) {
      var defaultOn = !!(res && res.screen_based_default);
      // 사용자가 아직 선택하지 않았다면 기본값 적용
      if (window._autoScreenCode === undefined) {
        window._autoScreenCode = defaultOn;
      }
      var topToggle = document.getElementById('tcIdModeToggle');
      if (topToggle) {
        topToggle.checked = !!window._autoScreenCode;
        setTcIdMode(topToggle.checked);
      }
    });

    document.getElementById('card3').scrollIntoView({ behavior: 'smooth', block: 'start' });
  }

  // 원칙 G — 중복 의심 TC 발견 알림 (v0.9.17~). 결과 카드(card5) 에서 표시되도록 sess에 보관.
  if (evt.type === 'duplicate_warning') {
    window._duplicateReport = evt.data;  // {patterns, total_duplicates, suggested_keep, suggested_remove}
    addLog('⚠️ 중복 의심 패턴 ' + (evt.data.patterns || []).length + '개 발견 — 결과 화면에서 상세 확인');
  }

  if (evt.type === 'done') {
    stopCountdown();
    const { filename, size, sid, total_tc, min_tc, smoke_tc } = evt.data;
    currentFilename = filename;
    // 모든 substep을 done 처리 (visual 완결)
    for (let i = 1; i <= 6; i++) {
      const el = document.getElementById('sub' + i);
      if (el) { el.classList.remove('active'); el.classList.add('done'); }
    }
    // 입력(card1)/파이프라인(card2) 숨기고 완료 카드(card5) 노출
    document.getElementById('card1').classList.add('hidden');
    document.getElementById('card2').classList.add('hidden');
    document.getElementById('card5').classList.remove('hidden');
    document.getElementById('resultFilename').textContent = filename;
    document.getElementById('resultMeta').textContent =
      `${(size / 1024).toFixed(1)} KB`;
    document.getElementById('statTotal').textContent = total_tc;
    // Smoke TC: 서버가 smoke_tc를 보내면 우선 사용, 없으면 min_tc 폴백 (하위호환)
    const smokeVal = (typeof smoke_tc === 'number') ? smoke_tc : min_tc;
    document.getElementById('statSmoke').textContent = (smokeVal !== undefined && smokeVal !== null) ? smokeVal : '—';
    // Drive 업로드 버튼 라벨 reset — 이전 결과의 'Drive 업로드 완료' 상태가 새 결과에
    // 그대로 남는 버그 차단. 새 파일은 아직 업로드 전이므로 원래 라벨로 복귀.
    resetDriveBtn();
    setStepBar(5);
    setStopButtonsDisabled(true);
    // 원칙 G — 중복 의심 TC 알림 렌더링 (있을 경우)
    try { renderDuplicateNotice(window._duplicateReport); } catch (_) {}
    document.getElementById('card5').scrollIntoView({ behavior: 'smooth', block: 'start' });
    showToast('🎉 TC 생성 완료!');
    if (eventSource) eventSource.close();
  }

  if (evt.type === 'stopped') {
    stopCountdown();
    setStopButtonsDisabled(true);
    // 상단 gateNavBox 가 보이면 그곳을 중단 상태로 갱신 (단일 네비게이션)
    const gateNavBox = document.getElementById('gateNavBox');
    const gateNavMsg = document.getElementById('gateNavMsg');
    const gateNavButtons = document.getElementById('gateNavButtons');
    if (gateNavBox && gateNavMsg && gateNavButtons) {
      // 시각 — 빨간 톤으로
      gateNavBox.style.background = '#FEE2E2';
      gateNavBox.style.borderColor = '#FCA5A5';
      gateNavBox.style.borderStyle = 'solid';
      gateNavMsg.style.color = '#991B1B';
      gateNavMsg.innerHTML = '⏹ <strong>파이프라인이 중단되었습니다</strong><br>'
        + '<span style="font-size:11px;">아래 버튼으로 다음 작업을 선택하세요.</span>';
      gateNavButtons.innerHTML = ''
        + '<button onclick="restartFromScratch()" style="padding:8px 16px; background:#2563EB; color:#fff; border:none; border-radius:8px; font-size:13px; font-weight:600; cursor:pointer;">🏠 처음부터 (입력 유지)</button>'
        + '<button onclick="retryPipeline()" style="padding:8px 14px; background:#fff; color:#1D4ED8; border:1.5px solid #93C5FD; border-radius:8px; font-size:12px; cursor:pointer;">🔄 이어서 재시작</button>';
    }
    // card2 에만 별도 banner (gateNavBox 는 card3 에만 있음)
    ['card2'].forEach(id => {
      const card = document.getElementById(id);
      if (!card.classList.contains('hidden')) {
        const banner = document.createElement('div');
        banner.className = 'stopped-banner';
        banner.innerHTML = '<div class="stopped-icon">⏹</div>' +
          '<div style="flex:1;">' +
            '<div class="stopped-msg">파이프라인이 중단되었습니다.</div>' +
            '<div class="stopped-sub">아래 버튼으로 다음 작업을 선택하세요.</div>' +
          '</div>' +
          '<div style="display:flex;gap:8px;flex-wrap:wrap;">' +
            '<button onclick="restartFromScratch()" style="padding:8px 16px;background:#2563EB;color:#fff;border:none;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;white-space:nowrap;">🏠 처음부터 시작</button>' +
            '<button onclick="retryPipeline()" style="padding:8px 14px;background:#fff;color:#1D4ED8;border:1.5px solid #93C5FD;border-radius:8px;font-size:12px;cursor:pointer;white-space:nowrap;">🔄 이어서 재시작</button>' +
          '</div>';
        banner.style.cssText = 'display:flex;align-items:center;gap:14px;flex-wrap:wrap;';
        card.appendChild(banner);
      }
    });
    document.getElementById('startBtn').disabled = false;
    document.getElementById('startModifyBtn').disabled = false;
    // card1 (입력 카드) 도 함께 노출 — 사용자가 처음부터 시작 / 입력 변경 등 옵션 갖도록
    document.getElementById('card1').classList.remove('hidden');
    setStepBar(1);
    showToast('⏹ 파이프라인이 중단되었습니다.', 'error');
    // 가드 해제 — isPipelineRunning() 이 true 로 남아 처음부터 시작이 막히던 버그 차단
    if (eventSource) { eventSource.close(); eventSource = null; }
    currentSid = null;
    if (typeof updateRestartButtonState === 'function') updateRestartButtonState();
  }

  if (evt.type === 'error') {
    stopCountdown();
    setStopButtonsDisabled(true);
    addLog('❌ 오류: ' + evt.data.msg, true);
    document.getElementById('stageLabel').textContent = '오류 발생';
    document.getElementById('startBtn').disabled = false;
    // 가드 해제 — error 후에도 처음부터 시작 / 입력 화면으로 가 막히지 않게
    if (eventSource) { eventSource.close(); eventSource = null; }
    currentSid = null;
    if (typeof updateRestartButtonState === 'function') updateRestartButtonState();
    // 오류 배너 + 재시작 버튼
    var card2 = document.getElementById('card2');
    var existing = card2.querySelector('.error-banner');
    if (existing) existing.remove();
    var banner = document.createElement('div');
    banner.className = 'error-banner';
    banner.style.cssText = 'background:#FEF2F2;border:1.5px solid #FECACA;border-radius:10px;padding:14px 18px;margin-top:14px;display:flex;align-items:center;gap:12px;flex-wrap:wrap;';
    banner.innerHTML = '<div style="font-size:24px;">⚠️</div>' +
      '<div style="flex:1;"><div style="font-size:13px;font-weight:600;color:#991B1B;">오류가 발생했습니다</div>' +
      '<div style="font-size:12px;color:#666;margin-top:2px;">' + (evt.data.msg || '').substring(0, 100) + '</div></div>' +
      '<button onclick="retryPipeline()" style="padding:8px 16px;background:#2563EB;color:#fff;border:none;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;white-space:nowrap;">🔄 이어서 재시작</button>' +
      '<button onclick="restartFromScratch()" style="padding:8px 14px;background:#fff;color:#1D4ED8;border:1.5px solid #93C5FD;border-radius:8px;font-size:12px;cursor:pointer;white-space:nowrap;">처음부터 시작</button>';
    card2.appendChild(banner);
  }

  if (evt.type === 'error_recovery') {
    setStopButtonsDisabled(true);
    if (eventSource) eventSource.close();
    const d = evt.data;
    document.getElementById('recoveryErrorMsg').textContent = '오류: ' + d.error;
    const resumeBtn = document.getElementById('btnResume');
    const cpInfo = document.getElementById('recoveryCheckpointInfo');
    if (d.has_checkpoint) {
      resumeBtn.style.display = 'block';
      cpInfo.classList.remove('hidden');
      cpInfo.innerHTML = '💾 저장된 체크포인트<br>프로젝트: <strong>' + d.checkpoint_project +
        '</strong><br>단계: ' + d.checkpoint_stage + '<br>저장 시각: ' + d.checkpoint_saved_at;
    } else {
      resumeBtn.style.display = 'none';
      cpInfo.classList.add('hidden');
    }
    document.getElementById('recovery-modal').classList.add('open');
  }
}

let _countdownTimer = null;
let _countdownSec = 0;

function stopCountdown() {
  if (_countdownTimer) { clearInterval(_countdownTimer); _countdownTimer = null; }
  var el = document.getElementById('countdownLabel');
  if (el) el.textContent = '';
}

function startCountdown(seconds) {
  if (_countdownTimer) clearInterval(_countdownTimer);
  _countdownSec = seconds;
  _countdownTimer = setInterval(function() {
    _countdownSec--;
    if (_countdownSec <= 0) {
      clearInterval(_countdownTimer);
      _countdownTimer = null;
      var el = document.getElementById('countdownLabel');
      if (el) el.textContent = '';
      return;
    }
    var min = Math.floor(_countdownSec / 60);
    var sec = _countdownSec % 60;
    var text = min > 0 ? min + '분 ' + sec + '초' : sec + '초';
    var el = document.getElementById('countdownLabel');
    if (el) el.textContent = '  (약 ' + text + ' 후 완료)';
  }, 1000);
}

function updateProgress(label, pct, etaSec) {
  document.getElementById('stageLabel').textContent = label + ' (' + pct + '%)';
  document.getElementById('progressBar').style.width = pct + '%';
  // 서버가 eta_sec을 명시하면 우선 사용
  if (typeof etaSec === 'number' && etaSec > 0) {
    startCountdown(etaSec);
    return;
  }
  // 폴백: label 기준으로 대략적인 예상 시간 추정 (pct 정확 매칭 의존 제거)
  const txt = String(label || '').toLowerCase();
  let estSec = 0;
  if (txt.includes('파싱')) estSec = 20;
  else if (txt.includes('정책·기능') || txt.includes('정책+기능') || txt.includes('통합 추출')) estSec = 90;
  else if (txt.includes('인벤토리')) estSec = 60;
  else if (txt.includes('정책') || txt.includes('기능 목록')) estSec = 60;
  else if (txt.includes('분류표')) estSec = 45;
  else if (txt.includes('gate') || txt.includes('검토 대기')) estSec = 0;  // 사용자 대기
  else if (txt.includes('tc 작성') || txt.includes('tc 생성')) estSec = 0; // 가변 — 중분류별로 다름
  else if (txt.includes('품질 검토') || txt.includes('누락')) estSec = 30;
  else if (txt.includes('보강')) estSec = 30;
  else if (txt.includes('excel') || txt.includes('빌드')) estSec = 60;
  else if (txt.includes('완료')) estSec = 0;
  if (estSec > 0) startCountdown(estSec);
  else stopCountdown();
}

// 백엔드 stage(1~5) + label 조합으로 현재 활성 substep 계산
// substep 구성: sub1 파싱 · sub2 정책·기능 · sub3 분류 · sub4 검토 · sub5 TC 생성 · sub6 Excel
function updateSubsteps(stage, label) {
  const ids = ['sub1','sub2','sub3','sub4','sub5','sub6'];
  // 모두 초기화
  ids.forEach(id => {
    const el = document.getElementById(id);
    if (el) el.classList.remove('active','done');
  });

  // stage와 label로 현재 substep 인덱스 결정 (1-based)
  const txt = String(label || '').toLowerCase();
  let active = 1;
  if (stage <= 1) {
    active = 1;  // 파싱
  } else if (stage === 2) {
    // 라벨 기준으로 정책·기능 vs 분류 구분
    if (txt.includes('분류표')) active = 3;  // "분류표 생성"
    else active = 2;  // "정책 분석", "기능 목록", "정책·기능 통합 추출" 등
  } else if (stage === 3) {
    active = 3;  // Gate 대기 — 분류 완료 직후
  } else if (stage === 4) {
    // 검토 vs TC 생성 구분
    if (txt.includes('검토') || txt.includes('보강')) active = 4;
    else if (txt.includes('excel') || txt.includes('빌드')) active = 6;
    else active = 5;  // "TC 작성 시작", "TC 작성: ..."
  } else if (stage === 5) {
    // Excel 빌드 또는 완료
    if (txt.includes('완료')) {
      active = 6;
      // 완료 시 sub6까지 done 처리
      for (let i = 1; i <= 6; i++) {
        const el = document.getElementById('sub' + i);
        if (el) el.classList.add('done');
      }
      return;
    }
    active = 6;
  }

  // 이전 substep들은 done, 현재만 active
  for (let i = 1; i < active; i++) {
    const el = document.getElementById('sub' + i);
    if (el) el.classList.add('done');
  }
  const cur = document.getElementById('sub' + active);
  if (cur) cur.classList.add('active');
}

function addLog(msg, isError=false) {
  const box = document.getElementById('logBox');
  const line = document.createElement('div');
  line.className = 'log-line' + (isError ? ' error' : '');
  line.textContent = '[' + new Date().toLocaleTimeString() + '] ' + msg;
  box.appendChild(line);
  box.scrollTop = box.scrollHeight;
}

// NOTE: addTcLog는 제거됨 — 통합 card2의 addLog 하나만 사용
// NOTE: 이전에 있던 "대분류 체크리스트"(parseDomains/renderDomainChecklist/selectAllDomains/getSelectedDomains)는
// 제거되었습니다. 범위 지정은 Step 1의 "TC 생성 범위"(focus_area)로 일원화되었습니다.

async function exportGateExcel() {
  const content = document.getElementById('gateContent').value.trim();
  if (!content) { alert('분류표 내용이 없습니다.'); return; }
  showToast('📥 Excel 생성 중...');
  try {
    const r = await fetch('/export-gate', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ content, mode: window._gateMode || 'new' })
    });
    if (!r.ok) { showToast('❌ Excel 생성 실패', 'error'); return; }
    const blob = await r.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url; a.download = 'gate_review.xlsx';
    document.body.appendChild(a); a.click();
    document.body.removeChild(a);
    URL.revokeObjectURL(url);
    showToast('✅ Excel 다운로드 완료!');
  } catch(e) {
    showToast('❌ 오류: ' + e.message, 'error');
  }
}

// ── 입력 소스 화면 식별자 (v0.9.13) — gate 이벤트로 받은 source_scr_map 안내 ──
// ── 원칙 G — 중복 의심 TC 알림 (v0.9.17~) ──
// duplicate_warning SSE 이벤트로 받은 결과를 결과 카드(card5) 에 렌더링
function renderDuplicateNotice(report) {
  var el = document.getElementById('duplicateNotice');
  if (!el) return;
  if (!report || !report.patterns || report.patterns.length === 0) {
    el.style.display = 'none';
    return;
  }
  var patterns = report.patterns;
  var totalDup = report.total_duplicates || 0;

  var html = '<div style="display:flex;align-items:center;gap:8px;margin-bottom:10px;">';
  html += '<span style="font-size:18px;">⚠️</span>';
  html += '<strong style="font-size:14px;color:#92400E;">중복 의심 패턴 ' + patterns.length + '개 발견 (원칙 G)</strong>';
  html += '<span style="font-size:11px;color:#78350F;">통합 시 약 ' + totalDup + '개 TC 감소 가능</span>';
  html += '</div>';
  html += '<div style="font-size:12px;color:#78350F;margin-bottom:10px;line-height:1.55;">';
  html += '같은 그룹 내에 동일한 비기능/에러 패턴이 여러 화면에 반복 작성되었습니다. ';
  html += '아래 패턴 중 통합이 필요한 것을 검토하세요. ';
  html += '<strong>스펙 표에 명시된 에러 케이스는 보존</strong>하고, AI 가 추가한 일반 비기능 TC 만 통합하는 것이 원칙입니다.';
  html += '</div>';

  // 패턴별 상세
  html += '<div style="background:#FFFFFF;border:1px solid #FBBF24;border-radius:8px;overflow:hidden;">';
  patterns.forEach(function(p, idx) {
    var border = idx > 0 ? 'border-top:1px solid #FCD34D;' : '';
    html += '<div style="padding:10px 12px;' + border + '">';
    html += '<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;">';
    html += '<span style="font-size:13px;font-weight:700;color:#92400E;">🔁 ' + escapeHtml(p.pattern) + '</span>';
    html += '<span style="font-size:11px;color:#78350F;">— ' + escapeHtml(p.major) + '</span>';
    html += '<span style="margin-left:auto;font-size:11px;background:#FEF3C7;color:#92400E;padding:2px 8px;border-radius:4px;font-weight:600;">' + p.count + '개 TC</span>';
    html += '</div>';
    html += '<div style="font-size:11.5px;color:#374151;line-height:1.6;">';
    html += '<strong>유지 권장</strong> (대표 화면): <code style="background:#D1FAE5;padding:1px 5px;border-radius:3px;color:#065F46;">' + escapeHtml(p.keep) + '</code><br>';
    html += '<strong>통합 후보</strong>: ';
    html += (p.remove || []).map(function(id) {
      return '<code style="background:#FEE2E2;padding:1px 5px;border-radius:3px;color:#991B1B;">' + escapeHtml(id) + '</code>';
    }).join(' ');
    html += '<br><span style="font-size:10px;color:#6B7280;">└ TC: ';
    html += (p.tcs || []).map(function(t) { return escapeHtml(t.id) + ' ' + escapeHtml(t.title || '').substring(0, 40); }).join(' / ');
    html += '</span>';
    html += '</div>';
    html += '</div>';
  });
  html += '</div>';

  // ── v0.9.22: 자동 통합 액션 (단계 가이드 단순화) ──
  html += '<div style="margin-top:14px;padding:12px;background:#ECFDF5;border:1.5px solid #6EE7B7;border-radius:8px;">';
  html += '<div style="font-size:12.5px;font-weight:700;color:#065F46;margin-bottom:6px;display:flex;align-items:center;gap:6px;">';
  html += '<span>💡</span><span>지금 바로 정리하시겠어요?</span>';
  html += '</div>';
  html += '<div style="font-size:11.5px;color:#374151;line-height:1.55;margin-bottom:10px;">';
  html += '아래 <strong>"🔁 자동 통합 + Excel 재생성"</strong> 버튼을 누르면 AI 가 통합 후보 TC 들을 제거하고 새 Excel 을 만들어줍니다. <strong>원본 Excel 은 보존</strong>됩니다 (파일명에 <code style="background:#FFFFFF;padding:1px 4px;border-radius:3px;">_merged</code> 추가).';
  html += '</div>';
  html += '<div style="display:flex;gap:8px;flex-wrap:wrap;">';
  html += '<button onclick="autoMergeTcs()" id="autoMergeTcsBtn" style="padding:9px 16px;background:#10B981;color:#FFFFFF;border:none;border-radius:6px;font-size:13px;font-weight:700;cursor:pointer;">🔁 자동 통합 + Excel 재생성</button>';
  html += '<button onclick="copyDuplicateReportToClipboard()" style="padding:9px 14px;background:#FFFFFF;color:#065F46;border:1.5px solid #6EE7B7;border-radius:6px;font-size:12px;font-weight:600;cursor:pointer;">📋 수동 — 분석 결과 복사</button>';
  html += '</div>';
  html += '<div style="font-size:10.5px;color:#6B7280;margin-top:8px;">자동 통합이 실패하거나 직접 검토하고 싶을 때만 복사 버튼을 사용하세요.</div>';
  html += '</div>';

  el.innerHTML = html;
  el.style.display = '';
}

// ── v0.9.22 Stage 2 — TC 자동 통합 + Excel 재생성 ──
async function autoMergeTcs() {
  var report = window._duplicateReport;
  if (!report || !report.patterns || report.patterns.length === 0) {
    alert('통합할 중복 패턴이 없습니다.');
    return;
  }
  if (!currentSid) { alert('세션이 없습니다.'); return; }

  var totalRemove = report.total_duplicates || 0;
  if (!confirm('AI 가 통합 후보 TC ' + totalRemove + '개를 제거하고 새 Excel 을 생성합니다.\\n원본 Excel 은 보존됩니다 (별도 파일).\\n약 30~60초 소요.\\n계속할까요?')) return;

  var btn = document.getElementById('autoMergeTcsBtn');
  if (btn) {
    btn.disabled = true;
    btn.textContent = '⏳ 통합 중... (30~60초)';
    btn.style.background = '#9CA3AF';
    btn.style.cursor = 'wait';
  }

  try {
    var r = await fetch('/merge-tcs/' + currentSid, {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ duplicate_report: report }),
    });
    var d = await r.json();
    if (!d.ok) {
      alert('통합 실패: ' + (d.error || 'unknown'));
      if (btn) {
        btn.disabled = false;
        btn.textContent = '🔁 자동 통합 + Excel 재생성';
        btn.style.background = '#10B981';
        btn.style.cursor = 'pointer';
      }
      return;
    }

    // 결과 카드 갱신 — 새 파일명, 새 TC 수
    var fileEl = document.getElementById('resultFilename');
    var metaEl = document.getElementById('resultMeta');
    var totalEl = document.getElementById('statTotal');
    var smokeEl = document.getElementById('statSmoke');
    if (fileEl) fileEl.textContent = d.filename;
    if (metaEl) metaEl.textContent = (d.size ? (d.size / 1024).toFixed(1) + ' KB' : '—') + ' · 통합본';
    if (totalEl) totalEl.textContent = d.new_total_tc;
    if (smokeEl) smokeEl.textContent = d.new_smoke_tc;
    currentFilename = d.filename;

    // 알림 박스 → 녹색 완료 표시
    var notice = document.getElementById('duplicateNotice');
    if (notice) {
      notice.style.background = '#ECFDF5';
      notice.style.borderColor = '#6EE7B7';
      notice.innerHTML = '<div style="display:flex;align-items:center;gap:10px;color:#065F46;font-size:13px;">';
      notice.innerHTML += '<span style="font-size:24px;">✅</span>';
      notice.innerHTML += '<div><strong>자동 통합 완료</strong><br>';
      notice.innerHTML += '<span style="font-size:11.5px;">';
      notice.innerHTML += '제거된 TC: <strong>' + d.removed_count + '개</strong> · ';
      notice.innerHTML += '새 Excel: <code style="background:#FFFFFF;padding:1px 5px;border-radius:3px;">' + escapeHtml(d.filename) + '</code><br>';
      notice.innerHTML += '원본 Excel 은 폴더에 그대로 보존되어 있습니다.';
      notice.innerHTML += '</span></div></div>';
    }
    showToast('✅ ' + d.removed_count + '개 TC 통합 완료 — 새 Excel: ' + d.filename, 'success');
  } catch (e) {
    alert('네트워크 오류: ' + e.message);
    if (btn) {
      btn.disabled = false;
      btn.textContent = '🔁 자동 통합 + Excel 재생성';
      btn.style.background = '#10B981';
      btn.style.cursor = 'pointer';
    }
  }
}

function copyDuplicateReportToClipboard() {
  var report = window._duplicateReport;
  if (!report || !report.patterns) return;

  // v0.9.21: 복사되는 텍스트에 사용 안내 + AI 즉시 실행 가능한 명령형 포함
  var text = '# 중복 의심 패턴 통합 요청 (원칙 G)\\n\\n';
  text += '> 사용 방법: 이 텍스트를 분류표 검토(Step 3) 의 하단 AI 도우미 채팅에 붙여넣고 전송하세요.\\n';
  text += '> AI 가 자동으로 통합 후보 TC 들을 제거하고 대표 화면에만 1개로 통합합니다.\\n\\n';
  text += '---\\n\\n';
  text += '## 분석 요약\\n';
  text += '- 총 ' + report.patterns.length + '개 중복 패턴 발견\\n';
  text += '- 통합 시 약 ' + (report.total_duplicates || 0) + '개 TC 감소 가능\\n\\n';
  text += '## 통합 지시\\n\\n';
  text += '아래 패턴별로 **통합 후보 TC 들을 분류표에서 제거**해줘. **유지 권장 TC** 는 그대로 두고, 그 TC 의 비고에 \\'[통합] 그룹 단위 비기능 검증\\' 태그를 추가해줘.\\n\\n';
  report.patterns.forEach(function(p, idx) {
    text += '### ' + (idx+1) + '. ' + p.pattern + ' (' + p.major + ')\\n';
    text += '- ✅ 유지: `' + p.keep + '`\\n';
    text += '- ❌ 제거: ' + (p.remove || []).map(function(id) { return '`' + id + '`'; }).join(', ') + '\\n';
    if (p.tcs && p.tcs.length) {
      text += '- 상세:\\n';
      p.tcs.forEach(function(t) {
        text += '  - ' + t.id + ' — ' + (t.middle || '') + ' / ' + (t.title || '').substring(0, 60) + '\\n';
      });
    }
    text += '\\n';
  });
  text += '---\\n\\n';
  text += '**스펙 표에 명시된 에러 케이스는 보존**하고, 위 비기능 패턴(네트워크/타임아웃/로딩 등)만 통합해줘.\\n';

  if (navigator.clipboard) {
    navigator.clipboard.writeText(text).then(function() {
      showToast('📋 복사 완료 — Step 3 의 AI 도우미 채팅에 Cmd+V 후 전송하세요', 'success');
    });
  } else {
    alert(text);
  }
}

// 헬퍼 — 안전한 HTML escape (이미 있으면 재사용 가능)
if (typeof escapeHtml === 'undefined') {
  window.escapeHtml = function(s) {
    if (s === null || s === undefined) return '';
    return String(s).replace(/[&<>"']/g, function(c) {
      return { '&':'&amp;', '<':'&lt;', '>':'&gt;', '"':'&quot;', "'":'&#39;' }[c];
    });
  };
}

// ── v0.9.22 Stage 1 — 분류표 중복 가능성 예측 알림 ──
function renderClassifyDuplicateNotice(prediction) {
  var el = document.getElementById('classifyDuplicateNotice');
  if (!el) return;
  prediction = prediction || { predictions: [], high_risk_count: 0 };
  if (!prediction.high_risk_count || !prediction.predictions || prediction.predictions.length === 0) {
    el.style.display = 'none';
    return;
  }
  var html = '<div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;">';
  html += '<span style="font-size:18px;">⚠️</span>';
  html += '<strong style="font-size:13.5px;color:#92400E;">중복 가능성 ' + prediction.predictions.length + '개 그룹 발견 (Stage 1 — TC 작성 전)</strong>';
  html += '</div>';
  html += '<div style="font-size:11.5px;color:#78350F;margin-bottom:10px;line-height:1.55;">';
  html += '같은 대분류 내에 중분류가 많거나 동일 비기능 키워드가 반복되어, AI 가 TC 작성 시 비기능 TC 를 여러 화면에 반복 작성할 가능성이 있습니다. 미리 분류표를 정리하면 TC 결과가 깔끔해집니다.';
  html += '</div>';

  // 예측별 상세
  html += '<div style="background:#FFFFFF;border:1px solid #FBBF24;border-radius:8px;overflow:hidden;margin-bottom:10px;">';
  prediction.predictions.forEach(function(p, idx) {
    var border = idx > 0 ? 'border-top:1px solid #FCD34D;' : '';
    html += '<div style="padding:10px 12px;' + border + '">';
    html += '<div style="font-size:12.5px;font-weight:700;color:#92400E;margin-bottom:4px;">';
    if (p.type === 'shared_minor') {
      html += '🔁 "' + escapeHtml(p.shared_keyword) + '" 키워드 — ' + escapeHtml(p.major);
    } else {
      html += '🔁 그룹 크기 — ' + escapeHtml(p.major) + ' (중분류 ' + p.middle_count + '개)';
    }
    html += '</div>';
    html += '<div style="font-size:11px;color:#374151;line-height:1.5;">';
    html += '<div>' + escapeHtml(p.predicted_pattern || '') + '</div>';
    html += '<div style="margin-top:2px;color:#6B7280;">→ ' + escapeHtml(p.recommendation || '') + '</div>';
    if (p.shared_in_middles && p.shared_in_middles.length) {
      html += '<div style="margin-top:4px;font-size:10.5px;color:#6B7280;">└ 영향 중분류: ' + p.shared_in_middles.map(escapeHtml).join(', ') + '</div>';
    }
    html += '</div>';
    html += '</div>';
  });
  html += '</div>';

  // 액션 버튼
  html += '<div style="display:flex;gap:8px;flex-wrap:wrap;align-items:center;">';
  html += '<button onclick="autoMergeClassification()" style="padding:8px 14px;background:#10B981;color:#FFFFFF;border:none;border-radius:6px;font-size:12.5px;font-weight:700;cursor:pointer;">🔁 분류표 자동 정리</button>';
  html += '<button onclick="dismissClassifyDuplicateNotice()" style="padding:8px 14px;background:#FFFFFF;color:#92400E;border:1.5px solid #FCD34D;border-radius:6px;font-size:12px;font-weight:600;cursor:pointer;">✕ 무시 (그대로 진행)</button>';
  html += '<span style="font-size:11px;color:#78350F;margin-left:6px;">자동 정리 실패 시 미니 채팅에서 직접 수정 가능합니다.</span>';
  html += '</div>';

  el.innerHTML = html;
  el.style.display = '';
}

function dismissClassifyDuplicateNotice() {
  var el = document.getElementById('classifyDuplicateNotice');
  if (el) el.style.display = 'none';
}

async function autoMergeClassification() {
  var prediction = window._classifyPrediction;
  if (!prediction || !prediction.predictions || prediction.predictions.length === 0) {
    alert('통합할 예측 정보가 없습니다.');
    return;
  }
  if (!currentSid) { alert('세션이 없습니다.'); return; }
  if (!confirm('AI 가 분류표를 정리합니다.\\n비기능 소분류(네트워크/타임아웃/로딩 등)를 그룹 대표 화면 1개에 통합합니다.\\n계속할까요?')) return;

  var notice = document.getElementById('classifyDuplicateNotice');
  if (notice) {
    notice.innerHTML = '<div style="display:flex;align-items:center;gap:10px;color:#92400E;font-size:13px;font-weight:600;">⏳ AI 가 분류표를 정리하고 있어요... (10~30초)</div>';
  }

  try {
    var currentDoc = document.getElementById('gateContent').value || '';
    var r = await fetch('/merge-classification/' + currentSid, {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({
        current_doc: currentDoc,
        predictions: prediction.predictions,
      }),
    });
    var d = await r.json();
    if (!d.ok) {
      alert('분류표 정리 실패: ' + (d.error || 'unknown'));
      // 알림 박스 복원
      renderClassifyDuplicateNotice(prediction);
      return;
    }
    // 분류표 갱신
    document.getElementById('gateContent').value = d.updated_doc;
    renderGateViewer(d.updated_doc);
    // 미니 채팅에 안내 메시지 추가
    if (typeof addGateChatMsg === 'function') {
      addGateChatMsg('system', '✅ 분류표 자동 정리 완료 — ' + (d.predictions_resolved || 0) + '개 패턴 처리. ' + (d.reply || ''));
    }
    // 알림 박스 → 녹색 완료 표시로 변경
    if (notice) {
      notice.style.background = '#ECFDF5';
      notice.style.borderColor = '#6EE7B7';
      notice.innerHTML = '<div style="display:flex;align-items:center;gap:10px;color:#065F46;font-size:13px;">';
      notice.innerHTML += '<span style="font-size:18px;">✅</span>';
      notice.innerHTML += '<div><strong>분류표 자동 정리 완료</strong><br>';
      notice.innerHTML += '<span style="font-size:11.5px;">' + escapeHtml(d.reply || '') + ' 분류표를 검토 후 승인하세요.</span></div>';
      notice.innerHTML += '</div>';
    }
    showToast('✅ 분류표 자동 정리 완료', 'success');
  } catch (e) {
    alert('네트워크 오류: ' + e.message);
    renderClassifyDuplicateNotice(prediction);
  }
}

function renderSourceScrNotice(scrMap) {
  var el = document.getElementById('sourceScrNotice');
  if (!el) return;
  scrMap = scrMap || {};
  var entries = Object.entries(scrMap);
  if (entries.length === 0) {
    // 입력 소스에서 화면 식별자 발견 못함 — 빈 칸 정책 안내
    el.style.display = '';
    el.style.background = '#FEF3C7';
    el.style.border = '1px solid #FCD34D';
    el.style.color = '#78350F';
    el.innerHTML =
      '<strong>📌 화면 코드 안내</strong> — ' +
      '입력 소스에서 화면 식별자(<code style="background:#FFFFFF;padding:1px 5px;border-radius:3px;">SCR-NNN</code> 형식)가 발견되지 않았습니다. ' +
      '<br>Excel 의 <strong>화면 코드</strong> 컬럼은 <strong>빈 칸</strong>으로 출력됩니다 (정상). ' +
      '식별자가 필요하면 입력 파일명을 <code style="background:#FFFFFF;padding:1px 5px;border-radius:3px;">SCR-001.md</code> 형식으로 변경하거나, 본문 H1 에 <code style="background:#FFFFFF;padding:1px 5px;border-radius:3px;">SCR-001:</code> 처럼 명시하세요.';
  } else {
    // 매핑 발견됨 — 사용자에게 안내
    el.style.display = '';
    el.style.background = '#ECFDF5';
    el.style.border = '1px solid #6EE7B7';
    el.style.color = '#065F46';
    var rows = entries.map(function(kv) {
      return '<code style="background:#FFFFFF;padding:1px 5px;border-radius:3px;font-size:11px;">' + kv[0] + '</code> → <strong>' + kv[1] + '</strong>';
    }).join(' &nbsp;·&nbsp; ');
    el.innerHTML =
      '<strong>📌 입력 소스 화면 코드 매핑 (' + entries.length + '개)</strong><br>' +
      rows +
      '<br><span style="font-size:11px;opacity:0.85;">생성된 TC 의 <strong>화면 코드</strong> 컬럼은 위 매핑을 따릅니다. 임의 추론 없음.</span>';
  }
}

// ── Human Gate 채팅 ──────────────────────────────────────────────────────────
let gateChatHistory = [];  // [{role, content}, ...]

function initGateChat(docContent) {
  // 문서 저장
  document.getElementById('gateContent').value = docContent;
  // 승인 버튼 상태 리셋
  var approveBtn = document.querySelector('#card3 .btn-success');
  if (approveBtn) { approveBtn.disabled = false; approveBtn.textContent = '✅ 승인 및 TC 생성 시작'; }
  setStopButtonsDisabled(false);
  // SuiteCode 섹션 초기화 (새 분류표마다 리셋)
  var suiteSection = document.getElementById('suiteCodeSection');
  if (suiteSection) { suiteSection.innerHTML = ''; suiteSection.style.display = 'none'; }
  // Viewer 렌더링
  renderGateViewer(docContent);
  // 채팅 초기화
  gateChatHistory = [];
  const msgs = document.getElementById('gateChatMessages');
  msgs.innerHTML = '';
  addGateChatMsg('assistant',
    'AI가 문서를 준비했습니다. 수정이 필요한 부분을 채팅으로 알려주세요. ' +
    '예) "AUTH 대분류의 비밀번호 변경 케이스 삭제해줘" / "PROD 대분류 이름을 상품관리로 바꿔줘"');
  document.getElementById('gateChatInput').focus();
}

function mdToHtml(md) {
  // 라인 단위 마크다운 → HTML (Python 문자열 호환 — 백슬래시 최소화)
  const lines = md.split('\\n');
  let out = [];
  let inList = false;
  for (let i = 0; i < lines.length; i++) {
    let line = lines[i]
      .replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;');
    // 헤딩
    if (line.startsWith('#### ')) {
      if (inList) { out.push('</ul>'); inList = false; }
      out.push('<h4 style="font-size:13px;font-weight:600;margin:8px 0 3px;">' + line.slice(5) + '</h4>');
    } else if (line.startsWith('### ')) {
      if (inList) { out.push('</ul>'); inList = false; }
      out.push('<h3>' + line.slice(4) + '</h3>');
    } else if (line.startsWith('## ')) {
      if (inList) { out.push('</ul>'); inList = false; }
      out.push('<h2>' + line.slice(3) + '</h2>');
    } else if (line.startsWith('# ')) {
      if (inList) { out.push('</ul>'); inList = false; }
      out.push('<h2 style="font-size:16px;">' + line.slice(2) + '</h2>');
    } else if (line.startsWith('> ')) {
      if (inList) { out.push('</ul>'); inList = false; }
      out.push('<blockquote>' + line.slice(2) + '</blockquote>');
    } else if (line.match(/^[-*] /) || line.match(/^\d+\. /)) {
      if (!inList) { out.push('<ul>'); inList = true; }
      const txt = line.replace(/^[-*] /, '').replace(/^\d+\. /, '');
      out.push('<li>' + txt + '</li>');
    } else if (line.trim() === '') {
      if (inList) { out.push('</ul>'); inList = false; }
      out.push('<br>');
    } else {
      if (inList) { out.push('</ul>'); inList = false; }
      out.push('<p style="margin:3px 0;">' + line + '</p>');
    }
  }
  if (inList) out.push('</ul>');
  // 인라인: **bold**, `code`
  let html = out.join('');
  html = html.replace(/\*\*([^*]+)\*\*/g, '<strong>$1</strong>');
  html = html.replace(/`([^`]+)`/g, '<code>$1</code>');
  return html;
}

function extractCategorySummary(mdText) {
  // 분류표에서 대분류/중분류/소분류 구조 추출
  // 결과 구조 (중분류별 그룹핑 — A안):
  //   [{
  //     major: '대분류 이름',
  //     middle: '중분류 이름',
  //     minors: ['소분류 1', '소분류 2', ...]
  //   }, ...]
  // 소분류는 다음 두 가지 패턴 모두 지원:
  //   1. `#### 소분류` 헤더 + 그 아래 `- 항목: 설명` 불릿
  //   2. `### 중분류` 직후 바로 `- 항목` 불릿 (헤더 없음)
  // `#### 소분류` 헤더 자체는 데이터가 아니므로 제외.
  var lines = mdText.split('\\n');
  var rows = [];
  var curMajor = '';
  var curMiddle = '';
  var curMinors = [];
  var inMinorSection = false;  // `#### 소분류` 헤더 직후 여부

  function flushMiddle() {
    if (curMajor && curMiddle) {
      rows.push({ major: curMajor, middle: curMiddle, minors: curMinors });
    }
    curMiddle = '';
    curMinors = [];
    inMinorSection = false;
  }

  for (var i = 0; i < lines.length; i++) {
    var line = lines[i].trim();
    // 대분류
    if (line.match(/^##\s+/) && !line.match(/^###\s+/)) {
      flushMiddle();
      curMajor = line.replace(/^##\s+/, '').replace(/\*\*/g, '').replace(/^대분류[:\s]*/, '').trim();
    }
    // 중분류
    else if (line.match(/^###\s+/) && !line.match(/^####\s+/)) {
      flushMiddle();
      curMiddle = line.replace(/^###\s+/, '').replace(/\*\*/g, '').replace(/^중분류[:\s]*/, '').trim();
      inMinorSection = false;
    }
    // 소분류 섹션 헤더 — '#### 소분류' 같은 헤더는 그 자체로 데이터 아님
    else if (line.match(/^####\s+/)) {
      inMinorSection = true;
    }
    // 불릿 항목 — 중분류 안에 있을 때만 소분류로 수집
    else if ((line.startsWith('- ') || line.startsWith('* ')) && curMiddle) {
      var txt = line.replace(/^[-*]\s+/, '').replace(/\*\*/g, '').trim();
      if (txt) curMinors.push(txt);
    }
  }
  flushMiddle();
  return rows;
}

function updateTcIdPreview(input) {
  // SuiteCode 모드: 같은 대분류(domIdx)에 속한 모든 미리보기 셀에 동일 코드 적용
  var code = input.value.trim().toUpperCase();
  var pcode = input.dataset.pcode || 'SC';
  // System Generated 모드면 사용자가 SuiteCode 바꿔도 ScreenCode 자동 적용이 우선 — 무시
  if (window._autoScreenCode) return;
  // 입력 input의 행을 기준으로, 다음 SuiteCode input(다음 대분류 행)을 만나기 전까지의 모든
  // tcIdPreview_* 셀에 적용
  var row = input.closest('tr');
  if (!row || !row.parentNode) return;
  var rows = Array.prototype.slice.call(row.parentNode.children);
  var startIdx = rows.indexOf(row);
  for (var i = startIdx; i < rows.length; i++) {
    var r = rows[i];
    if (i > startIdx && r.querySelector('input.suite-code-input')) break;  // 다음 대분류 시작 — 중단
    var prev = r.querySelector('[id^="tcIdPreview_"]');
    if (prev) prev.textContent = code ? pcode + '-' + code + '-001' : '-';
  }
}

// ── Screen Code Map 캐시 (Auto 모드 미리보기용) ──
var _screenCodeMap = null;
var _screenCodeMapLoading = null;

async function loadScreenCodeMap(projectName) {
  if (_screenCodeMap) return _screenCodeMap;
  if (_screenCodeMapLoading) return _screenCodeMapLoading;
  _screenCodeMapLoading = fetch('/screen-code-map?project=' + encodeURIComponent(projectName || ''))
    .then(r => r.json())
    .then(d => {
      _screenCodeMap = d.ok ? d : { map: {}, project_code: 'SC', screen_based_default: false };
      _screenCodeMapLoading = null;
      return _screenCodeMap;
    })
    .catch(() => {
      _screenCodeMap = { map: {}, project_code: 'SC', screen_based_default: false };
      _screenCodeMapLoading = null;
      return _screenCodeMap;
    });
  return _screenCodeMapLoading;
}

// Middle name → ScreenCode 프론트엔드 해석 (백엔드 resolve_screen_code와 동일 규칙, v0.9.7b)
// 규칙:
//   - 1단어:     앞 3자
//   - 2단어:     각 단어 앞 2자 연결
//   - 3단어 이상: 앞 단어들 2자 + 마지막 단어 1자
//   - 최대 8자 절단, 대문자화
// 충돌 처리는 백엔드에서만 수행 — FE 미리보기는 기본 파생값만 표시.
function resolveScreenCodeFE(middleName, screenMap) {
  if (!middleName) return 'SCR';
  if (screenMap && screenMap[middleName]) {
    var e = screenMap[middleName];
    if (typeof e === 'object' && e) return e.code || 'SCR';
    return e;
  }
  var words = String(middleName).split(/[\s\-_]+/)
    .map(function(w) { return w.replace(/[^A-Za-z]/g, ''); })
    .filter(function(w) { return w.length > 0; });
  if (words.length === 0) {
    var h = 0;
    for (var i = 0; i < middleName.length; i++) {
      h = ((h << 5) - h) + middleName.charCodeAt(i);
      h = h & h;
    }
    return 'MID' + String(Math.abs(h) % 1000).padStart(3, '0');
  }
  var base;
  if (words.length === 1) {
    base = words[0].substring(0, 3);
  } else if (words.length === 2) {
    base = words[0].substring(0, 2) + words[1].substring(0, 2);
  } else {
    // 3단어 이상: 앞 단어들 앞 2자 + 마지막 단어 앞 1자
    var head = words.slice(0, -1).map(function(w) { return w.substring(0, 2); }).join('');
    var tail = words[words.length - 1].substring(0, 1);
    base = head + tail;
  }
  return base.substring(0, 8).toUpperCase();
}

// 중분류명 정제
function cleanMiddleName(raw) {
  return String(raw || '').replace(/중분류[:\s]*/g, '').replace(/\(.*?\)/g, '').trim();
}

// 통합 TC ID 모드 설정 — 상단 독립 패널의 체크박스 핸들러
// (기존 toggleAutoScreenCode는 이 함수의 alias로 유지)
function setTcIdMode(systemGenerated) {
  window._autoScreenCode = !!systemGenerated;

  // 상단 독립 패널 — 설명 텍스트 업데이트
  var desc = document.getElementById('tcIdModeDesc');
  if (desc) {
    if (systemGenerated) {
      desc.innerHTML =
        '✅ <strong style="color:#1D4ED8;">System Generated 모드 ON</strong> — ' +
        '시스템이 각 화면을 <code style="background:#FFFFFF;padding:1px 5px;border-radius:3px;border:1px solid #93C5FD;font-family:monospace;">screen_code_map.md</code>에 따라 ' +
        '<code style="background:#FFFFFF;padding:1px 5px;border-radius:3px;border:1px solid #93C5FD;font-family:monospace;">SM-SPL-001</code>, ' +
        '<code style="background:#FFFFFF;padding:1px 5px;border-radius:3px;border:1px solid #93C5FD;font-family:monospace;">SM-LGI-001</code> 식으로 ' +
        '자동 매핑합니다. 아래 SuiteCode 입력란은 비활성화됩니다.';
    } else {
      desc.innerHTML =
        '✏️ <strong style="color:#B45309;">Manual 모드</strong> — ' +
        '아래 <strong>TC 분류 요약</strong> 표에서 각 대분류의 <strong>SuiteCode</strong>를 직접 입력하세요. ' +
        '입력한 값으로 <code style="background:#FFFFFF;padding:1px 5px;border-radius:3px;border:1px solid #FCD34D;font-family:monospace;">SM-{SuiteCode}-001</code> 형태의 TC ID가 생성됩니다.';
    }
  }

  // 상단 패널 라벨 색상 표현
  var label = document.getElementById('tcIdModeLabel');
  if (label) {
    label.style.background = systemGenerated ? '#3B82F6' : '#FFFFFF';
    label.style.color = systemGenerated ? '#FFFFFF' : '#1E3A5F';
    label.style.borderColor = systemGenerated ? '#1D4ED8' : '#93C5FD';
  }

  // 분류 요약 테이블 내 체크박스도 동기화 (역순 동기화 루프 방지용 이벤트 없이 값만 설정)
  var innerToggle = document.getElementById('autoScreenCodeToggle');
  if (innerToggle && innerToggle.checked !== systemGenerated) {
    innerToggle.checked = systemGenerated;
  }

  // SuiteCode 입력란 활성/비활성 + 스타일
  var inputs = document.querySelectorAll('.suite-code-input');
  inputs.forEach(function(inp) {
    inp.disabled = systemGenerated;
    inp.style.opacity = systemGenerated ? '0.4' : '1';
    inp.style.background = systemGenerated ? '#F3F4F6' : '#FFFFFF';
    inp.title = systemGenerated ? '자동 모드에서는 편집할 수 없습니다. 각 중분류는 screen_code_map에서 자동 매핑됩니다.' : '';
  });

  // 미리보기 재렌더
  var previewCells = document.querySelectorAll('[id^="tcIdPreview_"]');
  previewCells.forEach(function(cell) {
    var idx = cell.id.replace('tcIdPreview_', '');
    rerenderPreviewForRow(idx, systemGenerated);
  });

  // 헬프 문구 토글 (요약 테이블 내)
  var helpNormal = document.getElementById('suiteCodeHelpNormal');
  var helpAuto = document.getElementById('suiteCodeHelpAuto');
  if (helpNormal) helpNormal.style.display = systemGenerated ? 'none' : '';
  if (helpAuto) helpAuto.style.display = systemGenerated ? '' : 'none';
}

// ─ 분류 요약 표 — 소분류 행 개별/전체 토글 ─
function toggleMinorRow(idx) {
  var row = document.getElementById('minorRow_' + idx);
  var icon = document.getElementById('minorToggleIcon_' + idx);
  if (!row) return;
  var isHidden = row.style.display === 'none';
  row.style.display = isHidden ? '' : 'none';
  if (icon) icon.textContent = isHidden ? '▼' : '▶';
}

function toggleAllMinors(open) {
  // 모든 종속 행 일괄 토글
  var rows = document.querySelectorAll('.tc-summary-row-nested');
  rows.forEach(function(row) {
    row.style.display = open ? '' : 'none';
  });
  // 모든 토글 아이콘도 동기화
  var icons = document.querySelectorAll('[id^="minorToggleIcon_"]');
  icons.forEach(function(icon) {
    icon.textContent = open ? '▼' : '▶';
  });
}

// 기존 호환: 테이블 내 체크박스가 호출하는 함수 → 상단 토글과 동기화
function toggleAutoScreenCode(checkbox) {
  var topToggle = document.getElementById('tcIdModeToggle');
  if (topToggle) topToggle.checked = checkbox.checked;
  setTcIdMode(checkbox.checked);
}

function rerenderPreviewForRow(idx, auto) {
  // A안: 행마다 단일 중분류. data-middle 속성 사용.
  var cell = document.getElementById('tcIdPreview_' + idx);
  if (!cell) return;
  var pcode = cell.dataset.pcode || 'SC';
  if (auto) {
    // ScreenCode 자동 — 중분류 이름으로 코드 파생
    var mid = cell.dataset.middle || '';
    var mapData = (_screenCodeMap && _screenCodeMap.map) ? _screenCodeMap.map : {};
    if (!mid) {
      cell.textContent = '-';
    } else {
      var code = resolveScreenCodeFE(mid, mapData);
      cell.textContent = pcode + '-' + code + '-001';
    }
  } else {
    // Suite-based (수동): 같은 행/그룹의 SuiteCode input 값 사용
    // 행을 거슬러 올라가며 가장 가까운 SuiteCode input(같은 대분류 그룹) 검색
    var row = cell.closest('tr');
    var input = null;
    while (row) {
      input = row.querySelector('input.suite-code-input');
      if (input) break;
      row = row.previousElementSibling;
    }
    var code = input ? input.value.trim().toUpperCase() : '';
    cell.textContent = code ? pcode + '-' + code + '-001' : '-';
  }
}

function renderGateViewer(mdText) {
  var viewer = document.getElementById('gateViewer');
  // 분류 요약 카드
  var cats = extractCategorySummary(mdText);
  var summaryHtml = '';
  if (cats.length > 0) {
    var _projName = document.getElementById('projectName').value || '';
    var _pcode = _projName.toLowerCase().indexOf('mobile') >= 0 || _projName.indexOf('모바일') >= 0 ? 'SM' : 'SC';
    // Screen Code Map 비동기 로드 (렌더 후 상단 패널 모드 기준으로 미리보기 재렌더)
    loadScreenCodeMap(_projName).then(function(res) {
      // 상단 독립 패널 토글이 단일 소스 — 거기 상태 기준으로 미리보기 동기화
      var topToggle = document.getElementById('tcIdModeToggle');
      if (topToggle) {
        // 초기 기본값이 아직 결정 안 됐으면 screen_based_default 적용
        if (window._autoScreenCode === undefined) {
          window._autoScreenCode = !!(res && res.screen_based_default);
          topToggle.checked = window._autoScreenCode;
        }
        setTcIdMode(topToggle.checked);
      }
    });

    summaryHtml = '<details id="tcSummaryDetails" open style="background:#F0F9FF;border:1.5px solid #93C5FD;border-radius:10px;margin-bottom:16px;">';
    // <summary> 자체가 헤더 역할 — 클릭 시 펼침/접힘. list-style:none 으로 기본 마커 제거.
    summaryHtml += '<summary id="tcSummaryHeader" style="cursor:pointer;padding:14px 16px;list-style:none;display:flex;justify-content:space-between;align-items:center;gap:10px;user-select:none;">';
    summaryHtml += '<div style="font-size:14px;font-weight:700;color:#1E3A5F;display:flex;align-items:center;gap:8px;">';
    summaryHtml += '<span id="tcSummaryToggleIcon" style="font-size:12px;color:#3B82F6;">▼</span>';
    summaryHtml += '<span>TC 분류 요약</span>';
    summaryHtml += '<span id="tcSummaryFoldHint" style="font-size:11px;color:#6B7280;font-weight:400;">(클릭하여 접기)</span>';
    summaryHtml += '</div>';
    // 자동 적용 토글 (우측) — summary 내부에서 클릭 시 details 토글되지 않도록 stopPropagation 추가
    summaryHtml += '<label onclick="event.stopPropagation();" style="display:inline-flex;align-items:center;gap:6px;font-size:12px;color:#1E3A5F;cursor:pointer;user-select:none;background:#FFFFFF;padding:5px 10px;border:1.5px solid #93C5FD;border-radius:6px;">';
    summaryHtml += '<input type="checkbox" id="autoScreenCodeToggle" onchange="toggleAutoScreenCode(this)" onclick="event.stopPropagation();" style="margin:0;cursor:pointer;">';
    summaryHtml += '<span>🤖 <strong>시스템 규칙 자동 적용</strong></span>';
    summaryHtml += '</label>';
    summaryHtml += '</summary>';
    // 본문 영역 (접히는 부분)
    summaryHtml += '<div style="padding:0 16px 14px 16px;">';
    // 도움말 — 모드별 2종 (display toggle)
    summaryHtml += '<div id="suiteCodeHelpNormal" style="font-size:12px;color:#4B5563;margin-bottom:10px;">각 대분류의 <strong>SuiteCode</strong>를 입력하세요. 순번(001, 002...)은 자동 생성됩니다.<br>예: SuiteCode에 <code style="background:#DBEAFE;padding:1px 4px;border-radius:3px;">GNBF</code> 입력 → TC ID: <code style="background:#DBEAFE;padding:1px 4px;border-radius:3px;">' + _pcode + '-GNBF-001</code> ...</div>';
    summaryHtml += '<div id="suiteCodeHelpAuto" style="display:none;font-size:12px;color:#4B5563;margin-bottom:10px;">🤖 <strong>시스템 규칙 자동 적용</strong> — 각 중분류(화면)가 <code style="background:#DBEAFE;padding:1px 4px;border-radius:3px;">screen_code_map.md</code>에 등록된 ScreenCode로 자동 매핑됩니다.<br>예: Splash → <code style="background:#DBEAFE;padding:1px 4px;border-radius:3px;">' + _pcode + '-SPL-001</code>, Login Options → <code style="background:#DBEAFE;padding:1px 4px;border-radius:3px;">' + _pcode + '-LGI-001</code> ... (화면별 독립 001~)</div>';
    // 표 상단 컨트롤 — 전체 펼치기/접기
    summaryHtml += '<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:6px;gap:8px;flex-wrap:wrap;">';
    summaryHtml += '<div style="font-size:11px;color:#6B7280;">💡 각 중분류 행의 <strong>▼ 소분류</strong> 를 클릭하면 상세 항목이 펼쳐집니다.</div>';
    summaryHtml += '<div style="display:flex;gap:6px;">';
    summaryHtml += '<button type="button" onclick="toggleAllMinors(true)" style="padding:4px 10px;font-size:11px;background:#FFFFFF;color:#1E3A5F;border:1px solid #93C5FD;border-radius:6px;cursor:pointer;font-weight:600;">▼ 전체 펼치기</button>';
    summaryHtml += '<button type="button" onclick="toggleAllMinors(false)" style="padding:4px 10px;font-size:11px;background:#FFFFFF;color:#1E3A5F;border:1px solid #93C5FD;border-radius:6px;cursor:pointer;font-weight:600;">▶ 전체 접기</button>';
    summaryHtml += '</div></div>';
    summaryHtml += '<table class="tc-summary-table" style="width:100%;border-collapse:collapse;font-size:12px;">';
    summaryHtml += '<tr style="background:#DBEAFE;">';
    summaryHtml += '<th style="padding:6px 8px;text-align:left;border:1px solid #93C5FD;">시트명 (Excel)</th>';
    summaryHtml += '<th style="padding:6px 8px;text-align:left;border:1px solid #93C5FD;">SuiteCode</th>';
    summaryHtml += '<th style="padding:6px 8px;text-align:left;border:1px solid #93C5FD;">TC ID 미리보기</th>';
    summaryHtml += '<th style="padding:6px 8px;text-align:left;border:1px solid #93C5FD;">대분류</th>';
    summaryHtml += '<th style="padding:6px 8px;text-align:left;border:1px solid #93C5FD;">중분류 (화면)</th>';
    summaryHtml += '<th style="padding:6px 8px;text-align:left;border:1px solid #93C5FD;width:130px;">소분류</th>';
    summaryHtml += '</tr>';
    // 옵션 E: 메인 행(시트/Suite/TC ID/대분류/중분류 + 소분류 토글) + 종속 행(소분류 상세, colspan=6)
    var prevMajor = null;
    var groupCount = 0;
    var totalMiddles = cats.length;
    var majorIndex = -1;
    var seenMajors = {};
    for (var ci = 0; ci < cats.length; ci++) {
      var c = cats[ci];
      var majorClean = c.major;
      var midClean = (c.middle || '').trim();
      var minorList = (c.minors || []).filter(function(m){ return m && m.trim(); });
      var minorCount = minorList.length;

      var isFirstOfMajor = !(majorClean in seenMajors);
      if (isFirstOfMajor) {
        majorIndex++;
        seenMajors[majorClean] = majorIndex;
      }
      var domIdx = seenMajors[majorClean];

      var sheetName = majorClean;
      var autoCode = majorClean.replace(/[^A-Za-z]/g, '').toUpperCase().substring(0, 4);
      var screenCode = resolveScreenCodeFE(midClean, {});
      var previewId = _pcode + '-' + screenCode + '-001';

      var bg = ci % 2 === 0 ? '#F8FAFC' : '#FFFFFF';
      var nestedBg = ci % 2 === 0 ? '#F1F5F9' : '#F8FAFC';  // 종속 행은 살짝 더 어둡게
      // ─ 메인 행 (기본 정보) ─
      summaryHtml += '<tr class="tc-summary-row-main" style="background:' + bg + ';">';
      if (isFirstOfMajor) {
        summaryHtml += '<td style="padding:5px 8px;border:1px solid #D1D5DB;font-weight:700;color:#1E3A5F;">' + _escapeHtml(sheetName) + '</td>';
        summaryHtml += '<td style="padding:5px 8px;border:1px solid #D1D5DB;"><input type="text" class="suite-code-input" data-domain="' + domIdx + '" data-pcode="' + _pcode + '" value="' + autoCode + '" placeholder="예: GNBF" oninput="updateTcIdPreview(this)" style="width:80px;padding:4px 6px;border:1.5px solid #93C5FD;border-radius:4px;font-size:12px;font-weight:700;text-transform:uppercase;font-family:monospace;"></td>';
      } else {
        summaryHtml += '<td style="padding:5px 8px;border:1px solid #D1D5DB;color:#9CA3AF;font-size:11px;">↳ 동일 시트</td>';
        summaryHtml += '<td style="padding:5px 8px;border:1px solid #D1D5DB;color:#9CA3AF;">↳</td>';
      }
      summaryHtml += '<td style="padding:5px 8px;border:1px solid #D1D5DB;font-family:monospace;font-size:11px;color:#1D4ED8;" id="tcIdPreview_' + ci + '" data-pcode="' + _pcode + '" data-middle="' + _escapeHtml(midClean) + '">' + previewId + '</td>';
      summaryHtml += '<td style="padding:5px 8px;border:1px solid #D1D5DB;">' + (isFirstOfMajor ? _escapeHtml(majorClean) : '<span style="color:#9CA3AF;">↳</span>') + '</td>';
      summaryHtml += '<td style="padding:5px 8px;border:1px solid #D1D5DB;font-weight:600;">' + _escapeHtml(midClean) + '</td>';
      // 소분류 토글 셀
      if (minorCount > 0) {
        summaryHtml += '<td style="padding:5px 8px;border:1px solid #D1D5DB;text-align:center;cursor:pointer;user-select:none;" onclick="toggleMinorRow(' + ci + ')" title="클릭하여 소분류 펼치기/접기">';
        summaryHtml += '<span id="minorToggleIcon_' + ci + '" style="color:#3B82F6;font-weight:700;">▼</span> ';
        summaryHtml += '<span style="color:#1E3A5F;font-weight:600;font-size:11px;">' + minorCount + '개</span>';
        summaryHtml += '</td>';
      } else {
        summaryHtml += '<td style="padding:5px 8px;border:1px solid #D1D5DB;text-align:center;color:#9CA3AF;font-size:11px;">없음</td>';
      }
      summaryHtml += '</tr>';
      // ─ 종속 행 (소분류 상세) — 펼침 상태로 시작 (open) ─
      if (minorCount > 0) {
        var minorText = minorList.map(function(m){
          return '<li style="margin-bottom:4px;">' + _escapeHtml(m) + '</li>';
        }).join('');
        summaryHtml += '<tr class="tc-summary-row-nested" id="minorRow_' + ci + '" style="background:' + nestedBg + ';">';
        summaryHtml += '<td colspan="6" style="padding:0;border:1px solid #D1D5DB;border-top:none;">';
        summaryHtml += '<div style="padding:8px 12px 10px 28px;color:#374151;line-height:1.6;">';
        summaryHtml += '<div style="font-size:10.5px;color:#6B7280;font-weight:600;margin-bottom:4px;letter-spacing:0.3px;">↳ 소분류 (' + minorCount + ')</div>';
        summaryHtml += '<ul style="margin:0;padding:0 0 0 14px;list-style:disc;">' + minorText + '</ul>';
        summaryHtml += '</div></td></tr>';
      }
    }
    summaryHtml += '</table>';
    var estMin = totalMiddles * 5;
    var estMax = totalMiddles * 15;
    summaryHtml += '<div style="margin-top:10px;padding:8px 12px;background:#F0FDF4;border:1px solid #86EFAC;border-radius:8px;font-size:12px;color:#166534;display:flex;align-items:center;gap:8px;">';
    summaryHtml += '<span style="font-size:16px;">📊</span>';
    summaryHtml += '<span>예상 TC 수량: <strong>' + estMin + '~' + estMax + '개</strong> (중분류 ' + totalMiddles + '개 × 5~15개)</span>';
    summaryHtml += '</div>';
    summaryHtml += '</div>';   // 본문 div close
    summaryHtml += '</details>';  // 외부 details close
  }
  // 메인 영역(Viewer = gateViewer)에는 통합 표 + 분류표 헤더 표시 (옵션 A — 표 중심)
  var headerLine = '<div style="font-size:13px;color:var(--muted);margin-bottom:10px;">📋 분류표 결과 — 채팅으로 수정하거나 SuiteCode를 입력한 후 승인하세요.</div>';
  var fallbackHtml = '';
  if (!summaryHtml) {
    // 표를 만들 수 없으면 raw 마크다운을 즉시 보여주기 (토글 안 해도 보이게)
    fallbackHtml = '<div style="padding:14px; background:#FEF3C7; border:1px solid #FCD34D; border-radius:8px; font-size:12px; color:#92400E; margin-bottom:10px;">⚠️ 분류표 구조 자동 인식 실패 — 아래 원본 문서를 확인하시거나, AI에게 "표 형식으로 다시 정리해줘"로 요청하세요.</div>';
    fallbackHtml += '<pre style="margin:0; padding:14px; background:#FAFAFA; border:1px solid #E5E7EB; border-radius:6px; font-size:12px; line-height:1.7; max-height:400px; overflow:auto; white-space:pre-wrap; word-break:break-word; font-family:ui-monospace,SFMono-Regular,Menlo,monospace;">' + _escapeHtml(mdText) + '</pre>';
  }
  viewer.innerHTML = headerLine + (summaryHtml || fallbackHtml);
  viewer.classList.add('gate-doc-updated');
  setTimeout(() => viewer.classList.remove('gate-doc-updated'), 700);

  // 원본 마크다운은 토글 영역(rawDocContent)에 별도 렌더 — 펼쳐진 상태일 때만 사용자가 봄
  var rawCell = document.getElementById('rawDocContent');
  if (rawCell) rawCell.textContent = mdText;
  var rawBadge = document.getElementById('rawDocBadge');
  if (rawBadge) rawBadge.textContent = '최근 업데이트: ' + new Date().toLocaleTimeString();

  // 레거시: suiteCodeSection은 더 이상 사용하지 않지만 안전상 비워둠
  var suiteSection = document.getElementById('suiteCodeSection');
  if (suiteSection) {
    suiteSection.innerHTML = '';
    suiteSection.style.display = 'none';
  }
}

function addGateChatMsg(role, text) {
  const msgs = document.getElementById('gateChatMessages');
  const div = document.createElement('div');
  div.className = 'gate-msg ' + role;
  div.textContent = text;
  msgs.appendChild(div);
  msgs.scrollTop = msgs.scrollHeight;
}

async function sendGateChat() {
  const input = document.getElementById('gateChatInput');
  const sendBtn = document.getElementById('gateChatSend');
  const msg = input.value.trim();
  if (!msg || !currentSid) return;

  input.value = '';
  sendBtn.disabled = true;
  addGateChatMsg('user', msg);

  // 로딩 메시지
  const loadingDiv = document.createElement('div');
  loadingDiv.className = 'gate-msg assistant';
  loadingDiv.textContent = '⏳ 분석 중...';
  loadingDiv.id = 'gateLoadingMsg';
  document.getElementById('gateChatMessages').appendChild(loadingDiv);
  document.getElementById('gateChatMessages').scrollTop = 999999;

  const currentDoc = document.getElementById('gateContent').value;

  try {
    const r = await fetch('/gate-chat/' + currentSid, {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({
        message: msg,
        current_doc: currentDoc,
        gate_mode: window._gateMode || 'new',
        history: gateChatHistory.slice(-6)  // 최근 6턴만 전송
      })
    });
    const d = await r.json();

    // 로딩 제거
    document.getElementById('gateLoadingMsg')?.remove();

    if (d.ok) {
      addGateChatMsg('assistant', d.reply);
      // 문서 업데이트
      document.getElementById('gateContent').value = d.updated_doc;
      renderGateViewer(d.updated_doc);
      // 히스토리 추가 (컨텍스트 유지용 — 축약 버전)
      gateChatHistory.push({ role: 'user', content: '요청: ' + msg });
      gateChatHistory.push({ role: 'assistant', content: d.reply });
      // v0.9.20: 변경 여부에 따라 다른 메시지
      var statusMsg = '';
      if (d.changed) {
        statusMsg = '✅ 분류표가 업데이트되었습니다. 아래 표에서 확인하세요.';
        if (d.sync_status === 'synced') {
          statusMsg += ' 본문 정책·기능 문서도 자동 동기화 완료 — TC 작성 시 일관됨.';
        } else if (d.sync_status === 'failed') {
          statusMsg += ' (⚠️ 본문 동기화 실패 — 분류표 변경만 반영됨.)';
        }
      } else if (d.truncated) {
        statusMsg = '⚠️ AI 응답이 너무 길어 잘렸습니다. 분류표는 변경되지 않았어요. 더 작은 단위로 나눠 다시 요청해 주세요.';
      } else {
        // 변경 없음 — 명령이 모호했거나 AI 가 질문으로 해석한 경우
        statusMsg = 'ℹ️ 분류표가 변경되지 않았습니다. 요청이 모호했을 수 있어요. 더 명확하게 다시 요청해 주세요. (예: "Splash 중분류 삭제해줘", "AUTH 대분류의 3번 케이스 삭제")';
      }
      addGateChatMsg('system', statusMsg);
    } else {
      document.getElementById('gateLoadingMsg')?.remove();
      addGateChatMsg('system', '❌ 오류: ' + d.error);
    }
  } catch(e) {
    document.getElementById('gateLoadingMsg')?.remove();
    addGateChatMsg('system', '❌ 네트워크 오류: ' + e.message);
  }
  sendBtn.disabled = false;
  input.focus();
}

async function regenerateClassification() {
  if (!currentSid) return;
  if (!confirm('분류표를 처음부터 다시 생성합니다. 현재 분류표는 삭제됩니다. 계속할까요?')) return;
  try {
    // 현재 파이프라인 중단
    await fetch('/stop/' + currentSid, { method: 'POST' });
    if (eventSource) eventSource.close();
    // 분류표 재생성 요청
    const projectName = document.getElementById('projectName').value.trim() || selectedProject || '';
    const focusArea = document.getElementById('focusArea').value.trim();
    document.getElementById('card1').classList.add('hidden');
    document.getElementById('card3').classList.add('hidden');
    document.getElementById('card2').classList.remove('hidden');
    setStepBar(2);
    document.getElementById('card2').scrollIntoView({ behavior: 'smooth', block: 'start' });
    document.getElementById('stageLabel').textContent = '분류표 재생성 중...';
    const r = await fetch('/regenerate-classification', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ project_name: projectName, focus_area: focusArea || null })
    });
    const d = await r.json();
    if (!d.ok) { alert(d.error); return; }
    currentSid = d.sid;
    showToast('분류표를 다시 생성합니다.');
    connectStream(d.sid);
  } catch(e) {
    alert('오류: ' + e.message);
  }
}

// ── Excel 출력 옵션 (Full / Light / Custom) ──
const EXCEL_PRESETS = {
  full:   { cover:true,  stats:true,  smoke:true,  traceability:true,  tc_list:true, change_history:true },
  light:  { cover:false, stats:false, smoke:false, traceability:false, tc_list:true, change_history:false },
  // custom 은 마지막 사용자 체크 상태 그대로 — localStorage 에서 복원
};

function onExcelPresetChange() {
  const preset = document.querySelector('input[name="excelPreset"]:checked')?.value || 'full';
  const customPanel = document.getElementById('excelCustomPanel');
  if (preset === 'custom') {
    customPanel.style.display = '';
    // Custom 진입 시 마지막 저장된 체크 상태 복원
    try {
      const saved = JSON.parse(localStorage.getItem('tc_excel_custom_sheets') || 'null');
      if (saved) applySheetChecks(saved);
    } catch (_) {}
  } else {
    customPanel.style.display = 'none';
    // 프리셋 → 체크박스 동기화 (시각화)
    applySheetChecks(EXCEL_PRESETS[preset]);
  }
  updateExcelSheetSummary();
  // 사용자 선택 저장
  try { localStorage.setItem('tc_excel_preset', preset); } catch (_) {}
}

function onExcelSheetCheckChange() {
  // 체크박스 직접 변경 시 — 자동으로 Custom 모드로 전환
  const customRadio = document.querySelector('input[name="excelPreset"][value="custom"]');
  if (customRadio && !customRadio.checked) {
    customRadio.checked = true;
    document.getElementById('excelCustomPanel').style.display = '';
  }
  updateExcelSheetSummary();
  // Custom 체크 상태 저장
  try {
    localStorage.setItem('tc_excel_custom_sheets', JSON.stringify(collectExcelSheets()));
    localStorage.setItem('tc_excel_preset', 'custom');
  } catch (_) {}
}

function applySheetChecks(sheets) {
  document.querySelectorAll('.excel-sheet-cb').forEach(function(cb) {
    const key = cb.dataset.sheet;
    if (key === 'tc_list') return;  // 필수 — 건드리지 않음
    if (cb.disabled) return;
    if (key in sheets) cb.checked = !!sheets[key];
  });
}

function collectExcelSheets() {
  const sheets = {};
  document.querySelectorAll('.excel-sheet-cb').forEach(function(cb) {
    sheets[cb.dataset.sheet] = !!cb.checked;
  });
  sheets.tc_list = true;  // 강제
  return sheets;
}

function updateExcelSheetSummary() {
  const sheets = collectExcelSheets();
  const labels = {
    cover:'표지', stats:'통계', smoke:'Smoke',
    traceability:'Traceability', tc_list:'TC 목록', change_history:'변경 이력',
  };
  const enabled = Object.keys(sheets).filter(k => sheets[k]);
  const el = document.getElementById('excelSheetSummary');
  if (!el) return;
  if (enabled.length === Object.keys(labels).length) {
    el.innerHTML = '💡 요약: <strong>전체 시트 (' + enabled.length + '개)</strong>';
  } else {
    el.innerHTML = '💡 요약: <strong>' + enabled.map(k => labels[k]).join(' + ') + ' = ' + enabled.length + '개 시트</strong>';
  }
}

function getSelectedExcelSheets() {
  // approveGate 가 호출 — 현재 선택된 sheets dict 반환
  const preset = document.querySelector('input[name="excelPreset"]:checked')?.value || 'full';
  if (preset === 'full' || preset === 'light') {
    return { ...EXCEL_PRESETS[preset] };
  }
  return collectExcelSheets();
}

// 페이지 로드 시 — localStorage 에서 마지막 선택 복원
function restoreExcelOption() {
  try {
    const lastPreset = localStorage.getItem('tc_excel_preset') || 'full';
    const r = document.querySelector('input[name="excelPreset"][value="' + lastPreset + '"]');
    if (r) { r.checked = true; }
    if (lastPreset === 'custom') {
      const saved = JSON.parse(localStorage.getItem('tc_excel_custom_sheets') || 'null');
      if (saved) applySheetChecks(saved);
      document.getElementById('excelCustomPanel').style.display = '';
    } else {
      applySheetChecks(EXCEL_PRESETS[lastPreset] || EXCEL_PRESETS.full);
    }
    updateExcelSheetSummary();
  } catch (_) {}
}
// DOM ready 후 1회 복원
if (typeof document !== 'undefined') {
  document.addEventListener('DOMContentLoaded', function() {
    try { restoreExcelOption(); } catch(e) {}
  });
}

async function approveGate() {
  if (!currentSid) return;
  const content = document.getElementById('gateContent').value.trim();
  if (!content) { alert('내용이 비어있습니다.'); return; }

  // Auto 모드 여부 (체크박스 상태)
  var autoScreenCode = !!window._autoScreenCode;

  // SuiteCode 수집
  var suiteCodeInputs = document.querySelectorAll('.suite-code-input');
  var suiteCodes = {};
  var hasEmpty = false;
  suiteCodeInputs.forEach(function(input) {
    var code = input.value.trim().toUpperCase();
    if (!code) hasEmpty = true;
    suiteCodes[input.dataset.domain] = code;
  });
  // Auto 모드에서는 SuiteCode 입력 검증 생략 (screen_code_map이 중분류별로 자동 매핑)
  if (!autoScreenCode && suiteCodeInputs.length > 0 && hasEmpty) {
    alert('모든 대분류의 SuiteCode를 입력해주세요.\\n(또는 "🤖 시스템 규칙 자동 적용" 체크박스를 켜면 자동 매핑됩니다.)'); return;
  }
  // SuiteCode 목록 (순서대로) — Auto 모드에서도 백엔드 로그 표시용으로 전송
  var suiteCodeList = [];
  for (var i = 0; i < suiteCodeInputs.length; i++) {
    suiteCodeList.push(suiteCodeInputs[i].value.trim().toUpperCase().replace(/^-+|-+$/g, ''));
  }

  // 대분류 선택 UI는 제거됨 — 범위 제한은 Step 1의 "TC 생성 범위"(focus_area)로 일원화
  // selected_domains는 항상 null(=전체) 전송. 백엔드 하위 호환 유지.
  const selectedDomains = null;

  const approveBtn = document.querySelector('#card3 .btn-success');
  approveBtn.disabled = true;
  approveBtn.textContent = '⏳ 처리 중...';

  var codeMsg;
  if (autoScreenCode) {
    codeMsg = ' (🤖 시스템 규칙 자동 적용 — 화면별 ScreenCode로 자동 매핑)';
  } else {
    codeMsg = suiteCodeList.length > 0 ? ' (SuiteCode: ' + suiteCodeList.join(', ') + ')' : '';
  }
  const scopeMsg = '분류표 전체 범위로 TC를 생성합니다.' + codeMsg + ' 계속할까요?';
  if (!confirm(scopeMsg)) {
    approveBtn.disabled = false;
    approveBtn.textContent = '✅ 승인 및 TC 생성 시작';
    return;
  }

  try {
    const resp = await fetch('/approve/' + currentSid, {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({
        content,
        selected_domains: selectedDomains,
        suite_codes: suiteCodeList,
        auto_screen_code: autoScreenCode,
        excel_sheets: getSelectedExcelSheets()
      })
    });
    const data = await resp.json();
    if (data.ok) {
      // Gate 닫고 파이프라인 카드로 복귀 — 이후 TC 작성/빌드 로그는 card2에서 계속 표시
      document.getElementById('card1').classList.add('hidden');
      document.getElementById('card3').classList.add('hidden');
      document.getElementById('card2').classList.remove('hidden');
      document.getElementById('stageLabel').textContent = 'TC 작성 시작...';
      setStepBar(4);
      document.getElementById('card2').scrollIntoView({ behavior: 'smooth', block: 'start' });
    } else {
      alert('승인 오류: ' + data.error);
      approveBtn.disabled = false;
      approveBtn.textContent = '✅ 승인 및 TC 생성 시작';
    }
  } catch(e) {
    alert('오류: ' + e.message);
    approveBtn.disabled = false;
    approveBtn.textContent = '✅ 승인 및 TC 생성 시작';
  }
}

function setStepBar(active) {
  for (let i = 1; i <= 5; i++) {
    const el = document.getElementById('stepBar' + i);
    el.classList.remove('active', 'done');
    if (i < active) el.classList.add('done');
    else if (i === active) el.classList.add('active');
  }
}

async function stopPipeline() {
  if (!currentSid) return;
  if (!confirm('파이프라인을 중단할까요? 현재 단계가 끝난 직후 멈춥니다.')) return;
  setStopButtonsDisabled(true);
  try {
    await fetch('/stop/' + currentSid, { method: 'POST' });
    showToast('⏹ 중단 요청을 보냈습니다...', 'error');
    // 프론트에서도 SSE 즉시 끊기 — 가드 (isPipelineRunning) 가 풀리도록
    // 서버는 백그라운드에서 현 단계 종료 후 멈춤. 프론트는 더 이상 이벤트 받을 필요 없음.
    if (eventSource) {
      eventSource.close();
      eventSource = null;
      if (typeof updateRestartButtonState === 'function') updateRestartButtonState();
    }
  } catch(e) {
    showToast('❌ 중단 요청 실패: ' + e.message, 'error');
    setStopButtonsDisabled(false);
  }
}

function setStopButtonsDisabled(disabled) {
  ['stopBtn2', 'stopBtn3'].forEach(id => {
    const btn = document.getElementById(id);
    if (btn) btn.disabled = disabled;
  });
}

function downloadFile() {
  if (!currentSid || !currentFilename) return;
  window.location.href = '/download/' + currentSid + '/' + encodeURIComponent(currentFilename);
  showToast('⬇ 다운로드 시작!');
}

async function openFolder() {
  await fetch('/open-folder', { method: 'POST' });
  showToast('📁 폴더를 열었습니다');
}

const DRIVE_SVG = `<svg width="16" height="16" viewBox="0 0 87.3 78" xmlns="http://www.w3.org/2000/svg"><path d="m6.6 66.85 3.85 6.65c.8 1.4 1.95 2.5 3.3 3.3l13.75-23.8h-27.5c0 1.55.4 3.1 1.2 4.5z" fill="#0066da"/><path d="m43.65 25-13.75-23.8c-1.35.8-2.5 1.9-3.3 3.3l-25.4 44a9.06 9.06 0 0 0-1.2 4.5h27.5z" fill="#00ac47"/><path d="m73.55 76.8c1.35-.8 2.5-1.9 3.3-3.3l1.6-2.75 7.65-13.25c.8-1.4 1.2-2.95 1.2-4.5h-27.502l5.852 11.5z" fill="#ea4335"/><path d="m43.65 25 13.75-23.8c-1.35-.8-2.9-1.2-4.5-1.2h-18.5c-1.6 0-3.15.45-4.5 1.2z" fill="#00832d"/><path d="m59.8 53h-32.3l-13.75 23.8c1.35.8 2.9 1.2 4.5 1.2h50.8c1.6 0 3.15-.45 4.5-1.2z" fill="#2684fc"/><path d="m73.4 26.5-12.7-22c-.8-1.4-1.95-2.5-3.3-3.3l-13.75 23.8 16.15 27h27.45c0-1.55-.4-3.1-1.2-4.5z" fill="#ffba00"/></svg>`;

async function uploadToDrive() {
  if (!currentFilename) return;
  const btn = document.getElementById('driveBtn');
  btn.disabled = true;
  btn.innerHTML = '⏳ 인증 확인 중...';
  // 1. 인증 상태 확인
  try {
    const sr = await fetch('/drive/status');
    const sd = await sr.json();
    if (!sd.ok || !sd.authenticated) {
      if (sd.need_credentials) {
        openDriveModal();
        btn.innerHTML = '⚠️ Drive 연동 설정 필요';
        btn.style.borderColor = '#e53e3e';
        btn.style.color = '#e53e3e';
        btn.disabled = false;
        return;
      }
      // 토큰 만료 → 백엔드에서 재인증 필요 (get_drive_service 내부에서 브라우저 열림)
      btn.innerHTML = '⏳ 로그인 창이 열립니다...';
      showToast('Google 로그인이 필요합니다. 브라우저 창에서 로그인을 완료하세요.', 'info');
      // drive/status 재호출로 인증 유도
      await fetch('/drive/status');
      btn.innerHTML = DRIVE_SVG + ' Google Drive에 올리기';
      btn.disabled = false;
      showToast('로그인 완료 후 다시 Drive 버튼을 눌러주세요.');
      return;
    }
    // 2. 폴더 선택 모달 열기
    openDriveFolderModal(sd.email || '');
    btn.innerHTML = DRIVE_SVG + ' Google Drive에 올리기';
    btn.disabled = false;
  } catch(e) {
    showToast('❌ 오류: ' + e.message, 'error');
    btn.innerHTML = DRIVE_SVG + ' Google Drive에 올리기';
    btn.disabled = false;
  }
}

async function openDriveFolderModal(email) {
  const modal = document.getElementById('drive-folder-modal');
  modal.classList.add('open');
  document.getElementById('driveFolderEmail').textContent = email ? '로그인 계정: ' + email : '';
  document.getElementById('driveFolderSearch').value = '';
  // 업로드 버튼 상태 초기화 — 이전 업로드 직후 '⏳ 업로드 중...' 텍스트가
  // 잔존하던 버그 차단. 폴더 미선택 상태로 시작하므로 disabled 도 true.
  const uploadBtn = document.getElementById('driveUploadBtn');
  if (uploadBtn) {
    uploadBtn.textContent = '✅ 이 폴더에 업로드';
    uploadBtn.disabled = true;
  }
  // 이전 선택 폴더 표시도 리셋
  _selectedDriveFolder = null;
  const status = document.getElementById('driveSelectedFolder');
  if (status) status.textContent = '';
  await loadDriveFolders('');
  document.getElementById('driveFolderSearch').focus();
}

function closeDriveFolderModal() {
  document.getElementById('drive-folder-modal').classList.remove('open');
}

async function loadDriveFolders(query) {
  const list = document.getElementById('driveFolderList');
  list.innerHTML = '<div style="padding:20px;text-align:center;color:#888;">⏳ 폴더 조회 중...</div>';
  try {
    const url = '/drive/folders' + (query ? '?q=' + encodeURIComponent(query) : '');
    const r = await fetch(url);
    const d = await r.json();
    if (!d.ok) {
      list.innerHTML = '<div style="padding:20px;color:#c53030;">❌ ' + (d.error || '폴더 조회 실패') + '</div>';
      return;
    }
    if (d.folders.length === 0) {
      list.innerHTML = '<div style="padding:20px;text-align:center;color:#888;">검색 결과 없음</div>';
      return;
    }
    list.innerHTML = '';
    d.folders.forEach(function(f) {
      var div = document.createElement('div');
      div.className = 'drive-folder-item';
      div.onclick = function(e) { selectDriveFolder(f.id, f.name, e.currentTarget); };
      div.innerHTML = '<span style="font-size:18px;">📁</span> <span style="flex:1;">' + f.name + '</span>' +
        '<span style="font-size:11px;color:#888;">' + (f.modifiedTime || '').substring(0, 10) + '</span>';
      list.appendChild(div);
    });
  } catch(e) {
    list.innerHTML = '<div style="padding:20px;color:#c53030;">❌ ' + e.message + '</div>';
  }
}

let _selectedDriveFolder = null;
function selectDriveFolder(id, name, element) {
  _selectedDriveFolder = { id: id, name: name };
  document.querySelectorAll('.drive-folder-item').forEach(el => el.classList.remove('selected'));
  if (element) element.classList.add('selected');
  document.getElementById('driveUploadBtn').disabled = false;
  document.getElementById('driveSelectedFolder').textContent = '✓ 선택됨: ' + name;
}

// Drive 업로드 버튼을 원래 라벨/스타일로 복귀 — 새 결과 도착 또는 새 파이프라인 시작 시 호출
function resetDriveBtn() {
  const btn = document.getElementById('driveBtn');
  if (!btn) return;
  // SVG + "Google Drive에 올리기" 라벨 복원
  if (typeof DRIVE_SVG !== 'undefined') {
    btn.innerHTML = DRIVE_SVG + ' Google Drive에 올리기';
  } else {
    btn.textContent = 'Google Drive에 올리기';
  }
  btn.style.borderColor = '';
}

async function confirmDriveUpload() {
  if (!_selectedDriveFolder) { alert('폴더를 선택해주세요.'); return; }
  const uploadBtn = document.getElementById('driveUploadBtn');
  uploadBtn.disabled = true;
  uploadBtn.textContent = '⏳ 업로드 중...';
  try {
    const r = await fetch('/upload-to-drive', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({
        sid: currentSid,
        filename: currentFilename,
        folder_id: _selectedDriveFolder.id
      })
    });
    const d = await r.json();
    if (d.ok) {
      closeDriveFolderModal();
      showToast('✅ "' + _selectedDriveFolder.name + '" 폴더에 업로드 완료!');
      const btn = document.getElementById('driveBtn');
      btn.innerHTML = DRIVE_SVG + ' Drive 업로드 완료';
      btn.style.borderColor = '#34a853';
      if (d.link) window.open(d.link, '_blank');
    } else {
      alert('❌ ' + d.error);
      uploadBtn.disabled = false;
      uploadBtn.textContent = '✅ 이 폴더에 업로드';
    }
  } catch(e) {
    alert('오류: ' + e.message);
    uploadBtn.disabled = false;
    uploadBtn.textContent = '✅ 이 폴더에 업로드';
  }
}

function openDriveModal() { document.getElementById('drive-modal').classList.add('open'); }
function closeDriveModal() { document.getElementById('drive-modal').classList.remove('open'); }
function closeRecoveryModal() {
  document.getElementById('recovery-modal').classList.remove('open');
  fetch('/checkpoint', { method: 'DELETE' });  // 종료 선택 시 체크포인트 삭제
}

async function restartModify(mode) {
  document.getElementById('recovery-modal').classList.remove('open');
  const projectName = document.getElementById('projectSelect').value;
  const changeDesc  = document.getElementById('changeDesc')?.value?.trim() || '';
  document.getElementById('card1').classList.add('hidden');
  document.getElementById('card2').classList.remove('hidden');
  setStepBar(2);
  document.getElementById('card2').scrollIntoView({ behavior: 'smooth', block: 'start' });
  try {
    const r = await fetch('/restart-modify', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ mode, project_name: projectName, change_desc: changeDesc })
    });
    const d = await r.json();
    if (!d.ok) { alert('❌ ' + d.error); return; }
    currentSid = d.sid;
    showToast(mode === 'resume' ? '🔄 체크포인트에서 재시작합니다.' : '▶ 처음부터 재시작합니다.');
    connectStream(d.sid);
  } catch(e) {
    alert('재시작 오류: ' + e.message);
  }
}
document.getElementById('drive-modal').addEventListener('click', e => {
  if (e.target === e.currentTarget) closeDriveModal();
});

function showToast(msg, type = 'success') {
  const t = document.getElementById('toast');
  t.textContent = msg; t.className = type; t.style.display = 'block';
  setTimeout(() => t.style.display = 'none', 3000);
}

// ── 서버 재시작 (헤더 버튼) ──
async function restartServer() {
  const btn = document.getElementById('btnRestartServer');
  const overlay = document.getElementById('restartOverlay');
  const status = document.getElementById('restartOverlayStatus');

  // 1) 활성 세션 확인 (사전 체크)
  let activeSessions = 0;
  try {
    const r = await fetch('/admin/status');
    const d = await r.json();
    activeSessions = d.active_sessions || 0;
  } catch (_) {}

  // 2) confirm 다이얼로그 (활성 세션 있으면 강한 경고)
  let msg = '서버를 재시작하시겠습니까?\\n\\n';
  msg += '• 코드 변경사항이 반영됩니다.\\n';
  msg += '• 진행 중인 SSE 연결이 모두 끊깁니다.\\n';
  if (activeSessions > 0) {
    msg = '⚠️ 진행 중인 작업이 ' + activeSessions + '개 있습니다!\\n\\n';
    msg += '재시작하면 진행 중인 TC 생성이 모두 중단되며,\\n';
    msg += '복원이 어려울 수 있습니다.\\n\\n';
    msg += '정말 재시작하시겠습니까?';
  }
  if (!confirm(msg)) return;

  // 3) 재시작 요청 (활성 세션이 있으면 force=1)
  btn.disabled = true;
  overlay.classList.add('visible');
  status.textContent = '재시작 요청 전송 중...';

  try {
    const url = '/admin/restart' + (activeSessions > 0 ? '?force=1' : '');
    const r = await fetch(url, { method: 'POST' });
    const d = await r.json();
    if (!d.ok) {
      overlay.classList.remove('visible');
      btn.disabled = false;
      alert('재시작 실패: ' + (d.message || d.error || 'unknown error'));
      return;
    }
  } catch (e) {
    // 네트워크 오류 — 이미 서버가 죽었을 수 있음 (정상 흐름)
    status.textContent = '서버가 종료되었습니다. 살아나길 기다리는 중...';
  }

  // 4) 서버 살아남 폴링 → 새 버전이면 reload
  pollServerAlive();
}

// 서버가 다시 살아날 때까지 폴링 (최대 30초)
async function pollServerAlive() {
  const status = document.getElementById('restartOverlayStatus');
  const startedAt = Date.now();
  const maxMs = 30 * 1000;
  let attempt = 0;
  const initialVersion = window._INITIAL_APP_VERSION || '';

  while (Date.now() - startedAt < maxMs) {
    attempt++;
    status.textContent = '서버 응답 확인 중... (시도 ' + attempt + ')';
    try {
      const r = await fetch('/admin/status', { cache: 'no-store' });
      if (r.ok) {
        const d = await r.json();
        const newVersion = d.version || '';
        // 버전이 바뀌었거나 동일해도 살아있음 → reload
        status.textContent = '✅ 서버 응답 감지: ' + newVersion + ' — 새로고침합니다.';
        await new Promise(r => setTimeout(r, 600));
        // 강제 reload (캐시 우회)
        window.location.reload();
        return;
      }
    } catch (_) {
      // 아직 죽어있음 — 계속 폴링
    }
    await new Promise(r => setTimeout(r, 800));
  }

  // 30초 초과 → 사용자에게 수동 새로고침 요청
  status.textContent = '⚠️ 30초 동안 응답 없음. 직접 새로고침해 주세요.';
}

// ── Sticky Mini AI 채팅 패널 (Step 3 분류 검토 전용) ──
// - 메인 #gateChatMessages 에서 메시지를 미러링하여 하단에서도 응답 즉시 확인 가능
// - 3가지 상태: collapsed(입력만) / expanded(메시지+입력) / modal(확대)
// - v0.9.12: 다국어 placeholder + 모달 리사이즈 지원
(function() {
  // ─ 다국어 placeholder ─
  // localStorage 'tc_ui_lang' 우선, 없으면 navigator.language 자동 감지
  const I18N = {
    ko: {
      placeholder: '예) AUTH 대분류 케이스 3번 삭제해줘 — Enter로 전송, Shift+Enter 줄바꿈',
      placeholderModal: '예) AUTH 대분류 케이스 3번 삭제해줘 — Enter로 전송, Shift+Enter 줄바꿈',
      labelMain: '💬 AI 도우미',
      labelHint: '표 검토 중에도 바로 대화하세요',
      btnLarge: '⛶ 크게',
      btnClose: '✕ 닫기',
      modalTitle: '💬 AI 도우미 — 분류표 검토',
      empty: '아직 대화가 없어요. 아래에서 요청을 입력해보세요.',
    },
    en: {
      placeholder: 'e.g. Delete AUTH domain case #3 — Enter to send, Shift+Enter for newline',
      placeholderModal: 'e.g. Delete AUTH domain case #3 — Enter to send, Shift+Enter for newline',
      labelMain: '💬 AI Helper',
      labelHint: 'Chat anytime while reviewing the table',
      btnLarge: '⛶ Expand',
      btnClose: '✕ Close',
      modalTitle: '💬 AI Helper — Classification Review',
      empty: 'No conversation yet. Type a request below to start.',
    },
    ja: {
      placeholder: '例) AUTH ドメインのケース3を削除して — Enter で送信、Shift+Enter で改行',
      placeholderModal: '例) AUTH ドメインのケース3を削除して — Enter で送信、Shift+Enter で改行',
      labelMain: '💬 AI アシスタント',
      labelHint: '表を確認しながらいつでも対話',
      btnLarge: '⛶ 拡大',
      btnClose: '✕ 閉じる',
      modalTitle: '💬 AI アシスタント — 分類表レビュー',
      empty: 'まだ会話がありません。下に要求を入力してください。',
    },
    zh: {
      placeholder: '例) 删除 AUTH 域名案例 #3 — Enter 发送, Shift+Enter 换行',
      placeholderModal: '例) 删除 AUTH 域名案例 #3 — Enter 发送, Shift+Enter 换行',
      labelMain: '💬 AI 助手',
      labelHint: '审阅表格时随时对话',
      btnLarge: '⛶ 放大',
      btnClose: '✕ 关闭',
      modalTitle: '💬 AI 助手 — 分类表审阅',
      empty: '尚无对话。在下方输入请求开始。',
    },
  };
  function detectLang() {
    try {
      const saved = localStorage.getItem('tc_ui_lang');
      if (saved && I18N[saved]) return saved;
    } catch (_) {}
    const nav = (navigator.language || 'ko').toLowerCase();
    if (nav.startsWith('en')) return 'en';
    if (nav.startsWith('ja')) return 'ja';
    if (nav.startsWith('zh')) return 'zh';
    return 'ko';  // 기본값
  }
  const T = I18N[detectLang()];
  // 전역에 노출 — 다른 코드에서 언어 변경 시 호출 가능
  window.setStickyAiLang = function(lang) {
    if (!I18N[lang]) return false;
    try { localStorage.setItem('tc_ui_lang', lang); } catch(_) {}
    location.reload();
    return true;
  };
  // ─ DOM 구성 ─
  const bar = document.createElement('div');
  bar.id = 'floatingAiBar';
  bar.className = 'floating-ai-bar';
  bar.innerHTML =
    '<div class="floating-ai-header" id="floatingAiHeader">' +
      '<div class="floating-ai-header-label">' +
        '<span>' + T.labelMain + '</span>' +
        '<span class="floating-ai-msg-badge" id="floatingAiMsgBadge" style="display:none;">0</span>' +
        '<small>' + T.labelHint + '</small>' +
      '</div>' +
      '<button class="floating-ai-ctrl" id="floatingAiExpandBtn" title="' + T.btnLarge + '">' + T.btnLarge + '</button>' +
      '<button class="floating-ai-ctrl" id="floatingAiCollapseBtn" title="▾">▾</button>' +
    '</div>' +
    '<div class="floating-ai-messages" id="floatingAiMessages">' +
      '<div class="floating-ai-empty" id="floatingAiEmpty">' + T.empty + '</div>' +
    '</div>' +
    '<div class="floating-ai-inner">' +
      '<textarea class="floating-ai-input" id="floatingAiInput" rows="1" ' +
        'placeholder="' + T.placeholder + '"></textarea>' +
      '<button class="floating-ai-send" id="floatingAiSend">' + (detectLang() === 'ko' ? '전송' : detectLang() === 'en' ? 'Send' : detectLang() === 'ja' ? '送信' : '发送') + '</button>' +
    '</div>';
  document.body.appendChild(bar);

  // ─ 확대 모달 DOM ─
  const modal = document.createElement('div');
  modal.id = 'floatingAiModal';
  modal.className = 'floating-ai-modal';
  const sendLabel = detectLang() === 'ko' ? '전송' : detectLang() === 'en' ? 'Send' : detectLang() === 'ja' ? '送信' : '发送';
  modal.innerHTML =
    '<div class="floating-ai-modal-box" id="floatingAiModalBox">' +
      '<div class="floating-ai-modal-header">' +
        '<div class="floating-ai-modal-title">' + T.modalTitle + '</div>' +
        '<button class="floating-ai-modal-close" id="floatingAiModalClose">' + T.btnClose + '</button>' +
      '</div>' +
      '<div class="floating-ai-modal-messages" id="floatingAiModalMessages"></div>' +
      '<div class="floating-ai-modal-input-row">' +
        '<textarea class="floating-ai-input" id="floatingAiModalInput" rows="2" ' +
          'placeholder="' + T.placeholderModal + '" ' +
          'style="border-color:var(--border);"></textarea>' +
        '<button class="floating-ai-send" id="floatingAiModalSend">' + sendLabel + '</button>' +
      '</div>' +
      '<div class="floating-ai-modal-resize" id="floatingAiModalResize" title="크기 조절"></div>' +
    '</div>';
  document.body.appendChild(modal);

  const input = document.getElementById('floatingAiInput');
  const sendBtn = document.getElementById('floatingAiSend');
  const collapseBtn = document.getElementById('floatingAiCollapseBtn');
  const expandBtn = document.getElementById('floatingAiExpandBtn');
  const header = document.getElementById('floatingAiHeader');
  const miniMessages = document.getElementById('floatingAiMessages');
  const emptyMsg = document.getElementById('floatingAiEmpty');
  const msgBadge = document.getElementById('floatingAiMsgBadge');
  const modalMessages = document.getElementById('floatingAiModalMessages');
  const modalInput = document.getElementById('floatingAiModalInput');
  const modalSend = document.getElementById('floatingAiModalSend');
  const modalClose = document.getElementById('floatingAiModalClose');

  // 기본은 접힘 상태 — 사용자가 자주 쓰는 입력만 보임 (메시지는 헤더 클릭하면 펼침)
  bar.classList.add('collapsed');
  document.body.classList.add('has-floating-ai-collapsed');

  // 메인 입력창과 동기화하여 sendGateChat() 재사용
  async function submitFromInput(srcInput) {
    const msg = srcInput.value.trim();
    if (!msg) return;
    const mainInput = document.getElementById('gateChatInput');
    if (!mainInput) return;
    mainInput.value = msg;
    srcInput.value = '';
    sendBtn.disabled = true;
    modalSend.disabled = true;
    // 메시지를 보내는 순간 펼침 상태로 자동 전환 → 응답 보일 수 있게
    if (bar.classList.contains('collapsed')) toggleCollapsed(false);
    try {
      await sendGateChat();
    } finally {
      sendBtn.disabled = false;
      modalSend.disabled = false;
    }
  }
  sendBtn.addEventListener('click', () => submitFromInput(input));
  input.addEventListener('keydown', function(e) {
    // v0.9.24: IME (한글/일본어/중국어 등) 조합 중 Enter 무시
    // - keyCode === 229: 일부 브라우저에서 IME composition 중 표시
    // - e.isComposing: 표준 (Chrome/Firefox/Safari)
    if (e.isComposing || e.keyCode === 229) return;
    if (e.key === 'Enter' && !e.shiftKey) {
      e.preventDefault();
      submitFromInput(input);
    }
  });
  modalSend.addEventListener('click', () => submitFromInput(modalInput));
  modalInput.addEventListener('keydown', function(e) {
    if (e.isComposing || e.keyCode === 229) return;
    if (e.key === 'Enter' && !e.shiftKey) {
      e.preventDefault();
      submitFromInput(modalInput);
    }
  });

  // ─ collapse / expand 토글 ─
  function toggleCollapsed(forceCollapsed) {
    const willCollapse = (typeof forceCollapsed === 'boolean')
      ? forceCollapsed
      : !bar.classList.contains('collapsed');
    bar.classList.toggle('collapsed', willCollapse);
    document.body.classList.toggle('has-floating-ai-collapsed', willCollapse);
    collapseBtn.textContent = willCollapse ? '▴' : '▾';
    collapseBtn.title = willCollapse ? '펼치기' : '접기';
    if (!willCollapse) {
      // 펼쳤으니 메시지 영역 스크롤 맨 아래로
      setTimeout(() => { miniMessages.scrollTop = miniMessages.scrollHeight; }, 60);
    }
  }
  // 헤더 전체 클릭 → 토글 (단, 버튼 클릭 시는 무시)
  header.addEventListener('click', function(e) {
    if (e.target.closest('button')) return;
    toggleCollapsed();
  });
  collapseBtn.addEventListener('click', function(e) {
    e.stopPropagation();
    toggleCollapsed();
  });

  // ─ 모달 열기/닫기 + 크기 영속 ─
  const modalBox = document.getElementById('floatingAiModalBox');
  function applySavedModalSize() {
    try {
      const saved = JSON.parse(localStorage.getItem('tc_modal_size') || 'null');
      if (saved && saved.w && saved.h) {
        modalBox.style.width = saved.w + 'px';
        modalBox.style.height = saved.h + 'px';
        modalBox.style.maxWidth = 'none';  // saved size 가 max 보다 클 수 있게
      }
    } catch (_) {}
  }
  function saveModalSize() {
    try {
      const w = modalBox.offsetWidth;
      const h = modalBox.offsetHeight;
      if (w > 100 && h > 100) {
        localStorage.setItem('tc_modal_size', JSON.stringify({ w, h }));
      }
    } catch (_) {}
  }
  // ResizeObserver 로 사용자가 드래그 종료한 후 자동 저장
  if (window.ResizeObserver) {
    let resizeTimer = null;
    new ResizeObserver(function() {
      if (resizeTimer) clearTimeout(resizeTimer);
      resizeTimer = setTimeout(saveModalSize, 300);  // 드래그 멈춘 후 0.3초
    }).observe(modalBox);
  }

  function openModal() {
    syncMessages();  // 최신 동기화
    applySavedModalSize();
    modal.classList.add('open');
    setTimeout(() => modalInput.focus(), 60);
    setTimeout(() => { modalMessages.scrollTop = modalMessages.scrollHeight; }, 60);
  }
  function closeModal() { modal.classList.remove('open'); }
  expandBtn.addEventListener('click', function(e) {
    e.stopPropagation();
    openModal();
  });
  modalClose.addEventListener('click', closeModal);
  modal.addEventListener('click', function(e) {
    if (e.target === modal) closeModal();
  });
  document.addEventListener('keydown', function(e) {
    if (e.key === 'Escape' && modal.classList.contains('open')) closeModal();
  });

  // ─ 메시지 동기화: 메인 #gateChatMessages → mini + modal ─
  // 메인 채팅 패널의 메시지를 그대로 미러링 (단방향 — 메인이 단일 진실 소스)
  function syncMessages() {
    const main = document.getElementById('gateChatMessages');
    if (!main) return;
    const mainNodes = Array.from(main.querySelectorAll('.gate-msg'));
    // 미러 클래스 매핑
    function buildClone(node, prefix) {
      const div = document.createElement('div');
      let role = 'assistant';
      if (node.classList.contains('user')) role = 'user';
      else if (node.classList.contains('system')) role = 'system';
      div.className = prefix + ' ' + prefix + '-' + role;
      // mini-chat 클래스명 매핑
      if (prefix === 'floating-ai-msg') {
        div.className = 'floating-ai-msg ' + role;
      } else {
        // 모달은 메인 gate-msg 와 동일 스타일 그대로 활용
        div.className = 'gate-msg ' + role;
      }
      div.textContent = node.textContent;
      return div;
    }

    // mini 갱신 (innerHTML 한 번에 교체 — 단순)
    miniMessages.innerHTML = '';
    if (mainNodes.length === 0) {
      const empty = document.createElement('div');
      empty.className = 'floating-ai-empty';
      empty.textContent = '아직 대화가 없어요. 아래에서 요청을 입력해보세요.';
      miniMessages.appendChild(empty);
    } else {
      mainNodes.forEach(n => miniMessages.appendChild(buildClone(n, 'floating-ai-msg')));
    }
    // modal 갱신
    modalMessages.innerHTML = '';
    mainNodes.forEach(n => modalMessages.appendChild(buildClone(n, 'gate-msg')));

    // 메시지 개수 뱃지
    if (mainNodes.length > 0) {
      msgBadge.style.display = '';
      msgBadge.textContent = String(mainNodes.length);
    } else {
      msgBadge.style.display = 'none';
    }
    // 자동 스크롤 — 항상 최신 메시지 보이게
    setTimeout(() => {
      miniMessages.scrollTop = miniMessages.scrollHeight;
      modalMessages.scrollTop = modalMessages.scrollHeight;
    }, 30);
  }

  // 메인 채팅 패널 변경 감지
  const mainMessages = document.getElementById('gateChatMessages');
  if (mainMessages) {
    new MutationObserver(syncMessages).observe(mainMessages, {
      childList: true, subtree: true, characterData: true,
    });
    // 초기 동기화
    setTimeout(syncMessages, 200);
  }

  // 가시성 제어: Step 3 (분류표 검토) 화면일 때 항상 표시
  // v0.9.12: 메인 채팅 패널이 숨겨졌으므로 viewport 위치 기반 판정 제거.
  //          Step 3 진입 = 미니 채팅 항상 노출 (사용자가 항상 접근 가능)
  function updateVisibility() {
    const card3 = document.getElementById('card3');
    const stepBar3 = document.getElementById('stepBar3');
    const card3Visible = card3 && !card3.classList.contains('hidden');
    const stepBar3Active = stepBar3 && stepBar3.classList.contains('active');
    if (!card3Visible || !stepBar3Active) {
      bar.classList.remove('visible');
      document.body.classList.remove('has-floating-ai');
      return;
    }
    bar.classList.add('visible');
    document.body.classList.add('has-floating-ai');
  }

  // 스크롤/리사이즈/카드 토글에 반응
  let scrollRaf = null;
  function onScrollOrResize() {
    if (scrollRaf) return;
    scrollRaf = requestAnimationFrame(function() {
      updateVisibility();
      scrollRaf = null;
    });
  }
  window.addEventListener('scroll', onScrollOrResize, { passive: true });
  window.addEventListener('resize', onScrollOrResize);

  // card3 + stepBar3 클래스 변경 감지 (hidden / active 토글) + 첫 진입 안내 토스트
  const card3Watch = document.getElementById('card3');
  const stepBar3Watch = document.getElementById('stepBar3');
  if (card3Watch) {
    new MutationObserver(function() {
      updateVisibility();
      // 진짜 Step 3 진입 — card3 보이고 + stepBar3 active 일 때만 안내 토스트
      const isCard3Visible = !card3Watch.classList.contains('hidden');
      const isStep3Active = stepBar3Watch && stepBar3Watch.classList.contains('active');
      if (isCard3Visible && isStep3Active) {
        try {
          if (localStorage.getItem('tc_sticky_ai_hint_v098c') !== '1') {
            localStorage.setItem('tc_sticky_ai_hint_v098c', '1');
            if (typeof showToast === 'function') {
              setTimeout(function() {
                showToast('💡 분류표를 스크롤하면 하단에 AI 입력바가 자동으로 떠요', 'success');
              }, 1200);
            }
          }
        } catch (_) {}
      }
    }).observe(card3Watch, {
      attributes: true, attributeFilter: ['class']
    });
  }
  // stepBar3 의 active 클래스 변경도 감지 (SSE 재연결 등으로 card3 만 unhidden 되는 경우 방어)
  if (stepBar3Watch) {
    new MutationObserver(updateVisibility).observe(stepBar3Watch, {
      attributes: true, attributeFilter: ['class']
    });
  }
  // 초기 1회
  setTimeout(updateVisibility, 100);
})();

// ──────────────────────────────────────────────────────────
// Combo 명세 작성/갱신 모달
// ──────────────────────────────────────────────────────────
let _comboSpecMode = 'new';  // 'new' or 'edit'
let _comboSpecEditPath = '';

async function openComboSpecModal(mode) {
  _comboSpecMode = mode;
  const modal = document.getElementById('comboSpecModal');
  const title = document.getElementById('comboSpecModalTitle');
  const editSelect = document.getElementById('comboSpecEditSelect');
  const editArea = document.getElementById('comboSpecEditArea');
  const draftBtn = document.getElementById('comboSpecDraftBtn');
  const previewArea = document.getElementById('comboSpecPreview');
  const status = document.getElementById('comboSpecStatus');

  // spec 폴더 사전 검증
  const specPath = (document.getElementById('specFolderPath')?.value || '').trim();
  if (!specPath) {
    alert('먼저 위쪽 "구조화 spec 폴더 경로" 를 입력해주세요.\\nspec 폴더 분석이 필요합니다.');
    return;
  }

  title.textContent = mode === 'new' ? '➕ Combo 명세 새로 작성' : '✏️ Combo 명세 갱신';
  previewArea.value = '';
  status.textContent = '';
  status.style.color = '#6B7280';

  if (mode === 'edit') {
    editArea.style.display = 'block';
    // 기존 파일 목록 로드
    const projectName = (document.getElementById('projectName')?.value || 'supercycl').trim();
    const url = '/combo/list-files?project_name=' + encodeURIComponent(projectName) +
                (specPath ? '&spec_folder=' + encodeURIComponent(specPath) : '');
    const resp = await fetch(url);
    const data = await resp.json();
    if (!data.ok || !data.files || !data.files.length) {
      alert('편집할 명세가 없습니다. "새 명세 작성" 모드를 사용하세요.');
      return;
    }
    editSelect.innerHTML = data.files.map(f =>
      '<option value="' + f.path + '">' + f.name + ' (' + f.domain + ')</option>'
    ).join('');
  } else {
    editArea.style.display = 'none';
  }

  modal.style.display = 'flex';
}

function closeComboSpecModal() {
  document.getElementById('comboSpecModal').style.display = 'none';
}

async function draftComboSpec() {
  const status = document.getElementById('comboSpecStatus');
  const previewArea = document.getElementById('comboSpecPreview');
  const draftBtn = document.getElementById('comboSpecDraftBtn');
  const domain = document.getElementById('comboSpecDomain').value.trim() || 'Trade';
  const targetScrsRaw = document.getElementById('comboSpecTargetScrs').value.trim();
  const specPath = (document.getElementById('specFolderPath')?.value || '').trim();

  let existingMd = '';
  let editPath = '';
  if (_comboSpecMode === 'edit') {
    editPath = document.getElementById('comboSpecEditSelect').value;
    if (editPath) {
      // 기존 명세 로드
      const r = await fetch('/combo/read-spec?path=' + encodeURIComponent(editPath));
      const d = await r.json();
      if (d.ok) existingMd = d.content;
    }
  }
  _comboSpecEditPath = editPath;

  const targetScrs = targetScrsRaw
    ? targetScrsRaw.split(/[,\\s]+/).map(s => s.trim()).filter(s => s).map(s => s.toUpperCase())
    : null;

  draftBtn.disabled = true;
  status.style.color = '#1976D2';
  status.textContent = '⏳ AI 가 spec 분석 + 명세 초안 작성 중... (1~2분 소요)';
  previewArea.value = '';

  try {
    const resp = await fetch('/combo/draft-spec', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({
        spec_folder: specPath,
        domain: domain,
        target_scrs: targetScrs,
        existing_md: existingMd,
      })
    });
    const data = await resp.json();
    if (!data.ok) {
      status.style.color = '#DC2626';
      status.textContent = '❌ ' + (data.error || '오류');
      previewArea.value = data.trace || '';
      return;
    }
    previewArea.value = data.draft_md;
    status.style.color = '#10B981';
    let statusMsg = '✅ 초안 작성 완료. 검토 후 "저장" 버튼을 누르세요. ('
      + data.draft_md.length + '자)';
    if (data.meta) {
      const m = data.meta;
      if (m.reference_source) {
        statusMsg += ' · 📋 참조: ' + m.reference_source;
      }
      if (m.screens_count) {
        statusMsg += ' · 화면 ' + m.screens_count + '개 분석';
      }
    }
    status.textContent = statusMsg;
  } catch (e) {
    status.style.color = '#DC2626';
    status.textContent = '❌ 네트워크 오류: ' + e.message;
  } finally {
    draftBtn.disabled = false;
  }
}

async function saveComboSpec() {
  const content = document.getElementById('comboSpecPreview').value.trim();
  const status = document.getElementById('comboSpecStatus');
  if (!content) {
    alert('저장할 내용이 없습니다. "초안 생성" 을 먼저 실행하세요.');
    return;
  }

  let savePath = '';
  if (_comboSpecMode === 'edit' && _comboSpecEditPath) {
    savePath = _comboSpecEditPath;  // 덮어쓰기
  } else {
    // 새 파일 — 도메인 기반 파일명 + spec 폴더에 저장
    const domain = document.getElementById('comboSpecDomain').value.trim() || 'Trade';
    const specPath = (document.getElementById('specFolderPath')?.value || '').trim();
    const filename = domain.toLowerCase().replace(/[^a-z0-9]/g, '_') + '_combinations.md';
    savePath = specPath + '/' + filename;
  }

  if (!confirm('다음 위치에 저장합니다:\\n' + savePath + '\\n\\n진행할까요?')) return;

  status.style.color = '#1976D2';
  status.textContent = '💾 저장 중...';

  try {
    const resp = await fetch('/combo/save-spec', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ save_path: savePath, content: content })
    });
    const data = await resp.json();
    if (!data.ok) {
      status.style.color = '#DC2626';
      status.textContent = '❌ ' + (data.error || '저장 실패');
      if (data.allowed) {
        status.textContent += ' (허용: ' + data.allowed.join(', ') + ')';
      }
      return;
    }
    status.style.color = '#10B981';
    status.textContent = '✅ 저장 완료: ' + data.saved_path;
    // 파일 목록 새로고침
    setTimeout(function() {
      closeComboSpecModal();
      loadComboFiles();
    }, 1500);
  } catch (e) {
    status.style.color = '#DC2626';
    status.textContent = '❌ 네트워크 오류: ' + e.message;
  }
}
</script>

<!-- Combo 명세 작성/갱신 모달 -->
<div id="comboSpecModal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,0.5);z-index:9999;align-items:center;justify-content:center;">
  <div style="background:white;width:90%;max-width:900px;max-height:90vh;border-radius:10px;display:flex;flex-direction:column;overflow:hidden;">
    <div style="padding:16px 20px;background:#1976D2;color:white;display:flex;justify-content:space-between;align-items:center;">
      <h3 id="comboSpecModalTitle" style="margin:0;font-size:16px;">Combo 명세 작성</h3>
      <button onclick="closeComboSpecModal()" style="background:transparent;color:white;border:0;font-size:24px;cursor:pointer;line-height:1;">×</button>
    </div>
    <div style="padding:16px 20px;border-bottom:1px solid #E5E7EB;background:#F9FAFB;">
      <div id="comboSpecEditArea" style="display:none;margin-bottom:12px;">
        <label style="font-size:12px;font-weight:600;color:#374151;">갱신할 명세 파일</label>
        <select id="comboSpecEditSelect" style="width:100%;padding:6px 10px;border:1px solid #D1D5DB;border-radius:4px;font-size:13px;margin-top:4px;"></select>
      </div>
      <div style="display:flex;gap:10px;flex-wrap:wrap;">
        <div style="flex:1;min-width:180px;">
          <label style="font-size:12px;font-weight:600;color:#374151;">도메인</label>
          <input type="text" id="comboSpecDomain" value="Trade"
                 placeholder="예: Trade, Exchange, Portfolio"
                 style="width:100%;padding:6px 10px;border:1px solid #D1D5DB;border-radius:4px;font-size:13px;margin-top:4px;">
          <div style="font-size:11px;color:#6B7280;margin-top:2px;">→ 파일명 <code>{domain}_combinations.md</code> + 시트 <code>{domain}(Combo)</code></div>
        </div>
        <div style="flex:1;min-width:180px;">
          <label style="font-size:12px;font-weight:600;color:#374151;">분석 대상 SCR (선택)</label>
          <input type="text" id="comboSpecTargetScrs" placeholder="예: SCR-102,SCR-104,SCR-106 (비우면 자동)"
                 style="width:100%;padding:6px 10px;border:1px solid #D1D5DB;border-radius:4px;font-size:13px;margin-top:4px;">
          <div style="font-size:11px;color:#6B7280;margin-top:2px;">비우면 도메인 키워드로 자동 선택 (최대 10개)</div>
        </div>
      </div>
      <button id="comboSpecDraftBtn" onclick="draftComboSpec()"
              style="margin-top:12px;padding:8px 16px;background:#10B981;color:white;border:0;border-radius:6px;cursor:pointer;font-weight:600;">
        🤖 AI 초안 생성
      </button>
      <span id="comboSpecStatus" style="margin-left:10px;font-size:12px;"></span>
    </div>
    <div style="flex:1;padding:16px 20px;overflow:auto;">
      <label style="font-size:12px;font-weight:600;color:#374151;display:block;margin-bottom:6px;">
        명세 미리보기 / 편집 (markdown — 직접 수정 가능)
      </label>
      <textarea id="comboSpecPreview"
                style="width:100%;height:50vh;padding:12px;border:1px solid #D1D5DB;border-radius:6px;font-family:'Monaco','Menlo',monospace;font-size:12px;line-height:1.5;resize:vertical;"
                placeholder="AI 초안 생성 후 여기에 명세가 표시됩니다. 직접 수정 가능합니다."></textarea>
    </div>
    <div style="padding:12px 20px;background:#F9FAFB;border-top:1px solid #E5E7EB;display:flex;gap:8px;justify-content:flex-end;">
      <button onclick="closeComboSpecModal()"
              style="padding:8px 16px;background:#9CA3AF;color:white;border:0;border-radius:6px;cursor:pointer;">취소</button>
      <button onclick="saveComboSpec()"
              style="padding:8px 16px;background:#1976D2;color:white;border:0;border-radius:6px;cursor:pointer;font-weight:600;">
        💾 저장
      </button>
    </div>
  </div>
</div>

</body>
</html>
"""


# ── 진입점 ────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    print(f"🚀 TC 자동화 v2 시작: http://localhost:{PORT}")
    print(f"   BASE_DIR   : {BASE_DIR}")
    print(f"   AGENT_DIR  : {AGENT_DIR}")
    print(f"   BUILD_EXCEL: {BUILD_EXCEL} {'✓' if BUILD_EXCEL.exists() else '✗ (없음)'}")
    print(f"   RULES_FILE : {RULES_FILE} {'✓' if RULES_FILE.exists() else '✗ (없음)'}")
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if api_key:
        print(f"   API KEY    : sk-ant-...{api_key[-6:]} (설정됨)")
    else:
        print("   API KEY    : ⚠️  미설정 — .env 파일을 확인하세요")
    # ── SO_REUSEADDR/SO_REUSEPORT 활성화 + listen socket 캡처 (자기 재시작 지원) ──
    # 문제 1: 기본 소켓은 os.execv 후 새 프로세스가 'Address already in use' 로 즉시 bind 실패
    #         → SO_REUSEADDR/SO_REUSEPORT 활성화로 해결
    # 문제 2: execv 후에도 옛 listen socket fd가 살아있어서 OS가 connection을 임의로 두 fd 중 하나에
    #         배분 → 옛 fd로 가면 처리 안 됨
    #         → execv 직전에 우리 listen socket 을 명시적 close 해야 함 → app._listen_sock 캡처
    import socket as _socket_mod
    from werkzeug.serving import BaseWSGIServer as _BWS
    _orig_server_bind = _BWS.server_bind
    def _patched_server_bind(self):
        # bind 직전에 SO_REUSEADDR/SO_REUSEPORT 강제 활성화
        try:
            self.socket.setsockopt(_socket_mod.SOL_SOCKET, _socket_mod.SO_REUSEADDR, 1)
        except OSError:
            pass
        if hasattr(_socket_mod, "SO_REUSEPORT"):
            try:
                self.socket.setsockopt(_socket_mod.SOL_SOCKET, _socket_mod.SO_REUSEPORT, 1)
            except OSError:
                pass
        result = _orig_server_bind(self)
        # 자기 재시작 핸들러가 close 할 수 있도록 listen socket 캡처
        try:
            app._listen_sock = self.socket
        except Exception:
            pass
        return result
    _BWS.server_bind = _patched_server_bind
    # 일반 app.run 사용 — patched bind 가 자동 적용됨
    app.run(host="0.0.0.0", port=PORT, debug=False, threaded=True, use_reloader=False)
