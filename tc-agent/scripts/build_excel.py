"""
Supercycl TC Excel Builder — v3 (대분류/중분류/소분류 분류 체계 적용)
변경점 (v2 → v3):
  - 파서: 대분류/중분류/소분류 필드 추출 추가
  - TC 목록 컬럼: "도메인" → "대분류 | 중분류 | 소분류" 3열로 분리
  - 분류표 없는 TC는 기존 도메인 기반 폴백 동작
  - 통계 시트: 중분류별 TC 수 섹션 추가

사용법:
  python3 scripts/build_excel.py --phase P2_Mobile --tc phase2-mobile/_workspace/05_review/tc_final.md --output phase2-mobile/outputs
  python3 scripts/build_excel.py --phase P1_Youthmeta --tc phase1-youthmeta/workspace/youthmeta_3_tc.md --output phase1-youthmeta/outputs
"""

import re, sys
from datetime import datetime
from pathlib import Path
import argparse
from collections import defaultdict

# Windows cp949 콘솔 환경에서 이모지/한글 print 시 UnicodeEncodeError 방지:
# stdout/stderr를 UTF-8로 강제 재설정 (Python 3.7+)
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl 미설치. 실행: pip install openpyxl")
    sys.exit(1)

# ── Phase 설정 ─────────────────────────────────────────────────────

PHASE_CONFIG = {
    "P2_Mobile": {
        "name": "Supercycl 모바일 체험",
        "code": "P2_Mobile",
        "exchanges": [],
        "testers": {},
        "test_env": "Hyperliquid Testnet (Chain ID 998)",
        "platform": "Web(Mobile) — iOS Safari / Android Chrome",
        "target": "모바일 체험 시나리오 — 온보딩·거래·시그널·설정",
        "methodology": "User Journey  ·  BDD Given/When/Then  ·  블랙박스  ·  화면 관찰 기반",
    },
    "P_WebApp": {
        "name": "TC 자동화 웹앱",
        "code": "P_WebApp",
        "exchanges": [],
        "testers": {},
        "test_env": "Web 기반 TC 자동 생성",
        "platform": "Web(Mobile) / iOS Safari / Android Chrome / 공통",
        "target": "기획서/GitHub URL/텍스트 기반 자동 생성 TC",
        "methodology": "BDD Given/When/Then  ·  블랙박스  ·  화면 관찰 기반",
    },
    "P1_Youthmeta": {
        "name": "Supercycl × Youthmeta",
        "code": "P1_Youthmeta",
        "exchanges": ["Gate", "OKX", "Bybit", "Bitget", "Hyperliquid"],
        "testers": {
            "Gate":         ("Tester 2", "2E4057"),
            "OKX":          ("Tester 1", "1C6E38"),
            "Bybit":        ("Tester 1", "1C6E38"),
            "Bitget":       ("Tester 2", "2E4057"),
            "Hyperliquid":  ("Tester 2", "2E4057"),
        },
        "test_env": "https://aggr-dev.supercycl.io/?partner=Youthmeta",
        "platform": "Web (Chrome / Brave)",
        "target": "자동 TP/SL 설정 기능  +  유스메타 유입 유저 레버리지 2배 고정",
        "methodology": "User Journey  ·  BDD Given/When/Then  ·  블랙박스",
    },
}

# ── 도메인별 그룹 헤더 색상 ────────────────────────────────────────

DOMAIN_COLORS = {
    "AUTH": "1F3864",   # dark navy
    "REFF": "2E4057",   # dark blue-gray
    "LEVR": "7B2D8B",   # purple
    "TPSL": "C44D00",   # dark orange
    "TRAD": "0070C0",   # blue
    "SGNL": "1C6E38",   # dark green
    "MOBL": "4472C4",   # medium blue
    "SETG": "375623",   # dark green (settings)
    "ROUT": "833C00",   # brown
    "FUND": "375623",
    "ETC":  "636363",   # gray
}

DOMAIN_LABELS = {
    "AUTH": "인증 · 온보딩 · 자금 지급",
    "REFF": "레퍼럴 코드 자동 등록",
    "LEVR": "레버리지 2배 고정",
    "TPSL": "자동 TP/SL",
    "TRAD": "거래 · 주문 · 포지션",
    "SGNL": "YouthMeta 시그널",
    "MOBL": "Testnet UI · 모바일 반응형",
    "SETG": "설정 (SettingsPage)",
    "ROUT": "라우팅 · 접근 제어",
    "FUND": "자금 지급",
    "ETC":  "기타",
}

# ── 색상 팔레트 ────────────────────────────────────────────────────

C_DARK   = "1F3864"; C_WHITE  = "FFFFFF"
C_BLUE1  = "2E75B6"; C_BLUE2  = "DDEBF7"
C_GREEN  = "E2EFDA"; C_YELLOW = "FFF2CC"
C_RED    = "FCE4D6"; C_GRAY   = "E0E0E0"
C_ROW_A  = "FFFFFF"; C_ROW_B  = "F5F9FE"

PRIORITY_COLOR = {"High": "FCE4D6", "Medium": "FFF3E0", "Low": "E8F5E9"}
CATEGORY_COLOR = {"Positive": "E2EFDA", "Negative": "FCE4D6", "Edge": "FFF2CC"}

# ── 스타일 헬퍼 ────────────────────────────────────────────────────

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_font(bold=False, color="000000", size=10, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)

def make_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def make_align(wrap=True, h="left", v="center"):
    return Alignment(wrap_text=wrap, horizontal=h, vertical=v)

def set_cell(ws, row, col, value, bold=False, bg=None, font_color="000000",
             size=10, align_h="left", wrap=True, font_name="Arial"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = make_font(bold=bold, color=font_color, size=size, name=font_name)
    if bg:
        cell.fill = make_fill(bg)
    cell.alignment = make_align(wrap=wrap, h=align_h)
    cell.border = make_border()
    return cell

# ── TC 파싱 ────────────────────────────────────────────────────────

def parse_tc_markdown(filepath):
    with open(filepath, encoding="utf-8") as f:
        content = f.read()

    # ── v0.10.x: 화면 컨텍스트 추적 — step_write_tc_per_screen 이 TC 본문 앞에
    #    '<!-- SCR-XXX — Title -->' 주석을 삽입해 화면 단위 헤더로 사용함.
    #    AI 가 TC 의 "연관 화면" 필드에 SCR-ID 를 빼먹고 화면명만 적은 경우(예:
    #    'Portfolio Overview') screen_code 컬럼이 빈 칸이 되는 버그 방어.
    #    각 블록 직전의 가장 최근 SCR 컨텍스트를 추적해 빈 코드일 때 보충.
    scr_context_pat = re.compile(r"<!--\s*(SCR-[A-Z0-9]+)\s*[—\-:]", re.IGNORECASE)

    tcs = []
    blocks = re.split(r'\n(?=### )', content)
    current_scr_context = ""  # 현재 위치까지 등장한 가장 최근 SCR 컨텍스트

    for block in blocks:
        # 블록 안에 SCR 컨텍스트 헤더가 있으면 갱신 (다음 TC 들이 이 화면 소속)
        m_ctx = scr_context_pat.search(block)
        if m_ctx:
            current_scr_context = m_ctx.group(1).upper()

        if not block.strip().startswith('###'):
            continue

        first_line = block.split('\n')[0].strip()

        # 카테고리 구분 헤더 스킵 (### 카테고리 1: ..., ### 카테고리 2: ... 등)
        if re.match(r'###\s*카테고리\s*\d', first_line):
            continue

        # 분류 구분 헤더 스킵 (Claude가 섹션 구분용으로 남긴 "### 중분류: XXX" 등)
        # 이를 TC로 오인식하면 "중분류: Splash"가 TC ID가 되는 유령 TC가 생성됨
        if re.match(r'###\s*(대분류|중분류|소분류|Category|Middle|Minor)\s*[:：]', first_line, re.IGNORECASE):
            continue

        # TC ID 패턴 필수 검증 — AI가 ### 헤더로 설명문을 작성한 경우 차단
        # 유효 TC ID 예:  SC-AUTH-001, SM-SPL-001, SA-TRD-ORDR-012
        # 차단 예:        ### 화면 분석, ### 원칙 C 적용 결과, ### ⚠️ TC 생성 대상 제외
        # 패턴: `[A-Z]{2,}-[A-Z0-9]+-\d+` 이 헤더 라인에 등장해야 함 (bold 또는 plain)
        if not re.search(r'\b[A-Z]{2,}-[A-Z0-9]+-\d+\b', first_line):
            continue

        tc = {}

        # 헤더 패턴 — 다음 5가지 모두 지원:
        #   1) ### **ID** — 제목      (Smoke 표준)
        #   2) ### **ID — 제목**      (전체 bold — AI가 종종 만드는 변형)
        #   3) ### ID — 제목          (일반)
        #   4) ### **ID**             (제목 없음, bold)
        #   5) ### ID                 (제목 없음, plain)
        # 모두 처리하기 어려운 경우: ** 마크업을 라인 전체에서 우선 제거하고 통일된 패턴으로 매칭
        cleaned_line = re.sub(r'\*\*', '', first_line).strip()
        # cleaned_line 기준 매칭 (## prefix는 그대로 유지되어야 헤더로 인식)
        # 예: "### SM-VCD-003 — 숫자 키패드 입력"
        m_full = re.match(r'###\s+(.+?)\s+—\s+(.+?)\s*$', cleaned_line)
        m_no_title = re.match(r'###\s+(.+?)\s*$', cleaned_line)
        # 원본에서 bold가 있었는지 체크 (Smoke star 마킹용)
        had_bold = '**' in first_line

        if m_full:
            tc["star"]  = had_bold
            tc["id"]    = m_full.group(1).strip()
            tc["title"] = m_full.group(2).strip()
        elif m_no_title:
            tc["star"]  = had_bold
            tc["id"]    = m_no_title.group(1).strip()
            tc["title"] = tc["id"]
        else:
            continue

        # 안전장치 — 혹시 잔여 마크업이 남았으면 제거
        tc["id"] = re.sub(r'[`\*]', '', tc["id"]).strip()
        tc["title"] = re.sub(r'^\*+|\*+$', '', tc["title"]).strip()

        tc["is_new"]  = "[신규]" in tc["title"]
        tc["title"]   = tc["title"].replace("[신규]", "").strip()

        tc["category"] = extract_table_field(block, "분류") or ""
        tc["priority"] = extract_table_field(block, "우선순위") or ""
        tc["platform"] = extract_table_field(block, "플랫폼") or ""
        tc["screen"]   = extract_table_field(block, "연관 화면") or ""
        # 변경 이력 필드 (기획서 diff 기반 업데이트에서 주입됨)
        tc["status"]         = extract_table_field(block, "상태") or ""
        tc["change_reason"]  = extract_table_field(block, "수정 사유") or ""

        # 화면 코드 추출 (SCR-xxx / SCREEN-xxx / PAGE-xxx). 여러 개 가능 — 쉼표 분리.
        # "연관 화면" 값에서 코드만 뽑아낸다. 코드가 없으면 빈 문자열.
        code_pat = re.compile(r"(?:SCR|SCREEN|PAGE)-[A-Za-z0-9]+")
        codes_found = code_pat.findall(tc["screen"]) if tc["screen"] else []
        tc["screen_code"] = ", ".join(dict.fromkeys(codes_found))  # 중복 제거 · 순서 유지

        # 안전망: AI 가 "연관 화면" 에 SCR-ID 없이 화면명만 적은 경우(예: 'Portfolio
        # Overview') 화면 코드 컬럼이 빈 칸이 됨. step_write_tc_per_screen 의
        # '<!-- SCR-XXX -->' 컨텍스트 주석을 사용해 자동 보충.
        if not tc["screen_code"] and current_scr_context:
            tc["screen_code"] = current_scr_context

        # 대분류/중분류/소분류 추출 (신규 형식)
        # 대분류 필드 예: "Authentication & Onboarding (AUTH)" → "AUTH" 코드 파싱
        cat_raw = extract_table_field(block, "대분류") or ""
        tc["major"]  = cat_raw  # 대분류 (전체 텍스트)
        tc["middle"] = extract_table_field(block, "중분류") or ""
        tc["minor"]  = extract_table_field(block, "소분류") or ""

        # 도메인 코드 결정:
        # - 신규 Project-Suite-NNN 형식 (SC/SM/SA 등 2글자 ProjectCode): 두 번째 세그먼트가 도메인
        #   예: SC-AUTH-001 → AUTH, SM-SPL-001 → SPL, SC-TRD-ORDR-012 → TRD
        # - 구형 도메인-기능-번호 형식: 첫 번째 세그먼트
        #   예: AUTH-OAUTH-001 → AUTH, LEVR-POPUP-001 → LEVR
        parts = tc["id"].split("-")
        PROJECT_CODES = ("SC", "SM", "SA")  # 향후 다른 Project Code 추가 시 확장
        if parts and parts[0].upper() in PROJECT_CODES and len(parts) >= 3:
            tc["domain"] = parts[1]
        elif parts:
            tc["domain"] = parts[0]
        else:
            tc["domain"] = "ETC"

        tc["given"] = extract_section(block, "사전 조건")
        tc["when"]  = extract_section(block, "테스트 단계")
        tc["then"]  = extract_section(block, "예상 결과")
        tc["note"]  = extract_section(block, "비고")

        # 안전망: 본문이 모두 비었다면 응답 잘림 가능성 → 시각적 경고 마커 삽입
        # (Excel 셀이 단순 빈 칸이면 검토자가 놓치기 쉬움. 명시적 마커로 인지)
        if not tc["given"] and not tc["when"] and not tc["then"]:
            tc["given"] = "⚠ (원본 응답 누락 — 재생성 필요)"
            tc["when"]  = "⚠ (원본 응답 누락 — 재생성 필요)"
            tc["then"]  = "⚠ (원본 응답 누락 — 재생성 필요)"

        tc["exchange_na"] = parse_exchange_na(tc["note"])

        tcs.append(tc)

    # ── TC ID 기반 중복 제거 (최종 안전장치) ──
    # review 보강/재생성 등으로 같은 TC ID가 중복 수집된 경우 먼저 등장한 것을 유지한다.
    seen_ids = set()
    unique_tcs = []
    dup_count = 0
    for tc in tcs:
        tid = tc.get("id", "").strip()
        if not tid:
            unique_tcs.append(tc)
            continue
        if tid in seen_ids:
            dup_count += 1
            continue
        seen_ids.add(tid)
        unique_tcs.append(tc)
    if dup_count > 0:
        print(f"  ⚠️ 중복 TC ID {dup_count}개 제거됨 (TC ID 기반 dedup)")
    return unique_tcs

def extract_table_field(block, field_name):
    m = re.search(rf'\|\s*{field_name}\s*\|\s*(.+?)\s*\|', block)
    return m.group(1).strip() if m else ""

def extract_section(block, section_name):
    m = re.search(
        rf'\*\*{section_name}\*\*\s*\n(.*?)(?=\n\*\*|\n---|\Z)',
        block, re.DOTALL
    )
    if not m:
        return ""
    raw_lines = [l for l in m.group(1).strip().split('\n') if l.strip()]
    # 사전 조건만 번호 개조식, 나머지(테스트 단계/예상 결과/비고)는 불릿만 제거
    if section_name == "사전 조건":
        result = []
        n = 1
        for l in raw_lines:
            stripped = l.strip()
            if re.match(r'^\d+\.', stripped):
                result.append(stripped)
                n = int(re.match(r'^(\d+)', stripped).group(1)) + 1
            elif re.match(r'^[-*]\s+', stripped):
                body = re.sub(r'^[-*]\s+', '', stripped)
                result.append(f"{n}. {body}")
                n += 1
            else:
                result.append(f"{n}. {stripped}")
                n += 1
        return "\n".join(result)
    else:
        # 불릿/번호 제거 → 텍스트만
        result = []
        for l in raw_lines:
            stripped = l.strip()
            stripped = re.sub(r'^[-*]\s+', '', stripped)
            stripped = re.sub(r'^\d+\.\s*', '', stripped)
            result.append(stripped)
        return "\n".join(result)

def parse_exchange_na(note_text):
    na = {}
    if not note_text:
        return na
    for line in note_text.split('\n'):
        for ex in ["Gate", "OKX", "Bybit", "Bitget", "Hyperliquid"]:
            if re.search(rf'{ex}.*N/A', line, re.IGNORECASE):
                na[ex] = True
    return na

# ── 버전 자동 증가 ─────────────────────────────────────────────────

def get_next_version(output_dir, phase_code, date_str):
    pattern = re.compile(rf'SPCY_TC_{phase_code}_{date_str}_v(\d+)\.xlsx')
    existing = [int(m.group(1)) for f in Path(output_dir).glob("*.xlsx")
                if (m := pattern.match(f.name))]
    return max(existing) + 1 if existing else 1

# ── 표지 시트 ──────────────────────────────────────────────────────

def build_cover(ws, tcs, config, date_str, version):
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 52

    total = len(tcs)
    smoke_count = sum(1 for t in tcs if t.get("smoke"))

    # 메인 타이틀
    set_cell(ws, 2, 2, f"{config['name']}  |  테스트 케이스 명세서",
             bold=True, bg=C_DARK, font_color=C_WHITE, size=14, align_h="left")
    ws.merge_cells("B2:C2")
    ws.row_dimensions[2].height = 30

    set_cell(ws, 3, 2,
             f"테스트 케이스 명세서  ·  v{version}  ·  BDD Given/When/Then",
             bold=False, bg=C_BLUE1, font_color=C_WHITE, size=10)
    ws.merge_cells("B3:C3")
    ws.row_dimensions[3].height = 18

    # 기본 정보
    ws.row_dimensions[4].height = 8
    info = [
        ("대상 기능",   config["target"]),
        ("테스트 환경", config["test_env"]),
        ("플랫폼",      config["platform"]),
        ("TC 방법론",   config["methodology"]),
        ("작성일",      f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}  (v{version})"),
    ]
    for i, (k, v) in enumerate(info, start=5):
        set_cell(ws, i, 2, k, bold=True, bg=C_BLUE1, font_color=C_WHITE)
        set_cell(ws, i, 3, v, bold=False, bg=C_BLUE2)
        ws.row_dimensions[i].height = 18

    # TC 통계 요약
    ws.row_dimensions[10].height = 10
    set_cell(ws, 11, 2, "TC 통계 요약", bold=True, bg=C_DARK, font_color=C_WHITE, size=11)
    ws.merge_cells("B11:C11")
    ws.row_dimensions[11].height = 22

    high   = sum(1 for t in tcs if t["priority"] == "High")
    medium = sum(1 for t in tcs if t["priority"] == "Medium")
    low    = sum(1 for t in tcs if t["priority"] == "Low")
    pos    = sum(1 for t in tcs if t["category"] == "Positive")
    neg    = sum(1 for t in tcs if t["category"] == "Negative")
    edge   = sum(1 for t in tcs if t["category"] == "Edge")

    stats = [
        ("총 TC 수",      f"{total}개"),
        ("🔥 Smoke Test", f"{smoke_count}개  ({smoke_count*100//total if total else 0}%)"),
        ("High 우선순위",  f"{high}개  ({high*100//total if total else 0}%)"),
        ("Positive",      f"{pos}개"),
        ("Negative",      f"{neg}개"),
        ("Edge",          f"{edge}개"),
    ]
    for i, (k, v) in enumerate(stats, start=12):
        bg = C_BLUE2 if i % 2 == 0 else C_WHITE
        set_cell(ws, i, 2, k, bold=True, bg=C_BLUE1, font_color=C_WHITE)
        set_cell(ws, i, 3, v, bold=False, bg=bg)
        ws.row_dimensions[i].height = 18

    # 화면별 TC 수 — ScreenCode · 중분류 이름 형태로 표시 (옵션 B+C)
    row = 12 + len(stats) + 1
    ws.row_dimensions[row].height = 10
    row += 1
    set_cell(ws, row, 2, "화면별 TC 수", bold=True, bg=C_DARK, font_color=C_WHITE, size=11)
    ws.merge_cells(f"B{row}:C{row}")
    ws.row_dimensions[row].height = 22
    row += 1

    # 도메인(ScreenCode)별 TC 수 + 대표 중분류 이름 수집
    #   - 여러 중분류가 같은 ScreenCode를 공유하면 첫 등장 이름 사용 (보통 유일함)
    domain_counts: dict[str, int] = defaultdict(int)
    domain_labels_dyn: dict[str, str] = {}
    for tc in tcs:
        d = (tc.get("domain") or "").strip()
        if not d:
            continue
        # '**' 같은 마크업 잔여물 제거 (버그 방어)
        d_clean = d.replace("*", "").strip()
        if not d_clean:
            continue
        domain_counts[d_clean] += 1
        if d_clean not in domain_labels_dyn:
            # 중분류 이름 우선 → DOMAIN_LABELS 사전 → 없으면 도메인 코드
            mid = (tc.get("middle") or "").strip()
            mid = re.sub(r"^중분류[:\s]*", "", mid).strip()
            mid = re.sub(r"\s*[\(\（][A-Z0-9\-]+[\)\）]", "", mid).strip()
            if mid:
                domain_labels_dyn[d_clean] = mid
            elif d_clean in DOMAIN_LABELS:
                domain_labels_dyn[d_clean] = DOMAIN_LABELS[d_clean]
            else:
                domain_labels_dyn[d_clean] = d_clean  # fallback

    for i, (domain, cnt) in enumerate(sorted(domain_counts.items())):
        color = DOMAIN_COLORS.get(domain, "636363")
        label = domain_labels_dyn.get(domain, domain)
        # 코드와 label이 같으면 단독 표시 (중복 방지)
        if label == domain:
            text = domain
        else:
            text = f"{domain}  ·  {label}"
        set_cell(ws, row, 2, text, bold=True,
                 bg=color, font_color=C_WHITE)
        set_cell(ws, row, 3, f"{cnt}개", bold=False,
                 bg=C_BLUE2 if i % 2 == 0 else C_WHITE, align_h="center")
        ws.row_dimensions[row].height = 18
        row += 1

    # 판정 코드 범례
    ws.row_dimensions[row].height = 10
    row += 1
    set_cell(ws, row, 2, "판정 코드", bold=True, bg=C_DARK, font_color=C_WHITE, size=11)
    ws.merge_cells(f"B{row}:C{row}")
    ws.row_dimensions[row].height = 22
    row += 1

    for code, desc in [
        ("Pass",  "기대 결과와 일치"),
        ("Fail",  "버그 발생 → 결함 리포트 필수"),
        ("N/T",   "Not Tested  (미실시, 기본값)"),
        ("N/A",   "Not Applicable  (해당 없음)"),
    ]:
        set_cell(ws, row, 2, code, bold=True, bg=C_BLUE1, font_color=C_WHITE, align_h="center")
        set_cell(ws, row, 3, desc, bg=C_BLUE2)
        ws.row_dimensions[row].height = 18
        row += 1


# ── TC 전체목록 시트 ───────────────────────────────────────────────

def _tc_list_columns(config, include_change_columns: bool = True):
    """Phase에 따른 컬럼 정의 반환.
    include_change_columns=False면 상태/수정 사유 컬럼 제외 (신규 TC 생성 시).

    컬럼 순서 (v0.9.7b):
      TC ID | 대분류 | 중분류 | 소분류 | 사전조건 | 테스트 스텝 | 기대결과 |
      중요도 | 대상 거래소 | Smoke | 화면 코드
      (변경 이력 모드: +상태, +수정 사유)
    """
    exchanges = config["exchanges"]

    # 변경 이력 모드: TC ID 앞에 '상태' 추가, 끝에 '수정 사유' 추가
    base = []
    if include_change_columns:
        base.append(("상태", 11))
    base += [
        ("TC ID",              14),
        ("대분류",              13),
        ("중분류",              13),
        ("소분류",              16),
        ("사전조건",            28),
        ("테스트 스텝",         40),
        ("기대결과",            38),
        ("중요도",              9),
        ("대상 거래소",         12),
        ("Smoke",               8),
        ("화면 코드",           12),
    ]
    if include_change_columns:
        base.append(("수정 사유", 28))
    return base


def _priority_to_kr(priority):
    """우선순위 영문 → 한글 변환"""
    mapping = {"High": "높음", "Medium": "보통", "Low": "낮음",
               "높음": "높음", "보통": "보통", "낮음": "낮음"}
    return mapping.get(priority, priority or "보통")

# ── 컬럼 너비 정책 (v0.10.x — 하이브리드: 고정 + 자동) ─────────────────
# 짧은 라벨/패턴 고정 컬럼은 모든 시트에서 동일한 너비 → 시트 간 일관성 확보.
# 본문 멀티라인 컬럼은 콘텐츠 75 퍼센타일 + min/max 클램프 → 가독성 확보.
#
# 값은 Google Sheets/Excel 양쪽에서 시원하게 보이도록 1.4배 스케일 적용.
# Google Sheets 는 동일 width 를 더 좁게 렌더링하므로 (default 문자 폭 차이),
# Excel 기준 보다 약간 넓게 잡아두는 게 양쪽에 유리.
FIXED_COL_WIDTHS = {
    "상태":         17,   # 12 × 1.4
    "TC ID":        22,   # 16 × 1.4
    "대분류":        20,   # 14 × 1.4
    "중분류":        20,   # 14 × 1.4
    "중요도":        14,   # 10 × 1.4
    "대상 거래소":   20,   # 14 × 1.4
    "Smoke":        11,   # 8 × 1.4 (헤더 'Smoke' 5자 + 여유)
    "화면 코드":     18,   # 13 × 1.4 ("SCR-104" + 여유)
}
# 자동 계산 컬럼 — (min, max) 범위. 값이 길면 max 까지, 짧아도 min 보장.
AUTO_COL_RANGES = {
    "소분류":        (25, 50),   # (18, 36) × 1.4
    "사전조건":      (40, 70),   # (28, 50) × 1.4
    "테스트 스텝":   (50, 84),   # (36, 60) × 1.4
    "기대결과":      (45, 77),   # (32, 55) × 1.4
    "수정 사유":     (34, 63),   # (24, 45) × 1.4
}
DEFAULT_AUTO_RANGE = (17, 70)  # (12, 50) × 1.4 — 정의되지 않은 컬럼 폴백


def _calc_col_widths(col_names, all_row_data, min_widths=None, max_width=60):
    """하이브리드 너비 계산.
      - FIXED_COL_WIDTHS 에 있으면 고정 너비
      - AUTO_COL_RANGES 에 있으면 75 퍼센타일 + min/max 클램프
      - 그 외는 DEFAULT_AUTO_RANGE 적용
    min_widths/max_width 인자는 호환을 위해 유지 (오버라이드 가능).
    """
    col_lengths = [[] for _ in col_names]

    # 헤더 길이 (한글 2배 너비)
    header_widths = []
    for i, name in enumerate(col_names):
        hw = sum(2 if ord(c) > 127 else 1 for c in name) + 2
        header_widths.append(hw)

    # 각 셀의 가장 긴 줄 길이 수집 (자동 컬럼만)
    for row_data in all_row_data:
        for i, val in enumerate(row_data):
            if not val:
                col_lengths[i].append(0)
                continue
            lines = str(val).split('\n')
            max_line_w = max(
                sum(2 if ord(c) > 127 else 1 for c in line) + 2
                for line in lines if line.strip()
            ) if lines else 0
            col_lengths[i].append(max_line_w)

    widths = []
    for i, name in enumerate(col_names):
        # 1) 고정 너비 컬럼
        if name in FIXED_COL_WIDTHS:
            widths.append(FIXED_COL_WIDTHS[name])
            continue
        # 2) 자동 컬럼 — 75 퍼센타일 + min/max 클램프
        rng_min, rng_max = AUTO_COL_RANGES.get(name, DEFAULT_AUTO_RANGE)
        # min_widths 인자 호환: 더 큰 값 채택
        explicit_min = (min_widths or {}).get(name, 0)
        rng_min = max(rng_min, explicit_min)

        lengths = col_lengths[i]
        if not lengths:
            widths.append(max(header_widths[i], rng_min))
            continue
        sorted_l = sorted(lengths)
        p75_idx = int(len(sorted_l) * 0.75)
        p75 = sorted_l[min(p75_idx, len(sorted_l) - 1)]
        # 헤더가 더 길면 헤더 보장
        chosen = max(p75, header_widths[i])
        # min/max 클램프
        chosen = max(rng_min, min(chosen, rng_max))
        widths.append(chosen)

    return widths


def _mark_smoke(tcs):
    """Smoke Test 대상 마킹 — 커버리지와 효율 균형을 맞춘 선별.

    선별 기준 (목표: 전체의 20~30%):
      1. 중분류별 대표 Positive 1개: High+Positive 우선, 없으면 Medium+Positive
      2. 중분류별 대표 Negative 1개: High+Negative만 (서비스 차단급)
      3. 대분류별 최소 1개 보장
    """
    from collections import defaultdict

    # 중분류별 그룹핑 (domain + middle)
    mid_groups = defaultdict(list)
    for tc in tcs:
        key = (tc["domain"], tc.get("middle", ""))
        mid_groups[key].append(tc)

    for key, group in mid_groups.items():
        # 1. 대표 Positive 1개: High > Medium > Low 순
        pos_picked = False
        for pri_target in [("high", "높음"), ("medium", "보통"), ("low", "낮음")]:
            if pos_picked:
                break
            for tc in group:
                pri = tc.get("priority", "").lower()
                cat = tc.get("category", "").lower()
                if pri in pri_target and cat == "positive":
                    tc["smoke"] = True
                    pos_picked = True
                    break

        # 2. 대표 Negative 1개: High만
        neg_picked = False
        for tc in group:
            if neg_picked:
                break
            pri = tc.get("priority", "").lower()
            cat = tc.get("category", "").lower()
            if pri in ("high", "높음") and cat == "negative":
                tc["smoke"] = True
                neg_picked = True

    # 3. 대분류별 최소 1개 보장
    domain_has = {}
    for tc in tcs:
        if tc.get("smoke"):
            domain_has[tc["domain"]] = True
    for tc in tcs:
        d = tc["domain"]
        if d not in domain_has:
            tc["smoke"] = True
            domain_has[d] = True


def build_tc_list(ws, tcs, config, include_reason=False, group_by="domain",
                  include_change_columns: bool = True):
    """TC 목록 시트 생성.
    group_by:
      - "domain": TC 코드(도메인) 변경 시 섹션 헤더 (기존 동작)
      - "middle": 중분류 변경 시 섹션 헤더 (대분류별 시트 내부 분할용)
    """
    _mark_smoke(tcs)
    cols = _tc_list_columns(config, include_change_columns=include_change_columns)
    col_names = [c[0] for c in cols]

    # Row 1: 헤더 (너비는 데이터 수집 후 설정)
    for i, name in enumerate(col_names, 1):
        set_cell(ws, 1, i, name, bold=True, bg=C_DARK, font_color=C_WHITE,
                 size=10, align_h="center")
    ws.row_dimensions[1].height = 22

    data_start = 2

    # 데이터 행 (그룹 헤더 삽입)
    prev_group_key = None
    r = data_start
    row_idx = 0
    all_row_data = []  # 너비 계산용

    # 그룹 헤더(회색 구분 행) 표시 여부.
    # v0.10.x: TC 가 2개 이상이면 항상 표시 — 화면별/그룹별 시각 구분이 가독성 핵심.
    # 이전: 그룹 1개면 미표시 → 화면 1개만 처리하는 케이스에서 구분선 사라지던 문제 해결.
    if group_by == "middle":
        unique_groups = set((tc["domain"], tc.get("middle", "")) for tc in tcs)
    else:
        unique_groups = set(tc["domain"] for tc in tcs)
    show_group_headers = len(tcs) >= 2

    for tc in tcs:
        domain = tc["domain"]
        middle = tc.get("middle", "")

        # 그룹 키 결정
        if group_by == "middle":
            group_key = (domain, middle)
        else:
            group_key = domain

        # 그룹 변경 시 헤더 행 삽입 — 연한 회색 통일 (v0.10.x)
        if show_group_headers and group_key != prev_group_key:
            prev_group_key = group_key
            # 연한 회색 배경 + 진한 텍스트 — 가독성 + 시각적 조용함
            HEADER_BG = "BDBDBD"   # 중간 회색 (Material gray-400 톤) — 사용자 검증 완료 색상
            HEADER_FG = "1F2937"   # slate-800
            if group_by == "middle":
                # 중분류별 헤더: "Email Input  ·  SM-EML"
                screen_code = ""
                m_id = re.match(r"^([A-Z]{2})-([A-Z]{2,8})-", tc["id"])
                if m_id:
                    screen_code = f"{m_id.group(1)}-{m_id.group(2)}"
                label = middle or tc.get("major") or domain
                text = f"{label}  ·  {screen_code}" if screen_code else f"{label}"
            else:
                # label 결정 우선순위:
                #   1. DOMAIN_LABELS에 등록된 한글 설명 (AUTH→"인증·온보딩·자금 지급" 등)
                #   2. Screen-based TC라면 tc["middle"] (예: SPL → "Splash")
                #   3. tc["major"] (대분류명, 예: "Onboarding")
                #   4. domain 코드 자체
                label = (
                    DOMAIN_LABELS.get(domain)
                    or tc.get("middle")
                    or tc.get("major")
                    or domain
                )
                # 동일 텍스트 반복 방지 ("SM · SM", "중분류: Splash · 중분류: Splash" 방어)
                text = f"{domain}  ·  {label}" if domain != label else domain
            for i in range(1, len(col_names) + 1):
                set_cell(ws, r, i, text if i == 1 else "",
                         bold=True, bg=HEADER_BG, font_color=HEADER_FG,
                         size=10, align_h="left")
            ws.merge_cells(f"A{r}:{get_column_letter(len(col_names))}{r}")
            ws.row_dimensions[r].height = 22
            r += 1
            row_idx = 0

        bg_base = C_ROW_A if row_idx % 2 == 0 else C_ROW_B
        # 변경 이력 상태별 배경색 덮어쓰기 (zebra 보다 우선)
        _status = (tc.get("status") or "").strip()
        if "신규" in _status or "🆕" in _status:
            bg_base = "DCFCE7"   # 연한 초록
        elif "수정" in _status or "🔄" in _status:
            bg_base = "FEF3C7"   # 연한 노랑
        elif "Deprecated" in _status or "🗑" in _status or "폐기" in _status:
            bg_base = "E5E7EB"   # 회색

        # 대분류 표시
        major_display = tc.get("major", "") or tc["domain"]

        # 중분류/소분류
        middle_display = tc.get("middle") or DOMAIN_LABELS.get(tc["domain"], tc["domain"])
        minor_display  = tc.get("minor") or ""
        # 소분류 간결화 (v0.10.x):
        #   1) Markdown ** 강조 제거 — '**destructive 가드**' → 'destructive 가드'
        #   2) 메타 마커 제거 — '[통합 — 그룹 대표 화면 검증]' '[미결]' 등 대괄호 메타
        #   3) 중복 라벨 제거 — 'Normal 상태 표시 및 동작 — Normal 상태 UI 표시' 처럼
        #      앞뒤가 같은 키워드를 반복하면 시드 부분 제거
        #   4) 'A — B' 형태 두 문장 연결은 가독성을 위해 줄바꿈
        #   5) 너무 길면 절단 (35자 + 두 번째 줄도 35자)
        if minor_display:
            s = minor_display.strip()
            # 1) ** 강조 제거
            s = re.sub(r"\*\*([^*]+)\*\*", r"\1", s)
            # 2) 메타 마커 제거 — [통합 ...], [미결 ...], [그룹 ...] 등 대괄호 안 텍스트
            s = re.sub(r"\s*\[(통합|미결|그룹|메모|TODO|보류)[^\]]*\]\s*", " ", s)
            # 3) Legacy 부연설명 제거 — '…', 긴 괄호 텍스트 (15자 이상)
            #    예: '권한 목록 (Read account / Read balances / Withdrawals…)' → '권한 목록'
            s = re.sub(r"…+", "", s)
            s = re.sub(r"\s*\([^)]{15,}\)\s*", " ", s)
            # 3-b) Mismatched paren — 닫는 괄호 없이 긴 슬래시 나열 등은 '(' 이후 일부 보존
            #      예: '권한 목록 (Read account / Read balances / Withdrawals 항목 비활성 ...'
            #           → 닫는 ')' 부재 + 슬래시 2개 이상 → '(' 직전까지만 사용
            m_open = re.match(r"^(.+?)\s*\(([^)]+)$", s)
            if m_open and m_open.group(2).count("/") >= 2 and len(m_open.group(2)) > 20:
                s = m_open.group(1).strip()
            # 4) 'X — Y' 중복 키워드 제거: X 의 첫 단어가 Y 에 다시 등장하면 X 제거
            #    예: 'Normal 상태 표시 및 동작 — Normal 상태 UI 표시 확인' → 'Normal 상태 UI 표시 확인'
            m_dash = re.match(r"^(.+?)\s+[—–\-]{1,2}\s+(.+)$", s)
            if m_dash:
                seed, body = m_dash.group(1).strip(), m_dash.group(2).strip()
                seed_first = re.split(r"[\s/]", seed)[0]
                if len(seed_first) >= 3 and seed_first.lower() in body.lower():
                    s = body
            # 5) 짧은 괄호는 ', ' 로 변환 (단순 동의어/예시는 유지)
            s = re.sub(r"\s*\(([^)]{1,14})\)\s*", r" (\1) ", s)
            # 6) 양쪽 공백 dash 는 줄바꿈 (남아있는 경우만)
            s = re.sub(r"\s+[—–\-]{1,2}\s+", "\n", s)
            # 7) 줄별 길이 제한 32자 (한글 가독성, 30자 내 목표)
            cleaned_lines = []
            for line in s.split("\n"):
                line = re.sub(r"\s+", " ", line).strip(" .·-—:")
                if len(line) > 32:
                    # 마지막 공백 위치에서 절단해 단어 보존
                    cut = line[:32].rstrip()
                    last_space = cut.rfind(" ")
                    if last_space >= 20:
                        cut = cut[:last_space]
                    line = cut.rstrip(" .·-—:")
                if line:
                    cleaned_lines.append(line)
            minor_display = "\n".join(cleaned_lines)

        # 거래소 정보 추출
        exchange_text = ""
        note = tc.get("note", "")
        for ex in ["HL", "Hyperliquid", "BN", "Binance", "OKX", "Bybit", "Bitget", "Gate"]:
            if ex.lower() in note.lower():
                exchange_text = ex
                break

        col_data = {
            "상태":           tc.get("status", ""),  # 🆕/🔄/🗑️/빈값 (변경 이력 모드에서만)
            "TC ID":         tc["id"],
            "대분류":         major_display,
            "중분류":         middle_display,
            "소분류":         minor_display,
            "사전조건":       tc["given"],
            "테스트 스텝":    tc["when"],
            "기대결과":       tc["then"],
            "중요도":         _priority_to_kr(tc["priority"]),
            "대상 거래소":    exchange_text,
            "Smoke":         "Y" if tc.get("smoke") else "",
            "화면 코드":      tc.get("screen_code", ""),
            "수정 사유":      tc.get("change_reason", ""),
        }

        row_values = [col_data.get(name, "") for name in col_names]
        all_row_data.append(row_values)

        for i, (col_name, _) in enumerate(cols, 1):
            if col_name == "Smoke":
                cbg  = "E8F5E9" if tc.get("smoke") else bg_base
                bold = bool(tc.get("smoke"))
            elif col_name == "TC ID":
                cbg  = bg_base
                bold = tc["star"]
            elif col_name == "중요도":
                cbg  = PRIORITY_COLOR.get(tc["priority"], bg_base)
                bold = True
            else:
                cbg  = bg_base
                bold = tc["star"]
            align = "center" if col_name in ("Smoke", "상태", "중요도", "대상 거래소", "화면 코드") else "left"
            set_cell(ws, r, i, col_data.get(col_name, ""), bold=bold, bg=cbg, align_h=align)

        ws.row_dimensions[r].height = 80
        r += 1
        row_idx += 1

    # 컬럼 너비 자동 조정 (75퍼센타일 기준)
    min_widths = {"Smoke": 8, "상태": 11, "TC ID": 18, "중요도": 9, "대상 거래소": 12, "화면 코드": 12, "수정 사유": 24}
    widths = _calc_col_widths(col_names, all_row_data, min_widths=min_widths, max_width=55)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 필터 & 틀 고정
    last_col = get_column_letter(len(col_names))
    ws.auto_filter.ref = f"A1:{last_col}{r - 1}"
    ws.freeze_panes   = f"A{data_start + 1}"


# ── 통계 시트 ──────────────────────────────────────────────────────

def build_stats(ws, tcs, version):
    total  = len(tcs)
    smoke  = sum(1 for t in tcs if t.get("smoke"))

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10

    def sec_header(row, text):
        set_cell(ws, row, 2, text, bold=True, bg=C_DARK, font_color=C_WHITE, size=11)
        ws.merge_cells(f"B{row}:D{row}")
        ws.row_dimensions[row].height = 22

    def stat_row(row, label, count, pct_base=None, alt=False):
        bg = C_BLUE2 if alt else C_WHITE
        set_cell(ws, row, 2, label, bold=False, bg=bg)
        set_cell(ws, row, 3, count, bold=False, bg=bg, align_h="center")
        pct = f"{count*100//pct_base}%" if pct_base else ""
        set_cell(ws, row, 4, pct, bold=False, bg=bg, align_h="center")
        ws.row_dimensions[row].height = 18

    # 타이틀
    set_cell(ws, 2, 2,
             f"TC 통계  (v{version}  ·  전체 {total}개  ·  Smoke {smoke}개)",
             bold=True, bg=C_DARK, font_color=C_WHITE, size=12)
    ws.merge_cells("B2:D2")
    ws.row_dimensions[2].height = 26

    ws.row_dimensions[3].height = 8

    # ─ 전체 요약
    sec_header(4, "─ 전체 요약 ─")
    stat_row(5,  "전체 TC 수",    total,  total, alt=True)
    stat_row(6,  "🔥 Smoke Test", smoke,  total, alt=False)

    # ─ 화면별 TC 수 (ScreenCode · 중분류 이름)
    ws.row_dimensions[7].height = 8
    sec_header(8, "─ 화면별 TC 수 ─")
    domain_counts = defaultdict(int)
    domain_labels_dyn: dict[str, str] = {}
    for tc in tcs:
        d = (tc.get("domain") or "").strip().replace("*", "")  # '**' 마크업 잔여 방어
        if not d:
            continue
        domain_counts[d] += 1
        if d not in domain_labels_dyn:
            mid = (tc.get("middle") or "").strip()
            mid = re.sub(r"^중분류[:\s]*", "", mid).strip()
            mid = re.sub(r"\s*[\(\（][A-Z0-9\-]+[\)\）]", "", mid).strip()
            if mid:
                domain_labels_dyn[d] = mid
            elif d in DOMAIN_LABELS:
                domain_labels_dyn[d] = DOMAIN_LABELS[d]
            else:
                domain_labels_dyn[d] = d  # fallback — 아래에서 단독 표시
    for i, (domain, cnt) in enumerate(sorted(domain_counts.items())):
        label_val = domain_labels_dyn.get(domain, domain)
        # 코드와 label이 같으면 단독, 다르면 "코드 · 이름"
        label = domain if label_val == domain else f"{domain}  ·  {label_val}"
        stat_row(9 + i, label, cnt, total, alt=(i % 2 == 0))

    # ─ 우선순위별
    row = 9 + len(domain_counts) + 1
    ws.row_dimensions[row].height = 8
    row += 1
    sec_header(row, "─ 우선순위별 ─")
    row += 1
    for i, prio in enumerate(["High", "Medium", "Low"]):
        cnt = sum(1 for t in tcs if t["priority"] == prio)
        bg_map = {"High": "FCE4D6", "Medium": "FFF3E0", "Low": "E8F5E9"}
        cell_bg = bg_map.get(prio, C_WHITE)
        set_cell(ws, row, 2, prio, bold=True, bg=cell_bg)
        set_cell(ws, row, 3, cnt, bold=False, bg=cell_bg, align_h="center")
        set_cell(ws, row, 4, f"{cnt*100//total if total else 0}%", bg=cell_bg, align_h="center")
        ws.row_dimensions[row].height = 18
        row += 1

    # ─ 분류별
    ws.row_dimensions[row].height = 8
    row += 1
    sec_header(row, "─ 분류별 (TC 유형) ─")
    row += 1
    for i, cat in enumerate(["Positive", "Negative", "Edge"]):
        cnt = sum(1 for t in tcs if t["category"] == cat)
        cbg = CATEGORY_COLOR.get(cat, C_WHITE)
        set_cell(ws, row, 2, cat, bold=True, bg=cbg)
        set_cell(ws, row, 3, cnt, bold=False, bg=cbg, align_h="center")
        set_cell(ws, row, 4, f"{cnt*100//total if total else 0}%", bg=cbg, align_h="center")
        ws.row_dimensions[row].height = 18
        row += 1

    # NOTE: 이전에 있던 '─ 중분류별 ─' 섹션은 '─ 화면별 TC 수 ─' 와 중복되어 제거됨.
    # (화면별 섹션이 이미 ScreenCode · 중분류 이름을 함께 표시)

    # ─ [미결] / [신규]
    ws.row_dimensions[row].height = 8
    row += 1
    sec_header(row, "─ 태그별 ─")
    row += 1
    migyul = sum(1 for t in tcs if "[미결]" in t.get("note", ""))
    new_tc = sum(1 for t in tcs if t.get("is_new"))
    dev_ck = sum(1 for t in tcs if "[개발 확인]" in t.get("note", ""))
    for i, (label, cnt) in enumerate([
        ("[미결] 태그 (정책 미확정)", migyul),
        ("[신규] 태그 (코드 기반 도출)", new_tc),
        ("[개발 확인] 태그", dev_ck),
    ]):
        stat_row(row, label, cnt, total, alt=(i % 2 == 0))
        row += 1


# ── Smoke Test 시트 ─────────────────────────────────────────────────

def build_smoke(ws, tcs, config, include_change_columns: bool = True):
    smoke_tcs = [t for t in tcs if t.get("smoke")]
    build_tc_list(ws, smoke_tcs, config, include_change_columns=include_change_columns)


# ── Traceability Matrix 시트 ────────────────────────────────────────

def build_traceability(ws, tcs) -> int:
    """SCR(화면 코드) → TC 역참조 매트릭스 시트 생성. 반환: 유니크 SCR 수.
    screen_code가 비어있는 TC는 '(화면 코드 없음)' 그룹으로 집계.
    각 TC가 여러 코드를 가지면 쉼표 분리 후 각 코드에 모두 포함.
    """
    # SCR → [TC...] 맵핑 구성
    scr_map = defaultdict(list)  # code -> list of tc dict
    for tc in tcs:
        codes_str = (tc.get("screen_code") or "").strip()
        if not codes_str:
            scr_map["(화면 코드 없음)"].append(tc)
            continue
        for code in [c.strip() for c in codes_str.split(",") if c.strip()]:
            scr_map[code].append(tc)

    # 정렬: SCR-001, SCR-002... 숫자 파트 기준 / "없음" 그룹은 맨 뒤
    def sort_key(code):
        if code == "(화면 코드 없음)":
            return (1, 0, code)
        m = re.match(r"^(SCR|SCREEN|PAGE)-(.+)$", code)
        if m:
            suffix = m.group(2)
            # 숫자면 int, 아니면 큰 값 고정
            try:
                return (0, int(re.sub(r"\D", "", suffix) or 999999), code)
            except ValueError:
                return (0, 999999, code)
        return (0, 999999, code)

    sorted_codes = sorted(scr_map.keys(), key=sort_key)

    # 헤더
    headers = ["화면 코드", "대분류", "중분류", "TC 개수", "Smoke", "TC ID 목록"]
    widths = [14, 16, 18, 9, 8, 60]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        set_cell(ws, 1, i, h, bold=True, bg=C_DARK, font_color=C_WHITE,
                 size=10, align_h="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22

    r = 2
    row_idx = 0
    for code in sorted_codes:
        items = scr_map[code]
        # 대표 대분류/중분류 (다수결 또는 첫 등장)
        majors = [t.get("major", "") for t in items if t.get("major")]
        middles = [t.get("middle", "") for t in items if t.get("middle")]
        major_label = majors[0] if majors else ""
        middle_label = middles[0] if middles else ""
        # 섞여있는 경우 "외 N" 표기
        unique_majors = set(majors)
        unique_middles = set(middles)
        if len(unique_majors) > 1:
            major_label = f"{major_label} 외 {len(unique_majors)-1}"
        if len(unique_middles) > 1:
            middle_label = f"{middle_label} 외 {len(unique_middles)-1}"

        tc_ids = [t.get("id", "") for t in items if t.get("id")]
        smoke_count = sum(1 for t in items if t.get("smoke"))

        bg = C_ROW_A if row_idx % 2 == 0 else C_ROW_B
        set_cell(ws, r, 1, code, bold=True, bg=bg, align_h="center")
        set_cell(ws, r, 2, major_label, bg=bg, align_h="left")
        set_cell(ws, r, 3, middle_label, bg=bg, align_h="left")
        set_cell(ws, r, 4, len(tc_ids), bg=bg, align_h="center")
        set_cell(ws, r, 5, smoke_count if smoke_count else "", bg=bg, align_h="center")
        set_cell(ws, r, 6, ", ".join(tc_ids), bg=bg, align_h="left")
        ws.row_dimensions[r].height = max(20, min(100, 16 + 2 * (len(tc_ids) // 6)))
        r += 1
        row_idx += 1

    # 필터 + 틀 고정
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{r - 1}"
    ws.freeze_panes = "A2"

    # SCR 코드 있는 것만 센 유니크 수
    return sum(1 for c in sorted_codes if c != "(화면 코드 없음)")


# ── 변경 이력 시트 ───────────────────────────────────────────────────

def build_change_history(ws, tcs) -> int:
    """상태가 비어있지 않은 TC만 모아 변경 이력 시트 생성. 반환: 변경 TC 수.
    상태(🆕/🔄/🗑️), TC ID, 대분류/중분류, 화면 코드, 수정 사유 표기.
    """
    changed = [t for t in tcs if (t.get("status") or "").strip()]
    if not changed:
        return 0
    headers = ["상태", "TC ID", "화면 코드", "대분류", "중분류", "수정 사유"]
    widths  = [12, 16, 12, 16, 18, 50]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        set_cell(ws, 1, i, h, bold=True, bg=C_DARK, font_color=C_WHITE,
                 size=10, align_h="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22

    # 정렬: 🆕 신규 → 🔄 수정 → 🗑️ Deprecated 순
    def order_key(tc):
        s = tc.get("status", "")
        if "신규" in s or "🆕" in s: return (0, tc.get("id", ""))
        if "수정" in s or "🔄" in s: return (1, tc.get("id", ""))
        if "Deprecated" in s or "🗑" in s or "폐기" in s: return (2, tc.get("id", ""))
        return (3, tc.get("id", ""))
    changed.sort(key=order_key)

    r = 2
    for tc in changed:
        s = tc.get("status", "")
        if "신규" in s or "🆕" in s:
            bg = "DCFCE7"
        elif "수정" in s or "🔄" in s:
            bg = "FEF3C7"
        elif "Deprecated" in s or "🗑" in s or "폐기" in s:
            bg = "E5E7EB"
        else:
            bg = C_ROW_A
        set_cell(ws, r, 1, s, bold=True, bg=bg, align_h="center")
        set_cell(ws, r, 2, tc.get("id", ""), bg=bg, align_h="center")
        set_cell(ws, r, 3, tc.get("screen_code", ""), bg=bg, align_h="center")
        set_cell(ws, r, 4, tc.get("major", ""), bg=bg, align_h="left")
        set_cell(ws, r, 5, tc.get("middle", ""), bg=bg, align_h="left")
        set_cell(ws, r, 6, tc.get("change_reason", ""), bg=bg, align_h="left")
        ws.row_dimensions[r].height = 22
        r += 1

    ws.auto_filter.ref = f"A1:F{r-1}"
    ws.freeze_panes = "A2"
    return len(changed)


# ── 시트명 생성 ────────────────────────────────────────────────────
# Excel 제약: 31자 이내, `: \ / ? * [ ]` 금지, 같은 이름 중복 금지.

def parse_approved_classification(classification_md: str, screen_rows: list = None) -> tuple[set, dict]:
    """분류표 마크다운에서 ground truth 추출.
    반환: (allowed_majors, screen_to_major)
      - allowed_majors: 분류표의 모든 '## 대분류:' 값 set
      - screen_to_major: SCR-XXX → 정답 대분류 dict

    SCR 매핑 우선순위:
      1) screen_rows (parse_screen_list_table 결과) — 가장 정확한 ground truth
      2) classification_md 안의 소분류 라인에서 SCR-XXX 추출 (중분류명만 있어
         SCR ID 가 분류표에 없는 경우 대비 폴백)
    """
    allowed_majors: set[str] = set()
    screen_to_major: dict[str, str] = {}
    current_major = ""

    # 1) screen_rows 우선 사용 (가장 정확)
    if screen_rows:
        for r in screen_rows:
            scr = (r.get("id") or "").upper()
            major = (r.get("major") or "").strip()
            if scr and major:
                screen_to_major[scr] = major
                allowed_majors.add(major)

    # 2) classification_md 파싱 (대분류 set 보강 + screen_rows 누락 SCR 폴백)
    #    인식 패턴:
    #      A) '## 대분류: XXX' — allowed_majors 등록
    #      B) '<!-- SCR: SCR-001 -->' 메타 주석 — build_classification_from_screen_list 가 삽입
    #      C) 소분류 라인 안 'SCR-XXX' — 자유 형식 폴백
    scr_meta_pat = re.compile(r"<!--\s*SCR:\s*(SCR-[A-Za-z0-9]+)\s*-->", re.IGNORECASE)
    for line in classification_md.splitlines():
        m_major = re.match(r"^##\s+대분류[:\s]+(.+?)\s*$", line)
        if m_major:
            current_major = m_major.group(1).strip()
            allowed_majors.add(current_major)
            continue
        if current_major:
            # B) 메타 주석 우선
            m_meta = scr_meta_pat.search(line)
            if m_meta:
                scr = m_meta.group(1).upper()
                if scr not in screen_to_major:
                    screen_to_major[scr] = current_major
                continue
            # C) 소분류 라인의 SCR 패턴 폴백
            if line.lstrip().startswith("-"):
                for m in re.finditer(r"\bSCR-[A-Za-z0-9]+\b", line):
                    scr = m.group(0).upper()
                    if scr not in screen_to_major:
                        screen_to_major[scr] = current_major
    return allowed_majors, screen_to_major


def detect_suite_screen_mismatches(tcs: list, scr_to_screencode: dict) -> list:
    """TC ID 의 SuiteCode 와 '연관 화면' SCR 의 ScreenCode 가 일치하는지 검증.

    예: SM-OAC-015 (SuiteCode=OAC, OAuth Connect 화면 의미)
        연관 화면 = SCR-014 (Google Sign-in Complete, ScreenCode=GSC)
        → SuiteCode 'OAC' ≠ SCR-014 의 ScreenCode 'GSC' → 검출

    이런 케이스는 AI 가 화면 코드를 잘못 적은 결과 — 대분류·중분류·소분류 모두
    잘못된 화면 정보를 따라가서 일관된 듯 보이지만 실제론 다른 화면.

    Args:
        tcs:                parse_tc_markdown 결과
        scr_to_screencode:  {SCR-014: 'GSC', SCR-801: 'OAC', ...} 매핑 (screen_code_map.md 기반)
    Returns: 불일치 TC 리스트
    """
    mismatches = []
    for tc in tcs:
        tc_id = (tc.get("id") or "").strip()
        screen_code = (tc.get("screen_code") or "").strip()
        primary_scr = screen_code.split(",")[0].strip() if screen_code else ""
        if not tc_id or not primary_scr:
            continue

        # TC ID 의 SuiteCode 추출 — {ProjectCode}-{SuiteCode}-{NNN}
        # 예: SM-OAC-015 → OAC, SC-TRD-ORDR-001 → TRD-ORDR (다중 세그먼트 SuiteCode 도 지원)
        m = re.match(r"^[A-Z]{2,}-([A-Z][A-Z0-9\-]*)-\d+$", tc_id)
        if not m:
            continue
        tc_suite = m.group(1).upper()

        # 연관 화면 SCR 의 정답 ScreenCode
        correct_screencode = scr_to_screencode.get(primary_scr)
        if not correct_screencode:
            continue  # 매핑 없으면 검증 스킵

        if tc_suite != correct_screencode.upper():
            mismatches.append({
                "tc_id":            tc_id,
                "screen_code":      primary_scr,
                "current_suite":    tc_suite,
                "correct_suite":    correct_screencode,
                "reason":           "TC ID SuiteCode 와 연관 화면 ScreenCode 불일치",
            })
    return mismatches


def detect_major_mismatches(tcs: list, allowed_majors: set, screen_to_major: dict) -> list:
    """각 TC 의 대분류가 분류표와 일치하는지 검증.
    반환: 불일치 TC 리스트 [{tc_id, screen_code, current_major, correct_major, reason}, ...]
    자동 정정 안 함 — 검출만.

    검증 우선순위:
      1) 대분류 값이 분류표(allowed_majors)에 없음 → 명백한 오류
      2) 대분류 값은 분류표에 있지만, 이 TC 의 SCR 정답과 일치하지 않음 → 영역 혼동
    SCR 컨텍스트 없으면 (1)만 검증, (2)는 스킵.
    """
    mismatches = []
    for tc in tcs:
        current = (tc.get("major") or "").strip()
        scr_code = (tc.get("screen_code") or "").strip()
        # 첫 SCR 만 사용 (쉼표로 여러 개 있을 때)
        primary_scr = scr_code.split(",")[0].strip() if scr_code else ""

        if not current:
            continue  # 빈 대분류는 검증 스킵

        # (1) 분류표에 없는 대분류
        if current not in allowed_majors:
            correct = screen_to_major.get(primary_scr) if primary_scr else None
            mismatches.append({
                "tc_id":          tc.get("id", ""),
                "screen_code":    primary_scr or "(컨텍스트 없음)",
                "current_major":  current,
                "correct_major":  correct or "(분류표에서 SCR 매핑 못 찾음)",
                "reason":         "분류표 미등록 대분류",
            })
            continue

        # (2) 분류표에는 있지만 SCR 정답과 다름 (SCR 컨텍스트 있을 때만)
        if primary_scr and primary_scr in screen_to_major:
            correct = screen_to_major[primary_scr]
            if current != correct:
                mismatches.append({
                    "tc_id":          tc.get("id", ""),
                    "screen_code":    primary_scr,
                    "current_major":  current,
                    "correct_major":  correct,
                    "reason":         "SCR 소속 대분류와 불일치",
                })
    return mismatches


def build_classification_check(ws, major_mismatches: list, suite_mismatches: list) -> None:
    """이상 검출 시트 작성. 두 종류 불일치를 한 시트에 통합 표시.
      - 대분류 불일치 (AI 가 분류표 외 대분류 또는 다른 영역 대분류 채택)
      - SuiteCode/SCR 불일치 (TC ID 의 SuiteCode 가 연관 화면 SCR 의 ScreenCode 와 다름)
    호출부에서 (둘 합쳐서) 0건이면 시트 생성 안 함.
    """
    ws.sheet_view.showGridLines = False
    headers = ["TC ID", "화면 코드", "유형", "현재 값", "정답", "이유"]
    col_w = [18, 13, 22, 22, 22, 28]

    total = len(major_mismatches) + len(suite_mismatches)
    set_cell(ws, 1, 1,
             f"⚠ 분류 일관성 불일치 검출 — 총 {total}개 (대분류 {len(major_mismatches)} / SuiteCode-SCR {len(suite_mismatches)}). 검토 후 수정 필요 (자동 정정 X).",
             bold=True, bg=C_DARK, font_color=C_WHITE, size=11, align_h="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 24

    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        c = ws.cell(2, ci, h)
        c.font = make_font(bold=True, color=C_WHITE, size=10)
        c.fill = make_fill(C_DARK)
        c.alignment = make_align(wrap=True, h="center")
        c.border = make_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    # 정렬: 대분류 불일치 먼저, 그 다음 SuiteCode 불일치
    rows = []
    for m in major_mismatches:
        rows.append((m["tc_id"], m["screen_code"], "대분류 불일치",
                     m["current_major"], m["correct_major"], m["reason"]))
    for m in suite_mismatches:
        rows.append((m["tc_id"], m["screen_code"], "SuiteCode 불일치",
                     m["current_suite"], m["correct_suite"], m["reason"]))

    for ri, row in enumerate(rows, 3):
        bg = C_ROW_A if (ri - 3) % 2 == 0 else C_ROW_B
        for ci, val in enumerate(row, 1):
            align = "center" if ci in (2, 3) else "left"
            set_cell(ws, ri, ci, val, bg=bg, align_h=align)
        ws.row_dimensions[ri].height = 20

    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A2:{last_col}{len(rows) + 2}"
    ws.freeze_panes = "A3"


def _sheet_title_for_major(major_name: str, existing: list) -> str:
    """대분류명을 엑셀 시트명으로 변환. 중복이면 번호를 붙여 유일화."""
    s = re.sub(r'[:\\/\?\*\[\]]', ' ', major_name).strip() or "Sheet"
    # 괄호 코드 제거 — 시트명은 사람 친화형
    s = re.sub(r"\s*[\(\（][A-Z]{2,8}[\)\）]", "", s).strip()
    candidate = s
    if len(candidate) > 31:
        candidate = candidate[:31]
    base = candidate
    i = 2
    while candidate in existing:
        suffix = f" ({i})"
        candidate = (base[:31 - len(suffix)]) + suffix
        i += 1
    return candidate


# ── 메인 ───────────────────────────────────────────────────────────

def run_build(phase: str, tc_path, output_dir,
              verbose: bool = True,
              include_change_columns: bool = False,
              sheets: dict | None = None) -> dict:
    """재사용 가능한 Excel 빌드 엔트리 포인트.

    Args:
        phase:      "P_WebApp", "P2_Mobile" 등 PHASE_CONFIG 키
        tc_path:    tc_final.md 파일 경로 (str | Path)
        output_dir: 출력 디렉토리 (str | Path). 없으면 생성.
        verbose:    True면 진행 메시지 print (subprocess 호출 시 유용)
        include_change_columns: True면 상태/수정 사유 컬럼 + 변경 이력 시트 포함.
            기본 False — 신규 TC 생성에는 이 정보가 없으므로 생략.
            「기존 TC 수정」 플로우(/update-tc)에서만 True로 호출.
        sheets:     포함할 시트 선택 dict. None 이면 전체(Full Set) 생성.
            Keys (모두 bool):
              - "cover":          표지
              - "stats":          TC 통계
              - "smoke":          Smoke Test
              - "traceability":   Traceability Matrix
              - "tc_list":        TC 전체 목록 (필수 — False여도 항상 True 처리)
              - "change_history": 변경 이력 (include_change_columns 와 AND 조건)
            예) Light 모드: {"cover":False, "stats":False, "smoke":False,
                            "traceability":False, "tc_list":True, "change_history":False}

    Returns:
        {
          "ok":          bool,
          "out_path":    Path  — 생성된 xlsx 경로
          "total_tc":    int,
          "smoke_tc":    int,
          "high_tc":     int,
          "scr_count":   int,
          "major_count": int,
          "changed_tc":  int,
        }
    """
    # sheets 파라미터 기본값(Full Set) 처리 — TC 목록은 항상 True로 강제
    DEFAULT_SHEETS = {
        "cover": True, "stats": True, "smoke": True,
        "traceability": True, "tc_list": True, "change_history": True,
    }
    if sheets is None:
        sheets = DEFAULT_SHEETS.copy()
    else:
        # 누락된 키는 기본값 채움 + tc_list 는 항상 True 보장
        merged = DEFAULT_SHEETS.copy()
        merged.update({k: bool(v) for k, v in sheets.items()})
        merged["tc_list"] = True
        sheets = merged
    tc_path = Path(tc_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    config   = PHASE_CONFIG.get(phase, PHASE_CONFIG["P2_Mobile"])
    date_str = datetime.today().strftime("%Y%m%d")
    version  = get_next_version(output_dir, phase, date_str)
    filename = f"SPCY_TC_{phase}_{date_str}_v{version}.xlsx"
    out_path = output_dir / filename

    if verbose:
        print(f"TC 파싱 중: {tc_path}")
    tcs = parse_tc_markdown(tc_path)
    if verbose:
        print(f"  → {len(tcs)}개 TC 파싱 완료")

    wb = openpyxl.Workbook()

    # Smoke 마킹을 먼저 수행 — 모든 시트 빌더가 smoke 속성 참조 가능
    _mark_smoke(tcs)

    # ── 공통 시트: 사용자 옵션(sheets) 에 따라 선택적 생성 ──
    # 첫 시트는 wb.active (자동 생성된 'Sheet') — 표지가 꺼져있으면 임시 보관 후 마지막에 제거
    placeholder_ws = wb.active
    placeholder_ws.title = "_placeholder"

    # 표지
    if sheets.get("cover"):
        ws_cover = wb.create_sheet("표지", 0)  # 항상 첫 번째 위치
        build_cover(ws_cover, tcs, config, date_str, version)

    # TC 통계
    if sheets.get("stats"):
        ws_stats = wb.create_sheet("TC 통계")
        build_stats(ws_stats, tcs, version)

    # Smoke Test
    if sheets.get("smoke"):
        ws_smoke = wb.create_sheet("Smoke Test")
        build_smoke(ws_smoke, tcs, config, include_change_columns=include_change_columns)

    # Traceability Matrix
    scr_count = 0
    if sheets.get("traceability"):
        ws_trace = wb.create_sheet("Traceability")
        scr_count = build_traceability(ws_trace, tcs)
        if verbose:
            if scr_count > 0:
                print(f"  → Traceability Matrix 생성 — 유니크 화면 코드 {scr_count}개")
            else:
                print(f"  → Traceability Matrix 생성 — 화면 코드 없음 (TC 전체 '(화면 코드 없음)' 그룹)")

    # 변경 이력 시트: 사용자가 켰고 + include_change_columns=True 이고 + 변경된 TC 가 있어야 생성
    changed_tc = 0
    if (sheets.get("change_history") and include_change_columns
            and any((t.get("status") or "").strip() for t in tcs)):
        ws_changes = wb.create_sheet("변경 이력")
        changed_tc = build_change_history(ws_changes, tcs)
        if verbose:
            print(f"  → 변경 이력 시트 생성 — {changed_tc}개 변경 TC")

    # 대분류별 그룹핑 (major 컬럼 기준, 없으면 domain 코드 폴백)
    major_groups = defaultdict(list)
    major_order = []  # 등장 순서 유지
    for tc in tcs:
        key = (tc.get("major") or "").strip() or tc.get("domain") or "ETC"
        if key not in major_groups:
            major_order.append(key)
        major_groups[key].append(tc)

    # 각 대분류별 시트 생성
    for major_name in major_order:
        tcs_in_major = major_groups[major_name]
        sheet_title = _sheet_title_for_major(major_name, existing=wb.sheetnames)
        ws = wb.create_sheet(sheet_title)
        build_tc_list(ws, tcs_in_major, config, group_by="middle",
                      include_change_columns=include_change_columns)

    if verbose:
        print(f"  → 대분류별 시트 {len(major_order)}개 생성")

    # ── 분류 일관성 검출 (검출 전용, 자동 정정 X) ───────────────────────
    # 두 종류 불일치를 한 시트로 통합:
    #   1) 대분류 불일치  — AI 가 분류표 외 또는 다른 영역 대분류 채택
    #   2) SuiteCode/SCR 불일치 — TC ID 의 SuiteCode 와 연관 화면 SCR 의 ScreenCode 가 다름
    # 분류표는 tc_path 디렉토리의 classification_v1_APPROVED.md 에서 로드.
    classification_path = Path(tc_path).parent / "classification_v1_APPROVED.md"
    if classification_path.exists():
        try:
            classification_md = classification_path.read_text(encoding="utf-8")
            allowed_majors, screen_to_major = parse_approved_classification(classification_md)
            major_mm = detect_major_mismatches(tcs, allowed_majors, screen_to_major)

            # SCR → ScreenCode 매핑은 메타 주석 + 화면 ID 패턴에서 직접 추출
            # 분류표에 '<!-- SCR: SCR-XXX -->' 다음 줄에 중분류명, 또는 직접 화면 코드
            # screen_code_map.md 의 SCR ↔ ScreenCode 매핑이 필요해 별도 로드.
            scr_to_screencode = {}
            try:
                # screen_code_map.md 는 tc-agent/projects/{project}/ 위치
                # tc_path 가 workspace 안이므로 거기서는 직접 못 찾음. 환경 변수나
                # 휴리스틱으로 찾기 — 일단 sibling 인 spec/projects 폴더 확인.
                from . import build_excel as _self  # noqa
            except Exception:
                pass
            # 간단 휴리스틱: TC 의 screen_code 와 SuiteCode 직접 매핑 추출
            # (TC ID 가 보통 screen-based 면 ScreenCode = SuiteCode)
            # → TC 들 자체에서 SCR-X → SuiteCode 빈도 집계 (가장 흔한 게 정답)
            from collections import Counter
            scr_suite_count: dict[str, Counter] = {}
            for tc in tcs:
                sc = (tc.get("screen_code") or "").split(",")[0].strip()
                tc_id = tc.get("id", "")
                m = re.match(r"^[A-Z]{2,}-([A-Z][A-Z0-9\-]*)-\d+$", tc_id)
                if sc and m:
                    scr_suite_count.setdefault(sc, Counter())[m.group(1).upper()] += 1
            for scr, counter in scr_suite_count.items():
                # 가장 빈번한 SuiteCode 를 정답으로 (다수결)
                scr_to_screencode[scr] = counter.most_common(1)[0][0]

            suite_mm = detect_suite_screen_mismatches(tcs, scr_to_screencode)

            total_mm = len(major_mm) + len(suite_mm)
            if total_mm > 0:
                ws_check = wb.create_sheet("classification_check")
                build_classification_check(ws_check, major_mm, suite_mm)
                if verbose:
                    print(f"  → ⚠ 분류 불일치 {total_mm}개 검출 (대분류 {len(major_mm)} / SuiteCode {len(suite_mm)}) — 'classification_check' 시트 참고")
            elif verbose:
                print(f"  → ✓ 분류 일관성 검증 통과 (0개 불일치)")
        except Exception as e:
            if verbose:
                print(f"  → 분류 검증 스킵: {e}")

    # 임시 placeholder 시트 제거 (표지가 꺼졌을 때 첫 시트가 비어있는 상태 방지)
    if "_placeholder" in wb.sheetnames:
        # 다른 시트가 1개 이상 있을 때만 제거 (워크북에는 최소 1개 시트 필요)
        if len(wb.sheetnames) > 1:
            del wb["_placeholder"]
        else:
            # 모든 시트가 꺼진 극한 케이스 — placeholder 를 TC 목록 폴백으로 변환
            wb["_placeholder"].title = "📌 TC 목록 (폴백)"

    wb.save(out_path)

    smoke_n = sum(1 for t in tcs if t.get("smoke"))
    high    = sum(1 for t in tcs if t["priority"] == "High")
    if verbose:
        print(f"\n✅ Excel 저장 완료: {out_path}")
        print(f"   총 TC: {len(tcs)}개 | High: {high}개 | 🔥 Smoke Test: {smoke_n}개")
        print(f"   Phase: {config['code']} | 버전: v{version}")
        print(f"   거래소 컬럼: {config['exchanges'] if config['exchanges'] else '없음 (단일 거래소)'}")

    return {
        "ok":          True,
        "out_path":    out_path,
        "total_tc":    len(tcs),
        "smoke_tc":    smoke_n,
        "high_tc":     high,
        "scr_count":   scr_count,
        "major_count": len(major_order),
        "changed_tc":  changed_tc,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--phase",  default="P2_Mobile")
    parser.add_argument("--tc",     default="phase2-mobile/_workspace/05_review/tc_final.md")
    parser.add_argument("--output", default="phase2-mobile/outputs")
    args = parser.parse_args()

    base       = Path(__file__).parent.parent
    tc_path    = base / args.tc
    output_dir = base / args.output
    run_build(args.phase, tc_path, output_dir, verbose=True)


if __name__ == "__main__":
    main()
