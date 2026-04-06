#!/usr/bin/env python3
"""
TC 자동화 웹 인터페이스
실행: python3 scripts/upload_server.py
접속: http://localhost:5001
"""

import sys
import os
import json
import subprocess
from pathlib import Path
from flask import Flask, request, jsonify, render_template_string, send_file

BASE_DIR          = Path(__file__).parent.parent
SPECS_DIR         = BASE_DIR / "specs"
OUTPUTS_DIR       = BASE_DIR / "outputs"
POLICY_DIR        = BASE_DIR / "policy"
BUILD_SCRIPT      = BASE_DIR / "scripts" / "build_tc_v8.py"
CONFIG_FILE       = BASE_DIR / "config.json"
DRIVE_CREDS_FILE  = BASE_DIR / "credentials.json"
DRIVE_TOKEN_FILE  = BASE_DIR / ".drive_token.json"

SPECS_DIR.mkdir(exist_ok=True)
OUTPUTS_DIR.mkdir(exist_ok=True)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB

# ── Excel 생성 ─────────────────────────────────────────────────
def generate_excel_from_json(tc_data, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    tcs     = tc_data.get("tcs", [])
    feature = tc_data.get("feature_name", "Feature")

    NAVY, BLUE, TEAL = "1E2761", "3557A0", "028090"
    T1, T2, SUB     = "1C6E38", "2E4057", "1B3A5A"
    LIGHT, WHITE    = "EEF1F8", "FFFFFF"
    PHASE_PAL = ["EBF5FB","FEF9E7","FDECEA","EAF4F0","F4ECF7",
                 "FEF0E7","E8F8F5","FDF2F8","F0F3FF","FFFDE7"]

    def bdr():
        s = Side(style="thin", color="D0D7DE")
        return Border(left=s, right=s, top=s, bottom=s)
    def fill(c): return PatternFill("solid", fgColor=c)
    def hfont(c=WHITE): return Font(name="Calibri", bold=True, size=10, color=c)
    def dfont(c="222222"): return Font(name="Calibri", size=9, color=c)
    def center(w=False): return Alignment(horizontal="center", vertical="center", wrap_text=w)
    def left(w=True): return Alignment(horizontal="left", vertical="center", wrap_text=w)

    wb = Workbook()
    wb.remove(wb.active)

    # 📋 표지
    cov = wb.create_sheet("📋 표지")
    cov.sheet_view.showGridLines = False
    cov.merge_cells("A1:R5")
    c = cov["A1"]
    c.value = f"TC Checklist  ·  {feature}"
    c.font = Font(name="Calibri", bold=True, size=22, color=WHITE)
    c.fill = fill(NAVY); c.alignment = center()
    cov.row_dimensions[1].height = 90
    for ri, (lbl, val) in enumerate([
        ("작성일", datetime.now().strftime("%Y-%m-%d")),
        ("기능명", feature),
        ("총 TC 수", len(tcs)),
        ("최소 TC (★)", sum(1 for t in tcs if t.get("min_set"))),
        ("대상 거래소", "Gate / OKX / Bybit / Bitget / Hyperliquid"),
    ], 7):
        c = cov.cell(ri, 1, lbl)
        c.font = hfont(); c.fill = fill(BLUE); c.alignment = center()
        cov.merge_cells(f"A{ri}:C{ri}")
        c = cov.cell(ri, 4, val)
        c.font = Font(name="Calibri", size=11, color=NAVY)
        c.fill = fill(LIGHT); c.alignment = left(w=False)
        cov.merge_cells(f"D{ri}:R{ri}")
        cov.row_dimensions[ri].height = 26
    for i in range(1, 19): cov.column_dimensions[get_column_letter(i)].width = 8
    cov.column_dimensions["A"].width = 14
    cov.column_dimensions["D"].width = 30

    # 🧪 TC 전체목록
    ws = wb.create_sheet("🧪 TC 전체목록")
    ws.freeze_panes = "A3"
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(1, ci, h)
        c.font = hfont(); c.alignment = center(); c.border = bdr()
        if h in ["Gate","Bitget","Hyperliquid"]: c.fill = fill(T2)
        elif h in ["OKX","Bybit"]:               c.fill = fill(T1)
        else:                                    c.fill = fill(NAVY)
    ws.row_dimensions[1].height = 22
    tester = {"Gate":"Tester 2","OKX":"Tester 1","Bybit":"Tester 1",
              "Bitget":"Tester 2","Hyperliquid":"Tester 2"}
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(2, ci, tester.get(h, ""))
        c.font = Font(name="Calibri", size=8, color="AECCE8")
        c.alignment = center(); c.fill = fill(SUB); c.border = bdr()
    ws.row_dimensions[2].height = 14

    dv = DataValidation(type="list", formula1='"Pass,Fail,N/T,N/A"', allow_blank=True)
    ws.add_data_validation(dv)

    phases = list(dict.fromkeys(t.get("phase","") for t in tcs))
    phase_color = {p: PHASE_PAL[i % len(PHASE_PAL)] for i, p in enumerate(phases)}

    def write_row(sheet, ri, tc, bg):
        na = set(tc.get("na_exchanges", []))
        tid = ("★ " if tc.get("min_set") else "") + tc.get("tc_id","")
        vals = [
            tid, tc.get("phase",""), tc.get("main_category",""),
            tc.get("sub_category",""), tc.get("micro_category",""),
            tc.get("scenario",""), tc.get("priority",""), tc.get("importance",""),
            tc.get("given",""), tc.get("when",""), tc.get("then",""), "",
            "N/A" if "Gate"        in na else "N/T",
            "N/A" if "OKX"         in na else "N/T",
            "N/A" if "Bybit"       in na else "N/T",
            "N/A" if "Bitget"      in na else "N/T",
            "N/A" if "Hyperliquid" in na else "N/T",
            tc.get("note",""),
        ]
        for ci, val in enumerate(vals, 1):
            c = sheet.cell(ri, ci, val)
            c.fill = fill(bg); c.border = bdr()
            h = HEADERS[ci-1]
            if h in EXCHANGES:
                c.alignment = center()
                c.font = dfont("999999" if val == "N/A" else "222222")
                if val == "N/T": dv.add(c)
            elif h in ["TC ID","단계","우선순위","중요도"]:
                c.alignment = center(); c.font = dfont()
            else:
                c.alignment = left(); c.font = dfont()
        sheet.row_dimensions[ri].height = 45

    for ri, tc in enumerate(tcs, 3):
        write_row(ws, ri, tc, phase_color.get(tc.get("phase",""), "F8F9FA"))

    col_widths = [14,10,12,14,12,30,8,9,35,35,40,18,8,8,8,8,12,18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 🎯 최소 TC 세트
    min_tcs = [t for t in tcs if t.get("min_set")]
    if min_tcs:
        ms = wb.create_sheet("🎯 최소 TC 세트")
        ms.freeze_panes = "A2"
        for ci, h in enumerate(HEADERS, 1):
            c = ms.cell(1, ci, h)
            c.font = hfont(); c.alignment = center(); c.border = bdr()
            if h in ["Gate","Bitget","Hyperliquid"]: c.fill = fill(T2)
            elif h in ["OKX","Bybit"]:               c.fill = fill(T1)
            else:                                    c.fill = fill(TEAL)
        ms.row_dimensions[1].height = 22
        for ri, tc in enumerate(min_tcs, 2):
            write_row(ms, ri, tc, "E0F7FA")
        for i, w in enumerate(col_widths, 1):
            ms.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)

# ────────────────────────────────────────────────────────────────
#  HTML
# ────────────────────────────────────────────────────────────────
HTML = r"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>TC 자동화 시스템</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
  background:#EEF1F8;min-height:100vh;padding:32px 16px}
.header{text-align:center;margin-bottom:28px}
.header h1{font-size:22px;font-weight:800;color:#1E2761}
.header p{font-size:13px;color:#666;margin-top:4px}

/* 스텝 바 */
.steps{display:flex;align-items:center;justify-content:center;
  max-width:640px;margin:0 auto 28px}
.step{display:flex;flex-direction:column;align-items:center;gap:6px;flex:1}
.sc{width:36px;height:36px;border-radius:50%;display:flex;align-items:center;
  justify-content:center;font-size:14px;font-weight:700;
  background:#D1D9EE;color:#888;transition:all .3s}
.sc.active{background:#3557A0;color:#fff}
.sc.done{background:#028090;color:#fff}
.sl{font-size:11px;color:#888;font-weight:600;white-space:nowrap}
.sl.active{color:#3557A0}
.sl.done{color:#028090}
.step-line{flex:1;height:2px;background:#D1D9EE;margin-bottom:22px;transition:background .3s}
.step-line.done{background:#028090}

/* 카드 */
.card{background:#fff;border-radius:16px;padding:28px 32px;
  margin:0 auto 20px;max-width:720px;
  box-shadow:0 2px 10px rgba(0,0,0,.07);transition:opacity .3s}
.card.disabled{opacity:.4;pointer-events:none}
.card-hdr{display:flex;align-items:center;gap:10px;margin-bottom:20px}
.badge{width:26px;height:26px;border-radius:50%;background:#3557A0;color:#fff;
  font-size:12px;font-weight:700;display:flex;align-items:center;
  justify-content:center;flex-shrink:0}
.card-title{font-size:15px;font-weight:700;color:#1E2761}

/* 드롭존 */
.dropzone{border:2px dashed #3557A0;border-radius:12px;padding:40px 24px;
  text-align:center;cursor:pointer;transition:all .2s;background:#F7F9FF}
.dropzone:hover,.dropzone.drag-over{border-color:#028090;background:#EFF9FB}
.dz-icon{font-size:36px;margin-bottom:10px}
.dz-label{font-size:14px;font-weight:600;color:#1E2761;margin-bottom:4px}
.dz-hint{font-size:12px;color:#999}
.dropzone input{display:none}
.btn-pick{margin-top:14px;padding:8px 22px;background:#3557A0;color:#fff;
  border:none;border-radius:8px;font-size:13px;font-weight:600;
  cursor:pointer;transition:background .2s}
.btn-pick:hover{background:#1E2761}

/* 파일 목록 */
.file-list{margin-top:16px;display:flex;flex-direction:column;gap:8px}
.file-row{display:flex;align-items:center;gap:10px;
  padding:10px 14px;background:#F4F6FB;border-radius:10px}
.fname{flex:1;font-size:13px;font-weight:500;color:#222;
  overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.fsize{font-size:12px;color:#aaa;flex-shrink:0}
.btn-del{background:none;border:none;cursor:pointer;color:#ccc;
  font-size:15px;transition:color .15s;flex-shrink:0}
.btn-del:hover{color:#e53e3e}
.empty-hint{text-align:center;padding:18px;color:#bbb;font-size:13px}

/* STEP 2 */
.spec-summary{display:flex;align-items:center;gap:8px;
  padding:12px 16px;background:#F4F6FB;border-radius:10px;margin-bottom:16px}
.guide-box{background:#EFF6FF;border-left:4px solid #3557A0;
  border-radius:0 8px 8px 0;padding:12px 16px;font-size:13px;
  color:#444;line-height:1.8;margin-bottom:16px}
.guide-box strong{color:#1E2761}
.btn-gen{width:100%;padding:14px;background:#1E2761;color:#fff;border:none;
  border-radius:12px;font-size:15px;font-weight:700;cursor:pointer;
  transition:background .2s;display:flex;align-items:center;
  justify-content:center;gap:8px}
.btn-gen:hover:not(:disabled){background:#3557A0}
.btn-gen:disabled{background:#aaa;cursor:not-allowed}
.progress-bar{height:4px;background:#EEF1F8;border-radius:4px;
  overflow:hidden;margin-top:12px;display:none}
.progress-fill{height:100%;background:linear-gradient(90deg,#3557A0,#028090);
  width:0%;transition:width .4s;border-radius:4px}
.log-box{margin-top:14px;background:#0F1923;border-radius:10px;
  padding:16px;font-family:monospace;font-size:12px;color:#7EE8A2;
  line-height:1.8;max-height:180px;overflow-y:auto;display:none;
  white-space:pre-wrap;word-break:break-all}

/* STEP 3 */
.result-file{display:flex;align-items:center;gap:14px;
  padding:16px;background:#F0FBF8;border:1px solid #B2E8DC;
  border-radius:12px;margin-bottom:16px}
.ricon{font-size:32px;flex-shrink:0}
.rinfo{flex:1;overflow:hidden}
.rname{font-size:14px;font-weight:700;color:#1E2761;
  overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.rsize{font-size:12px;color:#888;margin-top:2px}
.result-empty{text-align:center;padding:24px;color:#bbb;font-size:13px}
.action-row{display:flex;gap:12px}
.btn-dl{flex:1;padding:12px;background:#028090;color:#fff;border:none;
  border-radius:10px;font-size:14px;font-weight:600;cursor:pointer;
  transition:background .2s;display:flex;align-items:center;
  justify-content:center;gap:6px}
.btn-dl:hover{background:#016570}
.btn-drive{flex:1;padding:12px;background:#fff;color:#1E2761;
  border:2px solid #D1D9EE;border-radius:10px;font-size:14px;
  font-weight:600;cursor:pointer;transition:all .2s;
  display:flex;align-items:center;justify-content:center;gap:6px}
.btn-drive:hover{border-color:#3557A0;background:#F7F9FF}

/* 드라이브 모달 */
.modal-bg{position:fixed;inset:0;background:rgba(0,0,0,.45);
  display:none;align-items:center;justify-content:center;z-index:999}
.modal-bg.open{display:flex}
.modal{background:#fff;border-radius:16px;padding:32px;max-width:440px;
  width:90%;box-shadow:0 8px 32px rgba(0,0,0,.18)}
.modal h2{font-size:17px;font-weight:700;color:#1E2761;margin-bottom:16px}
.modal ol{padding-left:20px;color:#444;font-size:13px;line-height:2.2}
.modal-btns{display:flex;gap:10px;margin-top:22px}
.btn-open-drive{flex:1;padding:11px;background:#4285F4;color:#fff;border:none;
  border-radius:8px;font-size:13px;font-weight:600;cursor:pointer}
.btn-close{flex:1;padding:11px;background:#F4F6FB;color:#444;border:none;
  border-radius:8px;font-size:13px;font-weight:600;cursor:pointer}

/* 토스트 */
#toast{position:fixed;bottom:28px;left:50%;transform:translateX(-50%);
  padding:12px 24px;border-radius:10px;font-size:14px;font-weight:600;
  display:none;z-index:1000;box-shadow:0 4px 16px rgba(0,0,0,.2);white-space:nowrap}
#toast.success{background:#028090;color:#fff}
#toast.error{background:#e53e3e;color:#fff}

@media(max-width:480px){.action-row{flex-direction:column}.card{padding:20px 18px}}
</style>
</head>
<body>

<div class="header">
  <h1>🤖 TC 자동화 시스템 <span style="font-size:13px;font-weight:500;color:#3557A0;background:#E8EDF8;padding:2px 10px;border-radius:12px;vertical-align:middle">v1.0</span></h1>
  <p>기획서 PDF 업로드 → TC Excel 생성 → 다운로드</p>
  <div style="margin-top:12px">
    <button onclick="shutdownServer()" style="background:none;border:1px solid #CCC;
      color:#999;font-size:11px;padding:4px 12px;border-radius:12px;cursor:pointer">
      ⏹ 서버 종료
    </button>
  </div>
</div>

<!-- 스텝 바 -->
<div class="steps">
  <div class="step">
    <div class="sc active" id="sc1">1</div>
    <div class="sl active" id="sl1">기획서 업로드</div>
  </div>
  <div class="step-line" id="line1"></div>
  <div class="step">
    <div class="sc" id="sc2">2</div>
    <div class="sl" id="sl2">TC AI 생성</div>
  </div>
  <div class="step-line" id="line2"></div>
  <div class="step">
    <div class="sc" id="sc3">3</div>
    <div class="sl" id="sl3">결과 다운로드</div>
  </div>
</div>

<!-- STEP 1 -->
<div class="card" id="card1">
  <div class="card-hdr">
    <div class="badge">1</div>
    <span class="card-title">기획서 PDF 업로드</span>
  </div>
  <div class="dropzone" id="dropzone">
    <div class="dz-icon">📄</div>
    <div class="dz-label">PDF 파일을 여기에 드래그 앤 드롭</div>
    <div class="dz-hint">또는 버튼을 클릭해서 파일 선택 (최대 50MB)</div>
    <input type="file" id="fileInput" accept=".pdf" multiple>
    <br>
    <button class="btn-pick" onclick="document.getElementById('fileInput').click()">📂 파일 선택</button>
  </div>
  <div class="file-list" id="spec-list"></div>
</div>

<!-- STEP 2 -->
<div class="card disabled" id="card2">
  <div class="card-hdr">
    <div class="badge" style="background:#028090">2</div>
    <span class="card-title">TC Excel 생성</span>
  </div>
  <div class="spec-summary">
    <span style="font-size:20px">📋</span>
    <div style="font-size:13px;color:#444;line-height:1.6">
      업로드된 기획서: <strong id="spec-names">없음</strong><br>
      <span style="color:#888;font-size:12px">업로드된 PDF를 기반으로 TC 정책 문서를 적용해 Excel을 생성합니다.</span>
    </div>
  </div>
  <button class="btn-gen" id="btn-gen" onclick="generate()">
    ⚡ TC Excel 생성 시작
  </button>
  <div class="log-box" id="log-box"></div>
</div>

<!-- STEP 3 -->
<div class="card disabled" id="card3">
  <div class="card-hdr">
    <div class="badge" style="background:#028090">3</div>
    <span class="card-title">결과 다운로드</span>
  </div>
  <div id="result-area">
    <div class="result-empty">TC 생성 후 여기에 파일이 표시됩니다.</div>
  </div>
</div>

<!-- credentials 안내 모달 -->
<div class="modal-bg" id="drive-modal">
  <div class="modal">
    <h2>🔑 Google Drive 연동 설정</h2>
    <p style="font-size:13px;color:#444;margin-bottom:14px">
      최초 1회 <strong>credentials.json</strong> 파일 설정이 필요합니다.
    </p>
    <ol style="padding-left:20px;color:#444;font-size:13px;line-height:2.4">
      <li><a href="https://console.cloud.google.com/apis/credentials" target="_blank" style="color:#3557A0">Google Cloud Console</a> 접속</li>
      <li><strong>+ 사용자 인증 정보 만들기 → OAuth 클라이언트 ID</strong></li>
      <li>애플리케이션 유형: <strong>데스크톱 앱</strong> 선택 후 만들기</li>
      <li>JSON 다운로드 → <strong>tc.automation 폴더에 <code>credentials.json</code>으로 저장</strong></li>
      <li>저장 후 다시 <strong>Drive에 올리기</strong> 클릭 → 브라우저에서 Google 계정 인증</li>
    </ol>
    <div class="modal-btns">
      <button class="btn-open-drive"
        onclick="window.open('https://console.cloud.google.com/apis/credentials','_blank')">
        Cloud Console 열기
      </button>
      <button class="btn-close" onclick="closeModal()">닫기</button>
    </div>
  </div>
</div>

<div id="toast"></div>

<script>
function fmt(b){
  if(b<1024)return b+' B';
  if(b<1024*1024)return (b/1024).toFixed(1)+' KB';
  return (b/1024/1024).toFixed(1)+' MB';
}
function toast(msg,type='success'){
  const t=document.getElementById('toast');
  t.textContent=msg;t.className=type;t.style.display='block';
  setTimeout(()=>t.style.display='none',3000);
}

function setStep(n){
  [1,2,3].forEach(i=>{
    const sc=document.getElementById('sc'+i);
    const sl=document.getElementById('sl'+i);
    if(i<n){sc.className='sc done';sl.className='sl done';}
    else if(i===n){sc.className='sc active';sl.className='sl active';}
    else{sc.className='sc';sl.className='sl';}
    if(i<3)document.getElementById('line'+i).className=i<n?'step-line done':'step-line';
  });
  document.getElementById('card2').classList.toggle('disabled',n<2);
  document.getElementById('card3').classList.toggle('disabled',n<3);
}

/* STEP 1 */
async function loadSpecs(){
  const r=await fetch('/files');
  const files=await r.json();
  const list=document.getElementById('spec-list');
  document.getElementById('spec-names').textContent=
    files.length?files.map(f=>f.name).join(', '):'없음';
  if(!files.length){
    list.innerHTML='<div class="empty-hint">업로드된 기획서가 없습니다.</div>';
    setStep(1); return;
  }
  list.innerHTML=files.map(f=>`
    <div class="file-row">
      <span style="font-size:18px">📄</span>
      <span class="fname" title="${f.name}">${f.name}</span>
      <span class="fsize">${fmt(f.size)}</span>
      <button class="btn-del" onclick="delSpec('${f.name}')" title="삭제">✕</button>
    </div>`).join('');
  setStep(2);
}

async function uploadFiles(files){
  for(const file of files){
    if(!file.name.toLowerCase().endsWith('.pdf')){
      toast('❌ PDF 파일만 업로드 가능합니다.','error'); continue;
    }
    const fd=new FormData(); fd.append('file',file);
    const r=await fetch('/upload',{method:'POST',body:fd});
    const d=await r.json();
    if(d.ok) toast('✅ '+file.name+' 업로드 완료!');
    else toast('❌ '+d.error,'error');
  }
  loadSpecs();
}

async function delSpec(name){
  if(!confirm('"'+name+'" 파일을 삭제할까요?'))return;
  const r=await fetch('/delete',{method:'POST',
    headers:{'Content-Type':'application/json'},body:JSON.stringify({name})});
  const d=await r.json();
  if(d.ok){toast('🗑 삭제 완료');loadSpecs();}
  else toast('❌ 삭제 실패','error');
}

const dz=document.getElementById('dropzone');
dz.addEventListener('dragover',e=>{e.preventDefault();dz.classList.add('drag-over')});
dz.addEventListener('dragleave',()=>dz.classList.remove('drag-over'));
dz.addEventListener('drop',e=>{e.preventDefault();dz.classList.remove('drag-over');
  uploadFiles(Array.from(e.dataTransfer.files));});
document.getElementById('fileInput').addEventListener('change',e=>{
  uploadFiles(Array.from(e.target.files));e.target.value='';});

/* STEP 2: TC 빌드 */
async function generate(){
  const btn=document.getElementById('btn-gen');
  const log=document.getElementById('log-box');
  btn.disabled=true;
  btn.textContent='⏳ 생성 중...';
  log.style.display='block';
  log.style.color='#7EE8A2';
  log.textContent='⚙️ TC Excel 생성 중...\n';
  try{
    const r=await fetch('/generate',{method:'POST'});
    const d=await r.json();
    if(d.ok){
      log.textContent+='✅ 완료!\n'+(d.output||'');
      toast('🎉 TC 생성 완료!');
      setStep(3);
      renderResults(d.files);
    }else{
      log.style.color='#FF6B6B';
      log.textContent+='❌ '+d.error;
      toast('❌ 생성 실패','error');
    }
  }catch(e){
    log.style.color='#FF6B6B';
    log.textContent+='❌ 오류: '+e.message;
    toast('❌ 오류 발생','error');
  }
  btn.disabled=false;
  btn.textContent='⚡ TC Excel 생성 시작';
}

/* STEP 3 */
function renderResults(files){
  const area=document.getElementById('result-area');
  if(!files||!files.length){
    area.innerHTML='<div class="result-empty">생성된 파일이 없습니다.</div>'; return;
  }
  area.innerHTML=files.map(f=>`
    <div class="result-file">
      <span class="ricon">📊</span>
      <div class="rinfo">
        <div class="rname">${f.name}</div>
        <div class="rsize">${fmt(f.size)}</div>
      </div>
      <button onclick="openFolder()" style="margin-left:auto;background:none;border:1px solid #CBD5E0;
        color:#555;font-size:12px;padding:5px 12px;border-radius:8px;cursor:pointer;white-space:nowrap">
        📁 폴더 열기
      </button>
    </div>
    <div class="action-row" style="margin-bottom:16px">
      <button class="btn-dl" onclick="downloadFile('${f.name}')">⬇ Excel 다운로드</button>
      <button class="btn-drive" id="btn-drive" onclick="uploadToDrive('${f.name}',this)">
        <svg width="16" height="16" viewBox="0 0 87.3 78" xmlns="http://www.w3.org/2000/svg">
          <path d="m6.6 66.85 3.85 6.65c.8 1.4 1.95 2.5 3.3 3.3l13.75-23.8h-27.5c0 1.55.4 3.1 1.2 4.5z" fill="#0066da"/>
          <path d="m43.65 25-13.75-23.8c-1.35.8-2.5 1.9-3.3 3.3l-25.4 44a9.06 9.06 0 0 0-1.2 4.5h27.5z" fill="#00ac47"/>
          <path d="m73.55 76.8c1.35-.8 2.5-1.9 3.3-3.3l1.6-2.75 7.65-13.25c.8-1.4 1.2-2.95 1.2-4.5h-27.502l5.852 11.5z" fill="#ea4335"/>
          <path d="m43.65 25 13.75-23.8c-1.35-.8-2.9-1.2-4.5-1.2h-18.5c-1.6 0-3.15.45-4.5 1.2z" fill="#00832d"/>
          <path d="m59.8 53h-32.3l-13.75 23.8c1.35.8 2.9 1.2 4.5 1.2h50.8c1.6 0 3.15-.45 4.5-1.2z" fill="#2684fc"/>
          <path d="m73.4 26.5-12.7-22c-.8-1.4-1.95-2.5-3.3-3.3l-13.75 23.8 16.15 27h27.45c0-1.55-.4-3.1-1.2-4.5z" fill="#ffba00"/>
        </svg>
        Google Drive에 올리기
      </button>
    </div>`).join('');
}

async function openFolder(){
  await fetch('/open-folder', {method:'POST'});
  toast('📁 Finder에서 폴더를 열었습니다');
}

function downloadFile(name){
  const a=document.createElement('a');
  a.href='/download/'+encodeURIComponent(name);
  a.download=name;
  document.body.appendChild(a);a.click();document.body.removeChild(a);
  toast('⬇ 다운로드 시작!');
}
const DRIVE_SVG=`<svg width="16" height="16" viewBox="0 0 87.3 78" xmlns="http://www.w3.org/2000/svg"><path d="m6.6 66.85 3.85 6.65c.8 1.4 1.95 2.5 3.3 3.3l13.75-23.8h-27.5c0 1.55.4 3.1 1.2 4.5z" fill="#0066da"/><path d="m43.65 25-13.75-23.8c-1.35.8-2.5 1.9-3.3 3.3l-25.4 44a9.06 9.06 0 0 0-1.2 4.5h27.5z" fill="#00ac47"/><path d="m73.55 76.8c1.35-.8 2.5-1.9 3.3-3.3l1.6-2.75 7.65-13.25c.8-1.4 1.2-2.95 1.2-4.5h-27.502l5.852 11.5z" fill="#ea4335"/><path d="m43.65 25 13.75-23.8c-1.35-.8-2.9-1.2-4.5-1.2h-18.5c-1.6 0-3.15.45-4.5 1.2z" fill="#00832d"/><path d="m59.8 53h-32.3l-13.75 23.8c1.35.8 2.9 1.2 4.5 1.2h50.8c1.6 0 3.15-.45 4.5-1.2z" fill="#2684fc"/><path d="m73.4 26.5-12.7-22c-.8-1.4-1.95-2.5-3.3-3.3l-13.75 23.8 16.15 27h27.45c0-1.55-.4-3.1-1.2-4.5z" fill="#ffba00"/></svg>`;

async function uploadToDrive(filename, btn){
  btn.disabled=true;
  btn.innerHTML='⏳ 업로드 중...';
  toast('☁️ Google Drive에 업로드 중...');
  try{
    const r=await fetch('/upload-to-drive',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({filename})
    });
    const d=await r.json();
    if(d.ok){
      toast('✅ Google Drive 업로드 완료!');
      if(d.link) window.open(d.link,'_blank');
    } else if(d.need_credentials){
      openDriveModal();
      toast('🔑 credentials.json 설정이 필요합니다','error');
    } else {
      toast('❌ '+d.error,'error');
    }
  }catch(e){
    toast('❌ 오류: '+e.message,'error');
  }
  btn.disabled=false;
  btn.innerHTML=DRIVE_SVG+' Google Drive에 올리기';
}

function openDriveModal(){document.getElementById('drive-modal').classList.add('open')}
function closeModal(){document.getElementById('drive-modal').classList.remove('open')}
document.getElementById('drive-modal').addEventListener('click',e=>{
  if(e.target===e.currentTarget)closeModal();});

async function loadOutputs(){
  const r=await fetch('/outputs');
  const files=await r.json();
  if(files.length) renderResults(files);
}

async function shutdownServer(){
  if(!confirm('서버를 종료하시겠습니까?\n브라우저 창도 함께 닫힙니다.')) return;
  await fetch('/shutdown', {method:'POST'}).catch(()=>{});
  document.body.innerHTML='<div style="display:flex;align-items:center;justify-content:center;height:100vh;font-size:18px;color:#666">✅ 서버가 종료되었습니다. 창을 닫아주세요.</div>';
}

loadSpecs();
loadOutputs();
</script>
</body>
</html>"""

# ────────────────────────────────────────────────────────────────
#  API 라우트
# ────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template_string(HTML)


@app.route("/files")
def list_specs():
    files = [
        {"name": f.name, "size": f.stat().st_size}
        for f in sorted(SPECS_DIR.iterdir())
        if f.suffix.lower() == ".pdf" and f.is_file()
    ]
    return jsonify(files)


@app.route("/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "파일이 없습니다."})
    f = request.files["file"]
    if not f.filename or Path(f.filename).suffix.lower() != ".pdf":
        return jsonify({"ok": False, "error": "PDF 파일만 업로드 가능합니다."})
    f.save(SPECS_DIR / f.filename)
    return jsonify({"ok": True})


@app.route("/delete", methods=["POST"])
def delete():
    name = (request.get_json() or {}).get("name", "")
    target = (SPECS_DIR / name).resolve()
    if not str(target).startswith(str(SPECS_DIR.resolve())):
        return jsonify({"ok": False, "error": "잘못된 경로"})
    if target.exists():
        target.unlink()
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "파일 없음"})


@app.route("/shutdown", methods=["POST"])
def shutdown():
    import threading, signal
    def _stop():
        import time; time.sleep(0.3)
        os.kill(os.getpid(), signal.SIGTERM)
    threading.Thread(target=_stop, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/generate", methods=["POST"])
def generate():
    if not BUILD_SCRIPT.exists():
        return jsonify({"ok": False, "error": f"빌드 스크립트 없음: {BUILD_SCRIPT}"})
    try:
        result = subprocess.run(
            [sys.executable, str(BUILD_SCRIPT)],
            cwd=str(BASE_DIR),
            capture_output=True,
            text=True,
            timeout=120,
        )
        output = (result.stdout or "") + (result.stderr or "")
        if result.returncode != 0:
            return jsonify({"ok": False, "error": output or "빌드 실패"})
        excels = sorted(OUTPUTS_DIR.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
        files  = [{"name": f.name, "size": f.stat().st_size} for f in excels[:1]]
        return jsonify({"ok": True, "output": output, "files": files})
    except subprocess.TimeoutExpired:
        return jsonify({"ok": False, "error": "빌드 타임아웃 (120초 초과)"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/outputs")
def list_outputs():
    excels = sorted(OUTPUTS_DIR.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    latest = excels[:1]  # 최신 파일 1개만
    return jsonify([{"name": f.name, "size": f.stat().st_size} for f in latest])


@app.route("/download/<filename>")
def download(filename):
    target = (OUTPUTS_DIR / filename).resolve()
    if not str(target).startswith(str(OUTPUTS_DIR.resolve())):
        return "Invalid path", 400
    if not target.exists():
        return "File not found", 404
    return send_file(str(target), as_attachment=True, download_name=filename)


@app.route("/open-folder", methods=["POST"])
def open_folder():
    import subprocess
    subprocess.Popen(["open", str(OUTPUTS_DIR)])
    return jsonify({"ok": True})


# ── Google Drive ─────────────────────────────────────────────────
def load_config():
    if CONFIG_FILE.exists():
        return json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
    return {}


def get_drive_service():
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build

    SCOPES = ["https://www.googleapis.com/auth/drive.file"]

    if not DRIVE_CREDS_FILE.exists():
        raise FileNotFoundError(
            "credentials.json이 없습니다.\n"
            "Google Cloud Console → API 및 서비스 → 사용자 인증 정보 →\n"
            "OAuth 2.0 클라이언트 ID (데스크톱 앱)를 만들고\n"
            f"다운로드한 파일을 '{DRIVE_CREDS_FILE}' 으로 저장하세요."
        )

    creds = None
    if DRIVE_TOKEN_FILE.exists():
        creds = Credentials.from_authorized_user_file(str(DRIVE_TOKEN_FILE), SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                str(DRIVE_CREDS_FILE), SCOPES
            )
            creds = flow.run_local_server(port=0)
        DRIVE_TOKEN_FILE.write_text(creds.to_json(), encoding="utf-8")

    return build("drive", "v3", credentials=creds)


@app.route("/upload-to-drive", methods=["POST"])
def upload_to_drive():
    data = request.get_json() or {}
    filename = data.get("filename", "")

    file_path = (OUTPUTS_DIR / filename).resolve()
    if not str(file_path).startswith(str(OUTPUTS_DIR.resolve())):
        return jsonify({"ok": False, "error": "잘못된 파일 경로"})
    if not file_path.exists():
        return jsonify({"ok": False, "error": "파일을 찾을 수 없습니다."})

    config = load_config()
    folder_id = config.get("google_drive", {}).get("upload_folder_id")
    if not folder_id:
        return jsonify({"ok": False, "error": "config.json에 google_drive.upload_folder_id가 없습니다."})

    try:
        from googleapiclient.http import MediaFileUpload
        service = get_drive_service()

        file_metadata = {"name": filename, "parents": [folder_id]}
        media = MediaFileUpload(
            str(file_path),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        uploaded = service.files().create(
            body=file_metadata,
            media_body=media,
            fields="id,webViewLink",
        ).execute()

        folder_url = config.get("google_drive", {}).get("folder_url", "")
        return jsonify({
            "ok": True,
            "file_id": uploaded.get("id"),
            "link": folder_url or uploaded.get("webViewLink"),
        })
    except FileNotFoundError as e:
        return jsonify({"ok": False, "error": str(e), "need_credentials": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5001))
    print(f"\n✅ TC 자동화 웹 인터페이스 시작!")
    print(f"   기획서 저장 경로 : {SPECS_DIR}")
    print(f"   결과물 저장 경로 : {OUTPUTS_DIR}")
    print(f"   브라우저에서 → http://localhost:{port}\n")
    app.run(host="0.0.0.0", port=port, debug=False)
