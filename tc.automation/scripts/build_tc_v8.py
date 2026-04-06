import sys, os
from datetime import datetime

# ── 경로 설정 (상대 경로 기반) ─────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR   = os.path.join(BASE_DIR, "data")
OUTPUT_DIR = os.path.join(BASE_DIR, "outputs")
sys.path.insert(0, DATA_DIR)

from tc_new_final import NEW_TC
from tc_old_final import OLD_TC

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from collections import Counter

ALL_TC = NEW_TC + OLD_TC

# ── 색상 정의 ─────────────────────────────────────────────────
C_DARK  = "1F3864"; C_WHITE = "FFFFFF"
C_BLUE1 = "2E75B6"; C_BLUE2 = "DDEBF7"
C_ROW_A = "FFFFFF"; C_ROW_B = "F5F9FE"
C_T1    = "1C6E38"   # Tester 1 (초록)
C_T2    = "2E4057"   # Tester 2 (어두운 파랑)

PHASE_COLORS = {
    "Phase 00A": ("7B2D8B","FFFFFF"), "Phase 00B": ("D44000","FFFFFF"),
    "Phase 00C": ("0070C0","FFFFFF"), "Phase 01":  ("1F3864","FFFFFF"),
    "Phase 02":  ("2E4057","FFFFFF"), "Phase 03":  ("1C6E38","FFFFFF"),
    "Phase 04":  ("7B3F00","FFFFFF"), "Phase 05":  ("C55A11","FFFFFF"),
    "Phase 06":  ("833C00","FFFFFF"), "Phase 07":  ("4472C4","FFFFFF"),
    "Phase 08":  ("7030A0","FFFFFF"), "Phase 09":  ("9C2B2B","FFFFFF"),
    "Phase 10":  ("385723","FFFFFF"), "Phase 11":  ("1F497D","FFFFFF"),
    "Phase 12":  ("404040","FFFFFF"), "Phase 13":  ("C00000","FFFFFF"),
    "Phase 14":  ("7F6000","FFFFFF"), "Phase 15":  ("2F5496","FFFFFF"),
    "Phase 16":  ("375623","FFFFFF"), "Phase 17":  ("843C0C","FFFFFF"),
    "Phase 18":  ("595959","FFFFFF"),
}

# 통계 시트 전용 파스텔 팔레트: (배경 파스텔, 텍스트 어두운 색)
# 스크린샷 참조 색상 직접 정의
PHASE_COLORS_PASTEL = {
    "Phase 00A": ("F9D0D0", "7B2D8B"),   # 연핑크
    "Phase 00B": ("F9D0D0", "D44000"),   # 연핑크
    "Phase 00C": ("F4CCCC", "0070C0"),   # 연로즈
    "Phase 01":  ("BDD7EE", "1F3864"),   # 연하늘
    "Phase 02":  ("DDEBF7", "2E4057"),   # 연하늘(밝)
    "Phase 03":  ("E2EFDA", "1C6E38"),   # 연세이지
    "Phase 04":  ("F4CCCC", "7B3F00"),   # 연주황
    "Phase 05":  ("FCE4D6", "C55A11"),   # 연살구
    "Phase 06":  ("FFD5B8", "833C00"),   # 연오렌지
    "Phase 07":  ("BDD7EE", "4472C4"),   # 연하늘
    "Phase 08":  ("EAD1DC", "7030A0"),   # 연라벤더
    "Phase 09":  ("F4CCCC", "9C2B2B"),   # 연레드
    "Phase 10":  ("D9EAD3", "385723"),   # 연그린
    "Phase 11":  ("BDD7EE", "1F497D"),   # 연하늘
    "Phase 12":  ("D9D9D9", "404040"),   # 연그레이
    "Phase 13":  ("F9D0D0", "C00000"),   # 연핑크레드
    "Phase 14":  ("FFF2CC", "7F6000"),   # 연옐로우
    "Phase 15":  ("E2EFDA", "2F5496"),   # 연연두
    "Phase 16":  ("EFEFEF", "375623"),   # 연실버
    "Phase 17":  ("FCE4D6", "843C0C"),   # 연살구
    "Phase 18":  ("D9D9D9", "595959"),   # 연그레이
}

PHASE_LABELS = {
    "Phase 00A": "Phase 00A  ·  Email 가입 / 로그인 시나리오",
    "Phase 00B": "Phase 00B  ·  Gmail 계정 가입 / 로그인 시나리오",
    "Phase 00C": "Phase 00C  ·  Web3 Wallet 가입 / 로그인 시나리오  (Metamask · Hanawallet · OKX Wallet · Rabby Wallet · Trust Wallet)",
    "Phase 01":  "Phase 01  ·  신규 가입 & 계정 태깅",
    "Phase 02":  "Phase 02  ·  최초 로그인 & 안내 팝업",
    "Phase 03":  "Phase 03  ·  초기 설정 – Auto TP/SL",
    "Phase 04":  "Phase 04  ·  레버리지 설정 (첫 거래 준비)",
    "Phase 05":  "Phase 05  ·  첫 번째 거래 – 시장가 롱 (Happy Path · S1)",
    "Phase 06":  "Phase 06  ·  지정가 주문 (S3)",
    "Phase 07":  "Phase 07  ·  숏 포지션",
    "Phase 08":  "Phase 08  ·  수동 TP/SL 우선 입력",
    "Phase 09":  "Phase 09  ·  수동 변경 (S5)",
    "Phase 10":  "Phase 10  ·  추가 진입 / 물타기 (S4)",
    "Phase 11":  "Phase 11  ·  설정 변경 후 영향도 (S6)",
    "Phase 12":  "Phase 12  ·  Auto TP/SL OFF 시나리오 (S2)",
    "Phase 13":  "Phase 13  ·  오류 / 실패 케이스",
    "Phase 14":  "Phase 14  ·  기존 고레버리지 포지션 (S9)",
    "Phase 15":  "Phase 15  ·  경계값 & 동치분할",
    "Phase 16":  "Phase 16  ·  탐색적 테스트",
    "Phase 17":  "Phase 17  ·  회귀 테스트",
    "Phase 18":  "Phase 18  ·  비기능 테스트",
}

PHASE_DAEBUNRYU = {
    "Phase 00A":"계정 · 인증","Phase 00B":"계정 · 인증","Phase 00C":"계정 · 인증",
    "Phase 01":"계정 · 인증","Phase 02":"계정 · 인증","Phase 03":"Settings",
    "Phase 04":"레버리지","Phase 05":"Trade · TP/SL","Phase 06":"Trade · TP/SL",
    "Phase 07":"Trade · TP/SL","Phase 08":"Trade · TP/SL","Phase 09":"Trade · TP/SL",
    "Phase 10":"Trade · TP/SL","Phase 11":"Trade · TP/SL","Phase 12":"Trade · TP/SL",
    "Phase 13":"오류 처리","Phase 14":"레버리지","Phase 15":"경계값",
    "Phase 16":"탐색적 테스트","Phase 17":"회귀 테스트","Phase 18":"비기능 테스트",
}

# ── 테스터 배정 설정 ─────────────────────────────────────────
# 실제 테스터 이름 또는 역할명으로 변경 가능
TESTER_ASSIGNMENT = {
    "Gate":        "Tester 2",
    "OKX":         "Tester 1",
    "Bybit":       "Tester 1",
    "Bitget":      "Tester 2",
    "Hyperliquid": "Tester 2",
}
TESTER_COLORS = {"Tester 1": C_T1, "Tester 2": C_T2}

# ── 컬럼 정의 (18개) ─────────────────────────────────────────
# Col 1:TC ID  2:단계  3:대분류  4:중분류  5:소분류  6:시나리오
#     7:우선순위  8:중요도  9:GIVEN  10:WHEN  11:THEN  12:실제결과
#     13:Gate  14:OKX  15:Bybit  16:Bitget  17:Hyperliquid  18:비고
EXCHANGE_COLS = {"Gate":13,"OKX":14,"Bybit":15,"Bitget":16,"Hyperliquid":17}
EXCHANGE_LIST = ["Gate","OKX","Bybit","Bitget","Hyperliquid"]

HEADERS = [
    "TC ID","단계","대분류","중분류","소분류",
    "시나리오 (BDD 요약)","우선순위","중요도",
    "GIVEN  (사전 조건)","WHEN  (테스트 단계)","THEN  (기대 결과)",
    "실제 결과",
    "Gate","OKX","Bybit","Bitget","Hyperliquid",
    "비고"
]
COL_W = [16,9,14,20,14,36,8,8,30,44,40,24,7,7,7,7,10,18]
NUM_COLS = len(HEADERS)  # 18

# ── 거래소별 N/A 사전 기입 (완전판) ─────────────────────────────
# 계정·인증·팝업 Phase (00A/B/C, 01, 02):
#   OKX만 테스트, 나머지 4개 거래소 N/A ("OKX 최초 거래소" 기준)
_AUTH_NA = ["Gate","Bybit","Bitget","Hyperliquid"]
_AUTH_IDS = [
    "SPCY-00A-001","SPCY-00A-002","SPCY-00A-003","SPCY-00A-005","SPCY-00A-011",
    "SPCY-00B-001","SPCY-00B-002","SPCY-00B-003","SPCY-00B-005","SPCY-00B-007",
    "SPCY-00C-001","SPCY-00C-002","SPCY-00C-004","SPCY-00C-006","SPCY-00C-008",
    "SPCY-00C-010","SPCY-00C-012","SPCY-00C-013","SPCY-00C-014","SPCY-00C-015",
    "SPCY-00C-016","SPCY-00C-019","SPCY-00C-020",
    "SPCY-01-001","SPCY-01-002","SPCY-01-003",
    "SPCY-02-001","SPCY-02-002","SPCY-02-003","SPCY-02-004",
    "SPCY-02-005","SPCY-02-006","SPCY-02-007",
]
EXCHANGE_NA = {tc_id: list(_AUTH_NA) for tc_id in _AUTH_IDS}

# Phase 05: Gate N/A (Gate는 Auto TP/SL 미지원)
for tc_id in ["SPCY-05-001","SPCY-05-002","SPCY-05-003","SPCY-05-004",
              "SPCY-05-005","SPCY-05-006","SPCY-05-008"]:
    EXCHANGE_NA[tc_id] = ["Gate"]

# Phase 06: 거래소별 세분화
EXCHANGE_NA["SPCY-06-001"] = ["Gate"]
EXCHANGE_NA["SPCY-06-002"] = ["Gate","Bitget","Hyperliquid"]   # 미체결 TP/SL 수정: OKX/Bybit만 지원
EXCHANGE_NA["SPCY-06-003"] = ["Gate"]
EXCHANGE_NA["SPCY-06-004"] = ["Gate"]
EXCHANGE_NA["SPCY-06-005"] = ["Gate"]
EXCHANGE_NA["SPCY-06-006"] = ["Gate","Bitget","Hyperliquid"]   # 수정 버튼 없음
EXCHANGE_NA["SPCY-06-007"] = ["Gate","Bitget","Hyperliquid"]   # 추가 버튼 없음
EXCHANGE_NA["SPCY-06-008"] = ["OKX","Bybit","Bitget"]          # Gate/HL 대상 TC

# Phase 07-08: Gate N/A (TP/SL 미지원)
for tc_id in ["SPCY-07-001","SPCY-07-002","SPCY-08-001","SPCY-08-002"]:
    EXCHANGE_NA[tc_id] = ["Gate"]

# ── 최소 TC 세트 (리스크 기반 필수 TC) ─────────────────────────
MIN_TC_SET = {
    # 계정·인증: 대표 Happy Path + 태깅 확인
    "SPCY-00A-001","SPCY-00A-011",
    "SPCY-00B-001","SPCY-00B-007",
    "SPCY-00C-002","SPCY-00C-020",
    # 태깅 정책
    "SPCY-01-001","SPCY-01-002","SPCY-01-003",
    # 최초 팝업
    "SPCY-02-001","SPCY-02-003","SPCY-02-005",
    # Auto TP/SL Settings 핵심
    "SPCY-03-001","SPCY-03-002","SPCY-03-003","SPCY-03-014","SPCY-03-018",
    # 레버리지 제한 핵심
    "SPCY-04-001","SPCY-04-003","SPCY-04-005","SPCY-04-006","SPCY-04-009",
    # Happy Path + TP/SL 계산
    "SPCY-05-004","SPCY-05-005","SPCY-05-006","SPCY-05-008",
    # 지정가 체결 + Open Orders TP/SL
    "SPCY-06-003","SPCY-06-006","SPCY-06-007","SPCY-06-008",
    # 숏 포지션
    "SPCY-07-001","SPCY-07-002",
    # 수동 TP/SL 우선
    "SPCY-08-001",
    # 수동 변경
    "SPCY-09-001","SPCY-09-002",
    # 물타기
    "SPCY-10-001","SPCY-10-002",
    # 설정 변경 영향
    "SPCY-11-001",
    # Auto TP/SL OFF
    "SPCY-12-001",
    # 오류 케이스
    "SPCY-13-002",
    # 기존 포지션
    "SPCY-14-001","SPCY-14-002","SPCY-14-003",
    # 경계값 핵심
    "SPCY-15-001","SPCY-15-007","SPCY-15-011","SPCY-15-013",
    # 탐색적
    "SPCY-16-002","SPCY-16-003",
    # 회귀
    "SPCY-17-001","SPCY-17-002",
    # 비기능
    "SPCY-18-002","SPCY-18-004",
}

# ── 스타일 헬퍼 ───────────────────────────────────────────────
def fill(c):   return PatternFill("solid", fgColor=c)
def font(bold=False, color="000000", size=9, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")
def border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)
def align(h="left", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════
# SHEET 1 – 표지
# ═══════════════════════════════════════════════════════════════
cv = wb.active; cv.title = "📋 표지"
cv.sheet_view.showGridLines = False

def cm(r1,c1,r2,c2,val,bg,fg,sz=11,bold=True,ha="center"):
    cv.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
    cell = cv.cell(r1,c1,val)
    cell.font = Font(bold=bold,color=fg,size=sz,name="Arial")
    cell.fill = fill(bg); cell.border = border()
    cell.alignment = Alignment(horizontal=ha,vertical="center",wrap_text=True)

cm(2,2,2,12,"Supercycl × Youthmeta  |  YouthMeta Partner Version 1.0",C_DARK,C_WHITE,16)
cm(3,2,3,12,"테스트 케이스 명세서  ·  v8  ·  BDD Given/When/Then  ·  대분류/중분류/소분류  ·  거래소별 테스터 배정",C_BLUE1,C_WHITE,11,False)
cv.row_dimensions[2].height = 32; cv.row_dimensions[3].height = 22

meta = [
    ("대상 기능","자동 TP/SL 설정 기능  +  유스메타 유입 유저 레버리지 2배 고정"),
    ("대상 거래소","Gate / OKX / Bybit / Bitget / Hyperliquid"),
    ("TC 방법론","User Journey  ·  BDD Given/When/Then  ·  블랙박스  ·  Risk-Based"),
    ("테스터 배정","Tester 1 → OKX / Bybit  |  Tester 2 → Gate / Bitget / Hyperliquid"),
    ("작성일","2026-03-18  (v8 – 테스터 배정 서브헤더 / 거래소 N/A 전체 정비 / 최소 TC 세트 시트 추가)"),
]
for i,(k,v) in enumerate(meta,5):
    cv.cell(i,2,k).font = Font(bold=True,color=C_WHITE,size=9,name="Arial")
    cv.cell(i,2).fill = fill(C_BLUE1); cv.cell(i,2).border = border()
    cv.cell(i,2).alignment = align("center",False)
    cv.merge_cells(start_row=i,start_column=3,end_row=i,end_column=12)
    cv.cell(i,3,v).font = font(size=9); cv.cell(i,3).fill = fill(C_BLUE2)
    cv.cell(i,3).border = border(); cv.cell(i,3).alignment = align()
    cv.row_dimensions[i].height = 18

# Phase 목차
r = 12
cv.merge_cells(start_row=r,start_column=2,end_row=r,end_column=12)
cv.cell(r,2,"User Journey 단계별 테스트 구성").font = Font(bold=True,color=C_WHITE,size=11,name="Arial")
cv.cell(r,2).fill = fill(C_DARK); cv.cell(r,2).border = border()
cv.cell(r,2).alignment = align("center",False); cv.row_dimensions[r].height = 20

phase_cnt = Counter(item[0] for item in ALL_TC)
phases_info = [
    ("Phase 00A","Email 가입/로그인","이메일 인증·약관·중단 케이스"),
    ("Phase 00B","Gmail 계정 가입/로그인","OAuth·계정선택·약관·중단 케이스"),
    ("Phase 00C","Web3 Wallet 가입/로그인","5종 지갑 Happy Path·연결취소·서명거부"),
    ("Phase 01","신규 가입 & 계정 태깅","레퍼럴 링크, partner_code 태깅"),
    ("Phase 02","최초 로그인 & 안내 팝업","약관 팝업 → 레버리지 2x 팝업 노출·동작"),
    ("Phase 03","초기 설정 – Auto TP/SL","Settings, 설정 모달, 토글, 동기화"),
    ("Phase 04","레버리지 설정","Adjust Leverage 2x 제한"),
    ("Phase 05","첫 번째 거래 – 시장가 롱","Happy Path, ROI 기반 TP/SL 자동 생성"),
    ("Phase 06","지정가 주문 (S3)","지정가 미체결/체결, Open Orders TP/SL 수정 (OKX/Bybit)"),
    ("Phase 07","숏 포지션","역방향 ROI 기반 TP/SL 계산"),
    ("Phase 08","수동 TP/SL 우선","수동값 우선 적용, 거래소별 동작"),
    ("Phase 09","수동 변경 (S5)","수정 모달, Bybit Full TP/SL"),
    ("Phase 10","물타기 (S4)","거래소별 Partial/Full TP/SL"),
    ("Phase 11","설정 변경 (S6)","기존 포지션 소급 없음"),
    ("Phase 12","Auto TP/SL OFF (S2)","OFF 상태 미생성·기존 유지"),
    ("Phase 13","오류/실패 케이스","네트워크 오류 처리"),
    ("Phase 14","기존 고레버리지 (S9)","포지션 유지·신규 2x 적용"),
    ("Phase 15","경계값 & 동치분할","TP/SL·레버리지 입력 범위"),
    ("Phase 16","탐색적 테스트","토글·극값·동시처리·네트워크"),
    ("Phase 17","회귀 테스트","기존 기능 영향 없음"),
    ("Phase 18","비기능 테스트","성능·UI 일관성"),
]
for i,(ph,name,desc) in enumerate(phases_info,r+1):
    bg,fg = PHASE_COLORS.get(ph,(C_DARK,C_WHITE))
    cnt = str(phase_cnt.get(ph,0))
    cv.cell(i,2,ph).font = Font(bold=True,color=fg,size=9,name="Arial")
    cv.cell(i,2).fill = fill(bg); cv.cell(i,2).border = border()
    cv.cell(i,2).alignment = align("center",False)
    cv.cell(i,3,name).font = font(bold=True,size=9)
    cv.cell(i,3).fill = fill(C_BLUE2 if i%2==0 else C_ROW_A)
    cv.cell(i,3).border = border(); cv.cell(i,3).alignment = align()
    cv.merge_cells(start_row=i,start_column=4,end_row=i,end_column=11)
    cv.cell(i,4,desc).font = font(size=9,italic=True)
    cv.cell(i,4).fill = fill(C_BLUE2 if i%2==0 else C_ROW_A)
    cv.cell(i,4).border = border(); cv.cell(i,4).alignment = align()
    cv.cell(i,12,cnt+"개").font = font(bold=True,size=9,color=C_DARK)
    cv.cell(i,12).fill = fill(C_BLUE2 if i%2==0 else C_ROW_A)
    cv.cell(i,12).border = border(); cv.cell(i,12).alignment = align("center",False)
    cv.row_dimensions[i].height = 16

lr = r+len(phases_info)+2
cv.merge_cells(start_row=lr,start_column=2,end_row=lr,end_column=12)
cv.cell(lr,2,"범례  ·  판정 코드").font = Font(bold=True,color=C_WHITE,size=10,name="Arial")
cv.cell(lr,2).fill = fill(C_DARK); cv.cell(lr,2).border = border()
cv.cell(lr,2).alignment = align("center",False); cv.row_dimensions[lr].height = 18

for j,rd in enumerate([
    ("분류 계층","대분류 = 기능 도메인","중분류 = 세부 기능 영역","소분류 = 케이스 유형"),
    ("테스터 배정","Tester 1 → OKX / Bybit","Tester 2 → Gate / Bitget / Hyperliquid","각 컬럼에 직접 기입"),
    ("판정 코드","Pass = 기대 결과 일치","Fail = 버그 → 리포트 필수","N/T = 미테스트 (기본값) / N/A = 해당 없음"),
    ("우선순위","P1 = 블로킹 위험","P2 = 핵심 흐름 영향","P3 = 선택적 개선"),
    ("★ 최소 TC","★ 표시 TC = 필수 최소 세트","'🎯 최소 TC 세트' 시트 참조","리스크 기반 선별 (P1/P2 + Happy Path)"),
],lr+1):
    cv.cell(j,2,rd[0]).font = font(bold=True,color=C_WHITE,size=9)
    cv.cell(j,2).fill = fill("595959"); cv.cell(j,2).border = border()
    cv.cell(j,2).alignment = align("center",False)
    for k,val in enumerate(rd[1:],1):
        cs=k*3+1; ce=k*3+3
        if ce>12: ce=12
        cv.merge_cells(start_row=j,start_column=cs,end_row=j,end_column=ce)
        cv.cell(j,cs,val).font = font(size=9)
        cv.cell(j,cs).fill = fill("F2F2F2"); cv.cell(j,cs).border = border()
        cv.cell(j,cs).alignment = align()
    cv.row_dimensions[j].height = 16

for c,w in [("A",2),("B",14),("C",20),("D",36),("E",2),("F",2),("G",2),("H",2),("I",2),("J",2),("K",2),("L",8)]:
    cv.column_dimensions[c].width = w

# ═══════════════════════════════════════════════════════════════
# SHEET 2 – TC 전체목록
# ═══════════════════════════════════════════════════════════════
tc_sheet = wb.create_sheet("🧪 TC 전체목록")
tc_sheet.sheet_view.showGridLines = False
tc_sheet.freeze_panes = "A3"  # Row 1=헤더, Row 2=테스터배정, Row3+=데이터

# ── Row 1: 메인 헤더 ──────────────────────────────────────────
for ci,(h,w) in enumerate(zip(HEADERS,COL_W),1):
    c = tc_sheet.cell(1,ci,h)
    c.font = Font(bold=True,color=C_WHITE,size=9,name="Arial")
    c.fill = fill(C_DARK); c.border = border()
    c.alignment = align("center",False)
    tc_sheet.column_dimensions[get_column_letter(ci)].width = w
tc_sheet.row_dimensions[1].height = 22

# 거래소 헤더: 테스터별 색상 적용
for exch, col_ci in EXCHANGE_COLS.items():
    tester = TESTER_ASSIGNMENT.get(exch,"")
    tc_sheet.cell(1,col_ci).fill = fill(TESTER_COLORS.get(tester,C_DARK))

# ── Row 2: 테스터 배정 서브헤더 ──────────────────────────────
tc_sheet.row_dimensions[2].height = 16
for ci in range(1, NUM_COLS+1):
    c = tc_sheet.cell(2, ci)
    c.border = border()
    if ci <= 12 or ci == 18:
        # 비판정 컬럼: 연한 배경, 비움
        c.fill = fill("E8E8E8")
        c.font = font(size=8, italic=True, color="888888")
        if ci == 9:
            c.value = "← 테스터 배정: 각 거래소 컬럼에 Pass/Fail/N/T 직접 기입 →"
            c.alignment = align("center", False)
    else:
        # 거래소 컬럼: 테스터 이름 표시
        exch = HEADERS[ci-1]
        tester = TESTER_ASSIGNMENT.get(exch, "")
        t_color = TESTER_COLORS.get(tester, C_DARK)
        c.value = tester
        c.fill = fill(t_color)
        c.font = Font(bold=True, color=C_WHITE, size=8, name="Arial")
        c.alignment = align("center", False)

# ── Row 2 테스터 배정 드롭다운 ────────────────────────────────
tester_dv = DataValidation(type="list", formula1='"Tester 1,Tester 2"',
                            allow_blank=False, showDropDown=False)
tester_dv.prompt  = "담당 테스터를 선택하세요"
tester_dv.promptTitle = "테스터 배정"
tc_sheet.add_data_validation(tester_dv)
for exch, col_ci in EXCHANGE_COLS.items():
    tester_dv.add(f"{get_column_letter(col_ci)}2")

# ── 판정 드롭다운 DataValidation ─────────────────────────────
dv = DataValidation(type="list", formula1='"Pass,Fail,N/T,N/A"',
                    allow_blank=False, showDropDown=False)
dv.error="Pass, Fail, N/T, N/A 중 하나를 선택하세요"
dv.errorTitle="입력 오류"
dv.prompt="판정 결과를 선택하세요"
dv.promptTitle="판정"
tc_sheet.add_data_validation(dv)

# 중앙 정렬 컬럼 (1-based)
CENTER_COLS = {1,2,3,5,7,8,13,14,15,16,17}

prev_phase = None
row_idx    = 3  # 데이터는 Row 3부터 시작

for item in ALL_TC:
    phase,tc_id,test_type,area,scenario,pri,risk,given,when_,then_,note = item
    is_min = "★ " if tc_id in MIN_TC_SET else ""

    # ── Phase 구분 헤더 ──────────────────────────────────────
    if phase != prev_phase:
        if prev_phase is not None:
            row_idx += 1
        bg,fg = PHASE_COLORS.get(phase,(C_DARK,C_WHITE))
        label = PHASE_LABELS.get(phase,phase)
        tc_sheet.merge_cells(start_row=row_idx,start_column=1,end_row=row_idx,end_column=NUM_COLS)
        hc = tc_sheet.cell(row_idx,1,label)
        hc.font = Font(bold=True,color=fg,size=10,name="Arial")
        hc.fill = fill(bg); hc.border = border()
        hc.alignment = align("left",False)
        tc_sheet.row_dimensions[row_idx].height = 20
        row_idx += 1
        prev_phase = phase

    tc_sheet.row_dimensions[row_idx].height = 80
    row_fill = fill(C_ROW_A if row_idx%2==0 else C_ROW_B)
    daebunryu = PHASE_DAEBUNRYU.get(phase,"─")

    # row_vals 18개 (0-based index 0~17)
    row_vals = [
        is_min+tc_id, phase, daebunryu, area, test_type,
        scenario, pri, risk, given, when_, then_, "",
        "N/T","N/T","N/T","N/T","N/T",
        note
    ]

    # N/A 사전 기입
    for exch in EXCHANGE_NA.get(tc_id, []):
        col_idx = EXCHANGE_COLS[exch] - 1  # 0-based
        row_vals[col_idx] = "N/A"

    for ci, val in enumerate(row_vals, 1):
        c = tc_sheet.cell(row_idx, ci, val)
        c.border = border(); c.fill = row_fill; c.font = font(size=9)
        c.alignment = align("center") if ci in CENTER_COLS else align("left")

    # TC ID 특별 스타일 (최소 세트 강조)
    id_cell = tc_sheet.cell(row_idx, 1)
    if tc_id in MIN_TC_SET:
        id_cell.font = Font(bold=True, color="1C6E38", size=9, name="Arial")

    # 우선순위 색상 (Col 7)
    pc = tc_sheet.cell(row_idx,7)
    if   pri=="P1": pc.fill=fill("FFE0E0"); pc.font=font(bold=True,color="C00000",size=9)
    elif pri=="P2": pc.fill=fill("FFF3E0"); pc.font=font(bold=True,color="C55A11",size=9)
    else:           pc.fill=fill("E8F5E9"); pc.font=font(bold=True,color="375623",size=9)

    # 중요도 색상 (Col 8)
    rc = tc_sheet.cell(row_idx,8)
    if   risk=="High":   rc.fill=fill("FFE0E0"); rc.font=font(bold=True,color="C00000",size=9)
    elif risk=="Medium": rc.fill=fill("FFF9E0"); rc.font=font(bold=True,color="996600",size=9)
    else:                rc.fill=fill("E8F5E9"); rc.font=font(bold=True,color="375623",size=9)

    # 거래소 판정 셀 스타일 + 드롭다운
    for exch, col_ci in EXCHANGE_COLS.items():
        jc = tc_sheet.cell(row_idx, col_ci)
        cell_val = row_vals[col_ci-1]
        tester = TESTER_ASSIGNMENT.get(exch,"")
        t_color = TESTER_COLORS.get(tester, C_DARK)
        if cell_val == "N/A":
            jc.fill = fill("E0E0E0")
            jc.font = font(bold=False, color="808080", size=9)
        else:
            # N/T: 연한 테스터 색상
            jc.fill = fill("F5F5F5")
        jc.alignment = align("center", False)
        dv.add(jc)

    row_idx += 1

tc_sheet.auto_filter.ref = f"A1:{get_column_letter(NUM_COLS)}1"

# ═══════════════════════════════════════════════════════════════
# SHEET 3 – TC 통계
# ═══════════════════════════════════════════════════════════════
st = wb.create_sheet("📊 TC 통계")
st.sheet_view.showGridLines = False
st.merge_cells("B2:H2")
total = len(ALL_TC)
min_total = len(MIN_TC_SET)
st["B2"] = f"TC 통계  (v8  ·  전체 {total}개  ·  최소 세트 {min_total}개)"
st["B2"].font = Font(bold=True,color=C_WHITE,size=14,name="Arial")
st["B2"].fill = fill(C_DARK); st["B2"].border = border()
st["B2"].alignment = align("center"); st.row_dimensions[2].height = 28

phases_ = [d[0] for d in ALL_TC]
types_  = [d[2] for d in ALL_TC]
pris_   = [d[5] for d in ALL_TC]
risks_  = [d[6] for d in ALL_TC]
p_cnt   = Counter(phases_); t_cnt=Counter(types_)
pr_cnt  = Counter(pris_);   rk_cnt=Counter(risks_)
daeb_   = [PHASE_DAEBUNRYU.get(d[0],"─") for d in ALL_TC]
daeb_cnt= Counter(daeb_)

pns = {
    "Phase 00A":"Email 가입/로그인","Phase 00B":"Gmail 가입/로그인",
    "Phase 00C":"Web3 Wallet","Phase 01":"신규 가입 & 태깅",
    "Phase 02":"최초 로그인 팝업","Phase 03":"Auto TP/SL Settings",
    "Phase 04":"레버리지 설정","Phase 05":"첫 번째 거래",
    "Phase 06":"지정가 주문","Phase 07":"숏 포지션",
    "Phase 08":"수동 TP/SL 우선","Phase 09":"수동 변경",
    "Phase 10":"물타기","Phase 11":"설정 변경 영향도",
    "Phase 12":"Auto OFF","Phase 13":"오류/실패",
    "Phase 14":"기존 고레버리지","Phase 15":"경계값/동치분할",
    "Phase 16":"탐색적","Phase 17":"회귀","Phase 18":"비기능"
}

srows = [("전체 TC 수",total,"100%","─"),("최소 TC 세트",min_total,f"{min_total/total*100:.0f}%","─")]
srows += [("","","",""),("─ 대분류별 ─","","","")]
for db in ["계정 · 인증","Settings","레버리지","Trade · TP/SL","오류 처리","경계값","탐색적 테스트","회귀 테스트","비기능 테스트"]:
    n = daeb_cnt.get(db,0)
    if n>0: srows.append((db,n,f"{n/total*100:.0f}%","─"))
srows += [("","","",""),("─ Phase별 ─","","","")]
for ph in sorted(set(phases_),key=lambda x:(x[:8],x)):
    n=p_cnt[ph]; pastel_bg,dark_fg=PHASE_COLORS_PASTEL.get(ph,(C_BLUE2,C_DARK))
    srows.append((f"{ph}  {pns.get(ph,'')}",n,f"{n/total*100:.0f}%",pastel_bg+"§"+dark_fg))
srows += [("","","",""),("─ 소분류(테스트 유형)별 ─","","","")]
for k in ["UI/UX","경계값/동치분할","중단 케이스","탐색적","회귀","비기능"]:
    n=t_cnt.get(k,0); srows.append((k,n,f"{n/total*100:.0f}%","─"))
srows += [("","","",""),("─ 우선순위별 ─","","","")]
for k,cl in [("P1","FFE0E0"),("P2","FFF3E0"),("P3","E8F5E9")]:
    n=pr_cnt.get(k,0); srows.append((k,n,f"{n/total*100:.0f}%",cl))
srows += [("","","",""),("─ 중요도별 ─","","","")]
for k,cl in [("High","FFE0E0"),("Medium","FFF9E0"),("Low","E8F5E9")]:
    n=rk_cnt.get(k,0); srows.append((k,n,f"{n/total*100:.0f}%",cl))

for i,(label,cnt,pct,bg) in enumerate(srows,4):
    st.row_dimensions[i].height = 16
    if isinstance(label,str) and label.startswith("─"):
        st.merge_cells(f"B{i}:H{i}")
        c=st.cell(i,2,label); c.font=Font(bold=True,color=C_WHITE,size=9,name="Arial")
        c.fill=fill(C_DARK); c.border=border(); c.alignment=align("center",False)
    elif label in ("전체 TC 수","최소 TC 세트"):
        for ci,val,sz in [(2,label,11),(3,cnt,14),(4,pct,9)]:
            c=st.cell(i,ci,val); c.fill=fill(C_BLUE2); c.border=border()
            c.font=Font(bold=True,size=sz,color=C_DARK,name="Arial")
            c.alignment=align("center" if ci>2 else "left",False)
    elif label:
        if isinstance(bg,str) and "§" in bg:
            rb, fg_color = bg.split("§")
        elif isinstance(bg,str) and len(bg)==6:
            rb, fg_color = bg, C_DARK
        else:
            rb, fg_color = (C_BLUE2 if i%2==0 else C_ROW_A), C_DARK
        for ci,val in enumerate([label,cnt,pct],2):
            c=st.cell(i,ci,val); c.fill=fill(rb); c.border=border()
            c.font=Font(size=9,color=fg_color,name="Arial")
            c.alignment=align("center" if ci>2 else "left",False)
for col,w in [("A",2),("B",34),("C",10),("D",10)]:
    st.column_dimensions[col].width = w

# ═══════════════════════════════════════════════════════════════
# SHEET 4 – 🎯 최소 TC 세트
# ═══════════════════════════════════════════════════════════════
ms = wb.create_sheet("🎯 최소 TC 세트")
ms.sheet_view.showGridLines = False
ms.freeze_panes = "A3"

# 헤더
MIN_HEADERS = ["TC ID","단계","대분류","중분류","시나리오 (BDD 요약)","우선순위","Gate","OKX","Bybit","Bitget","Hyperliquid","선정 이유"]
MIN_COL_W   = [20,9,14,22,44,8,7,7,7,7,10,28]
for ci,(h,w) in enumerate(zip(MIN_HEADERS,MIN_COL_W),1):
    c=ms.cell(1,ci,h)
    c.font=Font(bold=True,color=C_WHITE,size=9,name="Arial")
    c.fill=fill(C_DARK); c.border=border(); c.alignment=align("center",False)
    ms.column_dimensions[get_column_letter(ci)].width=w
ms.row_dimensions[1].height=22

# 서브헤더 (테스터 배정) + 드롭다운
ms.row_dimensions[2].height=16
ms_tester_dv = DataValidation(type="list", formula1='"Tester 1,Tester 2"',
                               allow_blank=False, showDropDown=False)
ms_tester_dv.prompt  = "담당 테스터를 선택하세요"
ms_tester_dv.promptTitle = "테스터 배정"
ms.add_data_validation(ms_tester_dv)
for ci in range(1,len(MIN_HEADERS)+1):
    c=ms.cell(2,ci); c.border=border()
    h=MIN_HEADERS[ci-1]
    if h in TESTER_ASSIGNMENT:
        tester=TESTER_ASSIGNMENT[h]
        c.value=tester; c.fill=fill(TESTER_COLORS.get(tester,C_DARK))
        c.font=Font(bold=True,color=C_WHITE,size=8,name="Arial")
        c.alignment=align("center",False)
        ms_tester_dv.add(f"{get_column_letter(ci)}2")
    else:
        c.fill=fill("E8E8E8")

MIN_REASONS = {
    "SPCY-00A-001":"Email 가입 Happy Path (대표 시나리오)",
    "SPCY-00A-011":"partner_code 태깅 확인 (핵심 정책)",
    "SPCY-00B-001":"Gmail 가입 Happy Path",
    "SPCY-00B-007":"Gmail partner_code 태깅 확인",
    "SPCY-00C-002":"Web3 지갑 대표 (Metamask) Happy Path",
    "SPCY-00C-020":"Web3 partner_code 태깅 확인",
    "SPCY-01-001":"레퍼럴 링크 태깅 핵심 (블로킹)",
    "SPCY-01-002":"일반 경로 태깅 없음 확인",
    "SPCY-01-003":"태깅 계정 레버리지 플래그 즉시 적용",
    "SPCY-02-001":"최초 팝업 정상 노출",
    "SPCY-02-003":"I Understand → 재노출 없음",
    "SPCY-02-005":"재로그인 팝업 미노출",
    "SPCY-03-001":"Settings OFF 초기 상태",
    "SPCY-03-002":"ON 전환 + 값 저장",
    "SPCY-03-003":"OFF 전환 저장",
    "SPCY-03-014":"재로그인 후 설정값 유지",
    "SPCY-03-018":"Settings ↔ OrderForm 동기화",
    "SPCY-04-001":"코인 변경 시 2x 자동 설정",
    "SPCY-04-003":"슬라이더 최대 2x 제한",
    "SPCY-04-005":"직접 입력 2 초과 불가",
    "SPCY-04-006":"직접 입력 2x 허용",
    "SPCY-04-009":"팝업 확인 후 재노출 없음",
    "SPCY-05-004":"시장가 롱 Happy Path + TP/SL 생성",
    "SPCY-05-005":"TP Trigger Price 계산 정확성",
    "SPCY-05-006":"SL Trigger Price 계산 정확성",
    "SPCY-05-008":"Toast 정상 노출",
    "SPCY-06-003":"지정가 체결 후 TP/SL 자동 생성",
    "SPCY-06-006":"Open Orders TP/SL 수정 (OKX/Bybit 신규기능)",
    "SPCY-06-007":"Open Orders TP/SL 재등록 (OKX/Bybit 신규기능)",
    "SPCY-06-008":"Gate/HL TP/SL 수정 불가 확인",
    "SPCY-07-001":"숏 포지션 TP/SL 역방향 계산",
    "SPCY-07-002":"숏 포지션 SL 역방향 계산",
    "SPCY-08-001":"수동 TP/SL 우선 적용",
    "SPCY-09-001":"수동 변경 후 반영 확인",
    "SPCY-09-002":"Bybit Full TP/SL 교체 동작",
    "SPCY-10-001":"동일 방향 추가 진입 시 TP/SL 동작",
    "SPCY-10-002":"반대 방향 소량 추가 시 TP/SL 유지",
    "SPCY-11-001":"설정 변경 → 기존 포지션 소급 없음",
    "SPCY-12-001":"OFF 상태 → 주문 시 TP/SL 미생성",
    "SPCY-13-002":"네트워크 오류 복구 후 값 유지",
    "SPCY-14-001":"기존 고레버 포지션 유지 확인",
    "SPCY-14-002":"기존 포지션 신규 주문 2x 적용",
    "SPCY-14-003":"레버리지 balance 충분 시 2x 적용",
    "SPCY-15-001":"TP 최솟값 0.1% 경계",
    "SPCY-15-007":"SL 최솟값 0.1% 경계",
    "SPCY-15-011":"레버리지 2x 상한 초과 입력 차단",
    "SPCY-15-013":"TP/SL 계산 극솟값 정확성",
    "SPCY-16-002":"Settings ↔ OrderForm 동기화 탐색",
    "SPCY-16-003":"다중 주문 동시 처리 안정성",
    "SPCY-17-001":"기존 기능 영향 없음 확인",
    "SPCY-17-002":"레버리지 기존 기능 영향 없음",
    "SPCY-18-002":"TP/SL 자동 생성 3초 이내",
    "SPCY-18-004":"로그인 포함 2초 이내",
}

dv2 = DataValidation(type="list",formula1='"Pass,Fail,N/T,N/A"',allow_blank=False,showDropDown=False)
ms.add_data_validation(dv2)

EXCH_MIN_COLS = {"Gate":7,"OKX":8,"Bybit":9,"Bitget":10,"Hyperliquid":11}
prev_phase_ms = None
ms_row = 3

min_tc_data = [item for item in ALL_TC if item[1] in MIN_TC_SET]

for item in min_tc_data:
    phase,tc_id,_,area,scenario,pri,risk,given,when_,then_,note = item
    if phase != prev_phase_ms:
        if prev_phase_ms is not None:
            ms_row += 1
        bg,fg = PHASE_COLORS.get(phase,(C_DARK,C_WHITE))
        ms.merge_cells(start_row=ms_row,start_column=1,end_row=ms_row,end_column=len(MIN_HEADERS))
        hc=ms.cell(ms_row,1,PHASE_LABELS.get(phase,phase))
        hc.font=Font(bold=True,color=fg,size=10,name="Arial")
        hc.fill=fill(bg); hc.border=border(); hc.alignment=align("left",False)
        ms.row_dimensions[ms_row].height=20
        ms_row+=1; prev_phase_ms=phase

    ms.row_dimensions[ms_row].height=60
    rf=fill(C_ROW_A if ms_row%2==0 else C_ROW_B)
    daebunryu=PHASE_DAEBUNRYU.get(phase,"─")
    reason=MIN_REASONS.get(tc_id,"─")

    row_vals_m=[tc_id,phase,daebunryu,area,scenario,pri,"N/T","N/T","N/T","N/T","N/T",reason]
    for exch in EXCHANGE_NA.get(tc_id,[]):
        if exch in EXCH_MIN_COLS:
            row_vals_m[EXCH_MIN_COLS[exch]-1]="N/A"

    for ci,val in enumerate(row_vals_m,1):
        c=ms.cell(ms_row,ci,val)
        c.border=border(); c.fill=rf; c.font=font(size=9)
        c.alignment=align("center") if ci in {1,2,3,6,7,8,9,10,11} else align("left")

    pc=ms.cell(ms_row,6)
    if   pri=="P1": pc.fill=fill("FFE0E0"); pc.font=font(bold=True,color="C00000",size=9)
    elif pri=="P2": pc.fill=fill("FFF3E0"); pc.font=font(bold=True,color="C55A11",size=9)

    for exch,col_ci in EXCH_MIN_COLS.items():
        jc=ms.cell(ms_row,col_ci)
        if row_vals_m[col_ci-1]=="N/A":
            jc.fill=fill("E0E0E0"); jc.font=font(color="808080",size=9)
        jc.alignment=align("center",False)
        dv2.add(jc)
    ms_row+=1

ms.auto_filter.ref=f"A1:{get_column_letter(len(MIN_HEADERS))}1"

# ── 저장 ─────────────────────────────────────────────────────
os.makedirs(OUTPUT_DIR, exist_ok=True)
date_str = datetime.now().strftime("%Y%m%d")
n = 1
while True:
    filename = f"SPCY_TC_SmokeTest_{date_str}_v{n}.xlsx"
    out = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(out):
        break
    n += 1
wb.save(out)
print(f"✅ Saved: {out}")
print(f"   전체 TC     : {total}개")
print(f"   최소 TC 세트 : {min_total}개 ({min_total/total*100:.0f}%)")
print(f"   컬럼        : {NUM_COLS}개 (테스터 배정 서브헤더 + 5 거래소)")
print(f"   시트        : 표지 / TC 전체목록 / TC 통계 / 최소 TC 세트")
